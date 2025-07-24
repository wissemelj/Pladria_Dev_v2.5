"""
Module 5: Contrôle Qualité - Pladria v3.0
Système de contrôle qualité pour l'analyse et la validation des données.
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import logging
import os
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict, List, Any, Tuple
import threading

# Imports pour le visualiseur
from config.constants import TeamsConfig

# Import des utilitaires Pladria
from config.constants import COLORS, AppInfo, TeamsConfig, UIConfig, AccessControl
from utils.logging_config import setup_logging
from utils.lazy_imports import get_pandas
from ui.styles import create_sofrecom_card

# Import password dialog with error handling
try:
    from ui.components.password_dialog import show_password_dialog
except ImportError as e:
    # Fallback function
    def show_password_dialog(parent, title="", message=""):
        from tkinter import simpledialog
        password = simpledialog.askstring(title, message, show='*')
        return password is not None, password or ""

# Imports pour la génération de rapports Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
from utils.performance import run_async_task
from utils.file_utils import check_file_access


class QualityControlModule:
    """Module de contrôle qualité pour Pladria v3.0."""
    
    def __init__(self, parent: tk.Widget, navigation_manager=None):
        """
        Initialise le module de contrôle qualité.

        Args:
            parent: Widget parent (fenêtre principale)
            navigation_manager: Gestionnaire de navigation (optionnel)
        """
        self.parent = parent
        self.navigation_manager = navigation_manager
        self.logger = logging.getLogger(__name__)

        # Variables de données
        self.qgis_data = None  # Données du fichier résultats QGis
        self.suivi_data = None  # Données du fichier suivi commune
        self.qc_results = None  # Résultats de l'analyse qualité
        self.current_qgis_file_path = None
        self.current_suivi_file_path = None

        # Variables d'interface
        self.main_frame = None
        self.notebook = None  # Notebook pour les onglets
        self.analysis_tab = None  # Onglet analyse
        self.viewer_tab = None  # Onglet visualiseur
        self.progress_var = None
        self.progress_bar = None
        self.status_label = None
        self.status_icon = None

        # Labels d'information des fichiers
        self.qgis_info_label = None
        self.suivi_info_label = None

        # Boutons d'interface
        self.analyze_button = None
        self.export_button = None
        self.results_label = None
        self.results_frame = None

        # Variables du visualiseur
        self.viewer_data = None
        self.viewer_tree = None
        self.viewer_filters = {}
        self.viewer_status_label = None
        self.viewer_access_granted = False  # Flag pour l'accès au visualiseur

        # Indicateurs de statut
        self.files_status = None
        self.analysis_status = None
        self.report_status = None

        # Dictionnaire pour stocker les affichages d'informations
        self.info_displays = {}

        # Variables de sélection
        self.collaborator_var = tk.StringVar()
        self.commune_var = tk.StringVar()
        self.insee_var = tk.StringVar()
        self.id_tache_var = tk.StringVar()

        # Données détectées
        self.detected_info = {}

        # Configuration Teams
        self.teams_folder_path = TeamsConfig.get_teams_base_path()

        # Configuration
        self.setup_ui()

        self.logger.info("Module Contrôle Qualité initialisé")
    
    def setup_ui(self):
        """Configure l'interface utilisateur avec onglets pour analyse et visualiseur."""
        try:
            # Frame principal avec style compact cohérent avec l'accueil
            self.main_frame = tk.Frame(self.parent, bg=COLORS['BG'])
            self.main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

            # Créer le notebook pour les onglets avec style compact
            style = ttk.Style()
            style.configure('Compact.TNotebook', tabposition='n')
            style.configure('Compact.TNotebook.Tab',
                          padding=[8, 4],  # Padding compact
                          font=('Segoe UI', 9))

            self.notebook = ttk.Notebook(self.main_frame, style='Compact.TNotebook')
            self.notebook.pack(fill=tk.BOTH, expand=True)

            # Onglet 1: Analyse Qualité
            self.analysis_tab = tk.Frame(self.notebook, bg=COLORS['BG'])
            self.notebook.add(self.analysis_tab, text="🔍 Analyse Qualité")

            # Onglet 2: Visualiseur
            self.viewer_tab = tk.Frame(self.notebook, bg=COLORS['BG'])
            self.notebook.add(self.viewer_tab, text="📊 Visualiseur")

            # Configurer l'onglet analyse
            self._setup_analysis_tab()

            # Configurer l'onglet visualiseur avec protection par mot de passe
            self._setup_viewer_tab()

            self.logger.info("Interface utilisateur créée avec onglets")

        except Exception as e:
            self.logger.error(f"Erreur lors de la création de l'interface: {e}")
            messagebox.showerror("Erreur", f"Erreur lors de la création de l'interface:\n{e}")

    def _setup_viewer_tab(self):
        """Configure l'onglet visualiseur de contrôle qualité avec protection par mot de passe."""
        try:
            # Créer l'interface de protection par mot de passe par défaut
            self._create_viewer_access_ui()

            self.logger.info("Onglet visualiseur configuré avec protection par mot de passe")

        except Exception as e:
            self.logger.error(f"Erreur configuration onglet visualiseur: {e}")

    def _create_viewer_access_ui(self):
        """Crée l'interface d'accès protégé pour le visualiseur."""
        # Clear any existing content
        for widget in self.viewer_tab.winfo_children():
            widget.destroy()

        # Main container with modern layout
        main_frame = tk.Frame(self.viewer_tab, bg=COLORS['BG'])
        main_frame.pack(fill=tk.BOTH, expand=True, padx=40, pady=40)

        # Center container with modern card design
        center_container = tk.Frame(main_frame, bg=COLORS['BG'])
        center_container.pack(expand=True)

        # Modern card with shadow effect
        card_frame = tk.Frame(center_container, bg=COLORS['CARD'], relief=tk.FLAT, bd=0)
        card_frame.pack(padx=20, pady=20)

        # Add subtle shadow effect
        shadow_frame = tk.Frame(center_container, bg=COLORS['BORDER'], height=2)
        shadow_frame.pack(fill=tk.X, padx=22, pady=(0, 2))

        # Content frame with padding
        content_frame = tk.Frame(card_frame, bg=COLORS['CARD'])
        content_frame.pack(padx=40, pady=30)

        # Icon and title section
        title_frame = tk.Frame(content_frame, bg=COLORS['CARD'])
        title_frame.pack(fill=tk.X, pady=(0, 20))

        # Security icon
        icon_label = tk.Label(
            title_frame,
            text="🔐",
            font=("Segoe UI", 32),
            bg=COLORS['CARD'],
            fg=COLORS['PRIMARY']
        )
        icon_label.pack()

        # Title
        title_label = tk.Label(
            title_frame,
            text="Accès Visualiseur Contrôle Qualité",
            font=UIConfig.FONT_HEADER,
            fg=COLORS['PRIMARY'],
            bg=COLORS['CARD']
        )
        title_label.pack(pady=(10, 0))

        # Subtitle
        subtitle_label = tk.Label(
            title_frame,
            text="Accès protégé par authentification",
            font=UIConfig.FONT_SUBTITLE,
            fg=COLORS['TEXT_SECONDARY'],
            bg=COLORS['CARD']
        )
        subtitle_label.pack(pady=(5, 0))

        # Message section
        message_frame = tk.Frame(content_frame, bg=COLORS['CARD'])
        message_frame.pack(fill=tk.X, pady=(0, 25))

        message_label = tk.Label(
            message_frame,
            text="Le visualiseur contient des données sensibles de contrôle qualité.\nVeuillez vous authentifier pour accéder aux fonctionnalités.",
            font=UIConfig.FONT_SUBTITLE,
            fg=COLORS['INFO'],
            bg=COLORS['CARD'],
            wraplength=450,
            justify=tk.CENTER
        )
        message_label.pack()

        # Action buttons
        buttons_frame = tk.Frame(content_frame, bg=COLORS['CARD'])
        buttons_frame.pack(fill=tk.X, pady=(0, 10))

        # Access button - primary action
        access_button = tk.Button(
            buttons_frame,
            text="🔑 Accéder au Visualiseur",
            font=UIConfig.FONT_BUTTON,
            bg=COLORS['PRIMARY'],
            fg=COLORS['WHITE'],
            relief=tk.FLAT,
            bd=0,
            padx=25,
            pady=12,
            command=self._request_viewer_access,
            cursor='hand2'
        )
        access_button.pack(side=tk.RIGHT, padx=(10, 0))

        # Add hover effects
        def on_access_enter(e):
            access_button.config(bg=COLORS['PRIMARY_DARK'])

        def on_access_leave(e):
            access_button.config(bg=COLORS['PRIMARY'])

        access_button.bind('<Enter>', on_access_enter)
        access_button.bind('<Leave>', on_access_leave)

    def _request_viewer_access(self):
        """Demande l'authentification pour accéder au visualiseur."""
        try:
            self.logger.info("Demande d'accès au visualiseur")

            # Show password dialog
            success, password = show_password_dialog(
                self.parent,
                title="🔐 Accès Visualiseur Contrôle Qualité",
                message="Ce visualiseur est protégé par mot de passe.\nVeuillez saisir le mot de passe pour continuer :"
            )

            if not success:
                self.logger.info("Authentification annulée par l'utilisateur")
                return

            # Verify password using the same system as statistics module
            if AccessControl.verify_stats_password(password):
                self.logger.info("Authentification réussie pour le visualiseur")
                self.viewer_access_granted = True
                self._create_viewer_interface()
            else:
                self.logger.warning("Tentative d'authentification échouée pour le visualiseur")
                messagebox.showerror(
                    "Accès Refusé",
                    "Accès refusé - mot de passe invalide"
                )

        except Exception as e:
            self.logger.error(f"Erreur lors de l'authentification du visualiseur: {e}")
            messagebox.showerror(
                "Erreur",
                f"Erreur lors de l'authentification:\n{e}"
            )

    def _create_viewer_interface(self):
        """Crée l'interface complète du visualiseur après authentification réussie."""
        try:
            # Clear the access UI
            for widget in self.viewer_tab.winfo_children():
                widget.destroy()

            # Layout en grille
            self.viewer_tab.grid_rowconfigure(0, weight=0)  # Header
            self.viewer_tab.grid_rowconfigure(1, weight=0)  # Filtres
            self.viewer_tab.grid_rowconfigure(2, weight=1)  # Tableau
            self.viewer_tab.grid_rowconfigure(3, weight=0)  # Status
            self.viewer_tab.grid_columnconfigure(0, weight=1)

            # Header du visualiseur
            self._create_viewer_header()

            # Section des filtres
            self._create_viewer_filters()

            # Tableau de données
            self._create_viewer_table()

            # Barre de statut
            self._create_viewer_status()

            # Charger les données initiales avec un délai pour s'assurer que l'interface est prête
            self.viewer_tab.after(500, self._load_viewer_data)

            self.logger.info("Interface du visualiseur créée avec succès")

        except Exception as e:
            self.logger.error(f"Erreur création interface visualiseur: {e}")
            messagebox.showerror("Erreur", f"Erreur lors de la création de l'interface:\n{e}")

    def _create_viewer_header(self):
        """Crée l'en-tête du visualiseur."""
        header_frame = tk.Frame(self.viewer_tab, bg=COLORS['PRIMARY'], height=45)
        header_frame.grid(row=0, column=0, sticky="ew", padx=2, pady=2)
        header_frame.grid_propagate(False)

        # Titre
        title_label = tk.Label(
            header_frame,
            text="📊 Visualiseur Contrôle Qualité",
            font=("Segoe UI", 14, "bold"),
            bg=COLORS['PRIMARY'],
            fg="white"
        )
        title_label.pack(side=tk.LEFT, padx=10, pady=10)

        # Bouton actualiser
        refresh_btn = tk.Button(
            header_frame,
            text="🔄 Actualiser",
            command=self._refresh_viewer_data,
            bg=COLORS['ACCENT'],
            fg="white",
            font=("Segoe UI", 9, "bold"),
            relief=tk.FLAT,
            padx=15
        )
        refresh_btn.pack(side=tk.RIGHT, padx=10, pady=8)



    def _create_viewer_filters(self):
        """Crée la section des filtres du visualiseur."""
        filter_frame = tk.Frame(self.viewer_tab, bg=COLORS['CARD'])
        filter_frame.grid(row=1, column=0, sticky="ew", padx=2, pady=2)

        # Variables de filtre
        self.viewer_filters = {
            'commune': tk.StringVar(),
            'domaine': tk.StringVar(),
            'affectation': tk.StringVar(),
            'controleur': tk.StringVar(),
            'statut_commune': tk.StringVar()
        }

        # Une seule ligne pour tous les filtres
        filters_row = tk.Frame(filter_frame, bg=COLORS['CARD'])
        filters_row.pack(fill=tk.X, padx=5, pady=2)

        # Filtre Commune
        tk.Label(filters_row, text="🏘️ Commune:", bg=COLORS['CARD'], font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(0,2))
        commune_entry = tk.Entry(filters_row, textvariable=self.viewer_filters['commune'], width=12)
        commune_entry.pack(side=tk.LEFT, padx=(0,8))

        # Filtre Domaine
        tk.Label(filters_row, text="🏢 Domaine:", bg=COLORS['CARD'], font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(0,2))
        domaine_combo = ttk.Combobox(filters_row, textvariable=self.viewer_filters['domaine'], width=10)
        domaine_combo.pack(side=tk.LEFT, padx=(0,8))

        # Filtre Affectation
        tk.Label(filters_row, text="👥 Affectation:", bg=COLORS['CARD'], font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(0,2))
        affectation_combo = ttk.Combobox(filters_row, textvariable=self.viewer_filters['affectation'], width=10)
        affectation_combo.pack(side=tk.LEFT, padx=(0,8))

        # Filtre Contrôleur
        tk.Label(filters_row, text="👤 Contrôleur:", bg=COLORS['CARD'], font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(0,2))
        controleur_combo = ttk.Combobox(filters_row, textvariable=self.viewer_filters['controleur'], width=12)
        controleur_combo.pack(side=tk.LEFT, padx=(0,8))

        # Filtre Statut Commune
        tk.Label(filters_row, text="📋 Statut:", bg=COLORS['CARD'], font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(0,2))
        statut_combo = ttk.Combobox(filters_row, textvariable=self.viewer_filters['statut_commune'], width=8)
        statut_combo['values'] = ('Tous', 'OK', 'KO')
        statut_combo.set('Tous')
        statut_combo.pack(side=tk.LEFT, padx=(0,8))

        # Bouton appliquer filtres
        apply_btn = tk.Button(
            filters_row,
            text="🔍 Filtrer",
            command=self._apply_viewer_filters,
            bg=COLORS['SUCCESS'],
            fg="white",
            font=("Segoe UI", 9),
            relief=tk.FLAT
        )
        apply_btn.pack(side=tk.RIGHT, padx=5)

        # Bind des événements pour filtrage automatique
        for var in self.viewer_filters.values():
            var.trace('w', lambda *args: self._apply_viewer_filters())

    def _create_viewer_table(self):
        """Crée le tableau de données du visualiseur."""
        try:
            table_frame = tk.Frame(self.viewer_tab, bg=COLORS['CARD'])
            table_frame.grid(row=2, column=0, sticky="nsew", padx=2, pady=2)

            # Titre du tableau
            table_title = tk.Label(
                table_frame,
                text="📋 Fichiers État de Lieu",
                font=("Segoe UI", 11, "bold"),
                bg=COLORS['CARD'],
                fg=COLORS['TEXT_PRIMARY']
            )
            table_title.pack(anchor=tk.W, padx=5, pady=2)

            # Container pour le treeview avec scrollbars
            tree_container = tk.Frame(table_frame, bg=COLORS['CARD'])
            tree_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=2)

            # Définir les colonnes dans l'ordre correct selon l'extraction
            columns = ('commune', 'id_tache', 'insee', 'domaine', 'affectation', 'controleur', 'score_total', 'statut_commune')
            column_names = {
                'commune': '🏘️ Commune',
                'id_tache': '🆔 ID Tâche PA',
                'insee': '📍 Code INSEE',
                'domaine': '🏢 Domaine',
                'affectation': '👥 Affectation',
                'controleur': '👤 Contrôleur',
                'score_total': '📊 Score Total',
                'statut_commune': '📋 Statut Commune'
            }

            # Créer le Treeview
            self.viewer_tree = ttk.Treeview(tree_container, columns=columns, show='headings', height=15)

            # Configurer les colonnes
            for col in columns:
                self.viewer_tree.heading(col, text=column_names[col])
                if col == 'commune':
                    self.viewer_tree.column(col, width=150, minwidth=120)
                elif col == 'id_tache':
                    self.viewer_tree.column(col, width=100, minwidth=80)
                elif col == 'insee':
                    self.viewer_tree.column(col, width=80, minwidth=70)
                elif col == 'domaine':
                    self.viewer_tree.column(col, width=120, minwidth=100)
                elif col == 'affectation':
                    self.viewer_tree.column(col, width=120, minwidth=100)
                elif col == 'controleur':
                    self.viewer_tree.column(col, width=120, minwidth=100)
                elif col == 'score_total':
                    self.viewer_tree.column(col, width=80, minwidth=70)
                elif col == 'statut_commune':
                    self.viewer_tree.column(col, width=120, minwidth=100)
                else:
                    self.viewer_tree.column(col, width=100, minwidth=80)

            # Scrollbars
            v_scrollbar = ttk.Scrollbar(tree_container, orient=tk.VERTICAL, command=self.viewer_tree.yview)
            h_scrollbar = ttk.Scrollbar(tree_container, orient=tk.HORIZONTAL, command=self.viewer_tree.xview)
            self.viewer_tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)

            # Placement
            self.viewer_tree.grid(row=0, column=0, sticky="nsew")
            v_scrollbar.grid(row=0, column=1, sticky="ns")
            h_scrollbar.grid(row=1, column=0, sticky="ew")

            tree_container.grid_rowconfigure(0, weight=1)
            tree_container.grid_columnconfigure(0, weight=1)

            # Bind double-click pour ouvrir fichier
            self.viewer_tree.bind('<Double-1>', self._on_viewer_double_click)

        except Exception as e:
            self.logger.error(f"Erreur création tableau visualiseur: {e}")
            # Créer un viewer_tree minimal en cas d'erreur
            self.viewer_tree = None

    def _create_viewer_status(self):
        """Crée la barre de statut du visualiseur."""
        status_frame = tk.Frame(self.viewer_tab, bg=COLORS['CARD'], height=30)
        status_frame.grid(row=3, column=0, sticky="ew", padx=2, pady=2)
        status_frame.grid_propagate(False)

        self.viewer_status_label = tk.Label(
            status_frame,
            text="📊 Prêt - 0 fichiers trouvés",
            font=("Segoe UI", 9),
            bg=COLORS['CARD'],
            fg=COLORS['TEXT_PRIMARY']
        )
        self.viewer_status_label.pack(side=tk.LEFT, padx=10, pady=5)

    def _load_viewer_data(self):
        """Charge les données du visualiseur depuis l'arborescence Teams."""
        try:
            # Vérifier l'accès au visualiseur
            if not self.viewer_access_granted:
                self.logger.warning("Tentative de chargement des données sans accès autorisé")
                return

            # Vérifier que l'interface existe
            if not hasattr(self, 'viewer_tree') or not self.viewer_tree:
                self._update_viewer_status("❌ Interface non initialisée")
                return

            self._update_viewer_status("🔄 Chargement des fichiers...")

            # Scanner l'arborescence Teams pour les fichiers état de lieu
            files_data = self._scan_quality_control_files()

            # Stocker les données
            self.viewer_data = files_data

            # Afficher dans le tableau
            self._populate_viewer_table(files_data)

            # Mettre à jour le statut
            count = len(files_data)
            self._update_viewer_status(f"📊 {count} fichier(s) trouvé(s)")

            # Debug: afficher les données chargées
            if files_data:
                self.logger.info("📝 Première entrée de données:")
                self.logger.info(f"   {files_data[0]}")

        except Exception as e:
            self.logger.error(f"❌ Erreur chargement données visualiseur: {e}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            self._update_viewer_status("❌ Erreur de chargement")
            messagebox.showerror("Erreur", f"Erreur lors du chargement:\n{e}")



    def _update_viewer_status(self, message: str):
        """Met à jour le statut du visualiseur."""
        try:
            if hasattr(self, 'viewer_status_label') and self.viewer_status_label:
                self.viewer_status_label.config(text=message)
        except Exception as e:
            self.logger.error(f"Erreur mise à jour statut: {e}")

    def _scan_quality_control_files(self) -> List[Dict[str, Any]]:
        """Scanne l'arborescence Teams pour trouver les fichiers état de lieu."""
        files_data = []

        try:
            # Chemin de base du contrôle qualité
            qc_base_path = TeamsConfig.get_quality_control_teams_path()

            self.logger.info(f"🔍 Début scan - Chemin base: {qc_base_path}")

            if not os.path.exists(qc_base_path):
                self.logger.warning(f"❌ Dossier contrôle qualité non trouvé: {qc_base_path}")
                messagebox.showwarning("Dossier non trouvé",
                                     f"Le dossier Contrôle Qualité n'existe pas:\n{qc_base_path}\n\n"
                                     f"Veuillez créer au moins un fichier état de lieu d'abord.")
                return files_data

            # Lister le contenu du dossier base
            base_contents = os.listdir(qc_base_path)
            self.logger.info(f"📁 Contenu dossier base: {base_contents}")

            if not base_contents:
                self.logger.info("📂 Dossier base vide")
                messagebox.showinfo("Aucun fichier",
                                  f"Le dossier Contrôle Qualité est vide.\n"
                                  f"Générez d'abord des fichiers état de lieu.")
                return files_data

            # Scanner chaque dossier collaborateur
            for collaborateur_folder in base_contents:
                collaborateur_path = os.path.join(qc_base_path, collaborateur_folder)

                self.logger.info(f"👤 Vérification collaborateur: {collaborateur_folder}")

                if not os.path.isdir(collaborateur_path):
                    self.logger.info(f"⚠️ Ignoré (pas un dossier): {collaborateur_folder}")
                    continue

                # Lister le contenu du dossier collaborateur
                collab_contents = os.listdir(collaborateur_path)
                self.logger.info(f"📁 Contenu {collaborateur_folder}: {collab_contents}")

                # Scanner chaque dossier commune du collaborateur
                for commune_folder in collab_contents:
                    commune_path = os.path.join(collaborateur_path, commune_folder)

                    self.logger.info(f"🏘️ Vérification commune: {commune_folder}")

                    if not os.path.isdir(commune_path):
                        self.logger.info(f"⚠️ Ignoré (pas un dossier): {commune_folder}")
                        continue

                    # Lister le contenu du dossier commune
                    commune_contents = os.listdir(commune_path)
                    self.logger.info(f"📁 Contenu {commune_folder}: {commune_contents}")

                    # Chercher les fichiers Excel état de lieu
                    for file_name in commune_contents:
                        self.logger.info(f"📄 Vérification fichier: {file_name}")

                        if file_name.startswith("Etat_De_Lieu_") and file_name.endswith(".xlsx"):
                            file_path = os.path.join(commune_path, file_name)
                            self.logger.info(f"✅ Fichier état de lieu trouvé: {file_path}")

                            # Extraire les données depuis le fichier Excel
                            file_info = self._extract_file_data(file_path)

                            if file_info:
                                files_data.append(file_info)
                                self.logger.info(f"✅ Fichier ajouté: {file_info}")
                            else:
                                self.logger.warning(f"❌ Échec extraction info: {file_name}")
                        else:
                            self.logger.info(f"⚠️ Fichier ignoré (ne correspond pas au pattern): {file_name}")

            self.logger.info(f"🎯 Scan terminé: {len(files_data)} fichiers trouvés")

            if len(files_data) == 0:
                messagebox.showinfo("Aucun fichier trouvé",
                                  f"Aucun fichier 'Etat_De_Lieu_*.xlsx' trouvé dans:\n{qc_base_path}\n\n"
                                  f"Vérifiez que les fichiers sont bien nommés et placés dans la bonne arborescence.")

            return files_data

        except Exception as e:
            self.logger.error(f"❌ Erreur scan fichiers QC: {e}")
            messagebox.showerror("Erreur", f"Erreur lors du scan des fichiers:\n{e}")
            return files_data



    def _populate_viewer_table(self, files_data: List[Dict[str, Any]]):
        """Remplit le tableau avec les données des fichiers."""
        try:
            # Vérifier que le tree existe
            if not hasattr(self, 'viewer_tree') or not self.viewer_tree:
                return

            # Vider le tableau
            for item in self.viewer_tree.get_children():
                self.viewer_tree.delete(item)

            # Ajouter les données
            for file_info in files_data:
                try:
                    # Les données extraites sont dans l'ordre correct selon les positions Excel
                    # Mais nous devons les réorganiser pour correspondre aux en-têtes de colonnes
                    values = (
                        file_info.get('commune', 'N/A'),      # B3 -> Commune
                        file_info.get('id_tache', 'N/A'),     # C3 -> ID Tâche
                        file_info.get('insee', 'N/A'),        # D3 -> INSEE
                        file_info.get('domaine', 'N/A'),      # E3 -> Domaine
                        file_info.get('affectation', 'N/A'),  # F3 -> Affectation
                        file_info.get('controleur', 'N/A'),   # G3 -> Contrôleur
                        file_info.get('score_total', 'N/A'),  # K11 -> Score Total
                        file_info.get('statut_commune', 'N/A') # L11 -> Statut Commune
                    )

                    self.viewer_tree.insert('', 'end', values=values)

                except Exception as e:
                    self.logger.error(f"Erreur ajout ligne: {e}")

        except Exception as e:
            self.logger.error(f"Erreur remplissage tableau: {e}")

        # Générer automatiquement le rapport Excel en arrière-plan
        self._generate_tracking_report_async(files_data)

    def _extract_file_data(self, file_path: str) -> Dict[str, Any]:
        """Extrait les données depuis un fichier Excel état de lieu."""
        try:
            pd = get_pandas()

            # Données par défaut
            file_data = {
                'commune': 'N/A',
                'id_tache': 'N/A',
                'insee': 'N/A',
                'domaine': 'N/A',
                'affectation': 'N/A',
                'controleur': 'N/A',
                'score_total': 'N/A',
                'statut_commune': 'N/A',
                'chemin': file_path
            }

            # Lire la première feuille (index 0) qui contient INFORMATIONS GÉNÉRALES
            try:
                df = pd.read_excel(file_path, sheet_name=0, header=None)

                # Extraire les données selon les positions exactes
                # Section INFORMATIONS GÉNÉRALES (ligne 3, index 2)
                if len(df) > 2:  # Vérifier qu'on a au moins 3 lignes
                    row_3 = df.iloc[2]  # Ligne 3 (index 2)

                    # Debug temporaire pour voir le contenu exact
                    self.logger.info(f"🔍 Contenu ligne 3: {row_3.tolist()}")

                    # Nom de commune (B3 = colonne 1, mais vérifions si c'est A3 = colonne 0)
                    if len(row_3) > 0 and not pd.isna(row_3.iloc[0]) and str(row_3.iloc[0]).strip() != 'INFORMATIONS GÉNÉRALES':
                        file_data['commune'] = str(row_3.iloc[0]).strip()
                    elif len(row_3) > 1 and not pd.isna(row_3.iloc[1]):
                        file_data['commune'] = str(row_3.iloc[1]).strip()

                    # ID tâche Plan Adressage (C3 = colonne 2, mais peut-être B3 = colonne 1)
                    if len(row_3) > 1 and not pd.isna(row_3.iloc[1]) and str(row_3.iloc[1]).strip().isdigit():
                        file_data['id_tache'] = str(row_3.iloc[1]).strip()
                    elif len(row_3) > 2 and not pd.isna(row_3.iloc[2]):
                        file_data['id_tache'] = str(row_3.iloc[2]).strip()

                    # Code INSEE (D3 = colonne 3, mais peut-être C3 = colonne 2)
                    if len(row_3) > 2 and not pd.isna(row_3.iloc[2]) and str(row_3.iloc[2]).strip().isdigit():
                        file_data['insee'] = str(row_3.iloc[2]).strip()
                    elif len(row_3) > 3 and not pd.isna(row_3.iloc[3]):
                        file_data['insee'] = str(row_3.iloc[3]).strip()

                    # Domaine (E3 = colonne 4, mais peut-être D3 = colonne 3)
                    if len(row_3) > 3 and not pd.isna(row_3.iloc[3]):
                        file_data['domaine'] = str(row_3.iloc[3]).strip()
                    elif len(row_3) > 4 and not pd.isna(row_3.iloc[4]):
                        file_data['domaine'] = str(row_3.iloc[4]).strip()

                    # AFFECTATION (F3 = colonne 5, mais peut-être E3 = colonne 4)
                    if len(row_3) > 4 and not pd.isna(row_3.iloc[4]):
                        file_data['affectation'] = str(row_3.iloc[4]).strip()
                    elif len(row_3) > 5 and not pd.isna(row_3.iloc[5]):
                        file_data['affectation'] = str(row_3.iloc[5]).strip()

                    # Contrôleur (G3 = colonne 6, mais peut-être F3 = colonne 5)
                    if len(row_3) > 5 and not pd.isna(row_3.iloc[5]):
                        file_data['controleur'] = str(row_3.iloc[5]).strip()
                    elif len(row_3) > 6 and not pd.isna(row_3.iloc[6]):
                        file_data['controleur'] = str(row_3.iloc[6]).strip()

                # Section SCORE TOTAL (ligne 11, index 10)
                if len(df) > 10:  # Vérifier qu'on a au moins 11 lignes
                    row_11 = df.iloc[10]  # Ligne 11 (index 10)

                    # Debug temporaire pour voir le contenu exact
                    self.logger.info(f"🔍 Contenu ligne 11: {row_11.tolist()}")

                    # Score Total - chercher dans les colonnes J, K, L (indices 9, 10, 11)
                    score_found = False
                    for col_idx in [9, 10, 11]:
                        if len(row_11) > col_idx and not pd.isna(row_11.iloc[col_idx]):
                            val = str(row_11.iloc[col_idx]).strip()
                            try:
                                # Convertir en float pour vérifier si c'est un nombre
                                float_val = float(val)
                                # Si c'est un nombre décimal entre 0 et 1, le convertir en pourcentage
                                if 0 <= float_val <= 1:
                                    percentage = float_val * 100
                                    file_data['score_total'] = f"{percentage:.2f}%"
                                else:
                                    # Si c'est déjà un nombre entier (comme 1, 2, etc.), le garder tel quel
                                    file_data['score_total'] = val
                                score_found = True
                                break
                            except ValueError:
                                # Si ce n'est pas un nombre, le garder tel quel
                                if val.isdigit():
                                    file_data['score_total'] = val
                                    score_found = True
                                    break

                    # Statut Commune - chercher dans les colonnes K, L, M (indices 10, 11, 12)
                    statut_found = False
                    for col_idx in [10, 11, 12]:
                        if len(row_11) > col_idx and not pd.isna(row_11.iloc[col_idx]):
                            val = str(row_11.iloc[col_idx]).strip()
                            if val.upper() in ['OK', 'KO', 'NOK']:
                                file_data['statut_commune'] = val.upper()  # Normaliser en majuscules
                                statut_found = True
                                break

            except Exception as e:
                self.logger.warning(f"Erreur lecture fichier Excel: {e}")

            return file_data

        except Exception as e:
            self.logger.error(f"Erreur extraction données fichier {file_path}: {e}")
            return {
                'commune': 'Erreur',
                'id_tache': 'Erreur',
                'insee': 'Erreur',
                'domaine': 'Erreur',
                'affectation': 'Erreur',
                'controleur': 'Erreur',
                'score_total': 'Erreur',
                'statut_commune': 'Erreur',
                'chemin': file_path
            }

    def _apply_viewer_filters(self):
        """Applique les filtres au tableau du visualiseur."""
        try:
            # Vérifier l'accès au visualiseur
            if not self.viewer_access_granted:
                return

            if not self.viewer_data:
                return

            # Récupérer les valeurs des filtres
            commune_filter = self.viewer_filters['commune'].get().strip().lower()
            domaine_filter = self.viewer_filters['domaine'].get().strip()
            affectation_filter = self.viewer_filters['affectation'].get().strip()
            controleur_filter = self.viewer_filters['controleur'].get().strip()
            statut_filter = self.viewer_filters['statut_commune'].get().strip()

            # Filtrer les données
            filtered_data = []
            for file_info in self.viewer_data:
                # Filtre commune
                if commune_filter and commune_filter not in file_info.get('commune', '').lower():
                    continue

                # Filtre domaine
                if domaine_filter and domaine_filter != 'Tous' and domaine_filter != file_info.get('domaine', ''):
                    continue

                # Filtre affectation
                if affectation_filter and affectation_filter != 'Tous' and affectation_filter != file_info.get('affectation', ''):
                    continue

                # Filtre contrôleur
                if controleur_filter and controleur_filter != 'Tous' and controleur_filter != file_info.get('controleur', ''):
                    continue

                # Filtre statut commune
                if statut_filter and statut_filter != 'Tous' and statut_filter != file_info.get('statut_commune', ''):
                    continue

                filtered_data.append(file_info)

            # Mettre à jour le tableau
            self._populate_viewer_table(filtered_data)

            # Mettre à jour le statut
            count = len(filtered_data)
            total = len(self.viewer_data)
            self._update_viewer_status(f"📊 {count}/{total} fichier(s) affiché(s)")

        except Exception as e:
            self.logger.error(f"Erreur application filtres: {e}")

    def _refresh_viewer_data(self):
        """Actualise les données du visualiseur."""
        # Vérifier l'accès au visualiseur
        if not self.viewer_access_granted:
            return
        self._load_viewer_data()

    def _on_viewer_double_click(self, event):
        """Gère le double-clic sur une ligne du tableau pour ouvrir le fichier."""
        try:
            # Vérifier l'accès au visualiseur
            if not self.viewer_access_granted:
                return

            selection = self.viewer_tree.selection()
            if not selection:
                return

            item = selection[0]
            values = self.viewer_tree.item(item, 'values')

            if not values or len(values) < 3:
                return

            # Récupérer les informations pour identifier le fichier
            commune = values[0]  # Nom de commune
            id_tache = values[1]  # ID tâche
            insee = values[2]    # Code INSEE

            # Chercher le fichier correspondant dans les données
            file_path = None
            if hasattr(self, 'viewer_data') and self.viewer_data:
                for file_info in self.viewer_data:
                    if (file_info.get('commune') == commune and
                        file_info.get('id_tache') == id_tache and
                        file_info.get('insee') == insee):
                        file_path = file_info.get('chemin')
                        break

            if file_path and os.path.exists(file_path):
                # Ouvrir le fichier Excel
                os.startfile(file_path)
                self._update_viewer_status(f"📂 Ouverture: {os.path.basename(file_path)}")
                self.logger.info(f"Fichier ouvert: {file_path}")
            else:
                messagebox.showerror("Erreur", f"Fichier non trouvé pour:\nCommune: {commune}\nID: {id_tache}")

        except Exception as e:
            self.logger.error(f"Erreur ouverture fichier: {e}")
            messagebox.showerror("Erreur", f"Erreur lors de l'ouverture:\n{e}")

    def _debug_viewer_paths(self):
        """Fonction de diagnostic pour vérifier les chemins et l'arborescence."""
        try:
            debug_info = []

            # 1. Vérifier le chemin de base Teams
            qc_base_path = TeamsConfig.get_quality_control_teams_path()
            debug_info.append(f"🔍 Chemin base Teams:")
            debug_info.append(f"   {qc_base_path}")
            debug_info.append(f"   Existe: {os.path.exists(qc_base_path)}")

            if os.path.exists(qc_base_path):
                # 2. Lister le contenu du dossier base
                try:
                    base_contents = os.listdir(qc_base_path)
                    debug_info.append(f"\n📁 Contenu dossier base ({len(base_contents)} éléments):")
                    for item in base_contents[:10]:  # Limiter à 10 pour éviter un message trop long
                        item_path = os.path.join(qc_base_path, item)
                        item_type = "📁" if os.path.isdir(item_path) else "📄"
                        debug_info.append(f"   {item_type} {item}")

                    if len(base_contents) > 10:
                        debug_info.append(f"   ... et {len(base_contents) - 10} autres")

                except Exception as e:
                    debug_info.append(f"   ❌ Erreur lecture: {e}")

            # 3. Vérifier les patterns de fichiers
            debug_info.append(f"\n🔍 Recherche de fichiers 'Etat_De_Lieu_*.xlsx':")

            if os.path.exists(qc_base_path):
                found_files = []
                for root, dirs, files in os.walk(qc_base_path):
                    for file in files:
                        if file.startswith("Etat_De_Lieu_") and file.endswith(".xlsx"):
                            rel_path = os.path.relpath(root, qc_base_path)
                            found_files.append(f"   📄 {rel_path}\\{file}")

                if found_files:
                    debug_info.append(f"   Trouvé {len(found_files)} fichier(s):")
                    debug_info.extend(found_files[:5])  # Limiter à 5
                    if len(found_files) > 5:
                        debug_info.append(f"   ... et {len(found_files) - 5} autres")
                else:
                    debug_info.append("   ❌ Aucun fichier trouvé")

            # 4. Afficher les informations
            debug_message = "\n".join(debug_info)

            # Créer une fenêtre de diagnostic
            debug_window = tk.Toplevel(self.parent)
            debug_window.title("🔧 Diagnostic Visualiseur")
            debug_window.geometry("800x600")
            debug_window.configure(bg=COLORS['BG'])

            # Zone de texte avec scrollbar
            text_frame = tk.Frame(debug_window, bg=COLORS['BG'])
            text_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

            text_widget = tk.Text(text_frame, wrap=tk.WORD, font=("Consolas", 10))
            scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=text_widget.yview)
            text_widget.configure(yscrollcommand=scrollbar.set)

            text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

            text_widget.insert(tk.END, debug_message)
            text_widget.config(state=tk.DISABLED)

            # Bouton fermer
            close_btn = tk.Button(
                debug_window,
                text="Fermer",
                command=debug_window.destroy,
                bg=COLORS['ACCENT'],
                fg="white",
                font=("Segoe UI", 10, "bold")
            )
            close_btn.pack(pady=10)

            self.logger.info("Diagnostic visualiseur affiché")

        except Exception as e:
            self.logger.error(f"Erreur diagnostic: {e}")
            messagebox.showerror("Erreur", f"Erreur lors du diagnostic:\n{e}")

    def _setup_analysis_tab(self):
        """Configure l'onglet d'analyse qualité."""
        try:
            # Layout en grille pour maximiser l'espace
            self.analysis_tab.grid_rowconfigure(0, weight=0)  # Header compact
            self.analysis_tab.grid_rowconfigure(1, weight=1)  # Contenu principal
            self.analysis_tab.grid_rowconfigure(2, weight=0)  # Status compact
            self.analysis_tab.grid_columnconfigure(0, weight=1)

            # Interface utilisateur modernisée et améliorée
            self._create_enhanced_header()

            # Contenu principal en grille 2x2 avec design amélioré
            self._create_enhanced_main_content()

            # Barre de statut modernisée
            self._create_enhanced_status_bar()

            self.logger.info("Interface utilisateur modernisée créée avec succès")

        except Exception as e:
            self.logger.error(f"Erreur lors de la création de l'interface QC: {e}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            messagebox.showerror("Erreur", f"Impossible de créer l'interface du module:\n{e}")

    def _create_ultra_compact_header(self):
        """Crée un header ultra-compact en une seule ligne."""
        header_frame = tk.Frame(self.main_frame, bg=COLORS['CARD'], height=35)
        header_frame.grid(row=0, column=0, sticky="ew", padx=2, pady=1)
        header_frame.pack_propagate(False)
        header_frame.config(highlightbackground=COLORS['BORDER'], highlightthickness=1)

        # Contenu en ligne horizontale
        content = tk.Frame(header_frame, bg=COLORS['CARD'])
        content.pack(fill=tk.BOTH, expand=True, padx=8, pady=3)

        # Icône et titre
        tk.Label(content, text="🔍", font=("Segoe UI", 12),
                fg=COLORS['PRIMARY'], bg=COLORS['CARD']).pack(side=tk.LEFT)

        tk.Label(content, text="Module 5 - Contrôle Qualité",
                font=("Segoe UI", 10, "bold"), fg=COLORS['PRIMARY'],
                bg=COLORS['CARD']).pack(side=tk.LEFT, padx=(5, 15))

        # Indicateurs de statut en ligne
        self._create_inline_status_indicators(content)

    def _create_inline_status_indicators(self, parent: tk.Widget):
        """Crée les indicateurs de statut en ligne."""
        # Séparateur
        tk.Label(parent, text="|", font=("Segoe UI", 10),
                fg=COLORS['BORDER'], bg=COLORS['CARD']).pack(side=tk.LEFT, padx=5)

        # Indicateur fichiers
        self.files_status = tk.Label(parent, text="📁 Fichiers: En attente",
                                   font=("Segoe UI", 8), fg=COLORS['INFO'],
                                   bg=COLORS['CARD'])
        self.files_status.pack(side=tk.LEFT, padx=5)

        # Indicateur analyse
        self.analysis_status = tk.Label(parent, text="🔍 Analyse: Non effectuée",
                                      font=("Segoe UI", 8), fg=COLORS['TEXT_SECONDARY'],
                                      bg=COLORS['CARD'])
        self.analysis_status.pack(side=tk.LEFT, padx=5)

        # Indicateur rapport
        self.report_status = tk.Label(parent, text="📊 Rapport: Non généré",
                                    font=("Segoe UI", 8), fg=COLORS['TEXT_SECONDARY'],
                                    bg=COLORS['CARD'])
        self.report_status.pack(side=tk.LEFT, padx=5)

    def _create_compact_main_content(self):
        """Crée le contenu principal en grille 2x2 ultra-compacte."""
        main_content = tk.Frame(self.main_frame, bg=COLORS['BG'])
        main_content.grid(row=1, column=0, sticky="nsew", padx=2, pady=1)

        # Configuration de la grille 2x2
        main_content.grid_rowconfigure(0, weight=1)
        main_content.grid_rowconfigure(1, weight=1)
        main_content.grid_columnconfigure(0, weight=1)
        main_content.grid_columnconfigure(1, weight=1)

        # Quadrant 1: Fichiers (haut gauche)
        self._create_files_quadrant(main_content, 0, 0)

        # Quadrant 2: Informations détectées (haut droite)
        self._create_info_quadrant(main_content, 0, 1)

        # Quadrant 3: Analyse et critères (bas gauche)
        self._create_analysis_quadrant(main_content, 1, 0)

        # Quadrant 4: Résultats (bas droite)
        self._create_results_quadrant(main_content, 1, 1)

    def _create_files_quadrant(self, parent: tk.Widget, row: int, col: int):
        """Quadrant 1: Chargement des fichiers ultra-compact."""
        frame = tk.Frame(parent, bg=COLORS['CARD'], relief='flat', bd=1)
        frame.grid(row=row, column=col, sticky="nsew", padx=1, pady=1)
        frame.config(highlightbackground=COLORS['BORDER'], highlightthickness=1)

        # Titre compact
        title_frame = tk.Frame(frame, bg=COLORS['ACCENT'], height=25)
        title_frame.pack(fill=tk.X)
        title_frame.pack_propagate(False)

        tk.Label(title_frame, text="📁 Chargement Fichiers",
                font=("Segoe UI", 9, "bold"), fg=COLORS['PRIMARY'],
                bg=COLORS['ACCENT']).pack(pady=3)

        # Contenu
        content = tk.Frame(frame, bg=COLORS['CARD'])
        content.pack(fill=tk.BOTH, expand=True, padx=5, pady=3)

        # QGis file - ultra compact
        qgis_frame = tk.Frame(content, bg=COLORS['CARD'])
        qgis_frame.pack(fill=tk.X, pady=1)

        tk.Label(qgis_frame, text="🗺️ QGis:", font=("Segoe UI", 8, "bold"),
                fg=COLORS['TEXT_PRIMARY'], bg=COLORS['CARD']).pack(anchor=tk.W)

        from tkinter import ttk
        self.load_qgis_button = ttk.Button(qgis_frame, text="📂 Importer QGis",
                                         command=self._load_qgis_file, style='Compact.TButton')
        self.load_qgis_button.pack(anchor=tk.W, pady=1)

        self.qgis_info_label = tk.Label(qgis_frame, text="❌ Non chargé",
                                      font=("Segoe UI", 7), fg=COLORS['TEXT_SECONDARY'],
                                      bg=COLORS['CARD'])
        self.qgis_info_label.pack(anchor=tk.W)

        # Séparateur mini
        tk.Frame(content, height=1, bg=COLORS['BORDER']).pack(fill=tk.X, pady=2)

        # Suivi file - ultra compact
        suivi_frame = tk.Frame(content, bg=COLORS['CARD'])
        suivi_frame.pack(fill=tk.X, pady=1)

        tk.Label(suivi_frame, text="📋 Suivi:", font=("Segoe UI", 8, "bold"),
                fg=COLORS['TEXT_PRIMARY'], bg=COLORS['CARD']).pack(anchor=tk.W)

        self.load_suivi_button = ttk.Button(suivi_frame, text="📂 Importer Suivi",
                                          command=self._load_suivi_file, style='Compact.TButton')
        self.load_suivi_button.pack(anchor=tk.W, pady=1)

        self.suivi_info_label = tk.Label(suivi_frame, text="❌ Non chargé",
                                       font=("Segoe UI", 7), fg=COLORS['TEXT_SECONDARY'],
                                       bg=COLORS['CARD'])
        self.suivi_info_label.pack(anchor=tk.W)

    def _create_info_quadrant(self, parent: tk.Widget, row: int, col: int):
        """Quadrant 2: Informations détectées ultra-compact."""
        frame = tk.Frame(parent, bg=COLORS['CARD'], relief='flat', bd=1)
        frame.grid(row=row, column=col, sticky="nsew", padx=1, pady=1)
        frame.config(highlightbackground=COLORS['BORDER'], highlightthickness=1)

        # Titre compact
        title_frame = tk.Frame(frame, bg=COLORS['ACCENT'], height=25)
        title_frame.pack(fill=tk.X)
        title_frame.pack_propagate(False)

        tk.Label(title_frame, text="📋 Informations Détectées",
                font=("Segoe UI", 9, "bold"), fg=COLORS['PRIMARY'],
                bg=COLORS['ACCENT']).pack(pady=3)

        # Contenu en grille 2x2
        content = tk.Frame(frame, bg=COLORS['CARD'])
        content.pack(fill=tk.BOTH, expand=True, padx=5, pady=3)
        content.grid_columnconfigure(0, weight=1)
        content.grid_columnconfigure(1, weight=1)

        # Ligne 1
        self._create_mini_info_field(content, 0, 0, "👤", "Collaborateur", self.collaborator_var)
        self._create_mini_info_field(content, 0, 1, "🏘️", "Commune", self.commune_var)

        # Ligne 2
        self._create_mini_info_field(content, 1, 0, "🏛️", "INSEE", self.insee_var)
        self._create_mini_info_field(content, 1, 1, "🆔", "ID Tâche", self.id_tache_var)

        # Note ultra-compacte
        note = tk.Label(content, text="ℹ️ Auto-détection depuis fichier suivi",
                       font=("Segoe UI", 7), fg=COLORS['INFO'], bg=COLORS['CARD'])
        note.grid(row=2, column=0, columnspan=2, sticky="w", pady=(3, 0))

    def _create_mini_info_field(self, parent: tk.Widget, row: int, col: int,
                               icon: str, label: str, var: tk.StringVar):
        """Crée un champ d'information mini."""
        field_frame = tk.Frame(parent, bg=COLORS['CARD'])
        field_frame.grid(row=row, column=col, sticky="ew", padx=2, pady=1)

        # Header en ligne
        header = tk.Frame(field_frame, bg=COLORS['CARD'])
        header.pack(fill=tk.X)

        tk.Label(header, text=icon, font=("Segoe UI", 8),
                fg=COLORS['SECONDARY'], bg=COLORS['CARD']).pack(side=tk.LEFT)

        tk.Label(header, text=f"{label}:", font=("Segoe UI", 7, "bold"),
                fg=COLORS['TEXT_PRIMARY'], bg=COLORS['CARD']).pack(side=tk.LEFT, padx=(2, 0))

        # Valeur - Texte en noir
        value_label = tk.Label(field_frame, textvariable=var, font=("Segoe UI", 7),
                              fg=COLORS['TEXT_PRIMARY'], bg=COLORS['CARD'], anchor='w')  # Changé de INFO à TEXT_PRIMARY
        value_label.pack(fill=tk.X)

        # Stocker pour mise à jour
        if not hasattr(self, 'info_displays'):
            self.info_displays = {}
        self.info_displays[label.lower()] = value_label

    def _create_analysis_quadrant(self, parent: tk.Widget, row: int, col: int):
        """Quadrant 3: Analyse et critères ultra-compact."""
        frame = tk.Frame(parent, bg=COLORS['CARD'], relief='flat', bd=1)
        frame.grid(row=row, column=col, sticky="nsew", padx=1, pady=1)
        frame.config(highlightbackground=COLORS['BORDER'], highlightthickness=1)

        # Titre compact
        title_frame = tk.Frame(frame, bg=COLORS['ACCENT'], height=25)
        title_frame.pack(fill=tk.X)
        title_frame.pack_propagate(False)

        tk.Label(title_frame, text="🔍 Analyse Qualité",
                font=("Segoe UI", 9, "bold"), fg=COLORS['PRIMARY'],
                bg=COLORS['ACCENT']).pack(pady=3)

        # Contenu
        content = tk.Frame(frame, bg=COLORS['CARD'])
        content.pack(fill=tk.BOTH, expand=True, padx=5, pady=3)

        # Critères ultra-compacts
        criteria_frame = tk.Frame(content, bg=COLORS['LIGHT'], relief='flat', bd=1)
        criteria_frame.pack(fill=tk.X, pady=(0, 3))
        criteria_frame.config(highlightbackground=COLORS['BORDER'], highlightthickness=1)

        criteria_content = tk.Frame(criteria_frame, bg=COLORS['LIGHT'])
        criteria_content.pack(fill=tk.X, padx=3, pady=2)

        # Critère 0
        c0_frame = tk.Frame(criteria_content, bg=COLORS['LIGHT'])
        c0_frame.pack(fill=tk.X, pady=1)

        tk.Label(c0_frame, text="0", font=("Segoe UI", 7, "bold"),
                fg='white', bg=COLORS['SECONDARY'], padx=3, pady=1).pack(side=tk.LEFT)
        tk.Label(c0_frame, text="Incohérences fichiers", font=("Segoe UI", 7),
                fg=COLORS['TEXT_PRIMARY'], bg=COLORS['LIGHT']).pack(side=tk.LEFT, padx=(3, 0))

        # Critère 1
        c1_frame = tk.Frame(criteria_content, bg=COLORS['LIGHT'])
        c1_frame.pack(fill=tk.X, pady=1)

        tk.Label(c1_frame, text="1", font=("Segoe UI", 7, "bold"),
                fg='white', bg=COLORS['SECONDARY'], padx=3, pady=1).pack(side=tk.LEFT)
        tk.Label(c1_frame, text="Doublons IMB", font=("Segoe UI", 7),
                fg=COLORS['TEXT_PRIMARY'], bg=COLORS['LIGHT']).pack(side=tk.LEFT, padx=(3, 0))

        # Critère 3
        c3_frame = tk.Frame(criteria_content, bg=COLORS['LIGHT'])
        c3_frame.pack(fill=tk.X, pady=1)
        tk.Label(c3_frame, text="CRITÈRE 3", font=("Segoe UI", 7, "bold"),
                fg='white', bg=COLORS['SECONDARY'], padx=3, pady=1).pack(side=tk.LEFT)
        tk.Label(c3_frame, text="Contrôle IMB Doublons", font=("Segoe UI", 7),
                fg=COLORS['TEXT_PRIMARY'], bg=COLORS['LIGHT']).pack(side=tk.LEFT, padx=(3, 0))

        # Critère 4
        c4_frame = tk.Frame(criteria_content, bg=COLORS['LIGHT'])
        c4_frame.pack(fill=tk.X, pady=1)
        tk.Label(c4_frame, text="CRITÈRE 4", font=("Segoe UI", 7, "bold"),
                fg='white', bg=COLORS['SECONDARY'], padx=3, pady=1).pack(side=tk.LEFT)
        tk.Label(c4_frame, text="AD à Analyser", font=("Segoe UI", 7),
                fg=COLORS['TEXT_PRIMARY'], bg=COLORS['LIGHT']).pack(side=tk.LEFT, padx=(3, 0))

        # Boutons d'action
        buttons_frame = tk.Frame(content, bg=COLORS['CARD'])
        buttons_frame.pack(fill=tk.X, pady=(3, 0))

        from tkinter import ttk
        self.analyze_button = ttk.Button(buttons_frame, text="🔍 Analyser",
                                       command=self._run_quality_analysis,
                                       style='CompactWarning.TButton', state='disabled')
        self.analyze_button.pack(side=tk.LEFT, padx=(0, 3))

        self.export_button = ttk.Button(buttons_frame, text="📤 Exporter",
                                      command=self._export_qc_report,
                                      style='Compact.TButton', state='disabled')
        self.export_button.pack(side=tk.LEFT)

    def _create_results_quadrant(self, parent: tk.Widget, row: int, col: int):
        """Quadrant 4: Résultats ultra-compact."""
        frame = tk.Frame(parent, bg=COLORS['CARD'], relief='flat', bd=1)
        frame.grid(row=row, column=col, sticky="nsew", padx=1, pady=1)
        frame.config(highlightbackground=COLORS['BORDER'], highlightthickness=1)

        # Titre compact
        title_frame = tk.Frame(frame, bg=COLORS['ACCENT'], height=25)
        title_frame.pack(fill=tk.X)
        title_frame.pack_propagate(False)

        tk.Label(title_frame, text="📊 Résultats",
                font=("Segoe UI", 9, "bold"), fg=COLORS['PRIMARY'],
                bg=COLORS['ACCENT']).pack(pady=3)

        # Zone de résultats
        self.results_frame = tk.Frame(frame, bg=COLORS['CARD'])
        self.results_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=3)

        # Message initial ultra-compact
        self.results_label = tk.Label(
            self.results_frame,
            text="⏳ En attente\n\nChargez les fichiers\net lancez l'analyse",
            font=("Segoe UI", 8),
            fg=COLORS['TEXT_SECONDARY'],
            bg=COLORS['CARD'],
            justify=tk.CENTER
        )
        self.results_label.pack(expand=True)

    def _create_ultra_compact_status(self):
        """Crée la barre de statut ultra-compacte."""
        status_frame = tk.Frame(self.main_frame, bg=COLORS['LIGHT'], height=25)
        status_frame.grid(row=2, column=0, sticky="ew", padx=2, pady=1)
        status_frame.pack_propagate(False)
        status_frame.config(highlightbackground=COLORS['BORDER'], highlightthickness=1)

        # Contenu en ligne
        content = tk.Frame(status_frame, bg=COLORS['LIGHT'])
        content.pack(fill=tk.BOTH, expand=True, padx=5, pady=2)

        # Progress bar mini
        from tkinter import ttk
        self.progress_var = tk.DoubleVar()

        tk.Label(content, text="Progression:", font=("Segoe UI", 8),
                fg=COLORS['TEXT_PRIMARY'], bg=COLORS['LIGHT']).pack(side=tk.LEFT)

        self.progress_bar = ttk.Progressbar(content, variable=self.progress_var,
                                          maximum=100, length=150)
        self.progress_bar.pack(side=tk.LEFT, padx=(5, 10))

        # Status label
        self.status_label = tk.Label(content, text="Prêt - En attente des fichiers",
                                   font=("Segoe UI", 8), fg=COLORS['INFO'],
                                   bg=COLORS['LIGHT'])
        self.status_label.pack(side=tk.LEFT, fill=tk.X, expand=True)

        # Créer un status_icon factice pour compatibilité
        self.status_icon = self.status_label
    
    def _create_modern_header(self, parent: tk.Widget):
        """Crée l'en-tête moderne du module avec style Sofrecom."""
        try:
            from ui.styles import create_sofrecom_card

            # Conteneur principal pour l'en-tête
            header_container = tk.Frame(parent, bg=COLORS['BG'])
            header_container.pack(fill=tk.X, pady=(0, 20))

            # Carte d'en-tête avec style Sofrecom
            header_card = create_sofrecom_card(
                header_container,
                title="Module 5 - Contrôle Qualité",
                subtitle="Système d'Analyse et de Validation de la Qualité des Données",
                description="Comparez les fichiers QGis et Teams pour détecter les incohérences et doublons automatiquement",
                icon="🔍",
                clickable=False
            )
            header_card.pack(fill=tk.X, padx=5, pady=5)

            # Indicateurs de statut rapide
            self._create_quick_status_indicators(header_container)

        except Exception as e:
            self.logger.warning(f"Fallback to basic header: {e}")
            # Fallback to basic header
            header_frame = tk.Frame(parent, bg=COLORS['CARD'])
            header_frame.pack(fill=tk.X, pady=(0, 20))

            title_label = tk.Label(
                header_frame,
                text="🔍 Module 5 - Contrôle Qualité",
                font=("Segoe UI", 16, "bold"),
                fg=COLORS['PRIMARY'],
                bg=COLORS['CARD']
            )
            title_label.pack(pady=10)

            desc_label = tk.Label(
                header_frame,
                text="Système d'analyse et de validation de la qualité des données",
                font=("Segoe UI", 10),
                fg=COLORS['TEXT_SECONDARY'],
                bg=COLORS['CARD']
            )
            desc_label.pack(pady=(0, 10))
    
    def _create_modern_data_loading_section(self, parent: tk.Widget):
        """Crée la section moderne de chargement des données avec cartes."""
        from ui.styles import create_sofrecom_card

        # Conteneur principal
        section_container = tk.Frame(parent, bg=COLORS['BG'])
        section_container.pack(fill=tk.X, pady=(0, 20))

        # Titre de section
        section_title = tk.Label(
            section_container,
            text="📁 Chargement des Données",
            font=UIConfig.FONT_HEADER,
            fg=COLORS['PRIMARY'],
            bg=COLORS['BG']
        )
        section_title.pack(anchor=tk.W, padx=5, pady=(0, 10))

        # Grille pour les cartes de fichiers
        files_grid = tk.Frame(section_container, bg=COLORS['BG'])
        files_grid.pack(fill=tk.X, padx=5)
        files_grid.grid_columnconfigure(0, weight=1)
        files_grid.grid_columnconfigure(1, weight=1)

        # Carte 1: Fichier QGis
        self._create_qgis_file_card(files_grid)

        # Carte 2: Fichier Suivi
        self._create_suivi_file_card(files_grid)

    def _create_quick_status_indicators(self, parent: tk.Widget):
        """Crée les indicateurs de statut rapide."""
        try:
            indicators_frame = tk.Frame(parent, bg=COLORS['BG'])
            indicators_frame.pack(fill=tk.X, padx=5, pady=(10, 0))

            # Grille pour les indicateurs
            indicators_frame.grid_columnconfigure(0, weight=1)
            indicators_frame.grid_columnconfigure(1, weight=1)
            indicators_frame.grid_columnconfigure(2, weight=1)

            # Indicateur 1: Fichiers chargés
            self.files_indicator = self._create_status_card(
                indicators_frame, 0, 0, "📁", "Fichiers", "En attente", COLORS['INFO']
            )

            # Indicateur 2: Analyse
            self.analysis_indicator = self._create_status_card(
                indicators_frame, 0, 1, "🔍", "Analyse", "Non effectuée", COLORS['TEXT_SECONDARY']
            )

            # Indicateur 3: Rapport
            self.report_indicator = self._create_status_card(
                indicators_frame, 0, 2, "📊", "Rapport", "Non généré", COLORS['TEXT_SECONDARY']
            )
        except Exception as e:
            self.logger.warning(f"Could not create status indicators: {e}")
            # Create dummy indicators to avoid errors
            self.files_indicator = tk.Frame(parent)
            self.analysis_indicator = tk.Frame(parent)
            self.report_indicator = tk.Frame(parent)

    def _create_status_card(self, parent: tk.Widget, row: int, col: int,
                           icon: str, title: str, status: str, color: str) -> tk.Frame:
        """Crée une carte d'indicateur de statut."""
        card_frame = tk.Frame(parent, bg=COLORS['CARD'], relief='flat', bd=1)
        card_frame.grid(row=row, column=col, sticky="ew", padx=3, pady=3)
        card_frame.config(highlightbackground=COLORS['BORDER'], highlightthickness=1)

        # Contenu de la carte
        content = tk.Frame(card_frame, bg=COLORS['CARD'])
        content.pack(fill=tk.BOTH, expand=True, padx=8, pady=6)

        # Icône et titre
        header = tk.Frame(content, bg=COLORS['CARD'])
        header.pack(fill=tk.X)

        icon_label = tk.Label(header, text=icon, font=("Segoe UI", 12),
                             fg=color, bg=COLORS['CARD'])
        icon_label.pack(side=tk.LEFT)

        title_label = tk.Label(header, text=title, font=UIConfig.FONT_SUBTITLE,
                              fg=COLORS['TEXT_PRIMARY'], bg=COLORS['CARD'])
        title_label.pack(side=tk.LEFT, padx=(5, 0))

        # Statut
        status_label = tk.Label(content, text=status, font=UIConfig.FONT_SMALL,
                               fg=color, bg=COLORS['CARD'])
        status_label.pack(anchor=tk.W, pady=(2, 0))

        # Stocker les labels pour mise à jour
        card_frame.icon_label = icon_label
        card_frame.status_label = status_label

        return card_frame

    def _create_qgis_file_card(self, parent: tk.Widget):
        """Crée la carte pour le fichier QGis."""
        from ui.styles import create_card_frame

        # Carte QGis
        qgis_card = create_card_frame(parent, shadow=True)
        qgis_card.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)

        # Contenu de la carte
        card_content = qgis_card.winfo_children()[0] if qgis_card.winfo_children() else qgis_card
        content_frame = tk.Frame(card_content, bg=COLORS['CARD'])
        content_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)

        # En-tête avec icône
        header_frame = tk.Frame(content_frame, bg=COLORS['CARD'])
        header_frame.pack(fill=tk.X, pady=(0, 10))

        tk.Label(
            header_frame,
            text="🗺️",
            font=("Segoe UI", 16),
            fg=COLORS['PRIMARY'],
            bg=COLORS['CARD']
        ).pack(side=tk.LEFT)

        tk.Label(
            header_frame,
            text="Fichier Résultats QGis",
            font=UIConfig.FONT_CARD_TITLE,
            fg=COLORS['PRIMARY'],
            bg=COLORS['CARD']
        ).pack(side=tk.LEFT, padx=(8, 0))

        # Description
        desc_label = tk.Label(
            content_frame,
            text="Fichier Excel contenant les résultats d'analyse QGis\n(Colonnes A-J requises)",
            font=UIConfig.FONT_SMALL,
            fg=COLORS['TEXT_SECONDARY'],
            bg=COLORS['CARD'],
            justify=tk.LEFT
        )
        desc_label.pack(anchor=tk.W, pady=(0, 15))

        # Bouton d'import moderne
        from tkinter import ttk
        self.load_qgis_button = ttk.Button(
            content_frame,
            text="📂 Importer Fichier QGis",
            command=self._load_qgis_file,
            style='Primary.TButton'
        )
        self.load_qgis_button.pack(anchor=tk.W, pady=(0, 10))

        # Statut du fichier
        self.qgis_info_label = tk.Label(
            content_frame,
            text="❌ Aucun fichier QGis chargé",
            font=UIConfig.FONT_SMALL,
            fg=COLORS['TEXT_SECONDARY'],
            bg=COLORS['CARD']
        )
        self.qgis_info_label.pack(anchor=tk.W)

    def _create_suivi_file_card(self, parent: tk.Widget):
        """Crée la carte pour le fichier Suivi."""
        from ui.styles import create_card_frame

        # Carte Suivi
        suivi_card = create_card_frame(parent, shadow=True)
        suivi_card.grid(row=0, column=1, sticky="nsew", padx=5, pady=5)

        # Contenu de la carte
        card_content = suivi_card.winfo_children()[0] if suivi_card.winfo_children() else suivi_card
        content_frame = tk.Frame(card_content, bg=COLORS['CARD'])
        content_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)

        # En-tête avec icône
        header_frame = tk.Frame(content_frame, bg=COLORS['CARD'])
        header_frame.pack(fill=tk.X, pady=(0, 10))

        tk.Label(
            header_frame,
            text="📋",
            font=("Segoe UI", 16),
            fg=COLORS['SUCCESS'],
            bg=COLORS['CARD']
        ).pack(side=tk.LEFT)

        tk.Label(
            header_frame,
            text="Fichier Suivi Commune",
            font=UIConfig.FONT_CARD_TITLE,
            fg=COLORS['SUCCESS'],
            bg=COLORS['CARD']
        ).pack(side=tk.LEFT, padx=(8, 0))

        # Description
        desc_label = tk.Label(
            content_frame,
            text="Fichier de suivi commune Teams\n(Feuille 3 utilisée pour l'analyse)",
            font=UIConfig.FONT_SMALL,
            fg=COLORS['TEXT_SECONDARY'],
            bg=COLORS['CARD'],
            justify=tk.LEFT
        )
        desc_label.pack(anchor=tk.W, pady=(0, 15))

        # Bouton d'import moderne
        from tkinter import ttk
        self.load_suivi_button = ttk.Button(
            content_frame,
            text="📂 Importer Fichier Suivi",
            command=self._load_suivi_file,
            style='Success.TButton'
        )
        self.load_suivi_button.pack(anchor=tk.W, pady=(0, 10))

        # Statut du fichier
        self.suivi_info_label = tk.Label(
            content_frame,
            text="❌ Aucun fichier suivi chargé",
            font=UIConfig.FONT_SMALL,
            fg=COLORS['TEXT_SECONDARY'],
            bg=COLORS['CARD']
        )
        self.suivi_info_label.pack(anchor=tk.W)
    
    def _create_modern_info_section(self, parent: tk.Widget):
        """Crée la section moderne d'informations détectées."""
        from ui.styles import create_card_frame

        # Conteneur principal
        section_container = tk.Frame(parent, bg=COLORS['BG'])
        section_container.pack(fill=tk.X, pady=(0, 20))

        # Titre de section
        section_title = tk.Label(
            section_container,
            text="📋 Informations Détectées Automatiquement",
            font=UIConfig.FONT_HEADER,
            fg=COLORS['PRIMARY'],
            bg=COLORS['BG']
        )
        section_title.pack(anchor=tk.W, padx=5, pady=(0, 10))

        # Carte d'informations
        info_card = create_card_frame(section_container, shadow=True)
        info_card.pack(fill=tk.X, padx=5, pady=5)

        # Contenu de la carte
        card_content = info_card.winfo_children()[0] if info_card.winfo_children() else info_card
        content_frame = tk.Frame(card_content, bg=COLORS['CARD'])
        content_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=15)

        # Grille pour les informations
        content_frame.grid_columnconfigure(0, weight=1)
        content_frame.grid_columnconfigure(1, weight=1)

        # Informations détectées
        self._create_info_field(content_frame, 0, 0, "👤", "Collaborateur", self.collaborator_var)
        self._create_info_field(content_frame, 0, 1, "🏘️", "Commune", self.commune_var)
        self._create_info_field(content_frame, 1, 0, "🏛️", "Code INSEE", self.insee_var)
        self._create_info_field(content_frame, 1, 1, "🆔", "ID Tâche", self.id_tache_var)

        # Note explicative
        note_label = tk.Label(
            content_frame,
            text="ℹ️ Ces informations sont extraites automatiquement du fichier de suivi commune",
            font=UIConfig.FONT_SMALL,
            fg=COLORS['INFO'],
            bg=COLORS['CARD']
        )
        note_label.grid(row=2, column=0, columnspan=2, sticky="w", pady=(15, 0))

    def _create_info_field(self, parent: tk.Widget, row: int, col: int,
                          icon: str, label: str, var: tk.StringVar):
        """Crée un champ d'information moderne."""
        field_frame = tk.Frame(parent, bg=COLORS['CARD'])
        field_frame.grid(row=row, column=col, sticky="ew", padx=10, pady=8)

        # En-tête avec icône
        header_frame = tk.Frame(field_frame, bg=COLORS['CARD'])
        header_frame.pack(fill=tk.X)

        icon_label = tk.Label(
            header_frame,
            text=icon,
            font=("Segoe UI", 12),
            fg=COLORS['SECONDARY'],
            bg=COLORS['CARD']
        )
        icon_label.pack(side=tk.LEFT)

        label_text = tk.Label(
            header_frame,
            text=label,
            font=UIConfig.FONT_SUBTITLE,
            fg=COLORS['TEXT_PRIMARY'],
            bg=COLORS['CARD']
        )
        label_text.pack(side=tk.LEFT, padx=(5, 0))

        # Valeur détectée - Texte en noir
        value_label = tk.Label(
            field_frame,
            textvariable=var,
            font=UIConfig.FONT_SMALL,
            fg=COLORS['TEXT_PRIMARY'],  # Changé de INFO à TEXT_PRIMARY pour texte noir
            bg=COLORS['CARD'],
            anchor='w'
        )
        value_label.pack(fill=tk.X, pady=(3, 0))

        # Stocker pour mise à jour
        if not hasattr(self, 'info_displays'):
            self.info_displays = {}
        self.info_displays[label.lower()] = value_label

    def _create_modern_analysis_section(self, parent: tk.Widget):
        """Crée la section moderne d'analyse et résultats."""
        from ui.styles import create_card_frame

        # Conteneur principal
        section_container = tk.Frame(parent, bg=COLORS['BG'])
        section_container.pack(fill=tk.X, pady=(0, 20))

        # Titre de section
        section_title = tk.Label(
            section_container,
            text="🔍 Analyse de Contrôle Qualité",
            font=UIConfig.FONT_HEADER,
            fg=COLORS['PRIMARY'],
            bg=COLORS['BG']
        )
        section_title.pack(anchor=tk.W, padx=5, pady=(0, 10))

        # Carte d'analyse
        analysis_card = create_card_frame(section_container, shadow=True)
        analysis_card.pack(fill=tk.X, padx=5, pady=5)

        # Contenu de la carte
        card_content = analysis_card.winfo_children()[0] if analysis_card.winfo_children() else analysis_card
        content_frame = tk.Frame(card_content, bg=COLORS['CARD'])
        content_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=15)

        # Description des critères
        criteria_frame = tk.Frame(content_frame, bg=COLORS['CARD'])
        criteria_frame.pack(fill=tk.X, pady=(0, 15))

        criteria_title = tk.Label(
            criteria_frame,
            text="Critères d'Analyse",
            font=UIConfig.FONT_CARD_TITLE,
            fg=COLORS['PRIMARY'],
            bg=COLORS['CARD']
        )
        criteria_title.pack(anchor=tk.W, pady=(0, 8))

        # Critère 0
        self._create_criteria_info(criteria_frame, "0", "Écart Plan Adressage",
                                  "Compare les motifs spécifiques: AD RAS, OK, NOK, UPR RAS, UPR OK, UPR NOK, Hors Commune")

        # Critère 2
        self._create_criteria_info(criteria_frame, "2", "Oubli Ticket UPR et 501/511",
                                  "Vérifie si les tickets UPR et 501/511 ont été déposés selon les motifs")

        # Critère 3
        self._create_criteria_info(criteria_frame, "3", "Contrôle IMB Doublons",
                                  "Détecte les doublons IMB suspects avec motifs identiques et adresses BAN différentes")

        # Critère 4
        self._create_criteria_info(criteria_frame, "4", "Détection AD à Analyser",
                                  "Détecte les motifs 'ad à analyser' avec IMB présent pour analyse approfondie")

        # Critère 5
        self._create_criteria_info(criteria_frame, "5", "Motif Incorrect",
                                  "Détecte les motifs non conformes (différents des 7 motifs autorisés)")

        # Bouton d'analyse moderne
        button_frame = tk.Frame(content_frame, bg=COLORS['CARD'])
        button_frame.pack(fill=tk.X, pady=(15, 0))

        from tkinter import ttk
        self.analyze_button = ttk.Button(
            button_frame,
            text="🔍 Lancer Analyse Qualité",
            command=self._run_quality_analysis,
            style='CompactWarning.TButton',
            state='disabled'
        )
        self.analyze_button.pack(side=tk.LEFT)

        # Bouton d'export
        self.export_button = ttk.Button(
            button_frame,
            text="📤 Exporter Rapport",
            command=self._export_qc_report,
            style='Compact.TButton',
            state='disabled'
        )
        self.export_button.pack(side=tk.LEFT, padx=(10, 0))

        # Zone de résultats
        self._create_results_display(content_frame)

    def _create_criteria_info(self, parent: tk.Widget, number: str, title: str, description: str):
        """Crée l'affichage d'un critère d'analyse."""
        criteria_frame = tk.Frame(parent, bg=COLORS['LIGHT'], relief='flat', bd=1)
        criteria_frame.pack(fill=tk.X, pady=3)
        criteria_frame.config(highlightbackground=COLORS['BORDER'], highlightthickness=1)

        content = tk.Frame(criteria_frame, bg=COLORS['LIGHT'])
        content.pack(fill=tk.X, padx=10, pady=8)

        # En-tête
        header = tk.Frame(content, bg=COLORS['LIGHT'])
        header.pack(fill=tk.X)

        # Badge du critère
        badge = tk.Label(
            header,
            text=f"CRITÈRE {number}",
            font=("Segoe UI", 8, "bold"),
            fg='white',
            bg=COLORS['SECONDARY'],
            padx=6,
            pady=2
        )
        badge.pack(side=tk.LEFT)

        # Titre
        title_label = tk.Label(
            header,
            text=title,
            font=UIConfig.FONT_SUBTITLE,
            fg=COLORS['TEXT_PRIMARY'],
            bg=COLORS['LIGHT']
        )
        title_label.pack(side=tk.LEFT, padx=(8, 0))

        # Description
        desc_label = tk.Label(
            content,
            text=description,
            font=UIConfig.FONT_SMALL,
            fg=COLORS['TEXT_SECONDARY'],
            bg=COLORS['LIGHT']
        )
        desc_label.pack(anchor=tk.W, pady=(5, 0))

    def _create_results_display(self, parent: tk.Widget):
        """Crée la zone d'affichage des résultats."""
        # Séparateur
        separator = tk.Frame(parent, height=1, bg=COLORS['BORDER'])
        separator.pack(fill=tk.X, pady=(20, 15))

        # Titre des résultats
        results_title = tk.Label(
            parent,
            text="📊 Résultats de l'Analyse",
            font=UIConfig.FONT_CARD_TITLE,
            fg=COLORS['PRIMARY'],
            bg=COLORS['CARD']
        )
        results_title.pack(anchor=tk.W, pady=(0, 10))

        # Zone de résultats (sera remplie après analyse)
        self.results_frame = tk.Frame(parent, bg=COLORS['CARD'])
        self.results_frame.pack(fill=tk.BOTH, expand=True)

        # Message initial
        self.results_label = tk.Label(
            self.results_frame,
            text="⏳ Aucune analyse effectuée\n\nChargez les fichiers requis et cliquez sur 'Lancer Analyse Qualité'",
            font=UIConfig.FONT_SMALL,
            fg=COLORS['TEXT_SECONDARY'],
            bg=COLORS['CARD'],
            justify=tk.CENTER
        )
        self.results_label.pack(expand=True, pady=20)

    def _create_modern_status_section(self, parent: tk.Widget):
        """Crée la section moderne de statut et progression."""
        try:
            from ui.styles import create_status_indicator

            # Conteneur de statut
            status_container = tk.Frame(parent, bg=COLORS['BG'])
            status_container.pack(fill=tk.X, pady=(10, 0))

            # Carte de statut
            status_card = tk.Frame(status_container, bg=COLORS['LIGHT'], relief='flat', bd=1)
            status_card.pack(fill=tk.X, padx=5, pady=5)
            status_card.config(highlightbackground=COLORS['BORDER'], highlightthickness=1)

            # Contenu de la carte de statut
            status_content = tk.Frame(status_card, bg=COLORS['LIGHT'])
            status_content.pack(fill=tk.X, padx=15, pady=10)

            # Barre de progression moderne
            from tkinter import ttk
            self.progress_var = tk.DoubleVar()

            progress_frame = tk.Frame(status_content, bg=COLORS['LIGHT'])
            progress_frame.pack(fill=tk.X, pady=(0, 8))

            progress_label = tk.Label(
                progress_frame,
                text="Progression:",
                font=UIConfig.FONT_SMALL,
                fg=COLORS['TEXT_PRIMARY'],
                bg=COLORS['LIGHT']
            )
            progress_label.pack(side=tk.LEFT)

            self.progress_bar = ttk.Progressbar(
                progress_frame,
                variable=self.progress_var,
                maximum=100,
                length=300
            )
            self.progress_bar.pack(side=tk.RIGHT, fill=tk.X, expand=True, padx=(10, 0))

            # Indicateur de statut moderne
            self.status_icon, self.status_label = create_status_indicator(status_content, COLORS['LIGHT'])

            # Initialiser le statut
            self._update_status("waiting", "Prêt - En attente des fichiers")

        except Exception as e:
            self.logger.warning(f"Fallback to basic status section: {e}")
            # Fallback to basic status section
            status_frame = tk.Frame(parent, bg=COLORS['CARD'])
            status_frame.pack(fill=tk.X, pady=(10, 0))

            # Barre de progression
            self.progress_var = tk.DoubleVar()
            self.progress_bar = ttk.Progressbar(
                status_frame,
                variable=self.progress_var,
                maximum=100,
                length=400
            )
            self.progress_bar.pack(pady=(0, 5))

            # Label de statut
            self.status_label = tk.Label(
                status_frame,
                text="Prêt - En attente des fichiers",
                font=("Segoe UI", 9),
                fg=COLORS['TEXT_SECONDARY'],
                bg=COLORS['CARD']
            )
            self.status_label.pack()

            # Créer un status_icon factice pour éviter les erreurs
            self.status_icon = self.status_label
    
    def _update_status(self, status_type: str, message: str):
        """Met à jour l'indicateur de statut moderne."""
        try:
            from ui.styles import update_status_indicator
            if hasattr(self, 'status_icon') and hasattr(self, 'status_label'):
                update_status_indicator(self.status_icon, self.status_label, status_type, message)
        except Exception as e:
            # Fallback to basic status update
            if hasattr(self, 'status_label'):
                self.status_label.config(text=message)
            self.logger.debug(f"Status update fallback: {e}")

    def _update_file_indicators(self):
        """Met à jour les indicateurs de fichiers."""
        qgis_loaded = self.qgis_data is not None
        suivi_loaded = self.suivi_data is not None

        if hasattr(self, 'files_status'):
            if qgis_loaded and suivi_loaded:
                self.files_status.config(text="📁 Fichiers: ✅ Tous chargés", fg=COLORS['SUCCESS'])
            elif qgis_loaded or suivi_loaded:
                self.files_status.config(text="📁 Fichiers: ⚠️ Partiels", fg=COLORS['WARNING'])
            else:
                self.files_status.config(text="📁 Fichiers: En attente", fg=COLORS['INFO'])

    def _update_status_card(self, card: tk.Frame, icon: str, title: str, status: str, color: str):
        """Met à jour une carte d'indicateur de statut."""
        if hasattr(card, 'icon_label') and hasattr(card, 'status_label'):
            card.icon_label.config(text=icon, fg=color)
            card.status_label.config(text=status, fg=color)
    
    def _load_qgis_file(self):
        """Charge un fichier résultats QGis pour analyse."""
        try:
            # Ouvrir dialogue de sélection de fichier
            file_path = filedialog.askopenfilename(
                title="Sélectionner le fichier Résultats QGis",
                filetypes=[
                    ("Fichiers Excel", "*.xlsx *.xls"),
                    ("Tous les fichiers", "*.*")
                ]
            )

            if not file_path:
                return

            # Utiliser les nouvelles améliorations visuelles si disponibles
            try:
                self._update_status_with_animation("Chargement du fichier QGis...", "📂", COLORS['INFO'])
                self._update_progress_bar(10)
            except:
                # Fallback vers l'ancienne méthode
                self._update_status("waiting", "Chargement du fichier QGis...")
                if hasattr(self, 'progress_var') and self.progress_var:
                    self.progress_var.set(10)

            def load_qgis():
                pd = get_pandas()

                # Lire le fichier Excel
                df = pd.read_excel(file_path, date_format=None)

                # Vérifier les colonnes requises
                required_columns = ['A', 'J']  # num dossier site, import fantome
                if len(df.columns) < 10:  # Au moins 10 colonnes (A-J)
                    raise ValueError("Le fichier ne contient pas assez de colonnes (minimum A-J requis)")

                return df

            def on_success(df):
                self.qgis_data = df
                self.current_qgis_file_path = file_path
                filename = os.path.basename(file_path)
                self.qgis_info_label.config(
                    text=f"✅ {filename} ({len(df)} lignes)",
                    fg=COLORS['SUCCESS']
                )
                # Utiliser les nouvelles améliorations visuelles si disponibles
                try:
                    self._update_status_with_animation("Fichier QGis chargé avec succès", "✅", COLORS['SUCCESS'])
                    self._update_progress_bar(100)
                    # Réinitialiser la barre après un délai
                    if self.main_frame and self.main_frame.winfo_exists():
                        self.main_frame.after(2000, lambda: self._update_progress_bar(0))
                except:
                    # Fallback vers l'ancienne méthode
                    self._update_status("success", "Fichier QGis chargé avec succès")

                # Mettre à jour les indicateurs de statut
                if hasattr(self, 'files_status'):
                    self.files_status.config(text="📁 Fichiers: QGis chargé", fg=COLORS['SUCCESS'])

                self._update_file_indicators()
                self._check_analysis_ready()
                self.logger.info(f"Fichier QGis chargé: {filename}")

            def on_error(error):
                self.qgis_info_label.config(
                    text="❌ Erreur de chargement",
                    fg=COLORS['ERROR']
                )
                # Utiliser les nouvelles améliorations visuelles si disponibles
                try:
                    self._update_status_with_animation("Erreur lors du chargement QGis", "❌", COLORS['ERROR'])
                    self._update_progress_bar(0)
                except:
                    # Fallback vers l'ancienne méthode
                    self._update_status("error", "Erreur lors du chargement QGis")

                # Mettre à jour les indicateurs de statut
                if hasattr(self, 'files_status'):
                    self.files_status.config(text="📁 Fichiers: Erreur QGis", fg=COLORS['ERROR'])

                self._update_file_indicators()
                messagebox.showerror("Erreur", f"Impossible de charger le fichier QGis:\n{error}")
                self.logger.error(f"Erreur chargement QGis: {error}")

            # Charger de manière asynchrone
            run_async_task(load_qgis, on_success, on_error, "Chargement QGis")

        except Exception as e:
            self.logger.error(f"Erreur lors de la sélection du fichier QGis: {e}")
            messagebox.showerror("Erreur", f"Erreur lors de la sélection du fichier:\n{e}")

    def _load_suivi_file(self):
        """Charge le fichier de suivi commune et détecte automatiquement les informations."""
        try:
            # Ouvrir dialogue de sélection de fichier
            file_path = filedialog.askopenfilename(
                title="Sélectionner le fichier Suivi Commune",
                filetypes=[
                    ("Fichiers Excel", "*.xlsx *.xls"),
                    ("Tous les fichiers", "*.*")
                ]
            )

            if not file_path:
                return

            # Utiliser les nouvelles améliorations visuelles si disponibles
            try:
                self._update_status_with_animation("Chargement du fichier suivi...", "📋", COLORS['INFO'])
                self._update_progress_bar(10)
            except:
                # Fallback vers l'ancienne méthode
                self._update_status("waiting", "Chargement du fichier suivi...")
                if hasattr(self, 'progress_var') and self.progress_var:
                    self.progress_var.set(10)

            def load_suivi():
                pd = get_pandas()

                # Lire la feuille 3 (index 2) du fichier de suivi
                df = pd.read_excel(file_path, sheet_name=2, date_format=None)

                # Détecter automatiquement les informations
                detected_info = self._detect_info_from_suivi(df, file_path)

                return df, detected_info

            def on_success(result):
                df, detected_info = result
                self.suivi_data = df
                self.current_suivi_file_path = file_path
                self.detected_info = detected_info

                # Mettre à jour l'affichage des informations détectées
                self.collaborator_var.set(detected_info.get('collaborateur', 'Non détecté'))
                self.commune_var.set(detected_info.get('commune', 'Non détecté'))
                self.insee_var.set(detected_info.get('insee', 'Non détecté'))
                self.id_tache_var.set(detected_info.get('id_tache', 'Non détecté'))

                # Évaluer et afficher le statut de conformité si des résultats QC existent
                if hasattr(self, 'qc_results') and self.qc_results:
                    evaluation_commune = self._evaluate_commune_status()
                    self._display_commune_status(evaluation_commune)

                filename = os.path.basename(file_path)
                self.suivi_info_label.config(
                    text=f"✅ {filename} ({len(df)} lignes)",
                    fg=COLORS['SUCCESS']
                )
                self._update_status("success", "Fichier suivi chargé avec succès")
                # Réinitialiser la barre de progression de manière sécurisée
                try:
                    self._update_progress_bar(0)
                except:
                    if hasattr(self, 'progress_var') and self.progress_var is not None:
                        self.progress_var.set(0)
                self._update_file_indicators()
                self._check_analysis_ready()
                self.logger.info(f"Fichier suivi chargé: {filename}")
                self.logger.info(f"Informations détectées: {detected_info}")

            def on_error(error):
                self.suivi_info_label.config(
                    text="❌ Erreur de chargement",
                    fg=COLORS['ERROR']
                )
                self._update_status("error", "Erreur lors du chargement suivi")
                # Réinitialiser la barre de progression de manière sécurisée
                try:
                    self._update_progress_bar(0)
                except:
                    if hasattr(self, 'progress_var') and self.progress_var is not None:
                        self.progress_var.set(0)
                self._update_file_indicators()
                messagebox.showerror("Erreur", f"Impossible de charger le fichier suivi:\n{error}")
                self.logger.error(f"Erreur chargement suivi: {error}")

            # Charger de manière asynchrone
            run_async_task(load_suivi, on_success, on_error, "Chargement suivi")

        except Exception as e:
            self.logger.error(f"Erreur lors de la sélection du fichier suivi: {e}")
            messagebox.showerror("Erreur", f"Erreur lors de la sélection du fichier:\n{e}")

    def _detect_info_from_suivi(self, df, file_path: str) -> Dict[str, str]:
        """Détecte automatiquement les informations depuis le fichier de suivi."""
        try:
            pd = get_pandas()
            detected = {}

            # 1. Détecter le nom de la commune depuis la colonne A (index 0) - Page 3
            if len(df.columns) >= 1:
                commune_column = df.iloc[:, 0]  # Colonne A
                # Chercher le premier nom de commune non vide
                for commune in commune_column:
                    if pd.notna(commune) and str(commune).strip() and str(commune).strip() not in ['Commune', 'Nom commune']:
                        detected['commune'] = str(commune).strip()
                        break

            # 2. Détecter le collaborateur depuis la colonne U (index 20) - Page 3
            if len(df.columns) >= 21:
                collab_column = df.iloc[:, 20]  # Colonne U
                # Chercher le premier collaborateur non vide
                for collab in collab_column:
                    if pd.notna(collab) and str(collab).strip() and str(collab).strip() != 'Collaborateur U':
                        detected['collaborateur'] = str(collab).strip()
                        break

            # 3. Détecter l'INSEE depuis la colonne C (index 2) - Page 3
            if len(df.columns) >= 3:
                insee_column = df.iloc[:, 2]  # Colonne C
                # Chercher le premier code INSEE valide (5 chiffres)
                for insee in insee_column:
                    if pd.notna(insee):
                        insee_str = str(insee).strip()
                        # Vérifier si c'est un code INSEE (5 chiffres)
                        if insee_str.isdigit() and len(insee_str) == 5:
                            detected['insee'] = insee_str
                            break

            # 4. Détecter l'ID Tâche depuis la colonne B (index 1) - Page 3
            if len(df.columns) >= 2:
                id_tache_column = df.iloc[:, 1]  # Colonne B
                # Chercher le premier ID tâche valide
                for id_tache in id_tache_column:
                    if pd.notna(id_tache):
                        id_tache_str = str(id_tache).strip()
                        # Vérifier si c'est un ID tâche valide (numérique, pas l'en-tête)
                        if id_tache_str.isdigit() and len(id_tache_str) >= 3 and id_tache_str != 'ID Tache':
                            detected['id_tache'] = id_tache_str
                            break

            # 5. Détecter le domaine depuis la colonne D (index 3) - Page 3
            if len(df.columns) >= 4 and len(df) > 0:
                # Essayer ligne 1 d'abord (index 0)
                if pd.notna(df.iloc[0, 3]):
                    domaine_val = str(df.iloc[0, 3]).strip()
                    if domaine_val not in ['Domaine', 'Domain', 'nan', ''] and len(domaine_val) > 1:
                        detected['domaine'] = domaine_val
                        self.logger.info(f"Domaine détecté ligne 1: {domaine_val}")

                # Si pas trouvé, essayer ligne 2 (index 1)
                if 'domaine' not in detected and len(df) > 1:
                    if pd.notna(df.iloc[1, 3]):
                        domaine_val = str(df.iloc[1, 3]).strip()
                        if domaine_val not in ['Domaine', 'Domain', 'nan', ''] and len(domaine_val) > 1:
                            detected['domaine'] = domaine_val
                            self.logger.info(f"Domaine détecté ligne 2: {domaine_val}")

                # Si pas trouvé, essayer ligne 3 (index 2)
                if 'domaine' not in detected and len(df) > 2:
                    if pd.notna(df.iloc[2, 3]):
                        domaine_val = str(df.iloc[2, 3]).strip()
                        if domaine_val not in ['Domaine', 'Domain', 'nan', ''] and len(domaine_val) > 1:
                            detected['domaine'] = domaine_val
                            self.logger.info(f"Domaine détecté ligne 3: {domaine_val}")

                # Si pas trouvé, chercher dans toutes les lignes
                if 'domaine' not in detected:
                    domaine_column = df.iloc[:, 3]  # Colonne D
                    for i, domaine in enumerate(domaine_column):
                        if pd.notna(domaine) and str(domaine).strip():
                            domaine_val = str(domaine).strip()
                            if domaine_val not in ['Domaine', 'Domain', 'nan', ''] and len(domaine_val) > 1:
                                detected['domaine'] = domaine_val
                                self.logger.info(f"Domaine détecté ligne {i+1}: {domaine_val}")
                                break

            # 6. Détecter Nbr voies CMS Total depuis la colonne G (index 6) - Page 3
            if len(df.columns) >= 7 and len(df) > 0:
                # Essayer ligne 1 d'abord (index 0)
                if pd.notna(df.iloc[0, 6]):
                    cms_str = str(df.iloc[0, 6]).strip()
                    try:
                        cms_num = float(cms_str)
                        if cms_num >= 0 and cms_str not in ['CMS', 'Total', 'Nbr', '0.0']:
                            # Correction pour les erreurs de précision flottante
                            if abs(cms_num - round(cms_num)) < 1e-10:
                                detected['cms_total'] = str(int(round(cms_num)))
                            else:
                                detected['cms_total'] = cms_str
                            self.logger.info(f"CMS Total détecté ligne 1: {detected['cms_total']}")
                    except ValueError:
                        pass

                # Si pas trouvé, essayer ligne 2 (index 1)
                if 'cms_total' not in detected and len(df) > 1:
                    if pd.notna(df.iloc[1, 6]):
                        cms_str = str(df.iloc[1, 6]).strip()
                        try:
                            cms_num = float(cms_str)
                            if cms_num >= 0 and cms_str not in ['CMS', 'Total', 'Nbr', '0.0']:
                                # Correction pour les erreurs de précision flottante
                                if abs(cms_num - round(cms_num)) < 1e-10:
                                    detected['cms_total'] = str(int(round(cms_num)))
                                else:
                                    detected['cms_total'] = cms_str
                                self.logger.info(f"CMS Total détecté ligne 2: {detected['cms_total']}")
                        except ValueError:
                            pass

                # Si pas trouvé, essayer ligne 3 (index 2)
                if 'cms_total' not in detected and len(df) > 2:
                    if pd.notna(df.iloc[2, 6]):
                        cms_str = str(df.iloc[2, 6]).strip()
                        try:
                            cms_num = float(cms_str)
                            if cms_num >= 0 and cms_str not in ['CMS', 'Total', 'Nbr', '0.0']:
                                # Correction pour les erreurs de précision flottante
                                if abs(cms_num - round(cms_num)) < 1e-10:
                                    detected['cms_total'] = str(int(round(cms_num)))
                                else:
                                    detected['cms_total'] = cms_str
                                self.logger.info(f"CMS Total détecté ligne 3: {detected['cms_total']}")
                        except ValueError:
                            pass

                # Si pas trouvé, chercher dans toutes les lignes
                if 'cms_total' not in detected:
                    cms_column = df.iloc[:, 6]  # Colonne G
                    for i, cms in enumerate(cms_column):
                        if pd.notna(cms):
                            cms_str = str(cms).strip()
                            try:
                                cms_num = float(cms_str)
                                if cms_num >= 0 and cms_str not in ['CMS', 'Total', 'Nbr', '0.0']:
                                    # Correction pour les erreurs de précision flottante
                                    if abs(cms_num - round(cms_num)) < 1e-10:
                                        detected['cms_total'] = str(int(round(cms_num)))
                                    else:
                                        detected['cms_total'] = cms_str
                                    self.logger.info(f"CMS Total détecté ligne {i+1}: {detected['cms_total']}")
                                    break
                            except ValueError:
                                continue

            # 7. Détecter Nbr IMB PA Total depuis la colonne H (index 7) - Page 3
            if len(df.columns) >= 8 and len(df) > 0:
                # Essayer ligne 1 d'abord (index 0)
                if pd.notna(df.iloc[0, 7]):
                    pa_str = str(df.iloc[0, 7]).strip()
                    try:
                        pa_num = float(pa_str)
                        if pa_num >= 0 and pa_str not in ['PA', 'Total', 'Nbr', '0.0']:
                            # Correction pour les erreurs de précision flottante
                            # Arrondir à l'entier le plus proche si très proche d'un entier
                            if abs(pa_num - round(pa_num)) < 1e-10:
                                detected['pa_total'] = str(int(round(pa_num)))
                            else:
                                detected['pa_total'] = pa_str
                            self.logger.info(f"PA Total détecté ligne 1: {detected['pa_total']}")
                    except ValueError:
                        pass

                # Si pas trouvé, essayer ligne 2 (index 1)
                if 'pa_total' not in detected and len(df) > 1:
                    if pd.notna(df.iloc[1, 7]):
                        pa_str = str(df.iloc[1, 7]).strip()
                        try:
                            pa_num = float(pa_str)
                            if pa_num >= 0 and pa_str not in ['PA', 'Total', 'Nbr', '0.0']:
                                # Correction pour les erreurs de précision flottante
                                # Arrondir à l'entier le plus proche si très proche d'un entier
                                if abs(pa_num - round(pa_num)) < 1e-10:
                                    detected['pa_total'] = str(int(round(pa_num)))
                                else:
                                    detected['pa_total'] = pa_str
                                self.logger.info(f"PA Total détecté ligne 2: {detected['pa_total']}")
                        except ValueError:
                            pass

                # Si pas trouvé, essayer ligne 3 (index 2)
                if 'pa_total' not in detected and len(df) > 2:
                    if pd.notna(df.iloc[2, 7]):
                        pa_str = str(df.iloc[2, 7]).strip()
                        try:
                            pa_num = float(pa_str)
                            if pa_num >= 0 and pa_str not in ['PA', 'Total', 'Nbr', '0.0']:
                                # Correction pour les erreurs de précision flottante
                                # Arrondir à l'entier le plus proche si très proche d'un entier
                                if abs(pa_num - round(pa_num)) < 1e-10:
                                    detected['pa_total'] = str(int(round(pa_num)))
                                else:
                                    detected['pa_total'] = pa_str
                                self.logger.info(f"PA Total détecté ligne 3: {detected['pa_total']}")
                        except ValueError:
                            pass

                # Si pas trouvé, chercher dans toutes les lignes
                if 'pa_total' not in detected:
                    pa_column = df.iloc[:, 7]  # Colonne H
                    for i, pa in enumerate(pa_column):
                        if pd.notna(pa):
                            pa_str = str(pa).strip()
                            try:
                                pa_num = float(pa_str)
                                if pa_num >= 0 and pa_str not in ['PA', 'Total', 'Nbr', '0.0']:
                                    # Correction pour les erreurs de précision flottante
                                    # Arrondir à l'entier le plus proche si très proche d'un entier
                                    if abs(pa_num - round(pa_num)) < 1e-10:
                                        detected['pa_total'] = str(int(round(pa_num)))
                                    else:
                                        detected['pa_total'] = pa_str
                                    self.logger.info(f"PA Total détecté ligne {i+1}: {detected['pa_total']}")
                                    break
                            except ValueError:
                                continue

            # 8. Détecter le collaborateur depuis la colonne U (index 20) - Page 3
            if len(df.columns) >= 21 and len(df) > 0:
                # Essayer ligne 1 d'abord (index 0)
                if pd.notna(df.iloc[0, 20]):
                    collab_val = str(df.iloc[0, 20]).strip()
                    if collab_val not in ['Collaborateur', 'Affectation', 'nan', ''] and len(collab_val) > 1:
                        detected['collaborateur'] = collab_val
                        self.logger.info(f"Collaborateur détecté ligne 1: {collab_val}")

                # Si pas trouvé, essayer ligne 2 (index 1)
                if 'collaborateur' not in detected and len(df) > 1:
                    if pd.notna(df.iloc[1, 20]):
                        collab_val = str(df.iloc[1, 20]).strip()
                        if collab_val not in ['Collaborateur', 'Affectation', 'nan', ''] and len(collab_val) > 1:
                            detected['collaborateur'] = collab_val
                            self.logger.info(f"Collaborateur détecté ligne 2: {collab_val}")

                # Si pas trouvé, essayer ligne 3 (index 2)
                if 'collaborateur' not in detected and len(df) > 2:
                    if pd.notna(df.iloc[2, 20]):
                        collab_val = str(df.iloc[2, 20]).strip()
                        if collab_val not in ['Collaborateur', 'Affectation', 'nan', ''] and len(collab_val) > 1:
                            detected['collaborateur'] = collab_val
                            self.logger.info(f"Collaborateur détecté ligne 3: {collab_val}")

                # Si pas trouvé, chercher dans toutes les lignes
                if 'collaborateur' not in detected:
                    collab_column = df.iloc[:, 20]  # Colonne U
                    for i, collab in enumerate(collab_column):
                        if pd.notna(collab) and str(collab).strip():
                            collab_val = str(collab).strip()
                            if collab_val not in ['Collaborateur', 'Affectation', 'nan', ''] and len(collab_val) > 1:
                                detected['collaborateur'] = collab_val
                                self.logger.info(f"Collaborateur détecté ligne {i+1}: {collab_val}")
                                break

            # 9. Valeurs par défaut si non détectées
            if 'commune' not in detected:
                detected['commune'] = 'Non détecté'
            if 'collaborateur' not in detected:
                detected['collaborateur'] = 'Non détecté'
            if 'insee' not in detected:
                detected['insee'] = 'Non détecté'
            if 'id_tache' not in detected:
                detected['id_tache'] = 'Non détecté'
            if 'domaine' not in detected:
                detected['domaine'] = ''
            if 'cms_total' not in detected:
                detected['cms_total'] = ''
            if 'pa_total' not in detected:
                detected['pa_total'] = ''

            # Log des informations détectées pour vérification
            self.logger.info("=== INFORMATIONS DÉTECTÉES ===")
            self.logger.info(f"Nom de commune (Col A): {detected.get('commune', 'N/A')}")
            self.logger.info(f"ID tâche (Col B): {detected.get('id_tache', 'N/A')}")
            self.logger.info(f"Code INSEE (Col C): {detected.get('insee', 'N/A')}")
            self.logger.info(f"Domaine (Col D): {detected.get('domaine', 'N/A')}")
            self.logger.info(f"Collaborateur (Col U): {detected.get('collaborateur', 'N/A')}")
            self.logger.info(f"CMS Total (Col G): {detected.get('cms_total', 'N/A')}")
            self.logger.info(f"PA Total (Col H): {detected.get('pa_total', 'N/A')}")
            self.logger.info("==============================")



            # Log détaillé des colonnes pour debug si des valeurs sont manquantes
            if not detected.get('domaine') or not detected.get('cms_total') or not detected.get('pa_total'):
                self.logger.info("=== DEBUG COLONNES MANQUANTES ===")
                if len(df.columns) >= 4:
                    self.logger.info(f"Colonne D (Domaine) - premières valeurs: {df.iloc[:5, 3].tolist()}")
                if len(df.columns) >= 7:
                    self.logger.info(f"Colonne G (CMS) - premières valeurs: {df.iloc[:5, 6].tolist()}")
                if len(df.columns) >= 8:
                    self.logger.info(f"Colonne H (PA) - premières valeurs: {df.iloc[:5, 7].tolist()}")
                self.logger.info("===================================")

            # Log détaillé spécifique pour ligne 2 (index 1)
            self.logger.info("=== DEBUG LIGNE 2 (INDEX 1) ===")
            if len(df) > 1:
                if len(df.columns) >= 4:
                    self.logger.info(f"Ligne 2, Colonne D (Domaine): '{df.iloc[1, 3]}'")
                if len(df.columns) >= 7:
                    self.logger.info(f"Ligne 2, Colonne G (CMS): '{df.iloc[1, 6]}'")
                if len(df.columns) >= 8:
                    self.logger.info(f"Ligne 2, Colonne H (PA): '{df.iloc[1, 7]}'")
            self.logger.info("===============================")

            return detected

        except Exception as e:
            self.logger.error(f"Erreur détection informations: {e}")
            return {
                'commune': 'Erreur détection',
                'collaborateur': 'Erreur détection',
                'insee': 'Erreur détection',
                'id_tache': 'Erreur détection'
            }
    
    def _check_analysis_ready(self):
        """Vérifie si tous les éléments sont prêts pour lancer l'analyse."""
        ready = (
            self.qgis_data is not None and
            self.suivi_data is not None and
            self.detected_info.get('collaborateur', '') != 'Non détecté' and
            self.detected_info.get('commune', '') != 'Non détecté'
        )

        if ready:
            self.analyze_button.config(state='normal')
            self._update_status("success", "Prêt pour l'analyse qualité")
        else:
            self.analyze_button.config(state='disabled')

        return ready
    
    def _run_quality_analysis(self):
        """Lance l'analyse de contrôle qualité."""
        # Log du mode utilisé
        mode = "Autoévaluation" if self.is_autoevaluation_mode() else "Contrôle Qualité"
        self.logger.info(f"Lancement analyse en mode: {mode}")

        if not self._check_analysis_ready():
            messagebox.showwarning("Attention", "Veuillez charger tous les fichiers requis avant de lancer l'analyse.")
            return

        try:
            self._update_status("waiting", "Analyse en cours...")
            # Mettre à jour la barre de progression de manière sécurisée
            try:
                self._update_progress_bar(10)
            except:
                if hasattr(self, 'progress_var') and self.progress_var is not None:
                    self.progress_var.set(10)

            def run_analysis():
                results = {
                    'critere_0': self._analyze_critere_0(),
                    'critere_2': self._analyze_critere_2(),
                    'critere_3': self._analyze_critere_3(),
                    'critere_4': self._analyze_critere_4(),
                    'critere_5': self._analyze_critere_5(),
                    'summary': {}
                }

                # Calculer le résumé
                results['summary'] = self._calculate_summary(results)

                return results

            def on_success(results):
                self.qc_results = results
                self._display_compact_results(results)
                self.export_button.config(state='normal')
                self._update_status("success", "Analyse terminée avec succès")

                # Mettre à jour les indicateurs de statut
                if hasattr(self, 'analysis_status'):
                    self.analysis_status.config(text="🔍 Analyse: ✅ Terminée", fg=COLORS['SUCCESS'])
                if hasattr(self, 'report_status'):
                    self.report_status.config(text="📊 Rapport: ⚠️ Prêt", fg=COLORS['WARNING'])

                # Mettre à jour la barre de progression de manière sécurisée
                try:
                    self._update_progress_bar(100)
                except:
                    if hasattr(self, 'progress_var') and self.progress_var is not None:
                        self.progress_var.set(100)
                self.logger.info("Analyse qualité terminée")

            def on_error(error):
                self._update_status("error", "Erreur lors de l'analyse")
                if hasattr(self, 'analysis_status'):
                    self.analysis_status.config(text="🔍 Analyse: ❌ Erreur", fg=COLORS['ERROR'])
                # Réinitialiser la barre de progression de manière sécurisée
                try:
                    self._update_progress_bar(0)
                except:
                    if hasattr(self, 'progress_var') and self.progress_var is not None:
                        self.progress_var.set(0)
                messagebox.showerror("Erreur", f"Erreur lors de l'analyse:\n{error}")
                self.logger.error(f"Erreur analyse qualité: {error}")

            # Lancer l'analyse de manière asynchrone
            run_async_task(run_analysis, on_success, on_error, "Analyse qualité")

        except Exception as e:
            self.logger.error(f"Erreur lors du lancement de l'analyse: {e}")
            messagebox.showerror("Erreur", f"Erreur lors du lancement:\n{e}")

    def _analyze_critere_0(self) -> Dict[str, Any]:
        """
        Critère 0: Incohérence entre fichier Résultats QGis et suivi commune.
        Compare le nombre des motifs spécifiés dans les deux fichiers.
        Motifs à vérifier: AD RAS, OK, NOK, UPR RAS, UPR OK, UPR NOK, Hors Commune
        """
        try:
            pd = get_pandas()

            # Motifs spécifiques à analyser pour l'écart Plan Adressage
            motifs_plan_adressage = [
                'AD RAS', 'OK', 'NOK', 'UPR RAS', 'UPR OK', 'UPR NOK', 'HORS COMMUNE'
            ]

            # Extraire les motifs du fichier QGis (colonne J - index 9)
            qgis_motifs = []
            if len(self.qgis_data.columns) > 9:
                qgis_motifs_column = self.qgis_data.iloc[:, 9]  # Colonne J (import fantome)
                qgis_motifs = [str(motif).strip().upper() for motif in qgis_motifs_column if pd.notna(motif) and str(motif).strip()]

            # Extraire les motifs du fichier suivi commune (colonne I - index 8) depuis la PAGE 2
            suivi_motifs = []
            try:
                # Lire spécifiquement la page 2 (index 1) pour les motifs
                pd = get_pandas()
                if hasattr(self, 'current_suivi_file_path') and self.current_suivi_file_path:
                    suivi_page2_df = pd.read_excel(self.current_suivi_file_path, sheet_name=1, date_format=None)  # Page 2
                    if len(suivi_page2_df.columns) > 8:
                        suivi_motifs_column = suivi_page2_df.iloc[:, 8]  # Colonne I (Motif)
                        suivi_motifs = [str(motif).strip().upper() for motif in suivi_motifs_column if pd.notna(motif) and str(motif).strip()]
                        self.logger.info(f"Motifs extraits de la page 2, colonne I: {len(suivi_motifs)} motifs trouvés")
                else:
                    self.logger.warning("Chemin du fichier suivi non disponible, impossible de lire la page 2")
            except Exception as e:
                self.logger.error(f"Erreur lors de la lecture de la page 2 pour les motifs: {e}")
                # Fallback: essayer avec les données actuelles (page 3)
                if len(self.suivi_data.columns) > 8:
                    suivi_motifs_column = self.suivi_data.iloc[:, 8]  # Colonne I (Motif)
                    suivi_motifs = [str(motif).strip().upper() for motif in suivi_motifs_column if pd.notna(motif) and str(motif).strip()]

            # Compter les occurrences de chaque motif spécifique
            qgis_counts = {}
            suivi_counts = {}

            # Initialiser tous les motifs à 0
            for motif in motifs_plan_adressage:
                qgis_counts[motif] = 0
                suivi_counts[motif] = 0

            # Compter les motifs QGis
            for motif in qgis_motifs:
                if motif in motifs_plan_adressage:
                    qgis_counts[motif] += 1

            # Compter les motifs Suivi
            for motif in suivi_motifs:
                if motif in motifs_plan_adressage:
                    suivi_counts[motif] += 1

            # Détecter les incohérences pour les motifs Plan Adressage
            incoherences = []
            ecart_plan_adressage = {}

            for motif in motifs_plan_adressage:
                qgis_count = qgis_counts[motif]
                suivi_count = suivi_counts[motif]
                difference = qgis_count - suivi_count

                # Stocker les données pour l'écart Plan Adressage
                ecart_plan_adressage[motif] = {
                    'suivi_count': suivi_count,
                    'qgis_count': qgis_count,
                    'difference': difference,
                    'has_ecart': difference != 0
                }

                # Ajouter aux incohérences si différence détectée
                if difference != 0:
                    incoherences.append({
                        'motif': motif,
                        'qgis_count': qgis_count,
                        'suivi_count': suivi_count,
                        'difference': difference,
                        'type': 'ECART_PLAN_ADRESSAGE',
                        'description': f"Écart détecté: {suivi_count} (Suivi) vs {qgis_count} (QGis) = {difference:+d}"
                    })

            # Calculer les totaux
            total_qgis = sum(qgis_counts.values())
            total_suivi = sum(suivi_counts.values())

            return {
                'status': 'COMPLETE',
                'total_incoherences': len(incoherences),
                'incoherences': incoherences,
                'ecart_plan_adressage': ecart_plan_adressage,
                'qgis_motifs_total': total_qgis,
                'suivi_motifs_total': total_suivi,
                'qgis_counts': qgis_counts,
                'suivi_counts': suivi_counts,
                'motifs_analyses': motifs_plan_adressage
            }

        except Exception as e:
            self.logger.error(f"Erreur dans l'analyse du critère 0: {e}")
            return {
                'status': 'ERROR',
                'error': str(e),
                'total_incoherences': 0,
                'incoherences': [],
                'ecart_plan_adressage': {}
            }
    


    def _analyze_critere_2(self) -> Dict[str, Any]:
        """
        Critère 2: Oubli Ticket UPR et 501/511
        Vérifie si les collaborateurs ont déposé les tickets requis selon les motifs utilisés.

        Règles:
        1. UPR OK (page 2, col I) -> doit avoir ticket UPR (page 3, col T)
        2. Création/Modification Voie (page 1, col E) OU OK (page 2, col I) -> doit avoir ticket 501/511 (page 3, col Q)
        """
        try:
            pd = get_pandas()

            # Initialiser les résultats
            ticket_upr_status = "N/A"
            ticket_501_511_status = "N/A"
            errors = []

            # Variables pour stocker les données des différentes pages
            page1_data = None
            page2_data = None
            page3_data = self.suivi_data  # Page 3 déjà chargée

            # Lire les pages 1 et 2 du fichier suivi commune
            if hasattr(self, 'current_suivi_file_path') and self.current_suivi_file_path:
                try:
                    # Page 1 pour vérifier colonne E (Motif Voie)
                    page1_data = pd.read_excel(self.current_suivi_file_path, sheet_name=0, date_format=None)
                    # Page 2 pour vérifier colonne I (Motif) - déjà utilisée dans critère 0
                    page2_data = pd.read_excel(self.current_suivi_file_path, sheet_name=1, date_format=None)

                    self.logger.info("Pages 1, 2 et 3 du fichier suivi chargées pour analyse critère 2")

                except Exception as e:
                    self.logger.error(f"Erreur lecture pages suivi pour critère 2: {e}")
                    return {
                        'status': 'ERROR',
                        'error': f'Impossible de lire les pages du fichier suivi: {e}',
                        'ticket_upr_status': 'ERROR',
                        'ticket_501_511_status': 'ERROR',
                        'errors': []
                    }
            else:
                return {
                    'status': 'ERROR',
                    'error': 'Chemin du fichier suivi non disponible',
                    'ticket_upr_status': 'ERROR',
                    'ticket_501_511_status': 'ERROR',
                    'errors': []
                }

            # VÉRIFICATION 1: Ticket UPR
            # Si UPR OK dans page 2, colonne I -> doit avoir ID dans page 3, colonne T
            upr_ok_found = False
            if page2_data is not None and len(page2_data.columns) > 8:
                motifs_page2 = page2_data.iloc[:, 8]  # Colonne I
                for motif in motifs_page2:
                    if pd.notna(motif) and str(motif).strip().upper() == 'UPR OK':
                        upr_ok_found = True
                        break

            if upr_ok_found:
                # Vérifier si ticket UPR déposé (page 3, colonne T - index 19)
                ticket_upr_deposited = False
                if len(page3_data.columns) > 19:
                    id_upr_column = page3_data.iloc[:, 19]  # Colonne T
                    for id_upr in id_upr_column:
                        if pd.notna(id_upr) and str(id_upr).strip() and str(id_upr).strip() not in ['', 'nan', 'ID UPR', 'Ticket UPR']:
                            ticket_upr_deposited = True
                            break

                ticket_upr_status = "OK" if ticket_upr_deposited else "NOK"
                if not ticket_upr_deposited:
                    errors.append({
                        'type': 'TICKET_UPR_MANQUANT',
                        'description': 'Motif UPR OK détecté mais aucun ticket UPR déposé (colonne T vide)',
                        'page': 'Page 2 -> Page 3',
                        'colonnes': 'I -> T'
                    })
            else:
                ticket_upr_status = "N/A"  # Pas de UPR OK, donc pas de ticket requis

            # VÉRIFICATION 2: Ticket 501/511
            # Si Création/Modification Voie (page 1, col E) OU OK (page 2, col I) -> doit avoir ID dans page 3, col Q
            ticket_501_511_required = False

            # Vérifier page 1, colonne E (Motif Voie)
            if page1_data is not None and len(page1_data.columns) > 4:
                motifs_voie = page1_data.iloc[:, 4]  # Colonne E
                for motif in motifs_voie:
                    if pd.notna(motif):
                        motif_str = str(motif).strip()
                        if motif_str in ['Création Voie', 'Modification Voie']:
                            ticket_501_511_required = True
                            break

            # Vérifier page 2, colonne I (motif OK)
            if not ticket_501_511_required and page2_data is not None and len(page2_data.columns) > 8:
                motifs_page2 = page2_data.iloc[:, 8]  # Colonne I
                for motif in motifs_page2:
                    if pd.notna(motif) and str(motif).strip().upper() == 'OK':
                        ticket_501_511_required = True
                        break

            if ticket_501_511_required:
                # Vérifier si ticket 501/511 déposé (page 3, colonne Q - index 16)
                ticket_501_511_deposited = False
                if len(page3_data.columns) > 16:
                    id_501_511_column = page3_data.iloc[:, 16]  # Colonne Q
                    for id_ticket in id_501_511_column:
                        if pd.notna(id_ticket) and str(id_ticket).strip() and str(id_ticket).strip() not in ['', 'nan', 'ID 501/511', 'Ticket 501/511']:
                            ticket_501_511_deposited = True
                            break

                ticket_501_511_status = "OK" if ticket_501_511_deposited else "NOK"
                if not ticket_501_511_deposited:
                    errors.append({
                        'type': 'TICKET_501_511_MANQUANT',
                        'description': 'Motif Création/Modification Voie ou OK détecté mais aucun ticket 501/511 déposé (colonne Q vide)',
                        'page': 'Page 1/2 -> Page 3',
                        'colonnes': 'E/I -> Q'
                    })
            else:
                ticket_501_511_status = "N/A"  # Pas de motif requis, donc pas de ticket nécessaire

            return {
                'status': 'COMPLETE',
                'ticket_upr_status': ticket_upr_status,
                'ticket_501_511_status': ticket_501_511_status,
                'total_errors': len(errors),
                'errors': errors,
                'upr_ok_found': upr_ok_found,
                'ticket_501_511_required': ticket_501_511_required
            }

        except Exception as e:
            self.logger.error(f"Erreur dans l'analyse du critère 2: {e}")
            return {
                'status': 'ERROR',
                'error': str(e),
                'ticket_upr_status': 'ERROR',
                'ticket_501_511_status': 'ERROR',
                'total_errors': 0,
                'errors': []
            }

    def _analyze_critere_3(self) -> Dict[str, Any]:
        """
        Critère 3: Contrôle IMB Doublons
        Détecte les doublons suspects dans le fichier Résultats QGis en identifiant
        les codes IMB identiques ayant le même motif de traitement avec des adresses BAN différentes.

        Colonnes analysées:
        - A: Num Dossier Site (codes IMB)
        - B: Numero Voie Site
        - C: Repondant Voie Site
        - D: Libelle Voie Site
        - J: Import Fantome (motifs)
        - U: Adresse BAN
        """
        try:
            pd = get_pandas()

            # Vérifier que les colonnes requises existent
            required_columns = ['A', 'B', 'C', 'D', 'J', 'U']  # Colonnes A, B, C, D, J, U
            if len(self.qgis_data.columns) < 21:  # Au moins 21 colonnes (A-U)
                raise ValueError("Le fichier QGis ne contient pas assez de colonnes (minimum A-U requis)")

            # Extraire les données des colonnes requises
            df = self.qgis_data.copy()

            # Renommer les colonnes pour faciliter le travail
            df_work = pd.DataFrame({
                'imb_code': df.iloc[:, 0],  # Colonne A - Num Dossier Site
                'numero_voie': df.iloc[:, 1],  # Colonne B - Numero Voie Site
                'repondant_voie': df.iloc[:, 2],  # Colonne C - Repondant Voie Site
                'libelle_voie': df.iloc[:, 3],  # Colonne D - Libelle Voie Site
                'motif': df.iloc[:, 9],  # Colonne J - Import Fantome
                'adresse_ban': df.iloc[:, 20]  # Colonne U - Adresse BAN
            })

            # Nettoyer les données
            df_work = df_work.dropna(subset=['imb_code'])  # Supprimer les lignes sans code IMB
            df_work['imb_code'] = df_work['imb_code'].astype(str).str.strip()
            df_work['motif'] = df_work['motif'].astype(str).str.strip().str.upper()
            df_work['adresse_ban'] = df_work['adresse_ban'].astype(str).str.strip()

            # Construire l'adresse optimum (B + C + D) en gérant les valeurs vides
            def construct_adresse_optimum(row):
                """Construit l'adresse optimum en gérant les valeurs vides et numériques."""
                # Nettoyer et convertir les valeurs
                numero = str(row['numero_voie']).strip() if pd.notna(row['numero_voie']) else ''
                repondant = str(row['repondant_voie']).strip() if pd.notna(row['repondant_voie']) else ''
                libelle = str(row['libelle_voie']).strip() if pd.notna(row['libelle_voie']) else ''

                # Nettoyer les valeurs numériques (enlever .0)
                if numero and numero != 'nan':
                    try:
                        # Si c'est un nombre entier, enlever les décimales
                        if '.' in numero and numero.replace('.', '').replace('-', '').isdigit():
                            numero = str(int(float(numero)))
                    except:
                        pass

                # Construire l'adresse selon les règles :
                # - Si C est vide : B + D uniquement
                # - Si B ou C est vide : D uniquement
                # - Sinon : B + C + D

                parts = []

                if numero and numero != 'nan':
                    parts.append(numero)

                if repondant and repondant != 'nan':
                    parts.append(repondant)

                if libelle and libelle != 'nan':
                    parts.append(libelle)

                # Appliquer les règles spécifiques
                if not repondant or repondant == 'nan':
                    # Si C est vide : B + D uniquement (ou D si B est vide)
                    if numero and numero != 'nan' and libelle and libelle != 'nan':
                        return f"{numero} {libelle}".strip()
                    elif libelle and libelle != 'nan':
                        return libelle.strip()
                    elif numero and numero != 'nan':
                        return numero.strip()
                    else:
                        return ''
                else:
                    # Cas normal : B + C + D
                    return ' '.join(parts).strip()

            df_work['adresse_optimum'] = df_work.apply(construct_adresse_optimum, axis=1)

            # Identifier les doublons IMB
            imb_counts = df_work['imb_code'].value_counts()
            doublons_imb = imb_counts[imb_counts > 1].index.tolist()

            doublons_suspects = []
            doublons_details = []
            erreurs_motif_ok = []
            ad_a_analyser = []
            lignes_deja_ajoutees = set()  # Pour éviter les doublons

            # ÉTAPE 1: Détecter les erreurs de motif "OK" avec adresses identiques (tous les IMB)
            for index, row in df_work.iterrows():
                motif = str(row['motif']).strip().upper()
                adresse_optimum = str(row['adresse_optimum']).strip()
                adresse_ban = str(row['adresse_ban']).strip()

                # Debug: Log pour vérifier les comparaisons
                self.logger.debug(f"Vérification motif OK - IMB: {row['imb_code']}, Motif: '{motif}', Adresse Opt: '{adresse_optimum}', Adresse BAN: '{adresse_ban}'")

                # Vérifier si motif = "OK" et adresses identiques
                if (motif == 'OK' and
                    adresse_optimum and adresse_optimum != '' and adresse_optimum != 'nan' and
                    adresse_ban and adresse_ban != '' and adresse_ban != 'nan' and
                    adresse_optimum == adresse_ban):

                    # Créer une clé unique pour cette ligne
                    ligne_key = f"{row['imb_code']}_{adresse_optimum}_{adresse_ban}_{motif}"

                    if ligne_key not in lignes_deja_ajoutees:
                        erreur_info = {
                            'imb_code': row['imb_code'],
                            'adresse_optimum': adresse_optimum,
                            'motif_initial': motif,
                            'adresse_ban': adresse_ban,
                            'type': 'ERREUR_MOTIF_OK_ADRESSE_IDENTIQUE',
                            'description': f"IMB {row['imb_code']} - Motif 'OK' mais adresse optimum = adresse BAN (erreur de saisie)"
                        }
                        erreurs_motif_ok.append(erreur_info)
                        doublons_details.append(erreur_info)
                        lignes_deja_ajoutees.add(ligne_key)

                        # Log pour confirmer la détection
                        self.logger.info(f"ERREUR MOTIF OK DÉTECTÉE - IMB: {row['imb_code']}, Adresse: '{adresse_optimum}'")

            # ÉTAPE 2: Analyser chaque groupe de doublons IMB
            for imb_code in doublons_imb:
                groupe_doublons = df_work[df_work['imb_code'] == imb_code].copy()

                # Vérifier s'il y a des motifs identiques avec des adresses BAN différentes
                for motif in groupe_doublons['motif'].unique():
                    if pd.isna(motif) or motif == '' or motif.upper() == 'NAN':
                        continue

                    lignes_meme_motif = groupe_doublons[groupe_doublons['motif'] == motif]

                    if len(lignes_meme_motif) > 1:
                        # Vérifier si les adresses BAN sont différentes
                        adresses_ban_uniques = lignes_meme_motif['adresse_ban'].unique()
                        adresses_ban_uniques = [addr for addr in adresses_ban_uniques
                                              if pd.notna(addr) and str(addr).strip() != '' and str(addr).upper() != 'NAN']

                        if len(adresses_ban_uniques) > 1:
                            # Doublon suspect détecté !
                            for _, ligne in lignes_meme_motif.iterrows():
                                # Créer une clé unique pour cette ligne
                                ligne_key = f"{ligne['imb_code']}_{ligne['adresse_optimum']}_{ligne['adresse_ban']}_{ligne['motif']}"

                                # Ajouter seulement si pas déjà ajoutée
                                if ligne_key not in lignes_deja_ajoutees:
                                    doublon_info = {
                                        'imb_code': ligne['imb_code'],
                                        'adresse_optimum': ligne['adresse_optimum'],
                                        'motif_initial': ligne['motif'],
                                        'adresse_ban': ligne['adresse_ban'],
                                        'type': 'DOUBLON_IMB_SUSPECT',
                                        'description': f"IMB {ligne['imb_code']} - Motif '{ligne['motif']}' avec adresses BAN différentes"
                                    }
                                    doublons_suspects.append(doublon_info)
                                    doublons_details.append(doublon_info)
                                    lignes_deja_ajoutees.add(ligne_key)

            # Calculer les statistiques
            total_doublons_imb = len(doublons_imb)
            total_doublons_suspects = len(doublons_suspects)
            total_erreurs_motif_ok = len(erreurs_motif_ok)
            total_erreurs_detectees = total_doublons_suspects + total_erreurs_motif_ok

            return {
                'status': 'COMPLETE',
                'total_doublons_imb': total_doublons_imb,
                'total_doublons_suspects': total_doublons_suspects,
                'total_erreurs_motif_ok': total_erreurs_motif_ok,
                'total_erreurs_detectees': total_erreurs_detectees,
                'doublons_suspects': doublons_suspects,
                'erreurs_motif_ok': erreurs_motif_ok,
                'doublons_details': doublons_details,
                'doublons_imb_codes': doublons_imb,
                'total_records_analyzed': len(df_work)
            }

        except Exception as e:
            self.logger.error(f"Erreur dans l'analyse du critère 3: {e}")
            return {
                'status': 'ERROR',
                'error': str(e),
                'total_doublons_imb': 0,
                'total_doublons_suspects': 0,
                'total_erreurs_motif_ok': 0,
                'total_erreurs_detectees': 0,
                'doublons_suspects': [],
                'erreurs_motif_ok': [],
                'doublons_details': []
            }

    def _analyze_critere_4(self) -> Dict[str, Any]:
        """
        Critère 4: Détection "ad à analyser"
        Détecte les entrées avec motif "ad à analyser" qui ont un IMB présent dans la colonne A.

        Colonnes analysées:
        - A: Num Dossier Site (codes IMB) - OBLIGATOIRE
        - B: Numero Voie Site
        - C: Repondant Voie Site
        - D: Libelle Voie Site
        - J: Import Fantome (motifs)
        - U: Adresse BAN
        """
        try:
            pd = get_pandas()

            # Vérifier que les colonnes requises existent
            if len(self.qgis_data.columns) < 21:  # Au moins 21 colonnes (A-U)
                raise ValueError("Le fichier QGis ne contient pas assez de colonnes (minimum A-U requis)")

            # Extraire les données des colonnes requises
            df = self.qgis_data.copy()

            # Renommer les colonnes pour faciliter le travail
            df_work = pd.DataFrame({
                'imb_code': df.iloc[:, 0],  # Colonne A - Num Dossier Site
                'numero_voie': df.iloc[:, 1],  # Colonne B - Numero Voie Site
                'repondant_voie': df.iloc[:, 2],  # Colonne C - Repondant Voie Site
                'libelle_voie': df.iloc[:, 3],  # Colonne D - Libelle Voie Site
                'motif': df.iloc[:, 9],  # Colonne J - Import Fantome
                'adresse_ban': df.iloc[:, 20]  # Colonne U - Adresse BAN
            })

            # Nettoyer les données
            df_work['imb_code'] = df_work['imb_code'].astype(str).str.strip()
            df_work['motif'] = df_work['motif'].astype(str).str.strip().str.upper()
            df_work['adresse_ban'] = df_work['adresse_ban'].astype(str).str.strip()

            # Construire l'adresse optimum (même logique que critère 3)
            def construct_adresse_optimum(row):
                """Construit l'adresse optimum en gérant les valeurs vides et numériques."""
                numero = str(row['numero_voie']).strip() if pd.notna(row['numero_voie']) else ''
                repondant = str(row['repondant_voie']).strip() if pd.notna(row['repondant_voie']) else ''
                libelle = str(row['libelle_voie']).strip() if pd.notna(row['libelle_voie']) else ''

                # Nettoyer les valeurs numériques (enlever .0)
                if numero and numero != 'nan':
                    try:
                        if '.' in numero and numero.replace('.', '').replace('-', '').isdigit():
                            numero = str(int(float(numero)))
                    except:
                        pass

                # Appliquer les règles de construction d'adresse
                if not repondant or repondant == 'nan':
                    if numero and numero != 'nan' and libelle and libelle != 'nan':
                        return f"{numero} {libelle}".strip()
                    elif libelle and libelle != 'nan':
                        return libelle.strip()
                    elif numero and numero != 'nan':
                        return numero.strip()
                    else:
                        return ''
                else:
                    parts = []
                    if numero and numero != 'nan':
                        parts.append(numero)
                    if repondant and repondant != 'nan':
                        parts.append(repondant)
                    if libelle and libelle != 'nan':
                        parts.append(libelle)
                    return ' '.join(parts).strip()

            df_work['adresse_optimum'] = df_work.apply(construct_adresse_optimum, axis=1)

            # Détecter les entrées "ad à analyser" avec IMB présent
            ad_a_analyser_entries = []

            for _, row in df_work.iterrows():
                motif = str(row['motif']).strip().upper()
                imb_code = str(row['imb_code']).strip()

                # Vérifier si motif = "AD À ANALYSER" et IMB présent
                if (motif == 'AD À ANALYSER' and
                    imb_code and imb_code != '' and imb_code != 'nan' and imb_code.upper() != 'NAN'):

                    entry_info = {
                        'imb_code': imb_code,
                        'adresse_optimum': row['adresse_optimum'],
                        'motif_initial': 'ad à analyser',
                        'adresse_ban': row['adresse_ban'],
                        'type': 'AD_A_ANALYSER_AVEC_IMB',
                        'description': f"IMB {imb_code} - Motif 'ad à analyser' nécessitant une analyse"
                    }
                    ad_a_analyser_entries.append(entry_info)

                    # Log pour confirmer la détection
                    self.logger.info(f"AD À ANALYSER DÉTECTÉ - IMB: {imb_code}, Adresse: '{row['adresse_optimum']}'")

            # Calculer les statistiques
            total_ad_a_analyser = len(ad_a_analyser_entries)

            return {
                'status': 'COMPLETE',
                'total_ad_a_analyser': total_ad_a_analyser,
                'ad_a_analyser_entries': ad_a_analyser_entries,
                'total_records_analyzed': len(df_work)
            }

        except Exception as e:
            self.logger.error(f"Erreur dans l'analyse du critère 4: {e}")
            return {
                'status': 'ERROR',
                'error': str(e),
                'total_ad_a_analyser': 0,
                'ad_a_analyser_entries': []
            }

    def _analyze_critere_5(self) -> Dict[str, Any]:
        """
        Critère 5: Motif Incorrect
        Détecte les motifs non conformes dans le fichier Résultats QGis (colonne J)
        qui ne correspondent pas aux 7 motifs autorisés.

        Colonnes analysées:
        - A: Num Dossier Site (codes IMB)
        - B: Numero Voie Site
        - C: Repondant Voie Site
        - D: Libelle Voie Site
        - J: Import Fantome (motifs à vérifier)
        - U: Adresse BAN
        """
        try:
            pd = get_pandas()

            # Motifs autorisés (liste de référence)
            motifs_autorises = [
                'AD RAS', 'OK', 'NOK', 'UPR RAS', 'UPR OK', 'UPR NOK', 'HORS COMMUNE'
            ]

            # Vérifier que les colonnes requises existent
            if len(self.qgis_data.columns) < 21:  # Au moins 21 colonnes (A-U)
                raise ValueError("Le fichier QGis ne contient pas assez de colonnes (minimum A-U requis)")

            # Extraire les données des colonnes requises
            df = self.qgis_data.copy()

            # Renommer les colonnes pour faciliter le travail
            df_work = pd.DataFrame({
                'imb_code': df.iloc[:, 0],  # Colonne A - Num Dossier Site
                'numero_voie': df.iloc[:, 1],  # Colonne B - Numero Voie Site
                'repondant_voie': df.iloc[:, 2],  # Colonne C - Repondant Voie Site
                'libelle_voie': df.iloc[:, 3],  # Colonne D - Libelle Voie Site
                'motif': df.iloc[:, 9],  # Colonne J - Import Fantome (index 9 pour colonne J)
                'adresse_ban': df.iloc[:, 20]  # Colonne U - Adresse BAN
            })

            # Nettoyer les données
            df_work = df_work.dropna(subset=['imb_code'])  # Supprimer les lignes sans code IMB
            df_work['imb_code'] = df_work['imb_code'].astype(str).str.strip()
            df_work['motif'] = df_work['motif'].astype(str).str.strip().str.upper()
            df_work['adresse_ban'] = df_work['adresse_ban'].astype(str).str.strip()

            # Construire l'adresse optimum (même logique que critère 3)
            def construct_adresse_optimum(row):
                """Construit l'adresse optimum en gérant les valeurs vides et numériques."""
                numero = str(row['numero_voie']).strip() if pd.notna(row['numero_voie']) else ''
                repondant = str(row['repondant_voie']).strip() if pd.notna(row['repondant_voie']) else ''
                libelle = str(row['libelle_voie']).strip() if pd.notna(row['libelle_voie']) else ''

                # Nettoyer les valeurs numériques (enlever .0)
                if numero and numero != 'nan':
                    try:
                        if '.' in numero and numero.replace('.', '').replace('-', '').isdigit():
                            numero = str(int(float(numero)))
                    except:
                        pass

                # Construire l'adresse selon les règles du critère 3
                parts = []
                if numero and numero != 'nan':
                    parts.append(numero)
                if repondant and repondant != 'nan':
                    parts.append(repondant)
                if libelle and libelle != 'nan':
                    parts.append(libelle)

                # Appliquer les règles spécifiques
                if not repondant or repondant == 'nan':
                    if numero and numero != 'nan' and libelle and libelle != 'nan':
                        return f"{numero} {libelle}".strip()
                    elif libelle and libelle != 'nan':
                        return libelle.strip()
                    elif numero and numero != 'nan':
                        return numero.strip()
                    else:
                        return ''
                else:
                    return ' '.join(parts).strip()

            df_work['adresse_optimum'] = df_work.apply(construct_adresse_optimum, axis=1)

            motifs_incorrects_entries = []

            # Analyser chaque ligne pour détecter les motifs incorrects
            for index, row in df_work.iterrows():
                motif = str(row['motif']).strip().upper()
                imb_code = str(row['imb_code']).strip()

                # Ignorer les lignes avec motifs vides ou NaN
                if not motif or motif == '' or motif.upper() == 'NAN':
                    continue

                # Vérifier si le motif n'est pas dans la liste autorisée
                if motif not in motifs_autorises:
                    entry_info = {
                        'imb_code': imb_code,
                        'adresse_optimum': row['adresse_optimum'],
                        'motif_incorrect': motif,
                        'adresse_ban': row['adresse_ban'],
                        'type': 'MOTIF_INCORRECT',
                        'description': f"IMB {imb_code} - Motif '{motif}' non autorisé"
                    }
                    motifs_incorrects_entries.append(entry_info)

                    # Log pour confirmer la détection
                    self.logger.info(f"MOTIF INCORRECT DÉTECTÉ - IMB: {imb_code}, Motif: '{motif}'")

            # Calculer les statistiques
            total_motifs_incorrects = len(motifs_incorrects_entries)

            return {
                'status': 'COMPLETE',
                'total_motifs_incorrects': total_motifs_incorrects,
                'motifs_incorrects_entries': motifs_incorrects_entries,
                'total_records_analyzed': len(df_work)
            }

        except Exception as e:
            self.logger.error(f"Erreur dans l'analyse du critère 5: {e}")
            return {
                'status': 'ERROR',
                'error': str(e),
                'total_motifs_incorrects': 0,
                'motifs_incorrects_entries': []
            }

    def _calculate_summary(self, results: Dict[str, Any]) -> Dict[str, Any]:
        """Calcule le résumé des résultats d'analyse."""
        try:
            critere_0 = results.get('critere_0', {})
            critere_2 = results.get('critere_2', {})
            critere_3 = results.get('critere_3', {})
            critere_4 = results.get('critere_4', {})
            critere_5 = results.get('critere_5', {})

            total_errors = (
                critere_0.get('total_incoherences', 0) +
                critere_2.get('total_errors', 0) +
                critere_3.get('total_erreurs_detectees', 0) +
                critere_4.get('total_ad_a_analyser', 0) +
                critere_5.get('total_motifs_incorrects', 0)
            )

            return {
                'total_errors': total_errors,
                'critere_0_errors': critere_0.get('total_incoherences', 0),
                'critere_2_errors': critere_2.get('total_errors', 0),
                'critere_3_errors': critere_3.get('total_erreurs_detectees', 0),
                'critere_3_doublons_suspects': critere_3.get('total_doublons_suspects', 0),
                'critere_3_erreurs_motif_ok': critere_3.get('total_erreurs_motif_ok', 0),
                'critere_4_errors': critere_4.get('total_ad_a_analyser', 0),
                'critere_5_errors': critere_5.get('total_motifs_incorrects', 0),
                'ticket_upr_status': critere_2.get('ticket_upr_status', 'N/A'),
                'ticket_501_511_status': critere_2.get('ticket_501_511_status', 'N/A'),
                'total_doublons_imb': critere_3.get('total_doublons_imb', 0),
                'total_doublons_suspects': critere_3.get('total_doublons_suspects', 0),
                'analysis_date': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                'commune': self.detected_info.get('commune', 'Non détecté'),
                'collaborateur': self.detected_info.get('collaborateur', 'Non détecté'),
                'insee': self.detected_info.get('insee', 'Non détecté'),
                'id_tache': self.detected_info.get('id_tache', 'Non détecté'),
                'domaine': self.detected_info.get('domaine', ''),
                'cms_total': self.detected_info.get('cms_total', ''),
                'pa_total': self.detected_info.get('pa_total', '')
            }

        except Exception as e:
            self.logger.error(f"Erreur calcul résumé: {e}")
            return {'total_errors': 0}

    def _export_qc_report(self):
        """Exporte le rapport de contrôle qualité automatiquement vers Teams."""
        if not self.qc_results:
            messagebox.showwarning("Attention", "Aucune analyse à exporter. Veuillez d'abord lancer l'analyse.")
            return

        try:
            # Générer le nom de fichier depuis les informations détectées
            commune = self.detected_info.get('commune', 'Commune')
            insee = self.detected_info.get('insee', 'INSEE')
            collaborateur = self.detected_info.get('collaborateur', 'Collaborateur')
            id_tache = self.detected_info.get('id_tache', 'ID_TACHE')

            # Nettoyer les noms pour le fichier (enlever caractères spéciaux)
            commune_clean = "".join(c for c in commune if c.isalnum() or c in (' ', '-', '_')).strip()
            collaborateur_clean = "".join(c for c in collaborateur if c.isalnum() or c in (' ', '-', '_')).strip()

            filename = f"Etat_De_Lieu_{commune_clean}_{insee}_{collaborateur_clean}.xlsx"

            # Utiliser l'enregistrement automatique Teams
            file_path = self._get_teams_save_path(commune, id_tache, insee, collaborateur, filename)

            if not file_path:
                return

            self._update_status("waiting", "Génération du rapport...")
            # Mettre à jour la barre de progression de manière sécurisée
            try:
                self._update_progress_bar(50)
            except:
                if hasattr(self, 'progress_var') and self.progress_var is not None:
                    self.progress_var.set(50)

            def generate_report():
                return self._generate_excel_report(file_path)

            def on_success(success):
                if success:
                    self._update_status("success", "Rapport exporté avec succès")
                    if hasattr(self, 'report_status'):
                        self.report_status.config(text="📊 Rapport: ✅ Exporté", fg=COLORS['SUCCESS'])
                    # Mettre à jour la barre de progression de manière sécurisée
                    try:
                        self._update_progress_bar(100)
                    except:
                        if hasattr(self, 'progress_var') and self.progress_var is not None:
                            self.progress_var.set(100)
                    messagebox.showinfo("Succès", f"Rapport exporté vers:\n{file_path}")
                    self.logger.info(f"Rapport exporté: {file_path}")
                else:
                    self._update_status("error", "Erreur lors de l'export")
                    # Réinitialiser la barre de progression de manière sécurisée
                    try:
                        self._update_progress_bar(0)
                    except:
                        if hasattr(self, 'progress_var') and self.progress_var is not None:
                            self.progress_var.set(0)
                    messagebox.showerror("Erreur", "Échec de la génération du rapport Excel.\nVérifiez les logs pour plus de détails.")

            def on_error(error):
                self._update_status("error", "Erreur lors de l'export")
                # Réinitialiser la barre de progression de manière sécurisée
                try:
                    self._update_progress_bar(0)
                except:
                    if hasattr(self, 'progress_var') and self.progress_var is not None:
                        self.progress_var.set(0)
                error_msg = f"Erreur lors de l'export:\n{str(error)}"
                if len(error_msg) > 200:
                    error_msg = error_msg[:200] + "...\n\nVoir les logs pour plus de détails."
                messagebox.showerror("Erreur", error_msg)
                self.logger.error(f"Erreur export rapport: {error}")
                import traceback
                self.logger.error(f"Traceback: {traceback.format_exc()}")

            # Générer de manière asynchrone
            run_async_task(generate_report, on_success, on_error, "Export rapport")

        except Exception as e:
            self.logger.error(f"Erreur lors de l'export: {e}")
            messagebox.showerror("Erreur", f"Erreur lors de l'export:\n{e}")

    def _get_teams_save_path(self, commune: str, id_tache: str, insee: str, collaborateur: str, filename: str) -> str:
        """
        Génère le chemin de sauvegarde pour le contrôle qualité.
        - Mode Autoévaluation: Sauvegarde locale
        - Mode Contrôle Qualité: Sauvegarde Teams

        Args:
            commune: Nom de la commune
            id_tache: ID de la tâche
            insee: Code INSEE
            collaborateur: Nom du collaborateur
            filename: Nom du fichier

        Returns:
            Chemin complet du fichier ou None si erreur
        """
        try:
            import os
            from utils.file_utils import create_quality_control_folder, get_quality_control_file_path
            from config.constants import TeamsConfig
            from tkinter import filedialog

            # Mode Autoévaluation: Sauvegarde locale
            if self.is_autoevaluation_mode():
                return self._get_local_save_path(commune, id_tache, insee, collaborateur, filename)

            # Mode Contrôle Qualité: Sauvegarde Teams
            # Vérifier que Teams est accessible
            quality_control_base = TeamsConfig.get_quality_control_teams_path()
            if not os.path.exists(quality_control_base):
                # Essayer de créer le dossier de base
                try:
                    os.makedirs(quality_control_base, exist_ok=True)
                    self.logger.info(f"Dossier Contrôle Qualité créé: {quality_control_base}")
                except Exception as e:
                    self.logger.error(f"Impossible de créer le dossier Contrôle Qualité: {e}")
                    messagebox.showerror("Erreur Teams",
                                       f"Impossible d'accéder au canal Teams:\n{quality_control_base}\n\n"
                                       f"Erreur: {e}")
                    return None

            # Créer la structure de dossiers
            folder_result = create_quality_control_folder(commune, id_tache, insee, collaborateur)

            if not folder_result['success']:
                self.logger.error(f"Erreur création dossier: {folder_result['error']}")
                messagebox.showerror("Erreur", f"Impossible de créer le dossier:\n{folder_result['error']}")
                return None

            # Générer le chemin complet du fichier
            file_path = get_quality_control_file_path(commune, id_tache, insee, collaborateur, filename)

            # Afficher confirmation à l'utilisateur
            commune_folder = f"{commune}_{id_tache}_{insee}"
            message = (f"Le fichier sera sauvegardé automatiquement dans Teams :\n\n"
                      f"👤 Collaborateur: {collaborateur}\n"
                      f"📁 Commune: {commune_folder}\n"
                      f"📄 Fichier: {filename}\n\n"
                      f"Arborescence: Contrôle Qualité > {collaborateur} > {commune_folder}\n\n"
                      f"Continuer ?")

            if messagebox.askyesno("Sauvegarde Teams", message, icon='question'):
                self.logger.info(f"Sauvegarde Teams confirmée: {file_path}")
                return file_path
            else:
                self.logger.info("Sauvegarde Teams annulée par l'utilisateur")
                return None

        except Exception as e:
            self.logger.error(f"Erreur génération chemin Teams: {e}")
            messagebox.showerror("Erreur", f"Erreur lors de la préparation de la sauvegarde:\n{e}")
            return None

    def _get_local_save_path(self, commune: str, id_tache: str, insee: str, collaborateur: str, filename: str) -> str:
        """
        Génère le chemin de sauvegarde locale pour le mode autoévaluation.

        Args:
            commune: Nom de la commune
            id_tache: ID de la tâche
            insee: Code INSEE
            collaborateur: Nom du collaborateur
            filename: Nom du fichier

        Returns:
            Chemin complet du fichier ou None si annulé
        """
        try:
            from tkinter import filedialog
            import os

            # Nom de fichier suggéré
            commune_folder = f"{commune}_{id_tache}_{insee}"
            suggested_filename = f"QC_Autoevaluation_{commune_folder}_{filename}"

            # Demander à l'utilisateur où sauvegarder
            file_path = filedialog.asksaveasfilename(
                title="Sauvegarder le rapport d'autoévaluation",
                defaultextension=".xlsx",
                filetypes=[("Fichiers Excel", "*.xlsx"), ("Tous les fichiers", "*.*")]
            )

            if file_path:
                # Afficher confirmation à l'utilisateur
                message = (f"Le fichier sera sauvegardé localement :\n\n"
                          f"📁 Mode: Autoévaluation (Local)\n"
                          f"👤 Collaborateur: {collaborateur}\n"
                          f"🏘️ Commune: {commune_folder}\n"
                          f"📄 Fichier: {os.path.basename(file_path)}\n"
                          f"📂 Dossier: {os.path.dirname(file_path)}\n\n"
                          f"Continuer ?")

                if messagebox.askyesno("Sauvegarde Locale", message, icon='question'):
                    self.logger.info(f"Sauvegarde locale confirmée: {file_path}")
                    return file_path
                else:
                    self.logger.info("Sauvegarde locale annulée par l'utilisateur")
                    return None
            else:
                self.logger.info("Aucun fichier sélectionné pour la sauvegarde locale")
                return None

        except Exception as e:
            self.logger.error(f"Erreur génération chemin local: {e}")
            messagebox.showerror("Erreur", f"Erreur lors de la sélection du fichier:\n{e}")
            return None

    def _generate_excel_report(self, file_path: str) -> bool:
        """Génère le rapport Excel avec 2 feuilles."""
        try:
            self.logger.info(f"Début génération rapport Excel: {file_path}")

            # Log du mode utilisé
            mode = "Autoévaluation (Local)" if self.is_autoevaluation_mode() else "Contrôle Qualité (Teams)"
            self.logger.info(f"Mode de sauvegarde: {mode}")

            # Vérifier que les résultats QC existent
            if not self.qc_results:
                raise ValueError("Aucun résultat d'analyse disponible pour l'export")

            # Nettoyer le chemin du fichier
            file_path = os.path.abspath(file_path)
            self.logger.info(f"Chemin absolu du fichier: {file_path}")

            # Vérifier que le répertoire existe
            directory = os.path.dirname(file_path)
            if not os.path.exists(directory):
                os.makedirs(directory, exist_ok=True)
                self.logger.info(f"Répertoire créé: {directory}")

            # Générer les données des 4 feuilles d'abord
            page1_data, page2_data, page3_data, page4_data = self._prepare_excel_data()
            self.logger.info("Données Excel préparées pour 4 feuilles")

            # Créer le fichier Excel de manière isolée
            success = self._write_excel_file(file_path, page1_data, page2_data, page3_data, page4_data)

            if success:
                # Vérifier que le fichier a été créé
                if os.path.exists(file_path):
                    file_size = os.path.getsize(file_path)
                    self.logger.info(f"Fichier créé avec succès: {file_path} ({file_size} bytes)")
                    return True
                else:
                    raise FileNotFoundError(f"Le fichier n'a pas été créé: {file_path}")
            else:
                return False

        except Exception as e:
            self.logger.error(f"Erreur génération rapport Excel: {e}")
            import traceback
            self.logger.error(f"Traceback complet: {traceback.format_exc()}")
            return False

    def _prepare_excel_data(self):
        """Prépare les données pour les 4 feuilles Excel selon la nouvelle structure."""
        try:
            summary = self.qc_results.get('summary', {})
            commune = summary.get('commune', 'COMMUNE')
            collaborateur = summary.get('collaborateur', 'COLLABORATEUR')
            insee = summary.get('insee', 'INSEE')
            id_tache = summary.get('id_tache', 'ID_TACHE')
            domaine = summary.get('domaine', '')
            cms_total = summary.get('cms_total', '')
            pa_total = summary.get('pa_total', '')

            # Page 1: Structure améliorée avec en-tête professionnel et métadonnées
            from datetime import datetime

            # Évaluer le statut de conformité de la commune
            evaluation_commune = self._evaluate_commune_status()
            statut_commune = evaluation_commune['statut']
            pourcentage_conformite = evaluation_commune['pourcentage_conformite']
            raisons_ko = evaluation_commune['raisons_ko']
            fautes_majeures = evaluation_commune['fautes_majeures']

            # Calculer les statistiques globales pour l'en-tête
            total_errors = sum([
                self.qc_results.get('critere_0', {}).get('total_incoherences', 0),
                self.qc_results.get('critere_2', {}).get('total_errors', 0),
                self.qc_results.get('critere_3', {}).get('total_erreurs_detectees', 0),
                self.qc_results.get('critere_4', {}).get('total_ad_a_analyser', 0)
            ]) if self.qc_results else 0

            # Déterminer le statut qualité global basé sur la conformité
            fichiers_manquants = evaluation_commune.get('fichiers_manquants', False)

            if statut_commune == "OK":
                if pourcentage_conformite >= 95:
                    statut_qualite = "EXCELLENT"
                    statut_couleur = "SUCCESS"
                else:
                    statut_qualite = "BON"
                    statut_couleur = "SUCCESS"
            else:  # statut_commune == "KO"
                if fichiers_manquants:
                    statut_qualite = "FICHIERS MANQUANTS"
                    statut_couleur = "ERROR"
                elif fautes_majeures:
                    statut_qualite = "CRITIQUE"
                    statut_couleur = "ERROR"
                else:
                    statut_qualite = "À AMÉLIORER"
                    statut_couleur = "ERROR"

            # Calculer les pourcentages avec pondérations (pour les valeurs statiques) AVANT de créer page1_data
            resume_erreurs_data = self._calculate_resume_erreurs()

            page1_data = [
                # Tableau 1: Informations de base (commence directement ligne 1)
                # Titre occupe uniquement la colonne A
                ['INFORMATIONS GÉNÉRALES', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''],
                ['Nom de commune', 'ID tâche Plan Adressage', 'Code INSEE', 'Domaine', 'AFFECTATION', 'Contrôleur', '', '', '', '', '', '', '', '', '', ''],
                [commune, id_tache, insee, domaine, collaborateur, '', '', '', '', '', '', '', '', '', '', ''],  # Contrôleur sera rempli par validation

                # Espacement de 1 ligne entre tableaux
                ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''],

                # Tableau 2: Qualité CMS Adresse (ligne 5, titre colonne A uniquement, tableau A vers D)
                ['Qualité CMS Adresse', '', '', '', '', '', 'Résumé Erreurs', '', '', '', '', '', '', '', '', ''],
                ['Nbr voies CMS Total', 'Nbr erreurs CMS KO', '% Erreur CMS', 'Statut Global CMS', '', '', 'Catégorie d\'Erreur', '% Brut', 'Pondération', 'Score', 'Statut Commune', '', '', '', '', ''],
                [cms_total, '=SUMPRODUCT(--(LEN(TRIM(Controle_Qualite_CMS!A2:A1000))>0))', '=IF(A7=0,0,B7/A7)', '=IF(C7>0,"Non Conforme","Conforme")', '', '', '% Erreur CMS', '=C7', resume_erreurs_data['ponderation_cms'], '=H7*I7', '', '', '', '', '', ''],
                # Tableau 3: Controle Plan Adressage (ligne 8, titre colonne A uniquement, tableau A vers D) - SANS ligne vide
                ['Controle Plan Adressage', '', '', '', '', '', '% Erreur PA', '=C10', resume_erreurs_data['ponderation_pa'], '=H8*I8', '', '', '', '', '', ''],
                ['Nbr IMB PA Total', 'Nbr IMB PA KO', '% Erreur PA', 'Statut Global PA', '', '', '% Erreur Banbou', resume_erreurs_data['pourcentage_banbou_brut'], resume_erreurs_data['ponderation_banbou'], '=H9*I9', '', '', '', '', '', ''],
                [pa_total, '=SUMPRODUCT(--(LEN(TRIM(Controle_Qualite_PA!A2:A1000))>0))', '=IF(A10=0,0,B10/A10)', '=IF(C10>0,"Non Conforme","Conforme")', '', '', '% Ecart Plan Adressage', '=B25', resume_erreurs_data['ponderation_ecart'], '=H10*I10', '', '', '', '', '', ''],
                # Fin Résumé Erreurs - SANS ligne vide
                ['', '', '', '', '', '', 'SCORE TOTAL', '-', '1', '=SUM(J7:J10)', '=IF(COUNTIF(Controle_Qualite_PA!G:G,"Faute Majeure")>0,"KO",IF(J11>=0.1,"KO","OK"))', '', '', '', '', ''],

                # Tableau 4: Contrôle Dépose Tickets (ligne 12, titre colonne A uniquement, tableau A vers D) - CORRIGÉ
                ['Contrôle Dépose Tickets', '', '', '', '', '', '', '', ''],
                ['Ticket 501/511', 'Ticket UPR', '% Erreur Banbou', 'Statut Global Tickets', '', '', '', '', ''],
            ]

            # Remplir les statuts des tickets et calculer % Erreur Banbou avec métadonnées
            if self.qc_results and 'critere_2' in self.qc_results:
                critere_2 = self.qc_results['critere_2']
                ticket_501_511_status = critere_2.get('ticket_501_511_status', '')
                ticket_upr_status = critere_2.get('ticket_upr_status', '')

                # Calculer le % Erreur Banbou
                erreur_banbou_percentage = self._calculate_erreur_banbou_percentage()
                erreur_banbou_str = f"{erreur_banbou_percentage:.0f}%"

                # Déterminer le statut global des tickets selon les nouvelles spécifications
                # Statut Global Tickets (Conforme ou Non Conforme si le % Erreur Banbou dépasse 0%)
                if erreur_banbou_percentage > 0:
                    statut_global_tickets = "Non Conforme"
                else:
                    statut_global_tickets = "Conforme"

                page1_data.append([ticket_501_511_status, ticket_upr_status, erreur_banbou_str, statut_global_tickets, '', '', '', '', '', '', '', '', ''])
            else:
                # Pas de données d'analyse, afficher vides
                page1_data.append(['', '', '', 'EN ATTENTE', '', '', '', '', '', '', '', '', ''])

            # Pas d'espacement supplémentaire ici pour éviter le décalage

            # Préparer les données d'écart Plan Adressage
            motifs_data = []
            if self.qc_results and 'critere_0' in self.qc_results:
                critere_0 = self.qc_results['critere_0']
                ecart_data = critere_0.get('ecart_plan_adressage', {})

                # Motifs dans l'ordre d'affichage
                motifs_ordre = ['AD RAS', 'OK', 'NOK', 'UPR RAS', 'UPR OK', 'UPR NOK', 'HORS COMMUNE']

                for motif in motifs_ordre:
                    if motif in ecart_data:
                        data = ecart_data[motif]
                        suivi_count = data['suivi_count']
                        qgis_count = data['qgis_count']
                        motif_display = motif.title() if motif != 'HORS COMMUNE' else 'Hors Commune'
                        # Garder les valeurs en format numérique pour les formules Excel
                        motifs_data.append([motif_display, suivi_count, qgis_count])
                    else:
                        motif_display = motif.title() if motif != 'HORS COMMUNE' else 'Hors Commune'
                        motifs_data.append([motif_display, 0, 0])
            else:
                # Pas de données d'analyse, afficher les motifs avec valeurs numériques 0
                motifs_ordre = ['AD RAS', 'OK', 'NOK', 'UPR RAS', 'UPR OK', 'UPR NOK', 'Hors Commune']
                for motif in motifs_ordre:
                    motifs_data.append([motif, 0, 0])

            # Espacement de 1 ligne avant le tableau Ecart Plan Adressage
            page1_data.extend([
                ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''],  # Ligne vide (ligne 16)
            ])

            # Tableau 5: Ecart Plan Adressage (ligne 17, titre colonne A uniquement, tableau A vers D)
            page1_data.extend([
                ['Ecart Plan Adressage', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''],  # Titre section (ligne 17, colonne A uniquement)
                ['Motif', 'Suivi', 'QGis', 'Écart', '', '', '', '', '', '', '', '', '', '', '', ''],  # En-têtes (ligne 18, A vers D)

                # Données des motifs avec calculs d'écart (tableau A vers D uniquement) - CORRIGÉ avec valeurs numériques
                [motifs_data[0][0], motifs_data[0][1], motifs_data[0][2],
                 '=ABS(B18-C18)',  # Ad Ras (ligne 18) - Toujours calculer l'écart
                 '', '', '', '', '', '', '', '', '', '', '', ''],

                [motifs_data[1][0], motifs_data[1][1], motifs_data[1][2],
                 '=ABS(B19-C19)',  # Ok (ligne 19) - Toujours calculer l'écart
                 '', '', '', '', '', '', '', '', '', '', '', ''],

                [motifs_data[2][0], motifs_data[2][1], motifs_data[2][2],
                 '=ABS(B20-C20)',  # Nok (ligne 20) - Toujours calculer l'écart
                 '', '', '', '', '', '', '', '', '', '', '', ''],

                [motifs_data[3][0], motifs_data[3][1], motifs_data[3][2],
                 '=ABS(B21-C21)',  # Upr Ras (ligne 21) - Toujours calculer l'écart
                 '', '', '', '', '', '', '', '', '', '', '', ''],

                [motifs_data[4][0], motifs_data[4][1], motifs_data[4][2],
                 '=ABS(B22-C22)',  # Upr Ok (ligne 22) - Toujours calculer l'écart
                 '', '', '', '', '', '', '', '', '', '', '', ''],

                [motifs_data[5][0], motifs_data[5][1], motifs_data[5][2],
                 '=ABS(B23-C23)',  # Upr Nok (ligne 23) - Toujours calculer l'écart
                 '', '', '', '', '', '', '', '', '', '', '', ''],

                [motifs_data[6][0], motifs_data[6][1], motifs_data[6][2],
                 '=ABS(B24-C24)',  # Hors Commune (ligne 24) - Toujours calculer l'écart
                 '', '', '', '', '', '', '', '', '', '', '', ''],

                # Ligne de pourcentage d'écart Plan Adressage avec calcul total - CORRIGÉ
                ['% Ecart Plan Adressage', '=IF(SUM(B18:B24)=0,0,SUM(D18:D24)/SUM(B18:B24))', '', '', '', '', '', '', '', '', '', '', '', '', '', ''],  # ligne 26
            ])

            # Fin du tableau Ecart Plan Adressage - pas d'espacement supplémentaire

            # Fin des données de la page 1 - plus de commentaires indésirables

            # Page 2: Controle Qualité CMS - Structure selon capture d'écran
            page2_data = [
                # En-tête selon la nouvelle structure demandée
                ['ID Tache', 'Voie demandé', 'Motif Voie Initial', 'Motif Voie Corrigé', 'Commentaire Controleur']
            ]

            # Ajouter des lignes vides pour la saisie manuelle
            for i in range(25):  # 25 lignes vides pour saisie manuelle
                page2_data.append([' ', ' ', ' ', ' ', ' '])  # 5 colonnes selon la nouvelle structure

            # Page 3: Controle Qualité PA - Structure avec colonnes spécifiées selon nouvelle demande
            page3_data = [
                # En-tête avec la nouvelle structure : C=Batiment (QGis col G), E=Motif Initial (QGis col J), F=Motif Corrigé, G=Etat
                ['Num Dossier Site', 'Adresse Optimum', 'Batiment', 'Adresse BAN', 'Motif Initial', 'Motif Corrigé', 'Etat', 'Commentaire Controleur']
            ]

            # Créer un dictionnaire pour mapper les codes IMB aux données de bâtiment (colonne G fichier résultats QGis)
            imb_to_batiment = {}
            if hasattr(self, 'qgis_data') and self.qgis_data is not None:
                try:
                    pd = get_pandas()

                    self.logger.info(f"Fichier QGis chargé pour bâtiments: {self.qgis_data.shape}")

                    # Extraire les données des colonnes A (IMB) et G (Batiment) du fichier QGis
                    if len(self.qgis_data.columns) >= 7:  # Au moins 7 colonnes (A-G)
                        for index, row in self.qgis_data.iterrows():
                            imb_code = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''  # Colonne A: IMB
                            batiment = str(row.iloc[6]).strip() if pd.notna(row.iloc[6]) else ''  # Colonne G: Batiment

                            # Ajouter au dictionnaire si les deux valeurs sont présentes
                            if imb_code and imb_code not in ['', 'nan', 'IMB', 'Num Dossier Site']:
                                imb_to_batiment[imb_code] = batiment

                    self.logger.info(f"Mapping IMB->Batiment créé depuis QGis colonne G: {len(imb_to_batiment)} entrées")

                except Exception as e:
                    self.logger.error(f"Erreur lecture fichier QGis pour bâtiments: {e}")
                    imb_to_batiment = {}

            # Ajouter les données des CRITÈRES 3, 4 et 5 si disponibles
            toutes_erreurs_detectees = []

            # CRITÈRE 3: Doublons IMB et erreurs motif OK
            if self.qc_results and 'critere_3' in self.qc_results:
                critere_3 = self.qc_results['critere_3']
                doublons_suspects = critere_3.get('doublons_suspects', [])
                erreurs_motif_ok = critere_3.get('erreurs_motif_ok', [])
                toutes_erreurs_detectees.extend(doublons_suspects + erreurs_motif_ok)

            # CRITÈRE 4: AD à analyser
            if self.qc_results and 'critere_4' in self.qc_results:
                critere_4 = self.qc_results['critere_4']
                ad_a_analyser_entries = critere_4.get('ad_a_analyser_entries', [])
                toutes_erreurs_detectees.extend(ad_a_analyser_entries)

            # CRITÈRE 5: Motifs incorrects
            if self.qc_results and 'critere_5' in self.qc_results:
                critere_5 = self.qc_results['critere_5']
                motifs_incorrects_entries = critere_5.get('motifs_incorrects_entries', [])
                toutes_erreurs_detectees.extend(motifs_incorrects_entries)

            if toutes_erreurs_detectees:
                # Compter par type
                c3_doublons = len([e for e in toutes_erreurs_detectees if e.get('type') == 'DOUBLON_IMB_SUSPECT'])
                c3_motif_ok = len([e for e in toutes_erreurs_detectees if e.get('type') == 'ERREUR_MOTIF_OK_ADRESSE_IDENTIQUE'])
                c4_ad_analyser = len([e for e in toutes_erreurs_detectees if e.get('type') == 'AD_A_ANALYSER_AVEC_IMB'])
                c5_motifs_incorrects = len([e for e in toutes_erreurs_detectees if e.get('type') == 'MOTIF_INCORRECT'])

                self.logger.info(f"Ajout de {len(toutes_erreurs_detectees)} erreurs à la page 3 (C3: {c3_doublons} doublons + {c3_motif_ok} motif OK, C4: {c4_ad_analyser} ad à analyser, C5: {c5_motifs_incorrects} motifs incorrects)")

                for erreur in toutes_erreurs_detectees:
                    # Pour le critère 5 (motifs incorrects), utiliser le motif incorrect comme motif initial
                    motif_initial = erreur.get('motif_incorrect', '') if erreur.get('type') == 'MOTIF_INCORRECT' else erreur.get('motif_initial', '')

                    # Récupérer les données de bâtiment depuis le mapping IMB->Batiment
                    imb_code = erreur.get('imb_code', '')
                    batiment = imb_to_batiment.get(imb_code, '') if imb_code else ''

                    page3_data.append([
                        imb_code,                             # Colonne A: Num Dossier Site
                        erreur.get('adresse_optimum', ''),    # Colonne B: Adresse Optimum
                        batiment,                             # Colonne C: Batiment (depuis colonne G fichier QGis)
                        erreur.get('adresse_ban', ''),        # Colonne D: Adresse BAN
                        motif_initial,                        # Colonne E: Motif Initial (depuis colonne J fichier QGis)
                        '',                                    # Colonne F: Motif Corrigé (vide pour saisie avec validation)
                        '',                                    # Colonne G: Etat (vide pour saisie avec validation Faute Mineure/Faute Majeure)
                        ''                                     # Colonne H: Commentaire Controleur (vide pour saisie)
                    ])

            # Ajouter des lignes vides supplémentaires pour la saisie manuelle
            lignes_vides_necessaires = max(0, 20 - (len(page3_data) - 1))  # -1 pour l'en-tête
            for i in range(lignes_vides_necessaires):
                page3_data.append([' ', ' ', ' ', ' ', ' ', ' ', ' ', ' '])  # Espaces au lieu de chaînes vides (8 colonnes : A-H)

            # Préparer les données pour la feuille Ecart (Page 4)
            page4_data = self._prepare_ecart_data()

            return page1_data, page2_data, page3_data, page4_data

        except Exception as e:
            self.logger.error(f"Erreur préparation données Excel: {e}")
            raise

    def _prepare_ecart_data(self):
        """Prépare les données pour la feuille Ecart (Page 4) avec analyse détaillée par IMB."""
        try:
            pd = get_pandas()

            # En-tête de la feuille Ecart
            ecart_data = [
                ['ANALYSE DES ÉCARTS ENTRE FICHIERS QGIS ET SUIVI COMMUNE', '', '', '', '', ''],
                ['', '', '', '', '', ''],
                ['📊 SECTION 1: RÉSUMÉ PAR MOTIF', '', '', '', '', ''],
                ['Type d\'Écart', 'Fichier QGis', 'Suivi Commune', 'Différence', 'Détails', 'Statut'],
                ['', '', '', '', '', '']
            ]

            # Récupérer les résultats du critère 0 (Écart Plan Adressage)
            critere_0_results = self.qc_results.get('critere_0', {}) if self.qc_results else {}

            if critere_0_results.get('status') == 'COMPLETE':
                # Ajouter les écarts détectés par motif
                motifs_qgis = critere_0_results.get('motifs_qgis', {})
                motifs_suivi = critere_0_results.get('motifs_suivi', {})
                ecarts_detectes = critere_0_results.get('ecarts_detectes', [])

                # Tous les motifs analysés
                tous_motifs = set(motifs_qgis.keys()) | set(motifs_suivi.keys())

                for motif in sorted(tous_motifs):
                    count_qgis = motifs_qgis.get(motif, 0)
                    count_suivi = motifs_suivi.get(motif, 0)
                    difference = count_qgis - count_suivi

                    # Déterminer le statut
                    if difference == 0:
                        statut = "✅ OK"
                        details = "Aucun écart"
                    else:
                        statut = "❌ ÉCART"
                        if difference > 0:
                            details = f"+{difference} dans QGis"
                        else:
                            details = f"{difference} dans QGis"

                    ecart_data.append([
                        f"Motif: {motif}",
                        str(count_qgis),
                        str(count_suivi),
                        str(difference) if difference != 0 else "0",
                        details,
                        statut
                    ])

                # Ajouter une ligne de séparation
                ecart_data.append(['', '', '', '', '', ''])

                # Résumé des écarts par motif
                total_ecarts = len([e for e in ecarts_detectes if e.get('difference', 0) != 0])
                ecart_data.append([
                    'RÉSUMÉ MOTIFS',
                    f"Total écarts détectés: {total_ecarts}",
                    '',
                    '',
                    '',
                    '✅ OK' if total_ecarts == 0 else '❌ ATTENTION'
                ])

                # Ajouter des lignes de séparation avant la section détaillée
                ecart_data.extend([
                    ['', '', '', '', '', ''],
                    ['', '', '', '', '', ''],
                    ['🔍 SECTION 2: ANALYSE DÉTAILLÉE PAR IMB', '', '', '', '', ''],
                    ['Code IMB', 'Motif QGis', 'Motif Suivi', 'Statut Comparaison', 'Détails', 'Action'],
                    ['', '', '', '', '', '']
                ])

                # Ajouter l'analyse détaillée par IMB et récupérer les statistiques
                try:
                    imb_analysis, imb_stats = self._analyze_imb_level_gaps()
                    ecart_data.extend(imb_analysis)
                except Exception as e:
                    self.logger.error(f"Erreur analyse IMB détaillée: {e}")
                    # Ajouter un message d'erreur propre au lieu de planter
                    ecart_data.extend([
                        ['⚠️ ERREUR ANALYSE IMB', '', '', '', '', ''],
                        [f'Impossible d\'analyser les écarts IMB: {str(e)}', '', '', '', '', ''],
                        ['Vérifiez les fichiers et relancez l\'analyse', '', '', '', '', ''],
                        ['', '', '', '', '', '']
                    ])
                    imb_analysis = []
                    imb_stats = {}

                # Mettre à jour le résumé des motifs avec le nouveau calcul basé sur l'analyse IMB
                if imb_stats:
                    total_ecarts_reel = imb_stats.get('total_ecarts_reel', 0)
                    nb_donnees_manquantes = imb_stats.get('nb_donnees_manquantes', 0)
                    nb_motifs_differents = imb_stats.get('nb_motifs_differents', 0)

                    # Trouver et mettre à jour la ligne de résumé des motifs
                    for i, row in enumerate(ecart_data):
                        if row[0] == 'RÉSUMÉ MOTIFS':
                            ecart_data[i] = [
                                'RÉSUMÉ MOTIFS',
                                f"Total écarts détectés: {total_ecarts_reel} (Manquants: {nb_donnees_manquantes}, Différents: {nb_motifs_differents})",
                                '',
                                '',
                                '',
                                '✅ OK' if total_ecarts_reel == 0 else '❌ ATTENTION'
                            ]
                            break

            else:
                # Si l'analyse n'a pas été effectuée, on peut quand même faire l'analyse IMB
                ecart_data.append([
                    'ERREUR',
                    'Analyse par motif non effectuée',
                    '',
                    '',
                    'Veuillez d\'abord lancer l\'analyse qualité',
                    '⚠️ PENDING'
                ])

                # Ajouter des lignes de séparation avant la section détaillée
                ecart_data.extend([
                    ['', '', '', '', '', ''],
                    ['', '', '', '', '', ''],
                    ['🔍 SECTION 2: ANALYSE DÉTAILLÉE PAR IMB', '', '', '', '', ''],
                    ['Code IMB', 'Motif QGis', 'Motif Suivi', 'Statut Comparaison', 'Détails', 'Action'],
                    ['', '', '', '', '', '']
                ])

                # Ajouter l'analyse détaillée par IMB même sans critère 0
                imb_analysis, imb_stats = self._analyze_imb_level_gaps()
                ecart_data.extend(imb_analysis)

            # Ajouter des lignes vides pour compléter
            while len(ecart_data) < 25:
                ecart_data.append(['', '', '', '', '', ''])

            return ecart_data

        except Exception as e:
            self.logger.error(f"Erreur préparation données feuille Ecart: {e}")
            # Retourner une feuille d'erreur basique
            return [
                ['ERREUR GÉNÉRATION FEUILLE ÉCART', '', '', '', '', ''],
                ['', '', '', '', '', ''],
                [f'Erreur: {str(e)}', '', '', '', '', ''],
                ['', '', '', '', '', '']
            ]

    def _analyze_imb_level_gaps(self):
        """Analyse détaillée des écarts au niveau des codes IMB individuels."""
        try:
            pd = get_pandas()
            imb_analysis_data = []

            if self.qgis_data is None or self.suivi_data is None:
                return [
                    ['⚠️ DONNÉES MANQUANTES', '', '', '', '', ''],
                    ['Fichiers QGis ou Suivi Commune non chargés', '', '', '', '', ''],
                    ['Veuillez charger les fichiers requis', '', '', '', '', ''],
                    ['', '', '', '', '', '']
                ], {
                    'total_imb': 0,
                    'matches': 0,
                    'differences': 0,
                    'error': 'Données manquantes'
                }

            # Extraire les données QGis (colonne A = IMB, colonne J = motif)
            qgis_imb_motifs = {}
            if len(self.qgis_data.columns) > 9:  # Au moins colonne J (index 9)
                for index, row in self.qgis_data.iterrows():
                    imb_code = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''  # Colonne A
                    motif = str(row.iloc[9]).strip().upper() if pd.notna(row.iloc[9]) else ''  # Colonne J

                    if imb_code and imb_code not in ['', 'nan', 'NAN']:
                        if imb_code not in qgis_imb_motifs:
                            qgis_imb_motifs[imb_code] = []
                        qgis_imb_motifs[imb_code].append(motif)

            # Extraire les données Suivi Commune (page 2, colonne C = IMB, colonne I = motif)
            suivi_imb_motifs = {}
            try:
                # Vérifier que le chemin du fichier Suivi existe
                if not hasattr(self, 'current_suivi_file_path') or not self.current_suivi_file_path:
                    raise FileNotFoundError("Chemin du fichier Suivi Commune non défini")

                # Lire la page 2 (index 1) du fichier Suivi Commune
                suivi_page2_df = pd.read_excel(self.current_suivi_file_path, sheet_name=1, date_format=None)

                if len(suivi_page2_df.columns) > 8:  # Au moins colonne I (index 8)
                    for index, row in suivi_page2_df.iterrows():
                        imb_code = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ''  # Colonne C
                        motif = str(row.iloc[8]).strip().upper() if pd.notna(row.iloc[8]) else ''  # Colonne I

                        if imb_code and imb_code not in ['', 'nan', 'NAN']:
                            if imb_code not in suivi_imb_motifs:
                                suivi_imb_motifs[imb_code] = []
                            suivi_imb_motifs[imb_code].append(motif)

            except Exception as e:
                self.logger.error(f"Erreur lecture page 2 Suivi Commune: {e}")
                # Retourner des données vides plutôt qu'une erreur dans les données Excel
                return [
                    ['⚠️ ANALYSE IMPOSSIBLE', '', '', '', '', ''],
                    ['Fichier Suivi Commune non accessible', '', '', '', '', ''],
                    ['Veuillez vérifier le fichier et relancer l\'analyse', '', '', '', '', ''],
                    ['', '', '', '', '', '']
                ], {
                    'total_imb': 0,
                    'matches': 0,
                    'differences': 0,
                    'error': f'Erreur lecture Suivi: {str(e)}'
                }

            # Analyser les écarts pour chaque IMB
            all_imb_codes = set(qgis_imb_motifs.keys()) | set(suivi_imb_motifs.keys())

            matches = 0
            mismatches = 0
            missing_in_qgis = 0
            missing_in_suivi = 0

            for imb_code in sorted(all_imb_codes):
                qgis_motifs = qgis_imb_motifs.get(imb_code, [])
                suivi_motifs = suivi_imb_motifs.get(imb_code, [])

                # Cas 1: IMB présent dans les deux fichiers
                if qgis_motifs and suivi_motifs:
                    # Comparer les motifs (prendre le premier de chaque liste pour simplifier)
                    qgis_motif = qgis_motifs[0] if qgis_motifs else ''
                    suivi_motif = suivi_motifs[0] if suivi_motifs else ''

                    if qgis_motif == suivi_motif:
                        # Motifs identiques
                        statut = '✅ MATCH'
                        details = 'Motifs identiques'
                        action = 'Aucune'
                        matches += 1

                        # Gérer les doublons pour les MATCH
                        if len(qgis_motifs) > 1 or len(suivi_motifs) > 1:
                            details += f' (QGis: {len(qgis_motifs)} entrées, Suivi: {len(suivi_motifs)} entrées)'
                            statut = '⚠️ MATCH+DOUBLONS'
                            action = 'Vérifier doublons'
                            # Ajouter seulement les MATCH avec doublons (pas les MATCH simples)
                            imb_analysis_data.append([
                                imb_code,
                                qgis_motif,
                                suivi_motif,
                                statut,
                                details,
                                action
                            ])
                        # Ne pas ajouter les MATCH simples (sans doublons) à la liste

                    else:
                        # Motifs différents
                        statut = '❌ MISMATCH'
                        details = f'Motifs différents'
                        action = 'Vérifier et corriger'
                        mismatches += 1

                        # Gérer les doublons pour les MISMATCH
                        if len(qgis_motifs) > 1 or len(suivi_motifs) > 1:
                            details += f' (QGis: {len(qgis_motifs)} entrées, Suivi: {len(suivi_motifs)} entrées)'

                        imb_analysis_data.append([
                            imb_code,
                            qgis_motif,
                            suivi_motif,
                            statut,
                            details,
                            action
                        ])

                # Cas 2: IMB présent seulement dans QGis
                elif qgis_motifs and not suivi_motifs:
                    qgis_motif = qgis_motifs[0] if qgis_motifs else ''
                    imb_analysis_data.append([
                        imb_code,
                        qgis_motif,
                        'ABSENT',
                        '⚠️ MANQUANT SUIVI',
                        'IMB présent dans QGis mais absent du Suivi Commune',
                        'Ajouter dans Suivi'
                    ])
                    missing_in_suivi += 1

                # Cas 3: IMB présent seulement dans Suivi Commune
                elif not qgis_motifs and suivi_motifs:
                    suivi_motif = suivi_motifs[0] if suivi_motifs else ''
                    imb_analysis_data.append([
                        imb_code,
                        'ABSENT',
                        suivi_motif,
                        '⚠️ MANQUANT QGIS',
                        'IMB présent dans Suivi Commune mais absent de QGis',
                        'Ajouter dans QGis'
                    ])
                    missing_in_qgis += 1

            # Calculer le total des écarts selon la nouvelle logique
            nb_donnees_manquantes = missing_in_qgis + missing_in_suivi
            nb_motifs_differents = mismatches
            total_ecarts_reel = nb_donnees_manquantes + nb_motifs_differents

            # Ajouter un résumé de l'analyse détaillée
            if imb_analysis_data:
                imb_analysis_data.extend([
                    ['', '', '', '', '', ''],
                    ['=== RÉSUMÉ ANALYSE IMB ===', '', '', '', '', ''],
                    [f'Total IMB analysés: {len(all_imb_codes)}', '', '', '', '', ''],
                    [f'✅ Matches parfaits: {matches} (non affichés)', '', '', '', '', ''],
                    [f'❌ Mismatches: {mismatches}', '', '', '', '', ''],
                    [f'⚠️ Manquants QGis: {missing_in_qgis}', '', '', '', '', ''],
                    [f'⚠️ Manquants Suivi: {missing_in_suivi}', '', '', '', '', ''],
                    ['', '', '', '', '', ''],
                    [f'📊 ÉLÉMENTS AFFICHÉS: Seuls les problèmes sont listés ci-dessus', '', '', '', '', ''],
                    [f'📊 TOTAL ÉCARTS: {total_ecarts_reel} (Manquants: {nb_donnees_manquantes}, Différents: {nb_motifs_differents})', '', '', '', '', '']
                ])
            else:
                imb_analysis_data.append(['Aucun problème détecté - Tous les IMB sont en MATCH parfait', '', '', '', '', ''])

            # Préparer les statistiques pour le retour
            stats = {
                'matches': matches,
                'mismatches': mismatches,
                'missing_in_qgis': missing_in_qgis,
                'missing_in_suivi': missing_in_suivi,
                'total_imb': len(all_imb_codes),
                'nb_donnees_manquantes': nb_donnees_manquantes,
                'nb_motifs_differents': nb_motifs_differents,
                'total_ecarts_reel': total_ecarts_reel
            }

            return imb_analysis_data, stats

        except Exception as e:
            self.logger.error(f"Erreur analyse IMB détaillée: {e}")
            return [['Erreur analyse IMB', f'Erreur: {str(e)}', '', '', '', '']], {}

    def _write_excel_file(self, file_path: str, page1_data: list, page2_data: list, page3_data: list, page4_data: list) -> bool:
        """Écrit le fichier Excel avec 4 feuilles et mise en forme."""
        try:
            pd = get_pandas()

            # Obtenir les informations pour les noms de feuilles
            summary = self.qc_results.get('summary', {}) if self.qc_results else {}
            commune = summary.get('commune', 'Commune')
            collaborateur = summary.get('collaborateur', 'Collaborateur')

            # Nettoyer les données avant création des DataFrames
            page1_data_clean = self._clean_excel_data(page1_data)
            page2_data_clean = self._clean_excel_data(page2_data)
            page3_data_clean = self._clean_excel_data(page3_data)
            page4_data_clean = self._clean_excel_data(page4_data)

            # Créer les DataFrames avec colonnes appropriées et données nettoyées
            # Déterminer le nombre maximum de colonnes pour page1 (minimum 13 pour le décalage colonne I)
            max_cols_page1 = max(len(row) for row in page1_data_clean) if page1_data_clean else 13
            max_cols_page1 = max(max_cols_page1, 13)  # S'assurer qu'on a au moins 13 colonnes
            page1_columns = [f'Col{i+1}' for i in range(max_cols_page1)]
            df_page1 = pd.DataFrame(page1_data_clean, columns=page1_columns)
            df_page2 = pd.DataFrame(page2_data_clean, columns=['Col1', 'Col2', 'Col3', 'Col4', 'Col5'])  # 5 colonnes selon nouvelle structure CMS
            df_page3 = pd.DataFrame(page3_data_clean, columns=['Col1', 'Col2', 'Col3', 'Col4', 'Col5', 'Col6', 'Col7', 'Col8'])  # 8 colonnes maintenant (ajout colonne Batiment)
            df_page4 = pd.DataFrame(page4_data_clean, columns=['Col1', 'Col2', 'Col3', 'Col4', 'Col5', 'Col6'])  # 6 colonnes pour la feuille Ecart

            self.logger.info("DataFrames créés pour 4 feuilles")

            # Noms des feuilles selon la nouvelle structure (nettoyés pour Excel)
            commune_clean = self._clean_sheet_name(commune)
            collaborateur_clean = self._clean_sheet_name(collaborateur)

            sheet1_name = f"Etat_de_lieu_{commune_clean}_{collaborateur_clean}"
            sheet2_name = "Controle_Qualite_CMS"
            sheet3_name = "Controle_Qualite_PA"
            sheet4_name = "Ecart"

            # Limiter la longueur des noms de feuilles (Excel limite à 31 caractères)
            if len(sheet1_name) > 31:
                sheet1_name = f"Etat_{commune_clean[:8]}_{collaborateur_clean[:8]}"
                if len(sheet1_name) > 31:
                    sheet1_name = "Etat_de_lieu"

            # Écrire le fichier Excel avec mise en forme
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # Page 1: Etat de lieu avec mise en forme
                df_page1.to_excel(writer, sheet_name=sheet1_name, index=False, header=False)

                # Appliquer la mise en forme à la page 1
                self._format_page1(writer.sheets[sheet1_name])

                # Page 2: Controle Qualité CMS (avec structure)
                # S'assurer que toutes les lignes ont 5 colonnes (nouvelle structure CMS)
                page2_data_fixed = []
                for row in page2_data:
                    row_copy = row.copy() if isinstance(row, list) else list(row)
                    if len(row_copy) < 5:
                        row_copy.extend([' '] * (5 - len(row_copy)))  # Ajouter des espaces si nécessaire
                    page2_data_fixed.append(row_copy[:5])  # Limiter à 5 colonnes

                df_page2_fixed = pd.DataFrame(page2_data_fixed, columns=['Col1', 'Col2', 'Col3', 'Col4', 'Col5'])
                df_page2_fixed.to_excel(writer, sheet_name=sheet2_name, index=False, header=False)

                # Appliquer la mise en forme à la page 2
                self._format_page2(writer.sheets[sheet2_name])

                # Page 3: Controle Qualité PA (avec structure)
                # S'assurer que toutes les lignes ont 8 colonnes (structure étendue)
                page3_data_fixed = []
                for row in page3_data:
                    row_copy = row.copy() if isinstance(row, list) else list(row)
                    if len(row_copy) < 8:
                        row_copy.extend([' '] * (8 - len(row_copy)))  # Ajouter des espaces si nécessaire
                    page3_data_fixed.append(row_copy[:8])  # Limiter à 8 colonnes

                df_page3_fixed = pd.DataFrame(page3_data_fixed, columns=['Col1', 'Col2', 'Col3', 'Col4', 'Col5', 'Col6', 'Col7', 'Col8'])
                df_page3_fixed.to_excel(writer, sheet_name=sheet3_name, index=False, header=False)

                # Appliquer la mise en forme à la page 3
                self._format_page3(writer.sheets[sheet3_name])

                # Page 4: Feuille Ecart
                df_page4.to_excel(writer, sheet_name=sheet4_name, index=False, header=False)

                # Appliquer la mise en forme à la page 4
                self._format_page4_ecart(writer.sheets[sheet4_name])

                self.logger.info(f"4 feuilles Excel écrites: {sheet1_name}, {sheet2_name}, {sheet3_name}, {sheet4_name}")

            self.logger.info("Fichier Excel écrit avec succès avec mise en forme")
            return True

        except Exception as e:
            self.logger.error(f"Erreur écriture fichier Excel: {e}")
            return False

    def _clean_sheet_name(self, name: str) -> str:
        """Nettoie un nom pour l'utiliser comme nom de feuille Excel."""
        if not name:
            return "Sheet"

        # Remplacer les caractères interdits par des underscores
        forbidden_chars = ['\\', '/', '*', '?', ':', '[', ']']
        clean_name = str(name)
        for char in forbidden_chars:
            clean_name = clean_name.replace(char, '_')

        # Remplacer les espaces par des underscores
        clean_name = clean_name.replace(' ', '_')

        # Supprimer les caractères spéciaux supplémentaires
        clean_name = ''.join(c for c in clean_name if c.isalnum() or c in ['_', '-'])

        # Limiter la longueur
        if len(clean_name) > 20:
            clean_name = clean_name[:20]

        return clean_name if clean_name else "Sheet"

    def _clean_excel_data(self, data: list) -> list:
        """Nettoie les données pour éviter les erreurs Excel et normalise le nombre de colonnes."""
        if not data:
            return []

        # Déterminer le nombre maximum de colonnes
        max_cols = max(len(row) for row in data)

        cleaned_data = []

        for row in data:
            cleaned_row = []
            for i in range(max_cols):
                if i < len(row):
                    cell = row[i]
                    if cell is None:
                        cleaned_row.append('')
                    elif isinstance(cell, str):
                        # Nettoyer les caractères problématiques
                        clean_cell = cell.replace('\x00', '').replace('\r', '').replace('\n', ' ')
                        # Limiter la longueur des cellules (Excel limite à 32767 caractères)
                        if len(clean_cell) > 32000:
                            clean_cell = clean_cell[:32000] + "..."
                        cleaned_row.append(clean_cell)
                    elif isinstance(cell, (int, float)):
                        # Vérifier les valeurs numériques
                        if str(cell).lower() in ['inf', '-inf', 'nan']:
                            cleaned_row.append(0)
                        else:
                            cleaned_row.append(cell)
                    else:
                        cleaned_row.append(str(cell))
                else:
                    # Ajouter des cellules vides pour normaliser la longueur
                    cleaned_row.append('')

            cleaned_data.append(cleaned_row)

        return cleaned_data

    def _format_page4_ecart(self, worksheet):
        """Applique la mise en forme à la feuille Ecart (Page 4)."""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            # Styles de base
            title_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
            header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            data_font = Font(name='Arial', size=10)

            title_fill = PatternFill(start_color='2F4F4F', end_color='2F4F4F', fill_type='solid')  # Gris foncé
            header_fill = PatternFill(start_color='4682B4', end_color='4682B4', fill_type='solid')  # Bleu acier

            # Couleurs selon les spécifications
            match_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Vert pour ✅ MATCH
            match_doublons_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Jaune pour ⚠️ MATCH+DOUBLONS
            mismatch_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')  # Orangé pour ❌ MISMATCH
            manquant_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Rouge pour ⚠️ MANQUANT

            center_alignment = Alignment(horizontal='center', vertical='center')
            left_alignment = Alignment(horizontal='left', vertical='center')

            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Titre principal (ligne 1)
            for col in range(1, 7):  # A à F
                cell = worksheet.cell(row=1, column=col)
                cell.font = title_font
                cell.fill = title_fill
                cell.alignment = center_alignment
                cell.border = thin_border

            # Fusionner les cellules du titre
            worksheet.merge_cells('A1:F1')

            # Identifier et formater les en-têtes de sections
            for row_num in range(1, worksheet.max_row + 1):
                cell_value = worksheet.cell(row=row_num, column=1).value
                if cell_value and ('===' in str(cell_value) or 'Type d\'Écart' in str(cell_value) or 'Code IMB' in str(cell_value)):
                    # C'est un en-tête de section ou de colonne
                    for col in range(1, 7):
                        cell = worksheet.cell(row=row_num, column=col)
                        if 'SECTION' in str(cell_value):
                            # En-tête de section
                            cell.font = title_font
                            cell.fill = title_fill
                        else:
                            # En-tête de colonne
                            cell.font = header_font
                            cell.fill = header_fill
                        cell.alignment = center_alignment
                        cell.border = thin_border

            # Mise en forme des données
            for row_num in range(1, worksheet.max_row + 1):
                cell_value = worksheet.cell(row=row_num, column=1).value

                # Ignorer les lignes d'en-têtes déjà formatées
                if cell_value and ('===' in str(cell_value) or 'Type d\'Écart' in str(cell_value) or 'Code IMB' in str(cell_value)):
                    continue

                for col in range(1, 7):
                    cell = worksheet.cell(row=row_num, column=col)
                    cell.font = data_font
                    cell.border = thin_border

                    # Alignement selon la colonne
                    if col == 1:  # Première colonne (Type d'écart ou Code IMB)
                        cell.alignment = left_alignment
                    else:
                        cell.alignment = center_alignment

                    # Coloration selon le statut (colonnes F et D pour les différentes sections)
                    if col in [4, 6]:  # Colonnes de statut
                        if cell.value:
                            cell_val = str(cell.value)

                            # Appliquer les couleurs selon les spécifications
                            if '✅' in cell_val and 'MATCH' in cell_val and 'DOUBLONS' not in cell_val:
                                # Vert pour ✅ MATCH (sans doublons)
                                cell.fill = match_fill
                            elif '⚠️' in cell_val and 'MATCH+DOUBLONS' in cell_val:
                                # Jaune pour ⚠️ MATCH+DOUBLONS
                                cell.fill = match_doublons_fill
                            elif '❌' in cell_val and 'MISMATCH' in cell_val:
                                # Orangé pour ❌ MISMATCH
                                cell.fill = mismatch_fill
                            elif '⚠️' in cell_val and 'MANQUANT' in cell_val:
                                # Rouge pour ⚠️ MANQUANT
                                cell.fill = manquant_fill
                            elif '❌' in cell_val and 'ÉCART' in cell_val:
                                # Orangé pour les écarts de la section 1
                                cell.fill = mismatch_fill
                            elif '✅' in cell_val and 'OK' in cell_val:
                                # Vert pour les OK de la section 1
                                cell.fill = match_fill
                            elif '❌' in cell_val and 'ATTENTION' in cell_val:
                                # Rouge pour les résumés d'attention
                                cell.fill = manquant_fill

            # Ajuster la largeur des colonnes automatiquement (auto-fit précis)
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if cell.value is not None and str(cell.value).strip():
                            cell_value = str(cell.value)
                            cell_length = len(cell_value)

                            if cell.font and cell.font.bold:
                                cell_length = int(cell_length * 1.1)

                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass

                if max_length > 0:
                    excel_width = (max_length * 1.2) + 1
                    adjusted_width = min(max(excel_width, 5), 20)
                else:
                    adjusted_width = 5

                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Figer la première ligne et les en-têtes
            worksheet.freeze_panes = 'A4'

            self.logger.info("Mise en forme appliquée à la feuille Ecart")

        except Exception as e:
            self.logger.error(f"Erreur mise en forme feuille Ecart: {e}")

    def _format_page1(self, worksheet):
        """Applique la mise en forme optimisée à la page 1 avec styling Module 1."""
        try:
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

            # Couleurs identiques au Module 1 (suivi commune)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Bleu froid
            green_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            orange_fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
            purple_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
            light_blue_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
            light_green_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
            light_orange_fill = PatternFill(start_color="F2CC8F", end_color="F2CC8F", fill_type="solid")
            light_purple_fill = PatternFill(start_color="D5A6BD", end_color="D5A6BD", fill_type="solid")  # Violet clair
            cyan_fill = PatternFill(start_color="4BACC6", end_color="4BACC6", fill_type="solid")  # Cyan/Turquoise
            light_cyan_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")  # Cyan clair

            # Polices identiques au Module 1
            header_font = Font(color="FFFFFF", bold=True, size=11, name="Calibri")  # Blanc, gras pour en-têtes
            bold_font = Font(bold=True, size=11, name="Calibri")
            normal_font = Font(size=11, name="Calibri")
            title_font = Font(color="FFFFFF", bold=True, size=14, name="Calibri")

            # Alignement justifié (gauche) par défaut
            left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

            # Bordures fines
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # APPLIQUER L'ALIGNEMENT JUSTIFIÉ À TOUTES LES CELLULES
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = left_alignment
                    cell.font = normal_font

            # MISE EN FORME AVEC ESPACEMENT OPTIMISÉ (sans en-tête fusionné)
            # Ne colorer que les cellules qui contiennent des données

            # Tableau 1: Informations Générales (lignes 1-3)
            # Titre "INFORMATIONS GÉNÉRALES" - colonne A uniquement avec couleur bleue
            cell_a1 = worksheet.cell(row=1, column=1)
            if cell_a1.value is not None and str(cell_a1.value).strip():
                cell_a1.fill = header_fill  # Bleu
                cell_a1.font = header_font
                cell_a1.alignment = center_alignment

            # En-têtes du tableau (ligne 2) - colonnes A à F avec couleur bleue claire
            for col in range(1, 7):  # A2:F2
                cell = worksheet.cell(row=2, column=col)
                if cell.value is not None and str(cell.value).strip():
                    cell.fill = light_blue_fill  # Bleu clair
                    cell.font = bold_font
                    cell.alignment = center_alignment

            # Données du tableau (ligne 3) - colonnes A à F
            for col in range(1, 7):  # A3:F3
                cell = worksheet.cell(row=3, column=col)
                if cell.value is not None and str(cell.value).strip():
                    cell.alignment = center_alignment

            # Tableau 2: Section CMS (lignes 5-7) - Vert avec indicateurs
            # Titre "Qualité CMS Adresse" - colonne A uniquement avec couleur verte
            cell_a5 = worksheet.cell(row=5, column=1)
            if cell_a5.value is not None and str(cell_a5.value).strip():
                cell_a5.fill = green_fill  # Vert
                cell_a5.font = header_font
                cell_a5.alignment = center_alignment

            # En-têtes du tableau CMS (ligne 6) - colonnes A à D avec couleur verte claire
            for col in range(1, 5):  # A6:D6 (tableau CMS limité à 4 colonnes)
                cell = worksheet.cell(row=6, column=col)
                if cell.value is not None and str(cell.value).strip():
                    cell.fill = light_green_fill  # Vert clair
                    cell.font = bold_font
                    cell.alignment = center_alignment

            # Données du tableau CMS (ligne 7) - colonnes A à D
            for col in range(1, 5):  # A7:D7
                cell = worksheet.cell(row=7, column=col)
                if cell.value is not None and str(cell.value).strip():
                    cell.alignment = center_alignment

            # Tableau 3: Controle Plan Adressage (ligne 8) - Orange avec indicateurs - CORRIGÉ après suppression ligne vide
            # Titre "Controle Plan Adressage" - colonne A uniquement avec couleur orange
            cell_a8 = worksheet.cell(row=8, column=1)
            if cell_a8.value is not None and str(cell_a8.value).strip():
                cell_a8.fill = orange_fill  # Orange
                cell_a8.font = header_font
                cell_a8.alignment = center_alignment

            # En-têtes du tableau PA (ligne 9) - colonnes A à D avec couleur orange claire - CORRIGÉ
            for col in range(1, 5):  # A9:D9 (tableau PA limité à 4 colonnes)
                cell = worksheet.cell(row=9, column=col)
                if cell.value is not None and str(cell.value).strip():
                    cell.fill = light_orange_fill  # Orange clair
                    cell.font = bold_font
                    cell.alignment = center_alignment

            # Données du tableau PA (ligne 10) - colonnes A à D - CORRIGÉ
            for col in range(1, 5):  # A10:D10
                cell = worksheet.cell(row=10, column=col)
                if cell.value is not None and str(cell.value).strip():
                    cell.alignment = center_alignment

            # Tableau 4: Contrôle Dépose Tickets (ligne 12) - Violet avec statuts - CORRIGÉ après suppression ligne vide
            # Titre "Contrôle Dépose Tickets" - colonne A uniquement avec couleur violette
            cell_a12 = worksheet.cell(row=12, column=1)
            if cell_a12.value is not None and str(cell_a12.value).strip():
                cell_a12.fill = purple_fill  # Violet
                cell_a12.font = header_font
                cell_a12.alignment = center_alignment

            # En-têtes du tableau Tickets (ligne 13) - colonnes A à D avec couleur violette claire - CORRIGÉ
            for col in range(1, 5):  # A13:D13 (tableau Tickets limité à 4 colonnes)
                cell = worksheet.cell(row=13, column=col)
                if cell.value is not None and str(cell.value).strip():
                    cell.fill = light_purple_fill  # Violet clair
                    cell.font = bold_font
                    cell.alignment = center_alignment

            # Données du tableau Tickets (ligne 14) - colonnes A à D - CORRIGÉ
            for col in range(1, 5):  # A14:D14
                cell = worksheet.cell(row=14, column=col)
                if cell.value is not None and str(cell.value).strip():
                    cell.alignment = center_alignment

            # Tableau 5: Ecart Plan Adressage - Formatage dynamique
            # Trouver automatiquement la ligne où commence "Ecart Plan Adressage"
            ecart_title_row = None
            for row in range(1, 50):  # Chercher dans les 50 premières lignes
                cell = worksheet.cell(row=row, column=1)
                if cell.value and 'Ecart Plan Adressage' in str(cell.value):
                    ecart_title_row = row
                    break

            if ecart_title_row:
                self.logger.info(f"Tableau 'Ecart Plan Adressage' trouvé ligne {ecart_title_row}")

                # Titre "Ecart Plan Adressage" - colonne A uniquement avec couleur cyan
                cell_title = worksheet.cell(row=ecart_title_row, column=1)
                cell_title.fill = cyan_fill  # Cyan/Turquoise
                cell_title.font = header_font
                cell_title.alignment = center_alignment

                # En-têtes du tableau Ecart (ligne suivante) - colonnes A à D avec couleur cyan claire
                headers_row = ecart_title_row + 1
                for col in range(1, 5):  # A:D (tableau Ecart limité à 4 colonnes)
                    cell = worksheet.cell(row=headers_row, column=col)
                    if cell.value is not None and str(cell.value).strip():
                        cell.fill = light_cyan_fill  # Cyan clair
                        cell.font = bold_font
                        cell.alignment = center_alignment

                # Données du tableau Ecart (7 lignes de motifs) - colonnes A à D
                for row_offset in range(2, 9):  # 7 motifs (Ad Ras, Ok, Nok, Upr Ras, Upr Ok, Upr Nok, Hors Commune)
                    data_row = ecart_title_row + row_offset
                    for col in range(1, 5):  # A:D
                        cell = worksheet.cell(row=data_row, column=col)
                        if cell.value is not None and str(cell.value).strip():
                            cell.alignment = center_alignment

                # Ligne "% Ecart Plan Adressage" (après les 7 motifs) - même couleur cyan que le titre
                percentage_row = ecart_title_row + 9  # Titre + En-têtes + 7 motifs = +9
                cell_percentage = worksheet.cell(row=percentage_row, column=1)
                if cell_percentage.value and '% Ecart Plan Adressage' in str(cell_percentage.value):
                    cell_percentage.fill = cyan_fill  # Cyan/Turquoise (même couleur que le titre)
                    cell_percentage.font = header_font
                    cell_percentage.alignment = center_alignment

                    # Formule du pourcentage avec formatage
                    cell_formula = worksheet.cell(row=percentage_row, column=2)
                    if cell_formula.value is not None and str(cell_formula.value).strip():
                        cell_formula.fill = light_cyan_fill  # Cyan clair
                        cell_formula.alignment = center_alignment

                self.logger.info(f"Formatage du tableau Ecart Plan Adressage appliqué (lignes {ecart_title_row}-{percentage_row})")
            else:
                self.logger.warning("Tableau 'Ecart Plan Adressage' non trouvé pour le formatage")

            # Données de l'analyse détaillée (lignes 22-28) - CORRIGÉ après suppression ligne vide
            for row in range(22, 29):  # Lignes 22-28
                for col in range(1, 6):  # A22:E28
                    cell = worksheet.cell(row=row, column=col)
                    if cell.value is not None and str(cell.value).strip():
                        cell.alignment = center_alignment

            # Tableau 6: Section Résumé Erreurs (lignes 5-11, colonnes G-K) - Corrigé après suppression ligne vide
            # Titre "Résumé Erreurs" seulement dans la colonne G (ligne 5)
            cell = worksheet.cell(row=5, column=7)  # G5
            if cell.value is not None and str(cell.value).strip():
                cell.fill = header_fill  # Fond bleu foncé comme dans le modèle
                cell.font = header_font  # Police blanche, gras
                cell.alignment = left_alignment

            # En-têtes du résumé erreurs (ligne 6, colonnes G-K)
            for col in range(7, 12):  # G6:K6
                cell = worksheet.cell(row=6, column=col)
                if cell.value is not None and str(cell.value).strip():
                    cell.fill = light_blue_fill  # Fond bleu clair comme dans le modèle
                    cell.font = bold_font
                    cell.alignment = left_alignment

            # Données du résumé erreurs (lignes 7-10, colonnes G-K)
            for row in range(7, 11):  # Lignes 7-10 (% Erreur CMS, PA, Banbou, Ecart)
                for col in range(7, 12):  # G7:K10
                    cell = worksheet.cell(row=row, column=col)
                    if cell.value is not None and str(cell.value).strip():
                        cell.alignment = left_alignment
                        # Coloration spéciale pour la colonne statut commune (K)
                        if col == 11:  # Colonne K (Statut Commune)
                            cell.fill = light_orange_fill

            # Ligne SCORE TOTAL (ligne 11, colonnes G-K) - Fond bleu foncé comme le titre
            for col in range(7, 12):  # G11:K11
                cell = worksheet.cell(row=11, column=col)
                if cell.value is not None and str(cell.value).strip():
                    if col == 11:  # Colonne K11 (Statut Commune) - formatage spécial
                        cell.fill = light_orange_fill  # Fond orange pour le statut
                        cell.font = bold_font  # Police noire, gras
                        cell.alignment = center_alignment  # Centré pour le statut
                    else:
                        cell.fill = header_fill  # Fond bleu foncé comme dans le modèle
                        cell.font = header_font  # Police blanche, gras
                        cell.alignment = left_alignment

            # Ajouter la validation de données pour la colonne Contrôleur (F3)
            from openpyxl.worksheet.datavalidation import DataValidation
            from config.constants import VALIDATION_LISTS

            # Liste des collaborateurs depuis constants.py
            collaborateurs = VALIDATION_LISTS.get("Collaborateur", [])
            collaborateurs_list = '"' + ','.join(collaborateurs) + '"'
            dv_controleur = DataValidation(type="list", formula1=collaborateurs_list, allow_blank=True)
            dv_controleur.error = "Veuillez sélectionner un collaborateur valide"
            dv_controleur.errorTitle = "Contrôleur incorrect"
            dv_controleur.prompt = "Sélectionnez le contrôleur responsable"
            dv_controleur.promptTitle = "Contrôleur"

            # Appliquer la validation à la cellule F3 (Contrôleur)
            dv_controleur.add("F3")
            worksheet.add_data_validation(dv_controleur)

            # Ajuster les largeurs de colonnes automatiquement (auto-fit amélioré pour toute la page)
            from openpyxl.utils import get_column_letter

            # Parcourir toutes les colonnes utilisées (A à P pour couvrir toute la page)
            for col_num in range(1, 17):  # Colonnes A à P
                column_letter = get_column_letter(col_num)
                max_length = 0

                # Parcourir toutes les lignes utilisées (jusqu'à ligne 30 pour couvrir tous les tableaux)
                for row_num in range(1, 31):
                    try:
                        cell = worksheet.cell(row=row_num, column=col_num)
                        if cell.value is not None and str(cell.value).strip():
                            # Calculer la largeur en tenant compte du formatage
                            cell_value = str(cell.value)
                            cell_length = len(cell_value)

                            # Facteur de correction pour Excel selon le formatage
                            if cell.font and cell.font.bold:
                                cell_length = int(cell_length * 1.15)  # Texte gras plus large

                            # Facteur supplémentaire pour les en-têtes colorés
                            if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb != "00000000":
                                cell_length = int(cell_length * 1.1)  # En-têtes colorés plus larges

                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass

                # Calcul optimisé de la largeur Excel
                if max_length > 0:
                    # Formule Excel améliorée avec marge de sécurité
                    excel_width = (max_length * 1.3) + 2
                    # Limiter entre 8 et 25 pour un meilleur affichage
                    adjusted_width = min(max(excel_width, 8), 25)
                else:
                    adjusted_width = 8  # Largeur minimale pour colonnes vides

                worksheet.column_dimensions[column_letter].width = adjusted_width

            # SUPPRESSION DES FUSIONS : Les titres occupent uniquement la colonne A
            # Plus de fusion de cellules pour les titres des tableaux
            # Chaque titre reste dans sa cellule A respective

            # Geler la première ligne (comme Module 1)
            worksheet.freeze_panes = 'A2'

            # Ajouter une validation des données pour la cellule Contrôleur (F3)
            self._add_controleur_validation(worksheet)

            # Appliquer le formatage des pourcentages aux cellules avec formules
            self._apply_percentage_formatting(worksheet)

            # Appliquer le formatage numérique au tableau Ecart Plan Adressage
            self._apply_numeric_formatting(worksheet)

            # Forcer le formatage des pourcentages après l'écriture des données
            self._force_percentage_formatting(worksheet)

            self.logger.info("Mise en forme optimisée appliquée à la page 1 (styling Module 1)")

        except Exception as e:
            self.logger.warning(f"Erreur lors de la mise en forme: {e}")
            # Continue sans mise en forme si erreur

    def _add_controleur_validation(self, worksheet):
        """Ajoute une validation des données pour la cellule Contrôleur."""
        try:
            from openpyxl.worksheet.datavalidation import DataValidation

            # Importer la liste des collaborateurs depuis constants.py
            from config.constants import VALIDATION_LISTS
            collaborateurs_list = VALIDATION_LISTS.get("Collaborateur", [])

            if not collaborateurs_list:
                self.logger.warning("Liste des collaborateurs vide, validation ignorée")
                return

            # Créer une feuille cachée pour la liste des collaborateurs
            # Cela évite les problèmes avec les formules longues
            if 'ValidationData' not in [ws.title for ws in worksheet.parent.worksheets]:
                validation_sheet = worksheet.parent.create_sheet('ValidationData')
                validation_sheet.sheet_state = 'hidden'

                # Écrire la liste des collaborateurs dans la feuille cachée
                for i, collab in enumerate(collaborateurs_list, 1):
                    validation_sheet.cell(row=i, column=1, value=collab)

                # Créer la validation des données avec référence à la feuille cachée
                range_ref = f"ValidationData!$A$1:$A${len(collaborateurs_list)}"
                dv = DataValidation(
                    type="list",
                    formula1=range_ref,
                    allow_blank=True
                )
            else:
                # Si la feuille existe déjà, utiliser la référence existante
                range_ref = f"ValidationData!$A$1:$A${len(collaborateurs_list)}"
                dv = DataValidation(
                    type="list",
                    formula1=range_ref,
                    allow_blank=True
                )

            dv.error = "Veuillez sélectionner un contrôleur dans la liste."
            dv.errorTitle = "Contrôleur invalide"
            dv.prompt = "Sélectionnez un contrôleur dans la liste déroulante."
            dv.promptTitle = "Sélection du contrôleur"

            # Appliquer la validation à la cellule F3 (Contrôleur - nouvelle position)
            dv.add('F3')
            worksheet.add_data_validation(dv)

            self.logger.info("Validation des données ajoutée pour le contrôleur")

        except Exception as e:
            self.logger.warning(f"Erreur lors de l'ajout de la validation des données: {e}")
            # Continue sans validation si erreur

    def _apply_percentage_formatting(self, worksheet):
        """Applique le formatage des pourcentages aux cellules contenant des formules de pourcentage."""
        try:
            # Appliquer le formatage des pourcentages aux cellules avec formules de pourcentage - AMÉLIORÉ
            percentage_cells = [
                'C7',   # % Erreur CMS (position ligne 7)
                'C10',  # % Erreur PA (position ligne 10 - CORRIGÉ après suppression ligne vide)
                'C14',  # % Erreur Banbou (position ligne 14 - affichage)
                'H7',   # % Brut CMS (Résumé Erreurs - colonne H, ligne 7)
                'H8',   # % Brut PA (Résumé Erreurs - colonne H, ligne 8)
                'H9',   # % Brut Banbou (Résumé Erreurs - colonne H, ligne 9)
                'H10',  # % Brut Ecart Plan Adressage (Résumé Erreurs - colonne H, ligne 10)
                'J7',   # Score CMS (Résumé Erreurs - colonne J, ligne 7)
                'J8',   # Score PA (Résumé Erreurs - colonne J, ligne 8)
                'J9',   # Score Banbou (Résumé Erreurs - colonne J, ligne 9)
                'J10',  # Score Ecart Plan Adressage (Résumé Erreurs - colonne J, ligne 10)
                'J11',  # Score Total (Résumé Erreurs - colonne J, ligne 11)
            ]

            # Ajouter dynamiquement la cellule du pourcentage Ecart Plan Adressage
            ecart_percentage_row = None
            for row in range(1, 50):
                cell = worksheet.cell(row=row, column=1)
                if cell.value and '% Ecart Plan Adressage' in str(cell.value):
                    ecart_percentage_row = row
                    break

            if ecart_percentage_row:
                percentage_cells.append(f'B{ecart_percentage_row}')  # Formule du pourcentage
                self.logger.info(f"Cellule pourcentage Ecart Plan Adressage trouvée: B{ecart_percentage_row}")

            # Appliquer le formatage avec une approche plus robuste
            formatted_count = 0
            for cell_ref in percentage_cells:
                try:
                    cell = worksheet[cell_ref]
                    if cell:
                        # Appliquer le formatage pourcentage directement
                        cell.number_format = '0.00%'  # Format avec 2 décimales pour plus de précision
                        formatted_count += 1
                        self.logger.debug(f"Formatage appliqué à {cell_ref}: {cell.value}")
                except Exception as cell_error:
                    self.logger.warning(f"Erreur formatage cellule {cell_ref}: {cell_error}")
                    continue

            self.logger.info(f"Formatage des pourcentages appliqué à {formatted_count} cellules")

        except Exception as e:
            self.logger.warning(f"Erreur lors du formatage des pourcentages: {e}")
            # Continue sans formatage si erreur

    def _apply_numeric_formatting(self, worksheet):
        """Applique le formatage numérique aux cellules du tableau Ecart Plan Adressage."""
        try:
            # Trouver dynamiquement la ligne où commence "Ecart Plan Adressage"
            ecart_title_row = None
            for row in range(1, 50):
                cell = worksheet.cell(row=row, column=1)
                if cell.value and 'Ecart Plan Adressage' in str(cell.value):
                    ecart_title_row = row
                    break

            if ecart_title_row:
                # Formater les colonnes Suivi (B) et QGis (C) en format numérique
                # Les données commencent à ecart_title_row + 2 (titre + en-têtes)
                data_start_row = ecart_title_row + 2
                data_end_row = data_start_row + 6  # 7 motifs (0-6)

                for row in range(data_start_row, data_end_row + 1):
                    # Colonne B (Suivi) - format numérique
                    cell_b = worksheet.cell(row=row, column=2)
                    if cell_b.value is not None:
                        cell_b.number_format = '0'  # Format entier

                    # Colonne C (QGis) - format numérique
                    cell_c = worksheet.cell(row=row, column=3)
                    if cell_c.value is not None:
                        cell_c.number_format = '0'  # Format entier

                    # Colonne D (Écart) - format numérique (formule)
                    cell_d = worksheet.cell(row=row, column=4)
                    if cell_d.value is not None:
                        cell_d.number_format = '0'  # Format entier

                self.logger.info(f"Formatage numérique appliqué au tableau Ecart Plan Adressage (lignes {data_start_row}-{data_end_row})")
            else:
                self.logger.warning("Tableau 'Ecart Plan Adressage' non trouvé pour le formatage numérique")

        except Exception as e:
            self.logger.warning(f"Erreur lors du formatage numérique: {e}")
            # Continue sans formatage si erreur

    def _force_percentage_formatting(self, worksheet):
        """Force le formatage des pourcentages sur toutes les cellules concernées après écriture des données."""
        try:
            # Liste exhaustive de toutes les cellules qui doivent être en format pourcentage
            all_percentage_cells = []

            # Parcourir toutes les cellules pour trouver celles qui contiennent des pourcentages
            for row in range(1, 50):  # Limiter la recherche aux 50 premières lignes
                for col in range(1, 15):  # Limiter aux 15 premières colonnes
                    cell = worksheet.cell(row=row, column=col)
                    if cell.value is not None:
                        cell_value = str(cell.value).strip()

                        # Détecter les formules de pourcentage ou les valeurs décimales qui devraient être des pourcentages
                        if (cell_value.startswith('=') and ('/' in cell_value or 'SUM' in cell_value)) or \
                           (cell_value.replace(',', '.').replace('-', '').replace('+', '').replace('.', '').isdigit() and
                            '0,' in cell_value and float(cell_value.replace(',', '.')) < 1):

                            # Vérifier si c'est dans une colonne/ligne de pourcentage
                            cell_ref = f"{chr(64 + col)}{row}"  # Convertir en référence A1

                            # Colonnes connues pour contenir des pourcentages
                            if col in [3, 8, 10] or 'Erreur' in str(worksheet.cell(row=row-1, column=col).value or '') or \
                               'Brut' in str(worksheet.cell(row=row-1, column=col).value or '') or \
                               'Score' in str(worksheet.cell(row=row-1, column=col).value or ''):
                                all_percentage_cells.append(cell_ref)

            # Ajouter les cellules spécifiques connues
            specific_cells = ['C7', 'C10', 'C14', 'H7', 'H8', 'H9', 'H10', 'J7', 'J8', 'J9', 'J10', 'J11']
            all_percentage_cells.extend(specific_cells)

            # Trouver la cellule % Ecart Plan Adressage
            for row in range(1, 50):
                cell = worksheet.cell(row=row, column=1)
                if cell.value and '% Ecart Plan Adressage' in str(cell.value):
                    all_percentage_cells.append(f'B{row}')
                    break

            # Supprimer les doublons
            all_percentage_cells = list(set(all_percentage_cells))

            # Appliquer le formatage pourcentage
            formatted_count = 0
            for cell_ref in all_percentage_cells:
                try:
                    cell = worksheet[cell_ref]
                    if cell and cell.value is not None:
                        # Forcer le formatage pourcentage
                        cell.number_format = '0.00%'
                        formatted_count += 1
                except Exception as cell_error:
                    self.logger.debug(f"Erreur formatage forcé cellule {cell_ref}: {cell_error}")
                    continue

            self.logger.info(f"Formatage pourcentage forcé appliqué à {formatted_count} cellules")

        except Exception as e:
            self.logger.warning(f"Erreur lors du formatage forcé des pourcentages: {e}")
            # Continue sans formatage si erreur

    def _format_page3(self, worksheet):
        """Applique la mise en forme à la page 3 - Contrôle Qualité PA."""
        try:
            from openpyxl.styles import PatternFill, Font, Alignment

            # Définir les styles
            blue_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            white_font = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
            center_alignment = Alignment(horizontal="center", vertical="center")

            # Mise en forme de l'en-tête (ligne 1) - seulement les cellules avec contenu (A-H)
            header_columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
            for col_letter in header_columns:
                cell = worksheet[f"{col_letter}1"]
                if cell.value is not None and str(cell.value).strip():
                    cell.fill = blue_fill
                    cell.font = white_font
                    cell.alignment = center_alignment

            # Ajuster la largeur des colonnes automatiquement (auto-fit précis)
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if cell.value is not None and str(cell.value).strip():
                            cell_value = str(cell.value)
                            cell_length = len(cell_value)

                            if cell.font and cell.font.bold:
                                cell_length = int(cell_length * 1.1)

                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass

                if max_length > 0:
                    excel_width = (max_length * 1.2) + 1
                    # Max 25 pour les adresses longues
                    adjusted_width = min(max(excel_width, 5), 25)
                else:
                    adjusted_width = 5

                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Ajouter la validation de données pour la colonne F (Motif Corrigé)
            from openpyxl.worksheet.datavalidation import DataValidation

            # Validation pour Motif Corrigé (colonne F)
            motif_options = '"AD RAS,OK,NOK,UPR RAS,UPR OK,UPR NOK,Hors Commune"'
            dv_motif = DataValidation(type="list", formula1=motif_options, allow_blank=True)
            dv_motif.error = "Veuillez sélectionner un motif valide"
            dv_motif.errorTitle = "Motif incorrect"
            dv_motif.prompt = "Sélectionnez: AD RAS, OK, NOK, UPR RAS, UPR OK, UPR NOK, Hors Commune"
            dv_motif.promptTitle = "Motif Corrigé"
            dv_motif.add(f"F2:F22")
            worksheet.add_data_validation(dv_motif)

            # Validation pour Etat (colonne G)
            etat_options = '"Faute Mineure,Faute Majeure"'
            dv_etat = DataValidation(type="list", formula1=etat_options, allow_blank=True)
            dv_etat.error = "Veuillez sélectionner une option valide"
            dv_etat.errorTitle = "Valeur incorrecte"
            dv_etat.prompt = "Sélectionnez: Faute Mineure ou Faute Majeure"
            dv_etat.promptTitle = "Etat"
            dv_etat.add(f"G2:G22")
            worksheet.add_data_validation(dv_etat)

            # Appliquer l'alignement centré seulement aux cellules avec contenu
            for row in range(1, 22):  # Lignes 1 à 21 (en-tête + 20 lignes de données)
                for col in range(1, 9):  # Colonnes A à H (8 colonnes maintenant)
                    cell = worksheet.cell(row=row, column=col)
                    if cell.value is not None and str(cell.value).strip():
                        cell.alignment = center_alignment

            self.logger.info("Mise en forme appliquée à la page 3 - Contrôle Qualité PA")

        except Exception as e:
            self.logger.warning(f"Erreur lors de la mise en forme de la page 3: {e}")
            # Continue sans mise en forme si erreur

    def _format_page2(self, worksheet):
        """Applique la mise en forme à la page 2 - Contrôle Qualité CMS avec validation des données."""
        try:
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.worksheet.datavalidation import DataValidation

            # Définir les styles améliorés
            header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")  # Bleu professionnel
            white_font = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
            center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            left_alignment = Alignment(horizontal="left", vertical="center")

            # Bordures pour structure professionnelle
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Mise en forme de l'en-tête (ligne 1) - seulement les cellules avec contenu
            header_columns = ['A', 'B', 'C', 'D', 'E']
            for col_letter in header_columns:
                cell = worksheet[f"{col_letter}1"]
                if cell.value is not None and str(cell.value).strip():
                    cell.fill = header_fill
                    cell.font = white_font
                    cell.alignment = center_alignment
                    cell.border = thin_border

            # Ajuster la largeur des colonnes automatiquement (auto-fit précis)
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if cell.value is not None and str(cell.value).strip():
                            cell_value = str(cell.value)
                            cell_length = len(cell_value)

                            if cell.font and cell.font.bold:
                                cell_length = int(cell_length * 1.1)

                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass

                if max_length > 0:
                    excel_width = (max_length * 1.2) + 1
                    # Max 30 pour les descriptions longues
                    adjusted_width = min(max(excel_width, 5), 30)
                else:
                    adjusted_width = 5

                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Ajouter les validations de données selon les spécifications

            # Validation pour Motif Voie Initial (colonne C)
            motif_voie_options = '"Rien à faire,Création Voie,Modification Voie"'
            dv_motif_initial = DataValidation(type="list", formula1=motif_voie_options, allow_blank=True)
            dv_motif_initial.error = "Veuillez sélectionner un motif valide"
            dv_motif_initial.errorTitle = "Motif incorrect"
            dv_motif_initial.prompt = "Sélectionnez le motif voie initial"
            dv_motif_initial.promptTitle = "Motif Voie Initial"
            dv_motif_initial.add("C2:C26")  # 25 lignes de données
            worksheet.add_data_validation(dv_motif_initial)

            # Validation pour Motif Voie Corrigé (colonne D)
            dv_motif_corrige = DataValidation(type="list", formula1=motif_voie_options, allow_blank=True)
            dv_motif_corrige.error = "Veuillez sélectionner un motif valide"
            dv_motif_corrige.errorTitle = "Motif incorrect"
            dv_motif_corrige.prompt = "Sélectionnez le motif voie corrigé"
            dv_motif_corrige.promptTitle = "Motif Voie Corrigé"
            dv_motif_corrige.add("D2:D26")  # 25 lignes de données
            worksheet.add_data_validation(dv_motif_corrige)

            # Appliquer les styles et alignements seulement aux cellules avec contenu
            for row in range(1, 27):  # Lignes 1 à 26 (en-tête + 25 lignes de données)
                for col in range(1, 6):  # Colonnes A à E (5 colonnes)
                    cell = worksheet.cell(row=row, column=col)

                    if cell.value is not None and str(cell.value).strip():
                        cell.border = thin_border

                        if row == 1:  # En-tête déjà formaté
                            continue
                        elif col in [1, 3, 4]:  # Colonnes centrées (ID Tache, Motif Initial, Motif Corrigé)
                            cell.alignment = center_alignment
                        else:  # Colonnes alignées à gauche (Voie demandé, Commentaire)
                            cell.alignment = left_alignment

            # Ajouter des couleurs conditionnelles pour les motifs
            from openpyxl.formatting.rule import CellIsRule

            # Couleurs pour les motifs
            creation_fill = PatternFill(start_color="E6F7E6", end_color="E6F7E6", fill_type="solid")      # Création Voie
            modification_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")   # Modification Voie
            rien_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")          # Rien à faire

            # Règles de formatage conditionnel pour les colonnes Motif (C et D)
            worksheet.conditional_formatting.add("C2:C26", CellIsRule(operator="equal", formula=['"Création Voie"'], fill=creation_fill))
            worksheet.conditional_formatting.add("C2:C26", CellIsRule(operator="equal", formula=['"Modification Voie"'], fill=modification_fill))
            worksheet.conditional_formatting.add("C2:C26", CellIsRule(operator="equal", formula=['"Rien à faire"'], fill=rien_fill))

            worksheet.conditional_formatting.add("D2:D26", CellIsRule(operator="equal", formula=['"Création Voie"'], fill=creation_fill))
            worksheet.conditional_formatting.add("D2:D26", CellIsRule(operator="equal", formula=['"Modification Voie"'], fill=modification_fill))
            worksheet.conditional_formatting.add("D2:D26", CellIsRule(operator="equal", formula=['"Rien à faire"'], fill=rien_fill))

            # Figer la première ligne pour faciliter la navigation
            worksheet.freeze_panes = 'A2'

            # Ajouter un filtre automatique sur l'en-tête
            worksheet.auto_filter.ref = "A1:E26"

            self.logger.info("Mise en forme avancée de la page 2 CMS appliquée avec succès")

        except Exception as e:
            self.logger.warning(f"Erreur lors de la mise en forme de la page 2: {e}")
            # Continue sans mise en forme si erreur





    def _calculate_error_percentage(self) -> float:
        """Calcule le pourcentage d'erreur."""
        try:
            summary = self.qc_results.get('summary', {})
            total_errors = summary.get('total_errors', 0)

            # Base de calcul (nombre total d'enregistrements analysés)
            total_records = 0
            if self.qgis_data is not None:
                total_records += len(self.qgis_data)
            if self.suivi_data is not None:
                total_records += len(self.suivi_data)

            if total_records > 0:
                return (total_errors / total_records) * 100
            return 0.0

        except Exception:
            return 0.0

    def _calculate_erreur_banbou_percentage(self) -> float:
        """
        Calcule le % Erreur Banbou basé sur les statuts des tickets UPR et 501/511.

        Logique:
        - Si les deux tickets sont OK ou N/A → 0%
        - Si les deux tickets sont NOK → 100%
        - Si un OK et un NOK → 50%

        Returns:
            float: Pourcentage d'erreur Banbou (0.0 à 100.0)
        """
        try:
            if not self.qc_results or 'critere_2' not in self.qc_results:
                return 0.0

            critere_2 = self.qc_results['critere_2']
            ticket_upr_status = critere_2.get('ticket_upr_status', 'N/A')
            ticket_501_511_status = critere_2.get('ticket_501_511_status', 'N/A')

            # Compter les tickets NOK
            nok_count = 0
            total_tickets = 0

            # Analyser ticket UPR
            if ticket_upr_status in ['OK', 'NOK']:
                total_tickets += 1
                if ticket_upr_status == 'NOK':
                    nok_count += 1

            # Analyser ticket 501/511
            if ticket_501_511_status in ['OK', 'NOK']:
                total_tickets += 1
                if ticket_501_511_status == 'NOK':
                    nok_count += 1

            # Calculer le pourcentage
            if total_tickets == 0:
                # Aucun ticket requis (tous N/A) → 0% d'erreur
                percentage = 0.0
            else:
                # Pourcentage = (nombre de NOK / total de tickets) * 100
                percentage = (nok_count / total_tickets) * 100.0

            self.logger.info(f"Calcul % Erreur Banbou: UPR={ticket_upr_status}, 501/511={ticket_501_511_status}, "
                           f"NOK={nok_count}/{total_tickets}, Résultat={percentage}%")

            return percentage

        except Exception as e:
            self.logger.error(f"Erreur calcul % Erreur Banbou: {e}")
            return 0.0

    def _evaluate_commune_status(self) -> Dict[str, Any]:
        """
        Évalue le statut de conformité de la commune selon les règles de validation.

        RÈGLES DE VALIDATION :
        1. Note globale < 90% → KO
        2. Fautes majeures détectées → KO immédiat

        Returns:
            Dict contenant le statut, pourcentage de conformité et raisons du KO
        """
        try:
            if not self.qc_results:
                return {
                    'statut': 'N/A',
                    'pourcentage_conformite': 0.0,
                    'raisons_ko': ['Aucune analyse effectuée'],
                    'fautes_majeures': [],
                    'note_globale_ko': False
                }

            # 1. CALCUL DU POURCENTAGE DE CONFORMITÉ GLOBAL
            pourcentage_conformite = self._calculate_conformite_percentage()

            # 2. DÉTECTION DES FAUTES MAJEURES
            fautes_majeures = self._detect_fautes_majeures()

            # 3. ÉVALUATION DU STATUT
            raisons_ko = []
            note_globale_ko = False
            fichiers_manquants = False

            # Vérifier les fautes majeures critiques (fichiers manquants) en priorité
            fautes_fichiers = [f for f in fautes_majeures if f['type'] in ['MANQUANT_QGIS', 'MANQUANT_SUIVI', 'STRUCTURE_QGIS_INVALIDE', 'STRUCTURE_SUIVI_INVALIDE']]
            if fautes_fichiers:
                fichiers_manquants = True
                for faute in fautes_fichiers:
                    raisons_ko.append(f"FICHIER CRITIQUE: {faute['type']} - {faute['description']}")

            # Vérifier le seuil de conformité (90%) seulement si les fichiers sont présents
            if not fichiers_manquants and pourcentage_conformite < 90.0:
                note_globale_ko = True
                raisons_ko.append(f"Note globale insuffisante: {pourcentage_conformite:.1f}% < 90%")

            # Vérifier les autres fautes majeures (KO immédiat)
            autres_fautes = [f for f in fautes_majeures if f['type'] not in ['MANQUANT_QGIS', 'MANQUANT_SUIVI', 'STRUCTURE_QGIS_INVALIDE', 'STRUCTURE_SUIVI_INVALIDE']]
            if autres_fautes:
                for faute in autres_fautes:
                    raisons_ko.append(f"Faute majeure: {faute['type']} - {faute['description']}")

            # Déterminer le statut final
            if fichiers_manquants:
                statut = "KO"  # KO immédiat si fichiers manquants
            elif note_globale_ko or autres_fautes:
                statut = "KO"  # KO si note insuffisante ou autres fautes
            else:
                statut = "OK"  # OK si tout est conforme

            self.logger.info(f"Évaluation commune - Statut: {statut}, Conformité: {pourcentage_conformite:.1f}%, "
                           f"Fautes majeures: {len(fautes_majeures)}")

            return {
                'statut': statut,
                'pourcentage_conformite': pourcentage_conformite,
                'raisons_ko': raisons_ko,
                'fautes_majeures': fautes_majeures,
                'note_globale_ko': note_globale_ko,
                'fichiers_manquants': fichiers_manquants,
                'fautes_fichiers': fautes_fichiers,
                'seuil_conformite': 90.0
            }

        except Exception as e:
            self.logger.error(f"Erreur évaluation statut commune: {e}")
            return {
                'statut': 'ERROR',
                'pourcentage_conformite': 0.0,
                'raisons_ko': [f'Erreur de calcul: {str(e)}'],
                'fautes_majeures': [],
                'note_globale_ko': False
            }

    def _calculate_conformite_percentage(self) -> float:
        """
        Calcule le pourcentage de conformité global basé sur les 5 critères avec pondérations.

        PONDÉRATIONS :
        - CMS: 30% (0.3)
        - PA: 60% (0.6)
        - Banbou: 5% (0.05)
        - Écart: 5% (0.05)

        Returns:
            Pourcentage de conformité (0-100%)
        """
        try:
            # Récupérer les résultats des critères
            summary = self.qc_results.get('summary', {})

            # Calculer les taux d'erreur par catégorie
            taux_erreur_cms = self._calculate_taux_erreur_cms()
            taux_erreur_pa = self._calculate_taux_erreur_pa()
            taux_erreur_banbou = self._calculate_erreur_banbou_percentage() / 100.0  # Convertir en décimal
            taux_erreur_ecart = self._calculate_ecart_plan_adressage_percentage() / 100.0  # Convertir en décimal

            # Appliquer les pondérations aux taux d'erreur
            score_erreur_pondere = (
                taux_erreur_cms * 0.3 +      # CMS: 30%
                taux_erreur_pa * 0.6 +       # PA: 60%
                taux_erreur_banbou * 0.05 +  # Banbou: 5%
                taux_erreur_ecart * 0.05     # Écart: 5%
            )

            # Convertir le score d'erreur en pourcentage de conformité
            # Conformité = 100% - (Score d'erreur pondéré * 100)
            pourcentage_conformite = max(0.0, 100.0 - (score_erreur_pondere * 100.0))

            self.logger.info(f"Calcul conformité - CMS: {taux_erreur_cms:.3f}, PA: {taux_erreur_pa:.3f}, "
                           f"Banbou: {taux_erreur_banbou:.3f}, Écart: {taux_erreur_ecart:.3f}, "
                           f"Score pondéré: {score_erreur_pondere:.3f}, Conformité: {pourcentage_conformite:.1f}%")

            return pourcentage_conformite

        except Exception as e:
            self.logger.error(f"Erreur calcul conformité: {e}")
            return 0.0

    def _calculate_taux_erreur_cms(self) -> float:
        """Calcule le taux d'erreur CMS (à implémenter selon vos critères CMS)."""
        try:
            # Pour l'instant, retourner 0.0 - à adapter selon vos critères CMS spécifiques
            # Vous pouvez ici implémenter la logique de calcul des erreurs CMS
            return 0.0
        except Exception:
            return 0.0

    def _calculate_taux_erreur_pa(self) -> float:
        """Calcule le taux d'erreur PA basé sur les critères 3, 4 et 5."""
        try:
            if not self.qc_results:
                return 0.0

            # Compter le total d'erreurs PA (critères 3, 4, 5)
            critere_3 = self.qc_results.get('critere_3', {})
            critere_4 = self.qc_results.get('critere_4', {})
            critere_5 = self.qc_results.get('critere_5', {})

            total_erreurs_pa = (
                critere_3.get('total_erreurs_detectees', 0) +
                critere_4.get('total_ad_a_analyser', 0) +
                critere_5.get('total_motifs_incorrects', 0)
            )

            # Base de calcul : nombre total d'enregistrements PA
            total_records_pa = 0
            if self.qgis_data is not None:
                total_records_pa = len(self.qgis_data)

            if total_records_pa > 0:
                taux_erreur = total_erreurs_pa / total_records_pa
                return min(1.0, taux_erreur)  # Limiter à 100%

            return 0.0

        except Exception as e:
            self.logger.error(f"Erreur calcul taux erreur PA: {e}")
            return 0.0

    def _calculate_statut_commune_nouvelle_logique(self) -> str:
        """
        Calcule le statut commune selon la nouvelle logique :
        1. Si faute majeure dans page 3 colonne G → KO
        2. Sinon si score total ≥ 10% → KO
        3. Sinon → OK
        """
        try:
            # Critère 1 : Vérifier les fautes majeures dans la page 3 colonne G
            if hasattr(self, 'qc_results') and self.qc_results:
                # Récupérer les données de la page 3 qui seront générées
                page3_data = self._get_page3_data_for_statut_check()

                # Vérifier s'il y a des "Faute Majeure" dans les données
                for row in page3_data:
                    if len(row) > 6:  # Colonne G (index 6)
                        etat_value = str(row[6]).strip()
                        if etat_value == "Faute Majeure":
                            self.logger.info("Statut commune: KO (Faute Majeure détectée dans page 3)")
                            return "KO"

            # Critère 2 : Vérifier le score total (sera calculé dans J11)
            # Calculer le score total selon la logique existante
            score_total = self._calculate_score_total_percentage()

            if score_total >= 10.0:
                self.logger.info(f"Statut commune: KO (Score total {score_total:.1f}% ≥ 10%)")
                return "KO"
            else:
                self.logger.info(f"Statut commune: OK (Score total {score_total:.1f}% < 10%)")
                return "OK"

        except Exception as e:
            self.logger.error(f"Erreur calcul statut commune nouvelle logique: {e}")
            return "ERROR"

    def _get_page3_data_for_statut_check(self) -> List[List[str]]:
        """Récupère les données de la page 3 pour vérification du statut."""
        try:
            # Cette fonction simule la génération des données page 3 pour vérifier les fautes majeures
            # En pratique, les fautes majeures seront saisies manuellement dans Excel
            # Pour l'instant, on retourne une liste vide car les données seront saisies après génération
            return []
        except Exception as e:
            self.logger.error(f"Erreur récupération données page 3: {e}")
            return []

    def _calculate_score_total_percentage(self) -> float:
        """Calcule le score total en pourcentage (équivalent à J11)."""
        try:
            # Calculer les taux d'erreur par catégorie
            taux_erreur_cms = self._calculate_taux_erreur_cms()
            taux_erreur_pa = self._calculate_taux_erreur_pa()
            taux_erreur_banbou = self._calculate_erreur_banbou_percentage() / 100.0
            taux_erreur_ecart = self._calculate_ecart_plan_adressage_percentage() / 100.0

            # Appliquer les pondérations aux taux d'erreur
            score_total = (
                taux_erreur_cms * 0.3 +      # CMS: 30%
                taux_erreur_pa * 0.6 +       # PA: 60%
                taux_erreur_banbou * 0.05 +  # Banbou: 5%
                taux_erreur_ecart * 0.05     # Écart: 5%
            )

            # Convertir en pourcentage
            score_total_percent = score_total * 100.0

            self.logger.info(f"Score total calculé: {score_total_percent:.2f}%")
            return score_total_percent

        except Exception as e:
            self.logger.error(f"Erreur calcul score total: {e}")
            return 0.0

    def _calculate_resume_erreurs(self) -> Dict[str, str]:
        """Calcule les pourcentages du résumé erreurs avec pondérations."""
        try:
            # Pondérations fixes (constantes)
            ponderation_cms = "0,3"
            ponderation_pa = "0,6"
            ponderation_banbou = "0,05"
            ponderation_ecart = "0,05"

            # Calculer les pourcentages bruts
            pourcentage_cms_brut = self._calculate_taux_erreur_cms() * 100.0  # Convertir en pourcentage
            pourcentage_pa_brut = self._calculate_taux_erreur_pa() * 100.0    # Convertir en pourcentage
            pourcentage_banbou_brut = self._calculate_erreur_banbou_percentage()  # Déjà en pourcentage
            pourcentage_ecart_brut = self._calculate_ecart_plan_adressage_percentage()  # Déjà en pourcentage

            # Convertir en décimal pour les calculs pondérés
            pourcentage_cms_decimal = pourcentage_cms_brut / 100.0
            pourcentage_pa_decimal = pourcentage_pa_brut / 100.0
            pourcentage_banbou_decimal = pourcentage_banbou_brut / 100.0
            pourcentage_ecart_decimal = pourcentage_ecart_brut / 100.0

            # Calculer les taux pondérés (pourcentage décimal * pondération)
            taux_cms = pourcentage_cms_decimal * 0.3
            taux_pa = pourcentage_pa_decimal * 0.6
            taux_banbou = pourcentage_banbou_decimal * 0.05
            taux_ecart = pourcentage_ecart_decimal * 0.05

            # Calculer le total
            total = taux_cms + taux_pa + taux_banbou + taux_ecart

            # Formater selon la capture d'écran : % Brut en décimal, Pondération en décimal, Score en entier
            return {
                'ponderation_cms': ponderation_cms,
                'ponderation_pa': ponderation_pa,
                'ponderation_banbou': ponderation_banbou,
                'ponderation_ecart': ponderation_ecart,
                # Pourcentages bruts (colonne % Brut) - Format décimal selon capture
                'pourcentage_cms_brut': f"{pourcentage_cms_decimal:.1f}".replace('.', ','),
                'pourcentage_pa_brut': f"{pourcentage_pa_decimal:.1f}".replace('.', ','),
                'pourcentage_banbou_brut': f"{pourcentage_banbou_decimal:.2f}".replace('.', ','),
                'pourcentage_ecart_brut': f"{pourcentage_ecart_decimal:.2f}".replace('.', ','),
                # Taux pondérés (colonne Score) - Format entier selon capture
                'taux_cms': f"{int(taux_cms * 100)}",
                'taux_pa': f"{int(taux_pa * 100)}",
                'taux_banbou': f"{int(taux_banbou * 100)}",
                'taux_ecart': f"{int(taux_ecart * 100)}",
                'total': f"{int(total * 100)}",
                # Données supplémentaires pour compatibilité
                'pourcentage_ecart_brut_percent': f"{pourcentage_ecart_brut:.2f}%",
                'pourcentage_banbou_brut_percent': f"{pourcentage_banbou_brut:.0f}%"
            }

        except Exception as e:
            self.logger.error(f"Erreur calcul résumé erreurs: {e}")
            return {
                'ponderation_cms': "0,3",
                'ponderation_pa': "0,6",
                'ponderation_banbou': "0,05",
                'ponderation_ecart': "0,05",
                'pourcentage_cms_brut': "0,0",
                'pourcentage_pa_brut': "0,0",
                'pourcentage_banbou_brut': "0,00",
                'pourcentage_ecart_brut': "0,00",
                'taux_cms': "0",
                'taux_pa': "0",
                'taux_banbou': "0",
                'taux_ecart': "0",
                'total': "0",
                'pourcentage_ecart_brut_percent': "0,00%",
                'pourcentage_banbou_brut_percent': "0%"
            }

    def _calculate_ecart_plan_adressage_percentage(self) -> float:
        """Calcule le pourcentage d'erreur spécifique à l'écart Plan Adressage."""
        try:
            if not self.qc_results or 'critere_0' not in self.qc_results:
                return 0.0

            critere_0 = self.qc_results['critere_0']
            ecart_data = critere_0.get('ecart_plan_adressage', {})

            total_ecarts = 0
            total_motifs = 0

            # Debug: afficher les données pour comprendre le calcul
            self.logger.debug("Calcul pourcentage écart Plan Adressage:")

            for motif, data in ecart_data.items():
                suivi_count = data.get('suivi_count', 0)
                qgis_count = data.get('qgis_count', 0)
                difference = data.get('difference', 0)
                has_ecart = data.get('has_ecart', False)

                # Utiliser le maximum comme base de calcul
                base_count = max(suivi_count, qgis_count)
                total_motifs += base_count

                # Ajouter l'écart absolu si il y en a un
                if has_ecart:
                    ecart_abs = abs(difference)
                    total_ecarts += ecart_abs
                    self.logger.debug(f"  {motif}: Suivi={suivi_count}, QGis={qgis_count}, Écart={difference:+d}, Base={base_count}")
                else:
                    self.logger.debug(f"  {motif}: Suivi={suivi_count}, QGis={qgis_count}, Pas d'écart, Base={base_count}")

            self.logger.debug(f"  Total écarts: {total_ecarts}, Total motifs: {total_motifs}")

            if total_motifs > 0:
                pourcentage = (total_ecarts / total_motifs) * 100
                self.logger.debug(f"  Pourcentage calculé: {pourcentage:.2f}%")
                return pourcentage
            return 0.0

        except Exception as e:
            self.logger.warning(f"Erreur calcul pourcentage écart Plan Adressage: {e}")
            return 0.0

    def _detect_fautes_majeures(self) -> List[Dict[str, Any]]:
        """
        Détecte les fautes majeures dans les résultats des 5 critères.

        FAUTES MAJEURES DÉTECTÉES :
        1. IMB supprimé (à implémenter selon vos critères)
        2. OK fautif (motif "OK" avec adresse optimum = adresse BAN)
        3. Autres fautes critiques

        Returns:
            Liste des fautes majeures détectées
        """
        try:
            fautes_majeures = []

            if not self.qc_results:
                return fautes_majeures

            # 1. DÉTECTION DES "OK FAUTIFS" (Critère 3)
            fautes_ok = self._detect_ok_fautifs()
            fautes_majeures.extend(fautes_ok)

            # 2. DÉTECTION DES FICHIERS MANQUANTS (FAUTES MAJEURES CRITIQUES)
            fautes_fichiers_manquants = self._detect_fichiers_manquants()
            fautes_majeures.extend(fautes_fichiers_manquants)

            # 3. DÉTECTION DES IMB SUPPRIMÉS (à implémenter selon vos critères)
            fautes_imb_supprimes = self._detect_imb_supprimes()
            fautes_majeures.extend(fautes_imb_supprimes)

            # 4. DÉTECTION D'AUTRES FAUTES MAJEURES
            autres_fautes = self._detect_autres_fautes_majeures()
            fautes_majeures.extend(autres_fautes)

            self.logger.info(f"Fautes majeures détectées: {len(fautes_majeures)}")
            for faute in fautes_majeures:
                self.logger.warning(f"FAUTE MAJEURE - {faute['type']}: {faute['description']}")

            return fautes_majeures

        except Exception as e:
            self.logger.error(f"Erreur détection fautes majeures: {e}")
            return []

    def _detect_ok_fautifs(self) -> List[Dict[str, Any]]:
        """Détecte les motifs 'OK' fautifs où adresse optimum = adresse BAN."""
        try:
            fautes_ok = []

            # Vérifier les résultats du critère 3
            critere_3 = self.qc_results.get('critere_3', {})
            erreurs_motif_ok = critere_3.get('erreurs_motif_ok', [])

            for erreur in erreurs_motif_ok:
                faute = {
                    'type': 'OK_FAUTIF',
                    'description': f"IMB {erreur.get('imb_code', 'N/A')} - Motif 'OK' incorrect (adresses identiques)",
                    'imb_code': erreur.get('imb_code', ''),
                    'adresse_optimum': erreur.get('adresse_optimum', ''),
                    'adresse_ban': erreur.get('adresse_ban', ''),
                    'critere_source': 3,
                    'gravite': 'MAJEURE'
                }
                fautes_ok.append(faute)

            return fautes_ok

        except Exception as e:
            self.logger.error(f"Erreur détection OK fautifs: {e}")
            return []

    def _detect_fichiers_manquants(self) -> List[Dict[str, Any]]:
        """
        Détecte les fautes majeures liées aux fichiers manquants.

        FAUTES MAJEURES CRITIQUES :
        1. MANQUANT QGIS - Fichier Résultats QGis non chargé
        2. MANQUANT SUIVI - Fichier Suivi Commune non chargé

        Returns:
            Liste des fautes majeures de fichiers manquants
        """
        try:
            fautes_fichiers = []

            # 1. VÉRIFICATION FICHIER QGIS
            if not hasattr(self, 'qgis_data') or self.qgis_data is None or len(self.qgis_data) == 0:
                faute_qgis = {
                    'type': 'MANQUANT_QGIS',
                    'description': 'Fichier Résultats QGis manquant ou vide - Analyse impossible',
                    'fichier_requis': 'Résultats QGis (Excel)',
                    'impact': 'Critères 3, 4, 5 non analysables',
                    'critere_source': 'FICHIER',
                    'gravite': 'CRITIQUE'
                }
                fautes_fichiers.append(faute_qgis)
                self.logger.error("FAUTE MAJEURE CRITIQUE: Fichier QGis manquant")

            # 2. VÉRIFICATION FICHIER SUIVI COMMUNE
            if not hasattr(self, 'suivi_data') or self.suivi_data is None or len(self.suivi_data) == 0:
                faute_suivi = {
                    'type': 'MANQUANT_SUIVI',
                    'description': 'Fichier Suivi Commune manquant ou vide - Analyse impossible',
                    'fichier_requis': 'Suivi Commune (Excel)',
                    'impact': 'Critères 0, 2 non analysables',
                    'critere_source': 'FICHIER',
                    'gravite': 'CRITIQUE'
                }
                fautes_fichiers.append(faute_suivi)
                self.logger.error("FAUTE MAJEURE CRITIQUE: Fichier Suivi Commune manquant")

            # 3. VÉRIFICATION COHÉRENCE DES FICHIERS
            if hasattr(self, 'qgis_data') and self.qgis_data is not None and len(self.qgis_data) > 0:
                if hasattr(self, 'suivi_data') and self.suivi_data is not None and len(self.suivi_data) > 0:
                    # Les deux fichiers sont présents - vérifier la cohérence basique
                    try:
                        # Vérifier que les fichiers ont des données exploitables
                        qgis_columns = list(self.qgis_data.columns) if hasattr(self.qgis_data, 'columns') else []
                        suivi_sheets = list(self.suivi_data.keys()) if isinstance(self.suivi_data, dict) else []

                        # Vérifier les colonnes critiques du fichier QGis
                        colonnes_critiques_qgis = ['A', 'J', 'U']  # IMB, Motif, Adresse BAN
                        colonnes_manquantes_qgis = [col for col in colonnes_critiques_qgis if col not in qgis_columns]

                        if colonnes_manquantes_qgis:
                            faute_structure_qgis = {
                                'type': 'STRUCTURE_QGIS_INVALIDE',
                                'description': f'Fichier QGis - Colonnes critiques manquantes: {", ".join(colonnes_manquantes_qgis)}',
                                'colonnes_manquantes': colonnes_manquantes_qgis,
                                'colonnes_requises': colonnes_critiques_qgis,
                                'critere_source': 'FICHIER',
                                'gravite': 'MAJEURE'
                            }
                            fautes_fichiers.append(faute_structure_qgis)

                        # Vérifier la structure du fichier Suivi (doit avoir au moins 2 pages)
                        if len(suivi_sheets) < 2:
                            faute_structure_suivi = {
                                'type': 'STRUCTURE_SUIVI_INVALIDE',
                                'description': f'Fichier Suivi Commune - Structure invalide: {len(suivi_sheets)} page(s) trouvée(s), minimum 2 requises',
                                'pages_trouvees': len(suivi_sheets),
                                'pages_requises': 2,
                                'critere_source': 'FICHIER',
                                'gravite': 'MAJEURE'
                            }
                            fautes_fichiers.append(faute_structure_suivi)

                    except Exception as e:
                        self.logger.warning(f"Erreur vérification cohérence fichiers: {e}")

            if fautes_fichiers:
                self.logger.warning(f"Fautes majeures fichiers détectées: {len(fautes_fichiers)}")
                for faute in fautes_fichiers:
                    self.logger.error(f"FICHIER MANQUANT/INVALIDE - {faute['type']}: {faute['description']}")

            return fautes_fichiers

        except Exception as e:
            self.logger.error(f"Erreur détection fichiers manquants: {e}")
            return []

    def _detect_imb_supprimes(self) -> List[Dict[str, Any]]:
        """
        Détecte les IMB supprimés de manière incorrecte.

        NOTE: Cette méthode est un placeholder - à implémenter selon vos critères spécifiques
        pour détecter les suppressions incorrectes d'IMB.
        """
        try:
            fautes_imb = []

            # TODO: Implémenter la logique de détection des IMB supprimés
            # Exemple de logique possible :
            # - Comparer avec un fichier de référence
            # - Détecter des motifs suspects de suppression
            # - Vérifier la cohérence des suppressions

            # Placeholder pour l'instant
            self.logger.debug("Détection IMB supprimés - À implémenter selon critères spécifiques")

            return fautes_imb

        except Exception as e:
            self.logger.error(f"Erreur détection IMB supprimés: {e}")
            return []

    def _detect_autres_fautes_majeures(self) -> List[Dict[str, Any]]:
        """Détecte d'autres types de fautes majeures selon les critères de qualité."""
        try:
            autres_fautes = []

            # 1. MOTIFS INCORRECTS CRITIQUES (Critère 5)
            critere_5 = self.qc_results.get('critere_5', {})
            motifs_incorrects = critere_5.get('motifs_incorrects_entries', [])

            # Considérer les motifs incorrects comme fautes majeures si nombreux
            if len(motifs_incorrects) > 5:  # Seuil configurable
                faute = {
                    'type': 'MOTIFS_INCORRECTS_MASSIFS',
                    'description': f"Nombre excessif de motifs incorrects: {len(motifs_incorrects)} détectés",
                    'count': len(motifs_incorrects),
                    'critere_source': 5,
                    'gravite': 'MAJEURE'
                }
                autres_fautes.append(faute)

            # 2. DOUBLONS SUSPECTS EXCESSIFS (Critère 3)
            critere_3 = self.qc_results.get('critere_3', {})
            doublons_suspects = critere_3.get('total_doublons_suspects', 0)

            if doublons_suspects > 10:  # Seuil configurable
                faute = {
                    'type': 'DOUBLONS_SUSPECTS_EXCESSIFS',
                    'description': f"Nombre excessif de doublons suspects: {doublons_suspects} détectés",
                    'count': doublons_suspects,
                    'critere_source': 3,
                    'gravite': 'MAJEURE'
                }
                autres_fautes.append(faute)

            # 3. ÉCARTS PLAN ADRESSAGE CRITIQUES (Critère 0)
            critere_0 = self.qc_results.get('critere_0', {})
            total_incoherences = critere_0.get('total_incoherences', 0)

            if total_incoherences > 20:  # Seuil configurable
                faute = {
                    'type': 'ECARTS_PLAN_ADRESSAGE_CRITIQUES',
                    'description': f"Écarts Plan Adressage excessifs: {total_incoherences} incohérences",
                    'count': total_incoherences,
                    'critere_source': 0,
                    'gravite': 'MAJEURE'
                }
                autres_fautes.append(faute)

            return autres_fautes

        except Exception as e:
            self.logger.error(f"Erreur détection autres fautes majeures: {e}")
            return []

    def _display_commune_status(self, evaluation_commune: Dict[str, Any]):
        """
        Affiche le statut de conformité de la commune dans l'interface utilisateur.

        Args:
            evaluation_commune: Résultat de l'évaluation de conformité
        """
        try:
            statut = evaluation_commune['statut']
            pourcentage_conformite = evaluation_commune['pourcentage_conformite']
            fautes_majeures = evaluation_commune['fautes_majeures']
            raisons_ko = evaluation_commune['raisons_ko']

            # Créer ou mettre à jour le widget de statut de conformité
            if not hasattr(self, 'statut_conformite_frame'):
                self._create_statut_conformite_widget()

            # Mettre à jour le statut
            fichiers_manquants = evaluation_commune.get('fichiers_manquants', False)

            if statut == "OK":
                statut_text = f"✅ CONFORME ({pourcentage_conformite:.1f}%)"
                statut_color = COLORS['SUCCESS']
                bg_color = "#E8F5E8"  # Vert clair
            elif statut == "KO":
                if fichiers_manquants:
                    statut_text = f"🚫 FICHIERS MANQUANTS - ANALYSE IMPOSSIBLE"
                    statut_color = COLORS['ERROR']
                    bg_color = "#FFE0E0"  # Rouge très clair
                elif fautes_majeures:
                    statut_text = f"🚨 NON CONFORME - FAUTES MAJEURES ({pourcentage_conformite:.1f}%)"
                    statut_color = COLORS['ERROR']
                    bg_color = "#FFE8E8"  # Rouge clair
                else:
                    statut_text = f"❌ NON CONFORME - NOTE INSUFFISANTE ({pourcentage_conformite:.1f}%)"
                    statut_color = COLORS['ERROR']
                    bg_color = "#FFE8E8"  # Rouge clair
            else:
                statut_text = f"⏳ EN ATTENTE D'ANALYSE"
                statut_color = COLORS['INFO']
                bg_color = COLORS['LIGHT']

            # Mettre à jour les variables d'affichage
            self.statut_conformite_var.set(statut_text)
            self.statut_conformite_label.config(fg=statut_color, bg=bg_color)

            # Mettre à jour les détails
            if fichiers_manquants:
                details_text = f"FICHIERS REQUIS MANQUANTS | Fautes critiques: {len(fautes_majeures)}"
            else:
                details_text = f"Seuil: 90% | Fautes majeures: {len(fautes_majeures)}"
                if raisons_ko:
                    details_text += f" | Raisons: {len(raisons_ko)}"

            self.statut_details_var.set(details_text)

            # Afficher un tooltip avec les détails si statut KO
            if statut == "KO" and (raisons_ko or fautes_majeures):
                tooltip_text = "DÉTAILS DU STATUT KO:\n\n"

                if raisons_ko:
                    tooltip_text += "RAISONS:\n"
                    for i, raison in enumerate(raisons_ko[:3], 1):
                        tooltip_text += f"{i}. {raison}\n"
                    if len(raisons_ko) > 3:
                        tooltip_text += f"... et {len(raisons_ko) - 3} autres\n"

                if fautes_majeures:
                    tooltip_text += "\nFAUTES MAJEURES:\n"
                    for i, faute in enumerate(fautes_majeures[:2], 1):
                        tooltip_text += f"{i}. {faute['type']}\n"
                    if len(fautes_majeures) > 2:
                        tooltip_text += f"... et {len(fautes_majeures) - 2} autres\n"

                # Stocker le tooltip pour affichage au survol
                self.statut_tooltip_text = tooltip_text

            self.logger.info(f"Statut conformité affiché: {statut} ({pourcentage_conformite:.1f}%)")

        except Exception as e:
            self.logger.error(f"Erreur affichage statut commune: {e}")

    def _create_statut_conformite_widget(self):
        """Crée le widget d'affichage du statut de conformité dans l'interface."""
        try:
            # Créer le frame pour le statut de conformité dans la section des informations détectées
            if hasattr(self, 'info_frame') and self.info_frame:
                # Créer un frame séparé pour le statut de conformité
                self.statut_conformite_frame = tk.Frame(self.info_frame, bg=COLORS['CARD'])
                self.statut_conformite_frame.pack(fill=tk.X, padx=5, pady=(10, 5))

                # Titre de la section
                title_label = tk.Label(
                    self.statut_conformite_frame,
                    text="🏆 STATUT DE CONFORMITÉ",
                    font=("Segoe UI", 10, "bold"),
                    fg=COLORS['PRIMARY'],
                    bg=COLORS['CARD']
                )
                title_label.pack(anchor=tk.W, pady=(0, 5))

                # Variables pour l'affichage
                self.statut_conformite_var = tk.StringVar(value="⏳ EN ATTENTE D'ANALYSE")
                self.statut_details_var = tk.StringVar(value="Seuil: 90% | Fautes majeures: 0")

                # Label principal du statut
                self.statut_conformite_label = tk.Label(
                    self.statut_conformite_frame,
                    textvariable=self.statut_conformite_var,
                    font=("Segoe UI", 9, "bold"),
                    fg=COLORS['INFO'],
                    bg=COLORS['LIGHT'],
                    relief=tk.RAISED,
                    bd=1,
                    padx=10,
                    pady=5
                )
                self.statut_conformite_label.pack(fill=tk.X, pady=(0, 3))

                # Label des détails
                details_label = tk.Label(
                    self.statut_conformite_frame,
                    textvariable=self.statut_details_var,
                    font=("Segoe UI", 8),
                    fg=COLORS['TEXT_PRIMARY'],
                    bg=COLORS['CARD']
                )
                details_label.pack(anchor=tk.W)

                # Bind pour afficher les détails au survol (si KO)
                self.statut_conformite_label.bind("<Enter>", self._on_statut_hover_enter)
                self.statut_conformite_label.bind("<Leave>", self._on_statut_hover_leave)

                self.logger.info("Widget statut conformité créé")

        except Exception as e:
            self.logger.error(f"Erreur création widget statut conformité: {e}")

    def _on_statut_hover_enter(self, event):
        """Affiche les détails du statut au survol de la souris."""
        try:
            if hasattr(self, 'statut_tooltip_text') and self.statut_tooltip_text:
                # Créer une fenêtre tooltip simple
                self.tooltip_window = tk.Toplevel()
                self.tooltip_window.wm_overrideredirect(True)
                self.tooltip_window.wm_geometry(f"+{event.x_root+10}+{event.y_root+10}")

                tooltip_label = tk.Label(
                    self.tooltip_window,
                    text=self.statut_tooltip_text,
                    background="#FFFFCC",
                    relief=tk.SOLID,
                    borderwidth=1,
                    font=("Segoe UI", 8),
                    justify=tk.LEFT
                )
                tooltip_label.pack()
        except Exception:
            pass

    def _on_statut_hover_leave(self, event):
        """Cache les détails du statut quand la souris quitte la zone."""
        try:
            if hasattr(self, 'tooltip_window'):
                self.tooltip_window.destroy()
                delattr(self, 'tooltip_window')
        except Exception:
            pass

    def _display_modern_results(self, results: Dict[str, Any]):
        """Affiche les résultats de l'analyse avec un design moderne."""
        try:
            # Effacer les résultats précédents
            for widget in self.results_frame.winfo_children():
                widget.destroy()

            summary = results.get('summary', {})
            critere_0 = results.get('critere_0', {})
            critere_1 = results.get('critere_1', {})

            total_errors = summary.get('total_errors', 0)

            # Carte de résumé principal
            self._create_summary_card(summary, total_errors)

            # Cartes de détails par critère
            critere_2 = results.get('critere_2', {})
            critere_3 = results.get('critere_3', {})
            critere_4 = results.get('critere_4', {})
            self._create_criteria_results_cards(critere_0, critere_2, critere_3, critere_4)

            # Informations de contexte
            self._create_context_info_card(summary)

        except Exception as e:
            self.logger.error(f"Erreur affichage résultats modernes: {e}")

    def _display_compact_results(self, results: Dict[str, Any]):
        """Affiche les résultats de l'analyse de manière ultra-compacte."""
        try:
            # Effacer les résultats précédents
            for widget in self.results_frame.winfo_children():
                widget.destroy()

            summary = results.get('summary', {})
            critere_0 = results.get('critere_0', {})
            critere_2 = results.get('critere_2', {})
            critere_3 = results.get('critere_3', {})

            total_errors = summary.get('total_errors', 0)

            # Statut principal ultra-compact
            status_frame = tk.Frame(self.results_frame, bg=COLORS['CARD'])
            status_frame.pack(fill=tk.X, pady=(0, 3))

            status_icon = "✅" if total_errors == 0 else "⚠️" if total_errors < 5 else "❌"
            status_color = COLORS['SUCCESS'] if total_errors == 0 else COLORS['WARNING'] if total_errors < 5 else COLORS['ERROR']

            tk.Label(status_frame, text=f"{status_icon} {total_errors} erreur(s)",
                    font=("Segoe UI", 9, "bold"), fg=status_color,
                    bg=COLORS['CARD']).pack()

            # Détails par critère ultra-compacts
            details_frame = tk.Frame(self.results_frame, bg=COLORS['LIGHT'], relief='flat', bd=1)
            details_frame.pack(fill=tk.X, pady=2)
            details_frame.config(highlightbackground=COLORS['BORDER'], highlightthickness=1)

            details_content = tk.Frame(details_frame, bg=COLORS['LIGHT'])
            details_content.pack(fill=tk.X, padx=3, pady=2)

            # Critère 0
            c0_errors = critere_0.get('total_incoherences', 0)
            c0_color = COLORS['SUCCESS'] if c0_errors == 0 else COLORS['ERROR']
            c0_frame = tk.Frame(details_content, bg=COLORS['LIGHT'])
            c0_frame.pack(fill=tk.X, pady=1)

            tk.Label(c0_frame, text="0", font=("Segoe UI", 7, "bold"),
                    fg='white', bg=c0_color, padx=2).pack(side=tk.LEFT)
            tk.Label(c0_frame, text=f"Incohérences: {c0_errors}",
                    font=("Segoe UI", 7), fg=COLORS['TEXT_PRIMARY'],
                    bg=COLORS['LIGHT']).pack(side=tk.LEFT, padx=(3, 0))

            # Critère 2
            c2_errors = critere_2.get('total_errors', 0)
            c2_color = COLORS['SUCCESS'] if c2_errors == 0 else COLORS['ERROR']
            c2_frame = tk.Frame(details_content, bg=COLORS['LIGHT'])
            c2_frame.pack(fill=tk.X, pady=1)

            tk.Label(c2_frame, text="2", font=("Segoe UI", 7, "bold"),
                    fg='white', bg=c2_color, padx=2).pack(side=tk.LEFT)
            tk.Label(c2_frame, text=f"Tickets: {c2_errors}",
                    font=("Segoe UI", 7), fg=COLORS['TEXT_PRIMARY'],
                    bg=COLORS['LIGHT']).pack(side=tk.LEFT, padx=(3, 0))

            # Critère 3
            c3_errors = critere_3.get('total_erreurs_detectees', 0)
            c3_doublons = critere_3.get('total_doublons_suspects', 0)
            c3_motif_ok = critere_3.get('total_erreurs_motif_ok', 0)
            c3_color = COLORS['SUCCESS'] if c3_errors == 0 else COLORS['ERROR']
            c3_frame = tk.Frame(details_content, bg=COLORS['LIGHT'])
            c3_frame.pack(fill=tk.X, pady=1)

            tk.Label(c3_frame, text="3", font=("Segoe UI", 7, "bold"),
                    fg='white', bg=c3_color, padx=2).pack(side=tk.LEFT)
            tk.Label(c3_frame, text=f"IMB: {c3_doublons}+{c3_motif_ok}={c3_errors}",
                    font=("Segoe UI", 7), fg=COLORS['TEXT_PRIMARY'],
                    bg=COLORS['LIGHT']).pack(side=tk.LEFT, padx=(3, 0))

            # Critère 4
            critere_4 = results.get('critere_4', {})
            c4_errors = critere_4.get('total_ad_a_analyser', 0)
            c4_color = COLORS['SUCCESS'] if c4_errors == 0 else COLORS['ERROR']
            c4_frame = tk.Frame(details_content, bg=COLORS['LIGHT'])
            c4_frame.pack(fill=tk.X, pady=1)

            tk.Label(c4_frame, text="4", font=("Segoe UI", 7, "bold"),
                    fg='white', bg=c4_color, padx=2).pack(side=tk.LEFT)
            tk.Label(c4_frame, text=f"AD à analyser: {c4_errors}",
                    font=("Segoe UI", 7), fg=COLORS['TEXT_PRIMARY'],
                    bg=COLORS['LIGHT']).pack(side=tk.LEFT, padx=(3, 0))

            # Informations contextuelles ultra-compactes
            context_frame = tk.Frame(self.results_frame, bg=COLORS['CARD'])
            context_frame.pack(fill=tk.BOTH, expand=True, pady=(3, 0))

            context_text = f"📅 {summary.get('analysis_date', 'N/A')[:16]}\n"
            context_text += f"📊 Taux erreur: {self._calculate_error_percentage():.1f}%\n"
            context_text += f"📄 Rapport: 3 feuilles vides prêtes"

            tk.Label(context_frame, text=context_text, font=("Segoe UI", 7),
                    fg=COLORS['TEXT_SECONDARY'], bg=COLORS['CARD'],
                    justify=tk.LEFT).pack(expand=True)

        except Exception as e:
            self.logger.error(f"Erreur affichage résultats compacts: {e}")
            # Fallback simple
            tk.Label(self.results_frame, text="❌ Erreur affichage",
                    font=("Segoe UI", 8), fg=COLORS['ERROR'],
                    bg=COLORS['CARD']).pack(expand=True)

    def _create_summary_card(self, summary: Dict[str, Any], total_errors: int):
        """Crée la carte de résumé principal."""
        from ui.styles import create_card_frame

        summary_card = create_card_frame(self.results_frame, shadow=True)
        summary_card.pack(fill=tk.X, pady=(0, 10))

        card_content = summary_card.winfo_children()[0] if summary_card.winfo_children() else summary_card
        content_frame = tk.Frame(card_content, bg=COLORS['CARD'])
        content_frame.pack(fill=tk.X, padx=20, pady=15)

        # En-tête avec statut
        header_frame = tk.Frame(content_frame, bg=COLORS['CARD'])
        header_frame.pack(fill=tk.X, pady=(0, 10))

        status_icon = "✅" if total_errors == 0 else "⚠️" if total_errors < 5 else "❌"
        status_color = COLORS['SUCCESS'] if total_errors == 0 else COLORS['WARNING'] if total_errors < 5 else COLORS['ERROR']

        tk.Label(
            header_frame,
            text=status_icon,
            font=("Segoe UI", 20),
            fg=status_color,
            bg=COLORS['CARD']
        ).pack(side=tk.LEFT)

        tk.Label(
            header_frame,
            text=f"Analyse Terminée - {total_errors} erreur(s) détectée(s)",
            font=UIConfig.FONT_CARD_TITLE,
            fg=status_color,
            bg=COLORS['CARD']
        ).pack(side=tk.LEFT, padx=(10, 0))

        # Métriques
        metrics_frame = tk.Frame(content_frame, bg=COLORS['CARD'])
        metrics_frame.pack(fill=tk.X)
        metrics_frame.grid_columnconfigure(0, weight=1)
        metrics_frame.grid_columnconfigure(1, weight=1)

        # Pourcentage d'erreur
        error_percentage = self._calculate_error_percentage()
        self._create_metric_display(metrics_frame, 0, 0, "📊", "Taux d'erreur", f"{error_percentage:.1f}%")

        # Date d'analyse
        analysis_date = summary.get('analysis_date', 'N/A')
        self._create_metric_display(metrics_frame, 0, 1, "📅", "Date d'analyse", analysis_date)

    def _create_metric_display(self, parent: tk.Widget, row: int, col: int,
                              icon: str, label: str, value: str):
        """Crée un affichage de métrique."""
        metric_frame = tk.Frame(parent, bg=COLORS['LIGHT'], relief='flat', bd=1)
        metric_frame.grid(row=row, column=col, sticky="ew", padx=5, pady=5)
        metric_frame.config(highlightbackground=COLORS['BORDER'], highlightthickness=1)

        content = tk.Frame(metric_frame, bg=COLORS['LIGHT'])
        content.pack(fill=tk.X, padx=10, pady=8)

        # Icône et label
        header = tk.Frame(content, bg=COLORS['LIGHT'])
        header.pack(fill=tk.X)

        tk.Label(header, text=icon, font=("Segoe UI", 12),
                fg=COLORS['SECONDARY'], bg=COLORS['LIGHT']).pack(side=tk.LEFT)

        tk.Label(header, text=label, font=UIConfig.FONT_SMALL,
                fg=COLORS['TEXT_SECONDARY'], bg=COLORS['LIGHT']).pack(side=tk.LEFT, padx=(5, 0))

        # Valeur
        tk.Label(content, text=value, font=UIConfig.FONT_SUBTITLE,
                fg=COLORS['TEXT_PRIMARY'], bg=COLORS['LIGHT']).pack(anchor=tk.W, pady=(3, 0))

    def _create_criteria_results_cards(self, critere_0: Dict[str, Any], critere_2: Dict[str, Any], critere_3: Dict[str, Any], critere_4: Dict[str, Any]):
        """Crée les cartes de résultats par critère."""
        from ui.styles import create_card_frame

        # Conteneur pour les cartes de critères
        criteria_container = tk.Frame(self.results_frame, bg=COLORS['CARD'])
        criteria_container.pack(fill=tk.X, pady=(0, 10))
        criteria_container.grid_columnconfigure(0, weight=1)
        criteria_container.grid_columnconfigure(1, weight=1)
        criteria_container.grid_columnconfigure(2, weight=1)
        criteria_container.grid_columnconfigure(3, weight=1)

        # Carte Critère 0 - Écart Plan Adressage
        self._create_single_criteria_card(criteria_container, 0, 0, "0", "Écart Plan Adressage",
                                         critere_0.get('total_incoherences', 0),
                                         critere_0.get('incoherences', []))

        # Carte Critère 2 - Oubli Tickets
        self._create_single_criteria_card(criteria_container, 0, 1, "2", "Oubli Tickets",
                                         critere_2.get('total_errors', 0),
                                         critere_2.get('errors', []))

        # Carte Critère 3 - Contrôle IMB
        total_c3_errors = critere_3.get('total_erreurs_detectees', 0)
        all_c3_details = critere_3.get('doublons_details', [])
        self._create_single_criteria_card(criteria_container, 0, 2, "3", "Contrôle IMB",
                                         total_c3_errors, all_c3_details)

        # Carte Critère 4 - AD à Analyser
        total_c4_errors = critere_4.get('total_ad_a_analyser', 0)
        all_c4_details = critere_4.get('ad_a_analyser_entries', [])
        self._create_single_criteria_card(criteria_container, 0, 3, "4", "AD à Analyser",
                                         total_c4_errors, all_c4_details)

    def _create_single_criteria_card(self, parent: tk.Widget, row: int, col: int,
                                    number: str, title: str, count: int, details: list):
        """Crée une carte pour un critère spécifique."""
        from ui.styles import create_card_frame

        card = create_card_frame(parent, shadow=True)
        card.grid(row=row, column=col, sticky="nsew", padx=5, pady=5)

        card_content = card.winfo_children()[0] if card.winfo_children() else card
        content_frame = tk.Frame(card_content, bg=COLORS['CARD'])
        content_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=12)

        # En-tête
        header_frame = tk.Frame(content_frame, bg=COLORS['CARD'])
        header_frame.pack(fill=tk.X, pady=(0, 8))

        # Badge
        badge_color = COLORS['SUCCESS'] if count == 0 else COLORS['WARNING'] if count < 3 else COLORS['ERROR']
        badge = tk.Label(
            header_frame,
            text=f"CRITÈRE {number}",
            font=("Segoe UI", 8, "bold"),
            fg='white',
            bg=badge_color,
            padx=6,
            pady=2
        )
        badge.pack(side=tk.LEFT)

        # Titre
        tk.Label(
            header_frame,
            text=title,
            font=UIConfig.FONT_SUBTITLE,
            fg=COLORS['TEXT_PRIMARY'],
            bg=COLORS['CARD']
        ).pack(side=tk.LEFT, padx=(8, 0))

        # Compteur
        count_color = COLORS['SUCCESS'] if count == 0 else COLORS['ERROR']
        tk.Label(
            content_frame,
            text=f"{count} erreur(s) détectée(s)",
            font=UIConfig.FONT_CARD_SUBTITLE,
            fg=count_color,
            bg=COLORS['CARD']
        ).pack(anchor=tk.W, pady=(0, 5))

        # Détails spécifiques pour le Critère 0 (Écart Plan Adressage)
        if number == "0" and details and count > 0:
            # Afficher les écarts détectés
            ecarts_text = "Écarts détectés:\n"
            for detail in details[:3]:  # Limiter à 3 exemples
                motif = detail.get('motif', 'N/A')
                suivi = detail.get('suivi_count', 0)
                qgis = detail.get('qgis_count', 0)
                diff = detail.get('difference', 0)
                ecarts_text += f"• {motif}: {suivi} vs {qgis} ({diff:+d})\n"

            if len(details) > 3:
                ecarts_text += f"... et {len(details) - 3} autre(s) écart(s)"

            details_label = tk.Label(
                content_frame,
                text=ecarts_text.strip(),
                font=UIConfig.FONT_SMALL,
                fg=COLORS['TEXT_SECONDARY'],
                bg=COLORS['CARD'],
                justify=tk.LEFT
            )
            details_label.pack(anchor=tk.W, pady=(5, 0))

        # Détails spécifiques pour le Critère 2 (Oubli Tickets)
        elif number == "2":
            # Récupérer les statuts des tickets depuis les résultats
            if hasattr(self, 'qc_results') and self.qc_results and 'critere_2' in self.qc_results:
                critere_2 = self.qc_results['critere_2']
                ticket_upr_status = critere_2.get('ticket_upr_status', 'N/A')
                ticket_501_511_status = critere_2.get('ticket_501_511_status', 'N/A')

                # Afficher les statuts des tickets
                tickets_text = f"Statuts des tickets:\n"
                tickets_text += f"• Ticket UPR: {ticket_upr_status}\n"
                tickets_text += f"• Ticket 501/511: {ticket_501_511_status}"

                # Couleur selon les statuts
                status_color = COLORS['SUCCESS']
                if ticket_upr_status == 'NOK' or ticket_501_511_status == 'NOK':
                    status_color = COLORS['ERROR']
                elif ticket_upr_status == 'N/A' and ticket_501_511_status == 'N/A':
                    status_color = COLORS['TEXT_SECONDARY']

                tickets_label = tk.Label(
                    content_frame,
                    text=tickets_text,
                    font=UIConfig.FONT_SMALL,
                    fg=status_color,
                    bg=COLORS['CARD'],
                    justify=tk.LEFT
                )
                tickets_label.pack(anchor=tk.W, pady=(5, 0))

                # Afficher les erreurs détectées si présentes
                if details and count > 0:
                    errors_text = "\nErreurs détectées:\n"
                    for detail in details[:2]:  # Limiter à 2 exemples
                        error_type = detail.get('type', 'N/A')
                        description = detail.get('description', 'N/A')
                        if error_type == 'TICKET_UPR_MANQUANT':
                            errors_text += "• Ticket UPR manquant\n"
                        elif error_type == 'TICKET_501_511_MANQUANT':
                            errors_text += "• Ticket 501/511 manquant\n"

                    if len(details) > 2:
                        errors_text += f"... et {len(details) - 2} autre(s) erreur(s)"

                    errors_label = tk.Label(
                        content_frame,
                        text=errors_text.strip(),
                        font=UIConfig.FONT_SMALL,
                        fg=COLORS['ERROR'],
                        bg=COLORS['CARD'],
                        justify=tk.LEFT
                    )
                    errors_label.pack(anchor=tk.W, pady=(2, 0))

        # Détails génériques pour les autres critères
        elif details and count > 0:
            details_text = f"Exemples: {len(details[:3])} élément(s) affichés"
            if len(details) > 3:
                details_text += f" (+ {len(details) - 3} autres)"

            tk.Label(
                content_frame,
                text=details_text,
                font=UIConfig.FONT_SMALL,
                fg=COLORS['TEXT_SECONDARY'],
                bg=COLORS['CARD']
            ).pack(anchor=tk.W)

    def _create_context_info_card(self, summary: Dict[str, Any]):
        """Crée la carte d'informations de contexte."""
        from ui.styles import create_card_frame

        context_card = create_card_frame(self.results_frame, shadow=True)
        context_card.pack(fill=tk.X)

        card_content = context_card.winfo_children()[0] if context_card.winfo_children() else context_card
        content_frame = tk.Frame(card_content, bg=COLORS['CARD'])
        content_frame.pack(fill=tk.X, padx=20, pady=15)

        # Titre
        tk.Label(
            content_frame,
            text="📋 Informations de Contexte",
            font=UIConfig.FONT_CARD_TITLE,
            fg=COLORS['PRIMARY'],
            bg=COLORS['CARD']
        ).pack(anchor=tk.W, pady=(0, 10))

        # Grille d'informations
        info_grid = tk.Frame(content_frame, bg=COLORS['CARD'])
        info_grid.pack(fill=tk.X)
        info_grid.grid_columnconfigure(0, weight=1)
        info_grid.grid_columnconfigure(1, weight=1)

        # Informations
        self._create_context_item(info_grid, 0, 0, "🏘️", "Commune", summary.get('commune', 'N/A'))
        self._create_context_item(info_grid, 0, 1, "👤", "Collaborateur", summary.get('collaborateur', 'N/A'))
        self._create_context_item(info_grid, 1, 0, "🏛️", "Code INSEE", summary.get('insee', 'N/A'))
        self._create_context_item(info_grid, 1, 1, "🆔", "ID Tâche", summary.get('id_tache', 'N/A'))

    def _create_context_item(self, parent: tk.Widget, row: int, col: int,
                            icon: str, label: str, value: str):
        """Crée un élément d'information de contexte."""
        item_frame = tk.Frame(parent, bg=COLORS['CARD'])
        item_frame.grid(row=row, column=col, sticky="w", padx=10, pady=3)

        # Icône et label
        header = tk.Frame(item_frame, bg=COLORS['CARD'])
        header.pack(anchor=tk.W)

        tk.Label(header, text=icon, font=("Segoe UI", 10),
                fg=COLORS['SECONDARY'], bg=COLORS['CARD']).pack(side=tk.LEFT)

        tk.Label(header, text=f"{label}:", font=UIConfig.FONT_SMALL,
                fg=COLORS['TEXT_SECONDARY'], bg=COLORS['CARD']).pack(side=tk.LEFT, padx=(3, 0))

        # Valeur
        tk.Label(item_frame, text=value, font=UIConfig.FONT_SUBTITLE,
                fg=COLORS['TEXT_PRIMARY'], bg=COLORS['CARD']).pack(anchor=tk.W, pady=(2, 0))

    def show(self):
        """Affiche le module."""
        if self.main_frame:
            self.main_frame.grid(row=0, column=0, sticky="nsew")

    def hide(self):
        """Cache le module."""
        if self.main_frame:
            self.main_frame.grid_forget()

    # ==========================================
    # INTERFACE UTILISATEUR MODERNISÉE
    # ==========================================

    def _create_enhanced_header(self):
        """Crée un en-tête modernisé avec design cohérent avec l'accueil."""
        # Header avec style Sofrecom compact
        header_frame = tk.Frame(self.analysis_tab, bg=COLORS['ACCENT'], height=40)
        header_frame.grid(row=0, column=0, sticky="ew", padx=0, pady=0)
        header_frame.pack_propagate(False)
        header_frame.config(highlightbackground=COLORS['PRIMARY'], highlightthickness=1)

        # Conteneur principal avec style Sofrecom compact
        content = tk.Frame(header_frame, bg=COLORS['ACCENT'])
        content.pack(fill=tk.BOTH, expand=True, padx=20, pady=8)

        # Titre principal avec style cohérent avec l'accueil
        title_frame = tk.Frame(content, bg=COLORS['ACCENT'])
        title_frame.pack(side=tk.LEFT)

        # Icône avec style Sofrecom
        icon_label = tk.Label(
            title_frame,
            text="🔍",
            font=("Segoe UI", 16),
            fg=COLORS['PRIMARY'],
            bg=COLORS['ACCENT']
        )
        icon_label.pack(side=tk.LEFT, padx=(0, 8))

        title_label = tk.Label(
            title_frame,
            text="Module 5 - Contrôle Qualité",
            font=("Segoe UI", 12, "bold"),
            fg=COLORS['PRIMARY'],
            bg=COLORS['ACCENT']
        )
        title_label.pack(side=tk.LEFT)

        # Sous-titre descriptif
        subtitle_label = tk.Label(
            content,
            text="Système d'analyse et de validation de la qualité des données",
            font=("Segoe UI", 9),
            fg=COLORS['INFO'],
            bg=COLORS['ACCENT']
        )
        subtitle_label.pack(side=tk.LEFT, padx=(15, 0))

        # Bouton de choix Mode (Autoévaluation / Contrôle Qualité) - Fonctionnalité future
        self._create_mode_selection_button(content)

        # Indicateurs de statut modernisés
        self._create_enhanced_status_indicators(content)

    def _create_mode_selection_button(self, parent: tk.Widget):
        """Crée le bouton de sélection du mode (Autoévaluation / Contrôle Qualité)."""
        try:
            # Séparateur avant le bouton de mode
            separator = tk.Frame(parent, width=2, bg=COLORS['BORDER'])
            separator.pack(side=tk.LEFT, fill=tk.Y, padx=10)

            # Frame pour le bouton de mode
            mode_frame = tk.Frame(parent, bg=COLORS['CARD'])
            mode_frame.pack(side=tk.LEFT)

            # Variable pour stocker le mode sélectionné
            self.selected_mode = tk.StringVar(value="Contrôle Qualité")

            # Label descriptif
            mode_label = tk.Label(
                mode_frame,
                text="Mode:",
                font=("Segoe UI", 9, "bold"),
                fg=COLORS['TEXT_PRIMARY'],
                bg=COLORS['CARD']
            )
            mode_label.pack(side=tk.LEFT, padx=(0, 5))

            # Bouton de sélection avec menu déroulant
            self.mode_button = tk.Menubutton(
                mode_frame,
                text="🔍 Contrôle Qualité",
                font=("Segoe UI", 9, "bold"),
                fg='white',
                bg=COLORS['PRIMARY'],
                activebackground=COLORS['ACCENT'],
                activeforeground='white',
                relief='flat',
                padx=12,
                pady=4,
                cursor='hand2',
                direction='below'
            )
            self.mode_button.pack(side=tk.LEFT)

            # Menu déroulant pour les options
            mode_menu = tk.Menu(self.mode_button, tearoff=0)
            self.mode_button.config(menu=mode_menu)

            # Options du menu
            mode_menu.add_command(
                label="🔍 Contrôle Qualité",
                command=lambda: self._select_mode("Contrôle Qualité", "🔍")
            )
            mode_menu.add_command(
                label="📊 Autoévaluation",
                command=lambda: self._select_mode("Autoévaluation", "📊")
            )

            # Tooltip supprimé pour éviter les popups

            self.logger.info("Bouton de sélection de mode créé avec succès")

        except Exception as e:
            self.logger.warning(f"Erreur création bouton mode: {e}")

    def _select_mode(self, mode: str, icon: str):
        """Sélectionne le mode d'analyse."""
        try:
            self.selected_mode.set(mode)
            self.mode_button.config(text=f"{icon} {mode}")

            # Feedback visuel uniquement sur le bouton
            if mode == "Autoévaluation":
                self.mode_button.config(bg=COLORS['INFO'])
            else:
                self.mode_button.config(bg=COLORS['PRIMARY'])

            self.logger.info(f"Mode sélectionné: {mode}")

        except Exception as e:
            self.logger.warning(f"Erreur sélection mode: {e}")





    def get_selected_mode(self) -> str:
        """Retourne le mode actuellement sélectionné."""
        try:
            if hasattr(self, 'selected_mode') and self.selected_mode:
                return self.selected_mode.get()
            return "Contrôle Qualité"  # Mode par défaut
        except Exception as e:
            self.logger.warning(f"Erreur récupération mode: {e}")
            return "Contrôle Qualité"

    def is_autoevaluation_mode(self) -> bool:
        """Vérifie si le mode Autoévaluation est sélectionné."""
        return self.get_selected_mode() == "Autoévaluation"

    def is_quality_control_mode(self) -> bool:
        """Vérifie si le mode Contrôle Qualité est sélectionné."""
        return self.get_selected_mode() == "Contrôle Qualité"

    def _handle_future_functionality(self, feature_name: str):
        """Gère les fonctionnalités futures non encore implémentées."""
        try:
            if self.is_autoevaluation_mode():
                messagebox.showinfo(
                    "Fonctionnalité Future",
                    f"🚀 {feature_name}\n\n"
                    f"Cette fonctionnalité sera disponible dans une future version.\n\n"
                    f"Mode Autoévaluation en cours de développement:\n"
                    f"• Auto-analyse intelligente des données\n"
                    f"• Suggestions d'amélioration automatiques\n"
                    f"• Rapports d'autoévaluation personnalisés\n\n"
                    f"Restez connecté pour les mises à jour !"
                )
                return False
            return True
        except Exception as e:
            self.logger.warning(f"Erreur gestion fonctionnalité future: {e}")
            return True

    def _create_enhanced_status_indicators(self, parent: tk.Widget):
        """Crée les indicateurs de statut avec design modernisé."""
        # Séparateur élégant
        separator = tk.Frame(parent, width=2, bg=COLORS['BORDER'])
        separator.pack(side=tk.LEFT, fill=tk.Y, padx=15)

        # Conteneur des indicateurs
        indicators_frame = tk.Frame(parent, bg=COLORS['CARD'])
        indicators_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)

        # Indicateur fichiers avec icône améliorée
        self.files_status = tk.Label(
            indicators_frame,
            text="📁 Fichiers: En attente",
            font=("Segoe UI", 9, "bold"),
            fg=COLORS['WARNING'],
            bg=COLORS['CARD']
        )
        self.files_status.pack(side=tk.LEFT, padx=8)

        # Point séparateur
        tk.Label(
            indicators_frame,
            text="•",
            font=("Segoe UI", 12),
            fg=COLORS['BORDER'],
            bg=COLORS['CARD']
        ).pack(side=tk.LEFT, padx=5)

        # Indicateur analyse
        self.analysis_status = tk.Label(
            indicators_frame,
            text="⚙️ Analyse: Prête",
            font=("Segoe UI", 9, "bold"),
            fg=COLORS['INFO'],
            bg=COLORS['CARD']
        )
        self.analysis_status.pack(side=tk.LEFT, padx=8)

        # Point séparateur
        tk.Label(
            indicators_frame,
            text="•",
            font=("Segoe UI", 12),
            fg=COLORS['BORDER'],
            bg=COLORS['CARD']
        ).pack(side=tk.LEFT, padx=5)

        # Indicateur rapport
        self.report_status = tk.Label(
            indicators_frame,
            text="📊 Rapport: En attente",
            font=("Segoe UI", 9, "bold"),
            fg=COLORS['TEXT_SECONDARY'],
            bg=COLORS['CARD']
        )
        self.report_status.pack(side=tk.LEFT, padx=8)

    def _create_enhanced_main_content(self):
        """Crée le contenu principal avec design cohérent avec l'accueil."""
        main_content = tk.Frame(self.analysis_tab, bg=COLORS['BG'])
        main_content.grid(row=1, column=0, sticky="nsew", padx=0, pady=0)

        # Configuration de la grille 2x2 avec espacement compact
        main_content.grid_rowconfigure(0, weight=1)
        main_content.grid_rowconfigure(1, weight=1)
        main_content.grid_columnconfigure(0, weight=1)
        main_content.grid_columnconfigure(1, weight=1)

        # Création des quadrants avec design amélioré
        self._create_enhanced_files_quadrant(main_content, 0, 0)
        self._create_enhanced_info_quadrant(main_content, 0, 1)
        self._create_enhanced_analysis_quadrant(main_content, 1, 0)
        self._create_enhanced_results_quadrant(main_content, 1, 1)

    def _create_enhanced_files_quadrant(self, parent: tk.Widget, row: int, col: int):
        """Quadrant 1: Chargement des fichiers avec style Sofrecom."""
        # Utiliser le style de carte Sofrecom
        card_container = tk.Frame(parent, bg=COLORS['BG'])
        card_container.grid(row=row, column=col, sticky="nsew", padx=10, pady=10)

        # Carte avec style Sofrecom
        frame = tk.Frame(card_container, bg=COLORS['CARD'], relief='flat', bd=0)
        frame.pack(fill=tk.BOTH, expand=True, padx=1, pady=1)
        frame.config(highlightbackground=COLORS['BORDER'], highlightthickness=1)

        # En-tête compact avec style Sofrecom
        title_frame = tk.Frame(frame, bg=COLORS['CARD'])
        title_frame.pack(fill=tk.X, padx=8, pady=(8, 4))

        # Icône et titre
        icon_label = tk.Label(
            title_frame,
            text="📁",
            font=("Segoe UI", 12),
            fg=COLORS['PRIMARY'],
            bg=COLORS['CARD']
        )
        icon_label.pack(side=tk.LEFT, padx=(0, 6))

        title_label = tk.Label(
            title_frame,
            text="Chargement des Fichiers",
            font=("Segoe UI", 10, "bold"),
            fg=COLORS['INFO'],
            bg=COLORS['CARD']
        )
        title_label.pack(side=tk.LEFT)

        # Contenu avec padding compact
        content = tk.Frame(frame, bg=COLORS['CARD'])
        content.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0, 8))

        # Créer les labels d'information s'ils n'existent pas
        if self.qgis_info_label is None:
            self.qgis_info_label = tk.Label(content, text="Aucun fichier chargé", bg=COLORS['CARD'])
        if self.suivi_info_label is None:
            self.suivi_info_label = tk.Label(content, text="Aucun fichier chargé", bg=COLORS['CARD'])

        # Section QGis avec design amélioré
        self._create_enhanced_file_section(content, "QGis", "🗺️", self._load_qgis_file, self.qgis_info_label)

        # Séparateur élégant
        separator = tk.Frame(content, height=1, bg=COLORS['BORDER'])
        separator.pack(fill=tk.X, pady=8)

        # Section Suivi Commune
        self._create_enhanced_file_section(content, "Suivi Commune", "📋", self._load_suivi_file, self.suivi_info_label)

    def _create_enhanced_file_section(self, parent: tk.Widget, title: str, icon: str, command, info_label):
        """Crée une section de fichier avec design amélioré."""
        section_frame = tk.Frame(parent, bg=COLORS['CARD'])
        section_frame.pack(fill=tk.X, pady=3)

        # En-tête de section
        header_frame = tk.Frame(section_frame, bg=COLORS['CARD'])
        header_frame.pack(fill=tk.X)

        tk.Label(
            header_frame,
            text=f"{icon} {title}",
            font=("Segoe UI", 10, "bold"),
            fg=COLORS['TEXT_PRIMARY'],
            bg=COLORS['CARD']
        ).pack(side=tk.LEFT)

        # Bouton avec style Sofrecom compact
        btn = tk.Button(
            header_frame,
            text="📂 Charger",
            font=("Segoe UI", 8, "bold"),
            fg='white',
            bg=COLORS['PRIMARY'],
            activebackground=COLORS['PRIMARY_LIGHT'],
            activeforeground='white',
            relief='flat',
            padx=8,
            pady=2,
            cursor='hand2',
            command=command
        )
        btn.pack(side=tk.RIGHT)

        # Effet hover Sofrecom
        def on_enter(e):
            btn.config(bg=COLORS['PRIMARY_LIGHT'])
        def on_leave(e):
            btn.config(bg=COLORS['PRIMARY'])

        btn.bind("<Enter>", on_enter)
        btn.bind("<Leave>", on_leave)

        # Label d'information avec style amélioré
        info_label.config(
            font=("Segoe UI", 8),
            fg=COLORS['TEXT_SECONDARY'],
            bg=COLORS['CARD'],
            wraplength=200,
            justify=tk.LEFT
        )
        info_label.pack(anchor=tk.W, pady=(5, 0))

    def _create_enhanced_info_quadrant(self, parent: tk.Widget, row: int, col: int):
        """Quadrant 2: Informations détectées avec design modernisé."""
        frame = tk.Frame(parent, bg=COLORS['CARD'], relief='flat', bd=0)
        frame.grid(row=row, column=col, sticky="nsew", padx=2, pady=2)
        frame.config(highlightbackground=COLORS['SUCCESS'], highlightthickness=2)

        # En-tête avec couleur distinctive
        title_frame = tk.Frame(frame, bg=COLORS['SUCCESS'], height=35)
        title_frame.pack(fill=tk.X)
        title_frame.pack_propagate(False)

        title_label = tk.Label(
            title_frame,
            text="ℹ️ Informations Détectées",
            font=("Segoe UI", 11, "bold"),
            fg='white',
            bg=COLORS['SUCCESS']
        )
        title_label.pack(expand=True)

        # Contenu avec grille améliorée
        content = tk.Frame(frame, bg=COLORS['CARD'])
        content.pack(fill=tk.BOTH, expand=True, padx=12, pady=10)

        # Configuration de la grille pour les champs d'information
        content.grid_columnconfigure(0, weight=1)
        content.grid_columnconfigure(1, weight=1)

        # Champs d'information avec design amélioré
        self._create_enhanced_info_field(content, 0, 0, "🏢", "Commune", self.commune_var)
        self._create_enhanced_info_field(content, 0, 1, "👤", "Collaborateur", self.collaborator_var)
        self._create_enhanced_info_field(content, 1, 0, "🆔", "INSEE", self.insee_var)
        self._create_enhanced_info_field(content, 1, 1, "📋", "ID Tâche", self.id_tache_var)

        # Note informative avec style amélioré
        note_frame = tk.Frame(content, bg=COLORS['LIGHT'], relief='flat', bd=1)
        note_frame.grid(row=2, column=0, columnspan=2, sticky="ew", pady=(10, 0))
        note_frame.config(highlightbackground=COLORS['BORDER'], highlightthickness=1)

        note_label = tk.Label(
            note_frame,
            text="💡 Ces informations sont détectées automatiquement lors du chargement des fichiers",
            font=("Segoe UI", 8, "italic"),
            fg=COLORS['INFO'],
            bg=COLORS['LIGHT'],
            wraplength=300,
            justify=tk.CENTER
        )
        note_label.pack(pady=8)

    def _create_enhanced_info_field(self, parent: tk.Widget, row: int, col: int,
                                   icon: str, label: str, var: tk.StringVar):
        """Crée un champ d'information avec design modernisé."""
        field_frame = tk.Frame(parent, bg=COLORS['LIGHT'], relief='flat', bd=1)
        field_frame.grid(row=row, column=col, sticky="ew", padx=3, pady=3)
        field_frame.config(highlightbackground=COLORS['BORDER'], highlightthickness=1)

        # Contenu avec padding amélioré
        content = tk.Frame(field_frame, bg=COLORS['LIGHT'])
        content.pack(fill=tk.X, padx=8, pady=6)

        # En-tête avec icône
        header = tk.Frame(content, bg=COLORS['LIGHT'])
        header.pack(fill=tk.X)

        tk.Label(
            header,
            text=f"{icon} {label}",
            font=("Segoe UI", 9, "bold"),
            fg=COLORS['TEXT_PRIMARY'],
            bg=COLORS['LIGHT']
        ).pack(side=tk.LEFT)

        # Valeur avec style amélioré - Texte en noir
        value_label = tk.Label(
            content,
            textvariable=var,
            font=("Segoe UI", 9),
            fg=COLORS['TEXT_PRIMARY'],  # Changé de ACCENT à TEXT_PRIMARY pour texte noir
            bg=COLORS['LIGHT'],
            wraplength=120,
            justify=tk.LEFT
        )
        value_label.pack(anchor=tk.W, pady=(3, 0))

        # Stocker la référence pour les mises à jour
        self.info_displays[label.lower()] = value_label

    def _create_enhanced_analysis_quadrant(self, parent: tk.Widget, row: int, col: int):
        """Quadrant 3: Analyse et critères avec style Sofrecom."""
        # Carte avec style Sofrecom
        card_container = tk.Frame(parent, bg=COLORS['BG'])
        card_container.grid(row=row, column=col, sticky="nsew", padx=10, pady=10)

        frame = tk.Frame(card_container, bg=COLORS['CARD'], relief='flat', bd=0)
        frame.pack(fill=tk.BOTH, expand=True, padx=1, pady=1)
        frame.config(highlightbackground=COLORS['BORDER'], highlightthickness=1)

        # En-tête avec style Sofrecom
        title_frame = tk.Frame(frame, bg=COLORS['CARD'])
        title_frame.pack(fill=tk.X, padx=8, pady=(8, 4))

        # Icône et titre
        icon_label = tk.Label(
            title_frame,
            text="⚙️",
            font=("Segoe UI", 12),
            fg=COLORS['SECONDARY'],
            bg=COLORS['CARD']
        )
        icon_label.pack(side=tk.LEFT, padx=(0, 6))

        title_label = tk.Label(
            title_frame,
            text="Analyse & Critères",
            font=("Segoe UI", 10, "bold"),
            fg=COLORS['INFO'],
            bg=COLORS['CARD']
        )
        title_label.pack(side=tk.LEFT)

        # Contenu compact
        content = tk.Frame(frame, bg=COLORS['CARD'])
        content.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0, 8))

        # Informations sur les critères avec design compact mais lisible
        criteria_info = tk.Frame(content, bg=COLORS['LIGHT'], relief='flat', bd=1)
        criteria_info.pack(fill=tk.X, pady=(0, 8))
        criteria_info.config(highlightbackground=COLORS['BORDER'], highlightthickness=1)

        criteria_label = tk.Label(
            criteria_info,
            text="🔍 5 Critères de Contrôle Qualité",
            font=("Segoe UI", 9, "bold"),
            fg=COLORS['TEXT_PRIMARY'],
            bg=COLORS['LIGHT']
        )
        criteria_label.pack(pady=6)

        # Boutons d'action avec design amélioré
        buttons_frame = tk.Frame(content, bg=COLORS['CARD'])
        buttons_frame.pack(fill=tk.X, pady=(5, 0))

        # Bouton Analyser avec style Sofrecom
        self.analyze_button = tk.Button(
            buttons_frame,
            text="🔍 Analyser",
            font=("Segoe UI", 9, "bold"),
            fg='white',
            bg=COLORS['PRIMARY'],
            activebackground=COLORS['PRIMARY_LIGHT'],
            activeforeground='white',
            relief='flat',
            padx=12,
            pady=4,
            cursor='hand2',
            command=self._run_quality_analysis
        )
        self.analyze_button.pack(side=tk.LEFT, padx=(0, 6))

        # Effet hover Sofrecom
        def on_analyze_enter(e):
            self.analyze_button.config(bg=COLORS['PRIMARY_LIGHT'])
        def on_analyze_leave(e):
            self.analyze_button.config(bg=COLORS['PRIMARY'])

        self.analyze_button.bind("<Enter>", on_analyze_enter)
        self.analyze_button.bind("<Leave>", on_analyze_leave)

        # Bouton Export avec style Sofrecom secondaire
        self.export_button = tk.Button(
            buttons_frame,
            text="📊 Exporter",
            font=("Segoe UI", 9, "bold"),
            fg='white',
            bg=COLORS['SECONDARY'],
            activebackground=COLORS['SECONDARY_LIGHT'],
            activeforeground='white',
            relief='flat',
            padx=12,
            pady=4,
            cursor='hand2',
            command=self._export_qc_report
        )
        self.export_button.pack(side=tk.LEFT)

        # Effet hover Sofrecom
        def on_export_enter(e):
            self.export_button.config(bg=COLORS['SECONDARY_LIGHT'])
        def on_export_leave(e):
            self.export_button.config(bg=COLORS['SECONDARY'])

        self.export_button.bind("<Enter>", on_export_enter)
        self.export_button.bind("<Leave>", on_export_leave)

        # Barre de progression compacte
        progress_frame = tk.Frame(content, bg=COLORS['CARD'])
        progress_frame.pack(fill=tk.X, pady=(8, 0))

        tk.Label(
            progress_frame,
            text="📈 Progression",
            font=("Segoe UI", 8, "bold"),
            fg=COLORS['TEXT_SECONDARY'],
            bg=COLORS['CARD']
        ).pack(anchor=tk.W)

        # Conteneur pour la barre de progression avec bordure
        progress_container = tk.Frame(progress_frame, bg=COLORS['BORDER'], height=8)
        progress_container.pack(fill=tk.X, pady=(5, 0))
        progress_container.pack_propagate(False)

        # Barre de progression avec couleur dynamique
        self.progress_bar = tk.Frame(progress_container, bg=COLORS['SUCCESS'], height=6)
        self.progress_bar.place(x=1, y=1, width=0, height=6)

    def _create_enhanced_results_quadrant(self, parent: tk.Widget, row: int, col: int):
        """Quadrant 4: Résultats avec style Sofrecom."""
        # Carte avec style Sofrecom
        card_container = tk.Frame(parent, bg=COLORS['BG'])
        card_container.grid(row=row, column=col, sticky="nsew", padx=10, pady=10)

        frame = tk.Frame(card_container, bg=COLORS['CARD'], relief='flat', bd=0)
        frame.pack(fill=tk.BOTH, expand=True, padx=1, pady=1)
        frame.config(highlightbackground=COLORS['BORDER'], highlightthickness=1)

        # En-tête avec couleur distinctive
        title_frame = tk.Frame(frame, bg=COLORS['INFO'], height=35)
        title_frame.pack(fill=tk.X)
        title_frame.pack_propagate(False)

        title_label = tk.Label(
            title_frame,
            text="📊 Résultats d'Analyse",
            font=("Segoe UI", 11, "bold"),
            fg='white',
            bg=COLORS['INFO']
        )
        title_label.pack(expand=True)

        # Contenu avec scrollbar pour les résultats
        content = tk.Frame(frame, bg=COLORS['CARD'])
        content.pack(fill=tk.BOTH, expand=True, padx=12, pady=10)

        # Zone de résultats avec style amélioré
        results_container = tk.Frame(content, bg=COLORS['LIGHT'], relief='flat', bd=1)
        results_container.pack(fill=tk.BOTH, expand=True)
        results_container.config(highlightbackground=COLORS['BORDER'], highlightthickness=1)

        # Créer le results_frame pour compatibilité avec les méthodes d'affichage
        self.results_frame = tk.Frame(results_container, bg=COLORS['LIGHT'])
        self.results_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Label de résultats avec style modernisé
        self.results_label = tk.Label(
            self.results_frame,
            text="🔄 En attente d'analyse...\n\n💡 Chargez les fichiers et lancez l'analyse pour voir les résultats détaillés",
            font=("Segoe UI", 9),
            fg=COLORS['TEXT_SECONDARY'],
            bg=COLORS['LIGHT'],
            wraplength=300,
            justify=tk.CENTER
        )
        self.results_label.pack(expand=True, pady=20)

    def _create_enhanced_status_bar(self):
        """Crée la barre de statut compacte avec style Sofrecom."""
        status_frame = tk.Frame(self.analysis_tab, bg=COLORS['LIGHT'], height=30)
        status_frame.grid(row=2, column=0, sticky="ew", padx=0, pady=0)
        status_frame.pack_propagate(False)
        status_frame.config(highlightbackground=COLORS['BORDER'], highlightthickness=1)

        # Contenu avec padding amélioré
        content = tk.Frame(status_frame, bg=COLORS['LIGHT'])
        content.pack(fill=tk.BOTH, expand=True, padx=15, pady=6)

        # Icône de statut avec animation potentielle
        self.status_icon = tk.Label(
            content,
            text="⚡",
            font=("Segoe UI", 12),
            fg=COLORS['SUCCESS'],
            bg=COLORS['LIGHT']
        )
        self.status_icon.pack(side=tk.LEFT)

        # Message de statut avec style amélioré
        self.status_label = tk.Label(
            content,
            text="Prêt - Module de Contrôle Qualité initialisé",
            font=("Segoe UI", 9),
            fg=COLORS['TEXT_PRIMARY'],
            bg=COLORS['LIGHT']
        )
        self.status_label.pack(side=tk.LEFT, padx=(8, 0))

        # Indicateur de temps/version à droite
        time_label = tk.Label(
            content,
            text="Pladria v3.0 | Module 5",
            font=("Segoe UI", 8),
            fg=COLORS['TEXT_SECONDARY'],
            bg=COLORS['LIGHT']
        )
        time_label.pack(side=tk.RIGHT)

    # ==========================================
    # MÉTHODES D'AMÉLIORATION DE L'EXPÉRIENCE UTILISATEUR
    # ==========================================

    def _update_progress_bar(self, percentage: float):
        """Met à jour la barre de progression avec animation."""
        try:
            if (hasattr(self, 'progress_bar') and
                self.progress_bar is not None and
                self.progress_bar.winfo_exists()):

                # Calculer la largeur basée sur le pourcentage
                container_width = self.progress_bar.master.winfo_width()
                if container_width > 2:  # S'assurer que le conteneur a une taille
                    new_width = max(0, min(container_width - 2, int((container_width - 2) * percentage / 100)))

                    # Changer la couleur selon le pourcentage
                    if percentage < 30:
                        color = COLORS['ERROR']
                    elif percentage < 70:
                        color = COLORS['WARNING']
                    else:
                        color = COLORS['SUCCESS']

                    self.progress_bar.config(bg=color)
                    self.progress_bar.place(width=new_width)
            else:
                # Fallback vers l'ancienne méthode si disponible
                if hasattr(self, 'progress_var') and self.progress_var is not None:
                    self.progress_var.set(percentage)
        except Exception as e:
            self.logger.error(f"Erreur mise à jour barre de progression: {e}")

    def _update_status_with_animation(self, message: str, icon: str = "⚡", color: str = None):
        """Met à jour le statut avec une animation visuelle."""
        try:
            if (hasattr(self, 'status_label') and
                self.status_label is not None and
                self.status_label.winfo_exists()):
                self.status_label.config(text=message)

            if (hasattr(self, 'status_icon') and
                self.status_icon is not None and
                self.status_icon.winfo_exists()):
                self.status_icon.config(text=icon)
                if color:
                    self.status_icon.config(fg=color)
            else:
                # Fallback vers l'ancienne méthode si disponible
                if hasattr(self, '_update_status'):
                    try:
                        self._update_status("info", message)
                    except:
                        pass  # Ignorer les erreurs de fallback

        except Exception as e:
            self.logger.error(f"Erreur mise à jour statut: {e}")

    def _animate_button_click(self, button: tk.Button):
        """Anime un bouton lors du clic."""
        try:
            if button is not None and button.winfo_exists():
                original_bg = button.cget('bg')
                button.config(bg=COLORS['ACCENT'])
                button.after(100, lambda: button.config(bg=original_bg) if button.winfo_exists() else None)
        except Exception as e:
            self.logger.error(f"Erreur animation bouton: {e}")

    # ==========================================
    # SYSTÈME DE SUIVI AUTOMATIQUE - RAPPORT EXCEL
    # ==========================================

    def _generate_tracking_report_async(self, files_data: List[Dict[str, Any]]):
        """Lance la génération du rapport Excel en arrière-plan."""
        if not files_data:
            return

        def generate_report():
            try:
                self._generate_tracking_report(files_data)
            except Exception as e:
                self.logger.error(f"Erreur génération rapport automatique: {e}")

        # Lancer en arrière-plan pour ne pas bloquer l'UI
        thread = threading.Thread(target=generate_report, daemon=True)
        thread.start()

    def _generate_tracking_report(self, files_data: List[Dict[str, Any]]):
        """Génère le rapport Excel de suivi automatique."""
        try:
            if not OPENPYXL_AVAILABLE:
                self.logger.warning("OpenPyXL non disponible - rapport Excel non généré")
                return

            # Définir le chemin du rapport
            report_dir = Path(r"C:\Users\welj\orange.com\BOT G2R - CM Adresses et Plan Adressage - CM Adresses et Plan Adressage\Suivis CMS Adresse_Plan Adressage\Contrôle Qualité\ZZZ_Suivi_Controle_Qualité")
            report_file = report_dir / "Suivi Controle Qualité.xlsx"

            # Créer le répertoire s'il n'existe pas
            report_dir.mkdir(parents=True, exist_ok=True)

            # Créer le workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Suivi Contrôle Qualité"

            # Définir les en-têtes
            headers = [
                "Commune",
                "ID Tâche PA",
                "Code INSEE",
                "Domaine",
                "Affectation",
                "Contrôleur",
                "Score Total",
                "Statut Commune"
            ]

            # Ajouter les en-têtes avec le style des fichiers état de lieu
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style="thin", color="000000"),
                    right=Side(style="thin", color="000000"),
                    top=Side(style="thin", color="000000"),
                    bottom=Side(style="thin", color="000000")
                )

            # Ajouter les données
            for row, file_info in enumerate(files_data, 2):
                data_row = [
                    file_info.get('commune', 'N/A'),
                    file_info.get('id_tache', 'N/A'),
                    file_info.get('insee', 'N/A'),
                    file_info.get('domaine', 'N/A'),
                    file_info.get('affectation', 'N/A'),
                    file_info.get('controleur', 'N/A'),
                    file_info.get('score_total', 'N/A'),
                    file_info.get('statut_commune', 'N/A')
                ]

                for col, value in enumerate(data_row, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.font = Font(size=10)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = Border(
                        left=Side(style="thin", color="000000"),
                        right=Side(style="thin", color="000000"),
                        top=Side(style="thin", color="000000"),
                        bottom=Side(style="thin", color="000000")
                    )

                    # Coloration conditionnelle pour le statut (style état de lieu)
                    if col == 8:  # Colonne Statut Commune
                        if value == "OK":
                            cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
                            cell.font = Font(bold=True, color="FFFFFF", size=10)
                        elif value in ["NOK", "NON CONFORME"]:
                            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                            cell.font = Font(bold=True, color="FFFFFF", size=10)

            # Ajuster la largeur des colonnes (style état de lieu)
            column_widths = [18, 14, 12, 12, 16, 14, 12, 16]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

            # Sauvegarder le fichier
            wb.save(report_file)

            self.logger.info(f"📊 Rapport Excel généré: {report_file}")

        except Exception as e:
            self.logger.error(f"Erreur génération rapport Excel: {e}")
