"""
Team Statistics module - Dashboard for analyzing team performance and statistics.
"""

import tkinter as tk
from tkinter import messagebox, filedialog, ttk
import logging
import os
import getpass
import re
from typing import Optional, List, Dict, Any
from pathlib import Path
import sys

# Ensure src directory is in path
src_path = Path(__file__).parent.parent.parent
if str(src_path) not in sys.path:
    sys.path.insert(0, str(src_path))

from config.constants import COLORS, UIConfig, TeamsConfig, AccessControl
from core import FileProcessor, DataValidator, ExcelGenerator
from utils.file_utils import get_icon_path, check_file_access, is_excel_file_open
from utils.lazy_imports import get_pandas
from utils.performance import run_async_task

from ui.styles import StyleManager, create_card_frame, create_section_header
from ui.responsive_utils import get_responsive_manager
from ui.components.scrollable_frame import create_scrollable_container
from ui.keyboard_shortcuts import KeyboardShortcutManager
from datetime import datetime, timedelta
import calendar

logger = logging.getLogger(__name__)

# Import password dialog with error handling
try:
    from ui.components.password_dialog import show_password_dialog
    logger.info("Successfully imported password_dialog")
except ImportError as e:
    logger.error(f"Failed to import password_dialog: {e}")
    # Fallback function
    def show_password_dialog(parent, title="", message=""):
        logger.warning("Using fallback password dialog")
        from tkinter import simpledialog
        password = simpledialog.askstring(title, message, show='*')
        return password is not None, password or ""
except Exception as e:
    logger.error(f"Unexpected error importing password_dialog: {e}")
    # Fallback function
    def show_password_dialog(parent, title="", message=""):
        logger.warning("Using fallback password dialog due to unexpected error")
        from tkinter import simpledialog
        password = simpledialog.askstring(title, message, show='*')
        return password is not None, password or ""


class TeamStatsModule:
    """Team Statistics module for analyzing team performance and statistics."""
    
    def __init__(self, parent: tk.Widget, navigation_manager=None):
        """
        Initialize the Team Statistics module.

        Args:
            parent: Parent widget
            navigation_manager: Navigation manager instance
        """
        self.parent = parent
        self.navigation_manager = navigation_manager
        self.logger = logging.getLogger(__name__)

        # Initialize responsive manager
        self.responsive_manager = get_responsive_manager()

        # Check password protection for Statistics Team module
        if not self._verify_password_access():
            self._create_access_denied_ui()
            return

        # Initialize core components
        try:
            self.file_processor = FileProcessor()
            self.data_validator = DataValidator()
            self.excel_generator = ExcelGenerator()
            self.logger.info("Core components initialized")
        except Exception as e:
            self.logger.error(f"Error initializing core components: {e}")
            self.file_processor = None
            self.data_validator = None
            self.excel_generator = None

        # Module data
        self.global_suivi_data = None
        self.team_statistics = {}
        self.collaborator_stats = {}
        self.ticket_status_breakdown = {}
        self.overall_averages = {}

        # DMT data
        self.dmt_data = {}
        self.monthly_dmt_file = None
        self.dmt_chart_data = {}

        # Chart components
        self.chart_frame = None
        self.chart_canvas = None
        self.collab_filter_var = None
        
        # Get dynamic Teams path for current user
        from config.constants import TeamsConfig
        self.teams_folder_path = TeamsConfig.get_global_teams_path()
        self.global_excel_filename = "Suivis Global Tickets CMS Adr_PA.xlsx"

        # UI components
        self.progress_var = None
        self.progress_bar = None
        self.status_label = None
        self.stats_display = None
        self.refresh_button = None
        self.export_buttons = None
        self.export_filters = None

        # Export filter variables
        self.collab_var = None
        self.month_var = None
        self.year_var = None
        self.collab_combo = None
        self.month_combo = None
        self.year_combo = None

        # Date range selection variables
        self.date_from_var = tk.StringVar()
        self.date_to_var = tk.StringVar()
        self.date_from_selected = None
        self.date_to_selected = None

        # Stats folder data
        self.stats_folder_data = {}
        self.filtered_statistics = {}
        self.dashboard_data = {}

        # Keyboard shortcuts (optional)
        self.keyboard_manager = None

        # Create UI first
        self._create_module_ui()

        # Initialize optional features after UI is created
        self.parent.after(100, self._initialize_optional_features)

        # Auto-load data after UI is ready - optimized timing
        self.parent.after(100, self._auto_load_data)

        # Force scroll update after UI is fully loaded
        self.parent.after(200, self._update_scroll_region)

        # Additional scroll updates to ensure all content is visible
        self.parent.after(500, self._update_scroll_region)
        self.parent.after(1000, self._update_scroll_region)

        # Debug layout after everything is loaded
        self.parent.after(1200, self._debug_layout)

    def _verify_password_access(self):
        """
        Verify password access for Statistics Team module.

        Returns:
            bool: True if access granted, False otherwise
        """
        try:
            self.logger.info("Starting password verification process")

            # Show password dialog
            self.logger.info("Showing password dialog")
            success, password = show_password_dialog(
                self.parent,
                title="🔐 Accès Module Statistiques Équipe",
                message="Ce module est protégé par mot de passe.\nVeuillez saisir le mot de passe pour continuer :"
            )

            self.logger.info(f"Password dialog result: success={success}, password_length={len(password) if password else 0}")

            if not success:
                self.logger.info("User cancelled password dialog")
                return False

            # Verify password
            self.logger.info("Verifying password with AccessControl")
            if AccessControl.verify_stats_password(password):
                self.logger.info("Password verification successful")
                return True
            else:
                self.logger.warning(f"Invalid password attempt: '{password}'")
                messagebox.showerror(
                    "Accès Refusé",
                    "Accès refusé - mot de passe invalide"
                )
                return False

        except Exception as e:
            self.logger.error(f"Error during password verification: {e}")
            import traceback
            self.logger.error(f"Full traceback: {traceback.format_exc()}")
            messagebox.showerror(
                "Erreur",
                f"Erreur lors de la vérification du mot de passe:\n{e}"
            )
            return False

    def _retry_password_access(self):
        """Retry password access."""
        try:
            # Clear the parent
            for widget in self.parent.winfo_children():
                widget.destroy()

            # Reinitialize with password check
            self.__init__(self.parent, self.navigation_manager)

        except Exception as e:
            self.logger.error(f"Retry password access failed: {e}")
            messagebox.showerror("Erreur", f"Impossible de réinitialiser le module:\n{e}")

    def _create_access_denied_ui(self):
        """Create modern UI for access denied scenario (password protection)."""
        # Clear any existing content
        for widget in self.parent.winfo_children():
            widget.destroy()

        # Main container with modern layout
        main_frame = tk.Frame(self.parent, bg=COLORS['BG'])
        main_frame.pack(fill=tk.BOTH, expand=True, padx=40, pady=40)

        # Center container with modern card design
        center_container = tk.Frame(main_frame, bg=COLORS['BG'])
        center_container.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

        # Modern card with subtle shadow
        card_frame = tk.Frame(center_container, bg=COLORS['CARD'], relief=tk.FLAT, bd=0)
        card_frame.pack(padx=20, pady=20)

        # Add subtle border
        border_frame = tk.Frame(card_frame, bg=COLORS['BORDER'], height=1)
        border_frame.pack(fill=tk.X, side=tk.TOP)

        # Content container
        content_frame = tk.Frame(card_frame, bg=COLORS['CARD'])
        content_frame.pack(fill=tk.BOTH, padx=40, pady=35)

        # Header section with Sofrecom branding
        header_frame = tk.Frame(content_frame, bg=COLORS['CARD'])
        header_frame.pack(fill=tk.X, pady=(0, 30))

        # Modern security icon with accent background
        icon_container = tk.Frame(header_frame, bg=COLORS['ACCENT'], width=80, height=80)
        icon_container.pack(pady=(0, 20))
        icon_container.pack_propagate(False)

        icon_label = tk.Label(
            icon_container,
            text="🔐",
            font=("Segoe UI", 32),
            bg=COLORS['ACCENT'],
            fg=COLORS['PRIMARY']
        )
        icon_label.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

        # Modern title with Sofrecom typography
        title_label = tk.Label(
            header_frame,
            text="Accès Refusé",
            font=UIConfig.FONT_HEADER,
            fg=COLORS['DANGER'],
            bg=COLORS['CARD']
        )
        title_label.pack()

        # Subtitle
        subtitle_label = tk.Label(
            header_frame,
            text="Module Statistiques Équipe",
            font=UIConfig.FONT_SUBTITLE,
            fg=COLORS['TEXT_SECONDARY'],
            bg=COLORS['CARD']
        )
        subtitle_label.pack(pady=(5, 0))

        # Message section with better typography
        message_frame = tk.Frame(content_frame, bg=COLORS['CARD'])
        message_frame.pack(fill=tk.X, pady=(0, 25))

        message_label = tk.Label(
            message_frame,
            text="L'authentification a échoué ou a été annulée.\nCe module contient des données sensibles et nécessite un accès autorisé.",
            font=UIConfig.FONT_SUBTITLE,
            fg=COLORS['INFO'],
            bg=COLORS['CARD'],
            wraplength=450,
            justify=tk.CENTER
        )
        message_label.pack()

        # Security information card
        info_card = tk.Frame(content_frame, bg=COLORS['LIGHT'], relief=tk.FLAT, bd=0)
        info_card.pack(fill=tk.X, pady=(0, 25))

        info_border = tk.Frame(info_card, bg=COLORS['BORDER'], height=1)
        info_border.pack(fill=tk.X, side=tk.TOP)

        info_content = tk.Frame(info_card, bg=COLORS['LIGHT'])
        info_content.pack(fill=tk.X, padx=20, pady=15)

        info_title = tk.Label(
            info_content,
            text="🛡️ Sécurité Renforcée",
            font=UIConfig.FONT_SUBTITLE,
            fg=COLORS['PRIMARY'],
            bg=COLORS['LIGHT']
        )
        info_title.pack(anchor=tk.W, pady=(0, 8))

        info_items = [
            "✓ Protection par authentification",
            "✓ Données sensibles sécurisées",
            "✓ Accès contrôlé et audité"
        ]

        for item in info_items:
            item_label = tk.Label(
                info_content,
                text=item,
                font=UIConfig.FONT_SMALL,
                fg=COLORS['TEXT_SECONDARY'],
                bg=COLORS['LIGHT']
            )
            item_label.pack(anchor=tk.W, pady=1)

        # Action buttons with modern styling
        buttons_frame = tk.Frame(content_frame, bg=COLORS['CARD'])
        buttons_frame.pack(fill=tk.X, pady=(0, 10))

        # Retry button - primary action
        retry_button = tk.Button(
            buttons_frame,
            text="🔄 Réessayer l'authentification",
            font=UIConfig.FONT_BUTTON,
            bg=COLORS['PRIMARY'],
            fg=COLORS['WHITE'],
            relief=tk.FLAT,
            bd=0,
            padx=25,
            pady=12,
            command=self._retry_password_access,
            cursor='hand2'
        )
        retry_button.pack(side=tk.RIGHT, padx=(10, 0))

        # Back button - secondary action
        back_button = tk.Button(
            buttons_frame,
            text="← Retour à l'accueil",
            font=UIConfig.FONT_BUTTON,
            bg=COLORS['LIGHT'],
            fg=COLORS['TEXT_SECONDARY'],
            relief=tk.FLAT,
            bd=0,
            padx=25,
            pady=12,
            command=self._go_back_home,
            cursor='hand2'
        )
        back_button.pack(side=tk.RIGHT)

        # Contact information
        contact_frame = tk.Frame(content_frame, bg=COLORS['CARD'])
        contact_frame.pack(fill=tk.X)

        contact_label = tk.Label(
            contact_frame,
            text="💬 Pour obtenir l'accès, contactez l'administrateur système",
            font=UIConfig.FONT_SMALL,
            fg=COLORS['TEXT_MUTED'],
            bg=COLORS['CARD'],
            justify=tk.CENTER
        )
        contact_label.pack()

        # Add hover effects to buttons
        self._add_access_denied_hover_effects(retry_button, back_button)

    def _add_access_denied_hover_effects(self, retry_button, back_button):
        """Add modern hover effects to access denied buttons."""
        # Retry button hover effects
        def on_retry_enter(e):
            retry_button.config(bg=COLORS['PRIMARY_DARK'])

        def on_retry_leave(e):
            retry_button.config(bg=COLORS['PRIMARY'])

        retry_button.bind('<Enter>', on_retry_enter)
        retry_button.bind('<Leave>', on_retry_leave)

        # Back button hover effects
        def on_back_enter(e):
            back_button.config(bg=COLORS['BORDER'])

        def on_back_leave(e):
            back_button.config(bg=COLORS['LIGHT'])

        back_button.bind('<Enter>', on_back_enter)
        back_button.bind('<Leave>', on_back_leave)

    def _go_back_home(self):
        """Navigate back to home screen."""
        try:
            if self.navigation_manager:
                from ui.navigation import NavigationState
                self.navigation_manager.navigate_to(NavigationState.HOME)
            else:
                self.logger.warning("No navigation manager available")
        except Exception as e:
            self.logger.error(f"Error navigating to home: {e}")

    def _initialize_optional_features(self):
        """Initialize optional features after UI is ready."""
        try:
            # Setup keyboard shortcuts
            if self.navigation_manager and hasattr(self.navigation_manager, 'root'):
                self.keyboard_manager = KeyboardShortcutManager(self.navigation_manager.root)
                self._setup_module_shortcuts()

            self.logger.info("Optional features initialized successfully")

        except Exception as e:
            self.logger.error(f"Error initializing optional features: {e}")

    def _auto_load_data(self):
        """Automatically load data when the module is opened - optimized."""
        try:
            self.logger.info("Auto-loading data...")
            self._load_global_data()

            # Force scroll update after data loading
            self.parent.after(100, self._update_scroll_region)

        except Exception as e:
            self.logger.error(f"Error during auto-load: {e}")

    def _update_scroll_region(self):
        """Force update of scroll region to ensure all content is accessible."""
        try:
            # Update the scrollable container if it exists
            if hasattr(self, 'scrollable_container') and self.scrollable_container:
                self.scrollable_container.update_scroll_region()
                self.logger.info("Scroll region updated successfully")
            else:
                self.logger.warning("Scrollable container not found")
        except Exception as e:
            self.logger.error(f"Error updating scroll region: {e}")

    def _debug_layout(self):
        """Debug layout information to help identify scrolling issues."""
        try:
            if hasattr(self, 'scrollable_container') and self.scrollable_container:
                canvas = self.scrollable_container.canvas
                scrollable_frame = self.scrollable_container.get_scrollable_frame()

                # Get dimensions
                canvas_height = canvas.winfo_height()
                frame_height = scrollable_frame.winfo_reqheight()
                scroll_region = canvas.cget("scrollregion")

                self.logger.info(f"Layout Debug - Canvas height: {canvas_height}, Frame height: {frame_height}, Scroll region: {scroll_region}")

                # Force another scroll update if needed
                if frame_height > canvas_height:
                    self.scrollable_container.update_scroll_region()
                    self.logger.info("Additional scroll update applied due to content overflow")

        except Exception as e:
            self.logger.error(f"Error in layout debug: {e}")

    def _create_module_ui(self):
        """Create the module user interface."""
        try:
            # Module title bar
            self._create_title_bar()

            # Main content area
            self._create_content_area()

            self.logger.info("Module UI created successfully")

        except Exception as e:
            self.logger.error(f"Error creating module UI: {e}")
            self._create_error_ui(str(e))

    def _create_error_ui(self, error_message: str):
        """Create a simple error display when module fails to load."""
        error_frame = tk.Frame(self.parent, bg=COLORS['BG'])
        error_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        # Error icon and message
        error_label = tk.Label(
            error_frame,
            text=f"❌ Erreur lors du chargement du module:\n\n{error_message}",
            font=UIConfig.FONT_SUBTITLE,
            fg=COLORS['DANGER'],
            bg=COLORS['BG'],
            justify=tk.CENTER
        )
        error_label.pack(expand=True)

        # Retry button
        retry_btn = tk.Button(
            error_frame,
            text="🔄 Réessayer",
            command=self._retry_initialization,
            bg=COLORS['PRIMARY'],
            fg='white',
            font=UIConfig.FONT_BUTTON,
            relief='flat',
            padx=20,
            pady=10
        )
        retry_btn.pack(pady=10)

    def _retry_initialization(self):
        """Retry module initialization."""
        try:
            # Clear the parent
            for widget in self.parent.winfo_children():
                widget.destroy()

            # Reinitialize
            self.__init__(self.parent, self.navigation_manager)

        except Exception as e:
            self.logger.error(f"Retry failed: {e}")
            messagebox.showerror("Erreur", f"Impossible de réinitialiser le module:\n{e}")

    def _create_title_bar(self):
        """Create the module title bar."""
        title_frame = tk.Frame(self.parent, bg=COLORS['CARD'], height=60)
        title_frame.pack(fill=tk.X, padx=10, pady=(10, 0))
        title_frame.pack_propagate(False)
        
        # Title content
        title_content = tk.Frame(title_frame, bg=COLORS['CARD'])
        title_content.pack(fill=tk.BOTH, expand=True, padx=20, pady=15)
        
        # Module icon and title
        title_left = tk.Frame(title_content, bg=COLORS['CARD'])
        title_left.pack(side=tk.LEFT, fill=tk.Y)
        
        # Icon
        icon_label = tk.Label(
            title_left,
            text="📊",
            font=("Segoe UI", 20),
            fg=COLORS['PRIMARY'],
            bg=COLORS['CARD']
        )
        icon_label.pack(side=tk.LEFT, padx=(0, 10))
        
        # Title and description
        title_text_frame = tk.Frame(title_left, bg=COLORS['CARD'])
        title_text_frame.pack(side=tk.LEFT, fill=tk.Y)
        
        title_label = tk.Label(
            title_text_frame,
            text="Statistiques Équipe",
            font=UIConfig.FONT_HEADER,
            fg=COLORS['PRIMARY'],
            bg=COLORS['CARD']
        )
        title_label.pack(anchor=tk.W)
        
        subtitle_label = tk.Label(
            title_text_frame,
            text="Tableau de bord des performances de l'équipe",
            font=UIConfig.FONT_SMALL,
            fg=COLORS['INFO'],
            bg=COLORS['CARD']
        )
        subtitle_label.pack(anchor=tk.W)
        
        # Action buttons on the right
        title_right = tk.Frame(title_content, bg=COLORS['CARD'])
        title_right.pack(side=tk.RIGHT, fill=tk.Y)

        # Reset button
        reset_btn = ttk.Button(
            title_right,
            text="🔄 Réinitialiser",
            command=self._reset_module,
            style='CompactWarning.TButton'
        )
        reset_btn.pack(side=tk.RIGHT, padx=(5, 0))

    def _create_content_area(self):
        """Create the main content area with responsive layout and scrolling."""
        # Get responsive configuration with corporate-optimized spacing
        column_config = self.responsive_manager.get_responsive_column_config()
        corporate_spacing = self.responsive_manager.get_corporate_spacing()
        padding = corporate_spacing['section_margin']

        # Create scrollable container for the entire content
        scrollable_container = create_scrollable_container(
            self.parent,
            self.responsive_manager,
            bg=COLORS['BG']
        )
        scrollable_container.pack(fill=tk.BOTH, expand=True, padx=padding, pady=padding)

        # Store reference to scrollable container for later updates
        self.scrollable_container = scrollable_container

        # Get the scrollable frame to add content to
        content_container = scrollable_container.get_scrollable_frame()
        content_container.configure(bg=COLORS['BG'])
        content_container.grid_rowconfigure(0, weight=1)
        content_container.grid_columnconfigure(0, weight=1)

        # Top row container
        top_row = tk.Frame(content_container, bg=COLORS['BG'])
        top_row.grid(row=0, column=0, sticky="nsew", pady=(0, 5))
        top_row.grid_rowconfigure(0, weight=1)

        # Configure columns based on responsive settings
        if column_config['columns'] == 1:
            # Single column layout for small screens
            top_row.grid_columnconfigure(0, weight=1, minsize=column_config['left_minsize'])
        elif column_config['columns'] == 2:
            # Two column layout for medium screens
            top_row.grid_columnconfigure(0, weight=column_config['left_weight'], minsize=column_config['left_minsize'])
            top_row.grid_columnconfigure(1, weight=column_config['middle_weight'], minsize=column_config['middle_minsize'])
        else:
            # Three column layout for large screens
            top_row.grid_columnconfigure(0, weight=column_config['left_weight'], minsize=column_config['left_minsize'])
            top_row.grid_columnconfigure(1, weight=column_config['middle_weight'], minsize=column_config['middle_minsize'])
            top_row.grid_columnconfigure(2, weight=column_config['right_weight'], minsize=column_config['right_minsize'])

        # Create columns based on responsive configuration
        if column_config['columns'] == 1:
            # Single column layout - stack everything vertically
            main_column = tk.Frame(top_row, bg=COLORS['BG'])
            main_column.grid(row=0, column=0, sticky="nsew")

            # Create sections in single column
            self._create_data_loading_section(main_column)
            self._create_overview_section(main_column)
            self._create_archive_section(main_column)
            self._create_statistics_section(main_column)
            self._create_export_section(main_column)

        elif column_config['columns'] == 2:
            # Two column layout
            left_column = tk.Frame(top_row, bg=COLORS['BG'])
            left_column.grid(row=0, column=0, sticky="nsew", padx=(0, 3))

            right_column = tk.Frame(top_row, bg=COLORS['BG'])
            right_column.grid(row=0, column=1, sticky="nsew", padx=(3, 0))

            # Distribute sections across two columns
            self._create_data_loading_section(left_column)
            self._create_overview_section(left_column)
            self._create_archive_section(left_column)

            self._create_statistics_section(right_column)
            self._create_export_section(right_column)

        else:
            # Three column layout (original)
            left_column = tk.Frame(top_row, bg=COLORS['BG'])
            left_column.grid(row=0, column=0, sticky="nsew", padx=(0, 3))

            middle_column = tk.Frame(top_row, bg=COLORS['BG'])
            middle_column.grid(row=0, column=1, sticky="nsew", padx=3)

            right_column = tk.Frame(top_row, bg=COLORS['BG'])
            right_column.grid(row=0, column=2, sticky="nsew", padx=(3, 0))

            # Create sections in three columns
            self._create_data_loading_section(left_column)
            self._create_overview_section(left_column)
            self._create_archive_section(left_column)

            self._create_statistics_section(middle_column)
            self._create_export_section(right_column)



    def _setup_module_shortcuts(self):
        """Set up keyboard shortcuts specific to this module."""
        self.keyboard_manager.set_callback("Control-l", self._load_shortcut)
        self.keyboard_manager.set_callback("Control-r", self._refresh_shortcut)
        self.keyboard_manager.set_callback("F5", self._refresh_shortcut)



    def _reset_module(self):
        """Reset the module to initial state."""
        try:
            self.global_suivi_data = None
            self.team_statistics.clear()
            self.collaborator_stats.clear()
            self.ticket_status_breakdown.clear()
            self.overall_averages.clear()

            if self.stats_display:
                # Clear statistics display
                for widget in self.stats_display.winfo_children():
                    widget.destroy()

            if self.status_label:
                self.status_label.config(text="Module réinitialisé - Redémarrez pour charger automatiquement")

            if self.progress_var:
                self.progress_var.set(0)

            self._disable_export_buttons()

            self.logger.info("Module reset successfully")

        except Exception as e:
            self.logger.error(f"Error resetting module: {e}")
            messagebox.showerror("Erreur", f"Erreur lors de la réinitialisation:\n{e}")

    def _create_data_loading_section(self, parent: tk.Widget):
        """Create the compact data loading section with auto-loading status."""
        # Loading card - more compact
        load_card = create_card_frame(parent)
        load_card.pack(fill=tk.X, pady=(0, 8))

        # Section header - reduced padding
        header_frame = create_section_header(load_card, "⚡", "Chargement automatique")
        header_frame.pack(fill=tk.X, padx=12, pady=(10, 5))

        # Content frame - reduced padding
        content_frame = tk.Frame(load_card, bg=COLORS['CARD'])
        content_frame.pack(fill=tk.X, padx=12, pady=(0, 10))

        # Progress bar - full width
        progress_frame = tk.Frame(content_frame, bg=COLORS['CARD'])
        progress_frame.pack(fill=tk.X)

        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(
            progress_frame,
            variable=self.progress_var,
            maximum=100,
            mode='determinate'
        )
        self.progress_bar.pack(fill=tk.X, pady=3)

        # Status label - smaller
        self.status_label = tk.Label(
            content_frame,
            text="Chargement automatique en cours...",
            font=("Segoe UI", 6),
            fg=COLORS['INFO'],
            bg=COLORS['CARD']
        )
        self.status_label.pack(anchor=tk.W, pady=(3, 0))

        # File status indicator
        status_frame = tk.Frame(content_frame, bg=COLORS['CARD'])
        status_frame.pack(fill=tk.X, pady=(5, 0))

        status_icon = tk.Label(
            status_frame,
            text="📁",
            font=("Segoe UI", 8),
            bg=COLORS['CARD'],
            fg=COLORS['INFO']
        )
        status_icon.pack(side=tk.LEFT, padx=(0, 3))

        self.file_status_label = tk.Label(
            status_frame,
            text="Vérification du statut du fichier...",
            font=("Segoe UI", 6),
            bg=COLORS['CARD'],
            fg=COLORS['INFO']
        )
        self.file_status_label.pack(side=tk.LEFT)

        # Refresh status button
        refresh_status_btn = tk.Button(
            status_frame,
            text="🔄",
            command=self._update_file_status_indicator,
            bg=COLORS['BORDER'],
            fg=COLORS['TEXT_SECONDARY'],
            font=("Segoe UI", 6),
            relief='flat',
            padx=3,
            pady=1
        )
        refresh_status_btn.pack(side=tk.RIGHT)

        # Initialize file status
        self._update_file_status_indicator()

    def _create_overview_section(self, parent: tk.Widget):
        """Create the overview section."""
        # Overview card
        overview_card = create_card_frame(parent)
        overview_card.pack(fill=tk.X, pady=(0, 10))

        # Section header
        header_frame = create_section_header(overview_card, "📈", "Aperçu général")
        header_frame.pack(fill=tk.X, padx=15, pady=(15, 10))

        # Content frame
        self.overview_content = tk.Frame(overview_card, bg=COLORS['CARD'])
        self.overview_content.pack(fill=tk.X, padx=15, pady=(0, 15))

        # Initial message
        initial_label = tk.Label(
            self.overview_content,
            text="Chargez les données pour voir l'aperçu général",
            font=UIConfig.FONT_SMALL,
            fg=COLORS['INFO'],
            bg=COLORS['CARD']
        )
        initial_label.pack(anchor=tk.W)

    def _create_archive_section(self, parent: tk.Widget):
        """Create the archive section for backing up treated communes."""
        try:
            self.logger.info("Creating archive section...")

            # Archive card - consistent with other sections
            archive_card = create_card_frame(parent)
            archive_card.pack(anchor=tk.N, fill=tk.X, padx=0, pady=(0, 5))  # Match export section style
            self.logger.debug("Archive card created")

            # Section header - using consistent header style
            header_frame = create_section_header(archive_card, "📦", "Archivage Communes")
            header_frame.pack(fill=tk.X, padx=15, pady=(15, 10))  # Match other sections
            self.logger.debug("Archive header created")

            # Content frame - consistent with other sections
            archive_content = tk.Frame(archive_card, bg=COLORS['CARD'])
            archive_content.pack(fill=tk.X, padx=15, pady=(0, 15))  # Match other sections

            # Archive info - more compact
            info_text = tk.Label(
                archive_content,
                text="Archive ZIP des communes traitées du mois sélectionné",
                font=UIConfig.FONT_SMALL,
                fg=COLORS['INFO'],
                bg=COLORS['CARD'],
                justify=tk.LEFT
            )
            info_text.pack(anchor=tk.W, pady=(0, 10))
            self.logger.debug("Archive info created")

            # Month selection frame
            month_frame = tk.Frame(archive_content, bg=COLORS['CARD'])
            month_frame.pack(fill=tk.X, pady=(0, 15))

            month_label = tk.Label(
                month_frame,
                text="Mois à archiver:",
                font=UIConfig.FONT_SMALL,
                fg=COLORS['TEXT_PRIMARY'],
                bg=COLORS['CARD']
            )
            month_label.pack(side=tk.LEFT, padx=(0, 10))

            # Month selection combobox
            from tkinter import ttk
            self.archive_month_var = tk.StringVar()
            self.archive_month_combo = ttk.Combobox(
                month_frame,
                textvariable=self.archive_month_var,
                values=["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                       "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"],
                state="readonly",
                width=12,
                font=UIConfig.FONT_SMALL
            )
            self.archive_month_combo.pack(side=tk.LEFT)

            # Set current month as default
            from datetime import datetime
            current_month = datetime.now().month
            month_names = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                          "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
            self.archive_month_var.set(month_names[current_month - 1])

            # Archive button
            self.archive_button = tk.Button(
                archive_content,
                text="📦 Créer Archive du Mois",
                font=UIConfig.FONT_BUTTON,
                bg=COLORS['PRIMARY'],
                fg='white',
                relief=tk.FLAT,
                padx=15,
                pady=6,
                cursor='hand2',
                command=self._create_monthly_archive,
                state=tk.DISABLED,  # Initially disabled until data is loaded
                activebackground=COLORS['PRIMARY_LIGHT'],
                borderwidth=0
            )
            self.archive_button.pack(pady=(5, 0))

            # Archive status label
            self.archive_status_label = tk.Label(
                archive_content,
                text="Chargez les données pour activer",
                font=UIConfig.FONT_SMALL,
                fg=COLORS['INFO'],
                bg=COLORS['CARD']
            )
            self.archive_status_label.pack(pady=(5, 0))

            self.logger.info("Archive section created successfully")

        except Exception as e:
            self.logger.error(f"Error creating archive section: {e}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")

            # Create a simple fallback section with consistent styling
            try:
                # Fallback archive card
                fallback_card = create_card_frame(parent)
                fallback_card.pack(anchor=tk.N, fill=tk.X, padx=0, pady=(0, 5))

                # Fallback section header
                fallback_header = create_section_header(fallback_card, "📦", "Archivage Communes Traitées")
                fallback_header.pack(fill=tk.X, padx=15, pady=(15, 10))

                # Fallback content frame
                fallback_content = tk.Frame(fallback_card, bg=COLORS['CARD'])
                fallback_content.pack(fill=tk.X, padx=15, pady=(0, 15))

                # Error message
                error_label = tk.Label(
                    fallback_content,
                    text="⚠️ Erreur de chargement de la section principale",
                    font=UIConfig.FONT_SMALL,
                    fg=COLORS['WARNING'],
                    bg=COLORS['CARD']
                )
                error_label.pack(anchor=tk.W, pady=(0, 10))

                # Month selection for fallback
                month_fallback_frame = tk.Frame(fallback_content, bg=COLORS['CARD'])
                month_fallback_frame.pack(fill=tk.X, pady=(0, 15))

                month_fallback_label = tk.Label(
                    month_fallback_frame,
                    text="Mois à archiver:",
                    font=UIConfig.FONT_SMALL,
                    fg=COLORS['TEXT_PRIMARY'],
                    bg=COLORS['CARD']
                )
                month_fallback_label.pack(side=tk.LEFT, padx=(0, 10))

                # Fallback month combobox
                if not hasattr(self, 'archive_month_var'):
                    self.archive_month_var = tk.StringVar()
                    from datetime import datetime
                    current_month = datetime.now().month
                    month_names = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                                  "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
                    self.archive_month_var.set(month_names[current_month - 1])

                self.archive_month_combo_fallback = ttk.Combobox(
                    month_fallback_frame,
                    textvariable=self.archive_month_var,
                    values=["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                           "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"],
                    state="readonly",
                    width=12,
                    font=UIConfig.FONT_SMALL
                )
                self.archive_month_combo_fallback.pack(side=tk.LEFT)

                # Fallback button frame
                fallback_button_frame = tk.Frame(fallback_content, bg=COLORS['CARD'])
                fallback_button_frame.pack(fill=tk.X, pady=(0, 5))

                self.archive_button = tk.Button(
                    fallback_button_frame,
                    text="📦 Créer Archive du Mois Sélectionné",
                    font=UIConfig.FONT_BUTTON,
                    bg=COLORS['PRIMARY'],
                    fg='white',
                    relief=tk.FLAT,
                    padx=20,
                    pady=8,
                    cursor='hand2',
                    command=self._create_monthly_archive,
                    state=tk.DISABLED,
                    activebackground=COLORS['PRIMARY_LIGHT'],
                    borderwidth=0
                )
                self.archive_button.pack(side=tk.LEFT)

                self.archive_status_label = tk.Label(
                    fallback_button_frame,
                    text="Chargez d'abord les données pour activer l'archivage",
                    font=UIConfig.FONT_SMALL,
                    fg=COLORS['INFO'],
                    bg=COLORS['CARD']
                )
                self.archive_status_label.pack()

                self.logger.info("Fallback archive section created")

            except Exception as fallback_error:
                self.logger.error(f"Failed to create fallback archive section: {fallback_error}")

    def _create_date_range_section(self, parent: tk.Widget):
        """Create the date range selection section for filtered statistics."""
        try:
            self.logger.info("Creating date range selection section...")

            # Date range card
            date_card = create_card_frame(parent)
            date_card.pack(fill=tk.X, padx=0, pady=(0, 10))

            # Section header
            header_frame = tk.Frame(date_card, bg=COLORS['CARD'])
            header_frame.pack(fill=tk.X, padx=15, pady=(8, 5))

            # Header icon and title
            header_icon = tk.Label(
                header_frame,
                text="📅",
                font=("Segoe UI", 12),
                fg=COLORS['PRIMARY'],
                bg=COLORS['CARD']
            )
            header_icon.pack(side=tk.LEFT, padx=(0, 8))

            header_label = tk.Label(
                header_frame,
                text="Filtrage par période",
                font=UIConfig.FONT_HEADER,
                fg=COLORS['PRIMARY'],
                bg=COLORS['CARD']
            )
            header_label.pack(side=tk.LEFT)

            # Content frame
            content_frame = tk.Frame(date_card, bg=COLORS['CARD'])
            content_frame.pack(fill=tk.X, padx=15, pady=(0, 15))

            # Date from section
            from_frame = tk.Frame(content_frame, bg=COLORS['CARD'])
            from_frame.pack(fill=tk.X, pady=(0, 8))

            from_label = tk.Label(
                from_frame,
                text="Date de début:",
                font=("Segoe UI", 8, "bold"),
                fg=COLORS['PRIMARY'],
                bg=COLORS['CARD']
            )
            from_label.pack(anchor=tk.W, pady=(0, 2))

            from_input_frame = tk.Frame(from_frame, bg=COLORS['CARD'])
            from_input_frame.pack(fill=tk.X)

            self.date_from_entry = tk.Entry(
                from_input_frame,
                textvariable=self.date_from_var,
                font=("Segoe UI", 8),
                width=12,
                state='readonly'
            )
            self.date_from_entry.pack(side=tk.LEFT, padx=(0, 5))

            from_button = tk.Button(
                from_input_frame,
                text="📅",
                font=("Segoe UI", 8),
                command=self._show_date_from_picker,
                bg=COLORS['ACCENT'],
                fg=COLORS['PRIMARY'],
                relief='flat',
                padx=8,
                pady=2,
                cursor='hand2'
            )
            from_button.pack(side=tk.LEFT)

            # Date to section
            to_frame = tk.Frame(content_frame, bg=COLORS['CARD'])
            to_frame.pack(fill=tk.X, pady=(0, 8))

            to_label = tk.Label(
                to_frame,
                text="Date de fin:",
                font=("Segoe UI", 8, "bold"),
                fg=COLORS['PRIMARY'],
                bg=COLORS['CARD']
            )
            to_label.pack(anchor=tk.W, pady=(0, 2))

            to_input_frame = tk.Frame(to_frame, bg=COLORS['CARD'])
            to_input_frame.pack(fill=tk.X)

            self.date_to_entry = tk.Entry(
                to_input_frame,
                textvariable=self.date_to_var,
                font=("Segoe UI", 8),
                width=12,
                state='readonly'
            )
            self.date_to_entry.pack(side=tk.LEFT, padx=(0, 5))

            to_button = tk.Button(
                to_input_frame,
                text="📅",
                font=("Segoe UI", 8),
                command=self._show_date_to_picker,
                bg=COLORS['ACCENT'],
                fg=COLORS['PRIMARY'],
                relief='flat',
                padx=8,
                pady=2,
                cursor='hand2'
            )
            to_button.pack(side=tk.LEFT)

            # Action buttons
            button_frame = tk.Frame(content_frame, bg=COLORS['CARD'])
            button_frame.pack(fill=tk.X, pady=(8, 0))

            # Generate filtered stats button
            self.generate_stats_button = tk.Button(
                button_frame,
                text="📊 Générer et ouvrir index",
                font=("Segoe UI", 8, "bold"),
                bg=COLORS['PRIMARY'],
                fg='white',
                relief='flat',
                padx=12,
                pady=6,
                cursor='hand2',
                command=self._generate_filtered_statistics,
                state=tk.DISABLED
            )
            self.generate_stats_button.pack(side=tk.LEFT, padx=(0, 5))

            # Clear dates button
            clear_button = tk.Button(
                button_frame,
                text="🗑️ Effacer",
                font=("Segoe UI", 8),
                bg=COLORS['SECONDARY'],
                fg='white',
                relief='flat',
                padx=8,
                pady=6,
                cursor='hand2',
                command=self._clear_date_range
            )
            clear_button.pack(side=tk.LEFT)

            # Status label
            self.date_range_status = tk.Label(
                content_frame,
                text="Sélectionnez une période pour générer et ouvrir l'index stats",
                font=UIConfig.FONT_SMALL,
                fg=COLORS['INFO'],
                bg=COLORS['CARD']
            )
            self.date_range_status.pack(anchor=tk.W, pady=(5, 0))

            self.logger.info("Date range selection section created successfully")

        except Exception as e:
            self.logger.error(f"Error creating date range section: {e}")

    def _show_date_from_picker(self):
        """Show date picker for start date."""
        try:
            from tkinter import simpledialog
            import datetime

            # Create a simple date input dialog
            date_str = simpledialog.askstring(
                "Date de début",
                "Entrez la date de début (YYYY-MM-DD):",
                initialvalue=datetime.date.today().strftime("%Y-%m-%d")
            )

            if date_str:
                try:
                    # Validate date format
                    date_obj = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
                    self.date_from_selected = date_obj
                    self.date_from_var.set(date_str)
                    self._update_date_range_status()
                    self.logger.info(f"Start date selected: {date_str}")
                except ValueError:
                    messagebox.showerror("Erreur", "Format de date invalide. Utilisez YYYY-MM-DD")

        except Exception as e:
            self.logger.error(f"Error showing date from picker: {e}")
            messagebox.showerror("Erreur", f"Erreur lors de la sélection de date: {e}")

    def _show_date_to_picker(self):
        """Show date picker for end date."""
        try:
            from tkinter import simpledialog
            import datetime

            # Create a simple date input dialog
            date_str = simpledialog.askstring(
                "Date de fin",
                "Entrez la date de fin (YYYY-MM-DD):",
                initialvalue=datetime.date.today().strftime("%Y-%m-%d")
            )

            if date_str:
                try:
                    # Validate date format
                    date_obj = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
                    self.date_to_selected = date_obj
                    self.date_to_var.set(date_str)
                    self._update_date_range_status()
                    self.logger.info(f"End date selected: {date_str}")
                except ValueError:
                    messagebox.showerror("Erreur", "Format de date invalide. Utilisez YYYY-MM-DD")

        except Exception as e:
            self.logger.error(f"Error showing date to picker: {e}")
            messagebox.showerror("Erreur", f"Erreur lors de la sélection de date: {e}")

    def _clear_date_range(self):
        """Clear the selected date range."""
        try:
            self.date_from_var.set("")
            self.date_to_var.set("")
            self.date_from_selected = None
            self.date_to_selected = None
            self._update_date_range_status()
            self.logger.info("Date range cleared")

        except Exception as e:
            self.logger.error(f"Error clearing date range: {e}")

    def _update_date_range_status(self):
        """Update the status label for date range selection."""
        try:
            if self.date_from_selected and self.date_to_selected:
                if self.date_from_selected <= self.date_to_selected:
                    days_diff = (self.date_to_selected - self.date_from_selected).days + 1
                    status_text = f"✅ Période sélectionnée: {days_diff} jour(s)"
                    self.date_range_status.config(text=status_text, fg=COLORS['SUCCESS'])

                    # Enable generate button if data is loaded
                    if hasattr(self, 'data') and self.data is not None:
                        self.generate_stats_button.config(state=tk.NORMAL)
                else:
                    self.date_range_status.config(
                        text="❌ La date de début doit être antérieure à la date de fin",
                        fg=COLORS['ERROR']
                    )
                    self.generate_stats_button.config(state=tk.DISABLED)
            elif self.date_from_selected or self.date_to_selected:
                self.date_range_status.config(
                    text="⚠️ Sélectionnez les deux dates pour filtrer",
                    fg=COLORS['WARNING']
                )
                self.generate_stats_button.config(state=tk.DISABLED)
            else:
                self.date_range_status.config(
                    text="Sélectionnez une période pour générer et ouvrir l'index stats",
                    fg=COLORS['INFO']
                )
                self.generate_stats_button.config(state=tk.DISABLED)

        except Exception as e:
            self.logger.error(f"Error updating date range status: {e}")

    def _create_statistics_section(self, parent: tk.Widget):
        """Create the detailed statistics section."""
        # Statistics card - balanced sizing with no shadow to eliminate spacing
        stats_card = create_card_frame(parent, shadow=False)
        stats_card.pack(fill=tk.BOTH, expand=True, padx=0, pady=0)  # Fill available space but controlled

        # Section header with refresh button - no padding at all
        header_frame = tk.Frame(stats_card, bg=COLORS['CARD'])
        header_frame.pack(fill=tk.X, padx=15, pady=0)  # Completely removed all padding

        # Header icon and title
        header_icon = tk.Label(
            header_frame,
            text="📊",
            font=("Segoe UI", 12),
            fg=COLORS['PRIMARY'],
            bg=COLORS['CARD']
        )
        header_icon.pack(side=tk.LEFT, padx=(0, 8))

        header_label = tk.Label(
            header_frame,
            text="Statistiques détaillées",
            font=UIConfig.FONT_HEADER,
            fg=COLORS['PRIMARY'],
            bg=COLORS['CARD']
        )
        header_label.pack(side=tk.LEFT)

        # Refresh button
        self.refresh_button = tk.Button(
            header_frame,
            text="🔄 Actualiser",
            command=self._refresh_statistics,
            bg=COLORS['SECONDARY'],
            fg='white',
            font=UIConfig.FONT_SMALL,
            relief='flat',
            padx=10,
            pady=3,
            state=tk.DISABLED
        )
        self.refresh_button.pack(side=tk.RIGHT)

        # Statistics display area with scrollbar - balanced sizing
        stats_frame = tk.Frame(stats_card, bg=COLORS['CARD'])
        stats_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=(0, 8))  # Fill available space

        # Create scrollable frame with mouse wheel support - adaptive height
        canvas = tk.Canvas(stats_frame, bg=COLORS['CARD'], highlightthickness=0)
        scrollbar = ttk.Scrollbar(stats_frame, orient="vertical", command=canvas.yview)
        self.stats_display = tk.Frame(canvas, bg=COLORS['CARD'])

        # Store canvas window reference for width configuration
        self.canvas_window = None

        # Configure canvas to update scroll region when content changes
        def _configure_scroll_region(event=None):
            canvas.configure(scrollregion=canvas.bbox("all"))
            # Also configure the canvas width to match the frame
            canvas_width = canvas.winfo_width()
            if canvas_width > 1 and self.canvas_window:  # Avoid setting width to 0
                canvas.itemconfig(self.canvas_window, width=canvas_width)

        self.stats_display.bind("<Configure>", _configure_scroll_region)

        # Create window and store reference for width configuration
        self.canvas_window = canvas.create_window((0, 0), window=self.stats_display, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # Store canvas reference for dynamic height adjustment
        self.stats_canvas = canvas

        # Add mouse wheel scrolling support
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")

        def _bind_mousewheel(event):
            canvas.bind_all("<MouseWheel>", _on_mousewheel)

        def _unbind_mousewheel(event):
            canvas.unbind_all("<MouseWheel>")

        canvas.bind('<Enter>', _bind_mousewheel)
        canvas.bind('<Leave>', _unbind_mousewheel)

        # Configure canvas to update when resized
        canvas.bind('<Configure>', _configure_scroll_region)

        canvas.pack(side="left", fill="both", expand=True)  # Fill available space
        scrollbar.pack(side="right", fill="y")

        # Initial message
        initial_stats_label = tk.Label(
            self.stats_display,
            text="Chargez les données pour voir les statistiques détaillées",
            font=UIConfig.FONT_SMALL,
            fg=COLORS['INFO'],
            bg=COLORS['CARD']
        )
        initial_stats_label.pack(anchor=tk.W, pady=8)  # Reduced padding

    def _create_export_section(self, parent: tk.Widget):
        """Create the export statistics section with integrated period filtering."""
        # Export card - compact and top-aligned
        export_card = create_card_frame(parent)
        export_card.pack(anchor=tk.N, fill=tk.X, padx=0, pady=(0, 5))  # Compact, top-aligned

        # Section header
        header_frame = create_section_header(export_card, "📤", "Export Stat & Période")
        header_frame.pack(fill=tk.X, padx=15, pady=(15, 10))

        # Content frame
        export_content = tk.Frame(export_card, bg=COLORS['CARD'])
        export_content.pack(fill=tk.X, padx=15, pady=(0, 15))

        # Period filtering controls (integrated from date range section)
        self._create_period_filtering_in_export(export_content)

        # Separator
        separator = tk.Frame(export_content, height=1, bg=COLORS['BORDER'])
        separator.pack(fill=tk.X, pady=(15, 10))

        # Export buttons
        self._create_export_buttons(export_content)

    def _create_period_filtering_in_export(self, parent: tk.Widget):
        """Create period filtering controls integrated within the export section."""
        try:
            # Period filtering subsection
            period_frame = tk.Frame(parent, bg=COLORS['CARD'])
            period_frame.pack(fill=tk.X, pady=(0, 10))

            # Period subsection header
            period_header = tk.Label(
                period_frame,
                text="📅 Filtrage par période",
                font=("Segoe UI", 9, "bold"),
                fg=COLORS['PRIMARY'],
                bg=COLORS['CARD']
            )
            period_header.pack(anchor=tk.W, pady=(0, 8))

            # Date selection frame
            dates_frame = tk.Frame(period_frame, bg=COLORS['CARD'])
            dates_frame.pack(fill=tk.X, pady=(0, 8))

            # Date from section
            from_frame = tk.Frame(dates_frame, bg=COLORS['CARD'])
            from_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))

            from_label = tk.Label(
                from_frame,
                text="Date de début:",
                font=("Segoe UI", 8, "bold"),
                fg=COLORS['PRIMARY'],
                bg=COLORS['CARD']
            )
            from_label.pack(anchor=tk.W, pady=(0, 2))

            from_input_frame = tk.Frame(from_frame, bg=COLORS['CARD'])
            from_input_frame.pack(fill=tk.X)

            self.date_from_entry = tk.Entry(
                from_input_frame,
                textvariable=self.date_from_var,
                font=("Segoe UI", 8),
                width=12,
                state='readonly'
            )
            self.date_from_entry.pack(side=tk.LEFT, padx=(0, 5))

            from_button = tk.Button(
                from_input_frame,
                text="📅",
                font=("Segoe UI", 8),
                command=self._show_date_from_picker,
                bg=COLORS['ACCENT'],
                fg=COLORS['PRIMARY'],
                relief='flat',
                padx=8,
                pady=2,
                cursor='hand2'
            )
            from_button.pack(side=tk.LEFT)

            # Date to section
            to_frame = tk.Frame(dates_frame, bg=COLORS['CARD'])
            to_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)

            to_label = tk.Label(
                to_frame,
                text="Date de fin:",
                font=("Segoe UI", 8, "bold"),
                fg=COLORS['PRIMARY'],
                bg=COLORS['CARD']
            )
            to_label.pack(anchor=tk.W, pady=(0, 2))

            to_input_frame = tk.Frame(to_frame, bg=COLORS['CARD'])
            to_input_frame.pack(fill=tk.X)

            self.date_to_entry = tk.Entry(
                to_input_frame,
                textvariable=self.date_to_var,
                font=("Segoe UI", 8),
                width=12,
                state='readonly'
            )
            self.date_to_entry.pack(side=tk.LEFT, padx=(0, 5))

            to_button = tk.Button(
                to_input_frame,
                text="📅",
                font=("Segoe UI", 8),
                command=self._show_date_to_picker,
                bg=COLORS['ACCENT'],
                fg=COLORS['PRIMARY'],
                relief='flat',
                padx=8,
                pady=2,
                cursor='hand2'
            )
            to_button.pack(side=tk.LEFT)

            # Action buttons frame
            action_frame = tk.Frame(period_frame, bg=COLORS['CARD'])
            action_frame.pack(fill=tk.X, pady=(8, 0))

            # Generate filtered stats button
            self.generate_stats_button = tk.Button(
                action_frame,
                text="📊 Générer et ouvrir index",
                font=("Segoe UI", 8, "bold"),
                bg=COLORS['PRIMARY'],
                fg='white',
                relief='flat',
                padx=12,
                pady=6,
                cursor='hand2',
                command=self._generate_filtered_statistics,
                state=tk.DISABLED
            )
            self.generate_stats_button.pack(side=tk.LEFT, padx=(0, 5))

            # Clear dates button
            clear_button = tk.Button(
                action_frame,
                text="🗑️ Effacer",
                font=("Segoe UI", 8),
                bg=COLORS['SECONDARY'],
                fg='white',
                relief='flat',
                padx=8,
                pady=6,
                cursor='hand2',
                command=self._clear_date_range
            )
            clear_button.pack(side=tk.LEFT)

            # Status label
            self.date_range_status = tk.Label(
                period_frame,
                text="Sélectionnez une période pour générer et ouvrir l'index stats",
                font=UIConfig.FONT_SMALL,
                fg=COLORS['INFO'],
                bg=COLORS['CARD'],
                wraplength=300,
                justify=tk.LEFT
            )
            self.date_range_status.pack(anchor=tk.W, pady=(8, 0))

            self.logger.info("Period filtering controls integrated into export section")

        except Exception as e:
            self.logger.error(f"Error creating period filtering in export section: {e}")

    def _create_export_buttons(self, parent: tk.Widget):
        """Create export filters and buttons."""
        try:
            # Initialize export components dictionaries early
            self.export_buttons = {}
            self.export_filters = {}

            # Filters section
            filters_frame = tk.Frame(parent, bg=COLORS['CARD'])
            filters_frame.pack(fill=tk.X, pady=(0, 10))

            # Collaborator filter
            collab_label = tk.Label(
                filters_frame,
                text="Collaborateur:",
                font=("Segoe UI", 8, "bold"),
                fg=COLORS['PRIMARY'],
                bg=COLORS['CARD']
            )
            collab_label.pack(anchor=tk.W, pady=(0, 2))

            self.collab_var = tk.StringVar()
            self.collab_combo = ttk.Combobox(
                filters_frame,
                textvariable=self.collab_var,
                font=("Segoe UI", 8),
                state="readonly",
                width=18
            )
            self.collab_combo.pack(fill=tk.X, pady=(0, 8))

            # Month filter
            month_label = tk.Label(
                filters_frame,
                text="Mois:",
                font=("Segoe UI", 8, "bold"),
                fg=COLORS['PRIMARY'],
                bg=COLORS['CARD']
            )
            month_label.pack(anchor=tk.W, pady=(0, 2))

            self.month_var = tk.StringVar()
            self.month_combo = ttk.Combobox(
                filters_frame,
                textvariable=self.month_var,
                font=("Segoe UI", 8),
                state="readonly",
                width=18,
                values=["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                       "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
            )
            self.month_combo.pack(fill=tk.X, pady=(0, 8))

            # Year filter
            year_label = tk.Label(
                filters_frame,
                text="Année:",
                font=("Segoe UI", 8, "bold"),
                fg=COLORS['PRIMARY'],
                bg=COLORS['CARD']
            )
            year_label.pack(anchor=tk.W, pady=(0, 2))

            from datetime import datetime
            current_year = datetime.now().year
            self.year_var = tk.StringVar(value=str(current_year))
            self.year_combo = ttk.Combobox(
                filters_frame,
                textvariable=self.year_var,
                font=("Segoe UI", 8),
                state="readonly",
                width=18,
                values=[str(year) for year in range(current_year-2, current_year+2)]
            )
            self.year_combo.pack(fill=tk.X, pady=(0, 10))

            # Export buttons frame - First row
            buttons_frame1 = tk.Frame(parent, bg=COLORS['CARD'])
            buttons_frame1.pack(fill=tk.X, pady=(10, 5))

            # Export CTJ Combiné (PA + CM) button
            excel_btn = tk.Button(
                buttons_frame1,
                text="📊 Export CTJ (PA + CM) + Motifs",
                command=self._export_ctj_combined_to_excel,
                bg=COLORS['SUCCESS'],
                fg='white',
                font=UIConfig.FONT_SMALL,
                relief='flat',
                padx=10,
                pady=6,
                state=tk.DISABLED
            )
            excel_btn.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 4))

            # Export anomalies button
            anomalies_btn = tk.Button(
                buttons_frame1,
                text="⚠️ Export Anomalies",
                command=self._export_anomalies,
                bg=COLORS['WARNING'],
                fg='white',
                font=UIConfig.FONT_BUTTON,
                relief='flat',
                padx=10,
                pady=6,
                cursor='hand2',
                state=tk.DISABLED
            )
            anomalies_btn.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(4, 0))

            # Store buttons immediately after creation
            self.export_buttons.update({'excel': excel_btn, 'anomalies': anomalies_btn})
            self.export_filters.update({
                'collaborator': self.collab_combo,
                'month': self.month_combo,
                'year': self.year_combo
            })



            self.logger.info("Export buttons and filters created successfully")

        except Exception as e:
            self.logger.error(f"Error creating export buttons: {e}")
            # Ensure dictionaries exist even if creation fails
            if not hasattr(self, 'export_buttons'):
                self.export_buttons = {}
            if not hasattr(self, 'export_filters'):
                self.export_filters = {}





    def _load_global_data(self):
        """Load and analyze data from the global suivi file - optimized."""
        try:
            self.status_label.config(text="Chargement optimisé des données...")
            self.progress_var.set(10)

            # Force UI update to show progress immediately
            self.parent.update_idletasks()

            # Construct path to global file
            global_file_path = os.path.join(self.teams_folder_path, self.global_excel_filename)

            # Store the excel path for stats injection functionality
            self.excel_path = global_file_path

            if not os.path.exists(global_file_path):
                messagebox.showerror(
                    "Fichier non trouvé",
                    f"Le fichier global n'a pas été trouvé:\n{global_file_path}\n\n"
                    "Assurez-vous que le module 'Suivi Global Tickets' a été exécuté au moins une fois."
                )
                self._reset_loading_state()
                return

            # Check file access before proceeding
            self.status_label.config(text="Vérification de l'accès au fichier...")
            self.progress_var.set(20)

            access_result = check_file_access(global_file_path, 'r')
            if not access_result['accessible']:
                self.logger.warning(f"File access issue: {access_result['error_message']}")

                # Show user-friendly dialog and get retry decision
                if access_result['error_type'] in ['file_locked', 'permission_denied']:
                    # This is a file-in-use situation, show the custom dialog
                    retry = self._show_file_access_dialog(access_result, global_file_path)
                    if not retry:
                        self._reset_loading_state()
                        return

                    # User wants to retry, check again
                    retry_access = check_file_access(global_file_path, 'r')
                    if not retry_access['accessible']:
                        # Still not accessible, show warning and abort
                        messagebox.showwarning(
                            "Fichier toujours en cours d'utilisation",
                            f"{retry_access['user_message']}\n\nVeuillez fermer Excel et redémarrer le module pour recharger automatiquement."
                        )
                        self._reset_loading_state()
                        return
                else:
                    # Other types of errors, show error and abort
                    messagebox.showerror("Erreur d'accès au fichier", access_result['user_message'])
                    self._reset_loading_state()
                    return

            self.status_label.config(text="Lecture des données Excel...")
            self.progress_var.set(30)

            # Load Excel data using pandas
            pd = get_pandas()

            # Read all sheets
            excel_data = {}
            sheet_names = ['Suivi Tickets', 'Traitement CMS Adr', 'Traitement PA']

            for i, sheet_name in enumerate(sheet_names):
                try:
                    self.status_label.config(text=f"Lecture de la page '{sheet_name}'...")
                    self.progress_var.set(30 + (i * 20))

                    df = pd.read_excel(global_file_path, sheet_name=sheet_name, date_format=None)
                    excel_data[sheet_name] = df
                    self.logger.info(f"Loaded sheet '{sheet_name}' with {len(df)} rows")

                except Exception as e:
                    self.logger.warning(f"Could not load sheet '{sheet_name}': {e}")
                    excel_data[sheet_name] = pd.DataFrame()

            self.global_suivi_data = excel_data

            self.status_label.config(text="Analyse des statistiques...")
            self.progress_var.set(80)

            # Analyze the data
            self._analyze_team_statistics()

            self.status_label.config(text="Mise à jour de l'affichage...")
            self.progress_var.set(90)

            # Update UI
            self._update_overview_display()
            self._update_statistics_display()

            self.progress_var.set(100)
            self.status_label.config(text="✅ Données chargées automatiquement avec succès")
            self.refresh_button.config(state=tk.NORMAL)

            # Debug logging before enabling buttons
            self.logger.info("Data loaded successfully, enabling export buttons...")
            self.logger.debug(f"Has export_buttons: {hasattr(self, 'export_buttons')}")
            if hasattr(self, 'export_buttons'):
                self.logger.debug(f"export_buttons keys: {list(self.export_buttons.keys()) if self.export_buttons else 'None'}")

            self._enable_export_buttons()
            self._update_export_filters()

            # Update file status indicator
            self._update_file_status_indicator()

            # Enable date range functionality
            self._enable_date_range_functionality()

            # Calculate DMT automatically after loading data
            self._calculate_dmt_automatically()



            # Enable archive button now that data is loaded
            self._enable_archive_button()

            # Update scroll region after data is loaded and UI is updated
            self.parent.after(100, self._update_scroll_region)

            self.logger.info("Global data loaded and analyzed successfully")

        except Exception as e:
            self.logger.error(f"Error loading global data: {e}")
            messagebox.showerror("Erreur", f"Erreur lors du chargement des données:\n{e}")
            self._reset_loading_state()

    def _enable_date_range_functionality(self):
        """Enable date range functionality after data is loaded."""
        try:
            # Convert loaded data to a format suitable for date filtering
            self.data = []

            if hasattr(self, 'global_suivi_data') and self.global_suivi_data:
                # Process each sheet
                for sheet_name, df in self.global_suivi_data.items():
                    if df is not None and not df.empty:
                        # Convert DataFrame rows to dictionaries
                        for index, row in df.iterrows():
                            record = row.to_dict()
                            record['sheet_name'] = sheet_name
                            record['row_index'] = index
                            self.data.append(record)

            # Update date range status
            self._update_date_range_status()

            self.logger.info(f"Date range functionality enabled with {len(self.data)} records")

        except Exception as e:
            self.logger.error(f"Error enabling date range functionality: {e}")

    def _enable_archive_button(self):
        """Enable the archive button and update status."""
        try:
            self.logger.info("Attempting to enable archive button...")

            # Check if archive button exists
            if hasattr(self, 'archive_button'):
                if self.archive_button:
                    self.archive_button.config(state=tk.NORMAL)
                    self.logger.info("Archive button enabled successfully")
                else:
                    self.logger.warning("Archive button exists but is None")
            else:
                self.logger.warning("Archive button does not exist")

            # Check if archive status label exists
            if hasattr(self, 'archive_status_label'):
                if self.archive_status_label:
                    # Count treated communes for current month
                    self.logger.info("Counting treated communes...")
                    treated_count = self._count_treated_communes_current_month()
                    self.logger.info(f"Found {treated_count} treated communes this month")

                    if treated_count > 0:
                        self.archive_status_label.config(
                            text=f"{treated_count} communes traitées ce mois - Prêt pour archivage",
                            fg=COLORS['SUCCESS']
                        )
                        self.logger.info("Archive status updated with success message")
                    else:
                        self.archive_status_label.config(
                            text="Aucune commune traitée ce mois",
                            fg=COLORS['WARNING']
                        )
                        self.logger.info("Archive status updated with warning message")
                else:
                    self.logger.warning("Archive status label exists but is None")
            else:
                self.logger.warning("Archive status label does not exist")

        except Exception as e:
            self.logger.error(f"Error enabling archive button: {e}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")

    def _calculate_dmt_automatically(self):
        """Calculate DMT automatically after loading data (silent version)."""
        try:
            if not self.global_suivi_data:
                return

            pd = get_pandas()

            # Calculate DMT for all collaborators and overall
            self.dmt_data = {}

            # Get the main data sheet (usually 'Suivi Tickets')
            main_sheet = None
            if 'Suivi Tickets' in self.global_suivi_data:
                main_sheet = self.global_suivi_data['Suivi Tickets']
            elif len(self.global_suivi_data) > 0:
                # Use the first sheet if 'Suivi Tickets' not found
                main_sheet = list(self.global_suivi_data.values())[0]

            if main_sheet is None or main_sheet.empty:
                self.logger.warning("No valid data sheet found for DMT calculation")
                return

            # Get unique collaborators from the main sheet
            if 'Collaborateur' not in main_sheet.columns:
                self.logger.warning("'Collaborateur' column not found in data")
                return

            collaborators = main_sheet['Collaborateur'].dropna().unique()

            # Calculate overall DMT using the main sheet
            overall_dmt_pa = self._calculate_dmt_pa(main_sheet)
            overall_dmt_cm = self._calculate_dmt_cm(main_sheet)

            self.dmt_data['Toute l\'équipe'] = {
                'dmt_pa': overall_dmt_pa,
                'dmt_cm': overall_dmt_cm
            }

            # Calculate DMT for each collaborator using individual calculation methods
            for collab in collaborators:
                if pd.notna(collab) and collab.strip():
                    # Calculate individual DMT PA and CM for this collaborator
                    dmt_pa = self._calculate_individual_dmt_pa(collab)
                    dmt_cm = self._calculate_individual_dmt_cm(collab)

                    self.dmt_data[collab] = {
                        'dmt_pa': dmt_pa,
                        'dmt_cm': dmt_cm
                    }

            # Update the detailed statistics display to show DMT values
            self._update_statistics_display()

            self.logger.info("DMT calculation completed automatically")

        except Exception as e:
            self.logger.error(f"Error calculating DMT automatically: {e}")
            # Don't show error dialog for automatic calculation

    def _update_statistics_display(self):
        """Update the statistics display to show current DMT values."""
        try:
            # Clear current statistics display
            for widget in self.stats_display.winfo_children():
                widget.destroy()

            # Recreate statistics display with updated DMT data
            if self.collaborator_stats:
                for collaborator, stats in self.collaborator_stats.items():
                    self._create_collaborator_card(collaborator, stats)

            self.logger.info("Statistics display updated with DMT values")

        except Exception as e:
            self.logger.error(f"Error updating statistics display: {e}")

    def _calculate_dmt_pa(self, data):
        """Calculate DMT PA par collaborateur - Feuille 3, Colonne F (collaborateur), Colonne H (durée)."""
        try:
            pd = get_pandas()

            # Vérifier si nous avons les données de la feuille 3 (Traitement PA)
            if 'Traitement PA' not in self.global_suivi_data:
                self.logger.warning("Traitement PA sheet data not found for DMT calculation")
                return 0

            df_pa = self.global_suivi_data['Traitement PA']
            if df_pa.empty:
                self.logger.warning("Traitement PA sheet is empty")
                return 0

            columns = list(df_pa.columns)
            if len(columns) <= 7:  # Need at least 8 columns (index 0-7)
                self.logger.warning(f"Not enough columns in PA sheet for DMT calculation. Found {len(columns)} columns")
                return 0

            # Colonne F (index 5) = Collaborateur, Colonne H (index 7) = Durée
            collaborateur_col = columns[5]  # Column F (index 5)
            duree_col = columns[7]          # Column H (index 7)

            self.logger.debug(f"DMT PA: Using collaborateur column '{collaborateur_col}' and durée column '{duree_col}'")

            # Calculer DMT par collaborateur en excluant les valeurs 0 et vides
            collaborateur_dmts = {}

            for index, row in df_pa.iterrows():
                collaborateur = row.get(collaborateur_col, None)
                duree_value = row.get(duree_col, None)

                # Vérifier que le collaborateur et la durée sont valides
                if (pd.notna(collaborateur) and pd.notna(duree_value) and
                    str(collaborateur).strip() != '' and str(duree_value).strip() != ''):

                    try:
                        duree_numeric = float(duree_value)
                        # Exclure les valeurs 0 et vides comme spécifié
                        if duree_numeric > 0:
                            collaborateur_name = str(collaborateur).strip()
                            if collaborateur_name not in collaborateur_dmts:
                                collaborateur_dmts[collaborateur_name] = []
                            collaborateur_dmts[collaborateur_name].append(duree_numeric)
                    except (ValueError, TypeError):
                        continue

            # Calculer la moyenne globale de tous les collaborateurs
            if collaborateur_dmts:
                all_durees = []
                for collab, durees in collaborateur_dmts.items():
                    if durees:  # S'assurer que la liste n'est pas vide
                        collab_moyenne = sum(durees) / len(durees)
                        all_durees.append(collab_moyenne)
                        self.logger.debug(f"DMT PA - {collab}: {collab_moyenne:.2f} (basé sur {len(durees)} valeurs non-nulles)")

                if all_durees:
                    dmt_global = sum(all_durees) / len(all_durees)
                    self.logger.info(f"DMT PA calculé: {dmt_global:.2f} (moyenne de {len(all_durees)} collaborateurs)")
                    return round(dmt_global, 2)
                else:
                    self.logger.warning("No valid collaborateur averages calculated for DMT PA")
                    return 0
            else:
                self.logger.warning("No valid duration data found for DMT PA calculation (all values were 0 or invalid)")
                return 0

        except Exception as e:
            self.logger.error(f"Error calculating DMT PA: {e}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            return 0

    def _calculate_dmt_cm(self, data):
        """Calculate DMT CM par collaborateur - Feuille 2, Colonne E (collaborateur), Colonne J (durée)."""
        try:
            pd = get_pandas()

            # Vérifier si nous avons les données de la feuille 2 (CM)
            if 'CM' not in self.global_suivi_data:
                self.logger.warning("CM sheet data not found for DMT calculation")
                return 0

            df_cm = self.global_suivi_data['CM']
            if df_cm.empty:
                self.logger.warning("CM sheet is empty")
                return 0

            columns = list(df_cm.columns)
            if len(columns) <= 9:  # Need at least 10 columns (index 0-9)
                self.logger.warning(f"Not enough columns in CM sheet for DMT calculation. Found {len(columns)} columns")
                return 0

            # Colonne E (index 4) = Collaborateur, Colonne J (index 9) = Durée
            collaborateur_col = columns[4]  # Column E (index 4) - CORRIGÉ
            duree_col = columns[9]          # Column J (index 9)

            self.logger.debug(f"DMT CM: Using collaborateur column '{collaborateur_col}' and durée column '{duree_col}'")

            # Calculer DMT par collaborateur en excluant les valeurs 0 et vides
            collaborateur_dmts = {}

            for index, row in df_cm.iterrows():
                collaborateur = row.get(collaborateur_col, None)
                duree_value = row.get(duree_col, None)

                # Vérifier que le collaborateur et la durée sont valides
                if (pd.notna(collaborateur) and pd.notna(duree_value) and
                    str(collaborateur).strip() != '' and str(duree_value).strip() != ''):

                    try:
                        duree_numeric = float(duree_value)
                        # Exclure les valeurs 0 et vides comme spécifié
                        if duree_numeric > 0:
                            collaborateur_name = str(collaborateur).strip()
                            if collaborateur_name not in collaborateur_dmts:
                                collaborateur_dmts[collaborateur_name] = []
                            collaborateur_dmts[collaborateur_name].append(duree_numeric)
                    except (ValueError, TypeError):
                        continue

            # Calculer la moyenne globale de tous les collaborateurs
            if collaborateur_dmts:
                all_durees = []
                for collab, durees in collaborateur_dmts.items():
                    if durees:  # S'assurer que la liste n'est pas vide
                        collab_moyenne = sum(durees) / len(durees)
                        all_durees.append(collab_moyenne)
                        self.logger.debug(f"DMT CM - {collab}: {collab_moyenne:.2f} (basé sur {len(durees)} valeurs non-nulles)")

                if all_durees:
                    dmt_global = sum(all_durees) / len(all_durees)
                    self.logger.info(f"DMT CM calculé: {dmt_global:.2f} (moyenne de {len(all_durees)} collaborateurs)")
                    return round(dmt_global, 2)
                else:
                    self.logger.warning("No valid collaborateur averages calculated for DMT CM")
                    return 0
            else:
                self.logger.warning("No valid duration data found for DMT CM calculation (all values were 0 or invalid)")
                return 0

        except Exception as e:
            self.logger.error(f"Error calculating DMT CM: {e}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            return 0

    def _calculate_individual_dmt_pa(self, collaborateur):
        """Calculate DMT PA for a specific collaborator - Feuille 3, Colonne F (collaborateur), Colonne H (durée)."""
        try:
            pd = get_pandas()

            # Vérifier si nous avons les données de la feuille 3 (Traitement PA)
            if 'Traitement PA' not in self.global_suivi_data:
                self.logger.debug(f"Traitement PA sheet data not found for DMT PA calculation for {collaborateur}")
                return 0

            df_pa = self.global_suivi_data['Traitement PA']
            if df_pa.empty:
                self.logger.debug(f"Traitement PA sheet is empty for {collaborateur}")
                return 0

            columns = list(df_pa.columns)
            if len(columns) <= 7:  # Need at least 8 columns (index 0-7)
                self.logger.debug(f"Not enough columns in PA sheet for DMT calculation for {collaborateur}. Found {len(columns)} columns")
                return 0

            # Colonne F (index 5) = Collaborateur, Colonne H (index 7) = Durée
            collaborateur_col = columns[5]  # Column F (index 5)
            duree_col = columns[7]          # Column H (index 7)

            # Filtrer les données pour ce collaborateur spécifique
            collaborateur_durees = []

            for index, row in df_pa.iterrows():
                row_collaborateur = row.get(collaborateur_col, None)
                duree_value = row.get(duree_col, None)

                # Vérifier que c'est le bon collaborateur et que la durée est valide
                if (pd.notna(row_collaborateur) and pd.notna(duree_value) and
                    str(row_collaborateur).strip() == str(collaborateur).strip() and
                    str(duree_value).strip() != ''):

                    try:
                        duree_numeric = float(duree_value)
                        # Exclure les valeurs 0 et vides comme spécifié
                        if duree_numeric > 0:
                            collaborateur_durees.append(duree_numeric)
                    except (ValueError, TypeError):
                        continue

            # Calculer la moyenne individuelle pour ce collaborateur
            if collaborateur_durees:
                dmt_individuel = sum(collaborateur_durees) / len(collaborateur_durees)
                self.logger.debug(f"DMT PA individuel - {collaborateur}: {dmt_individuel:.2f} (basé sur {len(collaborateur_durees)} valeurs non-nulles)")
                return round(dmt_individuel, 2)
            else:
                self.logger.debug(f"No valid duration data found for DMT PA calculation for {collaborateur}")
                return 0

        except Exception as e:
            self.logger.error(f"Error calculating individual DMT PA for {collaborateur}: {e}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            return 0

    def _calculate_individual_dmt_cm(self, collaborateur):
        """Calculate DMT CM for a specific collaborator - Feuille 2, Colonne E (collaborateur), Colonne J (durée)."""
        try:
            pd = get_pandas()

            # Chercher la feuille CM avec différents noms possibles
            df_cm = None
            cm_sheet_names = ['CM', 'Traitement CMS Adr', 'Traitement CM', 'Sheet2', 'Feuille2']

            for sheet_name in cm_sheet_names:
                if sheet_name in self.global_suivi_data:
                    df_cm = self.global_suivi_data[sheet_name]
                    self.logger.debug(f"Found CM sheet with name: {sheet_name}")
                    break

            if df_cm is None:
                available_sheets = list(self.global_suivi_data.keys())
                self.logger.warning(f"CM sheet not found for {collaborateur}. Available sheets: {available_sheets}")
                return 0
            if df_cm.empty:
                self.logger.debug(f"CM sheet is empty for {collaborateur}")
                return 0

            columns = list(df_cm.columns)
            if len(columns) <= 9:  # Need at least 10 columns (index 0-9)
                self.logger.debug(f"Not enough columns in CM sheet for DMT calculation for {collaborateur}. Found {len(columns)} columns")
                return 0

            # Colonne E (index 4) = Collaborateur, Colonne J (index 9) = Durée
            collaborateur_col = columns[4]  # Column E (index 4) - CORRIGÉ
            duree_col = columns[9]          # Column J (index 9)

            self.logger.debug(f"DMT CM individuel pour {collaborateur}: colonnes '{collaborateur_col}' et '{duree_col}'")

            # Filtrer les données pour ce collaborateur spécifique
            collaborateur_durees = []
            total_rows_checked = 0
            matching_collaborateur_rows = 0
            valid_duree_rows = 0

            for index, row in df_cm.iterrows():
                total_rows_checked += 1
                row_collaborateur = row.get(collaborateur_col, None)
                duree_value = row.get(duree_col, None)

                # Debug: log first few rows
                if total_rows_checked <= 3:
                    self.logger.debug(f"Row {total_rows_checked}: collaborateur='{row_collaborateur}', duree='{duree_value}'")

                # Vérifier que c'est le bon collaborateur
                if (pd.notna(row_collaborateur) and
                    str(row_collaborateur).strip() == str(collaborateur).strip()):
                    matching_collaborateur_rows += 1

                    # Vérifier que la durée est valide
                    if (pd.notna(duree_value) and str(duree_value).strip() != ''):
                        try:
                            duree_numeric = float(duree_value)
                            # Exclure les valeurs 0 et vides comme spécifié
                            if duree_numeric > 0:
                                collaborateur_durees.append(duree_numeric)
                                valid_duree_rows += 1
                        except (ValueError, TypeError):
                            self.logger.debug(f"Invalid duree value for {collaborateur}: '{duree_value}'")
                            continue

            self.logger.debug(f"DMT CM {collaborateur}: {total_rows_checked} rows checked, {matching_collaborateur_rows} matching collaborateur, {valid_duree_rows} valid durees")

            # Calculer la moyenne individuelle pour ce collaborateur
            if collaborateur_durees:
                dmt_individuel = sum(collaborateur_durees) / len(collaborateur_durees)
                self.logger.debug(f"DMT CM individuel - {collaborateur}: {dmt_individuel:.2f} (basé sur {len(collaborateur_durees)} valeurs non-nulles)")
                return round(dmt_individuel, 2)
            else:
                self.logger.debug(f"No valid duration data found for DMT CM calculation for {collaborateur}")
                return 0

        except Exception as e:
            self.logger.error(f"Error calculating individual DMT CM for {collaborateur}: {e}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            return 0

    def _reset_loading_state(self):
        """Reset the loading state after error or completion."""
        self.progress_var.set(0)
        self.status_label.config(text="Erreur de chargement - Réessayez en redémarrant le module")

    def _show_file_access_dialog(self, access_result: dict, file_path: str) -> bool:
        """
        Show a user-friendly dialog for file access issues.

        Args:
            access_result: Result from check_file_access()
            file_path: Path to the file

        Returns:
            True if user wants to retry, False otherwise
        """
        filename = os.path.basename(file_path)

        # Create custom dialog
        dialog = tk.Toplevel(self.parent)
        dialog.title("Fichier en cours d'utilisation")
        dialog.geometry("500x400")
        dialog.configure(bg=COLORS['BG'])
        dialog.resizable(False, False)

        # Center the dialog
        dialog.transient(self.parent.winfo_toplevel())
        dialog.grab_set()

        # Main frame
        main_frame = tk.Frame(dialog, bg=COLORS['BG'], padx=20, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Icon and title
        title_frame = tk.Frame(main_frame, bg=COLORS['BG'])
        title_frame.pack(fill=tk.X, pady=(0, 15))

        icon_label = tk.Label(
            title_frame,
            text="🔒",
            font=("Segoe UI", 24),
            bg=COLORS['BG'],
            fg=COLORS['WARNING']
        )
        icon_label.pack(side=tk.LEFT, padx=(0, 10))

        title_label = tk.Label(
            title_frame,
            text="Fichier Excel en cours d'utilisation",
            font=UIConfig.FONT_HEADER,
            bg=COLORS['BG'],
            fg=COLORS['PRIMARY']
        )
        title_label.pack(side=tk.LEFT)

        # Message
        message_text = access_result.get('user_message', f"Le fichier '{filename}' est actuellement utilisé.")
        message_label = tk.Label(
            main_frame,
            text=message_text,
            font=UIConfig.FONT_SUBTITLE,
            bg=COLORS['BG'],
            fg=COLORS['TEXT_SECONDARY'],
            wraplength=450,
            justify=tk.LEFT
        )
        message_label.pack(fill=tk.X, pady=(0, 15))

        # Suggestions
        suggestions = access_result.get('suggestions', [])
        if suggestions:
            suggestions_label = tk.Label(
                main_frame,
                text="💡 Solutions recommandées:",
                font=UIConfig.FONT_SUBTITLE,
                bg=COLORS['BG'],
                fg=COLORS['PRIMARY']
            )
            suggestions_label.pack(anchor=tk.W, pady=(0, 5))

            for i, suggestion in enumerate(suggestions, 1):
                suggestion_label = tk.Label(
                    main_frame,
                    text=f"  {i}. {suggestion}",
                    font=UIConfig.FONT_SMALL,
                    bg=COLORS['BG'],
                    fg=COLORS['TEXT_SECONDARY'],
                    wraplength=450,
                    justify=tk.LEFT
                )
                suggestion_label.pack(anchor=tk.W, pady=1)

        # Additional info
        info_frame = tk.Frame(main_frame, bg=COLORS['LIGHT'], relief=tk.RAISED, bd=1)
        info_frame.pack(fill=tk.X, pady=(15, 0))

        info_label = tk.Label(
            info_frame,
            text="ℹ️ Astuce: Vous pouvez consulter le fichier Excel pendant que l'application fonctionne,\nmais fermez-le avant d'ouvrir le module pour permettre le chargement automatique.",
            font=UIConfig.FONT_SMALL,
            bg=COLORS['LIGHT'],
            fg=COLORS['INFO'],
            justify=tk.CENTER,
            padx=10,
            pady=8
        )
        info_label.pack()

        # Buttons
        button_frame = tk.Frame(main_frame, bg=COLORS['BG'])
        button_frame.pack(fill=tk.X, pady=(20, 0))

        result = {'retry': False}

        def on_retry():
            result['retry'] = True
            dialog.destroy()

        def on_cancel():
            result['retry'] = False
            dialog.destroy()

        # Retry button
        retry_btn = tk.Button(
            button_frame,
            text="🔄 Réessayer",
            command=on_retry,
            bg=COLORS['PRIMARY'],
            fg='white',
            font=UIConfig.FONT_BUTTON,
            relief='flat',
            padx=20,
            pady=8
        )
        retry_btn.pack(side=tk.RIGHT, padx=(10, 0))

        # Cancel button
        cancel_btn = tk.Button(
            button_frame,
            text="❌ Annuler",
            command=on_cancel,
            bg=COLORS['BORDER'],
            fg=COLORS['TEXT_SECONDARY'],
            font=UIConfig.FONT_BUTTON,
            relief='flat',
            padx=20,
            pady=8
        )
        cancel_btn.pack(side=tk.RIGHT)

        # Wait for dialog to close
        dialog.wait_window()

        return result['retry']

    def _check_global_file_status(self):
        """Check if the global Excel file is accessible and update UI accordingly."""
        try:
            file_path = os.path.join(self.teams_folder_path, self.global_excel_filename)

            if not os.path.exists(file_path):
                return "not_exists"

            if is_excel_file_open(file_path):
                return "in_use"

            access_result = check_file_access(file_path, 'r')  # Only need read access for stats
            if access_result['accessible']:
                return "available"
            else:
                return "locked"

        except Exception as e:
            self.logger.error(f"Error checking global file status: {e}")
            return "unknown"

    def _update_file_status_indicator(self):
        """Update the file status indicator in the UI."""
        try:
            status = self._check_global_file_status()

            if hasattr(self, 'file_status_label'):
                status_config = {
                    'not_exists': {'text': '📄 Fichier global non trouvé', 'color': COLORS['WARNING']},
                    'available': {'text': '✅ Fichier disponible', 'color': COLORS['SUCCESS']},
                    'in_use': {'text': '🔒 Fichier ouvert dans Excel', 'color': COLORS['WARNING']},
                    'locked': {'text': '⚠️ Fichier verrouillé', 'color': COLORS['DANGER']},
                    'unknown': {'text': '❓ Statut inconnu', 'color': COLORS['TEXT_MUTED']}
                }

                config = status_config.get(status, status_config['unknown'])
                self.file_status_label.config(text=config['text'], fg=config['color'])

        except Exception as e:
            self.logger.error(f"Error updating file status indicator: {e}")

    def _analyze_team_statistics(self):
        """Analyze team statistics from the loaded data."""
        try:
            if not self.global_suivi_data:
                return

            pd = get_pandas()
            from datetime import datetime

            # Initialize statistics with KPI focus
            self.team_statistics = {
                'total_tickets': 0,
                'total_cms_records': 0,
                'total_pa_records': 0,
                'total_duration_cms': 0,
                'total_duration_pa': 0,
                'total_duration_finale': 0,
                'collaborators': set(),
                'communes': set(),
                'communes_traitees_mois_courant': 0,
                'communes_autres_statuts': {},
                # New KPI metrics
                'team_dmt': 0,  # Team-wide DMT (Average Treatment Duration)
                'team_ctj_today': 0,  # Team-wide CTJ for today (Daily Treatment Capacity)
                'total_elements_today': 0  # Total elements processed today by all collaborators
            }

            self.collaborator_stats = {}
            self.ticket_status_breakdown = {}
            self.overall_averages = {}

            # Analyze Page 1 (Suivi Tickets) - Main data source
            if 'Suivi Tickets' in self.global_suivi_data:
                df_tickets = self.global_suivi_data['Suivi Tickets']

                if not df_tickets.empty:
                    self.team_statistics['total_tickets'] = len(df_tickets)

                    # Extract collaborators and communes
                    if 'Collaborateur' in df_tickets.columns:
                        collaborators = df_tickets['Collaborateur'].dropna().unique()
                        self.team_statistics['collaborators'].update(collaborators)

                    if 'Nom Commune' in df_tickets.columns:
                        communes = df_tickets['Nom Commune'].dropna().unique()
                        self.team_statistics['communes'].update(communes)

                    # Analyze communes by status and date
                    self._analyze_commune_status_by_date(df_tickets, pd, datetime)

                    # Analyze ticket status breakdown (keep existing logic for compatibility)
                    if 'STATUT Ticket' in df_tickets.columns:
                        status_counts = df_tickets['STATUT Ticket'].value_counts()
                        self.ticket_status_breakdown = status_counts.to_dict()
                    elif 'Etat' in df_tickets.columns:
                        status_counts = df_tickets['Etat'].value_counts()
                        self.ticket_status_breakdown = status_counts.to_dict()

                    # Calculate total "Durée Finale" if available
                    if 'Durée Finale' in df_tickets.columns:
                        finale_duration = self._calculate_duration_sum(df_tickets['Durée Finale'])
                        self.team_statistics['total_duration_finale'] = finale_duration

                    # Analyze by collaborator
                    if 'Collaborateur' in df_tickets.columns:
                        for collaborator in df_tickets['Collaborateur'].dropna().unique():
                            collab_data = df_tickets[df_tickets['Collaborateur'] == collaborator]
                            collab_communes = set(collab_data['Nom Commune'].dropna().unique()) if 'Nom Commune' in collab_data.columns else set()

                            # Calculate DMT (Durée Moyenne de Traitement) for this collaborator
                            # Only for communes with status "Traité"
                            dmt_collaborator = self._calculate_dmt_for_treated_communes(collab_data, pd)

                            # Calculate CTJ (Capacité de Traitement Journalier) for today
                            # From page 3 (Traitement PA) column G
                            ctj_today = self._calculate_ctj_from_page3(collaborator, pd, datetime)

                            # Calculate CTJ CM (Capacité de Traitement Journalier CM) for today
                            # From page 2 (Traitement CMS Adr) column F
                            ctj_cm_today = self._calculate_ctj_cm_from_page2(collaborator, pd, datetime)

                            self.collaborator_stats[collaborator] = {
                                'tickets_count': len(collab_data),
                                'cms_records': 0,
                                'pa_records': 0,
                                'cms_duration': 0,
                                'pa_duration': 0,
                                'avg_cms_duration': 0,
                                'avg_pa_duration': 0,
                                'avg_finale_duration': dmt_collaborator,  # This is now the DMT
                                'communes': collab_communes,
                                'commune_count': len(collab_communes),
                                # New KPI metrics
                                'dmt': dmt_collaborator,  # DMT - Average Treatment Duration
                                'ctj_today': ctj_today,   # CTJ PA - Daily Treatment Capacity for today (PA)
                                'ctj_cm_today': ctj_cm_today,  # CTJ CM - Daily Treatment Capacity for today (CM)
                                'elements_today': ctj_today  # Elements processed today (same as CTJ for now)
                            }

            # Analyze Page 2 (Traitement CMS Adr) - Duration data
            if 'Traitement CMS Adr' in self.global_suivi_data:
                df_cms = self.global_suivi_data['Traitement CMS Adr']

                if not df_cms.empty:
                    self.team_statistics['total_cms_records'] = len(df_cms)

                    # Calculate total CMS duration
                    if 'Durée' in df_cms.columns:
                        duration_sum = self._calculate_duration_sum(df_cms['Durée'])
                        self.team_statistics['total_duration_cms'] = duration_sum

                    # Add CMS data to collaborator stats
                    if 'Collaborateur' in df_cms.columns:
                        for collaborator in df_cms['Collaborateur'].dropna().unique():
                            if collaborator in self.collaborator_stats:
                                collab_cms = df_cms[df_cms['Collaborateur'] == collaborator]
                                self.collaborator_stats[collaborator]['cms_records'] = len(collab_cms)

                                if 'Durée' in collab_cms.columns:
                                    duration_sum = self._calculate_duration_sum(collab_cms['Durée'])
                                    self.collaborator_stats[collaborator]['cms_duration'] = duration_sum

                                    # Calculate average CMS duration per commune
                                    commune_count = self.collaborator_stats[collaborator]['commune_count']
                                    if commune_count > 0:
                                        self.collaborator_stats[collaborator]['avg_cms_duration'] = duration_sum / commune_count

            # Analyze Page 3 (Traitement PA) - Duration data
            if 'Traitement PA' in self.global_suivi_data:
                df_pa = self.global_suivi_data['Traitement PA']

                if not df_pa.empty:
                    self.team_statistics['total_pa_records'] = len(df_pa)

                    # Calculate total PA duration
                    if 'Durée' in df_pa.columns:
                        duration_sum = self._calculate_duration_sum(df_pa['Durée'])
                        self.team_statistics['total_duration_pa'] = duration_sum

                    # Add PA data to collaborator stats
                    if 'Collaborateur' in df_pa.columns:
                        for collaborator in df_pa['Collaborateur'].dropna().unique():
                            if collaborator in self.collaborator_stats:
                                collab_pa = df_pa[df_pa['Collaborateur'] == collaborator]
                                self.collaborator_stats[collaborator]['pa_records'] = len(collab_pa)

                                if 'Durée' in collab_pa.columns:
                                    duration_sum = self._calculate_duration_sum(collab_pa['Durée'])
                                    self.collaborator_stats[collaborator]['pa_duration'] = duration_sum

                                    # Calculate average PA duration per commune
                                    commune_count = self.collaborator_stats[collaborator]['commune_count']
                                    if commune_count > 0:
                                        self.collaborator_stats[collaborator]['avg_pa_duration'] = duration_sum / commune_count

            # Calculate overall team averages and KPIs
            self._calculate_overall_averages()
            self._calculate_team_kpis()

            self.logger.info(f"Analyzed statistics for {len(self.collaborator_stats)} collaborators")

        except Exception as e:
            self.logger.error(f"Error analyzing team statistics: {e}")
            raise

    def _calculate_duration_sum(self, duration_series):
        """Calculate sum of duration values, handling different formats."""
        try:
            total_minutes = 0

            for duration in duration_series.dropna():
                if isinstance(duration, (int, float)):
                    # Assume it's already in minutes
                    total_minutes += duration
                elif isinstance(duration, str):
                    # Try to parse time format (HH:MM or MM)
                    duration = duration.strip()
                    if ':' in duration:
                        parts = duration.split(':')
                        if len(parts) == 2:
                            hours = int(parts[0])
                            minutes = int(parts[1])
                            total_minutes += (hours * 60) + minutes
                    else:
                        # Assume it's minutes
                        try:
                            total_minutes += float(duration)
                        except ValueError:
                            continue

            return total_minutes

        except Exception as e:
            self.logger.error(f"Error calculating duration sum: {e}")
            return 0

    def _calculate_dmt_for_treated_communes(self, collab_data, pd):
        """Calculate DMT only for communes with status 'Traité'."""
        try:
            # Find the correct status column (handle variations)
            status_column = None
            for col in collab_data.columns:
                if 'etat ticket pa' in col.lower().strip():
                    status_column = col
                    break

            if status_column is None:
                # Fallback to other possible status columns
                for col in collab_data.columns:
                    if any(keyword in col.lower() for keyword in ['etat', 'statut', 'status']):
                        status_column = col
                        break

            if status_column is None or 'Durée Finale' not in collab_data.columns:
                self.logger.warning(f"Status column or Durée Finale not found. Available columns: {list(collab_data.columns)}")
                return 0

            # Filter data for communes with status "Traité"
            treated_data = collab_data[collab_data[status_column].astype(str).str.strip() == 'Traité']

            if treated_data.empty:
                return 0

            # Calculate total duration for treated communes
            total_duration = self._calculate_duration_sum(treated_data['Durée Finale'])
            treated_count = len(treated_data)

            # Return average duration per treated commune
            return total_duration / treated_count if treated_count > 0 else 0

        except Exception as e:
            self.logger.error(f"Error calculating DMT for treated communes: {e}")
            return 0

    def _calculate_ctj_from_page3(self, collaborator, pd, datetime):
        """Calculate CTJ from page 3 (Traitement PA) for this collaborator - count ONLY lines with positive duration."""
        try:
            # Check if we have page 3 data
            if 'Traitement PA' not in self.global_suivi_data:
                return 0

            df_pa = self.global_suivi_data['Traitement PA']
            if df_pa.empty:
                return 0

            # Filter data for this collaborator
            collab_pa_data = df_pa[df_pa.get('Collaborateur', '') == collaborator]
            if collab_pa_data.empty:
                return 0

            today = datetime.now().date()
            elements_today = 0

            # Check date column for today's work
            if 'Date traitement' in collab_pa_data.columns:
                for index, row in collab_pa_data.iterrows():
                    date_value = row.get('Date traitement', None)
                    duree_value = row.get('Durée', None)

                    # Skip if date is completely missing
                    if pd.isna(date_value) or date_value == '':
                        continue

                    # Skip if duration is 0, empty, or invalid
                    if pd.isna(duree_value) or duree_value == '' or duree_value == 'nan':
                        continue

                    try:
                        # Convert duration to number and check if positive
                        duree_num = float(str(duree_value).replace(',', '.'))
                        if duree_num <= 0:
                            continue  # Skip lines with duration 0 or negative
                    except (ValueError, TypeError):
                        continue  # Skip lines with invalid duration

                    try:
                        # Convert date to datetime if it's not already
                        if isinstance(date_value, str):
                            # Try different date formats
                            for date_format in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
                                try:
                                    date_obj = datetime.strptime(date_value, date_format).date()
                                    break
                                except ValueError:
                                    continue
                            else:
                                continue  # Skip if no format matches
                        else:
                            date_obj = pd.to_datetime(date_value).date()

                        # Check if date is today
                        if date_obj == today:
                            # Count this element only if duration is positive
                            elements_today += 1
                            self.logger.debug(f"CTJ: Counted element for {collaborator} on {date_obj} (duration: {duree_value})")

                    except Exception as e:
                        self.logger.debug(f"Error parsing date {date_value}: {e}")
                        continue

            self.logger.info(f"CTJ calculation for {collaborator}: {elements_today} elements today (ONLY with positive duration)")
            return elements_today

        except Exception as e:
            self.logger.error(f"Error calculating CTJ from page 3: {e}")
            return 0

    def _calculate_ctj_cm_from_page2(self, collaborator, pd, datetime):
        """Calculate CTJ CM from page 2 (Traitement CMS Adr) for this collaborator - count lines with today's date in column F (now G due to Motif Voie)."""
        try:
            # Check if we have page 2 data
            if 'Traitement CMS Adr' not in self.global_suivi_data:
                return 0

            df_cms = self.global_suivi_data['Traitement CMS Adr']
            if df_cms.empty:
                return 0

            # Filter data for this collaborator
            collab_cms_data = df_cms[df_cms.get('Collaborateur', '') == collaborator]
            if collab_cms_data.empty:
                return 0

            today = datetime.now().date()
            elements_today = 0

            # Find column F (6th column, now index 6 due to new Motif Voie column in D) - this contains dates in ISO format
            if len(df_cms.columns) < 7:
                self.logger.warning(f"CTJ CM: Not enough columns in Traitement CMS Adr sheet (found {len(df_cms.columns)}, need at least 7)")
                return 0

            date_column_f = df_cms.columns[6]  # Column F (now index 6) - contains dates (shifted due to Motif Voie in D)
            self.logger.debug(f"CTJ CM: Using column F '{date_column_f}' for date filtering (ISO format)")

            for index, row in collab_cms_data.iterrows():
                date_value_f = row.get(date_column_f, None)

                # Skip if date is completely missing
                if pd.isna(date_value_f) or date_value_f == '':
                    continue

                try:
                    # Convert date from column F to datetime
                    if isinstance(date_value_f, str):
                        # Try ISO format first (YYYY-MM-DD), then other formats
                        for date_format in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
                            try:
                                date_obj = datetime.strptime(date_value_f, date_format).date()
                                break
                            except ValueError:
                                continue
                        else:
                            continue  # Skip if no format matches
                    else:
                        date_obj = pd.to_datetime(date_value_f).date()

                    # Check if date in column F is today
                    if date_obj == today:
                        elements_today += 1
                        self.logger.debug(f"CTJ CM: Counted element for {collaborator} on {date_obj} (date from column F)")

                except Exception as e:
                    self.logger.debug(f"Error parsing date {date_value_f} from column F: {e}")
                    continue

            self.logger.info(f"CTJ CM calculation for {collaborator}: {elements_today} elements today (based on column F dates)")
            return elements_today

        except Exception as e:
            self.logger.error(f"Error calculating CTJ CM from page 2: {e}")
            return 0

    def _calculate_team_kpis(self):
        """Calculate team-wide KPIs (DMT, CTJ PA and CTJ CM) using corrected logic."""
        try:
            if not self.collaborator_stats:
                return

            total_dmt = 0
            total_ctj_today = 0
            total_ctj_cm_today = 0
            collaborator_count_with_treated = 0

            # Calculate team totals
            for collaborator, stats in self.collaborator_stats.items():
                # For DMT: only count collaborators who have treated communes (DMT > 0)
                if stats['dmt'] > 0:
                    total_dmt += stats['dmt']
                    collaborator_count_with_treated += 1

                # For CTJ PA: sum all collaborators' daily capacity
                total_ctj_today += stats['ctj_today']

                # For CTJ CM: sum all collaborators' daily capacity
                total_ctj_cm_today += stats.get('ctj_cm_today', 0)

            # Calculate team DMT (average of collaborators with treated communes)
            if collaborator_count_with_treated > 0:
                self.team_statistics['team_dmt'] = total_dmt / collaborator_count_with_treated
            else:
                self.team_statistics['team_dmt'] = 0

            # Team CTJ PA is the sum of all individual CTJ PA
            self.team_statistics['team_ctj_today'] = total_ctj_today
            self.team_statistics['total_elements_today'] = total_ctj_today

            # Team CTJ CM is the sum of all individual CTJ CM
            self.team_statistics['team_ctj_cm_today'] = total_ctj_cm_today

            self.logger.info(f"Team KPIs calculated - DMT: {self.team_statistics['team_dmt']:.1f}min (from {collaborator_count_with_treated} collaborators with treated communes), CTJ PA Today: {total_ctj_today} elements, CTJ CM Today: {total_ctj_cm_today} elements")

        except Exception as e:
            self.logger.error(f"Error calculating team KPIs: {e}")

    def _analyze_commune_status_by_date(self, df_tickets, pd, datetime):
        """Analyze communes by status and delivery date for current month filtering."""
        try:
            current_month = datetime.now().month
            current_year = datetime.now().year

            # Debug: Show available columns
            self.logger.info(f"Available columns in Suivi Tickets: {list(df_tickets.columns)}")
            self.logger.info(f"DataFrame shape: {df_tickets.shape}")

            # Initialize counters
            communes_traitees_mois_courant = 0
            communes_autres_statuts = {
                'En Cours': 0,
                'En Attente': 0,
                'Rejeté': 0,
                'Autres': 0
            }

            # Check if required columns exist (handle column name variations)
            etat_column = None
            for col in df_tickets.columns:
                if 'etat ticket pa' in col.lower().strip():
                    etat_column = col
                    break

            has_etat_ticket_pa = etat_column is not None
            has_date_livraison = 'Date Livraison' in df_tickets.columns

            self.logger.info(f"Found status column: '{etat_column}' (exists: {has_etat_ticket_pa})")
            self.logger.info(f"Has 'Date Livraison' column: {has_date_livraison}")

            if not has_etat_ticket_pa:
                self.logger.warning("Column 'Etat Ticket PA' not found in Suivi Tickets sheet")
                # Try alternative column names
                alt_columns = [col for col in df_tickets.columns if 'etat' in col.lower() or 'statut' in col.lower()]
                self.logger.info(f"Alternative status columns found: {alt_columns}")
                return

            # Initialize counters
            current_month = datetime.now().month
            current_year = datetime.now().year
            communes_traitees_mois_courant = 0
            communes_autres_statuts = {}

            for index, row in df_tickets.iterrows():
                etat_ticket_pa = row.get(etat_column, '')
                date_livraison = row.get('Date Livraison', None)

                # Skip empty rows
                if pd.isna(etat_ticket_pa) or etat_ticket_pa == '':
                    continue

                # Check if commune is treated in current month
                if etat_ticket_pa == 'Traité' and has_date_livraison and not pd.isna(date_livraison):
                    try:
                        # Convertir la date en datetime si ce n'est pas déjà fait
                        if isinstance(date_livraison, str):
                            # Essayer plusieurs formats, y compris ISO
                            date_obj = None
                            for date_format in ['%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d', '%d-%m-%Y', '%Y.%m.%d']:
                                try:
                                    date_obj = datetime.strptime(date_livraison, date_format)
                                    break
                                except ValueError:
                                    continue
                            if date_obj is None:
                                # Fallback universel
                                try:
                                    date_obj = pd.to_datetime(date_livraison)
                                except Exception:
                                    continue
                        else:
                            date_obj = pd.to_datetime(date_livraison)

                        # Vérifier si la date est dans le mois courant
                        if date_obj.month == current_month and date_obj.year == current_year:
                            communes_traitees_mois_courant += 1
                    except Exception as e:
                        self.logger.debug(f"Error parsing date {date_livraison}: {e}")
                        continue

                # Count other statuses (not treated)
                elif etat_ticket_pa.strip() != 'Traité':
                    status_clean = etat_ticket_pa.strip()
                    if status_clean in communes_autres_statuts:
                        communes_autres_statuts[status_clean] += 1
                    else:
                        communes_autres_statuts[status_clean] = 1

            # Store results
            self.team_statistics['communes_traitees_mois_courant'] = communes_traitees_mois_courant
            self.team_statistics['communes_autres_statuts'] = communes_autres_statuts

            self.logger.info(f"Communes traitées ce mois: {communes_traitees_mois_courant}")
            self.logger.info(f"Communes autres statuts: {communes_autres_statuts}")

        except Exception as e:
            self.logger.error(f"Error analyzing commune status by date: {e}")

    def _calculate_overall_averages(self):
        """Calculate overall team averages for comparison."""
        try:
            if not self.collaborator_stats:
                return

            total_cms_avg = 0
            total_pa_avg = 0
            total_finale_avg = 0
            collaborator_count = 0

            for collaborator, stats in self.collaborator_stats.items():
                if stats['commune_count'] > 0:  # Only count collaborators with communes
                    total_cms_avg += stats['avg_cms_duration']
                    total_pa_avg += stats['avg_pa_duration']
                    total_finale_avg += stats['avg_finale_duration']
                    collaborator_count += 1

            if collaborator_count > 0:
                self.overall_averages = {
                    'avg_cms_duration': total_cms_avg / collaborator_count,
                    'avg_pa_duration': total_pa_avg / collaborator_count,
                    'avg_finale_duration': total_finale_avg / collaborator_count,
                    'collaborator_count': collaborator_count
                }
            else:
                self.overall_averages = {
                    'avg_cms_duration': 0,
                    'avg_pa_duration': 0,
                    'avg_finale_duration': 0,
                    'collaborator_count': 0
                }

        except Exception as e:
            self.logger.error(f"Error calculating overall averages: {e}")
            self.overall_averages = {}

    def _update_overview_display(self):
        """Update the overview section with commune status statistics by current month."""
        try:
            # Clear existing content
            for widget in self.overview_content.winfo_children():
                widget.destroy()

            if not self.team_statistics:
                # Show debug message
                debug_label = tk.Label(
                    self.overview_content,
                    text="Aucune statistique disponible - Vérifiez que les données sont chargées",
                    font=UIConfig.FONT_SMALL,
                    fg=COLORS['DANGER'],
                    bg=COLORS['CARD']
                )
                debug_label.pack(anchor=tk.W)
                return

            # Debug: Show what we have in team_statistics
            self.logger.info(f"Team statistics keys: {list(self.team_statistics.keys())}")
            self.logger.info(f"Communes traitées mois courant: {self.team_statistics.get('communes_traitees_mois_courant', 'NOT_FOUND')}")
            self.logger.info(f"Communes autres statuts: {self.team_statistics.get('communes_autres_statuts', 'NOT_FOUND')}")

            # Current month communes section
            from datetime import datetime
            current_month_name = datetime.now().strftime("%B %Y")

            # Communes status section
            status_section = tk.Frame(self.overview_content, bg=COLORS['CARD'])
            status_section.pack(fill=tk.X, pady=(0, 10))

            status_title = tk.Label(
                status_section,
                text=f"📊 Statut des Communes - {current_month_name}",
                font=UIConfig.FONT_SUBTITLE,
                fg=COLORS['PRIMARY'],
                bg=COLORS['CARD']
            )
            status_title.pack(anchor=tk.W, pady=(0, 10))

            # Status list - vertical layout with specific order
            status_container = tk.Frame(status_section, bg=COLORS['CARD'])
            status_container.pack(fill=tk.X)

            # Define status order and icons
            status_order = ['Traité', 'En Cours', 'En Attente', 'Bloqué', 'Rejeté']
            status_icons = {
                'Traité': '✅',
                'En Cours': '🔄',
                'En Attente': '⏳',
                'Bloqué': '🚫',
                'Rejeté': '❌'
            }

            # Get all status data including treated communes
            traited_count = self.team_statistics.get('communes_traitees_mois_courant', 0)
            autres_statuts = self.team_statistics.get('communes_autres_statuts', {})

            # Combine all statuses
            all_statuses = {'Traité': traited_count}
            all_statuses.update(autres_statuts)

            # Display statuses in specified order
            for status in status_order:
                count = all_statuses.get(status, 0)
                icon = status_icons.get(status, '📋')
                self._create_status_item(status_container, f"{icon} {status}", str(count), status)



        except Exception as e:
            self.logger.error(f"Error updating overview display: {e}")

    def _create_status_item(self, parent, label, value, status_type):
        """Create a status item with consistent styling."""
        # Status colors
        status_colors = {
            'Traité': COLORS['SUCCESS'],
            'En Cours': COLORS['WARNING'],
            'En Attente': COLORS['INFO'],
            'Bloqué': COLORS['DANGER'],
            'Rejeté': COLORS['DANGER']
        }

        color = status_colors.get(status_type, COLORS['PRIMARY'])

        # Item frame
        item_frame = tk.Frame(parent, bg=COLORS['LIGHT'], relief='flat', bd=1)
        item_frame.pack(fill=tk.X, pady=2)

        # Content frame
        content_frame = tk.Frame(item_frame, bg=COLORS['LIGHT'])
        content_frame.pack(fill=tk.X, padx=12, pady=8)

        # Label (left side)
        label_widget = tk.Label(
            content_frame,
            text=label,
            font=UIConfig.FONT_SUBTITLE,
            fg=color,
            bg=COLORS['LIGHT']
        )
        label_widget.pack(side=tk.LEFT)

        # Value (right side)
        value_widget = tk.Label(
            content_frame,
            text=f"{value} communes",
            font=UIConfig.FONT_SUBTITLE,
            fg=color,
            bg=COLORS['LIGHT']
        )
        value_widget.pack(side=tk.RIGHT)

    def _create_stat_item(self, parent, row, col, label, value):
        """Create a statistics item in the overview grid."""
        item_frame = tk.Frame(parent, bg=COLORS['LIGHT'], relief='flat', bd=1)
        item_frame.grid(row=row, column=col, sticky="ew", padx=5, pady=5)

        # Configure grid weight
        parent.grid_rowconfigure(row, weight=1)

        # Content
        content = tk.Frame(item_frame, bg=COLORS['LIGHT'])
        content.pack(fill=tk.X, padx=10, pady=8)

        # Label
        label_widget = tk.Label(
            content,
            text=label,
            font=UIConfig.FONT_SMALL,
            fg=COLORS['INFO'],
            bg=COLORS['LIGHT']
        )
        label_widget.pack(anchor=tk.W)

        # Value
        value_widget = tk.Label(
            content,
            text=value,
            font=UIConfig.FONT_SUBTITLE,
            fg=COLORS['PRIMARY'],
            bg=COLORS['LIGHT']
        )
        value_widget.pack(anchor=tk.W)

    def _update_statistics_display(self):
        """Update the detailed statistics display with overall averages and collaborator details."""
        try:
            # Clear existing content
            for widget in self.stats_display.winfo_children():
                widget.destroy()

            if not self.collaborator_stats:
                no_data_label = tk.Label(
                    self.stats_display,
                    text="Aucune donnée de collaborateur trouvée",
                    font=UIConfig.FONT_SMALL,
                    fg=COLORS['INFO'],
                    bg=COLORS['CARD']
                )
                no_data_label.pack(anchor=tk.W, pady=8)  # Reduced padding
                return

            # Overall Team Averages Section (at the top)
            if self.overall_averages:
                self._create_overall_averages_section()

                # Separator
                separator = tk.Frame(self.stats_display, bg=COLORS['BORDER'], height=1)
                separator.pack(fill=tk.X, pady=(8, 10))  # Reduced padding

            # Title for individual KPIs
            title_label = tk.Label(
                self.stats_display,
                text="📊 KPIs par Collaborateur - DMT & CTJ:",
                font=UIConfig.FONT_SUBTITLE,
                fg=COLORS['PRIMARY'],
                bg=COLORS['CARD']
            )
            title_label.pack(anchor=tk.W, pady=(0, 8))  # Reduced padding

            # Create collaborator cards
            for collaborator, stats in self.collaborator_stats.items():
                self._create_collaborator_card(collaborator, stats)

            # Adjust canvas height after content is loaded
            self._adjust_canvas_height()

        except Exception as e:
            self.logger.error(f"Error updating statistics display: {e}")

    def _adjust_canvas_height(self):
        """Adjust canvas height based on content to minimize empty space."""
        try:
            if not hasattr(self, 'stats_canvas') or not self.stats_canvas:
                return

            # Update the canvas to get accurate measurements
            self.stats_canvas.update_idletasks()

            # Get the bounding box of all content
            bbox = self.stats_canvas.bbox("all")
            if bbox:
                content_height = bbox[3] - bbox[1]  # height = bottom - top

                # Set reasonable limits
                min_height = 200
                max_height = 500

                # Calculate optimal height with some padding
                optimal_height = min(max(content_height + 40, min_height), max_height)

                # Apply the height
                self.stats_canvas.configure(height=optimal_height)

                self.logger.debug(f"Adjusted canvas height to {optimal_height}px (content: {content_height}px)")

        except Exception as e:
            self.logger.error(f"Error adjusting canvas height: {e}")

    def _create_overall_averages_section(self):
        """Create the team KPIs section at the top."""
        # Team KPIs title
        kpi_title = tk.Label(
            self.stats_display,
            text="🎯 KPIs Équipe - Indicateurs de Performance",
            font=UIConfig.FONT_SUBTITLE,
            fg=COLORS['PRIMARY'],
            bg=COLORS['CARD']
        )
        kpi_title.pack(anchor=tk.W, pady=(0, 6))  # Reduced padding

        # Team KPIs card
        kpi_card = tk.Frame(self.stats_display, bg=COLORS['SUCCESS'], relief='flat', bd=1)
        kpi_card.pack(fill=tk.X, pady=(0, 4))  # Reduced padding

        kpi_content = tk.Frame(kpi_card, bg=COLORS['SUCCESS'])
        kpi_content.pack(fill=tk.X, padx=10, pady=6)  # Reduced padding

        # KPIs grid
        kpi_grid = tk.Frame(kpi_content, bg=COLORS['SUCCESS'])
        kpi_grid.pack(fill=tk.X)

        kpi_grid.grid_columnconfigure(0, weight=1)
        kpi_grid.grid_columnconfigure(1, weight=1)
        kpi_grid.grid_columnconfigure(2, weight=1)
        kpi_grid.grid_columnconfigure(3, weight=1)

        # Team DMT (Average Treatment Duration)
        team_dmt = self._format_duration(self.team_statistics.get('team_dmt', 0))
        self._create_kpi_stat(kpi_grid, 0, 0, "⏱️ DMT Équipe", team_dmt, "Durée Moyenne de Traitement")

        # Team CTJ PA Today (Daily Treatment Capacity PA)
        team_ctj_pa = self.team_statistics.get('team_ctj_today', 0)
        self._create_kpi_stat(kpi_grid, 0, 1, "📈 CTJ PA Aujourd'hui", f"{team_ctj_pa} éléments", "Capacité de Traitement Journalier PA")

        # Team CTJ CM Today (Daily Treatment Capacity CM)
        team_ctj_cm = self.team_statistics.get('team_ctj_cm_today', 0)
        self._create_kpi_stat(kpi_grid, 0, 2, "🏠 CTJ CM Aujourd'hui", f"{team_ctj_cm} éléments", "Capacité de Traitement Journalier CM")

        # Total Collaborators
        total_collabs = len(self.collaborator_stats)
        self._create_kpi_stat(kpi_grid, 0, 3, "👥 Collaborateurs", f"{total_collabs} actifs", "Équipe en activité")

    def _create_kpi_stat(self, parent, row, col, label, value, description):
        """Create a KPI statistics item for the team section."""
        kpi_frame = tk.Frame(parent, bg='white', relief='flat', bd=1)
        kpi_frame.grid(row=row, column=col, sticky="ew", padx=3, pady=2)

        content = tk.Frame(kpi_frame, bg='white')
        content.pack(fill=tk.X, padx=6, pady=4)

        # Label
        label_widget = tk.Label(
            content,
            text=label,
            font=("Segoe UI", 8, "bold"),
            fg=COLORS['SUCCESS'],
            bg='white'
        )
        label_widget.pack(anchor=tk.W)

        # Value
        value_widget = tk.Label(
            content,
            text=value,
            font=("Segoe UI", 10, "bold"),
            fg=COLORS['PRIMARY'],
            bg='white'
        )
        value_widget.pack(anchor=tk.W)

        # Description
        desc_widget = tk.Label(
            content,
            text=description,
            font=("Segoe UI", 6),
            fg=COLORS['INFO'],
            bg='white'
        )
        desc_widget.pack(anchor=tk.W)

    def _create_avg_stat(self, parent, row, col, label, value):
        """Create an average statistics item for the overall section."""
        avg_frame = tk.Frame(parent, bg='white', relief='flat', bd=1)
        avg_frame.grid(row=row, column=col, sticky="ew", padx=3, pady=2)

        content = tk.Frame(avg_frame, bg='white')
        content.pack(fill=tk.X, padx=6, pady=4)

        # Label
        label_widget = tk.Label(
            content,
            text=label,
            font=("Segoe UI", 7, "bold"),
            fg=COLORS['SECONDARY'],
            bg='white'
        )
        label_widget.pack(anchor=tk.W)

        # Value
        value_widget = tk.Label(
            content,
            text=value,
            font=UIConfig.FONT_SMALL,
            fg=COLORS['PRIMARY'],
            bg='white'
        )
        value_widget.pack(anchor=tk.W)

    def _create_collaborator_card(self, collaborator, stats):
        """Create a card displaying KPI-focused statistics for a collaborator."""
        # Get corporate spacing
        corporate_spacing = self.responsive_manager.get_corporate_spacing()

        # Card frame - with compact spacing for corporate environments
        card_frame = tk.Frame(self.stats_display, bg=COLORS['LIGHT'], relief='flat', bd=1)
        card_frame.pack(fill=tk.X, pady=(0, corporate_spacing['element_spacing']))  # Compact spacing

        # Card content - corporate-optimized padding
        card_content = tk.Frame(card_frame, bg=COLORS['LIGHT'])
        card_content.pack(
            fill=tk.X,
            padx=corporate_spacing['card_padding_x'],
            pady=corporate_spacing['card_padding_y']
        )

        # Collaborator name and commune count - compact spacing
        header_frame = tk.Frame(card_content, bg=COLORS['LIGHT'])
        header_frame.pack(fill=tk.X, pady=(0, corporate_spacing['element_spacing'] // 2))

        name_label = tk.Label(
            header_frame,
            text=f"👤 {collaborator}",
            font=UIConfig.FONT_TITLE,
            fg=COLORS['PRIMARY'],
            bg=COLORS['LIGHT']
        )
        name_label.pack(side=tk.LEFT)

        # Communes count prominently displayed
        communes_label = tk.Label(
            header_frame,
            text=f"🏘️ {stats['commune_count']} communes",
            font=UIConfig.FONT_SMALL,
            fg=COLORS['INFO'],
            bg=COLORS['LIGHT']
        )
        communes_label.pack(side=tk.RIGHT)

        # KPI Statistics grid - optimized for better space usage
        kpi_grid = tk.Frame(card_content, bg=COLORS['LIGHT'])
        kpi_grid.pack(fill=tk.X)

        # Configure grid for 1 row x 5 columns - optimized sizing
        kpi_grid.grid_columnconfigure(0, weight=1, minsize=140)  # DMT original - reduced
        kpi_grid.grid_columnconfigure(1, weight=1, minsize=140)  # DMT PA - reduced
        kpi_grid.grid_columnconfigure(2, weight=1, minsize=140)  # DMT CM - reduced
        kpi_grid.grid_columnconfigure(3, weight=1, minsize=160)  # CTJ PA - reduced
        kpi_grid.grid_columnconfigure(4, weight=1, minsize=160)  # CTJ CM - reduced

        # Get DMT values from dmt_data if available
        dmt_pa_value = "N/A"
        dmt_cm_value = "N/A"

        if hasattr(self, 'dmt_data') and self.dmt_data and collaborator in self.dmt_data:
            dmt_pa_value = f"{self.dmt_data[collaborator]['dmt_pa']} min"
            dmt_cm_value = f"{self.dmt_data[collaborator]['dmt_cm']} min"

        # Original DMT and CTJ values
        dmt_original_value = self._format_duration(stats.get('dmt', 0))
        ctj_pa_value = f"{stats.get('ctj_today', 0)} éléments"
        ctj_cm_value = f"{stats.get('ctj_cm_today', 0)} éléments"

        # Display DMT original, DMT PA, DMT CM, CTJ PA and CTJ CM
        self._create_mini_stat(kpi_grid, 0, 0, "⏱️ DMT", dmt_original_value, highlight=True)
        self._create_mini_stat(kpi_grid, 0, 1, "📋 DMT PA", dmt_pa_value, highlight=True)
        self._create_mini_stat(kpi_grid, 0, 2, "🏠 DMT CM", dmt_cm_value, highlight=True)
        self._create_mini_stat(kpi_grid, 0, 3, "📈 CTJ PA", ctj_pa_value, highlight=False)
        self._create_mini_stat(kpi_grid, 0, 4, "🏠 CTJ CM", ctj_cm_value, highlight=False)

    def _create_mini_stat(self, parent, row, col, label, value, highlight=False):
        """Create a mini statistics item with optional highlighting."""
        bg_color = COLORS['PRIMARY_LIGHT'] if highlight else COLORS['WHITE']
        text_color = 'white' if highlight else COLORS['PRIMARY']
        label_color = 'white' if highlight else COLORS['INFO']

        mini_frame = tk.Frame(parent, bg=bg_color, relief='flat', bd=1)
        mini_frame.grid(row=row, column=col, sticky="ew", padx=2, pady=1)  # Reduced horizontal spacing

        # Content - optimized padding for compact appearance
        content = tk.Frame(mini_frame, bg=bg_color)
        content.pack(fill=tk.BOTH, expand=True, padx=4, pady=2)  # Reduced padding

        # Label
        label_widget = tk.Label(
            content,
            text=label,
            font=("Segoe UI", 6, "bold" if highlight else "normal"),
            fg=label_color,
            bg=bg_color
        )
        label_widget.pack(anchor=tk.W)

        # Value
        value_widget = tk.Label(
            content,
            text=value,
            font=("Segoe UI", 8, "bold" if highlight else "normal"),
            fg=text_color,
            bg=bg_color
        )
        value_widget.pack(anchor=tk.W)

    def _format_duration(self, minutes):
        """Format duration in minutes only."""
        if minutes == 0:
            return "0 min"

        # Always display in minutes only
        return f"{int(minutes)} min"

    def _refresh_statistics(self):
        """Refresh the statistics by reloading data."""
        self._load_global_data()

    def _load_shortcut(self, event=None):
        """Keyboard shortcut for loading data."""
        self._load_global_data()

    def _refresh_shortcut(self, event=None):
        """Keyboard shortcut for refreshing."""
        if self.global_suivi_data:
            self._refresh_statistics()



    def cleanup(self):
        """Clean up resources when module is destroyed."""
        try:
            self.logger.info("Team Stats module cleaned up")
        except Exception as e:
            self.logger.error(f"Error during cleanup: {e}")

    def _export_ctj_to_excel(self):
        """Export CTJ statistics to Excel format with filters."""
        try:
            # Verify dependencies first
            try:
                from utils.lazy_imports import get_pandas
                pd = get_pandas()
                import openpyxl
                self.logger.info("Dependencies verified successfully")
            except ImportError as ie:
                messagebox.showerror("Dépendance manquante", f"Module requis non trouvé: {ie}\nVeuillez installer les dépendances nécessaires.")
                return
            except Exception as de:
                messagebox.showerror("Erreur de dépendance", f"Erreur lors de la vérification des dépendances: {de}")
                return

            # Get filter values with safety checks
            if not hasattr(self, 'collab_var') or self.collab_var is None:
                messagebox.showerror("Erreur d'initialisation", "Les filtres d'export ne sont pas initialisés. Veuillez recharger le module.")
                return

            if not hasattr(self, 'month_var') or self.month_var is None:
                messagebox.showerror("Erreur d'initialisation", "Les filtres d'export ne sont pas initialisés. Veuillez recharger le module.")
                return

            if not hasattr(self, 'year_var') or self.year_var is None:
                messagebox.showerror("Erreur d'initialisation", "Les filtres d'export ne sont pas initialisés. Veuillez recharger le module.")
                return

            try:
                selected_collaborator = self.collab_var.get()
                selected_month = self.month_var.get()
                selected_year = self.year_var.get()
            except Exception as e:
                messagebox.showerror("Erreur de filtre", f"Erreur lors de la récupération des filtres: {e}")
                return

            if not selected_collaborator or not selected_month or not selected_year:
                messagebox.showwarning("Filtres requis", "Veuillez sélectionner un collaborateur, un mois et une année.")
                return

            self.logger.info(f"Starting CTJ export for {selected_collaborator}, {selected_month} {selected_year}")

            # Check if we have the required data first
            if not hasattr(self, 'global_suivi_data') or not self.global_suivi_data:
                messagebox.showerror("Aucune donnée", "Aucune donnée globale chargée. Veuillez d'abord charger les données avec le module 'Suivi Global Tickets'.")
                return

            if 'Traitement PA' not in self.global_suivi_data:
                available_sheets = list(self.global_suivi_data.keys()) if self.global_suivi_data else []
                messagebox.showerror("Données manquantes", f"Les données 'Traitement PA' ne sont pas disponibles.\n\nFeuilles disponibles: {available_sheets}\n\nVeuillez vérifier que le fichier global contient la feuille 'Traitement PA'.")
                return

            # Calculate CTJ data for the selected filters
            ctj_data = self._calculate_monthly_ctj(selected_collaborator, selected_month, selected_year)

            if not ctj_data:
                # Provide more specific error message for team vs individual
                if selected_collaborator == "Toute l'équipe":
                    messagebox.showinfo("Aucune donnée", f"Aucune donnée CTJ trouvée pour l'équipe en {selected_month} {selected_year}.\n\nVérifiez que :\n- Les données 'Traitement PA' contiennent des collaborateurs\n- Les dates de traitement correspondent à la période sélectionnée")
                else:
                    messagebox.showinfo("Aucune donnée", f"Aucune donnée CTJ trouvée pour {selected_collaborator} en {selected_month} {selected_year}.\n\nVérifiez que :\n- Le collaborateur existe dans les données\n- Les dates de traitement correspondent à la période sélectionnée")
                return

            self.logger.info(f"CTJ data calculated: {len(ctj_data)} records")

            # Validate data structure
            if not isinstance(ctj_data, list) or len(ctj_data) == 0:
                messagebox.showerror("Erreur de données", "Les données CTJ ne sont pas dans le format attendu.")
                return

            # Create Excel file
            self._create_ctj_excel_file(ctj_data, selected_collaborator, selected_month, selected_year)

        except Exception as e:
            self.logger.error(f"Error exporting CTJ to Excel: {e}")
            import traceback
            self.logger.error(f"Full traceback: {traceback.format_exc()}")
            messagebox.showerror("Erreur", f"Erreur lors de l'export CTJ Excel:\n{str(e)}")

    def _export_ctj_combined_to_excel(self):
        """Export CTJ PA + CM combined statistics to Excel format with filters."""
        try:
            # Verify dependencies first
            try:
                from utils.lazy_imports import get_pandas
                pd = get_pandas()
                import openpyxl
                self.logger.info("Dependencies verified successfully for CTJ Combined export")
            except ImportError as ie:
                messagebox.showerror("Dépendance manquante", f"Module requis non trouvé: {ie}\nVeuillez installer les dépendances nécessaires.")
                return
            except Exception as de:
                messagebox.showerror("Erreur de dépendance", f"Erreur lors de la vérification des dépendances: {de}")
                return

            # Get filter values with safety checks
            if not hasattr(self, 'collab_var') or self.collab_var is None:
                messagebox.showerror("Erreur d'initialisation", "Les filtres d'export ne sont pas initialisés. Veuillez recharger le module.")
                return

            if not hasattr(self, 'month_var') or self.month_var is None:
                messagebox.showerror("Erreur d'initialisation", "Les filtres d'export ne sont pas initialisés. Veuillez recharger le module.")
                return

            if not hasattr(self, 'year_var') or self.year_var is None:
                messagebox.showerror("Erreur d'initialisation", "Les filtres d'export ne sont pas initialisés. Veuillez recharger le module.")
                return

            try:
                selected_collaborator = self.collab_var.get()
                selected_month = self.month_var.get()
                selected_year = self.year_var.get()
            except Exception as e:
                messagebox.showerror("Erreur de filtre", f"Erreur lors de la récupération des filtres: {e}")
                return

            if not selected_collaborator or not selected_month or not selected_year:
                messagebox.showwarning("Filtres requis", "Veuillez sélectionner un collaborateur, un mois et une année.")
                return

            self.logger.info(f"Starting CTJ Combined export for {selected_collaborator}, {selected_month} {selected_year}")

            # Check if we have the required data first
            if not hasattr(self, 'global_suivi_data') or not self.global_suivi_data:
                messagebox.showerror("Aucune donnée", "Aucune donnée globale chargée. Veuillez d'abord charger les données avec le module 'Suivi Global Tickets'.")
                return

            # Check for both required sheets
            missing_sheets = []
            if 'Traitement PA' not in self.global_suivi_data:
                missing_sheets.append('Traitement PA')
            if 'Traitement CMS Adr' not in self.global_suivi_data:
                missing_sheets.append('Traitement CMS Adr')

            if missing_sheets:
                available_sheets = list(self.global_suivi_data.keys()) if self.global_suivi_data else []
                messagebox.showerror("Données manquantes", f"Les données suivantes ne sont pas disponibles: {', '.join(missing_sheets)}\n\nFeuilles disponibles: {available_sheets}\n\nVeuillez vérifier que le fichier global contient les feuilles requises.")
                return

            # Calculate CTJ PA data for the selected filters
            ctj_pa_data = self._calculate_monthly_ctj(selected_collaborator, selected_month, selected_year)

            # Calculate CTJ CM data for the selected filters
            ctj_cm_data = self._calculate_monthly_ctj_cm(selected_collaborator, selected_month, selected_year)

            # Check if we have any data
            if not ctj_pa_data and not ctj_cm_data:
                if selected_collaborator == "Toute l'équipe":
                    messagebox.showinfo("Aucune donnée", f"Aucune donnée CTJ (PA ou CM) trouvée pour l'équipe en {selected_month} {selected_year}.\n\nVérifiez que :\n- Les données 'Traitement PA' et 'Traitement CMS Adr' contiennent des collaborateurs\n- Les dates de traitement correspondent à la période sélectionnée")
                else:
                    messagebox.showinfo("Aucune donnée", f"Aucune donnée CTJ (PA ou CM) trouvée pour {selected_collaborator} en {selected_month} {selected_year}.\n\nVérifiez que :\n- Le collaborateur existe dans les données\n- Les dates de traitement correspondent à la période sélectionnée")
                return

            self.logger.info(f"CTJ Combined data calculated: PA={len(ctj_pa_data) if ctj_pa_data else 0} records, CM={len(ctj_cm_data) if ctj_cm_data else 0} records")

            # Create Excel file with combined data
            self._create_ctj_combined_excel_file(ctj_pa_data, ctj_cm_data, selected_collaborator, selected_month, selected_year)

        except Exception as e:
            self.logger.error(f"Error exporting CTJ Combined to Excel: {e}")
            import traceback
            self.logger.error(f"Full traceback: {traceback.format_exc()}")
            messagebox.showerror("Erreur", f"Erreur lors de l'export CTJ Combiné Excel:\n{str(e)}")

    def _export_ctj_cm_to_excel(self):
        """Export CTJ CM statistics to Excel format with filters."""
        try:
            # Verify dependencies first
            try:
                from utils.lazy_imports import get_pandas
                pd = get_pandas()
                import openpyxl
                self.logger.info("Dependencies verified successfully for CTJ CM export")
            except ImportError as ie:
                messagebox.showerror("Dépendance manquante", f"Module requis non trouvé: {ie}\nVeuillez installer les dépendances nécessaires.")
                return
            except Exception as de:
                messagebox.showerror("Erreur de dépendance", f"Erreur lors de la vérification des dépendances: {de}")
                return

            # Get filter values with safety checks
            if not hasattr(self, 'collab_var') or self.collab_var is None:
                messagebox.showerror("Erreur d'initialisation", "Les filtres d'export ne sont pas initialisés. Veuillez recharger le module.")
                return

            if not hasattr(self, 'month_var') or self.month_var is None:
                messagebox.showerror("Erreur d'initialisation", "Les filtres d'export ne sont pas initialisés. Veuillez recharger le module.")
                return

            if not hasattr(self, 'year_var') or self.year_var is None:
                messagebox.showerror("Erreur d'initialisation", "Les filtres d'export ne sont pas initialisés. Veuillez recharger le module.")
                return

            try:
                selected_collaborator = self.collab_var.get()
                selected_month = self.month_var.get()
                selected_year = self.year_var.get()
            except Exception as e:
                messagebox.showerror("Erreur de filtre", f"Erreur lors de la récupération des filtres: {e}")
                return

            if not selected_collaborator or not selected_month or not selected_year:
                messagebox.showwarning("Filtres requis", "Veuillez sélectionner un collaborateur, un mois et une année.")
                return

            self.logger.info(f"Starting CTJ CM export for {selected_collaborator}, {selected_month} {selected_year}")

            # Check if we have the required data first
            if not hasattr(self, 'global_suivi_data') or not self.global_suivi_data:
                messagebox.showerror("Aucune donnée", "Aucune donnée globale chargée. Veuillez d'abord charger les données avec le module 'Suivi Global Tickets'.")
                return

            if 'Traitement CMS Adr' not in self.global_suivi_data:
                available_sheets = list(self.global_suivi_data.keys()) if self.global_suivi_data else []
                messagebox.showerror("Données manquantes", f"Les données 'Traitement CMS Adr' ne sont pas disponibles.\n\nFeuilles disponibles: {available_sheets}\n\nVeuillez vérifier que le fichier global contient la feuille 'Traitement CMS Adr'.")
                return

            # Calculate CTJ CM data for the selected filters
            ctj_cm_data = self._calculate_monthly_ctj_cm(selected_collaborator, selected_month, selected_year)

            if not ctj_cm_data:
                # Provide more specific error message for team vs individual
                if selected_collaborator == "Toute l'équipe":
                    messagebox.showinfo("Aucune donnée", f"Aucune donnée CTJ CM trouvée pour l'équipe en {selected_month} {selected_year}.\n\nVérifiez que :\n- Les données 'Traitement CMS Adr' contiennent des collaborateurs\n- Les dates de traitement correspondent à la période sélectionnée")
                else:
                    messagebox.showinfo("Aucune donnée", f"Aucune donnée CTJ CM trouvée pour {selected_collaborator} en {selected_month} {selected_year}.\n\nVérifiez que :\n- Le collaborateur existe dans les données\n- Les dates de traitement correspondent à la période sélectionnée")
                return

            self.logger.info(f"CTJ CM data calculated: {len(ctj_cm_data)} records")

            # Validate data structure
            if not isinstance(ctj_cm_data, list) or len(ctj_cm_data) == 0:
                messagebox.showerror("Erreur de données", "Les données CTJ CM ne sont pas dans le format attendu.")
                return

            # Create Excel file
            self._create_ctj_cm_excel_file(ctj_cm_data, selected_collaborator, selected_month, selected_year)

        except Exception as e:
            self.logger.error(f"Error exporting CTJ CM to Excel: {e}")
            import traceback
            self.logger.error(f"Full traceback: {traceback.format_exc()}")
            messagebox.showerror("Erreur", f"Erreur lors de l'export CTJ CM Excel:\n{str(e)}")



    def _calculate_monthly_ctj(self, collaborator, month_name, year):
        """Calculate CTJ data for a specific collaborator and month."""
        try:
            from datetime import datetime
            import calendar
            from utils.lazy_imports import get_pandas

            # Convert month name to number
            month_names = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                          "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
            month_num = month_names.index(month_name) + 1
            year_num = int(year)

            # Get page 3 data (Traitement PA)
            if 'Traitement PA' not in self.global_suivi_data:
                self.logger.warning("No 'Traitement PA' data found in global suivi data")
                return []

            df_pa = self.global_suivi_data['Traitement PA']
            if df_pa.empty:
                self.logger.warning("'Traitement PA' data is empty")
                return []

            # Check for required columns
            required_columns = ['Date traitement', 'Collaborateur']
            missing_columns = [col for col in required_columns if col not in df_pa.columns]
            if missing_columns:
                self.logger.warning(f"Missing required columns in PA data: {missing_columns}")
                self.logger.info(f"Available columns: {list(df_pa.columns)}")
                return []

            # Filter data for this collaborator (or all if "Toute l'équipe")
            if collaborator == "Toute l'équipe":
                collab_data = df_pa
                self.logger.info(f"Team export: Total PA data rows: {len(collab_data)}")
            else:
                collab_data = df_pa[df_pa.get('Collaborateur', '') == collaborator]
                if collab_data.empty:
                    self.logger.warning(f"No data found for collaborator: {collaborator}")
                    return []
                self.logger.info(f"Individual export: Found {len(collab_data)} rows for {collaborator}")

            pd = get_pandas()
            # Get number of days in the month
            days_in_month = calendar.monthrange(year_num, month_num)[1]

            if collaborator == "Toute l'équipe":
                # For team export, create data by collaborator and day
                team_ctj = {}

                # Check if 'Collaborateur' column exists
                if 'Collaborateur' not in collab_data.columns:
                    self.logger.warning("No 'Collaborateur' column found in PA data")
                    return []

                collaborators_in_data = collab_data['Collaborateur'].unique()
                valid_collaborators = [str(collab).strip() for collab in collaborators_in_data
                                     if pd.notna(collab) and str(collab).strip()]

                self.logger.info(f"Team export: Found collaborators: {valid_collaborators}")

                if not valid_collaborators:
                    self.logger.warning("No valid collaborators found in PA data")
                    return []

                # Initialize structure for all collaborators and days
                for collab in valid_collaborators:
                    team_ctj[collab] = {}
                    for day in range(1, days_in_month + 1):
                        team_ctj[collab][day] = 0

                # Count elements processed each day by each collaborator (ONLY with positive duration)
                processed_rows = 0
                valid_dates = 0

                for index, row in collab_data.iterrows():
                    date_value = row.get('Date traitement', None)
                    collab_name = row.get('Collaborateur', '')
                    duree_value = row.get('Durée', None)

                    if pd.isna(date_value) or date_value == '' or pd.isna(collab_name) or str(collab_name).strip() == '':
                        continue

                    collab_name = str(collab_name).strip()
                    if collab_name not in team_ctj:
                        continue

                    # Skip if duration is 0, empty, or invalid
                    if pd.isna(duree_value) or duree_value == '' or duree_value == 'nan':
                        continue

                    try:
                        # Convert duration to number and check if positive
                        duree_num = float(str(duree_value).replace(',', '.'))
                        if duree_num <= 0:
                            continue  # Skip lines with duration 0 or negative
                    except (ValueError, TypeError):
                        continue  # Skip lines with invalid duration

                    processed_rows += 1

                    try:
                        # Convert date to datetime
                        if isinstance(date_value, str):
                            for date_format in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
                                try:
                                    date_obj = datetime.strptime(date_value, date_format)
                                    break
                                except ValueError:
                                    continue
                            else:
                                continue
                        else:
                            date_obj = pd.to_datetime(date_value)

                        # Check if date is in the selected month/year
                        if date_obj.month == month_num and date_obj.year == year_num:
                            day = date_obj.day
                            team_ctj[collab_name][day] += 1
                            valid_dates += 1

                    except Exception as e:
                        self.logger.debug(f"Error parsing date {date_value}: {e}")
                        continue

                self.logger.info(f"Team export: Processed {processed_rows} rows, found {valid_dates} valid dates for {month_name} {year}")

                # Convert to list format for Excel export (team format)
                ctj_data = []
                total_team_ctj = 0

                for day in range(1, days_in_month + 1):
                    date_str = f"{day:02d}/{month_num:02d}/{year_num}"
                    row_data = {'Date': date_str}

                    # Add each collaborator's CTJ for this day
                    total_day = 0
                    for collab in sorted(team_ctj.keys()):
                        ctj_value = team_ctj[collab][day]
                        row_data[collab] = ctj_value
                        total_day += ctj_value

                    row_data['Total'] = total_day
                    total_team_ctj += total_day
                    ctj_data.append(row_data)

                self.logger.info(f"Team export: Generated {len(ctj_data)} days of data, total CTJ: {total_team_ctj}")

                # If no data was found, create a minimal structure for the export to work
                if total_team_ctj == 0 and not team_ctj:
                    self.logger.warning("No CTJ data found for team export, creating minimal structure")
                    # Create a minimal structure with at least one dummy collaborator
                    for day in range(1, days_in_month + 1):
                        date_str = f"{day:02d}/{month_num:02d}/{year_num}"
                        ctj_data[day-1] = {'Date': date_str, 'Aucune donnée': 0, 'Total': 0}

                return ctj_data

            else:
                # Individual collaborator export
                daily_ctj = {}

                # Initialize all days with 0
                for day in range(1, days_in_month + 1):
                    daily_ctj[day] = 0

                # Count elements processed each day (ONLY with positive duration)
                for index, row in collab_data.iterrows():
                    date_value = row.get('Date traitement', None)
                    duree_value = row.get('Durée', None)

                    if pd.isna(date_value) or date_value == '':
                        continue

                    # Skip if duration is 0, empty, or invalid
                    if pd.isna(duree_value) or duree_value == '' or duree_value == 'nan':
                        continue

                    try:
                        # Convert duration to number and check if positive
                        duree_num = float(str(duree_value).replace(',', '.'))
                        if duree_num <= 0:
                            continue  # Skip lines with duration 0 or negative
                    except (ValueError, TypeError):
                        continue  # Skip lines with invalid duration

                    try:
                        # Convert date to datetime
                        if isinstance(date_value, str):
                            for date_format in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
                                try:
                                    date_obj = datetime.strptime(date_value, date_format)
                                    break
                                except ValueError:
                                    continue
                            else:
                                continue
                        else:
                            date_obj = pd.to_datetime(date_value)

                        # Check if date is in the selected month/year
                        if date_obj.month == month_num and date_obj.year == year_num:
                            day = date_obj.day
                            daily_ctj[day] += 1

                    except Exception as e:
                        self.logger.debug(f"Error parsing date {date_value}: {e}")
                        continue

                # Convert to list format for Excel export
                ctj_data = []
                for day in range(1, days_in_month + 1):
                    date_str = f"{day:02d}/{month_num:02d}/{year_num}"
                    ctj_data.append({
                        'Date': date_str,
                        'CTJ': daily_ctj[day]
                    })

                return ctj_data

        except Exception as e:
            self.logger.error(f"Error calculating monthly CTJ: {e}")
            return []

    def _calculate_monthly_ctj_cm(self, collaborator, month_name, year):
        """Calculate CTJ CM data for a specific collaborator and month from Traitement CMS Adr sheet (column F now G due to Motif Voie)."""
        try:
            from datetime import datetime
            import calendar
            from utils.lazy_imports import get_pandas

            # Convert month name to number
            month_names = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                          "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
            month_num = month_names.index(month_name) + 1
            year_num = int(year)

            # Get page 2 data (Traitement CMS Adr)
            if 'Traitement CMS Adr' not in self.global_suivi_data:
                self.logger.warning("No 'Traitement CMS Adr' data found in global suivi data")
                return []

            df_cms = self.global_suivi_data['Traitement CMS Adr']
            if df_cms.empty:
                self.logger.warning("'Traitement CMS Adr' data is empty")
                return []

            # Check for required columns
            required_columns = ['Collaborateur']
            missing_columns = [col for col in required_columns if col not in df_cms.columns]
            if missing_columns:
                self.logger.warning(f"Missing required columns in CMS data: {missing_columns}")
                self.logger.info(f"Available columns: {list(df_cms.columns)}")
                return []

            # Find column F (6th column, now index 6 due to new Motif Voie column in D) - this contains dates in ISO format
            if len(df_cms.columns) < 7:
                self.logger.warning(f"CTJ CM: Not enough columns in Traitement CMS Adr sheet (found {len(df_cms.columns)}, need at least 7)")
                return []

            date_column_f = df_cms.columns[6]  # Column F (now index 6) - contains dates (shifted due to Motif Voie in D)
            self.logger.info(f"CTJ CM: Using column F '{date_column_f}' for date filtering (ISO format)")

            # Filter data for this collaborator (or all if "Toute l'équipe")
            if collaborator == "Toute l'équipe":
                collab_data = df_cms
                self.logger.info(f"Team export: Total CMS data rows: {len(collab_data)}")
            else:
                collab_data = df_cms[df_cms.get('Collaborateur', '') == collaborator]
                self.logger.info(f"Individual export: Filtered CMS data for {collaborator}: {len(collab_data)} rows")

            if collab_data.empty:
                self.logger.warning(f"No CMS data found for collaborator: {collaborator}")
                return []

            pd = get_pandas()

            # Get number of days in the month
            days_in_month = calendar.monthrange(year_num, month_num)[1]

            if collaborator == "Toute l'équipe":
                # For team export, create data by collaborator and day
                team_ctj_cm = {}

                # Check if 'Collaborateur' column exists
                if 'Collaborateur' not in collab_data.columns:
                    self.logger.warning("No 'Collaborateur' column found in CMS data")
                    return []

                collaborators_in_data = collab_data['Collaborateur'].unique()
                valid_collaborators = [str(collab).strip() for collab in collaborators_in_data
                                     if pd.notna(collab) and str(collab).strip()]

                if not valid_collaborators:
                    self.logger.warning("No valid collaborators found in CMS data")
                    return []

                # Initialize structure for all collaborators and days
                for collab in valid_collaborators:
                    team_ctj_cm[collab] = {}
                    for day in range(1, days_in_month + 1):
                        team_ctj_cm[collab][day] = 0

                # Count elements processed each day by each collaborator (ONLY with positive duration)
                processed_rows = 0
                valid_dates = 0

                for index, row in collab_data.iterrows():
                    date_value_f = row.get(date_column_f, None)
                    collab_name = row.get('Collaborateur', '')

                    if pd.isna(date_value_f) or date_value_f == '' or pd.isna(collab_name) or str(collab_name).strip() == '':
                        continue

                    collab_name = str(collab_name).strip()
                    if collab_name not in team_ctj_cm:
                        continue

                    processed_rows += 1

                    try:
                        # Convert date from column F to datetime
                        if isinstance(date_value_f, str):
                            # Try ISO format first (YYYY-MM-DD), then other formats
                            for date_format in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
                                try:
                                    date_obj = datetime.strptime(date_value_f, date_format)
                                    break
                                except ValueError:
                                    continue
                            else:
                                continue  # Skip if no format matches
                        else:
                            date_obj = pd.to_datetime(date_value_f)

                        # Check if date is in the selected month/year
                        if date_obj.month == month_num and date_obj.year == year_num:
                            day = date_obj.day
                            team_ctj_cm[collab_name][day] += 1
                            valid_dates += 1

                    except Exception as e:
                        self.logger.debug(f"Error parsing date {date_value_f} from column F in CMS data: {e}")
                        continue

                self.logger.info(f"Team CTJ CM export: Processed {processed_rows} rows, found {valid_dates} valid dates for {month_name} {year}")

                # Convert to list format for Excel export (team format)
                ctj_cm_data = []
                total_team_ctj_cm = 0

                for day in range(1, days_in_month + 1):
                    date_str = f"{day:02d}/{month_num:02d}/{year_num}"
                    row_data = {'Date': date_str}

                    # Add each collaborator's CTJ CM for this day
                    total_day = 0
                    for collab in sorted(team_ctj_cm.keys()):
                        ctj_cm_value = team_ctj_cm[collab][day]
                        row_data[collab] = ctj_cm_value
                        total_day += ctj_cm_value

                    row_data['Total'] = total_day
                    total_team_ctj_cm += total_day
                    ctj_cm_data.append(row_data)

                self.logger.info(f"Team CTJ CM export: Generated {len(ctj_cm_data)} days of data, total CTJ CM: {total_team_ctj_cm}")

                # If no data was found, create a minimal structure for the export to work
                if total_team_ctj_cm == 0 and not team_ctj_cm:
                    self.logger.warning("No CTJ CM data found for team export, creating minimal structure")
                    # Create a minimal structure with at least one dummy collaborator
                    for day in range(1, days_in_month + 1):
                        date_str = f"{day:02d}/{month_num:02d}/{year_num}"
                        ctj_cm_data[day-1] = {'Date': date_str, 'Aucune donnée': 0, 'Total': 0}

                return ctj_cm_data

            else:
                # Individual collaborator export
                daily_ctj_cm = {}

                # Initialize all days with 0
                for day in range(1, days_in_month + 1):
                    daily_ctj_cm[day] = 0

                # Count elements processed each day (based on dates in column F)
                for index, row in collab_data.iterrows():
                    date_value_f = row.get(date_column_f, None)

                    if pd.isna(date_value_f) or date_value_f == '':
                        continue

                    try:
                        # Convert date from column F to datetime
                        if isinstance(date_value_f, str):
                            # Try ISO format first (YYYY-MM-DD), then other formats
                            for date_format in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
                                try:
                                    date_obj = datetime.strptime(date_value_f, date_format)
                                    break
                                except ValueError:
                                    continue
                            else:
                                continue  # Skip if no format matches
                        else:
                            date_obj = pd.to_datetime(date_value_f)

                        # Check if date is in the selected month/year
                        if date_obj.month == month_num and date_obj.year == year_num:
                            day = date_obj.day
                            daily_ctj_cm[day] += 1

                    except Exception as e:
                        self.logger.debug(f"Error parsing date {date_value_f} from column F: {e}")
                        continue

                # Convert to list format for Excel export
                ctj_cm_data = []
                for day in range(1, days_in_month + 1):
                    date_str = f"{day:02d}/{month_num:02d}/{year_num}"
                    ctj_cm_data.append({
                        'Date': date_str,
                        'CTJ_CM': daily_ctj_cm[day]
                    })

                return ctj_cm_data

        except Exception as e:
            self.logger.error(f"Error calculating monthly CTJ CM: {e}")
            return []

    def _create_ctj_excel_file(self, ctj_data, collaborator, month_name, year):
        """Create Excel file with CTJ data in horizontal date format."""
        try:
            import os
            from datetime import datetime, timedelta
            from tkinter import filedialog
            import calendar

            # Generate filename
            current_date = datetime.now().strftime("%Y%m%d")
            safe_collaborator = collaborator.replace(' ', '_').replace("'", "")
            filename = f"Stat_{safe_collaborator}_{current_date}.xlsx"

            # Ask user for save location
            file_path = filedialog.asksaveasfilename(
                title="Enregistrer les statistiques CTJ",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=filename
            )

            if not file_path:
                return

            # Create the new format Excel file
            self._create_horizontal_ctj_excel(file_path, ctj_data, collaborator, month_name, year)

            messagebox.showinfo("Export réussi", f"Fichier CTJ exporté avec succès:\n{file_path}")
            self.logger.info(f"CTJ Excel file created: {file_path}")

        except Exception as e:
            self.logger.error(f"Error creating CTJ Excel file: {e}")

            # Provide user-friendly error messages
            error_str = str(e)
            if "Permission denied" in error_str or "WinError 33" in error_str or "annulé par l'utilisateur" in error_str:
                # File access issues - user already saw the dialog or cancelled
                if "annulé par l'utilisateur" not in error_str:
                    messagebox.showwarning(
                        "Fichier en cours d'utilisation",
                        "🔒 Le fichier d'export est actuellement ouvert dans Excel.\n\n"
                        "Veuillez fermer Excel et réessayer l'export."
                    )
            else:
                # Other errors
                messagebox.showerror("Erreur Export", f"Erreur lors de la création du fichier Excel:\n{str(e)}")
            raise

    def _create_ctj_combined_excel_file(self, ctj_pa_data, ctj_cm_data, collaborator, month_name, year):
        """Create Excel file with combined CTJ PA + CM data."""
        try:
            import os
            from datetime import datetime, timedelta
            from tkinter import filedialog
            import calendar

            # Generate filename
            current_date = datetime.now().strftime("%Y%m%d")
            safe_collaborator = collaborator.replace(' ', '_').replace("'", "")
            filename = f"Stat_CTJ_Combined_{safe_collaborator}_{current_date}.xlsx"

            # Ask user for save location
            file_path = filedialog.asksaveasfilename(
                title="Enregistrer les statistiques CTJ Combiné (PA + CM)",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=filename
            )

            if not file_path:
                return

            # Create the combined Excel file with motifs statistics
            self._create_horizontal_ctj_combined_excel_with_motifs(file_path, ctj_pa_data, ctj_cm_data, collaborator, month_name, year)

            messagebox.showinfo("Export réussi", f"Fichier CTJ Combiné (PA + CM) exporté avec succès:\n{file_path}\n\n✨ Inclut maintenant les statistiques motifs mensuels!")
            self.logger.info(f"CTJ Combined Excel file with motifs statistics created: {file_path}")

        except Exception as e:
            self.logger.error(f"Error creating CTJ Combined Excel file: {e}")

            # Provide user-friendly error messages
            error_str = str(e)
            if "Permission denied" in error_str or "WinError 33" in error_str or "annulé par l'utilisateur" in error_str:
                # File access issues - user already saw the dialog or cancelled
                if "annulé par l'utilisateur" not in error_str:
                    messagebox.showwarning(
                        "Fichier en cours d'utilisation",
                        "🔒 Le fichier d'export est actuellement ouvert dans Excel.\n\n"
                        "Veuillez fermer Excel et réessayer l'export."
                    )
            else:
                # Other errors
                messagebox.showerror("Erreur Export", f"Erreur lors de la création du fichier Excel:\n{str(e)}")
            raise

    def _create_ctj_cm_excel_file(self, ctj_cm_data, collaborator, month_name, year):
        """Create Excel file with CTJ CM data in horizontal date format."""
        try:
            import os
            from datetime import datetime, timedelta
            from tkinter import filedialog
            import calendar

            # Generate filename
            current_date = datetime.now().strftime("%Y%m%d")
            safe_collaborator = collaborator.replace(' ', '_').replace("'", "")
            filename = f"Stat_CTJ_CM_{safe_collaborator}_{current_date}.xlsx"

            # Ask user for save location
            file_path = filedialog.asksaveasfilename(
                title="Enregistrer les statistiques CTJ CM",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=filename
            )

            if not file_path:
                return

            # Create the new format Excel file
            self._create_horizontal_ctj_cm_excel(file_path, ctj_cm_data, collaborator, month_name, year)

            messagebox.showinfo("Export réussi", f"Fichier CTJ CM exporté avec succès:\n{file_path}")
            self.logger.info(f"CTJ CM Excel file created: {file_path}")

        except Exception as e:
            self.logger.error(f"Error creating CTJ CM Excel file: {e}")

            # Provide user-friendly error messages
            error_str = str(e)
            if "Permission denied" in error_str or "WinError 33" in error_str or "annulé par l'utilisateur" in error_str:
                # File access issues - user already saw the dialog or cancelled
                if "annulé par l'utilisateur" not in error_str:
                    messagebox.showwarning(
                        "Fichier en cours d'utilisation",
                        "🔒 Le fichier d'export est actuellement ouvert dans Excel.\n\n"
                        "Veuillez fermer Excel et réessayer l'export."
                    )
            else:
                # Other errors
                messagebox.showerror("Erreur Export", f"Erreur lors de la création du fichier Excel:\n{str(e)}")
            raise

    def _create_horizontal_ctj_cm_excel(self, file_path, ctj_cm_data, collaborator, month_name, year):
        """Create streamlined Excel file with only CTJ CM Quotidien sheet."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from datetime import datetime, timedelta
            import calendar

            # Create workbook
            workbook = openpyxl.Workbook()

            # Get the default sheet to remove it later
            default_sheet = workbook.active

            # Create only the CTJ CM Quotidien sheet as requested
            success_daily = self._create_streamlined_daily_ctj_cm_sheet(workbook, ctj_cm_data, collaborator, month_name, year)

            # Always remove the default sheet if we successfully created our custom sheet
            if success_daily:
                try:
                    workbook.remove(default_sheet)
                    self.logger.info("Default sheet removed successfully")
                except Exception as e:
                    self.logger.warning(f"Could not remove default sheet: {e}")
            else:
                # If creation failed, provide detailed error message
                error_msg = f"Failed to create CTJ CM export sheet for {collaborator}. Please check if data is available for the selected period."
                self.logger.error(error_msg)
                raise Exception(error_msg)

            # Check file access before saving
            if os.path.exists(file_path):
                access_result = check_file_access(file_path, 'w')
                if not access_result['accessible']:
                    self.logger.warning(f"File access issue during export: {access_result['error_message']}")

                    # Show user-friendly dialog and get retry decision
                    retry_decision = messagebox.askyesnocancel(
                        "Fichier en cours d'utilisation",
                        f"🔒 Le fichier d'export est actuellement ouvert dans Excel.\n\n"
                        f"Fichier: {file_path}\n\n"
                        f"Voulez-vous réessayer après avoir fermé Excel?\n\n"
                        f"• Oui: Réessayer maintenant\n"
                        f"• Non: Choisir un autre nom de fichier\n"
                        f"• Annuler: Abandonner l'export"
                    )

                    if retry_decision is None:  # Cancel
                        raise Exception("Export annulé par l'utilisateur")
                    elif retry_decision is False:  # No - choose different file
                        raise Exception(access_result['user_message'])

            # Save the streamlined workbook
            workbook.save(file_path)

            self.logger.info(f"Streamlined CTJ CM Excel file created successfully: {file_path}")

        except Exception as e:
            self.logger.error(f"Error creating streamlined CTJ CM Excel: {e}")
            raise

    def _create_horizontal_ctj_combined_excel(self, file_path, ctj_pa_data, ctj_cm_data, collaborator, month_name, year):
        """Create Excel file with combined CTJ PA + CM data."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from datetime import datetime, timedelta
            import calendar

            # Create workbook
            workbook = openpyxl.Workbook()

            # Get the default sheet to remove it later
            default_sheet = workbook.active

            # Create combined CTJ sheet
            success_combined = self._create_combined_ctj_sheet(workbook, ctj_pa_data, ctj_cm_data, collaborator, month_name, year)

            # Always remove the default sheet if we successfully created our custom sheet
            if success_combined:
                try:
                    workbook.remove(default_sheet)
                    self.logger.info("Default sheet removed successfully")
                except Exception as e:
                    self.logger.warning(f"Could not remove default sheet: {e}")
            else:
                # If creation failed, provide detailed error message
                error_msg = f"Failed to create CTJ Combined export sheet for {collaborator}. Please check if data is available for the selected period."
                self.logger.error(error_msg)
                raise Exception(error_msg)

            # Check file access before saving
            if os.path.exists(file_path):
                access_result = check_file_access(file_path, 'w')
                if not access_result['accessible']:
                    self.logger.warning(f"File access issue during export: {access_result['error_message']}")

                    # Show user-friendly dialog and get retry decision
                    retry_decision = messagebox.askyesnocancel(
                        "Fichier en cours d'utilisation",
                        f"🔒 Le fichier d'export est actuellement ouvert dans Excel.\n\n"
                        f"Fichier: {file_path}\n\n"
                        f"Voulez-vous réessayer après avoir fermé Excel?\n\n"
                        f"• Oui: Réessayer maintenant\n"
                        f"• Non: Choisir un autre nom de fichier\n"
                        f"• Annuler: Abandonner l'export"
                    )

                    if retry_decision is None:  # Cancel
                        raise Exception("Export annulé par l'utilisateur")
                    elif retry_decision is False:  # No - choose different file
                        raise Exception(access_result['user_message'])

            # Save the combined workbook
            workbook.save(file_path)

            self.logger.info(f"CTJ Combined Excel file created successfully: {file_path}")

        except Exception as e:
            self.logger.error(f"Error creating CTJ Combined Excel: {e}")
            raise

    def _create_horizontal_ctj_combined_excel_with_motifs(self, file_path, ctj_pa_data, ctj_cm_data, collaborator, month_name, year):
        """Create Excel file with combined CTJ PA + CM data and motifs statistics."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from datetime import datetime, timedelta

            # Create workbook
            workbook = openpyxl.Workbook()

            # Remove default sheet
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])

            # 1. Create combined CTJ sheet (existing functionality)
            success = self._create_combined_ctj_sheet(workbook, ctj_pa_data, ctj_cm_data, collaborator, month_name, year)
            if not success:
                self.logger.warning("Failed to create combined CTJ sheet")

            # 2. Extract and create motifs statistics sheet (NEW)
            motifs_data = self._extract_monthly_motifs_data(month_name, year)
            if motifs_data:
                motifs_success = self._create_motifs_statistics_sheet(workbook, motifs_data, month_name, year)
                if motifs_success:
                    self.logger.info(f"Added motifs statistics sheet with {len(motifs_data)} records")
                else:
                    self.logger.warning("Failed to create motifs statistics sheet")
            else:
                self.logger.info("No motifs data found for the selected month/year")
                # Create empty motifs sheet with explanation
                self._create_empty_motifs_sheet(workbook, month_name, year)

            # Save workbook
            workbook.save(file_path)
            workbook.close()

            self.logger.info(f"CTJ Combined Excel with motifs statistics created: {file_path}")

        except Exception as e:
            self.logger.error(f"Error creating CTJ Combined Excel with motifs: {e}")
            raise

    def _create_empty_motifs_sheet(self, workbook, month_name, year):
        """Create empty motifs sheet with explanation when no data is available."""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment

            # Create the sheet
            motifs_sheet = workbook.create_sheet(title="Statistiques Motifs Mensuels")

            # Title
            motifs_sheet.cell(row=1, column=1).value = f"Statistiques Motifs Mensuels - {month_name} {year}"
            motifs_sheet.cell(row=1, column=1).font = Font(size=14, bold=True, name="Calibri")
            motifs_sheet.merge_cells('A1:D1')

            # Explanation
            motifs_sheet.cell(row=3, column=1).value = "ℹ️ Aucune donnée de motifs disponible pour cette période"
            motifs_sheet.cell(row=3, column=1).font = Font(size=12, name="Calibri")
            motifs_sheet.merge_cells('A3:D3')

            motifs_sheet.cell(row=5, column=1).value = "Causes possibles :"
            motifs_sheet.cell(row=5, column=1).font = Font(size=11, bold=True, name="Calibri")

            explanations = [
                "• Aucun traitement effectué ce mois-ci",
                "• Données de motifs manquantes dans les fichiers sources",
                "• Colonnes de motifs non renseignées",
                "• Dates de traitement en dehors de la période sélectionnée"
            ]

            for i, explanation in enumerate(explanations, 6):
                motifs_sheet.cell(row=i, column=1).value = explanation
                motifs_sheet.cell(row=i, column=1).font = Font(size=10, name="Calibri")

            # Optimize column width
            motifs_sheet.column_dimensions['A'].width = 50

            self.logger.info("Created empty motifs sheet with explanation")

        except Exception as e:
            self.logger.error(f"Error creating empty motifs sheet: {e}")

    def _create_combined_ctj_sheet(self, workbook, ctj_pa_data, ctj_cm_data, collaborator, month_name, year):
        """Create combined CTJ sheet with both PA and CM data."""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from datetime import datetime
            import calendar

            # Create combined sheet
            combined_sheet = workbook.create_sheet(title="CTJ Combiné (PA + CM)")

            # Define enhanced styles with better colors and formatting
            header_font = Font(size=12, bold=True, color="FFFFFF", name="Calibri")
            data_font = Font(size=10, name="Calibri")
            collaborator_font = Font(size=11, bold=True, name="Calibri")
            total_font = Font(size=10, bold=True, name="Calibri")

            # Enhanced color scheme with better contrast and professional look
            header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")  # Darker blue
            pa_fill = PatternFill(start_color="E1F5FE", end_color="E1F5FE", fill_type="solid")  # Light cyan for PA
            cm_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")  # Light amber for CM
            weekend_fill = PatternFill(start_color="ECEFF1", end_color="ECEFF1", fill_type="solid")  # Light gray
            zero_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")  # Light red
            separator_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")  # Very light gray

            center_align = Alignment(horizontal="center", vertical="center")
            left_align = Alignment(horizontal="left", vertical="center")

            # Enhanced border styles
            thin_border = Border(
                left=Side(style='thin', color="666666"),
                right=Side(style='thin', color="666666"),
                top=Side(style='thin', color="666666"),
                bottom=Side(style='thin', color="666666")
            )

            thick_border = Border(
                left=Side(style='medium', color="2F5597"),
                right=Side(style='medium', color="2F5597"),
                top=Side(style='medium', color="2F5597"),
                bottom=Side(style='medium', color="2F5597")
            )

            # Get month info
            month_names = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                          "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
            month_num = month_names.index(month_name) + 1
            year_num = int(year)
            days_in_month = calendar.monthrange(year_num, month_num)[1]

            # Month abbreviations for headers
            month_abbrev = {
                1: "Jan", 2: "Fév", 3: "Mar", 4: "Avr", 5: "Mai", 6: "Jun",
                7: "Jul", 8: "Aoû", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Déc"
            }

            # Title
            combined_sheet.merge_cells('A1:AH1')
            title_cell = combined_sheet['A1']
            title_cell.value = f"CTJ COMBINÉ (PA + CM) - {collaborator.upper()} - {month_name} {year}"
            title_cell.font = Font(size=14, bold=True, color="FFFFFF")
            title_cell.fill = header_fill
            title_cell.alignment = center_align

            # Headers with date format
            headers = ['Type CTJ']
            for day in range(1, days_in_month + 1):
                headers.append(f"{day} {month_abbrev[month_num]}")
            headers.append('Total')

            # Write headers with borders
            for col, header in enumerate(headers, 1):
                cell = combined_sheet.cell(row=3, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border

            # Process data for individual or team export with enhanced styles
            if collaborator == "Toute l'équipe":
                return self._create_team_combined_ctj_export(combined_sheet, ctj_pa_data, ctj_cm_data, month_name, year,
                                                   header_font, data_font, header_fill, pa_fill, cm_fill,
                                                   weekend_fill, zero_fill, center_align, thin_border,
                                                   collaborator_font, total_font, separator_fill, thick_border)
            else:
                return self._create_individual_combined_ctj_export(combined_sheet, ctj_pa_data, ctj_cm_data, collaborator, month_name, year,
                                                         header_font, data_font, header_fill, pa_fill, cm_fill,
                                                         weekend_fill, zero_fill, center_align, thin_border,
                                                         collaborator_font, total_font, separator_fill, thick_border)

        except Exception as e:
            self.logger.error(f"Error creating combined CTJ sheet: {e}")
            import traceback
            self.logger.error(f"Full traceback: {traceback.format_exc()}")
            return False

    def _create_individual_combined_ctj_export(self, combined_sheet, ctj_pa_data, ctj_cm_data, collaborator, month_name, year,
                                     header_font, data_font, header_fill, pa_fill, cm_fill,
                                     weekend_fill, zero_fill, center_align, thin_border):
        """Create individual collaborator combined CTJ export with PA and CM data."""
        try:
            from datetime import datetime
            import calendar
            from openpyxl.styles import Font

            self.logger.info(f"Creating individual combined CTJ export for {collaborator}")

            # Get month info
            month_names = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                          "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
            month_num = month_names.index(month_name) + 1
            year_num = int(year)
            days_in_month = calendar.monthrange(year_num, month_num)[1]

            # Prepare data arrays
            pa_values = [0] * days_in_month
            cm_values = [0] * days_in_month

            # Fill PA data
            if ctj_pa_data:
                for day_data in ctj_pa_data:
                    if 'Date' in day_data and 'CTJ' in day_data:
                        try:
                            # Extract day from date string (format: DD/MM/YYYY)
                            date_str = day_data['Date']
                            day = int(date_str.split('/')[0])
                            if 1 <= day <= days_in_month:
                                pa_values[day - 1] = day_data['CTJ']
                        except (ValueError, IndexError):
                            continue

            # Fill CM data
            if ctj_cm_data:
                for day_data in ctj_cm_data:
                    if 'Date' in day_data and 'CTJ_CM' in day_data:
                        try:
                            # Extract day from date string (format: DD/MM/YYYY)
                            date_str = day_data['Date']
                            day = int(date_str.split('/')[0])
                            if 1 <= day <= days_in_month:
                                cm_values[day - 1] = day_data['CTJ_CM']
                        except (ValueError, IndexError):
                            continue

            # Write CTJ PA row
            combined_sheet.cell(row=4, column=1).value = "CTJ PA"
            combined_sheet.cell(row=4, column=1).font = data_font
            combined_sheet.cell(row=4, column=1).fill = pa_fill
            combined_sheet.cell(row=4, column=1).border = thin_border

            total_pa = 0
            for day in range(1, days_in_month + 1):
                col = day + 1
                pa_value = pa_values[day - 1]
                total_pa += pa_value

                cell = combined_sheet.cell(row=4, column=col)
                cell.value = pa_value if pa_value > 0 else ""
                cell.font = data_font
                cell.alignment = center_align
                cell.border = thin_border
                cell.fill = pa_fill

                # Apply weekend formatting
                date_obj = datetime(year_num, month_num, day)
                if date_obj.weekday() >= 5:  # Weekend
                    cell.fill = weekend_fill
                elif pa_value == 0:
                    cell.fill = zero_fill

            # Total PA column
            total_cell = combined_sheet.cell(row=4, column=days_in_month + 2)
            total_cell.value = total_pa
            total_cell.font = data_font
            total_cell.alignment = center_align
            total_cell.border = thin_border
            total_cell.fill = pa_fill

            # Write CTJ CM row
            combined_sheet.cell(row=5, column=1).value = "CTJ CM"
            combined_sheet.cell(row=5, column=1).font = data_font
            combined_sheet.cell(row=5, column=1).fill = cm_fill
            combined_sheet.cell(row=5, column=1).border = thin_border

            total_cm = 0
            for day in range(1, days_in_month + 1):
                col = day + 1
                cm_value = cm_values[day - 1]
                total_cm += cm_value

                cell = combined_sheet.cell(row=5, column=col)
                cell.value = cm_value if cm_value > 0 else ""
                cell.font = data_font
                cell.alignment = center_align
                cell.border = thin_border
                cell.fill = cm_fill

                # Apply weekend formatting
                date_obj = datetime(year_num, month_num, day)
                if date_obj.weekday() >= 5:  # Weekend
                    cell.fill = weekend_fill
                elif cm_value == 0:
                    cell.fill = zero_fill

            # Total CM column
            total_cell = combined_sheet.cell(row=5, column=days_in_month + 2)
            total_cell.value = total_cm
            total_cell.font = data_font
            total_cell.alignment = center_align
            total_cell.border = thin_border
            total_cell.fill = cm_fill

            # Write Total Combined row
            combined_sheet.cell(row=6, column=1).value = "TOTAL (PA + CM)"
            combined_sheet.cell(row=6, column=1).font = Font(bold=True)
            combined_sheet.cell(row=6, column=1).fill = header_fill
            combined_sheet.cell(row=6, column=1).border = thin_border

            total_combined = 0
            for day in range(1, days_in_month + 1):
                col = day + 1
                combined_value = pa_values[day - 1] + cm_values[day - 1]
                total_combined += combined_value

                cell = combined_sheet.cell(row=6, column=col)
                cell.value = combined_value if combined_value > 0 else ""
                cell.font = Font(bold=True)
                cell.alignment = center_align
                cell.border = thin_border
                cell.fill = header_fill

            # Total Combined column
            total_cell = combined_sheet.cell(row=6, column=days_in_month + 2)
            total_cell.value = total_combined
            total_cell.font = Font(bold=True)
            total_cell.alignment = center_align
            total_cell.border = thin_border
            total_cell.fill = header_fill

            # Add summary section
            self._add_combined_summary(combined_sheet, total_pa, total_cm, total_combined, collaborator, month_name, year,
                                     days_in_month, thin_border, header_font, data_font, center_align)

            # Auto-fit columns
            for col in range(1, days_in_month + 3):
                from openpyxl.utils import get_column_letter
                column_letter = get_column_letter(col)
                max_width = 0
                for row in range(1, 15):
                    try:
                        cell_value = combined_sheet[f"{column_letter}{row}"].value
                        if cell_value:
                            cell_length = len(str(cell_value))
                            max_width = max(max_width, cell_length)
                    except:
                        continue
                column_width = min(max(max_width + 1, 8), 15)
                combined_sheet.column_dimensions[column_letter].width = column_width

            self.logger.info(f"Individual combined CTJ export completed: PA={total_pa}, CM={total_cm}, Total={total_combined}")
            return True

        except Exception as e:
            self.logger.error(f"Error creating individual combined CTJ export: {e}")
            return False

    def _add_combined_summary(self, combined_sheet, total_pa, total_cm, total_combined, collaborator, month_name, year,
                            days_in_month, thin_border, header_font, data_font, center_align):
        """Add summary section to combined CTJ export."""
        try:
            from openpyxl.styles import Font, PatternFill

            # Summary section starting at row 8
            summary_row = 8

            # Summary title
            combined_sheet.merge_cells(f'A{summary_row}:C{summary_row}')
            title_cell = combined_sheet[f'A{summary_row}']
            title_cell.value = f"RÉSUMÉ - {collaborator.upper()}"
            title_cell.font = Font(size=12, bold=True)
            title_cell.alignment = center_align
            title_cell.border = thin_border

            # Summary data
            summary_data = [
                ("CTJ PA Total", total_pa),
                ("CTJ CM Total", total_cm),
                ("CTJ Combiné Total", total_combined),
                ("Jours dans le mois", days_in_month),
                ("Moyenne PA/jour", round(total_pa / days_in_month, 1) if days_in_month > 0 else 0),
                ("Moyenne CM/jour", round(total_cm / days_in_month, 1) if days_in_month > 0 else 0),
                ("Moyenne Combinée/jour", round(total_combined / days_in_month, 1) if days_in_month > 0 else 0)
            ]

            for i, (label, value) in enumerate(summary_data):
                row = summary_row + 1 + i

                # Label
                label_cell = combined_sheet.cell(row=row, column=1)
                label_cell.value = label
                label_cell.font = data_font
                label_cell.border = thin_border

                # Value
                value_cell = combined_sheet.cell(row=row, column=2)
                value_cell.value = value
                value_cell.font = data_font
                value_cell.alignment = center_align
                value_cell.border = thin_border

            self.logger.info(f"Combined summary added: PA={total_pa}, CM={total_cm}, Total={total_combined}")

        except Exception as e:
            self.logger.error(f"Error adding combined summary: {e}")

    def _create_team_combined_ctj_export(self, combined_sheet, ctj_pa_data, ctj_cm_data, month_name, year,
                               header_font, data_font, header_fill, pa_fill, cm_fill,
                               weekend_fill, zero_fill, center_align, thin_border,
                               collaborator_font=None, total_font=None, separator_fill=None, thick_border=None):
        """Create team combined CTJ export with PA and CM data for all collaborators."""
        try:
            from datetime import datetime
            import calendar
            from openpyxl.styles import Font, PatternFill

            # Set default values for optional parameters
            if collaborator_font is None:
                collaborator_font = Font(size=11, bold=True)
            if total_font is None:
                total_font = Font(size=10, bold=True)
            if separator_fill is None:
                separator_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            if thick_border is None:
                thick_border = thin_border

            self.logger.info(f"Creating team combined CTJ export with enhanced formatting")

            # Get month info
            month_names = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                          "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
            month_num = month_names.index(month_name) + 1
            year_num = int(year)
            days_in_month = calendar.monthrange(year_num, month_num)[1]

            # Extract collaborators from both datasets
            collaborators_pa = set()
            collaborators_cm = set()

            if ctj_pa_data and len(ctj_pa_data) > 0:
                # PA data format: list of daily records with collaborator columns
                first_record = ctj_pa_data[0]
                for key in first_record.keys():
                    if key not in ['Date', 'Total']:
                        collaborators_pa.add(key)

            if ctj_cm_data and len(ctj_cm_data) > 0:
                # CM data format: list of daily records with collaborator columns
                first_record = ctj_cm_data[0]
                for key in first_record.keys():
                    if key not in ['Date', 'Total']:
                        collaborators_cm.add(key)

            # Combine all collaborators
            all_collaborators = sorted(list(collaborators_pa.union(collaborators_cm)))

            if not all_collaborators:
                self.logger.warning("No collaborators found in team data")
                return False

            self.logger.info(f"Found collaborators: PA={list(collaborators_pa)}, CM={list(collaborators_cm)}, Combined={all_collaborators}")

            current_row = 4

            # For each collaborator, add PA and CM rows with separator
            for i, collaborator in enumerate(all_collaborators):
                # Add separator line between collaborators (except for the first one)
                if i > 0:
                    # Add empty row for visual separation with enhanced styling
                    current_row += 1
                    # Create a subtle separator line with enhanced formatting
                    separator_cell = combined_sheet.cell(row=current_row, column=1)
                    separator_cell.value = ""
                    # Apply enhanced separator styling to the entire row
                    for col in range(1, days_in_month + 3):  # Include all columns
                        sep_cell = combined_sheet.cell(row=current_row, column=col)
                        sep_cell.fill = separator_fill
                        sep_cell.border = thin_border
                    current_row += 1

                # CTJ PA row for this collaborator with enhanced styling
                combined_sheet.cell(row=current_row, column=1).value = f"{collaborator} - PA"
                combined_sheet.cell(row=current_row, column=1).font = collaborator_font  # Use enhanced collaborator font
                combined_sheet.cell(row=current_row, column=1).fill = pa_fill
                combined_sheet.cell(row=current_row, column=1).border = thick_border  # Use thicker border for collaborator names

                total_pa = 0
                for day in range(1, days_in_month + 1):
                    col = day + 1
                    pa_value = 0

                    # Find PA value for this day and collaborator
                    if ctj_pa_data and day <= len(ctj_pa_data):
                        day_record = ctj_pa_data[day - 1]
                        pa_value = day_record.get(collaborator, 0)

                    total_pa += pa_value

                    cell = combined_sheet.cell(row=current_row, column=col)
                    cell.value = pa_value if pa_value > 0 else ""
                    cell.font = data_font
                    cell.alignment = center_align
                    cell.border = thin_border
                    cell.fill = pa_fill

                    # Apply weekend formatting
                    date_obj = datetime(year_num, month_num, day)
                    if date_obj.weekday() >= 5:  # Weekend
                        cell.fill = weekend_fill
                    elif pa_value == 0:
                        cell.fill = zero_fill

                # Total PA column with enhanced styling
                total_cell = combined_sheet.cell(row=current_row, column=days_in_month + 2)
                total_cell.value = total_pa
                total_cell.font = total_font  # Use enhanced total font
                total_cell.alignment = center_align
                total_cell.border = thick_border  # Use thicker border for totals
                total_cell.fill = pa_fill

                current_row += 1

                # CTJ CM row for this collaborator with enhanced styling
                combined_sheet.cell(row=current_row, column=1).value = f"{collaborator} - CM"
                combined_sheet.cell(row=current_row, column=1).font = collaborator_font  # Use enhanced collaborator font
                combined_sheet.cell(row=current_row, column=1).fill = cm_fill
                combined_sheet.cell(row=current_row, column=1).border = thick_border  # Use thicker border for collaborator names

                total_cm = 0
                for day in range(1, days_in_month + 1):
                    col = day + 1
                    cm_value = 0

                    # Find CM value for this day and collaborator
                    if ctj_cm_data and day <= len(ctj_cm_data):
                        day_record = ctj_cm_data[day - 1]
                        cm_value = day_record.get(collaborator, 0)

                    total_cm += cm_value

                    cell = combined_sheet.cell(row=current_row, column=col)
                    cell.value = cm_value if cm_value > 0 else ""
                    cell.font = data_font
                    cell.alignment = center_align
                    cell.border = thin_border
                    cell.fill = cm_fill

                    # Apply weekend formatting
                    date_obj = datetime(year_num, month_num, day)
                    if date_obj.weekday() >= 5:  # Weekend
                        cell.fill = weekend_fill
                    elif cm_value == 0:
                        cell.fill = zero_fill

                # Total CM column with enhanced styling
                total_cell = combined_sheet.cell(row=current_row, column=days_in_month + 2)
                total_cell.value = total_cm
                total_cell.font = total_font  # Use enhanced total font
                total_cell.alignment = center_align
                total_cell.border = thick_border  # Use thicker border for totals
                total_cell.fill = cm_fill

                current_row += 1

            # Add team totals with enhanced spacing
            current_row += 2  # Add extra space before team totals

            # Team PA Total
            combined_sheet.cell(row=current_row, column=1).value = "ÉQUIPE TOTAL PA"
            combined_sheet.cell(row=current_row, column=1).font = Font(bold=True)
            combined_sheet.cell(row=current_row, column=1).fill = header_fill
            combined_sheet.cell(row=current_row, column=1).border = thin_border

            team_total_pa = 0
            for day in range(1, days_in_month + 1):
                col = day + 1
                day_total_pa = 0

                if ctj_pa_data and day <= len(ctj_pa_data):
                    day_record = ctj_pa_data[day - 1]
                    day_total_pa = day_record.get('Total', 0)

                team_total_pa += day_total_pa

                cell = combined_sheet.cell(row=current_row, column=col)
                cell.value = day_total_pa if day_total_pa > 0 else ""
                cell.font = Font(bold=True)
                cell.alignment = center_align
                cell.border = thin_border
                cell.fill = header_fill

            # Team PA Total column
            total_cell = combined_sheet.cell(row=current_row, column=days_in_month + 2)
            total_cell.value = team_total_pa
            total_cell.font = Font(bold=True)
            total_cell.alignment = center_align
            total_cell.border = thin_border
            total_cell.fill = header_fill

            current_row += 1

            # Team CM Total
            combined_sheet.cell(row=current_row, column=1).value = "ÉQUIPE TOTAL CM"
            combined_sheet.cell(row=current_row, column=1).font = Font(bold=True)
            combined_sheet.cell(row=current_row, column=1).fill = header_fill
            combined_sheet.cell(row=current_row, column=1).border = thin_border

            team_total_cm = 0
            for day in range(1, days_in_month + 1):
                col = day + 1
                day_total_cm = 0

                if ctj_cm_data and day <= len(ctj_cm_data):
                    day_record = ctj_cm_data[day - 1]
                    day_total_cm = day_record.get('Total', 0)

                team_total_cm += day_total_cm

                cell = combined_sheet.cell(row=current_row, column=col)
                cell.value = day_total_cm if day_total_cm > 0 else ""
                cell.font = Font(bold=True)
                cell.alignment = center_align
                cell.border = thin_border
                cell.fill = header_fill

            # Team CM Total column
            total_cell = combined_sheet.cell(row=current_row, column=days_in_month + 2)
            total_cell.value = team_total_cm
            total_cell.font = Font(bold=True)
            total_cell.alignment = center_align
            total_cell.border = thin_border
            total_cell.fill = header_fill

            current_row += 1

            # Team Combined Total
            combined_sheet.cell(row=current_row, column=1).value = "ÉQUIPE TOTAL COMBINÉ"
            combined_sheet.cell(row=current_row, column=1).font = Font(bold=True)
            combined_sheet.cell(row=current_row, column=1).fill = header_fill
            combined_sheet.cell(row=current_row, column=1).border = thin_border

            team_total_combined = 0
            for day in range(1, days_in_month + 1):
                col = day + 1
                day_total_pa = 0
                day_total_cm = 0

                if ctj_pa_data and day <= len(ctj_pa_data):
                    day_record = ctj_pa_data[day - 1]
                    day_total_pa = day_record.get('Total', 0)

                if ctj_cm_data and day <= len(ctj_cm_data):
                    day_record = ctj_cm_data[day - 1]
                    day_total_cm = day_record.get('Total', 0)

                day_combined = day_total_pa + day_total_cm
                team_total_combined += day_combined

                cell = combined_sheet.cell(row=current_row, column=col)
                cell.value = day_combined if day_combined > 0 else ""
                cell.font = Font(bold=True)
                cell.alignment = center_align
                cell.border = thin_border
                cell.fill = header_fill

            # Team Combined Total column
            total_cell = combined_sheet.cell(row=current_row, column=days_in_month + 2)
            total_cell.value = team_total_combined
            total_cell.font = Font(bold=True)
            total_cell.alignment = center_align
            total_cell.border = thin_border
            total_cell.fill = header_fill

            # Auto-fit columns
            for col in range(1, days_in_month + 3):
                from openpyxl.utils import get_column_letter
                column_letter = get_column_letter(col)
                max_width = 0
                for row in range(1, current_row + 5):
                    try:
                        cell_value = combined_sheet[f"{column_letter}{row}"].value
                        if cell_value:
                            cell_length = len(str(cell_value))
                            max_width = max(max_width, cell_length)
                    except:
                        continue
                column_width = min(max(max_width + 1, 8), 20)
                combined_sheet.column_dimensions[column_letter].width = column_width

            self.logger.info(f"Team combined CTJ export completed: {len(all_collaborators)} collaborators, PA={team_total_pa}, CM={team_total_cm}, Total={team_total_combined}")
            return True

        except Exception as e:
            self.logger.error(f"Error creating team combined CTJ export: {e}")
            return False

    def _optimize_column_widths(self, sheet, days_in_month):
        """Optimize column widths for better presentation."""
        try:
            from openpyxl.utils import get_column_letter

            # Set specific widths for different column types
            for col in range(1, days_in_month + 3):
                column_letter = get_column_letter(col)

                if col == 1:  # Collaborator name column
                    sheet.column_dimensions[column_letter].width = 25
                elif col == days_in_month + 2:  # Total column
                    sheet.column_dimensions[column_letter].width = 12
                else:  # Date columns
                    sheet.column_dimensions[column_letter].width = 8

            self.logger.info("Column widths optimized successfully")

        except Exception as e:
            self.logger.warning(f"Could not optimize column widths: {e}")

    def _add_enhanced_formatting(self, sheet, start_row, end_row, days_in_month):
        """Add enhanced formatting to improve readability."""
        try:
            from openpyxl.styles import PatternFill

            # Add alternating row colors for better readability
            light_fill = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")

            for row in range(start_row, end_row + 1):
                if row % 2 == 0:  # Even rows
                    for col in range(1, days_in_month + 3):
                        cell = sheet.cell(row=row, column=col)
                        if cell.fill.start_color.rgb == "00000000":  # Only if no fill is set
                            cell.fill = light_fill

            self.logger.info("Enhanced formatting applied successfully")

        except Exception as e:
            self.logger.warning(f"Could not apply enhanced formatting: {e}")

    def _extract_monthly_motifs_data(self, month_name, year):
        """Extract monthly motifs data from CM, PA and RIP sheets for statistics."""
        try:
            from datetime import datetime
            import calendar
            from utils.lazy_imports import get_pandas

            # Convert month name to number
            month_names = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                          "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
            month_num = month_names.index(month_name) + 1
            year_num = int(year)

            motifs_data = []

            # 1. Extract CM motifs from "Traitement CMS Adr" sheet
            if 'Traitement CMS Adr' in self.global_suivi_data:
                df_cms = self.global_suivi_data['Traitement CMS Adr']
                if not df_cms.empty:
                    motifs_data.extend(self._extract_motifs_from_cms(df_cms, month_num, year_num))

            # 2. Extract PA motifs from "Traitement PA" sheet
            if 'Traitement PA' in self.global_suivi_data:
                df_pa = self.global_suivi_data['Traitement PA']
                if not df_pa.empty:
                    motifs_data.extend(self._extract_motifs_from_pa(df_pa, month_num, year_num))

            # 3. Extract RIP motifs from "Traitement RIP" sheet
            if 'Traitement RIP' in self.global_suivi_data:
                df_rip = self.global_suivi_data['Traitement RIP']
                if not df_rip.empty:
                    motifs_data.extend(self._extract_motifs_from_rip(df_rip, month_num, year_num))

            self.logger.info(f"Extracted {len(motifs_data)} motifs records for {month_name} {year}")
            return motifs_data

        except Exception as e:
            self.logger.error(f"Error extracting monthly motifs data: {e}")
            return []

    def _extract_motifs_from_cms(self, df_cms, month_num, year_num):
        """Extract motifs data from CM (Traitement CMS Adr) sheet."""
        try:
            from datetime import datetime

            motifs_data = []

            # Check if required columns exist
            if len(df_cms.columns) < 8:  # Need at least 8 columns (A-H)
                self.logger.warning(f"CM sheet: Not enough columns (found {len(df_cms.columns)}, need at least 8)")
                return motifs_data

            # Column D: Motif Voie, Column G: Date traitement (index 3 and 6)
            motif_column = df_cms.columns[3]  # Column D (Motif Voie)
            date_column = df_cms.columns[6]   # Column G (Date traitement) - shifted due to Motif Voie in D
            collaborator_column = 'Collaborateur' if 'Collaborateur' in df_cms.columns else None

            self.logger.debug(f"CM motifs extraction: Using columns '{motif_column}' and '{date_column}'")

            for index, row in df_cms.iterrows():
                try:
                    # Get motif value and normalize to uppercase
                    motif_value = row.get(motif_column, '')
                    if not motif_value or str(motif_value).strip() == '' or str(motif_value).lower() == 'nan':
                        continue
                    motif_value = str(motif_value).strip().upper()  # Normalize to uppercase

                    # Get date value
                    date_value = row.get(date_column, '')
                    if not date_value or str(date_value).strip() == '' or str(date_value).lower() == 'nan':
                        continue

                    # Parse date (try multiple formats)
                    date_obj = None
                    for date_format in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
                        try:
                            date_obj = datetime.strptime(str(date_value), date_format)
                            break
                        except ValueError:
                            continue

                    if not date_obj:
                        continue

                    # Check if date is in selected month/year
                    if date_obj.month == month_num and date_obj.year == year_num:
                        collaborator = row.get(collaborator_column, 'Non spécifié') if collaborator_column else 'Non spécifié'

                        motifs_data.append({
                            'Type': 'CM',
                            'Motif': motif_value,  # Already normalized to uppercase
                            'Date': date_obj.strftime('%Y-%m-%d'),
                            'Collaborateur': str(collaborator).strip(),
                            'Mois': month_num,
                            'Année': year_num
                        })

                except Exception as e:
                    self.logger.debug(f"Error processing CM row {index}: {e}")
                    continue

            self.logger.info(f"Extracted {len(motifs_data)} CM motifs for {month_num}/{year_num}")
            return motifs_data

        except Exception as e:
            self.logger.error(f"Error extracting CM motifs: {e}")
            return []

    def _extract_motifs_from_pa(self, df_pa, month_num, year_num):
        """Extract motifs data from PA (Traitement PA) sheet."""
        try:
            from datetime import datetime

            motifs_data = []

            # Check if required columns exist
            if len(df_pa.columns) < 8:  # Need at least 8 columns (A-H)
                self.logger.warning(f"PA sheet: Not enough columns (found {len(df_pa.columns)}, need at least 8)")
                return motifs_data

            # Column D: Motif, Column G: Date traitement (index 3 and 6)
            motif_column = df_pa.columns[3]  # Column D (Motif)
            date_column = df_pa.columns[6]   # Column G (Date traitement)
            collaborator_column = 'Collaborateur' if 'Collaborateur' in df_pa.columns else None

            self.logger.debug(f"PA motifs extraction: Using columns '{motif_column}' and '{date_column}'")

            for index, row in df_pa.iterrows():
                try:
                    # Get motif value and normalize to uppercase
                    motif_value = row.get(motif_column, '')
                    if not motif_value or str(motif_value).strip() == '' or str(motif_value).lower() == 'nan':
                        continue
                    motif_value = str(motif_value).strip().upper()  # Normalize to uppercase

                    # Get date value
                    date_value = row.get(date_column, '')
                    if not date_value or str(date_value).strip() == '' or str(date_value).lower() == 'nan':
                        continue

                    # Parse date (try multiple formats)
                    date_obj = None
                    for date_format in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
                        try:
                            date_obj = datetime.strptime(str(date_value), date_format)
                            break
                        except ValueError:
                            continue

                    if not date_obj:
                        continue

                    # Check if date is in selected month/year
                    if date_obj.month == month_num and date_obj.year == year_num:
                        collaborator = row.get(collaborator_column, 'Non spécifié') if collaborator_column else 'Non spécifié'

                        motifs_data.append({
                            'Type': 'PA',
                            'Motif': motif_value,  # Already normalized to uppercase
                            'Date': date_obj.strftime('%Y-%m-%d'),
                            'Collaborateur': str(collaborator).strip(),
                            'Mois': month_num,
                            'Année': year_num
                        })

                except Exception as e:
                    self.logger.debug(f"Error processing PA row {index}: {e}")
                    continue

            self.logger.info(f"Extracted {len(motifs_data)} PA motifs for {month_num}/{year_num}")
            return motifs_data

        except Exception as e:
            self.logger.error(f"Error extracting PA motifs: {e}")
            return []

    def _extract_motifs_from_rip(self, df_rip, month_num, year_num):
        """Extract motifs data from RIP (Traitement RIP) sheet."""
        try:
            from datetime import datetime

            motifs_data = []

            # Check if required columns exist
            if len(df_rip.columns) < 9:  # Need at least 9 columns (A-I)
                self.logger.warning(f"RIP sheet: Not enough columns (found {len(df_rip.columns)}, need at least 9)")
                return motifs_data

            # Column E: Acte de traitement, Column H: Date de traitement (index 4 and 7)
            motif_column = df_rip.columns[4]  # Column E (Acte de traitement)
            date_column = df_rip.columns[7]   # Column H (Date de traitement)
            collaborator_column = 'Collaborateur' if 'Collaborateur' in df_rip.columns else None

            self.logger.debug(f"RIP motifs extraction: Using columns '{motif_column}' and '{date_column}'")

            for index, row in df_rip.iterrows():
                try:
                    # Get motif value and normalize to uppercase
                    motif_value = row.get(motif_column, '')
                    if not motif_value or str(motif_value).strip() == '' or str(motif_value).lower() == 'nan':
                        continue
                    motif_value = str(motif_value).strip().upper()  # Normalize to uppercase

                    # Get date value
                    date_value = row.get(date_column, '')
                    if not date_value or str(date_value).strip() == '' or str(date_value).lower() == 'nan':
                        continue

                    # Parse date (try multiple formats)
                    date_obj = None
                    for date_format in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
                        try:
                            date_obj = datetime.strptime(str(date_value), date_format)
                            break
                        except ValueError:
                            continue

                    if not date_obj:
                        continue

                    # Check if date is in selected month/year
                    if date_obj.month == month_num and date_obj.year == year_num:
                        collaborator = row.get(collaborator_column, 'Non spécifié') if collaborator_column else 'Non spécifié'

                        motifs_data.append({
                            'Type': 'RIP',
                            'Motif': motif_value,  # Already normalized to uppercase
                            'Date': date_obj.strftime('%Y-%m-%d'),
                            'Collaborateur': str(collaborator).strip(),
                            'Mois': month_num,
                            'Année': year_num
                        })

                except Exception as e:
                    self.logger.debug(f"Error processing RIP row {index}: {e}")
                    continue

            self.logger.info(f"Extracted {len(motifs_data)} RIP motifs for {month_num}/{year_num}")
            return motifs_data

        except Exception as e:
            self.logger.error(f"Error extracting RIP motifs: {e}")
            return []

    def _create_motifs_statistics_sheet(self, workbook, motifs_data, month_name, year):
        """Create the 'Statistiques Motifs Mensuels' sheet with detailed motifs analysis."""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from collections import defaultdict, Counter

            # Create the sheet
            motifs_sheet = workbook.create_sheet(title="Statistiques Motifs Mensuels")

            # Define styles
            header_font = Font(size=12, bold=True, color="FFFFFF", name="Calibri")
            subheader_font = Font(size=11, bold=True, name="Calibri")
            data_font = Font(size=10, name="Calibri")

            header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            cm_fill = PatternFill(start_color="E1F5FE", end_color="E1F5FE", fill_type="solid")
            pa_fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
            rip_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")

            center_align = Alignment(horizontal="center", vertical="center")
            left_align = Alignment(horizontal="left", vertical="center")

            thin_border = Border(
                left=Side(style='thin', color="666666"),
                right=Side(style='thin', color="666666"),
                top=Side(style='thin', color="666666"),
                bottom=Side(style='thin', color="666666")
            )

            # Title
            motifs_sheet.cell(row=1, column=1).value = f"Statistiques Motifs Mensuels - {month_name} {year}"
            motifs_sheet.cell(row=1, column=1).font = Font(size=14, bold=True, name="Calibri")
            motifs_sheet.merge_cells('A1:F1')

            current_row = 3

            # 1. Summary by Type
            self._create_motifs_summary_section(motifs_sheet, motifs_data, current_row, header_font, subheader_font,
                                              data_font, header_fill, center_align, left_align, thin_border)
            current_row += 10

            # 2. Separate detailed sections for each type
            # CM Section
            cm_data = [item for item in motifs_data if item['Type'] == 'CM']
            if cm_data:
                current_row = self._create_motifs_type_section(motifs_sheet, cm_data, 'CM', 'Communes Orange',
                                                             current_row, header_font, subheader_font, data_font,
                                                             header_fill, cm_fill, center_align, left_align, thin_border)
                current_row += 3

            # PA Section
            pa_data = [item for item in motifs_data if item['Type'] == 'PA']
            if pa_data:
                current_row = self._create_motifs_type_section(motifs_sheet, pa_data, 'PA', 'Plan Adressage',
                                                             current_row, header_font, subheader_font, data_font,
                                                             header_fill, pa_fill, center_align, left_align, thin_border)
                current_row += 3

            # RIP Section
            rip_data = [item for item in motifs_data if item['Type'] == 'RIP']
            if rip_data:
                current_row = self._create_motifs_type_section(motifs_sheet, rip_data, 'RIP', 'Communes RIP',
                                                             current_row, header_font, subheader_font, data_font,
                                                             header_fill, rip_fill, center_align, left_align, thin_border)
                current_row += 8  # More space before collaborator analysis

            # 3. Collaborator analysis (positioned at the bottom with more spacing)
            self._create_motifs_collaborator_section(motifs_sheet, motifs_data, current_row, header_font, subheader_font,
                                                   data_font, header_fill, center_align, left_align, thin_border)

            # Optimize column widths
            self._optimize_motifs_sheet_columns(motifs_sheet)

            self.logger.info(f"Created motifs statistics sheet with {len(motifs_data)} records")
            return True

        except Exception as e:
            self.logger.error(f"Error creating motifs statistics sheet: {e}")
            return False

    def _create_motifs_summary_section(self, sheet, motifs_data, start_row, header_font, subheader_font,
                                     data_font, header_fill, center_align, left_align, thin_border):
        """Create summary section with counts by type."""
        try:
            from collections import Counter

            # Section title
            sheet.cell(row=start_row, column=1).value = "📊 Résumé par Type de Traitement"
            sheet.cell(row=start_row, column=1).font = subheader_font
            sheet.merge_cells(f'A{start_row}:D{start_row}')

            # Headers
            headers = ["Type", "Nombre de Motifs", "Pourcentage", "Motifs Uniques"]
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=start_row + 2, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border

            # Calculate statistics
            type_counts = Counter([item['Type'] for item in motifs_data])
            total_records = len(motifs_data)

            type_motifs = {}
            for item in motifs_data:
                if item['Type'] not in type_motifs:
                    type_motifs[item['Type']] = set()
                type_motifs[item['Type']].add(item['Motif'])

            # Data rows
            row = start_row + 3
            for type_name in ['CM', 'PA', 'RIP']:
                count = type_counts.get(type_name, 0)
                percentage = (count / total_records * 100) if total_records > 0 else 0
                unique_motifs = len(type_motifs.get(type_name, set()))

                sheet.cell(row=row, column=1).value = type_name
                sheet.cell(row=row, column=1).font = data_font
                sheet.cell(row=row, column=1).alignment = center_align
                sheet.cell(row=row, column=1).border = thin_border

                sheet.cell(row=row, column=2).value = count
                sheet.cell(row=row, column=2).font = data_font
                sheet.cell(row=row, column=2).alignment = center_align
                sheet.cell(row=row, column=2).border = thin_border

                sheet.cell(row=row, column=3).value = f"{percentage:.1f}%"
                sheet.cell(row=row, column=3).font = data_font
                sheet.cell(row=row, column=3).alignment = center_align
                sheet.cell(row=row, column=3).border = thin_border

                sheet.cell(row=row, column=4).value = unique_motifs
                sheet.cell(row=row, column=4).font = data_font
                sheet.cell(row=row, column=4).alignment = center_align
                sheet.cell(row=row, column=4).border = thin_border

                row += 1

            # Total row
            total_font = Font(size=10, bold=True, name="Calibri")

            sheet.cell(row=row, column=1).value = "TOTAL"
            sheet.cell(row=row, column=1).font = total_font
            sheet.cell(row=row, column=1).alignment = center_align
            sheet.cell(row=row, column=1).border = thin_border

            sheet.cell(row=row, column=2).value = total_records
            sheet.cell(row=row, column=2).font = total_font
            sheet.cell(row=row, column=2).alignment = center_align
            sheet.cell(row=row, column=2).border = thin_border

            sheet.cell(row=row, column=3).value = "100.0%"
            sheet.cell(row=row, column=3).font = total_font
            sheet.cell(row=row, column=3).alignment = center_align
            sheet.cell(row=row, column=3).border = thin_border

            total_unique = len(set([item['Motif'] for item in motifs_data]))
            sheet.cell(row=row, column=4).value = total_unique
            sheet.cell(row=row, column=4).font = total_font
            sheet.cell(row=row, column=4).alignment = center_align
            sheet.cell(row=row, column=4).border = thin_border

        except Exception as e:
            self.logger.error(f"Error creating motifs summary section: {e}")

    def _create_motifs_type_section(self, sheet, type_data, type_code, type_name, start_row,
                                  header_font, subheader_font, data_font, header_fill, type_fill,
                                  center_align, left_align, thin_border):
        """Create detailed section for a specific type (CM, PA, or RIP)."""
        try:
            from collections import Counter
            from openpyxl.styles import Font

            if not type_data:
                return start_row

            # Section title with emoji
            emoji_map = {'CM': '🏠', 'PA': '📍', 'RIP': '🔧'}
            emoji = emoji_map.get(type_code, '📊')

            sheet.cell(row=start_row, column=1).value = f"{emoji} {type_name} - Détail des Motifs"
            sheet.cell(row=start_row, column=1).font = subheader_font
            sheet.merge_cells(f'A{start_row}:E{start_row}')

            # Headers
            headers = ["Motif", "Occurrences", "Pourcentage", "Premier Usage", "Dernier Usage"]
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=start_row + 2, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border

            # Group data by motif
            motif_data = {}
            for item in type_data:
                motif = item['Motif']
                if motif not in motif_data:
                    motif_data[motif] = []
                motif_data[motif].append(item['Date'])

            # Sort and create rows
            row = start_row + 3
            total_records = len(type_data)

            for motif in sorted(motif_data.keys()):
                dates = motif_data[motif]
                count = len(dates)
                percentage = (count / total_records * 100) if total_records > 0 else 0
                first_date = min(dates)
                last_date = max(dates)

                sheet.cell(row=row, column=1).value = motif
                sheet.cell(row=row, column=1).font = data_font
                sheet.cell(row=row, column=1).alignment = left_align
                sheet.cell(row=row, column=1).border = thin_border
                sheet.cell(row=row, column=1).fill = type_fill

                sheet.cell(row=row, column=2).value = count
                sheet.cell(row=row, column=2).font = data_font
                sheet.cell(row=row, column=2).alignment = center_align
                sheet.cell(row=row, column=2).border = thin_border

                sheet.cell(row=row, column=3).value = f"{percentage:.1f}%"
                sheet.cell(row=row, column=3).font = data_font
                sheet.cell(row=row, column=3).alignment = center_align
                sheet.cell(row=row, column=3).border = thin_border

                sheet.cell(row=row, column=4).value = first_date
                sheet.cell(row=row, column=4).font = data_font
                sheet.cell(row=row, column=4).alignment = center_align
                sheet.cell(row=row, column=4).border = thin_border

                sheet.cell(row=row, column=5).value = last_date
                sheet.cell(row=row, column=5).font = data_font
                sheet.cell(row=row, column=5).alignment = center_align
                sheet.cell(row=row, column=5).border = thin_border

                row += 1

            # Add total row for this type
            total_font = Font(size=10, bold=True, name="Calibri")

            sheet.cell(row=row, column=1).value = f"TOTAL {type_code}"
            sheet.cell(row=row, column=1).font = total_font
            sheet.cell(row=row, column=1).alignment = left_align
            sheet.cell(row=row, column=1).border = thin_border
            sheet.cell(row=row, column=1).fill = type_fill

            sheet.cell(row=row, column=2).value = total_records
            sheet.cell(row=row, column=2).font = total_font
            sheet.cell(row=row, column=2).alignment = center_align
            sheet.cell(row=row, column=2).border = thin_border

            sheet.cell(row=row, column=3).value = "100.0%"
            sheet.cell(row=row, column=3).font = total_font
            sheet.cell(row=row, column=3).alignment = center_align
            sheet.cell(row=row, column=3).border = thin_border

            unique_motifs = len(motif_data)
            sheet.cell(row=row, column=4).value = f"{unique_motifs} motifs"
            sheet.cell(row=row, column=4).font = total_font
            sheet.cell(row=row, column=4).alignment = center_align
            sheet.cell(row=row, column=4).border = thin_border

            return row + 1  # Return next available row

        except Exception as e:
            self.logger.error(f"Error creating motifs type section for {type_code}: {e}")
            return start_row

    def _create_motifs_collaborator_section(self, sheet, motifs_data, start_row, header_font, subheader_font,
                                          data_font, header_fill, center_align, left_align, thin_border):
        """Create collaborator analysis section."""
        try:
            from collections import defaultdict, Counter
            from openpyxl.styles import Font

            # Section title
            sheet.cell(row=start_row, column=1).value = "👥 Analyse par Collaborateur"
            sheet.cell(row=start_row, column=1).font = subheader_font
            sheet.merge_cells(f'A{start_row}:E{start_row}')

            # Headers
            headers = ["Collaborateur", "Total Traitements", "CM", "PA", "RIP"]
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=start_row + 2, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border

            # Group data by collaborator
            collab_data = defaultdict(lambda: {'CM': 0, 'PA': 0, 'RIP': 0, 'Total': 0})

            for item in motifs_data:
                collaborator = item['Collaborateur']
                type_name = item['Type']
                collab_data[collaborator][type_name] += 1
                collab_data[collaborator]['Total'] += 1

            # Create rows
            row = start_row + 3
            for collaborator in sorted(collab_data.keys()):
                data = collab_data[collaborator]

                sheet.cell(row=row, column=1).value = collaborator
                sheet.cell(row=row, column=1).font = data_font
                sheet.cell(row=row, column=1).alignment = left_align
                sheet.cell(row=row, column=1).border = thin_border

                sheet.cell(row=row, column=2).value = data['Total']
                sheet.cell(row=row, column=2).font = data_font
                sheet.cell(row=row, column=2).alignment = center_align
                sheet.cell(row=row, column=2).border = thin_border

                sheet.cell(row=row, column=3).value = data['CM']
                sheet.cell(row=row, column=3).font = data_font
                sheet.cell(row=row, column=3).alignment = center_align
                sheet.cell(row=row, column=3).border = thin_border

                sheet.cell(row=row, column=4).value = data['PA']
                sheet.cell(row=row, column=4).font = data_font
                sheet.cell(row=row, column=4).alignment = center_align
                sheet.cell(row=row, column=4).border = thin_border

                sheet.cell(row=row, column=5).value = data['RIP']
                sheet.cell(row=row, column=5).font = data_font
                sheet.cell(row=row, column=5).alignment = center_align
                sheet.cell(row=row, column=5).border = thin_border

                row += 1

        except Exception as e:
            self.logger.error(f"Error creating motifs collaborator section: {e}")

    def _optimize_motifs_sheet_columns(self, sheet):
        """Optimize column widths for motifs statistics sheet."""
        try:
            from openpyxl.utils import get_column_letter

            # Set specific widths for different columns
            column_widths = {
                'A': 15,  # Type/Collaborateur
                'B': 30,  # Motif
                'C': 12,  # Occurrences/Total
                'D': 12,  # Pourcentage/CM
                'E': 12,  # Premier Usage/PA
                'F': 12   # Dernier Usage/RIP
            }

            for col_letter, width in column_widths.items():
                sheet.column_dimensions[col_letter].width = width

            self.logger.debug("Motifs sheet column widths optimized")

        except Exception as e:
            self.logger.warning(f"Could not optimize motifs sheet column widths: {e}")

    def _create_streamlined_daily_ctj_cm_sheet(self, workbook, ctj_cm_data, collaborator, month_name, year):
        """Create streamlined daily CTJ CM sheet with only essential data."""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from datetime import datetime
            import calendar

            # Create streamlined daily sheet
            daily_sheet = workbook.create_sheet(title="CTJ CM Quotidien")

            # Define styles
            header_font = Font(size=11, bold=True, color="FFFFFF")
            data_font = Font(size=10)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            weekend_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            zero_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")

            center_align = Alignment(horizontal="center", vertical="center")

            # Month abbreviations for headers
            month_abbrev = {
                1: "Jan", 2: "Fév", 3: "Mar", 4: "Avr", 5: "Mai", 6: "Jun",
                7: "Jul", 8: "Aoû", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Déc"
            }

            if collaborator == "Toute l'équipe":
                # Team export: stack individual collaborators vertically
                return self._create_team_stacked_ctj_cm_export(daily_sheet, ctj_cm_data, month_name, year,
                                                      header_font, data_font, header_fill,
                                                      weekend_fill, zero_fill, center_align)
            else:
                # Individual collaborator export
                return self._create_individual_ctj_cm_export(daily_sheet, ctj_cm_data, collaborator, month_name, year,
                                                    header_font, data_font, header_fill,
                                                    weekend_fill, zero_fill, center_align)

        except Exception as e:
            self.logger.error(f"Error creating streamlined daily CTJ CM sheet: {e}")
            import traceback
            self.logger.error(f"Full traceback: {traceback.format_exc()}")
            return False

    def _create_individual_ctj_cm_export(self, daily_sheet, ctj_cm_data, collaborator, month_name, year,
                                header_font, data_font, header_fill, weekend_fill, zero_fill, center_align):
        """Create individual collaborator CTJ CM export with only CTJ CM data."""
        try:
            from datetime import datetime
            import calendar
            from openpyxl.styles import Font, Border, Side, PatternFill

            self.logger.info(f"Creating individual CTJ CM export for {collaborator}, data length: {len(ctj_cm_data) if ctj_cm_data else 0}")

            # Define border style
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Get month info
            month_names = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                          "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
            month_num = month_names.index(month_name) + 1
            year_num = int(year)
            days_in_month = calendar.monthrange(year_num, month_num)[1]

            # Month abbreviations for headers
            month_abbrev = {
                1: "Jan", 2: "Fév", 3: "Mar", 4: "Avr", 5: "Mai", 6: "Jun",
                7: "Jul", 8: "Aoû", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Déc"
            }

            # Title
            daily_sheet.merge_cells('A1:AH1')
            title_cell = daily_sheet['A1']
            title_cell.value = f"CTJ CM QUOTIDIEN - {collaborator.upper()} - {month_name} {year}"
            title_cell.font = Font(size=14, bold=True, color="FFFFFF")
            title_cell.fill = header_fill
            title_cell.alignment = center_align

            # Headers with new date format
            headers = ['Collaborateur']
            for day in range(1, days_in_month + 1):
                headers.append(f"{day} {month_abbrev[month_num]}")

            # Write headers with borders
            for col, header in enumerate(headers, 1):
                cell = daily_sheet.cell(row=3, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border

            # Write collaborator name with border
            daily_sheet.cell(row=4, column=1).value = collaborator
            daily_sheet.cell(row=4, column=1).font = Font(bold=True)
            daily_sheet.cell(row=4, column=1).border = thin_border

            # Write CTJ CM data and calculate cumulative
            total_ctj_cm_written = 0
            cumulative_ctj_cm = 0
            daily_values = []

            for day in range(1, days_in_month + 1):
                col = day + 1
                ctj_cm_value = 0

                if day <= len(ctj_cm_data):
                    row_data = ctj_cm_data[day - 1]
                    ctj_cm_value = row_data.get('CTJ_CM', 0)
                    if ctj_cm_value > 0:
                        total_ctj_cm_written += ctj_cm_value

                # Store daily value for calculations
                daily_values.append(ctj_cm_value)
                cumulative_ctj_cm += ctj_cm_value

                cell = daily_sheet.cell(row=4, column=col)
                cell.value = ctj_cm_value if ctj_cm_value > 0 else ""
                cell.font = data_font
                cell.alignment = center_align
                cell.border = thin_border

                # Apply conditional formatting
                date_obj = datetime(year_num, month_num, day)
                is_weekend = date_obj.weekday() >= 5

                if is_weekend:
                    cell.fill = weekend_fill
                elif ctj_cm_value == 0:
                    cell.fill = zero_fill

            # Add cumulative CTJ CM row
            daily_sheet.cell(row=5, column=1).value = "Cumulé CTJ CM"
            daily_sheet.cell(row=5, column=1).font = Font(bold=True, italic=True)
            daily_sheet.cell(row=5, column=1).border = thin_border
            daily_sheet.cell(row=5, column=1).fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")  # Light blue

            running_total = 0
            for day in range(1, days_in_month + 1):
                col = day + 1
                running_total += daily_values[day - 1]

                cell = daily_sheet.cell(row=5, column=col)
                cell.value = running_total if running_total > 0 else ""
                cell.font = Font(italic=True, size=9)
                cell.alignment = center_align
                cell.border = thin_border
                cell.fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")  # Light blue

            # Add global summary section
            self._add_individual_ctj_cm_summary(daily_sheet, daily_values, collaborator, month_name, year,
                                       days_in_month, thin_border, header_font, data_font, center_align)

            self.logger.info(f"Individual CTJ CM export: wrote {total_ctj_cm_written} total CTJ CM values for {collaborator}")

            # Auto-fit all columns based on content
            for col in range(1, days_in_month + 2):
                from openpyxl.utils import get_column_letter
                column_letter = get_column_letter(col)

                # Calculate the maximum width needed for this column
                max_width = 0
                for row in range(1, 15):  # Check more rows including summary
                    try:
                        cell_value = daily_sheet[f"{column_letter}{row}"].value
                        if cell_value:
                            cell_length = len(str(cell_value))
                            max_width = max(max_width, cell_length)
                    except:
                        continue

                # Set column width with reasonable limits
                column_width = min(max(max_width + 1, 8), 15)
                daily_sheet.column_dimensions[column_letter].width = column_width

            return True

        except Exception as e:
            self.logger.error(f"Error creating individual CTJ CM export: {e}")
            return False

    def _create_horizontal_ctj_excel(self, file_path, ctj_data, collaborator, month_name, year):
        """Create streamlined Excel file with only CTJ Quotidien sheet."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from datetime import datetime, timedelta
            import calendar

            # Create workbook
            workbook = openpyxl.Workbook()

            # Get the default sheet to remove it later
            default_sheet = workbook.active

            # Create only the CTJ Quotidien sheet as requested
            success_daily = self._create_streamlined_daily_sheet(workbook, ctj_data, collaborator, month_name, year)

            # Always remove the default sheet if we successfully created our custom sheet
            if success_daily:
                try:
                    workbook.remove(default_sheet)
                    self.logger.info("Default sheet removed successfully")
                except Exception as e:
                    self.logger.warning(f"Could not remove default sheet: {e}")
            else:
                # If creation failed, provide detailed error message
                error_msg = f"Failed to create CTJ export sheet for {collaborator}. Please check if data is available for the selected period."
                self.logger.error(error_msg)
                raise Exception(error_msg)

            # Check file access before saving
            if os.path.exists(file_path):
                access_result = check_file_access(file_path, 'w')
                if not access_result['accessible']:
                    self.logger.warning(f"File access issue during export: {access_result['error_message']}")

                    # Show user-friendly dialog and get retry decision
                    if access_result['error_type'] in ['file_locked', 'permission_denied']:
                        # This is a file-in-use situation, show the custom dialog
                        retry = self._show_file_access_dialog(access_result, file_path)
                        if not retry:
                            raise Exception("Export annulé par l'utilisateur")

                        # User wants to retry, check again
                        retry_access = check_file_access(file_path, 'w')
                        if not retry_access['accessible']:
                            # Still not accessible, raise with user-friendly message
                            raise Exception(f"{retry_access['user_message']}\n\nVeuillez fermer Excel et réessayer l'export.")
                    else:
                        # Other types of errors, raise with user message
                        raise Exception(access_result['user_message'])

            # Save the streamlined workbook
            workbook.save(file_path)

            self.logger.info(f"Streamlined CTJ Excel file created successfully: {file_path}")

        except Exception as e:
            self.logger.error(f"Error creating streamlined CTJ Excel: {e}")
            raise




    def _apply_horizontal_formatting(self, worksheet, num_days, month_num, year, num_collaborators):
        """Apply formatting to horizontal CTJ Excel layout."""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from datetime import datetime
            import calendar

            # Define colors
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            weekend_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Gray
            zero_ctj_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")  # Pink

            # Define fonts
            header_font = Font(bold=True, color="FFFFFF")
            bold_font = Font(bold=True)

            # Define alignment
            center_alignment = Alignment(horizontal="center", vertical="center")

            # Define borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Format header row (row 1 - collaborator name)
            worksheet.cell(row=1, column=1).font = bold_font
            worksheet.cell(row=1, column=1).alignment = center_alignment

            # Format second row (row 2 - "Eléments Traités" and dates)
            worksheet.cell(row=2, column=1, value="Eléments Traités")
            worksheet.cell(row=2, column=1).font = header_font
            worksheet.cell(row=2, column=1).fill = header_fill
            worksheet.cell(row=2, column=1).alignment = center_alignment
            worksheet.cell(row=2, column=1).border = thin_border

            # Format date headers and apply weekend/zero formatting
            for day in range(1, num_days + 1):
                col = day + 1
                date_obj = datetime(year, month_num, day)

                # Header cell formatting
                header_cell = worksheet.cell(row=2, column=col)
                header_cell.font = header_font
                header_cell.fill = header_fill
                header_cell.alignment = center_alignment
                header_cell.border = thin_border

                # Check if weekend (Saturday=5, Sunday=6)
                is_weekend = date_obj.weekday() >= 5

                # Format data cells for this day
                for row in range(3, 3 + num_collaborators):
                    data_cell = worksheet.cell(row=row, column=col)
                    data_cell.alignment = center_alignment
                    data_cell.border = thin_border

                    # Apply weekend formatting
                    if is_weekend:
                        data_cell.fill = weekend_fill
                    # Apply zero CTJ formatting (pink)
                    elif data_cell.value == 0:
                        data_cell.fill = zero_ctj_fill

            # Format collaborator names column
            for row in range(3, 3 + num_collaborators):
                name_cell = worksheet.cell(row=row, column=1)
                name_cell.font = bold_font
                name_cell.alignment = Alignment(horizontal="left", vertical="center")
                name_cell.border = thin_border

            # Auto-adjust column widths
            # Column A (collaborator names) - wider
            worksheet.column_dimensions['A'].width = 20

            # Date columns - narrower
            for day in range(1, num_days + 1):
                col_letter = worksheet.cell(row=2, column=day + 1).column_letter
                worksheet.column_dimensions[col_letter].width = 4

            # Add total row if multiple collaborators
            if num_collaborators > 1:
                total_row = 3 + num_collaborators
                worksheet.cell(row=total_row, column=1, value="TOTAL")
                worksheet.cell(row=total_row, column=1).font = bold_font
                worksheet.cell(row=total_row, column=1).fill = header_fill
                worksheet.cell(row=total_row, column=1).alignment = center_alignment
                worksheet.cell(row=total_row, column=1).border = thin_border

                # Calculate totals for each day
                for day in range(1, num_days + 1):
                    col = day + 1
                    date_obj = datetime(year, month_num, day)
                    is_weekend = date_obj.weekday() >= 5

                    # Sum values from all collaborators for this day
                    total_value = 0
                    for row in range(3, 3 + num_collaborators):
                        cell_value = worksheet.cell(row=row, column=col).value
                        if isinstance(cell_value, (int, float)):
                            total_value += cell_value

                    total_cell = worksheet.cell(row=total_row, column=col)
                    total_cell.value = total_value
                    total_cell.font = bold_font
                    total_cell.alignment = center_alignment
                    total_cell.border = thin_border

                    # Apply formatting
                    if is_weekend:
                        total_cell.fill = weekend_fill
                    elif total_value == 0:
                        total_cell.fill = zero_ctj_fill

        except Exception as e:
            self.logger.warning(f"Could not apply Excel formatting: {e}")
            # Continue without formatting if there's an error







    def _create_streamlined_daily_sheet(self, workbook, ctj_data, collaborator, month_name, year):
        """Create streamlined daily CTJ sheet with only essential data."""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from datetime import datetime
            import calendar

            # Create streamlined daily sheet
            daily_sheet = workbook.create_sheet(title="CTJ Quotidien")

            # Define styles
            header_font = Font(size=11, bold=True, color="FFFFFF")
            data_font = Font(size=10)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            weekend_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            zero_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")

            center_align = Alignment(horizontal="center", vertical="center")

            # Get month info
            month_names = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                          "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
            month_num = month_names.index(month_name) + 1
            year_num = int(year)
            days_in_month = calendar.monthrange(year_num, month_num)[1]

            # Create date headers with day and abbreviated month format
            month_abbrev = {
                1: "Jan", 2: "Fév", 3: "Mar", 4: "Avr", 5: "Mai", 6: "Jun",
                7: "Jul", 8: "Aoû", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Déc"
            }

            if collaborator == "Toute l'équipe":
                # Team export: stack individual collaborators vertically
                return self._create_team_stacked_export(daily_sheet, ctj_data, month_name, year,
                                                      header_font, data_font, header_fill,
                                                      weekend_fill, zero_fill, center_align)
            else:
                # Individual collaborator export
                return self._create_individual_export(daily_sheet, ctj_data, collaborator, month_name, year,
                                                    header_font, data_font, header_fill,
                                                    weekend_fill, zero_fill, center_align)

        except Exception as e:
            self.logger.error(f"Error creating streamlined daily sheet: {e}")
            import traceback
            self.logger.error(f"Full traceback: {traceback.format_exc()}")
            return False

    def _create_individual_export(self, daily_sheet, ctj_data, collaborator, month_name, year,
                                header_font, data_font, header_fill, weekend_fill, zero_fill, center_align):
        """Create individual collaborator export with only CTJ data."""
        try:
            from datetime import datetime
            import calendar
            from openpyxl.styles import Font, Border, Side, PatternFill

            self.logger.info(f"Creating individual export for {collaborator}, data length: {len(ctj_data) if ctj_data else 0}")

            # Define border style
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Get month info
            month_names = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                          "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
            month_num = month_names.index(month_name) + 1
            year_num = int(year)
            days_in_month = calendar.monthrange(year_num, month_num)[1]

            # Create date headers with day and abbreviated month format
            month_abbrev = {
                1: "Jan", 2: "Fév", 3: "Mar", 4: "Avr", 5: "Mai", 6: "Jun",
                7: "Jul", 8: "Aoû", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Déc"
            }

            # Title
            daily_sheet.merge_cells('A1:AH1')
            title_cell = daily_sheet['A1']
            title_cell.value = f"CTJ QUOTIDIEN - {collaborator.upper()} - {month_name} {year}"
            title_cell.font = Font(size=14, bold=True, color="FFFFFF")
            title_cell.fill = header_fill
            title_cell.alignment = center_align

            # Headers with new date format
            headers = ['Collaborateur']
            for day in range(1, days_in_month + 1):
                headers.append(f"{day} {month_abbrev[month_num]}")

            # Write headers with borders
            for col, header in enumerate(headers, 1):
                cell = daily_sheet.cell(row=3, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border

            # Write collaborator name with border
            daily_sheet.cell(row=4, column=1).value = collaborator
            daily_sheet.cell(row=4, column=1).font = Font(bold=True)
            daily_sheet.cell(row=4, column=1).border = thin_border

            # Write CTJ data and calculate cumulative
            total_ctj_written = 0
            cumulative_ctj = 0
            daily_values = []

            for day in range(1, days_in_month + 1):
                col = day + 1
                ctj_value = 0

                if day <= len(ctj_data):
                    row_data = ctj_data[day - 1]
                    ctj_value = row_data.get('CTJ', 0)
                    if ctj_value > 0:
                        total_ctj_written += ctj_value

                # Store daily value for calculations
                daily_values.append(ctj_value)
                cumulative_ctj += ctj_value

                cell = daily_sheet.cell(row=4, column=col)
                cell.value = ctj_value if ctj_value > 0 else ""
                cell.font = data_font
                cell.alignment = center_align
                cell.border = thin_border

                # Apply conditional formatting
                date_obj = datetime(year_num, month_num, day)
                is_weekend = date_obj.weekday() >= 5

                if is_weekend:
                    cell.fill = weekend_fill
                elif ctj_value == 0:
                    cell.fill = zero_fill

            # Add cumulative CTJ row
            daily_sheet.cell(row=5, column=1).value = "Cumulé CTJ"
            daily_sheet.cell(row=5, column=1).font = Font(bold=True, italic=True)
            daily_sheet.cell(row=5, column=1).border = thin_border
            daily_sheet.cell(row=5, column=1).fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")  # Light blue

            running_total = 0
            for day in range(1, days_in_month + 1):
                col = day + 1
                running_total += daily_values[day - 1]

                cell = daily_sheet.cell(row=5, column=col)
                cell.value = running_total if running_total > 0 else ""
                cell.font = Font(size=9, italic=True)
                cell.alignment = center_align
                cell.border = thin_border
                cell.fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")  # Light blue

            # Add global summary section
            self._add_individual_summary(daily_sheet, daily_values, collaborator, month_name, year,
                                       days_in_month, thin_border, header_font, data_font, center_align)

            self.logger.info(f"Individual export: wrote {total_ctj_written} total CTJ values for {collaborator}")

            # Auto-fit all columns based on content (now checking more rows including summary)
            for col in range(1, days_in_month + 2):
                # Get column letter safely
                from openpyxl.utils import get_column_letter
                column_letter = get_column_letter(col)

                # Calculate the maximum width needed for this column
                max_width = 0
                # Check more rows now that we have summary section
                for row in range(2, daily_sheet.max_row + 1):  # Check all rows with content
                    cell = daily_sheet.cell(row=row, column=col)
                    if cell.value and not hasattr(cell, 'coordinate'):  # Skip merged cells
                        cell_width = len(str(cell.value)) + 2  # Add padding
                        max_width = max(max_width, cell_width)

                # Set minimum and maximum widths
                width = max(8, min(max_width, 25))  # Min 8, Max 25
                daily_sheet.column_dimensions[column_letter].width = width

            # Make first column wider for collaborator name and labels
            daily_sheet.column_dimensions['A'].width = max(25, daily_sheet.column_dimensions['A'].width)

            self.logger.info("Individual export sheet created successfully")
            return True

        except Exception as e:
            self.logger.error(f"Error creating individual export: {e}")
            import traceback
            self.logger.error(f"Full traceback: {traceback.format_exc()}")
            return False

    def _add_individual_summary(self, daily_sheet, daily_values, collaborator, month_name, year,
                              days_in_month, thin_border, header_font, data_font, center_align):
        """Add global summary section for individual collaborator."""
        try:
            from openpyxl.styles import Font, PatternFill

            # Calculate statistics
            total_ctj = sum(daily_values)
            working_days = len([v for v in daily_values if v > 0])
            avg_ctj_per_working_day = total_ctj / working_days if working_days > 0 else 0
            avg_ctj_per_month_day = total_ctj / days_in_month if days_in_month > 0 else 0
            max_daily_ctj = max(daily_values) if daily_values else 0

            # Calculate working days in month (excluding weekends)
            month_names = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                          "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
            month_num = month_names.index(month_name) + 1
            year_num = int(year)
            working_days_in_month = self._calculate_working_days_in_month(month_num, year_num)

            # Summary section styling
            summary_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")  # Light blue
            label_font = Font(size=10, bold=True)
            value_font = Font(size=10)

            # Start summary section 2 rows below the cumulative row
            start_row = 8

            # Title for summary section
            daily_sheet.merge_cells(f'A{start_row}:D{start_row}')
            title_cell = daily_sheet[f'A{start_row}']
            title_cell.value = f"📊 RÉSUMÉ GLOBAL - {month_name} {year}"
            title_cell.font = Font(size=12, bold=True, color="1F4E79")
            title_cell.alignment = center_align
            title_cell.fill = summary_fill
            title_cell.border = thin_border

            # Summary data
            summary_data = [
                ("CTJ Total du Mois", f"{total_ctj:,}"),
                ("Jours Travaillés", f"{working_days}"),
                ("CTJ Moyen/Jour Travaillé", f"{avg_ctj_per_working_day:.1f}"),
                ("CTJ Moyen/Jour du Mois", f"{avg_ctj_per_month_day:.1f}"),
                ("CTJ Maximum en 1 Jour", f"{max_daily_ctj}"),
                ("Taux d'Activité", f"{(working_days/working_days_in_month)*100:.1f}%")
            ]

            # Write summary data
            for i, (label, value) in enumerate(summary_data):
                row = start_row + 1 + i

                # Label column
                label_cell = daily_sheet.cell(row=row, column=1)
                label_cell.value = label
                label_cell.font = label_font
                label_cell.border = thin_border
                label_cell.fill = summary_fill

                # Value column
                value_cell = daily_sheet.cell(row=row, column=2)
                value_cell.value = value
                value_cell.font = value_font
                value_cell.alignment = center_align
                value_cell.border = thin_border

                # Empty cells for formatting consistency
                for col in range(3, 5):
                    empty_cell = daily_sheet.cell(row=row, column=col)
                    empty_cell.border = thin_border
                    empty_cell.fill = summary_fill

            self.logger.info(f"Added individual summary for {collaborator}: Total CTJ={total_ctj}, Working days={working_days}")

        except Exception as e:
            self.logger.error(f"Error adding individual summary: {e}")

    def _create_team_stacked_export(self, daily_sheet, ctj_data, month_name, year,
                                  header_font, data_font, header_fill, weekend_fill, zero_fill, center_align):
        """Create team export with individual collaborators stacked vertically."""
        try:
            from datetime import datetime
            import calendar
            from openpyxl.styles import Font, Border, Side, PatternFill, Alignment

            self.logger.info(f"Creating team export, data length: {len(ctj_data) if ctj_data else 0}")

            # Define border style
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Get month info
            month_names = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                          "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
            month_num = month_names.index(month_name) + 1
            year_num = int(year)
            days_in_month = calendar.monthrange(year_num, month_num)[1]

            # Create date headers with day and abbreviated month format
            month_abbrev = {
                1: "Jan", 2: "Fév", 3: "Mar", 4: "Avr", 5: "Mai", 6: "Jun",
                7: "Jul", 8: "Aoû", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Déc"
            }

            # Title
            daily_sheet.merge_cells('A1:AH1')
            title_cell = daily_sheet['A1']
            title_cell.value = f"CTJ QUOTIDIEN - TOUTE L'ÉQUIPE - {month_name} {year}"
            title_cell.font = Font(size=14, bold=True, color="FFFFFF")
            title_cell.fill = header_fill
            title_cell.alignment = center_align

            # Headers with new date format
            headers = ['Collaborateur']
            for day in range(1, days_in_month + 1):
                headers.append(f"{day} {month_abbrev[month_num]}")

            # Write headers with borders
            for col, header in enumerate(headers, 1):
                cell = daily_sheet.cell(row=3, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border

            # Initialize current_row for auto-fitting
            current_row = 4

            # Extract collaborators from ctj_data and stack them vertically
            if ctj_data and len(ctj_data) > 0:
                # Get all collaborators from the first day's data (excluding 'Date' and 'Total')
                first_day_data = ctj_data[0]
                collaborators = [key for key in first_day_data.keys() if key not in ['Date', 'Total']]

                self.logger.info(f"Team export: found {len(collaborators)} collaborators: {collaborators}")

                # Check if we have any real collaborators (not just 'Aucune donnée')
                real_collaborators = [c for c in collaborators if c != 'Aucune donnée']
                if not real_collaborators and 'Aucune donnée' in collaborators:
                    self.logger.warning("Only dummy data found for team export")

                team_daily_values = {}  # Store daily values for each collaborator

                for i, collaborator in enumerate(sorted(collaborators)):
                    # Add separator line between collaborators (except for the first one)
                    if i > 0:
                        # Add empty row for visual separation
                        separator_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                        for col in range(1, days_in_month + 2):  # Include all columns
                            sep_cell = daily_sheet.cell(row=current_row, column=col)
                            sep_cell.fill = separator_fill
                            sep_cell.border = thin_border
                        current_row += 1

                    # Write collaborator name with enhanced formatting
                    daily_sheet.cell(row=current_row, column=1).value = collaborator
                    daily_sheet.cell(row=current_row, column=1).font = Font(bold=True, size=11)  # Slightly larger and bold
                    daily_sheet.cell(row=current_row, column=1).border = thin_border
                    # Add light blue background for collaborator names
                    daily_sheet.cell(row=current_row, column=1).fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")

                    # Write CTJ data for this collaborator
                    total_ctj_for_collab = 0
                    daily_values = []

                    for day in range(1, days_in_month + 1):
                        col = day + 1
                        ctj_value = 0

                        if day <= len(ctj_data):
                            row_data = ctj_data[day - 1]
                            ctj_value = row_data.get(collaborator, 0)
                            if ctj_value > 0:
                                total_ctj_for_collab += ctj_value

                        daily_values.append(ctj_value)

                        cell = daily_sheet.cell(row=current_row, column=col)
                        cell.value = ctj_value if ctj_value > 0 else ""
                        cell.font = data_font
                        cell.alignment = center_align
                        cell.border = thin_border

                        # Apply conditional formatting
                        date_obj = datetime(year_num, month_num, day)
                        is_weekend = date_obj.weekday() >= 5

                        if is_weekend:
                            cell.fill = weekend_fill
                        elif ctj_value == 0:
                            cell.fill = zero_fill

                    # Store daily values for team summary
                    team_daily_values[collaborator] = daily_values

                    self.logger.info(f"Team export: wrote {total_ctj_for_collab} total CTJ values for {collaborator}")
                    current_row += 1

                    # Add cumulative row for this collaborator with enhanced formatting
                    daily_sheet.cell(row=current_row, column=1).value = f"Cumulé {collaborator}"
                    daily_sheet.cell(row=current_row, column=1).font = Font(bold=True, italic=True, size=9)
                    daily_sheet.cell(row=current_row, column=1).border = thin_border
                    daily_sheet.cell(row=current_row, column=1).fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")  # Light blue

                    running_total = 0
                    for day in range(1, days_in_month + 1):
                        col = day + 1
                        running_total += daily_values[day - 1]

                        cell = daily_sheet.cell(row=current_row, column=col)
                        cell.value = running_total if running_total > 0 else ""
                        cell.font = Font(size=8, italic=True)
                        cell.alignment = center_align
                        cell.border = thin_border
                        cell.fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")

                    current_row += 1  # Move to next row after cumulative row

                # Add team global summary
                self._add_team_summary(daily_sheet, team_daily_values, month_name, year,
                                     days_in_month, current_row, thin_border, header_font, data_font, center_align)
            else:
                self.logger.warning("Team export: No CTJ data available")
                # Add a message in the sheet when no data is available
                daily_sheet.cell(row=4, column=1).value = "Aucune donnée disponible"
                daily_sheet.cell(row=4, column=1).font = Font(italic=True)
                daily_sheet.cell(row=4, column=1).border = thin_border
                current_row = 5

            # Auto-fit all columns based on content (now checking all rows including summary)
            for col in range(1, days_in_month + 2):
                # Get column letter safely
                from openpyxl.utils import get_column_letter
                column_letter = get_column_letter(col)

                # Calculate the maximum width needed for this column
                max_width = 0
                # Check all rows with content including summary sections
                for row in range(2, daily_sheet.max_row + 1):
                    cell = daily_sheet.cell(row=row, column=col)
                    if cell.value and not hasattr(cell, 'coordinate'):  # Skip merged cells
                        cell_width = len(str(cell.value)) + 2  # Add padding
                        max_width = max(max_width, cell_width)

                # Set minimum and maximum widths
                width = max(8, min(max_width, 30))  # Min 8, Max 30 (increased for summary content)
                daily_sheet.column_dimensions[column_letter].width = width

            # Make first column wider for collaborator names and labels
            daily_sheet.column_dimensions['A'].width = max(30, daily_sheet.column_dimensions['A'].width)

            self.logger.info("Team stacked export sheet created successfully")
            return True

        except Exception as e:
            self.logger.error(f"Error creating team stacked export: {e}")
            import traceback
            self.logger.error(f"Full traceback: {traceback.format_exc()}")
            return False

    def _add_team_summary(self, daily_sheet, team_daily_values, month_name, year,
                         days_in_month, start_row, thin_border, header_font, data_font, center_align):
        """Add global summary section for team export."""
        try:
            from openpyxl.styles import Font, PatternFill

            # Calculate team statistics
            team_totals = []
            team_working_days = []
            collaborator_summaries = {}

            for collaborator, daily_values in team_daily_values.items():
                total_ctj = sum(daily_values)
                working_days = len([v for v in daily_values if v > 0])
                avg_ctj_per_working_day = total_ctj / working_days if working_days > 0 else 0
                max_daily_ctj = max(daily_values) if daily_values else 0

                collaborator_summaries[collaborator] = {
                    'total': total_ctj,
                    'working_days': working_days,
                    'avg_per_working_day': avg_ctj_per_working_day,
                    'max_daily': max_daily_ctj
                }

                team_totals.append(total_ctj)
                team_working_days.append(working_days)

            # Overall team statistics
            total_team_ctj = sum(team_totals)
            avg_team_working_days = sum(team_working_days) / len(team_working_days) if team_working_days else 0
            avg_team_ctj_per_month_day = total_team_ctj / days_in_month if days_in_month > 0 else 0
            max_team_daily = max([max(values) for values in team_daily_values.values()]) if team_daily_values else 0

            # Summary section styling
            summary_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
            team_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")  # Light green
            label_font = Font(size=10, bold=True)
            value_font = Font(size=10)

            # Title for summary section
            daily_sheet.merge_cells(f'A{start_row}:E{start_row}')
            title_cell = daily_sheet[f'A{start_row}']
            title_cell.value = f"📊 RÉSUMÉ GLOBAL ÉQUIPE - {month_name} {year}"
            title_cell.font = Font(size=12, bold=True, color="1F4E79")
            title_cell.alignment = center_align
            title_cell.fill = team_fill
            title_cell.border = thin_border

            current_row = start_row + 2

            # Individual collaborator summaries
            daily_sheet.merge_cells(f'A{current_row}:E{current_row}')
            collab_title_cell = daily_sheet[f'A{current_row}']
            collab_title_cell.value = "👥 DÉTAIL PAR COLLABORATEUR"
            collab_title_cell.font = Font(size=11, bold=True, color="2E7D32")
            collab_title_cell.alignment = center_align
            collab_title_cell.fill = summary_fill
            collab_title_cell.border = thin_border

            current_row += 1

            # Headers for collaborator details
            headers = ["Collaborateur", "CTJ Total", "Jours Travaillés", "CTJ Moy/Jour", "CTJ Max/Jour"]
            for col, header in enumerate(headers, 1):
                header_cell = daily_sheet.cell(row=current_row, column=col)
                header_cell.value = header
                header_cell.font = Font(size=9, bold=True)
                header_cell.alignment = center_align
                header_cell.border = thin_border
                header_cell.fill = summary_fill

            current_row += 1

            # Write collaborator details
            for collaborator, stats in sorted(collaborator_summaries.items()):
                collab_data = [
                    collaborator,
                    f"{stats['total']:,}",
                    f"{stats['working_days']}",
                    f"{stats['avg_per_working_day']:.1f}",
                    f"{stats['max_daily']}"
                ]

                for col, data in enumerate(collab_data, 1):
                    cell = daily_sheet.cell(row=current_row, column=col)
                    cell.value = data
                    cell.font = Font(size=9)
                    if col == 1:  # Collaborator name
                        cell.font = Font(size=9, bold=True)
                    cell.alignment = center_align if col > 1 else None
                    cell.border = thin_border

                current_row += 1

            self.logger.info(f"Added team summary: Total CTJ={total_team_ctj}, {len(team_daily_values)} collaborators")

        except Exception as e:
            self.logger.error(f"Error adding team summary: {e}")

    def _calculate_enhanced_daily_data(self, ctj_data, collaborator, month_name, year):
        """Calculate enhanced daily data with additional KPIs."""
        try:
            from datetime import datetime
            import calendar

            # Get month info
            month_names = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                          "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
            month_num = month_names.index(month_name) + 1
            year_num = int(year)
            days_in_month = calendar.monthrange(year_num, month_num)[1]

            enhanced_data = {}
            cumulative_ctj = 0

            # Get DMT for the collaborator
            collaborator_dmt = 0
            if collaborator in self.collaborator_stats:
                collaborator_dmt = self.collaborator_stats[collaborator].get('dmt', 0)
            elif collaborator == "Toute l'équipe":
                # Calculate average team DMT
                total_dmt = sum(stats.get('dmt', 0) for stats in self.collaborator_stats.values())
                active_collabs = len([stats for stats in self.collaborator_stats.values() if stats.get('dmt', 0) > 0])
                collaborator_dmt = total_dmt / active_collabs if active_collabs > 0 else 0

            for day in range(1, days_in_month + 1):
                day_data = {}

                if day <= len(ctj_data):
                    row_data = ctj_data[day - 1]

                    if collaborator == "Toute l'équipe":
                        # Team metrics
                        ctj_value = row_data.get('Total', 0)
                        day_data['Total'] = ctj_value
                        day_data['Communes_Started'] = self._estimate_communes_started(day, month_num, year_num)
                        day_data['Communes_Completed'] = self._estimate_communes_completed(day, month_num, year_num)
                        day_data['Active_Collabs'] = len([k for k, v in row_data.items() if k != 'Total' and v > 0])
                        day_data['Efficiency'] = ctj_value / collaborator_dmt if collaborator_dmt > 0 else 0
                    else:
                        # Individual metrics
                        ctj_value = row_data.get('CTJ', 0)
                        day_data['CTJ'] = ctj_value
                        day_data['DMT'] = collaborator_dmt
                        day_data['Active_Communes'] = self._estimate_active_communes(collaborator, day, month_num, year_num)
                        day_data['Efficiency'] = ctj_value / collaborator_dmt if collaborator_dmt > 0 else 0

                        cumulative_ctj += ctj_value
                        day_data['Cumulative_CTJ'] = cumulative_ctj
                else:
                    # No data for this day
                    if collaborator == "Toute l'équipe":
                        day_data = {'Total': 0, 'Communes_Started': 0, 'Communes_Completed': 0, 'Active_Collabs': 0, 'Efficiency': 0}
                    else:
                        day_data = {'CTJ': 0, 'DMT': collaborator_dmt, 'Active_Communes': 0, 'Efficiency': 0, 'Cumulative_CTJ': cumulative_ctj}

                enhanced_data[day] = day_data

            return enhanced_data

        except Exception as e:
            self.logger.error(f"Error calculating enhanced daily data: {e}")
            return {}

    def _estimate_communes_started(self, day, month, year):
        """Estimate communes started on a specific day (simplified)."""
        # This is a simplified estimation - in a real implementation,
        # you would track actual start dates from the data
        try:
            if 'Suivi Tickets' in self.global_suivi_data:
                df_tickets = self.global_suivi_data['Suivi Tickets']
                if not df_tickets.empty and 'Date d\'affectation' in df_tickets.columns:
                    # Filter by date (simplified - would need proper date parsing)
                    return 1 if day % 3 == 0 else 0  # Placeholder logic
            return 0
        except:
            return 0

    def _estimate_communes_completed(self, day, month, year):
        """Estimate communes completed on a specific day (simplified)."""
        # This is a simplified estimation - in a real implementation,
        # you would track actual completion dates from the data
        try:
            if 'Suivi Tickets' in self.global_suivi_data:
                df_tickets = self.global_suivi_data['Suivi Tickets']
                if not df_tickets.empty and 'Date Livraison' in df_tickets.columns:
                    # Filter by date (simplified - would need proper date parsing)
                    return 1 if day % 5 == 0 else 0  # Placeholder logic
            return 0
        except:
            return 0

    def _estimate_active_communes(self, collaborator, day, month, year):
        """Estimate active communes for a collaborator on a specific day."""
        # This is a simplified estimation - in a real implementation,
        # you would track actual active communes from the data
        try:
            if collaborator in self.collaborator_stats:
                # Estimate based on total workload (simplified)
                base_communes = 3  # Base number of active communes
                variation = (day % 7) - 3  # Daily variation
                return max(1, base_communes + variation)
            return 0
        except:
            return 0

    def _calculate_working_days_in_month(self, month_num, year_num):
        """
        Calculate the number of working days (excluding weekends) in a month.

        Args:
            month_num (int): Month number (1-12)
            year_num (int): Year number

        Returns:
            int: Number of working days (Monday to Friday)

        Examples:
            - June 2025 (30 days): 21 working days (30 - 9 weekend days)
            - February 2025 (28 days): 20 working days (28 - 8 weekend days)
        """
        try:
            import calendar
            from datetime import datetime

            # Get total days in month
            days_in_month = calendar.monthrange(year_num, month_num)[1]
            working_days = 0

            # Count working days (Monday=0 to Friday=4, Saturday=5, Sunday=6)
            for day in range(1, days_in_month + 1):
                date_obj = datetime(year_num, month_num, day)
                if date_obj.weekday() < 5:  # Monday to Friday (0-4)
                    working_days += 1

            self.logger.debug(f"Month {month_num}/{year_num}: {working_days} working days out of {days_in_month} total days")
            return working_days

        except Exception as e:
            self.logger.error(f"Error calculating working days: {e}")
            # Fallback: approximate 22 working days per month
            return 22

    def _create_commune_tracking_sheet(self, workbook, collaborator, month_name, year):
        """Create commune tracking sheet with detailed commune information."""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            # Create commune tracking sheet
            commune_sheet = workbook.create_sheet(title="Suivi Communes")

            # Define styles
            header_font = Font(size=11, bold=True, color="FFFFFF")
            data_font = Font(size=10)
            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            completed_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            in_progress_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            pending_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")

            center_align = Alignment(horizontal="center", vertical="center")
            left_align = Alignment(horizontal="left", vertical="center")

            # Title
            commune_sheet.merge_cells('A1:J1')
            title_cell = commune_sheet['A1']
            title_cell.value = f"🏘️ SUIVI DÉTAILLÉ DES COMMUNES - {collaborator.upper()}"
            title_cell.font = Font(size=14, bold=True, color="FFFFFF")
            title_cell.fill = header_fill
            title_cell.alignment = center_align

            # Headers
            headers = [
                'Commune', 'Code INSEE', 'Collaborateur', 'Statut', 'Date Début',
                'Date Fin', 'Durée (min)', 'Éléments', 'DMT', 'Progression'
            ]

            for col, header in enumerate(headers, 1):
                cell = commune_sheet.cell(row=3, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align

            # Get commune data
            commune_data = self._get_commune_tracking_data(collaborator, month_name, year)

            # Write data
            for row_idx, commune_info in enumerate(commune_data, 4):
                commune_sheet.cell(row=row_idx, column=1).value = commune_info.get('nom', 'N/A')
                commune_sheet.cell(row=row_idx, column=2).value = commune_info.get('insee', 'N/A')
                commune_sheet.cell(row=row_idx, column=3).value = commune_info.get('collaborateur', 'N/A')
                commune_sheet.cell(row=row_idx, column=4).value = commune_info.get('statut', 'N/A')
                commune_sheet.cell(row=row_idx, column=5).value = commune_info.get('date_debut', 'N/A')
                commune_sheet.cell(row=row_idx, column=6).value = commune_info.get('date_fin', 'N/A')
                commune_sheet.cell(row=row_idx, column=7).value = commune_info.get('duree', 0)
                commune_sheet.cell(row=row_idx, column=8).value = commune_info.get('elements', 0)
                commune_sheet.cell(row=row_idx, column=9).value = f"{commune_info.get('dmt', 0):.1f}"
                commune_sheet.cell(row=row_idx, column=10).value = f"{commune_info.get('progression', 0):.1f}%"

                # Apply status-based formatting
                status = commune_info.get('statut', '')
                for col in range(1, 11):
                    cell = commune_sheet.cell(row=row_idx, column=col)
                    cell.font = data_font
                    if col in [1, 3]:  # Left align for text columns
                        cell.alignment = left_align
                    else:
                        cell.alignment = center_align

                    if status == 'Traité':
                        cell.fill = completed_fill
                    elif status == 'En Cours':
                        cell.fill = in_progress_fill
                    elif status in ['En Attente', 'Rejeté']:
                        cell.fill = pending_fill

            # Auto-adjust column widths
            column_widths = [20, 12, 18, 12, 12, 12, 12, 10, 8, 12]
            for col, width in enumerate(column_widths, 1):
                commune_sheet.column_dimensions[commune_sheet.cell(row=1, column=col).column_letter].width = width

            self.logger.info("Commune tracking sheet created successfully")
            return True

        except Exception as e:
            self.logger.error(f"Error creating commune tracking sheet: {e}")
            return False

    def _get_commune_tracking_data(self, collaborator, month_name, year):
        """Get commune tracking data for the specified collaborator and period."""
        try:
            commune_data = []

            if 'Suivi Tickets' not in self.global_suivi_data:
                return commune_data

            df_tickets = self.global_suivi_data['Suivi Tickets']
            if df_tickets.empty:
                return commune_data

            # Filter data for collaborator
            if collaborator == "Toute l'équipe":
                filtered_data = df_tickets
            else:
                filtered_data = df_tickets[df_tickets.get('Collaborateur', '') == collaborator]

            # Process each commune
            for _, row in filtered_data.iterrows():
                commune_info = {
                    'nom': row.get('Nom Commune', 'N/A'),
                    'insee': row.get('Code INSEE', 'N/A'),
                    'collaborateur': row.get('Collaborateur', 'N/A'),
                    'statut': row.get('Etat Ticket PA', 'N/A'),
                    'date_debut': row.get('Date d\'affectation', 'N/A'),
                    'date_fin': row.get('Date Livraison', 'N/A'),
                    'duree': row.get('Durée Finale', 0),
                    'elements': self._estimate_commune_elements(row),
                    'dmt': self._calculate_commune_dmt(row),
                    'progression': self._calculate_commune_progression(row)
                }
                commune_data.append(commune_info)

            return commune_data

        except Exception as e:
            self.logger.error(f"Error getting commune tracking data: {e}")
            return []

    def _estimate_commune_elements(self, commune_row):
        """Estimate number of elements for a commune."""
        # This would be calculated from actual data in a real implementation
        return 50  # Placeholder

    def _calculate_commune_dmt(self, commune_row):
        """Calculate DMT for a specific commune."""
        duree = commune_row.get('Durée Finale', 0)
        elements = self._estimate_commune_elements(commune_row)
        return duree / elements if elements > 0 else 0

    def _calculate_commune_progression(self, commune_row):
        """Calculate progression percentage for a commune."""
        status = commune_row.get('Etat Ticket PA', '')
        if status == 'Traité':
            return 100.0
        elif status == 'En Cours':
            return 75.0  # Estimated
        elif status == 'En Attente':
            return 25.0  # Estimated
        else:
            return 0.0

    def _update_export_filters(self):
        """Update export filter options based on loaded data."""
        try:
            # Check if we have the necessary data and UI components
            if not self.collaborator_stats:
                self.logger.debug("No collaborator stats available, skipping filter update")
                return

            # Check if export UI components are created
            if not hasattr(self, 'collab_combo') or self.collab_combo is None:
                self.logger.debug("Export UI components not yet created, skipping filter update")
                return

            # Update collaborator list with "Toute l'équipe" option
            collaborators = ["Toute l'équipe"] + sorted(list(self.collaborator_stats.keys()))

            # Safely update collaborator filter
            try:
                self.collab_combo['values'] = collaborators
                self.logger.debug(f"Updated collaborator combobox with {len(collaborators)} values")
            except Exception as e:
                self.logger.warning(f"Could not update collaborator combobox values: {e}")

            # Set "Toute l'équipe" as default selection
            try:
                if hasattr(self, 'collab_var') and self.collab_var is not None:
                    self.collab_var.set("Toute l'équipe")
                    self.logger.debug("Set default collaborator selection to 'Toute l'équipe'")
            except Exception as e:
                self.logger.warning(f"Could not set default collaborator selection: {e}")

            # Set current month as default
            try:
                from datetime import datetime
                current_month = datetime.now().month
                month_names = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                              "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
                if hasattr(self, 'month_var') and self.month_var is not None:
                    self.month_var.set(month_names[current_month - 1])
                    self.logger.debug(f"Set default month selection to {month_names[current_month - 1]}")
            except Exception as e:
                self.logger.warning(f"Could not set default month selection: {e}")

            self.logger.info(f"Export filters updated with {len(collaborators)} collaborators (including 'Toute l'équipe')")

        except Exception as e:
            self.logger.error(f"Error updating export filters: {e}")
            # Additional error details for debugging
            self.logger.error(f"collaborator_stats exists: {bool(self.collaborator_stats)}")
            self.logger.error(f"collab_combo exists: {hasattr(self, 'collab_combo')}")
            self.logger.error(f"collab_var exists: {hasattr(self, 'collab_var')}")
            self.logger.error(f"month_var exists: {hasattr(self, 'month_var')}")
            if hasattr(self, 'export_filters'):
                self.logger.error(f"export_filters content: {self.export_filters}")

    def _enable_export_buttons(self):
        """Enable export buttons and filters when data is loaded."""
        try:
            self.logger.debug("Attempting to enable export buttons...")

            # Enable export buttons
            if hasattr(self, 'export_buttons') and self.export_buttons:
                self.logger.debug(f"Found export_buttons: {list(self.export_buttons.keys())}")
                for button_name, button in self.export_buttons.items():
                    if button is not None:
                        try:
                            button.config(state=tk.NORMAL)
                            self.logger.debug(f"Enabled button: {button_name}")
                        except Exception as e:
                            self.logger.warning(f"Could not enable button {button_name}: {e}")
            else:
                self.logger.warning("export_buttons not found or empty")

            # Enable export filters
            if hasattr(self, 'export_filters') and self.export_filters:
                self.logger.debug(f"Found export_filters: {list(self.export_filters.keys())}")
                for filter_name, filter_widget in self.export_filters.items():
                    if filter_widget is not None:
                        try:
                            filter_widget.config(state="readonly")
                            self.logger.debug(f"Enabled filter: {filter_name}")
                        except Exception as e:
                            self.logger.warning(f"Could not enable filter {filter_name}: {e}")
            else:
                self.logger.warning("export_filters not found or empty")

            self.logger.info("Export buttons and filters enabled successfully")

        except Exception as e:
            self.logger.error(f"Error enabling export buttons: {e}")
            # Additional debugging information
            self.logger.error(f"Has export_buttons: {hasattr(self, 'export_buttons')}")
            self.logger.error(f"Has export_filters: {hasattr(self, 'export_filters')}")
            if hasattr(self, 'export_buttons'):
                self.logger.error(f"export_buttons content: {self.export_buttons}")
            if hasattr(self, 'export_filters'):
                self.logger.error(f"export_filters content: {self.export_filters}")

    def _disable_export_buttons(self):
        """Disable export buttons and filters when no data is available."""
        try:
            if hasattr(self, 'export_buttons') and self.export_buttons:
                for button_name, button in self.export_buttons.items():
                    if button is not None:
                        try:
                            button.config(state=tk.DISABLED)
                        except Exception as e:
                            self.logger.warning(f"Could not disable button {button_name}: {e}")

            if hasattr(self, 'export_filters') and self.export_filters:
                for filter_name, filter_widget in self.export_filters.items():
                    if filter_widget is not None:
                        try:
                            filter_widget.config(state=tk.DISABLED)
                        except Exception as e:
                            self.logger.warning(f"Could not disable filter {filter_name}: {e}")

        except Exception as e:
            self.logger.error(f"Error disabling export buttons: {e}")

    def _export_anomalies(self):
        """Export anomalies from Suivis Global Tickets CMS Adr_PA file."""
        try:
            from tkinter import filedialog
            from datetime import datetime
            import os

            # Check if global data is loaded
            if not self.global_suivi_data:
                messagebox.showwarning("Données manquantes", "Veuillez d'abord charger les données globales.")
                return

            self.logger.info("Starting anomalies export...")

            # Find anomalies
            anomalies = self._detect_anomalies()

            if not anomalies:
                messagebox.showinfo("Aucune anomalie", "Aucune anomalie détectée dans les données.")
                return

            # Ask user for save location
            current_date = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"Anomalies_Export_{current_date}.xlsx"

            file_path = filedialog.asksaveasfilename(
                title="Enregistrer l'export des anomalies",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=filename
            )

            if not file_path:
                return

            # Create Excel file with anomalies
            self._create_anomalies_excel(file_path, anomalies)

            messagebox.showinfo("Export réussi", f"Export des anomalies terminé avec succès !\n\n{len(anomalies)} anomalie(s) détectée(s)\n\nFichier : {file_path}")
            self.logger.info(f"Anomalies export completed: {file_path}")

        except Exception as e:
            self.logger.error(f"Error exporting anomalies: {e}")
            messagebox.showerror("Erreur Export", f"Erreur lors de l'export des anomalies :\n{str(e)}")

    def _detect_anomalies(self):
        """Detect anomalies in the global data according to specified criteria."""
        try:
            from utils.lazy_imports import get_pandas
            from datetime import datetime

            pd = get_pandas()
            anomalies = []
            today = datetime.now().date()

            self.logger.info("Detecting anomalies in global data...")

            # Check each sheet
            for sheet_name, df in self.global_suivi_data.items():
                if df.empty:
                    continue

                self.logger.info(f"Checking anomalies in sheet: {sheet_name}")

                if sheet_name == 'Suivi Tickets':
                    # Feuille 1: Vérifier cases vides colonnes A à U (sauf O, Q, R, T)
                    anomalies.extend(self._check_empty_cells_sheet1(df, pd))
                    # Vérifier dates futures
                    anomalies.extend(self._check_future_dates_sheet1(df, pd, today))
                    # Vérifier incohérence dates affectation/livraison
                    anomalies.extend(self._check_assignment_delivery_dates_sheet1(df, pd))

                elif sheet_name == 'Traitement CMS Adr':
                    # Feuille 2: Vérifier dates futures
                    anomalies.extend(self._check_future_dates_sheet2(df, pd, today))
                    # Vérifier tickets CM traités sans motif voie
                    anomalies.extend(self._check_cm_treated_without_motif_sheet2(df, pd))

                elif sheet_name == 'Traitement PA':
                    # Feuille 3: Vérifier dates futures
                    anomalies.extend(self._check_future_dates_sheet3(df, pd, today))
                    # Vérifier adresses non jointes avec temps de traitement
                    anomalies.extend(self._check_unjoined_address_with_time_sheet3(df, pd))

            self.logger.info(f"Total anomalies detected: {len(anomalies)}")
            return anomalies

        except Exception as e:
            self.logger.error(f"Error detecting anomalies: {e}")
            return []

    def _check_empty_cells_sheet1(self, df, pd):
        """Check for empty cells in sheet 1 (columns A to U except O, Q, R, T)."""
        anomalies = []
        try:
            # Colonnes à vérifier (A à U sauf O, Q, R, T)
            columns_to_check = []
            excluded_cols = ['O', 'Q', 'R', 'T']  # Colonnes exclues

            # Générer les lettres de colonnes A à U
            for i in range(ord('A'), ord('U') + 1):
                col_letter = chr(i)
                if col_letter not in excluded_cols:
                    columns_to_check.append(col_letter)

            # Mapper les lettres aux noms de colonnes réels
            actual_columns = []
            for i, col_name in enumerate(df.columns):
                col_letter = chr(ord('A') + i)
                if col_letter in columns_to_check:
                    actual_columns.append(col_name)

            self.logger.info(f"Checking empty cells in columns: {actual_columns}")

            # Vérifier chaque ligne
            for index, row in df.iterrows():
                insee_code = row.get(df.columns[2], '') if len(df.columns) > 2 else ''  # Colonne C

                # Vérifier les cellules vides dans les colonnes requises
                empty_columns = []
                for col_name in actual_columns:
                    if col_name in row.index:
                        value = row[col_name]
                        if pd.isna(value) or value == '' or str(value).strip() == '':
                            empty_columns.append(col_name)

                if empty_columns:
                    anomalies.append({
                        'type': 'Cases vides',
                        'feuille': 'Suivi Tickets',
                        'ligne': index + 2,  # +2 car Excel commence à 1 et il y a un header
                        'code_insee': str(insee_code),
                        'colonnes_vides': ', '.join(empty_columns),
                        'commentaire': f'Cases vides dans colonnes: {", ".join(empty_columns)}'
                    })

            self.logger.info(f"Found {len([a for a in anomalies if a['type'] == 'Cases vides'])} empty cell anomalies in sheet 1")

        except Exception as e:
            self.logger.error(f"Error checking empty cells in sheet 1: {e}")

        return anomalies

    def _check_future_dates_sheet1(self, df, pd, today):
        """Check for future dates in sheet 1."""
        anomalies = []
        try:
            # Colonnes de dates à vérifier dans la feuille 1
            date_columns = []
            for col_name in df.columns:
                if 'date' in col_name.lower():
                    date_columns.append(col_name)

            self.logger.info(f"Checking future dates in sheet 1 columns: {date_columns}")

            for index, row in df.iterrows():
                insee_code = row.get(df.columns[2], '') if len(df.columns) > 2 else ''  # Colonne C

                for col_name in date_columns:
                    if col_name in row.index:
                        date_value = row[col_name]

                        if pd.isna(date_value) or date_value == '':
                            continue

                        try:
                            # Essayer de parser la date
                            if isinstance(date_value, str):
                                for date_format in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
                                    try:
                                        date_obj = datetime.strptime(date_value, date_format).date()
                                        break
                                    except ValueError:
                                        continue
                                else:
                                    continue
                            else:
                                date_obj = pd.to_datetime(date_value).date()

                            # Vérifier si la date est dans le futur
                            if date_obj > today:
                                anomalies.append({
                                    'type': 'Date future',
                                    'feuille': 'Suivi Tickets',
                                    'ligne': index + 2,
                                    'code_insee': str(insee_code),
                                    'colonne': col_name,
                                    'date': str(date_value),
                                    'commentaire': 'Incohérence date'
                                })

                        except Exception as e:
                            self.logger.debug(f"Error parsing date {date_value} in sheet 1: {e}")
                            continue

            future_dates_count = len([a for a in anomalies if a['type'] == 'Date future' and a['feuille'] == 'Suivi Tickets'])
            self.logger.info(f"Found {future_dates_count} future date anomalies in sheet 1")

        except Exception as e:
            self.logger.error(f"Error checking future dates in sheet 1: {e}")

        return anomalies

    def _check_assignment_delivery_dates_sheet1(self, df, pd):
        """Check for assignment dates posterior to delivery dates in sheet 1."""
        anomalies = []
        try:
            self.logger.info("Checking assignment/delivery date inconsistencies in sheet 1")

            # Colonnes I (date d'affectation) et O (date de livraison)
            assignment_col = None
            delivery_col = None

            # Trouver les colonnes par index (I = index 8, O = index 14)
            if len(df.columns) > 8:
                assignment_col = df.columns[8]  # Colonne I
            if len(df.columns) > 14:
                delivery_col = df.columns[14]  # Colonne O

            if not assignment_col or not delivery_col:
                self.logger.warning("Assignment or delivery columns not found in sheet 1")
                return anomalies

            self.logger.info(f"Checking assignment column: {assignment_col}, delivery column: {delivery_col}")

            for index, row in df.iterrows():
                insee_code = row.get(df.columns[2], '') if len(df.columns) > 2 else ''  # Colonne C

                assignment_value = row.get(assignment_col, '')
                delivery_value = row.get(delivery_col, '')

                # Ignorer si l'une des dates est vide
                if pd.isna(assignment_value) or assignment_value == '' or pd.isna(delivery_value) or delivery_value == '':
                    continue

                try:
                    # Parser les dates
                    assignment_date = None
                    delivery_date = None

                    # Parser date d'affectation
                    if isinstance(assignment_value, str):
                        for date_format in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
                            try:
                                assignment_date = datetime.strptime(assignment_value, date_format).date()
                                break
                            except ValueError:
                                continue
                    else:
                        assignment_date = pd.to_datetime(assignment_value).date()

                    # Parser date de livraison
                    if isinstance(delivery_value, str):
                        for date_format in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
                            try:
                                delivery_date = datetime.strptime(delivery_value, date_format).date()
                                break
                            except ValueError:
                                continue
                    else:
                        delivery_date = pd.to_datetime(delivery_value).date()

                    # Vérifier si date d'affectation > date de livraison
                    if assignment_date and delivery_date and assignment_date > delivery_date:
                        anomalies.append({
                            'type': 'Incohérence dates affectation/livraison',
                            'feuille': 'Suivi Tickets',
                            'ligne': index + 2,
                            'code_insee': str(insee_code),
                            'colonnes_concernees': f'{assignment_col}, {delivery_col}',
                            'valeurs_problematiques': f'Affectation: {assignment_value}, Livraison: {delivery_value}',
                            'commentaire': f'Date d\'affectation ({assignment_value}) postérieure à la date de livraison ({delivery_value})'
                        })

                except Exception as e:
                    self.logger.debug(f"Error parsing assignment/delivery dates {assignment_value}/{delivery_value}: {e}")
                    continue

            inconsistency_count = len([a for a in anomalies if a['type'] == 'Incohérence dates affectation/livraison'])
            self.logger.info(f"Found {inconsistency_count} assignment/delivery date inconsistencies in sheet 1")

        except Exception as e:
            self.logger.error(f"Error checking assignment/delivery dates in sheet 1: {e}")

        return anomalies

    def _check_future_dates_sheet2(self, df, pd, today):
        """Check for future dates in sheet 2 (Traitement CMS Adr)."""
        anomalies = []
        try:
            from datetime import datetime

            # Colonnes de dates à vérifier dans la feuille 2
            date_columns = []
            for col_name in df.columns:
                if 'date' in col_name.lower():
                    date_columns.append(col_name)

            self.logger.info(f"Checking future dates in sheet 2 columns: {date_columns}")

            for index, row in df.iterrows():
                insee_code = row.get(df.columns[1], '') if len(df.columns) > 1 else ''  # Colonne B

                for col_name in date_columns:
                    if col_name in row.index:
                        date_value = row[col_name]

                        if pd.isna(date_value) or date_value == '':
                            continue

                        try:
                            # Essayer de parser la date
                            if isinstance(date_value, str):
                                for date_format in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
                                    try:
                                        date_obj = datetime.strptime(date_value, date_format).date()
                                        break
                                    except ValueError:
                                        continue
                                else:
                                    continue
                            else:
                                date_obj = pd.to_datetime(date_value).date()

                            # Vérifier si la date est dans le futur
                            if date_obj > today:
                                anomalies.append({
                                    'type': 'Date future',
                                    'feuille': 'Traitement CMS Adr',
                                    'ligne': index + 2,
                                    'code_insee': str(insee_code),
                                    'colonne': col_name,
                                    'date': str(date_value),
                                    'commentaire': 'Incohérence date'
                                })

                        except Exception as e:
                            self.logger.debug(f"Error parsing date {date_value} in sheet 2: {e}")
                            continue

            future_dates_count = len([a for a in anomalies if a['type'] == 'Date future' and a['feuille'] == 'Traitement CMS Adr'])
            self.logger.info(f"Found {future_dates_count} future date anomalies in sheet 2")

        except Exception as e:
            self.logger.error(f"Error checking future dates in sheet 2: {e}")

        return anomalies

    def _check_cm_treated_without_motif_sheet2(self, df, pd):
        """Check for CM tickets treated without motif voie in sheet 2."""
        anomalies = []
        try:
            self.logger.info("Checking CM tickets treated without motif voie in sheet 2")

            # Colonnes D (motif voie) et I (date de traitement ou statut traité)
            motif_col = None
            treatment_col = None

            # Trouver les colonnes par index (D = index 3, I = index 8)
            if len(df.columns) > 3:
                motif_col = df.columns[3]  # Colonne D
            if len(df.columns) > 8:
                treatment_col = df.columns[8]  # Colonne I

            if not motif_col or not treatment_col:
                self.logger.warning("Motif or treatment columns not found in sheet 2")
                return anomalies

            self.logger.info(f"Checking motif column: {motif_col}, treatment column: {treatment_col}")

            for index, row in df.iterrows():
                insee_code = row.get(df.columns[1], '') if len(df.columns) > 1 else ''  # Colonne B

                motif_value = row.get(motif_col, '')
                treatment_value = row.get(treatment_col, '')

                # Vérifier si le ticket est traité (colonne I non vide) mais motif voie vide (colonne D)
                is_treated = not (pd.isna(treatment_value) or treatment_value == '' or str(treatment_value).strip() == '')
                is_motif_empty = pd.isna(motif_value) or motif_value == '' or str(motif_value).strip() == ''

                if is_treated and is_motif_empty:
                    anomalies.append({
                        'type': 'Ticket CM traité sans motif voie',
                        'feuille': 'Traitement CMS Adr',
                        'ligne': index + 2,
                        'code_insee': str(insee_code),
                        'colonnes_concernees': f'{motif_col}, {treatment_col}',
                        'valeurs_problematiques': f'Motif: {motif_value}, Traitement: {treatment_value}',
                        'commentaire': f'Ticket marqué comme traité (colonne {treatment_col}: {treatment_value}) mais motif voie vide (colonne {motif_col})'
                    })

            cm_without_motif_count = len([a for a in anomalies if a['type'] == 'Ticket CM traité sans motif voie'])
            self.logger.info(f"Found {cm_without_motif_count} CM tickets treated without motif voie in sheet 2")

        except Exception as e:
            self.logger.error(f"Error checking CM tickets treated without motif voie in sheet 2: {e}")

        return anomalies

    def _check_future_dates_sheet3(self, df, pd, today):
        """Check for future dates in sheet 3 (Traitement PA)."""
        anomalies = []
        try:
            from datetime import datetime

            # Colonnes de dates à vérifier dans la feuille 3
            date_columns = []
            for col_name in df.columns:
                if 'date' in col_name.lower():
                    date_columns.append(col_name)

            self.logger.info(f"Checking future dates in sheet 3 columns: {date_columns}")

            for index, row in df.iterrows():
                insee_code = row.get(df.columns[1], '') if len(df.columns) > 1 else ''  # Colonne B

                for col_name in date_columns:
                    if col_name in row.index:
                        date_value = row[col_name]

                        if pd.isna(date_value) or date_value == '':
                            continue

                        try:
                            # Essayer de parser la date
                            if isinstance(date_value, str):
                                for date_format in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
                                    try:
                                        date_obj = datetime.strptime(date_value, date_format).date()
                                        break
                                    except ValueError:
                                        continue
                                else:
                                    continue
                            else:
                                date_obj = pd.to_datetime(date_value).date()

                            # Vérifier si la date est dans le futur
                            if date_obj > today:
                                anomalies.append({
                                    'type': 'Date future',
                                    'feuille': 'Traitement PA',
                                    'ligne': index + 2,
                                    'code_insee': str(insee_code),
                                    'colonne': col_name,
                                    'date': str(date_value),
                                    'commentaire': 'Incohérence date'
                                })

                        except Exception as e:
                            self.logger.debug(f"Error parsing date {date_value} in sheet 3: {e}")
                            continue

            future_dates_count = len([a for a in anomalies if a['type'] == 'Date future' and a['feuille'] == 'Traitement PA'])
            self.logger.info(f"Found {future_dates_count} future date anomalies in sheet 3")

        except Exception as e:
            self.logger.error(f"Error checking future dates in sheet 3: {e}")

        return anomalies

    def _check_unjoined_address_with_time_sheet3(self, df, pd):
        """Check for unjoined addresses with processing time > 0 in sheet 3."""
        anomalies = []
        try:
            self.logger.info("Checking unjoined addresses with processing time in sheet 3")

            # Colonnes D (statut adresse jointe) et H (temps de traitement)
            address_status_col = None
            processing_time_col = None

            # Trouver les colonnes par index (D = index 3, H = index 7)
            if len(df.columns) > 3:
                address_status_col = df.columns[3]  # Colonne D
            if len(df.columns) > 7:
                processing_time_col = df.columns[7]  # Colonne H

            if not address_status_col or not processing_time_col:
                self.logger.warning("Address status or processing time columns not found in sheet 3")
                return anomalies

            self.logger.info(f"Checking address status column: {address_status_col}, processing time column: {processing_time_col}")

            for index, row in df.iterrows():
                insee_code = row.get(df.columns[1], '') if len(df.columns) > 1 else ''  # Colonne B

                address_status_value = row.get(address_status_col, '')
                processing_time_value = row.get(processing_time_col, '')

                # Vérifier si l'adresse n'est pas jointe (colonne D = "Non" ou vide)
                is_address_not_joined = (
                    pd.isna(address_status_value) or
                    address_status_value == '' or
                    str(address_status_value).strip().lower() in ['non', 'no', '']
                )

                # Vérifier si le temps de traitement > 0
                processing_time_numeric = 0
                try:
                    if not (pd.isna(processing_time_value) or processing_time_value == ''):
                        processing_time_numeric = float(str(processing_time_value).replace(',', '.'))
                except (ValueError, TypeError):
                    processing_time_numeric = 0

                # Détecter l'anomalie : adresse non jointe mais temps de traitement > 0
                if is_address_not_joined and processing_time_numeric > 0:
                    anomalies.append({
                        'type': 'Adresse non jointe avec temps de traitement',
                        'feuille': 'Traitement PA',
                        'ligne': index + 2,
                        'code_insee': str(insee_code),
                        'colonnes_concernees': f'{address_status_col}, {processing_time_col}',
                        'valeurs_problematiques': f'Statut adresse: {address_status_value}, Temps: {processing_time_value}',
                        'commentaire': f'Adresse non jointe (colonne {address_status_col}: {address_status_value}) mais temps de traitement > 0 (colonne {processing_time_col}: {processing_time_value})'
                    })

            unjoined_with_time_count = len([a for a in anomalies if a['type'] == 'Adresse non jointe avec temps de traitement'])
            self.logger.info(f"Found {unjoined_with_time_count} unjoined addresses with processing time in sheet 3")

        except Exception as e:
            self.logger.error(f"Error checking unjoined addresses with processing time in sheet 3: {e}")

        return anomalies

    def _create_anomalies_excel(self, file_path, anomalies):
        """Create Excel file with detected anomalies using global tickets styling."""
        try:
            from utils.lazy_imports import get_pandas
            from datetime import datetime

            pd = get_pandas()

            # Préparer les données pour l'export
            export_data = []

            for anomaly in anomalies:
                if anomaly['type'] == 'Cases vides':
                    export_data.append({
                        'Type Anomalie': anomaly['type'],
                        'Feuille': anomaly['feuille'],
                        'Ligne': anomaly['ligne'],
                        'Code INSEE': anomaly['code_insee'],
                        'Colonnes Vides': anomaly.get('colonnes_vides', ''),
                        'Colonne Date': '',
                        'Date Problématique': '',
                        'Colonnes Concernées': '',
                        'Valeurs Problématiques': '',
                        'Commentaire': anomaly['commentaire']
                    })
                elif anomaly['type'] == 'Date future':
                    export_data.append({
                        'Type Anomalie': anomaly['type'],
                        'Feuille': anomaly['feuille'],
                        'Ligne': anomaly['ligne'],
                        'Code INSEE': anomaly['code_insee'],
                        'Colonnes Vides': '',
                        'Colonne Date': anomaly.get('colonne', ''),
                        'Date Problématique': anomaly.get('date', ''),
                        'Colonnes Concernées': '',
                        'Valeurs Problématiques': '',
                        'Commentaire': anomaly['commentaire']
                    })
                elif anomaly['type'] in ['Incohérence dates affectation/livraison', 'Ticket CM traité sans motif voie', 'Adresse non jointe avec temps de traitement']:
                    export_data.append({
                        'Type Anomalie': anomaly['type'],
                        'Feuille': anomaly['feuille'],
                        'Ligne': anomaly['ligne'],
                        'Code INSEE': anomaly['code_insee'],
                        'Colonnes Vides': '',
                        'Colonne Date': '',
                        'Date Problématique': '',
                        'Colonnes Concernées': anomaly.get('colonnes_concernees', ''),
                        'Valeurs Problématiques': anomaly.get('valeurs_problematiques', ''),
                        'Commentaire': anomaly['commentaire']
                    })

            # Créer le DataFrame
            df_anomalies = pd.DataFrame(export_data)

            # Trier par type d'anomalie puis par feuille
            df_anomalies = df_anomalies.sort_values(['Type Anomalie', 'Feuille', 'Ligne'])

            # Créer le fichier Excel avec formatage
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df_anomalies.to_excel(writer, sheet_name='Anomalies', index=False)

                # Obtenir la feuille pour le formatage
                worksheet = writer.sheets['Anomalies']

                # Apply global tickets styling
                self._apply_global_tickets_styling_to_anomalies(worksheet, df_anomalies, anomalies)

            self.logger.info(f"Anomalies Excel file created with global tickets styling: {file_path}")

        except Exception as e:
            self.logger.error(f"Error creating anomalies Excel file: {e}")
            raise

    def _apply_global_tickets_styling_to_anomalies(self, worksheet, df_anomalies, anomalies):
        """Apply the same styling as global tickets Excel file to anomalies export."""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            from datetime import datetime

            # Use the same styling as global tickets file (blue header with white text)
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')  # Blue froid
            header_font = Font(color='FFFFFF', bold=True)  # White text, bold
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)

            # Ajuster la largeur des colonnes
            column_widths = {
                'A': 20,  # Type Anomalie
                'B': 20,  # Feuille
                'C': 8,   # Ligne
                'D': 12,  # Code INSEE
                'E': 25,  # Colonnes Vides
                'F': 15,  # Colonne Date
                'G': 18,  # Date Problématique
                'H': 25,  # Colonnes Concernées
                'I': 30,  # Valeurs Problématiques
                'J': 35   # Commentaire
            }

            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width

            # Apply center alignment to all cells
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = center_alignment

            # Apply header styling to first row (same as global tickets)
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment

            # Freeze header row (same as global tickets)
            worksheet.freeze_panes = 'A2'

            # Formatage conditionnel pour les types d'anomalies (plus subtil)
            warning_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Light yellow
            error_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")    # Light red

            # Formatage conditionnel pour les nouveaux types d'anomalies
            inconsistency_fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")  # Light orange

            for row in range(2, len(df_anomalies) + 2):
                anomaly_type = worksheet[f'A{row}'].value
                if anomaly_type == 'Cases vides':
                    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
                        cell = worksheet[f'{col}{row}']
                        cell.fill = warning_fill
                        cell.alignment = center_alignment
                elif anomaly_type == 'Date future':
                    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
                        cell = worksheet[f'{col}{row}']
                        cell.fill = error_fill
                        cell.alignment = center_alignment
                elif anomaly_type in ['Incohérence dates affectation/livraison', 'Ticket CM traité sans motif voie', 'Adresse non jointe avec temps de traitement']:
                    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
                        cell = worksheet[f'{col}{row}']
                        cell.fill = inconsistency_fill
                        cell.alignment = center_alignment

            # Ajouter un résumé en bas avec le même style
            summary_row = len(df_anomalies) + 3
            worksheet[f'A{summary_row}'] = "RÉSUMÉ"
            worksheet[f'A{summary_row}'].font = Font(bold=True)
            worksheet[f'A{summary_row}'].alignment = center_alignment

            cases_vides_count = len([a for a in anomalies if a['type'] == 'Cases vides'])
            dates_futures_count = len([a for a in anomalies if a['type'] == 'Date future'])
            dates_inconsistency_count = len([a for a in anomalies if a['type'] == 'Incohérence dates affectation/livraison'])
            cm_without_motif_count = len([a for a in anomalies if a['type'] == 'Ticket CM traité sans motif voie'])
            unjoined_with_time_count = len([a for a in anomalies if a['type'] == 'Adresse non jointe avec temps de traitement'])

            summary_data = [
                f"Cases vides: {cases_vides_count}",
                f"Dates futures: {dates_futures_count}",
                f"Incohérences dates affectation/livraison: {dates_inconsistency_count}",
                f"Tickets CM traités sans motif voie: {cm_without_motif_count}",
                f"Adresses non jointes avec temps de traitement: {unjoined_with_time_count}",
                f"Total anomalies: {len(anomalies)}",
                f"Généré le: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
            ]

            for i, data in enumerate(summary_data):
                cell = worksheet[f'A{summary_row + 1 + i}']
                cell.value = data
                cell.alignment = center_alignment

            self.logger.info("Applied global tickets styling to anomalies export")

        except Exception as e:
            self.logger.error(f"Error applying global tickets styling to anomalies: {e}")
            # Don't raise - styling is optional









    def _open_generated_file(self, file_path: str):
        """Open the generated Excel file."""
        try:
            import subprocess
            import platform

            if platform.system() == 'Windows':
                os.startfile(file_path)
            elif platform.system() == 'Darwin':  # macOS
                subprocess.call(['open', file_path])
            else:  # Linux
                subprocess.call(['xdg-open', file_path])

            self.logger.info(f"Opened generated file: {file_path}")

        except Exception as e:
            self.logger.error(f"Error opening file: {e}")
            messagebox.showerror("Erreur", f"Impossible d'ouvrir le fichier:\n{e}")




















    def _count_treated_communes_current_month(self) -> int:
        """Count communes treated in the current month."""
        try:
            if not self.global_suivi_data or 'Suivi Tickets' not in self.global_suivi_data:
                return 0

            pd = get_pandas()
            from datetime import datetime

            df_tickets = self.global_suivi_data['Suivi Tickets']
            if df_tickets.empty:
                return 0

            current_month = datetime.now().month
            current_year = datetime.now().year

            # Find status and date columns
            status_column = None
            date_column = None

            for col in df_tickets.columns:
                col_lower = col.lower().strip()
                if 'etat' in col_lower and 'ticket' in col_lower:
                    status_column = col
                elif 'date' in col_lower and 'livraison' in col_lower:
                    date_column = col

            if not status_column:
                return 0

            # Count treated communes in current month
            treated_count = 0
            for _, row in df_tickets.iterrows():
                status = str(row.get(status_column, '')).strip()
                if status.lower() == 'traité':
                    # Check if delivery date is in current month
                    if date_column and pd.notna(row.get(date_column)):
                        try:
                            delivery_date = pd.to_datetime(row[date_column])
                            if delivery_date.month == current_month and delivery_date.year == current_year:
                                treated_count += 1
                        except:
                            # If date parsing fails, still count as treated
                            treated_count += 1
                    else:
                        # No date column or date is empty, still count as treated
                        treated_count += 1

            return treated_count

        except Exception as e:
            self.logger.error(f"Error counting treated communes: {e}")
            return 0

    def _create_monthly_archive(self):
        """Create monthly archive of treated communes with progress indicator."""
        try:
            if not self.global_suivi_data:
                messagebox.showwarning("Aucune donnée", "Veuillez d'abord charger les données.")
                return

            # Get selected month
            selected_month = self.archive_month_var.get() if hasattr(self, 'archive_month_var') else None
            if not selected_month:
                from datetime import datetime
                month_names = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                              "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
                current_month = datetime.now().month
                selected_month = month_names[current_month - 1]

            # Disable button during processing
            self.archive_button.config(state=tk.DISABLED)
            self.archive_status_label.config(text=f"Préparation de l'archive pour {selected_month}...", fg=COLORS['INFO'])

            # Get treated communes for selected month
            treated_communes = self._get_treated_communes_for_month(selected_month)

            if not treated_communes:
                messagebox.showinfo("Aucune commune", f"Aucune commune traitée trouvée pour {selected_month}.")
                self._enable_archive_button()
                return

            # Show confirmation dialog with details
            total_size = sum(commune.get('folder_size', 0) for commune in treated_communes)
            total_files = sum(commune.get('file_count', 0) for commune in treated_communes)

            confirm_msg = (
                f"Créer une archive avec {len(treated_communes)} communes ?\n\n"
                f"📁 Nombre total de fichiers: {total_files}\n"
                f"💾 Taille totale: {total_size/1024/1024:.1f} MB\n\n"
                f"L'archive sera sauvegardée sur D:\\BackUP Plan Adressage\\"
            )

            if not messagebox.askyesno("Confirmer l'archivage", confirm_msg):
                self._enable_archive_button()
                return

            # Create progress window
            progress_window = self._create_progress_window(treated_communes)

            # Create archive with progress tracking
            archive_path = self._create_archive_zip_with_progress(treated_communes, progress_window)

            # Close progress window
            if progress_window and progress_window.winfo_exists():
                progress_window.destroy()

            if archive_path:
                # Show success message with archive details
                archive_size = os.path.getsize(archive_path)
                compression_ratio = (1 - archive_size / total_size) * 100 if total_size > 0 else 0

                success_msg = (
                    f"✅ Archive créée avec succès!\n\n"
                    f"📁 Emplacement: {archive_path}\n"
                    f"📊 {len(treated_communes)} communes archivées\n"
                    f"💾 Taille finale: {archive_size/1024/1024:.1f} MB\n"
                    f"🗜️ Compression: {compression_ratio:.1f}%"
                )

                messagebox.showinfo("Archive créée", success_msg)

                self.archive_status_label.config(
                    text=f"✅ Archive créée: {len(treated_communes)} communes ({archive_size/1024/1024:.1f} MB)",
                    fg=COLORS['SUCCESS']
                )

                # Ask if user wants to open the backup folder
                if messagebox.askyesno("Ouvrir le dossier", "Voulez-vous ouvrir le dossier de sauvegarde ?"):
                    try:
                        import subprocess
                        subprocess.Popen(f'explorer "{os.path.dirname(archive_path)}"')
                    except Exception as e:
                        self.logger.warning(f"Could not open backup folder: {e}")
            else:
                messagebox.showerror("Erreur", "Erreur lors de la création de l'archive.")
                self._enable_archive_button()

        except Exception as e:
            self.logger.error(f"Error creating monthly archive: {e}")
            messagebox.showerror("Erreur", f"Erreur lors de la création de l'archive:\n{e}")
            self._enable_archive_button()

    def _create_progress_window(self, treated_communes: list):
        """Create a progress window for archive creation."""
        try:
            # Create progress window
            progress_window = tk.Toplevel(self.parent)
            progress_window.title("Création d'archive en cours...")
            progress_window.geometry("500x300")
            progress_window.resizable(False, False)
            progress_window.transient(self.parent)
            progress_window.grab_set()

            # Center the window
            progress_window.update_idletasks()
            x = (progress_window.winfo_screenwidth() // 2) - (500 // 2)
            y = (progress_window.winfo_screenheight() // 2) - (300 // 2)
            progress_window.geometry(f"500x300+{x}+{y}")

            # Main frame
            main_frame = tk.Frame(progress_window, bg='white', padx=20, pady=20)
            main_frame.pack(fill=tk.BOTH, expand=True)

            # Title
            title_label = tk.Label(
                main_frame,
                text="📦 Création d'archive en cours",
                font=("Arial", 16, "bold"),
                fg="#0066cc",
                bg="white"
            )
            title_label.pack(pady=(0, 20))

            # Progress info
            progress_window.info_label = tk.Label(
                main_frame,
                text="Initialisation...",
                font=("Arial", 10),
                fg="#333333",
                bg="white"
            )
            progress_window.info_label.pack(pady=(0, 10))

            # Progress bar
            progress_window.progress_var = tk.DoubleVar()
            progress_window.progress_bar = ttk.Progressbar(
                main_frame,
                variable=progress_window.progress_var,
                maximum=100,
                length=400,
                mode='determinate'
            )
            progress_window.progress_bar.pack(pady=(0, 10))

            # Progress percentage
            progress_window.percent_label = tk.Label(
                main_frame,
                text="0%",
                font=("Arial", 12, "bold"),
                fg="#0066cc",
                bg="white"
            )
            progress_window.percent_label.pack(pady=(0, 20))

            # Details frame
            details_frame = tk.Frame(main_frame, bg="white")
            details_frame.pack(fill=tk.X, pady=(0, 10))

            # Current commune
            progress_window.current_label = tk.Label(
                details_frame,
                text="Commune actuelle: -",
                font=("Arial", 9),
                fg="#666666",
                bg="white",
                anchor="w"
            )
            progress_window.current_label.pack(fill=tk.X)

            # Files processed
            progress_window.files_label = tk.Label(
                details_frame,
                text="Fichiers traités: 0",
                font=("Arial", 9),
                fg="#666666",
                bg="white",
                anchor="w"
            )
            progress_window.files_label.pack(fill=tk.X)

            # Estimated time
            progress_window.time_label = tk.Label(
                details_frame,
                text="Temps estimé: Calcul...",
                font=("Arial", 9),
                fg="#666666",
                bg="white",
                anchor="w"
            )
            progress_window.time_label.pack(fill=tk.X)

            # Cancel button (disabled for now)
            cancel_button = tk.Button(
                main_frame,
                text="Annuler",
                font=UIConfig.FONT_BUTTON,
                bg=COLORS['DANGER'],
                fg="white",
                state=tk.DISABLED,
                relief=tk.FLAT,
                borderwidth=0,
                padx=20,
                pady=5
            )
            cancel_button.pack(pady=(10, 0))

            progress_window.update()
            return progress_window

        except Exception as e:
            self.logger.error(f"Error creating progress window: {e}")
            return None

    def _create_archive_zip_with_progress(self, treated_communes: list, progress_window) -> str:
        """Create ZIP archive with progress tracking."""
        try:
            import zipfile
            from datetime import datetime
            import time

            # Create backup directory on D:\ drive
            backup_dir = "D:\\BackUP Plan Adressage"
            os.makedirs(backup_dir, exist_ok=True)

            # Generate archive filename using selected month
            selected_month = self.archive_month_var.get() if hasattr(self, 'archive_month_var') else None
            if not selected_month:
                current_date = datetime.now()
                month_name = current_date.strftime("%B")  # Full month name
            else:
                month_name = selected_month
            commune_count = len(treated_communes)

            archive_filename = f"{commune_count}.CommunesTraités_{month_name}.zip"
            archive_path = os.path.join(backup_dir, archive_filename)

            self.logger.info(f"Creating archive with progress: {archive_filename}")

            # Calculate totals for progress tracking
            total_size = sum(commune.get('folder_size', 0) for commune in treated_communes)
            total_files = sum(commune.get('file_count', 0) for commune in treated_communes)
            processed_size = 0
            processed_files = 0
            start_time = time.time()

            # Update progress window
            if progress_window and progress_window.winfo_exists():
                progress_window.info_label.config(text=f"Création de l'archive: {archive_filename}")
                progress_window.files_label.config(text=f"Fichiers à traiter: {total_files}")
                progress_window.update()

            # Create ZIP file with optimal compression
            with zipfile.ZipFile(archive_path, 'w', zipfile.ZIP_DEFLATED, compresslevel=6) as zipf:

                # Add each commune's complete folder
                for i, commune in enumerate(treated_communes):
                    commune_name = commune['nom']
                    folder_path = commune.get('folder_path')

                    if not folder_path or not os.path.exists(folder_path):
                        self.logger.warning(f"❌ Folder not found for commune {commune_name}: {folder_path}")
                        continue

                    # Update progress window
                    if progress_window and progress_window.winfo_exists():
                        progress_window.current_label.config(text=f"Commune actuelle: {commune_name}")
                        progress_window.update()

                    self.logger.info(f"📁 Adding commune {i+1}/{commune_count}: {commune_name}")

                    try:
                        # Add complete folder recursively with progress
                        added_files = self._add_folder_to_zip_with_progress(
                            zipf, folder_path, commune_name, progress_window
                        )

                        # Add separate commune files if any
                        separate_files = commune.get('separate_files', [])
                        if separate_files:
                            for separate_file in separate_files:
                                try:
                                    # Add separate file to commune folder in archive
                                    file_name = os.path.basename(separate_file)
                                    archive_path = f"{commune_name}/Fichiers_Separes/{file_name}"
                                    zipf.write(separate_file, archive_path)
                                    added_files += 1
                                except Exception as e:
                                    self.logger.warning(f"Could not add separate file {separate_file}: {e}")

                        # Update progress
                        processed_size += commune.get('folder_size', 0)
                        processed_files += added_files
                        progress = (i + 1) / commune_count * 90  # Reserve 10% for final steps

                        # Calculate estimated time
                        elapsed_time = time.time() - start_time
                        if i > 0:
                            avg_time_per_commune = elapsed_time / (i + 1)
                            remaining_communes = commune_count - (i + 1)
                            estimated_remaining = avg_time_per_commune * remaining_communes
                            time_str = f"{estimated_remaining:.0f}s restantes"
                        else:
                            time_str = "Calcul..."

                        # Update progress window
                        if progress_window and progress_window.winfo_exists():
                            progress_window.progress_var.set(progress)
                            progress_window.percent_label.config(text=f"{progress:.1f}%")
                            progress_window.files_label.config(text=f"Fichiers traités: {processed_files}/{total_files}")
                            progress_window.time_label.config(text=f"Temps estimé: {time_str}")
                            progress_window.update()

                        total_files_msg = f"{added_files} fichiers"
                        if separate_files:
                            total_files_msg += f" (+ {len(separate_files)} séparés)"

                        self.logger.info(f"✅ Added {total_files_msg} for {commune_name} "
                                       f"(Progress: {progress:.1f}%)")

                    except Exception as e:
                        self.logger.error(f"❌ Error adding folder for {commune_name}: {e}")
                        continue

                # Add global tracking file
                if progress_window and progress_window.winfo_exists():
                    progress_window.current_label.config(text="Ajout du fichier de suivi global...")
                    progress_window.progress_var.set(95)
                    progress_window.percent_label.config(text="95%")
                    progress_window.update()

                self.logger.info("📊 Adding global tracking file...")
                self._add_treated_communes_tracking_to_zip(zipf, treated_communes)

            # Final progress update
            if progress_window and progress_window.winfo_exists():
                progress_window.progress_var.set(100)
                progress_window.percent_label.config(text="100%")
                progress_window.current_label.config(text="Archive créée avec succès!")
                progress_window.update()

            # Get final archive size and stats
            archive_size = os.path.getsize(archive_path)
            compression_ratio = (1 - archive_size / total_size) * 100 if total_size > 0 else 0
            total_time = time.time() - start_time

            self.logger.info(f"✅ Archive created successfully: {archive_path}")
            self.logger.info(f"📊 Archive size: {archive_size/1024/1024:.1f} MB "
                           f"(Compression: {compression_ratio:.1f}%)")
            self.logger.info(f"⏱️ Total time: {total_time:.1f}s")

            return archive_path

        except Exception as e:
            self.logger.error(f"❌ Error creating archive ZIP with progress: {e}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            return None

    def _add_folder_to_zip_with_progress(self, zipf, folder_path: str, commune_name: str, progress_window) -> int:
        """Add a complete folder to ZIP with progress updates - copy everything."""
        try:
            added_files = 0
            error_files = 0

            # Walk through ALL files and subdirectories without filtering
            for root, dirs, files in os.walk(folder_path):
                # Calculate relative path from the commune folder
                rel_path = os.path.relpath(root, folder_path)

                # Create the archive path structure
                if rel_path == '.':
                    archive_dir = commune_name
                else:
                    archive_dir = f"{commune_name}/{rel_path.replace(os.sep, '/')}"

                # Add ALL files in current directory without any filtering
                for file in files:
                    file_path = os.path.join(root, file)

                    try:
                        # Create archive path
                        archive_file_path = f"{archive_dir}/{file}"

                        # Add file to ZIP - copy everything without analysis
                        zipf.write(file_path, archive_file_path)
                        added_files += 1

                        # Update progress every 25 files
                        if added_files % 25 == 0 and progress_window and progress_window.winfo_exists():
                            try:
                                progress_window.update()
                            except:
                                pass  # Ignore UI update errors

                    except (OSError, IOError, PermissionError) as e:
                        error_files += 1
                        self.logger.warning(f"Could not access file {file_path}: {e}")
                        continue
                    except Exception as e:
                        error_files += 1
                        self.logger.warning(f"Error adding file {file_path}: {e}")
                        continue

                # Add empty directories to preserve complete structure
                if not files and not dirs:
                    try:
                        zipf.writestr(f"{archive_dir}/", "")
                    except Exception as e:
                        self.logger.debug(f"Could not add empty directory {archive_dir}: {e}")

            if error_files > 0:
                self.logger.info(f"Folder {commune_name}: {added_files} files added, {error_files} errors")

            return added_files

        except Exception as e:
            self.logger.error(f"Error adding folder to ZIP with progress: {e}")
            return 0

    def _get_treated_communes_for_month(self, month_name: str) -> list:
        """Get list of treated communes for specified month with their file paths."""
        try:
            pd = get_pandas()
            from datetime import datetime

            treated_communes = []

            if 'Suivi Tickets' not in self.global_suivi_data:
                return treated_communes

            df_tickets = self.global_suivi_data['Suivi Tickets']
            if df_tickets.empty:
                return treated_communes

            # Convert month name to month number
            month_names = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                          "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
            try:
                target_month = month_names.index(month_name) + 1
            except ValueError:
                self.logger.error(f"Invalid month name: {month_name}")
                return treated_communes

            current_year = datetime.now().year

            # Find required columns
            status_column = None
            date_column = None
            commune_column = None
            insee_column = None
            id_tache_column = None

            for col in df_tickets.columns:
                col_lower = col.lower().strip()
                if 'etat' in col_lower and 'ticket' in col_lower:
                    status_column = col
                elif 'date' in col_lower and 'livraison' in col_lower:
                    date_column = col
                elif 'nom' in col_lower and 'commune' in col_lower:
                    commune_column = col
                elif 'insee' in col_lower or 'code' in col_lower:
                    insee_column = col
                elif 'id' in col_lower and 'tâche' in col_lower and 'plan' in col_lower:
                    id_tache_column = col

            if not all([status_column, commune_column]):
                self.logger.warning("Required columns not found for archive creation")
                return treated_communes

            # Process each row
            for _, row in df_tickets.iterrows():
                status = str(row.get(status_column, '')).strip()
                if status.lower() == 'traité':
                    # Check if delivery date is in target month
                    is_target_month = True
                    if date_column and pd.notna(row.get(date_column)):
                        try:
                            delivery_date = pd.to_datetime(row[date_column])
                            is_target_month = (delivery_date.month == target_month and
                                             delivery_date.year == current_year)
                        except:
                            pass  # Keep as target month if date parsing fails

                    if is_target_month:
                        commune_name = str(row.get(commune_column, '')).strip()
                        insee_code = str(row.get(insee_column, '')).strip()
                        id_tache = str(row.get(id_tache_column, '')).strip() if id_tache_column else ''

                        if commune_name:
                            commune_info = {
                                'nom': commune_name,
                                'insee': insee_code,
                                'id_tache': id_tache,
                                'status': status,
                                'date_livraison': row.get(date_column, ''),
                                'file_path': None,  # Will be found later
                                'folder_path': None  # Will be found later
                            }
                            treated_communes.append(commune_info)

            # Find file paths for each commune
            self._find_commune_file_paths(treated_communes)

            # Find separate commune files
            self._find_commune_separate_files(treated_communes)

            return treated_communes

        except Exception as e:
            self.logger.error(f"Error getting treated communes for {month_name}: {e}")
            return []

    def _get_treated_communes_current_month(self) -> list:
        """Get list of treated communes for current month with their file paths."""
        try:
            pd = get_pandas()
            from datetime import datetime

            treated_communes = []

            if 'Suivi Tickets' not in self.global_suivi_data:
                return treated_communes

            df_tickets = self.global_suivi_data['Suivi Tickets']
            if df_tickets.empty:
                return treated_communes

            current_month = datetime.now().month
            current_year = datetime.now().year

            # Find required columns
            status_column = None
            date_column = None
            commune_column = None
            insee_column = None
            id_tache_column = None

            for col in df_tickets.columns:
                col_lower = col.lower().strip()
                if 'etat' in col_lower and 'ticket' in col_lower:
                    status_column = col
                elif 'date' in col_lower and 'livraison' in col_lower:
                    date_column = col
                elif 'nom' in col_lower and 'commune' in col_lower:
                    commune_column = col
                elif 'insee' in col_lower or 'code' in col_lower:
                    insee_column = col
                elif 'id' in col_lower and 'tâche' in col_lower and 'plan' in col_lower:
                    id_tache_column = col

            if not all([status_column, commune_column]):
                self.logger.warning("Required columns not found for archive creation")
                return treated_communes

            # Process each row
            for _, row in df_tickets.iterrows():
                status = str(row.get(status_column, '')).strip()
                if status.lower() == 'traité':
                    # Check if delivery date is in current month
                    is_current_month = True
                    if date_column and pd.notna(row.get(date_column)):
                        try:
                            delivery_date = pd.to_datetime(row[date_column])
                            is_current_month = (delivery_date.month == current_month and
                                              delivery_date.year == current_year)
                        except:
                            pass  # Keep as current month if date parsing fails

                    if is_current_month:
                        commune_name = str(row.get(commune_column, '')).strip()
                        insee_code = str(row.get(insee_column, '')).strip()
                        id_tache = str(row.get(id_tache_column, '')).strip() if id_tache_column else ''

                        if commune_name:
                            commune_info = {
                                'nom': commune_name,
                                'insee': insee_code,
                                'id_tache': id_tache,
                                'status': status,
                                'date_livraison': row.get(date_column, ''),
                                'file_path': None,  # Will be found later
                                'folder_path': None  # Will be found later
                            }
                            treated_communes.append(commune_info)

            # Find file paths for each commune
            self._find_commune_file_paths(treated_communes)

            # Find separate commune files
            self._find_commune_separate_files(treated_communes)

            return treated_communes

        except Exception as e:
            self.logger.error(f"Error getting treated communes: {e}")
            return []

    def _find_commune_file_paths(self, treated_communes: list):
        """Find complete folder paths for treated communes using exact nom+ID matching."""
        try:
            # CORRECTION: Les dossiers communes sont directement dans le dossier Teams configuré
            # qui pointe déjà vers "Actes des Traitements" via TeamsConfig.get_teams_base_path()
            from config.constants import TeamsConfig

            # Le chemin Teams pointe déjà vers le bon dossier "Actes des Traitements"
            actes_path = TeamsConfig.get_teams_base_path()

            if not os.path.exists(actes_path):
                self.logger.warning(f"Actes des Traitements folder not found: {actes_path}")
                return

            self.logger.info(f"Scanning Actes des Traitements folder for {len(treated_communes)} treated communes...")
            self.logger.info(f"Search path: {actes_path}")

            # Create a map of all available folders for better matching
            available_folders = {}
            for item in os.listdir(actes_path):
                item_path = os.path.join(actes_path, item)
                if os.path.isdir(item_path):
                    available_folders[item.lower()] = {
                        'original_name': item,
                        'path': item_path
                    }

            self.logger.info(f"Found {len(available_folders)} folders in Actes des Traitements directory")

            # Debug analysis (can be removed in production)
            self._debug_available_folders(available_folders, treated_communes)

            # Process each commune with exact matching logic
            for commune in treated_communes:
                commune_name = commune['nom']
                insee_code = commune['insee']
                id_tache = commune['id_tache']

                # Create unique identifier for logging
                commune_id = f"{commune_name} (ID: {id_tache})"
                self.logger.info(f"🔍 Searching for folder: {commune_id}")

                # Initialize paths
                commune['file_path'] = None
                commune['folder_path'] = None
                commune['folder_size'] = 0
                commune['file_count'] = 0
                commune['matched_folder_name'] = None

                # Try exact matching strategies in order of priority
                found_folder = self._find_exact_folder_match(
                    commune_name, id_tache, insee_code, available_folders
                )

                if found_folder:
                    folder_path = found_folder['path']
                    folder_name = found_folder['original_name']

                    commune['folder_path'] = folder_path
                    commune['matched_folder_name'] = folder_name

                    self.logger.info(f"✅ Found exact match for {commune_id}: {folder_name}")

                    # Analyze folder contents
                    try:
                        folder_info = self._analyze_folder_contents(folder_path)
                        commune['file_count'] = folder_info['file_count']
                        commune['folder_size'] = folder_info['total_size']
                        commune['has_excel'] = folder_info['has_excel']
                        commune['main_excel_file'] = folder_info['main_excel_file']

                        if folder_info['main_excel_file']:
                            commune['file_path'] = folder_info['main_excel_file']

                        self.logger.info(f"📊 Folder analysis: {folder_info['file_count']} files, "
                                       f"{folder_info['total_size']/1024/1024:.1f} MB")

                    except Exception as e:
                        self.logger.warning(f"Error analyzing folder {folder_path}: {e}")
                        # Still keep the folder path even if analysis fails
                        commune['file_count'] = 0
                        commune['folder_size'] = 0
                else:
                    self.logger.warning(f"❌ No matching folder found for {commune_id}")
                    self.logger.debug(f"Expected folder formats: {commune_name}_{id_tache}")

                    # Log available folders that might be similar for debugging
                    similar_folders = self._find_similar_folders(commune_name, available_folders)
                    if similar_folders:
                        self.logger.debug(f"Similar folders found: {', '.join(similar_folders)}")

                    # Log all available folders that contain the commune name (for debugging)
                    commune_name_clean = commune_name.upper().replace(' ', '').replace('-', '').replace('_', '')
                    matching_folders = []
                    for folder_key, folder_info in available_folders.items():
                        folder_clean = folder_key.upper().replace(' ', '').replace('-', '').replace('_', '')
                        if commune_name_clean in folder_clean or folder_clean.startswith(commune_name_clean[:5]):
                            matching_folders.append(folder_info['original_name'])

                    if matching_folders:
                        self.logger.warning(f"🔍 Potential matches for {commune_name}: {', '.join(matching_folders)}")
                    else:
                        self.logger.warning(f"🔍 No potential matches found for {commune_name} in {len(available_folders)} available folders")

        except Exception as e:
            self.logger.error(f"Error finding commune file paths: {e}")

    def _debug_available_folders(self, available_folders: dict, target_communes: list):
        """Debug method to analyze available folders and target communes."""
        try:
            self.logger.info("=== FOLDER MATCHING DEBUG ===")
            self.logger.info(f"Total available folders: {len(available_folders)}")
            self.logger.info(f"Target communes: {len(target_communes)}")

            # Log first 10 available folders as examples
            self.logger.info("Sample available folders:")
            for i, (key, info) in enumerate(list(available_folders.items())[:10]):
                self.logger.info(f"  {i+1}. Key: '{key}' -> Original: '{info['original_name']}'")

            # Log target communes with spaces
            communes_with_spaces = [c for c in target_communes if ' ' in c['nom']]
            self.logger.info(f"Communes with spaces in name: {len(communes_with_spaces)}")
            for commune in communes_with_spaces:
                self.logger.info(f"  - {commune['nom']} (ID: {commune['id_tache']})")

            # Try to find pattern matches
            self.logger.info("Pattern analysis:")
            for commune in target_communes[:5]:  # Test first 5
                commune_name = commune['nom']
                id_tache = commune['id_tache']

                # Test different patterns
                patterns = [
                    f"{commune_name}_{id_tache}".upper(),
                    f"{commune_name.replace(' ', '_')}_{id_tache}".upper(),
                    f"{commune_name.replace(' ', '')}_{id_tache}".upper(),
                    f"{commune_name.replace(' ', '_').replace('-', '_')}_{id_tache}".upper(),
                ]

                self.logger.info(f"Testing patterns for {commune_name} (ID: {id_tache}):")
                for pattern in patterns:
                    pattern_lower = pattern.lower()
                    found = pattern_lower in available_folders
                    self.logger.info(f"  Pattern: '{pattern}' -> Found: {found}")
                    if found:
                        self.logger.info(f"    Matched folder: {available_folders[pattern_lower]['original_name']}")

            self.logger.info("=== END DEBUG ===")

        except Exception as e:
            self.logger.error(f"Error in debug analysis: {e}")

    def _find_exact_folder_match(self, commune_name: str, id_tache: str, insee_code: str, available_folders: dict) -> dict:
        """Find exact folder match using multiple strategies with priority order."""
        try:
            # CORRECTION: Amélioration du matching pour les noms comme ORIGNAC_2838560, CAMPUZAN_2838656

            # Strategy 1: Exact format nomcommune_idtache (HIGHEST PRIORITY)
            if id_tache:
                # Nettoyer les noms pour le matching exact
                clean_commune = self._clean_commune_name_for_matching(commune_name)
                clean_id = str(id_tache).strip()

                exact_format = f"{clean_commune}_{clean_id}".lower()

                # Recherche directe
                if exact_format in available_folders:
                    self.logger.debug(f"✅ Strategy 1 success: Exact format {exact_format}")
                    return available_folders[exact_format]

                # Strategy 2: Recherche avec variations de casse et caractères (AMÉLIORÉE)
                variations = [
                    # Variations de base
                    f"{clean_commune}_{clean_id}".upper(),
                    f"{clean_commune}_{clean_id}".lower(),
                    f"{clean_commune.upper()}_{clean_id}",
                    f"{clean_commune.lower()}_{clean_id}",

                    # Variations avec espaces/underscores
                    f"{commune_name.replace(' ', '_').replace('-', '_')}_{clean_id}".upper(),
                    f"{commune_name.replace(' ', '_').replace('-', '_')}_{clean_id}".lower(),
                    f"{commune_name.replace(' ', '').replace('-', '')}_{clean_id}".upper(),
                    f"{commune_name.replace(' ', '').replace('-', '')}_{clean_id}".lower(),

                    # Variations avec espaces préservés (au cas où)
                    f"{commune_name}_{clean_id}".upper(),
                    f"{commune_name}_{clean_id}".lower(),

                    # Variations mixtes
                    f"{commune_name.upper().replace(' ', '_').replace('-', '_')}_{clean_id}",
                    f"{commune_name.lower().replace(' ', '_').replace('-', '_')}_{clean_id}",
                    f"{commune_name.upper().replace(' ', '').replace('-', '')}_{clean_id}",
                    f"{commune_name.lower().replace(' ', '').replace('-', '')}_{clean_id}",

                    # Variations avec caractères spéciaux nettoyés
                    f"{self._clean_commune_name_for_matching(commune_name).upper()}_{clean_id}",
                    f"{self._clean_commune_name_for_matching(commune_name).lower()}_{clean_id}",
                ]

                # Supprimer les doublons tout en préservant l'ordre
                seen = set()
                unique_variations = []
                for var in variations:
                    if var not in seen:
                        seen.add(var)
                        unique_variations.append(var)

                # Log des variations testées pour debug
                self.logger.debug(f"Testing {len(unique_variations)} variations for {commune_name} (ID: {clean_id}):")
                for i, var in enumerate(unique_variations[:5]):  # Log seulement les 5 premières
                    self.logger.debug(f"  Variation {i+1}: {var}")

                variations = unique_variations

                for variation in variations:
                    variation_lower = variation.lower()
                    if variation_lower in available_folders:
                        self.logger.debug(f"✅ Strategy 2 success: Format variation {variation}")
                        return available_folders[variation_lower]

                # Strategy 3: Recherche par correspondance exacte dans les clés existantes
                commune_lower = clean_commune.lower()
                id_lower = clean_id.lower()

                for folder_key, folder_info in available_folders.items():
                    # Vérifier si le dossier correspond exactement au pattern nom_id
                    if '_' in folder_key:
                        parts = folder_key.split('_')
                        if len(parts) >= 2:
                            folder_commune = '_'.join(parts[:-1])  # Tout sauf le dernier élément
                            folder_id = parts[-1]  # Dernier élément

                            if (folder_commune == commune_lower and folder_id == id_lower):
                                self.logger.debug(f"✅ Strategy 3 success: Exact pattern match {folder_key}")
                                return folder_info

                # Strategy 4: Recherche contenant les deux éléments (plus permissive)
                for folder_key, folder_info in available_folders.items():
                    if commune_lower in folder_key and id_lower in folder_key:
                        # Vérifier que c'est bien un match valide (pas juste une sous-chaîne)
                        if (folder_key.startswith(commune_lower + '_') or
                            f'_{commune_lower}_' in folder_key or
                            folder_key.endswith('_' + id_lower)):
                            self.logger.debug(f"✅ Strategy 4 success: Contains both name and ID")
                            return folder_info

            # Strategy 5: Recherche flexible avec correspondance partielle (NOUVEAU)
            if id_tache:
                # Créer des versions nettoyées du nom de commune
                commune_variants = [
                    commune_name.upper(),
                    commune_name.lower(),
                    commune_name.upper().replace(' ', '_'),
                    commune_name.upper().replace(' ', ''),
                    self._clean_commune_name_for_matching(commune_name).upper(),
                    self._clean_commune_name_for_matching(commune_name).lower(),
                ]

                # Supprimer les doublons
                commune_variants = list(set(commune_variants))

                for variant in commune_variants:
                    for folder_key, folder_info in available_folders.items():
                        # Vérifier si le dossier contient le nom de commune et l'ID
                        if (variant.lower() in folder_key.lower() and
                            str(id_tache) in folder_key):
                            # Vérification supplémentaire pour éviter les faux positifs
                            folder_parts = folder_key.replace('_', ' ').replace('-', ' ').split()
                            commune_parts = variant.replace('_', ' ').replace('-', ' ').split()

                            # Vérifier que tous les mots de la commune sont dans le dossier
                            if all(part.lower() in [fp.lower() for fp in folder_parts] for part in commune_parts):
                                self.logger.debug(f"✅ Strategy 5 success: Flexible partial match")
                                self.logger.debug(f"   Commune variant: {variant}")
                                self.logger.debug(f"   Matched folder: {folder_info['original_name']}")
                                return folder_info

            # Strategy 6: Fallback to INSEE code if available (LOWER PRIORITY)
            if insee_code:
                insee_clean = str(insee_code).strip()
                for folder_key, folder_info in available_folders.items():
                    if insee_clean.lower() in folder_key:
                        self.logger.debug(f"✅ Strategy 6 success: INSEE code match")
                        return folder_info

            return None

        except Exception as e:
            self.logger.error(f"Error in exact folder matching: {e}")
            return None

    def _clean_commune_name_for_matching(self, commune_name: str) -> str:
        """Clean commune name for better matching with folder names."""
        try:
            if not commune_name:
                return ""

            # Nettoyer le nom de commune pour le matching
            cleaned = commune_name.strip()

            # Remplacer les caractères problématiques
            replacements = {
                ' ': '_',
                '-': '_',
                "'": '',
                'é': 'e',
                'è': 'e',
                'ê': 'e',
                'à': 'a',
                'ç': 'c',
                'ô': 'o',
                'û': 'u',
                'î': 'i',
                'ù': 'u',
                'â': 'a',
                'ë': 'e',
                'ï': 'i',
                'ü': 'u',
                'ÿ': 'y',
                'ñ': 'n'
            }

            for old, new in replacements.items():
                cleaned = cleaned.replace(old, new)

            # Supprimer les caractères spéciaux restants
            import re
            cleaned = re.sub(r'[^a-zA-Z0-9_]', '', cleaned)

            # Supprimer les underscores multiples
            cleaned = re.sub(r'_+', '_', cleaned)

            # Supprimer les underscores en début/fin
            cleaned = cleaned.strip('_')

            return cleaned

        except Exception as e:
            self.logger.warning(f"Error cleaning commune name '{commune_name}': {e}")
            return commune_name.strip()

    def _find_similar_folders(self, commune_name: str, available_folders: dict) -> list:
        """Find folders with similar names for debugging purposes."""
        try:
            similar = []
            commune_lower = commune_name.lower()

            for folder_key, folder_info in available_folders.items():
                if commune_lower in folder_key or any(word in folder_key for word in commune_lower.split()):
                    similar.append(folder_info['original_name'])

            return similar[:5]  # Return max 5 for readability

        except Exception as e:
            self.logger.debug(f"Error finding similar folders: {e}")
            return []

    def _find_commune_separate_files(self, treated_communes: list):
        """Find separate commune files that might exist outside the main folders."""
        try:
            # CORRECTION: Utiliser le bon chemin Teams qui pointe déjà vers "Actes des Traitements"
            from config.constants import TeamsConfig

            actes_path = TeamsConfig.get_teams_base_path()
            if not os.path.exists(actes_path):
                return

            # Chercher aussi dans le dossier parent (global Teams path)
            global_teams_path = TeamsConfig.get_global_teams_path()
            search_paths = [actes_path]
            if os.path.exists(global_teams_path) and global_teams_path != actes_path:
                search_paths.append(global_teams_path)

            self.logger.info("🔍 Searching for separate commune files...")
            self.logger.info(f"Search paths: {search_paths}")

            # Look for Excel files directly in Teams folder that match commune patterns
            for commune in treated_communes:
                commune_name = commune['nom']
                id_tache = commune['id_tache']
                insee_code = commune['insee']

                commune['separate_files'] = []

                # Search for files matching patterns
                patterns_to_search = [
                    f"{commune_name}_{id_tache}*.xlsx",
                    f"{commune_name}_*{id_tache}*.xlsx",
                    f"*{commune_name}*{id_tache}*.xlsx",
                ]

                if insee_code:
                    patterns_to_search.extend([
                        f"{commune_name}*{insee_code}*.xlsx",
                        f"*{commune_name}*{insee_code}*.xlsx"
                    ])

                # Search in all specified paths
                for search_path in search_paths:
                    for root, _, files in os.walk(search_path):
                        # Limit search to 2 levels deep to avoid performance issues
                        level = root.replace(search_path, '').count(os.sep)
                        if level > 1:
                            continue

                    for file in files:
                        if file.endswith('.xlsx') and not file.startswith('~'):
                            file_path = os.path.join(root, file)
                            file_lower = file.lower()
                            commune_lower = commune_name.lower()

                            # Check if file matches commune patterns
                            matches = False

                            # Exact pattern matching
                            if id_tache:
                                id_lower = id_tache.lower()
                                if (commune_lower in file_lower and id_lower in file_lower):
                                    matches = True

                            # INSEE pattern matching
                            if not matches and insee_code:
                                if (commune_lower in file_lower and insee_code in file_lower):
                                    matches = True

                            # Add file if it matches and is not already in the main folder
                            if matches:
                                # Check if this file is not already covered by the main folder
                                main_folder = commune.get('folder_path', '')
                                if not main_folder or not file_path.startswith(main_folder):
                                    commune['separate_files'].append(file_path)
                                    self.logger.debug(f"Found separate file for {commune_name}: {file}")

                if commune['separate_files']:
                    self.logger.info(f"📄 Found {len(commune['separate_files'])} separate files for {commune_name}")

        except Exception as e:
            self.logger.error(f"Error finding separate commune files: {e}")

    def _analyze_folder_contents(self, folder_path: str) -> dict:
        """Quick analysis of folder contents without reading file contents."""
        try:
            info = {
                'file_count': 0,
                'total_size': 0,
                'has_excel': False,
                'main_excel_file': None,
                'file_types': set(),
                'subdirs': []
            }

            # Quick walk through all files and subdirectories
            for root, dirs, files in os.walk(folder_path):
                info['subdirs'].extend(dirs)

                for file in files:
                    file_path = os.path.join(root, file)
                    try:
                        # Quick size check without opening file
                        file_size = os.path.getsize(file_path)
                        info['total_size'] += file_size
                        info['file_count'] += 1

                        # Track file extensions for statistics only
                        _, ext = os.path.splitext(file)
                        if ext:
                            info['file_types'].add(ext.lower())

                        # Look for Excel files for main file identification
                        if file.endswith('.xlsx') and not info['main_excel_file']:
                            info['has_excel'] = True
                            info['main_excel_file'] = file_path

                    except (OSError, IOError):
                        # Count file even if we can't get size
                        info['file_count'] += 1
                        continue

            return info

        except Exception as e:
            self.logger.error(f"Error analyzing folder contents: {e}")
            return {
                'file_count': 0,
                'total_size': 0,
                'has_excel': False,
                'main_excel_file': None,
                'file_types': set(),
                'subdirs': []
            }

    def _generate_filtered_statistics(self):
        """Generate statistics filtered by the selected date range."""
        try:
            if not self.date_from_selected or not self.date_to_selected:
                messagebox.showwarning("Attention", "Veuillez sélectionner une période complète")
                return

            if not hasattr(self, 'data') or self.data is None:
                messagebox.showwarning("Attention", "Veuillez d'abord charger les données")
                return

            self.logger.info(f"Generating filtered statistics from {self.date_from_selected} to {self.date_to_selected}")

            # Update status
            self.date_range_status.config(
                text="🔄 Génération des statistiques en cours...",
                fg=COLORS['INFO']
            )
            self.generate_stats_button.config(state=tk.DISABLED)

            # Use async task for heavy computation
            def generate_stats():
                try:
                    # Filter data by date range
                    filtered_data = self._filter_data_by_date_range()

                    # Load stats folder data
                    stats_folder_data = self._load_stats_folder_data()

                    # Generate comprehensive statistics
                    statistics = self._compute_filtered_statistics(filtered_data, stats_folder_data)

                    # Prepare dashboard data
                    dashboard_data = self._prepare_dashboard_data(statistics)

                    return {
                        'filtered_data': filtered_data,
                        'statistics': statistics,
                        'dashboard_data': dashboard_data,
                        'stats_folder_data': stats_folder_data
                    }

                except Exception as e:
                    self.logger.error(f"Error in generate_stats: {e}")
                    raise

            def on_success(result):
                try:
                    # Store results
                    self.filtered_statistics = result['statistics']
                    self.dashboard_data = result['dashboard_data']
                    self.stats_folder_data = result['stats_folder_data']

                    # Update UI
                    self.date_range_status.config(
                        text=f"✅ Statistiques générées ({len(result['filtered_data'])} enregistrements)",
                        fg=COLORS['SUCCESS']
                    )
                    self.generate_stats_button.config(state=tk.NORMAL)

                    # Update statistics display
                    self._update_statistics_display_with_filtered_data()

                    # Inject statistics into stats folder index and open it
                    stats_index_path = self._inject_statistics_to_stats_index()

                    # Open stats index file directly - NO dashboard modal
                    if stats_index_path:
                        self._open_stats_index_file(stats_index_path)
                    else:
                        # If no stats index found, show error message
                        messagebox.showwarning(
                            "Index non trouvé",
                            "Aucun fichier index trouvé dans le dossier stats.\n\n"
                            "Veuillez créer un fichier index.html ou index.xlsx dans le dossier stats."
                        )

                    self.logger.info("Filtered statistics generated successfully")

                except Exception as e:
                    self.logger.error(f"Error in on_success: {e}")
                    messagebox.showerror("Erreur", f"Erreur lors de la mise à jour: {e}")

            def on_error(error):
                self.logger.error(f"Error generating filtered statistics: {error}")
                self.date_range_status.config(
                    text="❌ Erreur lors de la génération",
                    fg=COLORS['ERROR']
                )
                self.generate_stats_button.config(state=tk.NORMAL)
                messagebox.showerror("Erreur", f"Erreur lors de la génération des statistiques:\n{error}")

            # Run async task
            run_async_task(
                generate_stats,
                callback=on_success,
                error_callback=on_error,
                task_name="Génération des statistiques filtrées"
            )

        except Exception as e:
            self.logger.error(f"Error in _generate_filtered_statistics: {e}")
            messagebox.showerror("Erreur", f"Erreur lors du lancement de la génération: {e}")

    def _filter_data_by_date_range(self):
        """Filter the loaded data by the selected date range."""
        try:
            if not hasattr(self, 'data') or self.data is None:
                return []

            filtered_data = []

            for record in self.data:
                try:
                    # Extract date from record - adapt based on your data structure
                    record_date = None

                    # Try different date field names that might exist
                    date_fields = ['date', 'Date', 'DATE', 'date_creation', 'timestamp', 'created_at']

                    for field in date_fields:
                        if field in record and record[field]:
                            date_value = record[field]

                            # Handle different date formats
                            if isinstance(date_value, str):
                                # Try parsing common date formats
                                for date_format in ['%Y-%m-%d', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S']:
                                    try:
                                        record_date = datetime.datetime.strptime(date_value, date_format).date()
                                        break
                                    except ValueError:
                                        continue
                            elif hasattr(date_value, 'date'):
                                # Handle datetime objects
                                record_date = date_value.date()
                            elif hasattr(date_value, 'year'):
                                # Handle date objects
                                record_date = date_value

                            if record_date:
                                break

                    # If no date found, try to extract from filename or other fields
                    if not record_date and 'filename' in record:
                        record_date = self._extract_date_from_filename(record['filename'])

                    # Check if record falls within date range
                    if record_date and self.date_from_selected <= record_date <= self.date_to_selected:
                        filtered_data.append(record)

                except Exception as e:
                    self.logger.debug(f"Error processing record for date filtering: {e}")
                    continue

            self.logger.info(f"Filtered {len(filtered_data)} records from {len(self.data)} total records")
            return filtered_data

        except Exception as e:
            self.logger.error(f"Error filtering data by date range: {e}")
            return []

    def _load_stats_folder_data(self):
        """Load data from the stats folder if it exists."""
        try:
            stats_folder_data = {}

            # Use the same search logic as other methods
            search_locations = []

            # Location 1: Relative to Excel file (original logic)
            if hasattr(self, 'excel_path') and self.excel_path:
                base_dir = os.path.dirname(self.excel_path)
                search_locations.append(base_dir)

            # Location 2: In the application's src directory
            import sys
            if hasattr(sys, '_MEIPASS'):
                app_dir = sys._MEIPASS
                search_locations.append(app_dir)
            else:
                current_file = os.path.abspath(__file__)
                src_dir = os.path.dirname(os.path.dirname(current_file))
                search_locations.append(src_dir)

            # Location 3: Current working directory
            search_locations.append(os.getcwd())

            # Try different possible folder names
            possible_folders = ['pres stats', 'stats', 'Stats', 'STATS']
            stats_folder = None

            for base_dir in search_locations:
                if not base_dir or not os.path.exists(base_dir):
                    continue

                for folder_name in possible_folders:
                    test_folder = os.path.join(base_dir, folder_name)
                    if os.path.exists(test_folder) and os.path.isdir(test_folder):
                        stats_folder = test_folder
                        self.logger.info(f"Found stats folder for data loading: {folder_name} at {test_folder}")
                        break

                if stats_folder:
                    break

                if stats_folder and os.path.exists(stats_folder) and os.path.isdir(stats_folder):
                    self.logger.info(f"Loading data from stats folder: {stats_folder}")

                    # Load all files in stats folder
                    for root, dirs, files in os.walk(stats_folder):
                        for file in files:
                            if file.endswith(('.xlsx', '.xls', '.csv')):
                                file_path = os.path.join(root, file)
                                try:
                                    # Load file data
                                    if file.endswith('.csv'):
                                        import pandas as pd
                                        data = pd.read_csv(file_path)
                                    else:
                                        import pandas as pd
                                        data = pd.read_excel(file_path)

                                    stats_folder_data[file] = {
                                        'path': file_path,
                                        'data': data,
                                        'rows': len(data),
                                        'columns': list(data.columns)
                                    }

                                    self.logger.info(f"Loaded stats file: {file} ({len(data)} rows)")

                                except Exception as e:
                                    self.logger.warning(f"Could not load stats file {file}: {e}")
                else:
                    self.logger.info("No stats folder found")

            return stats_folder_data

        except Exception as e:
            self.logger.error(f"Error loading stats folder data: {e}")
            return {}

    def _compute_filtered_statistics(self, filtered_data, stats_folder_data):
        """Compute comprehensive statistics from filtered data and stats folder."""
        try:
            statistics = {
                'period': {
                    'start_date': self.date_from_selected,
                    'end_date': self.date_to_selected,
                    'total_days': (self.date_to_selected - self.date_from_selected).days + 1
                },
                'data_summary': {
                    'total_records': len(filtered_data),
                    'stats_files': len(stats_folder_data)
                },
                'motifs': {},
                'processing_times': {},
                'collaborateurs': {},
                'communes': {},
                'daily_stats': {}
            }

            # Analyze motifs using existing methods
            motifs_analysis = self._analyze_motifs_in_filtered_data(filtered_data)
            statistics['motifs'] = motifs_analysis

            # Analyze processing times using existing methods
            processing_analysis = self._analyze_processing_times_in_filtered_data(filtered_data)
            statistics['processing_times'] = processing_analysis

            # Analyze by collaborateur
            collab_analysis = self._analyze_by_collaborateur(filtered_data)
            statistics['collaborateurs'] = collab_analysis

            # Analyze by commune
            commune_analysis = self._analyze_by_commune(filtered_data)
            statistics['communes'] = commune_analysis

            # Daily statistics
            daily_analysis = self._analyze_daily_statistics(filtered_data)
            statistics['daily_stats'] = daily_analysis

            # Integrate stats folder data
            if stats_folder_data:
                statistics['stats_folder_analysis'] = self._analyze_stats_folder_data(stats_folder_data)

            self.logger.info(f"Computed statistics for {len(filtered_data)} records")
            return statistics

        except Exception as e:
            self.logger.error(f"Error computing filtered statistics: {e}")
            return {}

    def _analyze_motifs_in_filtered_data(self, filtered_data):
        """Analyze motifs in the filtered data using existing methods."""
        try:
            motifs_count = {}
            motifs_details = {}

            for record in filtered_data:
                # Extract motif from record - adapt based on your data structure
                motif = None

                # Try different motif field names
                motif_fields = ['motif', 'Motif', 'MOTIF', 'reason', 'type']

                for field in motif_fields:
                    if field in record and record[field]:
                        motif = str(record[field]).strip()
                        break

                if motif:
                    # Count motifs
                    if motif not in motifs_count:
                        motifs_count[motif] = 0
                        motifs_details[motif] = []

                    motifs_count[motif] += 1
                    motifs_details[motif].append(record)

            # Sort by frequency
            sorted_motifs = sorted(motifs_count.items(), key=lambda x: x[1], reverse=True)

            return {
                'count': motifs_count,
                'details': motifs_details,
                'sorted': sorted_motifs,
                'total_unique': len(motifs_count)
            }

        except Exception as e:
            self.logger.error(f"Error analyzing motifs: {e}")
            return {}

    def _analyze_processing_times_in_filtered_data(self, filtered_data):
        """Analyze processing times in the filtered data."""
        try:
            processing_times = []

            for record in filtered_data:
                # Extract processing time - adapt based on your data structure
                processing_time = None

                # Try different time field names
                time_fields = ['processing_time', 'duration', 'temps_traitement', 'time']

                for field in time_fields:
                    if field in record and record[field]:
                        try:
                            processing_time = float(record[field])
                            break
                        except (ValueError, TypeError):
                            continue

                if processing_time is not None:
                    processing_times.append(processing_time)

            if processing_times:
                import statistics as stats
                return {
                    'count': len(processing_times),
                    'average': stats.mean(processing_times),
                    'median': stats.median(processing_times),
                    'min': min(processing_times),
                    'max': max(processing_times),
                    'total': sum(processing_times)
                }
            else:
                return {'count': 0}

        except Exception as e:
            self.logger.error(f"Error analyzing processing times: {e}")
            return {}

    def _analyze_by_collaborateur(self, filtered_data):
        """Analyze data by collaborateur."""
        try:
            collab_stats = {}

            for record in filtered_data:
                # Extract collaborateur
                collaborateur = None

                collab_fields = ['collaborateur', 'Collaborateur', 'user', 'operator']

                for field in collab_fields:
                    if field in record and record[field]:
                        collaborateur = str(record[field]).strip()
                        break

                if collaborateur:
                    if collaborateur not in collab_stats:
                        collab_stats[collaborateur] = {
                            'count': 0,
                            'records': []
                        }

                    collab_stats[collaborateur]['count'] += 1
                    collab_stats[collaborateur]['records'].append(record)

            return collab_stats

        except Exception as e:
            self.logger.error(f"Error analyzing by collaborateur: {e}")
            return {}

    def _analyze_by_commune(self, filtered_data):
        """Analyze data by commune."""
        try:
            commune_stats = {}

            for record in filtered_data:
                # Extract commune
                commune = None

                commune_fields = ['commune', 'Commune', 'city', 'municipality']

                for field in commune_fields:
                    if field in record and record[field]:
                        commune = str(record[field]).strip()
                        break

                if commune:
                    if commune not in commune_stats:
                        commune_stats[commune] = {
                            'count': 0,
                            'records': []
                        }

                    commune_stats[commune]['count'] += 1
                    commune_stats[commune]['records'].append(record)

            return commune_stats

        except Exception as e:
            self.logger.error(f"Error analyzing by commune: {e}")
            return {}

    def _analyze_daily_statistics(self, filtered_data):
        """Analyze daily statistics for the filtered period."""
        try:
            daily_stats = {}

            # Initialize all days in the range
            current_date = self.date_from_selected
            while current_date <= self.date_to_selected:
                daily_stats[current_date.strftime('%Y-%m-%d')] = {
                    'count': 0,
                    'records': []
                }
                current_date += datetime.timedelta(days=1)

            # Populate with actual data
            for record in filtered_data:
                record_date = self._extract_date_from_record(record)
                if record_date:
                    date_str = record_date.strftime('%Y-%m-%d')
                    if date_str in daily_stats:
                        daily_stats[date_str]['count'] += 1
                        daily_stats[date_str]['records'].append(record)

            return daily_stats

        except Exception as e:
            self.logger.error(f"Error analyzing daily statistics: {e}")
            return {}

    def _analyze_stats_folder_data(self, stats_folder_data):
        """Analyze data from the stats folder."""
        try:
            analysis = {
                'files_summary': {},
                'combined_metrics': {},
                'data_quality': {}
            }

            for filename, file_info in stats_folder_data.items():
                data = file_info['data']

                # Basic file analysis
                analysis['files_summary'][filename] = {
                    'rows': len(data),
                    'columns': len(data.columns),
                    'column_names': list(data.columns),
                    'data_types': data.dtypes.to_dict() if hasattr(data, 'dtypes') else {}
                }

                # Look for key metrics in the data
                if hasattr(data, 'columns'):
                    for col in data.columns:
                        col_lower = col.lower()
                        if any(keyword in col_lower for keyword in ['count', 'total', 'sum', 'average']):
                            try:
                                if filename not in analysis['combined_metrics']:
                                    analysis['combined_metrics'][filename] = {}
                                analysis['combined_metrics'][filename][col] = {
                                    'sum': data[col].sum() if data[col].dtype in ['int64', 'float64'] else None,
                                    'mean': data[col].mean() if data[col].dtype in ['int64', 'float64'] else None,
                                    'count': data[col].count()
                                }
                            except Exception as e:
                                self.logger.debug(f"Could not analyze column {col}: {e}")

            return analysis

        except Exception as e:
            self.logger.error(f"Error analyzing stats folder data: {e}")
            return {}

    def _extract_date_from_record(self, record):
        """Extract date from a record."""
        try:
            # Try different date field names
            date_fields = ['date', 'Date', 'DATE', 'date_creation', 'timestamp', 'created_at']

            for field in date_fields:
                if field in record and record[field]:
                    date_value = record[field]

                    if isinstance(date_value, str):
                        # Try parsing common date formats
                        for date_format in ['%Y-%m-%d', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S']:
                            try:
                                return datetime.datetime.strptime(date_value, date_format).date()
                            except ValueError:
                                continue
                    elif hasattr(date_value, 'date'):
                        return date_value.date()
                    elif hasattr(date_value, 'year'):
                        return date_value

            return None

        except Exception as e:
            self.logger.debug(f"Error extracting date from record: {e}")
            return None

    def _extract_date_from_filename(self, filename):
        """Extract date from filename if possible."""
        try:
            import re

            # Look for date patterns in filename
            date_patterns = [
                r'(\d{4}-\d{2}-\d{2})',  # YYYY-MM-DD
                r'(\d{2}/\d{2}/\d{4})',  # DD/MM/YYYY
                r'(\d{4}\d{2}\d{2})',    # YYYYMMDD
            ]

            for pattern in date_patterns:
                match = re.search(pattern, filename)
                if match:
                    date_str = match.group(1)

                    # Try to parse the found date
                    for date_format in ['%Y-%m-%d', '%d/%m/%Y', '%Y%m%d']:
                        try:
                            return datetime.datetime.strptime(date_str, date_format).date()
                        except ValueError:
                            continue

            return None

        except Exception as e:
            self.logger.debug(f"Error extracting date from filename: {e}")
            return None

    def _inject_statistics_to_stats_index(self):
        """Inject the generated statistics into the stats folder index file.

        Returns:
            str: Path to the stats index file if successful, None otherwise
        """
        try:
            self.logger.info("Starting statistics injection process...")

            if not hasattr(self, 'filtered_statistics') or not self.filtered_statistics:
                self.logger.warning("No filtered statistics available for injection")
                return None

            self.logger.info("Filtered statistics available, analyzing pres stats folder...")

            # First, analyze the pres stats folder to understand its structure
            folder_analysis = self._analyze_pres_stats_folder()
            if not folder_analysis:
                self.logger.warning("No pres stats folder found or analysis failed")
                return None

            # Find the best index file to use
            stats_index_path = self._find_stats_index_file()
            if not stats_index_path:
                self.logger.warning("No stats index file found for injection")
                self.logger.info(f"Excel path: {getattr(self, 'excel_path', 'NOT SET')}")

                # Show detailed information about what was found
                if folder_analysis:
                    self.logger.info(f"Folder analysis results:")
                    self.logger.info(f"  Folder: {folder_analysis['folder_name']}")
                    self.logger.info(f"  Total files: {folder_analysis['total_files']}")
                    self.logger.info(f"  HTML files: {len(folder_analysis['html_files'])}")
                    self.logger.info(f"  Excel files: {len(folder_analysis['excel_files'])}")
                    self.logger.info(f"  Potential index files: {len(folder_analysis['potential_index_files'])}")

                    # If we have potential index files, try to use the first one
                    if folder_analysis['potential_index_files']:
                        stats_index_path = folder_analysis['potential_index_files'][0]['path']
                        self.logger.info(f"Using potential index file: {stats_index_path}")
                    elif folder_analysis['html_files']:
                        stats_index_path = folder_analysis['html_files'][0]['path']
                        self.logger.info(f"Using first HTML file as index: {stats_index_path}")
                    elif folder_analysis['excel_files']:
                        stats_index_path = folder_analysis['excel_files'][0]['path']
                        self.logger.info(f"Using first Excel file as index: {stats_index_path}")

                if not stats_index_path:
                    return None

            self.logger.info(f"Injecting statistics into: {stats_index_path}")

            # Examine the structure of the index file
            file_structure = self._examine_index_file_structure(stats_index_path)
            if file_structure:
                self.logger.info(f"File structure analysis completed for {file_structure['file_type']} file")
                self.logger.info(f"Available injection strategies: {file_structure['injection_strategies']}")

            # Prepare statistics data for injection
            stats_data = self._prepare_statistics_for_injection()

            # Inject into the index file using the analyzed structure
            success = self._write_statistics_to_index(stats_index_path, stats_data, file_structure)

            if success:
                self.logger.info("Statistics successfully injected into stats index")
                # Update status
                if hasattr(self, 'date_range_status'):
                    current_text = self.date_range_status.cget('text')
                    self.date_range_status.config(
                        text=f"{current_text} | 📝 Injecté et ouvert",
                        fg=COLORS['SUCCESS']
                    )
                return stats_index_path
            else:
                self.logger.error("Failed to inject statistics into stats index")
                return None

        except Exception as e:
            self.logger.error(f"Error injecting statistics to stats index: {e}")
            return None

    def _open_stats_index_file(self, stats_index_path):
        """Open the stats index file with the injected statistics."""
        try:
            import os
            import subprocess
            import platform

            if not os.path.exists(stats_index_path):
                self.logger.error(f"Stats index file not found: {stats_index_path}")
                messagebox.showerror("Erreur", f"Fichier index introuvable: {stats_index_path}")
                return False

            self.logger.info(f"Opening stats index file: {stats_index_path}")

            # Determine the system and open the file with the default application
            system = platform.system()

            try:
                if system == "Windows":
                    # Windows - use os.startfile
                    os.startfile(stats_index_path)
                elif system == "Darwin":
                    # macOS - use open command
                    subprocess.run(["open", stats_index_path], check=True)
                elif system == "Linux":
                    # Linux - use xdg-open
                    subprocess.run(["xdg-open", stats_index_path], check=True)
                else:
                    # Fallback - try to use webbrowser for HTML files
                    file_ext = os.path.splitext(stats_index_path)[1].lower()
                    if file_ext in ['.html', '.htm']:
                        import webbrowser
                        webbrowser.open(f"file://{os.path.abspath(stats_index_path)}")
                    else:
                        raise OSError(f"Unsupported system: {system}")

                # Show success message
                filename = os.path.basename(stats_index_path)
                messagebox.showinfo(
                    "Succès",
                    f"📊 Statistiques injectées et fichier ouvert:\n\n{filename}\n\nLes statistiques de la période sélectionnée ont été intégrées dans votre index."
                )

                self.logger.info(f"Successfully opened stats index: {filename}")
                return True

            except subprocess.CalledProcessError as e:
                self.logger.error(f"Failed to open file with system command: {e}")
                raise
            except Exception as e:
                self.logger.error(f"Failed to open file: {e}")
                raise

        except Exception as e:
            self.logger.error(f"Error opening stats index file: {e}")

            # Show error message with fallback option
            filename = os.path.basename(stats_index_path) if stats_index_path else "fichier"
            result = messagebox.askyesno(
                "Erreur d'ouverture",
                f"Impossible d'ouvrir automatiquement le fichier {filename}.\n\n"
                f"Les statistiques ont été injectées avec succès.\n\n"
                f"Voulez-vous ouvrir le dossier contenant le fichier?"
            )

            if result:
                try:
                    # Open the containing folder
                    folder_path = os.path.dirname(stats_index_path)
                    if platform.system() == "Windows":
                        os.startfile(folder_path)
                    elif platform.system() == "Darwin":
                        subprocess.run(["open", folder_path], check=True)
                    elif platform.system() == "Linux":
                        subprocess.run(["xdg-open", folder_path], check=True)

                    self.logger.info(f"Opened containing folder: {folder_path}")

                except Exception as folder_error:
                    self.logger.error(f"Failed to open containing folder: {folder_error}")
                    messagebox.showerror(
                        "Erreur",
                        f"Impossible d'ouvrir le dossier.\n\nChemin: {folder_path}"
                    )

            return False

    def _find_stats_index_file(self):
        """Find the index file in the stats folder."""
        try:
            if not hasattr(self, 'excel_path') or not self.excel_path:
                return None

            # Try multiple search locations for the pres stats folder
            search_locations = []

            # Location 1: Relative to Excel file (original logic)
            if hasattr(self, 'excel_path') and self.excel_path:
                base_dir = os.path.dirname(self.excel_path)
                search_locations.append(base_dir)

            # Location 2: In the application's src directory
            import sys
            if hasattr(sys, '_MEIPASS'):
                # Running as PyInstaller bundle
                app_dir = sys._MEIPASS
                search_locations.append(app_dir)
            else:
                # Running as script - get src directory
                current_file = os.path.abspath(__file__)
                src_dir = os.path.dirname(os.path.dirname(current_file))  # Go up from ui/modules to src
                search_locations.append(src_dir)

            # Location 3: Current working directory
            search_locations.append(os.getcwd())

            # Location 4: Application root directory (parent of src)
            if len(search_locations) > 1 and search_locations[1]:
                app_root = os.path.dirname(search_locations[1])  # Parent of src
                search_locations.append(app_root)

            # Try different possible folder names
            possible_stats_folders = [
                'pres stats',  # The actual folder name
                'stats',       # Fallback
                'Stats',       # Case variation
                'STATS'        # Case variation
            ]

            stats_folder = None
            for base_dir in search_locations:
                if not base_dir or not os.path.exists(base_dir):
                    continue

                self.logger.info(f"Searching for pres stats folder in: {base_dir}")

                for folder_name in possible_stats_folders:
                    test_folder = os.path.join(base_dir, folder_name)
                    if os.path.exists(test_folder) and os.path.isdir(test_folder):
                        stats_folder = test_folder
                        self.logger.info(f"Found pres stats folder: {folder_name} at {test_folder}")
                        break

                if stats_folder:
                    break

            if not stats_folder:
                self.logger.warning(f"No pres stats folder found in any search location")
                self.logger.info(f"Searched locations: {search_locations}")
                self.logger.info(f"Searched folder names: {possible_stats_folders}")
                return None

            # Look for index files (common names)
            index_names = [
                'index.html',
                'index.htm',
                'dashboard.html',
                'main.html',
                'stats.html',
                'rapport.html',
                'index.xlsx',
                'index.xls',
                'dashboard.xlsx',
                'stats.xlsx'
            ]

            for index_name in index_names:
                index_path = os.path.join(stats_folder, index_name)
                if os.path.exists(index_path):
                    self.logger.info(f"Found stats index file: {index_name}")
                    return index_path

            # If no specific index file found, look for any HTML or Excel file
            for root, dirs, files in os.walk(stats_folder):
                for file in files:
                    if file.lower().endswith(('.html', '.htm', '.xlsx', '.xls')):
                        file_path = os.path.join(root, file)
                        self.logger.info(f"Using stats file as index: {file}")
                        return file_path

            return None

        except Exception as e:
            self.logger.error(f"Error finding stats index file: {e}")
            return None

    def _analyze_pres_stats_folder(self):
        """Analyze the contents of the 'pres stats' folder to understand its structure."""
        try:
            # Try multiple search locations for the pres stats folder
            search_locations = []

            # Location 1: Relative to Excel file (original logic)
            if hasattr(self, 'excel_path') and self.excel_path:
                base_dir = os.path.dirname(self.excel_path)
                search_locations.append(base_dir)

            # Location 2: In the application's src directory
            import sys
            if hasattr(sys, '_MEIPASS'):
                # Running as PyInstaller bundle
                app_dir = sys._MEIPASS
            else:
                # Running as script - get src directory
                current_file = os.path.abspath(__file__)
                src_dir = os.path.dirname(os.path.dirname(current_file))  # Go up from ui/modules to src
                search_locations.append(src_dir)

            # Location 3: Current working directory
            search_locations.append(os.getcwd())

            # Location 4: Application root directory (parent of src)
            if len(search_locations) > 1:
                app_root = os.path.dirname(search_locations[1])  # Parent of src
                search_locations.append(app_root)

            # Try to find the pres stats folder in all locations
            possible_folders = ['pres stats', 'stats', 'Stats', 'STATS']
            stats_folder = None

            for base_dir in search_locations:
                if not base_dir or not os.path.exists(base_dir):
                    continue

                for folder_name in possible_folders:
                    test_folder = os.path.join(base_dir, folder_name)
                    if os.path.exists(test_folder) and os.path.isdir(test_folder):
                        stats_folder = test_folder
                        self.logger.info(f"Found pres stats folder: {folder_name} at {test_folder}")
                        break

                if stats_folder:
                    break

            if not stats_folder:
                self.logger.info("No pres stats folder found for analysis")
                return None

            # Analyze folder contents
            analysis = {
                'folder_path': stats_folder,
                'folder_name': os.path.basename(stats_folder),
                'files': [],
                'html_files': [],
                'excel_files': [],
                'other_files': [],
                'total_files': 0,
                'potential_index_files': []
            }

            # Walk through the folder
            for root, dirs, files in os.walk(stats_folder):
                for file in files:
                    file_path = os.path.join(root, file)
                    relative_path = os.path.relpath(file_path, stats_folder)
                    file_size = os.path.getsize(file_path)

                    file_info = {
                        'name': file,
                        'path': file_path,
                        'relative_path': relative_path,
                        'size': file_size,
                        'extension': os.path.splitext(file)[1].lower()
                    }

                    analysis['files'].append(file_info)
                    analysis['total_files'] += 1

                    # Categorize files
                    ext = file_info['extension']
                    if ext in ['.html', '.htm']:
                        analysis['html_files'].append(file_info)
                        # Check if it might be an index file
                        if any(keyword in file.lower() for keyword in ['index', 'main', 'dashboard', 'home']):
                            analysis['potential_index_files'].append(file_info)
                    elif ext in ['.xlsx', '.xls']:
                        analysis['excel_files'].append(file_info)
                        if any(keyword in file.lower() for keyword in ['index', 'main', 'dashboard', 'stats']):
                            analysis['potential_index_files'].append(file_info)
                    else:
                        analysis['other_files'].append(file_info)

            # Log the analysis
            self.logger.info(f"Pres stats folder analysis:")
            self.logger.info(f"  Total files: {analysis['total_files']}")
            self.logger.info(f"  HTML files: {len(analysis['html_files'])}")
            self.logger.info(f"  Excel files: {len(analysis['excel_files'])}")
            self.logger.info(f"  Other files: {len(analysis['other_files'])}")
            self.logger.info(f"  Potential index files: {len(analysis['potential_index_files'])}")

            # List all files found
            for file_info in analysis['files']:
                self.logger.info(f"    {file_info['relative_path']} ({file_info['size']} bytes)")

            return analysis

        except Exception as e:
            self.logger.error(f"Error analyzing pres stats folder: {e}")
            return None

    def _examine_index_file_structure(self, file_path):
        """Examine the structure of an index file to understand injection points."""
        try:
            self.logger.info(f"Examining index file structure: {file_path}")

            file_ext = os.path.splitext(file_path)[1].lower()

            if file_ext in ['.html', '.htm']:
                return self._examine_html_structure(file_path)
            elif file_ext in ['.xlsx', '.xls']:
                return self._examine_excel_structure(file_path)
            else:
                self.logger.warning(f"Unsupported file type for examination: {file_ext}")
                return None

        except Exception as e:
            self.logger.error(f"Error examining index file structure: {e}")
            return None

    def _examine_html_structure(self, html_path):
        """Examine HTML file structure for injection points."""
        try:
            with open(html_path, 'r', encoding='utf-8') as f:
                content = f.read()

            structure = {
                'file_path': html_path,
                'file_type': 'html',
                'content_length': len(content),
                'has_pladria_injection_point': False,
                'has_body_tag': False,
                'has_head_tag': False,
                'injection_strategies': []
            }

            # Check for existing Pladria injection points
            if '<!-- PLADRIA_STATS_INJECTION -->' in content:
                structure['has_pladria_injection_point'] = True
                structure['injection_strategies'].append('existing_pladria_marker')
                self.logger.info("Found existing Pladria injection marker")

            # Check for HTML structure
            if '<body' in content.lower():
                structure['has_body_tag'] = True
                structure['injection_strategies'].append('before_body_close')

            if '<head' in content.lower():
                structure['has_head_tag'] = True

            # Look for other potential injection points
            potential_points = [
                ('<!-- stats -->', 'stats_comment'),
                ('<!-- statistics -->', 'statistics_comment'),
                ('<div id="stats"', 'stats_div'),
                ('<div class="stats"', 'stats_class_div'),
                ('<main', 'main_tag'),
                ('<section', 'section_tag')
            ]

            for marker, strategy in potential_points:
                if marker.lower() in content.lower():
                    structure['injection_strategies'].append(strategy)
                    self.logger.info(f"Found potential injection point: {marker}")

            # If no specific points found, use body fallback
            if not structure['injection_strategies'] and structure['has_body_tag']:
                structure['injection_strategies'].append('body_fallback')

            self.logger.info(f"HTML structure analysis: {len(structure['injection_strategies'])} injection strategies available")

            return structure

        except Exception as e:
            self.logger.error(f"Error examining HTML structure: {e}")
            return None

    def _examine_excel_structure(self, excel_path):
        """Examine Excel file structure for injection points."""
        try:
            pd = get_pandas()

            # Read Excel file to understand its structure
            excel_file = pd.ExcelFile(excel_path)
            sheet_names = excel_file.sheet_names

            structure = {
                'file_path': excel_path,
                'file_type': 'excel',
                'sheet_names': sheet_names,
                'sheet_count': len(sheet_names),
                'injection_strategies': []
            }

            # Check for existing Pladria sheets
            pladria_sheets = [name for name in sheet_names if 'pladria' in name.lower()]
            if pladria_sheets:
                structure['existing_pladria_sheets'] = pladria_sheets
                structure['injection_strategies'].append('update_existing_pladria_sheets')
                self.logger.info(f"Found existing Pladria sheets: {pladria_sheets}")

            # Always allow adding new sheets
            structure['injection_strategies'].append('add_new_sheets')

            self.logger.info(f"Excel structure analysis: {structure['sheet_count']} sheets, {len(structure['injection_strategies'])} injection strategies")

            return structure

        except Exception as e:
            self.logger.error(f"Error examining Excel structure: {e}")
            return None

    def _prepare_statistics_for_injection(self):
        """Prepare statistics data in a format suitable for injection."""
        try:
            stats = self.filtered_statistics

            # Create structured data for injection
            injection_data = {
                'metadata': {
                    'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'period_start': str(stats['period']['start_date']),
                    'period_end': str(stats['period']['end_date']),
                    'total_days': stats['period']['total_days'],
                    'total_records': stats['data_summary']['total_records']
                },
                'summary': {
                    'total_records': stats['data_summary']['total_records'],
                    'unique_motifs': stats['motifs'].get('total_unique', 0),
                    'stats_files': stats['data_summary']['stats_files']
                },
                'top_motifs': [],
                'collaborateurs': [],
                'communes': [],
                'processing_times': {},
                'daily_stats': []
            }

            # Top motifs
            if stats['motifs'].get('sorted'):
                for motif, count in stats['motifs']['sorted'][:10]:
                    percentage = round((count / stats['data_summary']['total_records']) * 100, 1)
                    injection_data['top_motifs'].append({
                        'motif': motif,
                        'count': count,
                        'percentage': percentage
                    })

            # Collaborateurs
            if stats['collaborateurs']:
                sorted_collabs = sorted(stats['collaborateurs'].items(),
                                      key=lambda x: x[1]['count'], reverse=True)
                for collab, data in sorted_collabs[:10]:
                    percentage = round((data['count'] / stats['data_summary']['total_records']) * 100, 1)
                    injection_data['collaborateurs'].append({
                        'name': collab,
                        'count': data['count'],
                        'percentage': percentage
                    })

            # Communes
            if stats['communes']:
                sorted_communes = sorted(stats['communes'].items(),
                                       key=lambda x: x[1]['count'], reverse=True)
                for commune, data in sorted_communes[:10]:
                    percentage = round((data['count'] / stats['data_summary']['total_records']) * 100, 1)
                    injection_data['communes'].append({
                        'name': commune,
                        'count': data['count'],
                        'percentage': percentage
                    })

            # Processing times
            if stats['processing_times'].get('count', 0) > 0:
                pt = stats['processing_times']
                injection_data['processing_times'] = {
                    'average': round(pt['average'], 2),
                    'median': round(pt['median'], 2),
                    'min': round(pt['min'], 2),
                    'max': round(pt['max'], 2),
                    'total_processed': pt['count']
                }

            # Daily stats (last 7 days)
            if stats['daily_stats']:
                sorted_days = sorted(stats['daily_stats'].items())[-7:]
                for date_str, data in sorted_days:
                    injection_data['daily_stats'].append({
                        'date': date_str,
                        'count': data['count']
                    })

            return injection_data

        except Exception as e:
            self.logger.error(f"Error preparing statistics for injection: {e}")
            return {}

    def _write_statistics_to_index(self, index_path, stats_data, file_structure=None):
        """Write statistics data to the index file using analyzed structure."""
        try:
            file_ext = os.path.splitext(index_path)[1].lower()

            if file_ext in ['.html', '.htm']:
                return self._inject_to_html_index(index_path, stats_data, file_structure)
            elif file_ext in ['.xlsx', '.xls']:
                return self._inject_to_excel_index(index_path, stats_data, file_structure)
            else:
                self.logger.warning(f"Unsupported index file format: {file_ext}")
                return False

        except Exception as e:
            self.logger.error(f"Error writing statistics to index: {e}")
            return False

    def _inject_to_html_index(self, html_path, stats_data, file_structure=None):
        """Inject statistics into HTML index file using analyzed structure."""
        try:
            # Read existing HTML content
            with open(html_path, 'r', encoding='utf-8') as f:
                html_content = f.read()

            self.logger.info(f"Original HTML content length: {len(html_content)} characters")

            # Check if this is the pres stats dashboard (has stats-grid and Chart.js)
            if 'stats-grid' in html_content and 'script.js' in html_content:
                self.logger.info("Detected pres stats dashboard - using value update strategy")

                # Use the new value update strategy for pres stats dashboard
                updated_content = self._update_existing_dashboard_values(html_content, html_path)
                if updated_content:
                    # Write updated content
                    with open(html_path, 'w', encoding='utf-8') as f:
                        f.write(updated_content)

                    self.logger.info("Successfully updated pres stats dashboard with filtered data")
                    return True
                else:
                    self.logger.warning("Failed to update dashboard values, falling back to HTML generation")

            # Fallback to original HTML generation method
            self.logger.info("Using HTML generation strategy")

            # Generate HTML statistics section
            stats_html = self._generate_html_statistics(stats_data)

            # Use file structure analysis to determine best injection strategy
            injection_successful = False

            if file_structure and file_structure.get('has_pladria_injection_point'):
                # Use existing Pladria injection point
                injection_successful = self._inject_at_pladria_marker(html_content, stats_html, html_path)
                if injection_successful:
                    self.logger.info("Successfully injected at existing Pladria marker")

            if not injection_successful:
                # Try alternative injection strategies based on file structure
                injection_successful = self._try_alternative_html_injection(html_content, stats_html, html_path, file_structure)

            return injection_successful

        except Exception as e:
            self.logger.error(f"Error injecting to HTML index: {e}")
            return False

    def _inject_at_pladria_marker(self, html_content, stats_html, html_path):
        """Inject statistics at existing Pladria marker."""
        try:
            injection_point = '<!-- PLADRIA_STATS_INJECTION -->'
            end_marker = '<!-- END_PLADRIA_STATS -->'

            start_idx = html_content.find(injection_point)
            end_idx = html_content.find(end_marker)

            if start_idx != -1 and end_idx != -1:
                # Replace existing content between markers
                new_content = (html_content[:start_idx] +
                             injection_point + '\n' + stats_html + '\n' +
                             html_content[end_idx:])
            elif start_idx != -1:
                # Add end marker and content
                new_content = html_content.replace(injection_point,
                                                 injection_point + '\n' + stats_html + '\n<!-- END_PLADRIA_STATS -->')
            else:
                return False

            # Write updated content
            with open(html_path, 'w', encoding='utf-8') as f:
                f.write(new_content)

            self.logger.info("Successfully injected at Pladria marker")
            return True

        except Exception as e:
            self.logger.error(f"Error injecting at Pladria marker: {e}")
            return False

    def _try_alternative_html_injection(self, html_content, stats_html, html_path, file_structure):
        """Try alternative injection strategies for HTML files - Update existing values without changing structure."""
        try:
            new_content = None
            strategy_used = None

            # Strategy 1: Update existing values in place (for pres stats dashboard)
            if 'stats-grid' in html_content and 'script.js' in html_content:
                new_content = self._update_existing_dashboard_values(html_content, html_path)
                if new_content:
                    strategy_used = "update_existing_values"

            # Strategy 2: Inject before closing body tag (fallback)
            if not new_content and '</body>' in html_content.lower():
                body_close_idx = html_content.lower().rfind('</body>')
                if body_close_idx != -1:
                    new_content = (html_content[:body_close_idx] +
                                 '\n<!-- PLADRIA_STATS_INJECTION -->\n' + stats_html + '\n<!-- END_PLADRIA_STATS -->\n' +
                                 html_content[body_close_idx:])
                    strategy_used = "before_body_close"

            # Strategy 3: Append to end of file if no body tag
            if not new_content:
                new_content = html_content + '\n<!-- PLADRIA_STATS_INJECTION -->\n' + stats_html + '\n<!-- END_PLADRIA_STATS -->'
                strategy_used = "append_to_end"

            # Write the updated content
            if new_content:
                with open(html_path, 'w', encoding='utf-8') as f:
                    f.write(new_content)

                self.logger.info(f"Successfully injected using strategy: {strategy_used}")
                return True

            return False

        except Exception as e:
            self.logger.error(f"Error in alternative HTML injection: {e}")
            return False

            self.logger.info(f"Statistics injected into HTML index: {html_path}")
            return True

        except Exception as e:
            self.logger.error(f"Error injecting to HTML index: {e}")
            return False

    def _inject_to_excel_index(self, excel_path, stats_data, file_structure=None):
        """Inject statistics into Excel index file using analyzed structure."""
        try:
            pd = get_pandas()

            self.logger.info(f"Injecting statistics into Excel file: {excel_path}")

            # Create a new sheet with statistics
            with pd.ExcelWriter(excel_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
                # Summary sheet
                summary_df = pd.DataFrame([
                    ['Période', f"{stats_data['metadata']['period_start']} à {stats_data['metadata']['period_end']}"],
                    ['Durée (jours)', stats_data['metadata']['total_days']],
                    ['Total enregistrements', stats_data['summary']['total_records']],
                    ['Motifs uniques', stats_data['summary']['unique_motifs']],
                    ['Fichiers stats', stats_data['summary']['stats_files']],
                    ['Généré le', stats_data['metadata']['generated_at']]
                ], columns=['Métrique', 'Valeur'])

                summary_df.to_excel(writer, sheet_name='Pladria_Stats_Summary', index=False)

                # Top motifs sheet
                if stats_data['top_motifs']:
                    motifs_df = pd.DataFrame(stats_data['top_motifs'])
                    motifs_df.to_excel(writer, sheet_name='Pladria_Top_Motifs', index=False)

                # Collaborateurs sheet
                if stats_data['collaborateurs']:
                    collabs_df = pd.DataFrame(stats_data['collaborateurs'])
                    collabs_df.to_excel(writer, sheet_name='Pladria_Collaborateurs', index=False)

                # Communes sheet
                if stats_data['communes']:
                    communes_df = pd.DataFrame(stats_data['communes'])
                    communes_df.to_excel(writer, sheet_name='Pladria_Communes', index=False)

                # Daily stats sheet
                if stats_data['daily_stats']:
                    daily_df = pd.DataFrame(stats_data['daily_stats'])
                    daily_df.to_excel(writer, sheet_name='Pladria_Daily_Stats', index=False)

            self.logger.info(f"Statistics injected into Excel index: {excel_path}")
            return True

        except Exception as e:
            self.logger.error(f"Error injecting to Excel index: {e}")
            return False

    def _update_existing_dashboard_values(self, html_content, html_path):
        """Update existing values in the pres stats dashboard without changing structure."""
        try:
            if not hasattr(self, 'filtered_statistics') or not self.filtered_statistics:
                self.logger.warning("No filtered statistics available for value updates")
                return None

            stats = self.filtered_statistics
            updated_html = html_content

            # Update the subtitle with the filtered period
            period_text = f"Analyse des données de traitement - {stats['period']['start_date']} à {stats['period']['end_date']}"
            updated_html = self._update_html_text(updated_html,
                r'<p class="subtitle">.*?</p>',
                f'<p class="subtitle">{period_text}</p>')

            # Update total records count in various places if we can map them
            total_records = stats['data_summary']['total_records']

            # Update HTML elements with new data
            self.logger.info("Updating HTML elements with filtered data...")
            updated_html = self._update_html_elements_with_data(updated_html, stats)

            # Update JavaScript file with new data
            script_path = os.path.join(os.path.dirname(html_path), 'script.js')
            if os.path.exists(script_path):
                self._update_script_js_values(script_path, stats)

            # Add a comment to track when this was last updated
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            comment = f'\n<!-- Pladria Statistics Updated: {timestamp} | Period: {stats["period"]["start_date"]} to {stats["period"]["end_date"]} | Records: {total_records} -->\n'

            # Insert comment before closing body tag
            body_close_idx = updated_html.lower().rfind('</body>')
            if body_close_idx != -1:
                updated_html = updated_html[:body_close_idx] + comment + updated_html[body_close_idx:]

            self.logger.info(f"Updated dashboard values for period {stats['period']['start_date']} to {stats['period']['end_date']}")
            return updated_html

        except Exception as e:
            self.logger.error(f"Error updating existing dashboard values: {e}")
            return None

    def _update_html_text(self, html_content, pattern, replacement):
        """Update HTML content using regex pattern."""
        try:
            import re
            return re.sub(pattern, replacement, html_content, flags=re.DOTALL)
        except Exception as e:
            self.logger.error(f"Error updating HTML text: {e}")
            return html_content

    def _update_html_elements_with_data(self, html_content, stats):
        """Update HTML elements in the dashboard with filtered data."""
        try:
            self.logger.info("Updating HTML elements with dashboard data...")
            updated_html = html_content

            # Get dashboard mapping to extract the data
            dashboard_mapping = self._map_stats_to_dashboard_categories(stats)

            # Validate data before injection
            if dashboard_mapping:
                validation_result = self._validate_injection_data(dashboard_mapping)
                if not validation_result['valid']:
                    self.logger.error(f"Data validation failed before HTML injection:")
                    for error in validation_result['errors']:
                        self.logger.error(f"  • {error}")
                    self.logger.warning("Proceeding with injection despite validation errors")

                if validation_result['warnings']:
                    self.logger.warning("Data validation warnings:")
                    for warning in validation_result['warnings']:
                        self.logger.warning(f"  • {warning}")

                self.logger.info(f"Data validation summary: {validation_result.get('data_summary', {})}")
            else:
                self.logger.warning("No dashboard mapping available for validation")

            if dashboard_mapping and 'cm' in dashboard_mapping:
                cm_data = dashboard_mapping['cm']
                cm_values = cm_data['data']  # [RAF_count, MODIF_count, CREA_count]
                total_cm = sum(cm_values)

                self.logger.info(f"Updating CM HTML elements: RAF={cm_values[0]}, MODIF={cm_values[1]}, CREA={cm_values[2]}, Total={total_cm}")

                # Update CM card title with total count
                # Pattern: <h2>CM (894)</h2>
                cm_title_pattern = r'(<h2>CM\s*\()[^)]*(\)</h2>)'
                cm_title_replacement = rf'\g<1>{total_cm}\g<2>'
                updated_html = self._update_html_text(updated_html, cm_title_pattern, cm_title_replacement)

                # Verify title update
                title_check = re.search(cm_title_pattern, updated_html)
                if title_check and str(total_cm) in title_check.group(0):
                    self.logger.info(f"✅ CM title updated to: {title_check.group(0)}")
                else:
                    self.logger.warning(f"⚠️ CM title update may have failed")

                # Update individual CM stat values
                # RAF value: <span class="stat-value raf">806</span>
                raf_pattern = r'(<span class="stat-value raf">)[^<]*(</span>)'
                raf_replacement = rf'\g<1>{cm_values[0]}\g<2>'
                updated_html = self._update_html_text(updated_html, raf_pattern, raf_replacement)

                # MODIF value: <span class="stat-value modif">17</span>
                modif_pattern = r'(<span class="stat-value modif">)[^<]*(</span>)'
                modif_replacement = rf'\g<1>{cm_values[1]}\g<2>'
                updated_html = self._update_html_text(updated_html, modif_pattern, modif_replacement)

                # CREA value: <span class="stat-value crea">71</span>
                crea_pattern = r'(<span class="stat-value crea">)[^<]*(</span>)'
                crea_replacement = rf'\g<1>{cm_values[2]}\g<2>'
                updated_html = self._update_html_text(updated_html, crea_pattern, crea_replacement)

                # Verify individual value updates
                raf_check = re.search(raf_pattern, updated_html)
                modif_check = re.search(modif_pattern, updated_html)
                crea_check = re.search(crea_pattern, updated_html)

                if raf_check and str(cm_values[0]) in raf_check.group(0):
                    self.logger.info(f"✅ RAF value updated to: {raf_check.group(0)}")
                else:
                    self.logger.warning(f"⚠️ RAF value update may have failed")

                if modif_check and str(cm_values[1]) in modif_check.group(0):
                    self.logger.info(f"✅ MODIF value updated to: {modif_check.group(0)}")
                else:
                    self.logger.warning(f"⚠️ MODIF value update may have failed")

                if crea_check and str(cm_values[2]) in crea_check.group(0):
                    self.logger.info(f"✅ CREA value updated to: {crea_check.group(0)}")
                else:
                    self.logger.warning(f"⚠️ CREA value update may have failed")

                self.logger.info("CM HTML elements update process completed")
            else:
                self.logger.warning("No CM data available for HTML element updates")

            # Update Communes section if data is available
            if dashboard_mapping and 'communes' in dashboard_mapping:
                communes_data = dashboard_mapping['communes']
                communes_values = communes_data['data']  # [Orange_count, RIP_count]
                total_communes = sum(communes_values)

                self.logger.info(f"Updating Communes HTML elements: Orange={communes_values[0]}, RIP={communes_values[1]}, Total={total_communes}")

                # Update Communes card title with total count
                # Pattern: <h2>Communes Livrées (60)</h2>
                communes_title_pattern = r'(<h2>Communes Livrées\s*\()[^)]*(\)</h2>)'
                communes_title_replacement = rf'\g<1>{total_communes}\g<2>'
                updated_html = self._update_html_text(updated_html, communes_title_pattern, communes_title_replacement)

                # Update individual Communes stat values
                # Orange value: <span class="stat-value orange">56</span>
                orange_pattern = r'(<span class="stat-value orange">)[^<]*(</span>)'
                orange_replacement = rf'\g<1>{communes_values[0]}\g<2>'
                updated_html = self._update_html_text(updated_html, orange_pattern, orange_replacement)

                # RIP value: <span class="stat-value rip">4</span>
                rip_pattern = r'(<span class="stat-value rip">)[^<]*(</span>)'
                rip_replacement = rf'\g<1>{communes_values[1]}\g<2>'
                updated_html = self._update_html_text(updated_html, rip_pattern, rip_replacement)

                # Verify Communes updates
                communes_title_check = re.search(communes_title_pattern, updated_html)
                orange_check = re.search(orange_pattern, updated_html)
                rip_check = re.search(rip_pattern, updated_html)

                if communes_title_check and str(total_communes) in communes_title_check.group(0):
                    self.logger.info(f"✅ Communes title updated to: {communes_title_check.group(0)}")
                else:
                    self.logger.warning(f"⚠️ Communes title update may have failed")

                if orange_check and str(communes_values[0]) in orange_check.group(0):
                    self.logger.info(f"✅ Orange value updated to: {orange_check.group(0)}")
                else:
                    self.logger.warning(f"⚠️ Orange value update may have failed")

                if rip_check and str(communes_values[1]) in rip_check.group(0):
                    self.logger.info(f"✅ RIP value updated to: {rip_check.group(0)}")
                else:
                    self.logger.warning(f"⚠️ RIP value update may have failed")

                self.logger.info("Communes HTML elements update process completed")
            else:
                self.logger.warning("No Communes data available for HTML element updates")

            # Update Acts section if data is available
            if dashboard_mapping and 'acts' in dashboard_mapping:
                acts_data = dashboard_mapping['acts']
                acts_values = acts_data['data']  # [count1, count2, ...]
                acts_labels = acts_data['labels']  # ['category1', 'category2', ...]
                total_acts = sum(acts_values)

                self.logger.info(f"Updating Acts HTML elements: Total={total_acts}, Categories={len(acts_labels)}")

                # Update Acts card title with total count
                # Pattern: <h2>Acts Traitement PA (11,396)</h2>
                acts_title_pattern = r'(<h2>Acts Traitement PA\s*\()[^)]*(\)</h2>)'
                acts_title_replacement = rf'\g<1>{total_acts:,}\g<2>'
                updated_html = self._update_html_text(updated_html, acts_title_pattern, acts_title_replacement)

                # Update individual Acts stat values AND percentages based on the HTML structure
                # The HTML has specific summary-value elements for each category with static percentages
                acts_html_mappings = {
                    'AD RAS sans temps': r'(<span class="summary-value">)[^<]*(</span>\s*<span class="summary-label">AD RAS sans temps \()[^)]*(\)</span>)',
                    'AD RAS avec temps': r'(<span class="summary-value">)[^<]*(</span>\s*<span class="summary-label">AD RAS avec temps \()[^)]*(\)</span>)',
                    'OK': r'(<span class="summary-value">)[^<]*(</span>\s*<span class="summary-label">OK \()[^)]*(\)</span>)',
                    'NOK': r'(<span class="summary-value">)[^<]*(</span>\s*<span class="summary-label">NOK \()[^)]*(\)</span>)',
                    'AD Non jointe': r'(<span class="summary-value">)[^<]*(</span>\s*<span class="summary-label">AD Non jointe \()[^)]*(\)</span>)',
                    'UPR RAS': r'(<span class="summary-value">)[^<]*(</span>\s*<span class="summary-label">UPR RAS \()[^)]*(\)</span>)',
                    'AD Non trouvée': r'(<span class="summary-value">)[^<]*(</span>\s*<span class="summary-label">AD Non trouvée \()[^)]*(\)</span>)',
                    'Hors commune': r'(<span class="summary-value">)[^<]*(</span>\s*<span class="summary-label">Hors commune \()[^)]*(\)</span>)',
                    'UPR NOK': r'(<span class="summary-value">)[^<]*(</span>\s*<span class="summary-label">UPR NOK \()[^)]*(\)</span>)',
                    'UPR OK': r'(<span class="summary-value">)[^<]*(</span>\s*<span class="summary-label">UPR OK \()[^)]*(\)</span>)'
                }

                # Update each category that we have data for with calculated percentages
                for i, (label, count) in enumerate(zip(acts_labels, acts_values)):
                    if label in acts_html_mappings:
                        # Calculate percentage
                        percentage = (count / total_acts * 100) if total_acts > 0 else 0

                        pattern = acts_html_mappings[label]
                        replacement = rf'\g<1>{count:,}\g<2>{percentage:.1f}%\g<3>'
                        updated_html = self._update_html_text(updated_html, pattern, replacement)
                        self.logger.info(f"   Updated {label}: {count:,} ({percentage:.1f}%)")

                # Verify Acts title update
                acts_title_check = re.search(acts_title_pattern, updated_html)
                if acts_title_check and str(total_acts) in acts_title_check.group(0).replace(',', ''):
                    self.logger.info(f"✅ Acts title updated to: {acts_title_check.group(0)}")
                else:
                    self.logger.warning(f"⚠️ Acts title update may have failed")

                self.logger.info("Acts HTML elements update process completed")
            else:
                self.logger.warning("No Acts data available for HTML element updates")

            # Update UPR section if data is available
            if dashboard_mapping and 'upr' in dashboard_mapping:
                upr_data = dashboard_mapping['upr']
                upr_values = upr_data['data']  # [cree_count, non_count]
                upr_labels = upr_data['labels']  # ['Créé', 'Non']
                total_upr = sum(upr_values)

                self.logger.info(f"Updating UPR HTML elements: Créé tickets={upr_values[0] if len(upr_values) > 0 else 0}")

                # Update only the Créé count (simplified UPR display)
                if len(upr_values) >= 1:
                    # Update Créé count with new metric-value class
                    cree_pattern = r'(<span class="metric-value upr-cree">)[^<]*(</span>)'
                    cree_replacement = rf'\g<1>{upr_values[0]:,}\g<2>'
                    updated_html = self._update_html_text(updated_html, cree_pattern, cree_replacement)

                    self.logger.info(f"   Updated UPR Créé tickets: {upr_values[0]:,}")
                else:
                    self.logger.warning("   No UPR data available for Créé count")

                self.logger.info("UPR HTML elements update process completed")
            else:
                self.logger.warning("No UPR data available for HTML element updates")

            # Update 501/511 section if data is available
            if dashboard_mapping and 'tickets_501511' in dashboard_mapping:
                tickets_501511_data = dashboard_mapping['tickets_501511']
                tickets_501511_values = tickets_501511_data['data']  # [count]
                total_501511 = sum(tickets_501511_values)

                self.logger.info(f"Updating 501/511 HTML elements: Total tickets={total_501511}")

                # Update 501/511 stat value with new metric-value class
                if len(tickets_501511_values) >= 1:
                    tickets_501511_pattern = r'(<span class="metric-value tickets-501511">)[^<]*(</span>)'
                    tickets_501511_replacement = rf'\g<1>{tickets_501511_values[0]:,}\g<2>'
                    updated_html = self._update_html_text(updated_html, tickets_501511_pattern, tickets_501511_replacement)

                    self.logger.info(f"   Updated 501/511 tickets count: {tickets_501511_values[0]:,}")
                else:
                    self.logger.warning("   No 501/511 data available")

                self.logger.info("501/511 HTML elements update process completed")
            else:
                self.logger.warning("No 501/511 data available for HTML element updates")

            # Update RIP section if data is available
            if dashboard_mapping and 'rip' in dashboard_mapping:
                rip_data = dashboard_mapping['rip']
                rip_values = rip_data['data']  # [rien_count, modification_count, creation_count]
                rip_labels = rip_data['labels']  # ['Rien à faire', 'Modification', 'Création']
                total_rip = sum(rip_values)

                self.logger.info(f"Updating RIP HTML elements: Total={total_rip}, Categories={len(rip_labels)}")

                # Update RIP card title with total count
                # Pattern: <h2>RIP (P0 P1) (0)</h2>
                rip_title_pattern = r'(<h2>RIP \(P0 P1\)\s*\()[^)]*(\)</h2>)'
                rip_title_replacement = rf'\g<1>{total_rip:,}\g<2>'
                updated_html = self._update_html_text(updated_html, rip_title_pattern, rip_title_replacement)

                # Update individual RIP stat values (using stat-value class like CM section)
                if len(rip_values) >= 3:
                    # Update Rien à faire count
                    rien_pattern = r'(<span class="stat-value rip-rien">)[^<]*(</span>)'
                    rien_replacement = rf'\g<1>{rip_values[0]:,}\g<2>'
                    updated_html = self._update_html_text(updated_html, rien_pattern, rien_replacement)

                    # Update Modification count
                    modification_pattern = r'(<span class="stat-value rip-modification">)[^<]*(</span>)'
                    modification_replacement = rf'\g<1>{rip_values[1]:,}\g<2>'
                    updated_html = self._update_html_text(updated_html, modification_pattern, modification_replacement)

                    # Update Création count
                    creation_pattern = r'(<span class="stat-value rip-creation">)[^<]*(</span>)'
                    creation_replacement = rf'\g<1>{rip_values[2]:,}\g<2>'
                    updated_html = self._update_html_text(updated_html, creation_pattern, creation_replacement)

                    self.logger.info(f"   Updated RIP Rien à faire: {rip_values[0]:,}")
                    self.logger.info(f"   Updated RIP Modification: {rip_values[1]:,}")
                    self.logger.info(f"   Updated RIP Création: {rip_values[2]:,}")

                # Verify RIP title update
                rip_title_check = re.search(rip_title_pattern, updated_html)
                if rip_title_check and str(total_rip) in rip_title_check.group(0).replace(',', ''):
                    self.logger.info(f"✅ RIP title updated to: {rip_title_check.group(0)}")
                else:
                    self.logger.warning(f"⚠️ RIP title update may have failed")

                self.logger.info("RIP HTML elements update process completed")
            else:
                self.logger.warning("No RIP data available for HTML element updates")

            # Update other sections if data is available
            # TODO: Add updates for Quality Control sections

            # Update Facturation section with real data
            self._update_facturation_data(updated_html, dashboard_mapping)

            return updated_html

        except Exception as e:
            self.logger.error(f"Error updating HTML elements with data: {e}")
            return html_content

    def _validate_injection_data(self, dashboard_mapping):
        """Validate data before injecting into HTML dashboard."""
        try:
            from core.data_validator import DataValidator

            validator = DataValidator()
            return validator.validate_dashboard_injection_data(dashboard_mapping)

        except ImportError:
            self.logger.warning("DataValidator not available - skipping validation")
            return {
                'valid': True,
                'errors': [],
                'warnings': ['DataValidator not available - validation skipped'],
                'data_summary': {}
            }
        except Exception as e:
            self.logger.error(f"Error during data validation: {e}")
            return {
                'valid': False,
                'errors': [f"Validation error: {str(e)}"],
                'warnings': [],
                'data_summary': {}
            }

    def _update_facturation_data(self, html_content, dashboard_mapping):
        """Update the facturation section with real CM and PA data by individual motifs."""
        try:
            self.logger.info("Updating Facturation section with detailed motif data...")

            if dashboard_mapping:
                # Extract PA data (Acts) - individual motif counts
                pa_motif_data = []
                if 'acts' in dashboard_mapping:
                    acts_data = dashboard_mapping['acts']
                    acts_values = acts_data.get('data', [])
                    acts_labels = acts_data.get('labels', [])

                    # Store individual motif counts (should match the order in JavaScript)
                    pa_motif_data = acts_values if acts_values else [0] * 10

                    self.logger.info(f"PA (Acts) motif data for facturation:")
                    for i, (label, count) in enumerate(zip(acts_labels, pa_motif_data)):
                        if i < len(acts_labels):
                            self.logger.info(f"   {label}: {count:,}")

                # Extract CM data - individual motif counts
                cm_motif_data = []
                if 'cm' in dashboard_mapping:
                    cm_data = dashboard_mapping['cm']
                    cm_values = cm_data.get('data', [])
                    cm_labels = cm_data.get('labels', [])

                    # Store individual motif counts (should match the order in JavaScript)
                    cm_motif_data = cm_values if cm_values else [0] * 3

                    self.logger.info(f"CM motif data for facturation:")
                    for i, (label, count) in enumerate(zip(cm_labels, cm_motif_data)):
                        if i < len(cm_labels):
                            self.logger.info(f"   {label}: {count:,}")

                # Extract UPR data - individual motif counts
                upr_motif_data = []
                if 'upr' in dashboard_mapping:
                    upr_data = dashboard_mapping['upr']
                    upr_values = upr_data.get('data', [])
                    upr_labels = upr_data.get('labels', [])

                    # Store individual motif counts (should match the order in JavaScript)
                    upr_motif_data = upr_values if upr_values else [0] * 2

                    self.logger.info(f"UPR motif data for facturation:")
                    for i, (label, count) in enumerate(zip(upr_labels, upr_motif_data)):
                        if i < len(upr_labels):
                            self.logger.info(f"   {label}: {count:,}")

                # Extract 501/511 data - single value
                tickets_501511_data = []
                if 'tickets_501511' in dashboard_mapping:
                    tickets_501511_mapping = dashboard_mapping['tickets_501511']
                    tickets_501511_values = tickets_501511_mapping.get('data', [])

                    # Store single value as array for consistency
                    tickets_501511_data = tickets_501511_values if tickets_501511_values else [0]

                    self.logger.info(f"501/511 tickets data for facturation:")
                    if tickets_501511_data:
                        self.logger.info(f"   Tickets 501/511: {tickets_501511_data[0]:,}")

                # Store the detailed data for script.js update
                self.facturation_data = {
                    'pa_motifs': pa_motif_data,
                    'cm_motifs': cm_motif_data,
                    'upr_motifs': upr_motif_data,
                    'tickets_501511_motifs': tickets_501511_data,
                    'pa_total': sum(pa_motif_data) if pa_motif_data else 0,
                    'cm_total': sum(cm_motif_data) if cm_motif_data else 0,
                    'upr_total': sum(upr_motif_data) if upr_motif_data else 0,
                    'tickets_501511_total': sum(tickets_501511_data) if tickets_501511_data else 0
                }

                total_pa = sum(pa_motif_data) if pa_motif_data else 0
                total_cm = sum(cm_motif_data) if cm_motif_data else 0
                total_upr = sum(upr_motif_data) if upr_motif_data else 0
                total_501511 = sum(tickets_501511_data) if tickets_501511_data else 0
                self.logger.info(f"Facturation detailed data prepared:")
                self.logger.info(f"   PA motifs={len(pa_motif_data)}, CM motifs={len(cm_motif_data)}")
                self.logger.info(f"   UPR motifs={len(upr_motif_data)}, 501/511 motifs={len(tickets_501511_data)}")
                self.logger.info(f"   Total PA: {total_pa:,}, Total CM: {total_cm:,}")
                self.logger.info(f"   Total UPR: {total_upr:,}, Total 501/511: {total_501511:,}")

            else:
                self.logger.warning("No dashboard mapping available for facturation")
                self.facturation_data = {
                    'pa_motifs': [0] * 10,
                    'cm_motifs': [0] * 3,
                    'upr_motifs': [0] * 2,
                    'tickets_501511_motifs': [0] * 1,
                    'pa_total': 0,
                    'cm_total': 0,
                    'upr_total': 0,
                    'tickets_501511_total': 0
                }

        except Exception as e:
            self.logger.error(f"Error updating facturation data: {e}")
            self.facturation_data = {
                'pa_motifs': [0] * 10,
                'cm_motifs': [0] * 3,
                'upr_motifs': [0] * 2,
                'tickets_501511_motifs': [0] * 1,
                'pa_total': 0,
                'cm_total': 0,
                'upr_total': 0,
                'tickets_501511_total': 0
            }

    def _update_script_js_values(self, script_path, stats):
        """Update the script.js file with new statistical values."""
        try:
            with open(script_path, 'r', encoding='utf-8') as f:
                script_content = f.read()

            updated_script = script_content

            # Map filtered statistics to dashboard categories
            self.logger.info("Mapping filtered statistics to dashboard categories...")
            dashboard_mapping = self._map_stats_to_dashboard_categories(stats)
            self.logger.info(f"Dashboard mapping result: {dashboard_mapping}")

            # Update Chart.js data arrays based on the mapping
            if dashboard_mapping:
                # Update communes data if available
                if 'communes' in dashboard_mapping:
                    communes_data = dashboard_mapping['communes']
                    communes_chart_data = communes_data.get('data', [56, 4])
                    self.logger.info(f"Updating communes chart data: {communes_chart_data}")
                    # Update the communes chart data: [56, 4] -> new values
                    updated_script = self._update_chart_data(updated_script, 'communesCtx', communes_chart_data)

                # Update CM data if available
                if 'cm' in dashboard_mapping:
                    cm_data = dashboard_mapping['cm']
                    cm_chart_data = cm_data.get('data', [806, 17, 71])
                    self.logger.info(f"Updating CM chart data: {cm_chart_data}")
                    # Update the CM chart data: [806, 17, 71] -> new values
                    updated_script = self._update_chart_data(updated_script, 'cmCtx', cm_chart_data)

                # Update quality control data if available
                if 'quality' in dashboard_mapping:
                    quality_data = dashboard_mapping['quality']
                    # Update quality chart data: [37, 25] -> new values
                    updated_script = self._update_chart_data(updated_script, 'qualityCtx', quality_data.get('data', [37, 25]))

                # Update acts data if available
                if 'acts' in dashboard_mapping:
                    acts_data = dashboard_mapping['acts']
                    acts_chart_data = acts_data.get('data', [2324, 6023, 584, 143, 23, 930, 1084, 180, 16, 14])
                    self.logger.info(f"Updating acts chart data: {len(acts_chart_data)} categories")
                    self.logger.info(f"   Acts data: {acts_chart_data}")
                    # Update acts chart data with real filtered values
                    updated_script = self._update_chart_data(updated_script, 'actsCtx', acts_chart_data)

                # Update RIP data if available
                if 'rip' in dashboard_mapping:
                    rip_data = dashboard_mapping['rip']
                    rip_chart_data = rip_data.get('data', [0, 0, 0])
                    self.logger.info(f"Updating RIP chart data: {len(rip_chart_data)} categories")
                    self.logger.info(f"   RIP data: {rip_chart_data}")
                    # Update RIP chart data with real filtered values
                    updated_script = self._update_chart_data(updated_script, 'ripCtx', rip_chart_data)

                # Note: UPR and 501/511 sections use simple number displays instead of charts
                # No script.js updates needed for these sections

            # Update detailed facturation data if available
            if hasattr(self, 'facturation_data') and self.facturation_data:
                pa_motifs = self.facturation_data.get('pa_motifs', [])
                cm_motifs = self.facturation_data.get('cm_motifs', [])
                upr_motifs = self.facturation_data.get('upr_motifs', [])
                tickets_501511_motifs = self.facturation_data.get('tickets_501511_motifs', [])

                self.logger.info(f"Updating detailed facturation data in script.js:")
                self.logger.info(f"   PA motifs: {len(pa_motifs)} values")
                self.logger.info(f"   CM motifs: {len(cm_motifs)} values")
                self.logger.info(f"   UPR motifs: {len(upr_motifs)} values")
                self.logger.info(f"   501/511 motifs: {len(tickets_501511_motifs)} values")

                # Create JavaScript arrays for the data
                pa_data_str = ', '.join(str(count) for count in pa_motifs) if pa_motifs else '0'
                cm_data_str = ', '.join(str(count) for count in cm_motifs) if cm_motifs else '0'
                upr_data_str = ', '.join(str(count) for count in upr_motifs) if upr_motifs else '0'
                tickets_501511_data_str = ', '.join(str(count) for count in tickets_501511_motifs) if tickets_501511_motifs else '0'

                # Add detailed facturation data update at the end
                facturation_update = f'''
// Update detailed facturation data with real values
if (typeof detailedBillingCalculator !== 'undefined' && detailedBillingCalculator) {{
    detailedBillingCalculator.updateWithRealData([{pa_data_str}], [{cm_data_str}], [{upr_data_str}], [{tickets_501511_data_str}]);
}}
'''
                updated_script += facturation_update
                self.logger.info(f"   Added detailed facturation data update")
                self.logger.info(f"   PA data: [{pa_data_str}]")
                self.logger.info(f"   CM data: [{cm_data_str}]")
                self.logger.info(f"   UPR data: [{upr_data_str}]")
                self.logger.info(f"   501/511 data: [{tickets_501511_data_str}]")

            # Add timestamp comment
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            comment = f'\n// Pladria Statistics Updated: {timestamp}\n// Period: {stats["period"]["start_date"]} to {stats["period"]["end_date"]}\n// Total Records: {stats["data_summary"]["total_records"]}\n'
            updated_script = comment + updated_script

            # Write updated script
            with open(script_path, 'w', encoding='utf-8') as f:
                f.write(updated_script)

            self.logger.info(f"Updated script.js with new statistical values")

        except Exception as e:
            self.logger.error(f"Error updating script.js values: {e}")

    def _map_stats_to_dashboard_categories(self, stats):
        """Map filtered statistics to dashboard categories using actual Suivi Global data."""
        try:
            mapping = {}

            # Extract CM data from Sheet 2 (Traitement CMS Adr) of the Suivi Global file
            self.logger.info("Attempting to extract CM data from Suivi Global...")
            cm_data = self._extract_cm_data_for_dashboard()
            if cm_data:
                mapping['cm'] = cm_data
                self.logger.info(f"✅ CM data successfully mapped: {cm_data['data']}")
                self.logger.info(f"   Total CM records processed: {cm_data['total_records']}")
                self.logger.info(f"   CM motif breakdown: {cm_data['motif_breakdown']}")
            else:
                self.logger.warning("❌ No CM data extracted - using fallback values")

            # Extract Communes data from Sheet 1 (Suivi Tickets) of the Suivi Global file
            self.logger.info("Attempting to extract Communes data from Suivi Global...")
            communes_data = self._extract_communes_data_for_dashboard()
            if communes_data:
                mapping['communes'] = communes_data
                self.logger.info(f"✅ Communes data successfully mapped: {communes_data['data']}")
                self.logger.info(f"   Total Communes records processed: {communes_data['total_records']}")
                self.logger.info(f"   Communes breakdown: {communes_data['commune_breakdown']}")
            else:
                self.logger.warning("❌ No Communes data extracted - using fallback values")

            # Extract Acts data from Sheet 3 (Traitement PA) of the Suivi Global file
            self.logger.info("Attempting to extract Acts data from Suivi Global...")
            acts_data = self._extract_acts_data_for_dashboard()
            if acts_data:
                mapping['acts'] = acts_data
                self.logger.info(f"✅ Acts data successfully mapped: {len(acts_data['data'])} categories")
                self.logger.info(f"   Total Acts records processed: {acts_data['total_records']}")
                self.logger.info(f"   Acts categories: {acts_data['labels']}")
                self.logger.info(f"   Acts counts: {acts_data['data']}")
            else:
                self.logger.warning("❌ No Acts data extracted - using fallback values")

            # Extract UPR tickets data from Sheet 1 (Suivi Tickets) of the Suivi Global file
            self.logger.info("Attempting to extract UPR tickets data from Suivi Global...")
            upr_data = self._extract_upr_data_for_dashboard()
            if upr_data:
                mapping['upr'] = upr_data
                self.logger.info(f"✅ UPR data successfully mapped: {len(upr_data['data'])} categories")
                self.logger.info(f"   Total UPR records processed: {upr_data['total_records']}")
                self.logger.info(f"   UPR categories: {upr_data['labels']}")
                self.logger.info(f"   UPR counts: {upr_data['data']}")
            else:
                self.logger.warning("❌ No UPR data extracted - using fallback values")

            # Extract 501/511 tickets data from Sheet 1 (Suivi Tickets) of the Suivi Global file
            self.logger.info("Attempting to extract 501/511 tickets data from Suivi Global...")
            tickets_501511_data = self._extract_501511_data_for_dashboard()
            if tickets_501511_data:
                mapping['tickets_501511'] = tickets_501511_data
                self.logger.info(f"✅ 501/511 data successfully mapped: {len(tickets_501511_data['data'])} categories")
                self.logger.info(f"   Total 501/511 records processed: {tickets_501511_data['total_records']}")
                self.logger.info(f"   501/511 categories: {tickets_501511_data['labels']}")
                self.logger.info(f"   501/511 counts: {tickets_501511_data['data']}")
            else:
                self.logger.warning("❌ No 501/511 data extracted - using fallback values")

            # Extract RIP (P0 P1) data from Sheet 4 (Traitement RIP) of the Suivi Global file
            self.logger.info("Attempting to extract RIP (P0 P1) data from Suivi Global...")
            rip_data = self._extract_rip_data_for_dashboard()
            if rip_data:
                mapping['rip'] = rip_data
                self.logger.info(f"✅ RIP data successfully mapped: {len(rip_data['data'])} categories")
                self.logger.info(f"   Total RIP records processed: {rip_data['total_records']}")
                self.logger.info(f"   RIP categories: {rip_data['labels']}")
                self.logger.info(f"   RIP counts: {rip_data['data']}")
            else:
                self.logger.warning("❌ No RIP data extracted - using fallback values")

            # Keep existing logic for other categories as fallback
            if stats.get('motifs', {}).get('sorted'):
                top_motifs = dict(stats['motifs']['sorted'][:10])

                # Map to communes (example - you'll need to adjust based on your data)
                if 'Orange' in top_motifs or 'RIP' in top_motifs:
                    mapping['communes'] = {
                        'data': [
                            top_motifs.get('Orange', 56),
                            top_motifs.get('RIP', 4)
                        ]
                    }

                # Map to quality control (example)
                total_records = stats['data_summary']['total_records']
                conformes = int(total_records * 0.6)  # Example: 60% conformes
                non_conformes = total_records - conformes

                mapping['quality'] = {
                    'data': [conformes, non_conformes]
                }

                # Map to acts data (example - you'll need to map your actual categories)
                mapping['acts'] = {
                    'data': [
                        top_motifs.get('AD_RAS_AVEC_TEMPS', 2324),
                        top_motifs.get('AD_RAS_SANS_TEMPS', 6023),
                        top_motifs.get('AD_NON_JOINTE', 584),
                        top_motifs.get('AD_NON_TROUVEE', 143),
                        top_motifs.get('HORS_COMMUNE', 23),
                        top_motifs.get('NOK', 930),
                        top_motifs.get('OK', 1084),
                        top_motifs.get('UPR_RAS', 180),
                        top_motifs.get('UPR_NOK', 16),
                        top_motifs.get('UPR_OK', 14)
                    ]
                }

            return mapping

        except Exception as e:
            self.logger.error(f"Error mapping stats to dashboard categories: {e}")
            return {}

    def _extract_cm_data_for_dashboard(self):
        """Extract CM data from Sheet 2 (Traitement CMS Adr) for dashboard population."""
        try:
            from datetime import datetime
            # Check if we have the required data
            if not hasattr(self, 'global_suivi_data') or not self.global_suivi_data:
                self.logger.warning("No global_suivi_data available for CM extraction")
                return None

            if 'Traitement CMS Adr' not in self.global_suivi_data:
                self.logger.warning("Sheet 'Traitement CMS Adr' not found in global_suivi_data")
                return None

            df_cms = self.global_suivi_data['Traitement CMS Adr']
            if df_cms.empty:
                self.logger.warning("Sheet 'Traitement CMS Adr' is empty")
                return None

            # Check if we have the required date range
            if not hasattr(self, 'date_from_selected') or not hasattr(self, 'date_to_selected'):
                self.logger.warning("Date range not set for CM data extraction")
                return None

            if not self.date_from_selected or not self.date_to_selected:
                self.logger.warning("Date range values are None")
                return None

            self.logger.info(f"Extracting CM data for period: {self.date_from_selected} to {self.date_to_selected}")

            # Check column structure
            if len(df_cms.columns) < 8:  # Need at least 8 columns (A-H)
                self.logger.warning(f"CM sheet: Not enough columns (found {len(df_cms.columns)}, need at least 8)")
                return None

            # Column mapping based on the existing code structure
            motif_column = df_cms.columns[3]  # Column D (Motif Voie)

            # Try to find delivery date column (Column H, index 7)
            delivery_date_column = None
            if len(df_cms.columns) > 7:
                delivery_date_column = df_cms.columns[7]  # Column H (delivery date)
                self.logger.info(f"Using delivery date column H: '{delivery_date_column}'")
            else:
                # Fallback to processing date (Column G, index 6)
                delivery_date_column = df_cms.columns[6]  # Column G (processing date)
                self.logger.info(f"Fallback to processing date column G: '{delivery_date_column}'")

            self.logger.info(f"CM extraction: Using motif column '{motif_column}' and date column '{delivery_date_column}'")

            # Extract and count motifs within the date range
            motif_counts = {}
            total_processed = 0

            for index, row in df_cms.iterrows():
                try:
                    # Get motif value
                    motif_value = row.get(motif_column, '')
                    if not motif_value or str(motif_value).strip() == '' or str(motif_value).lower() == 'nan':
                        continue
                    motif_value = str(motif_value).strip().upper()  # Normalize to uppercase

                    # Get date value
                    date_value = row.get(delivery_date_column, '')
                    if not date_value or str(date_value).strip() == '' or str(date_value).lower() == 'nan':
                        continue

                    # Parse date (try multiple formats)
                    date_obj = None
                    for date_format in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
                        try:
                            date_obj = datetime.strptime(str(date_value), date_format).date()
                            break
                        except ValueError:
                            continue

                    if not date_obj:
                        continue

                    # Check if date is within selected range
                    if self.date_from_selected <= date_obj <= self.date_to_selected:
                        # Count this motif
                        if motif_value not in motif_counts:
                            motif_counts[motif_value] = 0
                        motif_counts[motif_value] += 1
                        total_processed += 1

                except Exception as e:
                    self.logger.debug(f"Error processing CM row {index}: {e}")
                    continue

            self.logger.info(f"CM extraction completed: {total_processed} records processed, {len(motif_counts)} unique motifs found")

            # Log the motif counts for debugging
            for motif, count in sorted(motif_counts.items(), key=lambda x: x[1], reverse=True):
                self.logger.info(f"  {motif}: {count}")

            # Map motifs to CM chart categories
            # The original dashboard has RAF, MODIF, CREA categories
            # We'll map the actual motifs found to these categories or use the top motifs

            cm_chart_data = self._map_motifs_to_cm_categories(motif_counts)

            return {
                'data': cm_chart_data,
                'total_records': total_processed,
                'motif_breakdown': motif_counts
            }

        except Exception as e:
            self.logger.error(f"Error extracting CM data for dashboard: {e}")
            return None

    def _extract_communes_data_for_dashboard(self):
        """Extract Communes data from Sheet 1 (Suivi Tickets) for dashboard population."""
        try:
            from datetime import datetime

            # Check if we have the required data
            if not hasattr(self, 'global_suivi_data') or not self.global_suivi_data:
                self.logger.warning("No global_suivi_data available for Communes extraction")
                return None

            if 'Suivi Tickets' not in self.global_suivi_data:
                self.logger.warning("Sheet 'Suivi Tickets' not found in global_suivi_data")
                return None

            df_tickets = self.global_suivi_data['Suivi Tickets']
            if df_tickets.empty:
                self.logger.warning("Sheet 'Suivi Tickets' is empty")
                return None

            # Check if we have the required date range
            if not hasattr(self, 'date_from_selected') or not hasattr(self, 'date_to_selected'):
                self.logger.warning("Date range not set for Communes data extraction")
                return None

            if not self.date_from_selected or not self.date_to_selected:
                self.logger.warning("Date range values are None")
                return None

            self.logger.info(f"Extracting Communes data for period: {self.date_from_selected} to {self.date_to_selected}")

            # Check column structure - need at least 15 columns (A-O)
            if len(df_tickets.columns) < 15:
                self.logger.warning(f"Tickets sheet: Not enough columns (found {len(df_tickets.columns)}, need at least 15)")
                return None

            # Column mapping based on requirements
            commune_type_column = df_tickets.columns[3]  # Column D (Commune Type)
            delivery_date_column = df_tickets.columns[14]  # Column O (Delivery Date)

            self.logger.info(f"Communes extraction: Using commune type column '{commune_type_column}' and delivery date column '{delivery_date_column}'")

            # Extract and count commune types within the date range
            commune_counts = {'Orange': 0, 'RIP': 0}
            total_processed = 0

            for index, row in df_tickets.iterrows():
                try:
                    # Get commune type value
                    commune_type = row.get(commune_type_column, '')
                    if not commune_type or str(commune_type).strip() == '' or str(commune_type).lower() == 'nan':
                        continue
                    commune_type = str(commune_type).strip()

                    # Normalize commune type (case-insensitive)
                    commune_type_upper = commune_type.upper()
                    if commune_type_upper not in ['ORANGE', 'RIP']:
                        continue  # Skip if not Orange or RIP

                    # Get delivery date value
                    delivery_date = row.get(delivery_date_column, '')
                    if not delivery_date or str(delivery_date).strip() == '' or str(delivery_date).lower() == 'nan':
                        continue

                    # Parse delivery date (handle datetime format: "2025-05-22 00:00:00")
                    date_obj = None
                    delivery_date_str = str(delivery_date).strip()

                    # Try multiple datetime formats
                    date_formats = [
                        '%Y-%m-%d %H:%M:%S',  # "2025-05-22 00:00:00"
                        '%Y-%m-%d',           # "2025-05-22"
                        '%d/%m/%Y %H:%M:%S',  # "22/05/2025 00:00:00"
                        '%d/%m/%Y',           # "22/05/2025"
                        '%d-%m-%Y %H:%M:%S',  # "22-05-2025 00:00:00"
                        '%d-%m-%Y'            # "22-05-2025"
                    ]

                    for date_format in date_formats:
                        try:
                            date_obj = datetime.strptime(delivery_date_str, date_format).date()
                            break
                        except ValueError:
                            continue

                    if not date_obj:
                        continue

                    # Check if date is within selected range
                    if self.date_from_selected <= date_obj <= self.date_to_selected:
                        # Count this commune type
                        if commune_type_upper == 'ORANGE':
                            commune_counts['Orange'] += 1
                        elif commune_type_upper == 'RIP':
                            commune_counts['RIP'] += 1

                        total_processed += 1

                except Exception as e:
                    self.logger.debug(f"Error processing Communes row {index}: {e}")
                    continue

            self.logger.info(f"Communes extraction completed: {total_processed} records processed")
            self.logger.info(f"  Orange: {commune_counts['Orange']}")
            self.logger.info(f"  RIP: {commune_counts['RIP']}")

            # Prepare data for dashboard (Orange first, then RIP to match chart labels)
            communes_chart_data = [commune_counts['Orange'], commune_counts['RIP']]

            return {
                'data': communes_chart_data,
                'total_records': total_processed,
                'commune_breakdown': commune_counts
            }

        except Exception as e:
            self.logger.error(f"Error extracting Communes data for dashboard: {e}")
            return None

    def _extract_acts_data_for_dashboard(self):
        """Extract Acts data from Sheet 3 (Traitement PA) for dashboard population."""
        try:
            from datetime import datetime

            # Check if we have the required data
            if not hasattr(self, 'global_suivi_data') or not self.global_suivi_data:
                self.logger.warning("No global_suivi_data available for Acts extraction")
                return None

            if 'Traitement PA' not in self.global_suivi_data:
                self.logger.warning("Sheet 'Traitement PA' not found in global_suivi_data")
                return None

            df_acts = self.global_suivi_data['Traitement PA']
            if df_acts.empty:
                self.logger.warning("Sheet 'Traitement PA' is empty")
                return None

            # Check if we have the required date range
            if not hasattr(self, 'date_from_selected') or not hasattr(self, 'date_to_selected'):
                self.logger.warning("Date range not set for Acts data extraction")
                return None

            if not self.date_from_selected or not self.date_to_selected:
                self.logger.warning("Date range values are None")
                return None

            self.logger.info(f"Extracting Acts data for period: {self.date_from_selected} to {self.date_to_selected}")

            # Check column structure - need at least 8 columns (A-H)
            if len(df_acts.columns) < 8:
                self.logger.warning(f"Acts sheet: Not enough columns (found {len(df_acts.columns)}, need at least 8)")
                return None

            # Column mapping based on requirements
            motif_column = df_acts.columns[3]  # Column D (Motif)
            processing_date_column = df_acts.columns[6]  # Column G (Date traitement)
            duration_column = df_acts.columns[7]  # Column H (Durée)

            self.logger.info(f"Acts extraction: Using motif column '{motif_column}', date column '{processing_date_column}', duration column '{duration_column}'")

            # Extract and count motifs within the date range
            motif_counts = {}
            total_processed = 0

            for index, row in df_acts.iterrows():
                try:
                    # Get motif value
                    motif = row.get(motif_column, '')
                    if not motif or str(motif).strip() == '' or str(motif).lower() == 'nan':
                        continue
                    motif = str(motif).strip()

                    # Get processing date value
                    processing_date = row.get(processing_date_column, '')
                    if not processing_date or str(processing_date).strip() == '' or str(processing_date).lower() == 'nan':
                        continue

                    # Parse processing date (handle date format: "2025-07-02")
                    date_obj = None
                    processing_date_str = str(processing_date).strip()

                    # Try multiple date formats
                    date_formats = [
                        '%Y-%m-%d',           # "2025-07-02"
                        '%Y-%m-%d %H:%M:%S',  # "2025-07-02 00:00:00"
                        '%d/%m/%Y',           # "02/07/2025"
                        '%d/%m/%Y %H:%M:%S',  # "02/07/2025 00:00:00"
                        '%d-%m-%Y',           # "02-07-2025"
                        '%d-%m-%Y %H:%M:%S'   # "02-07-2025 00:00:00"
                    ]

                    for date_format in date_formats:
                        try:
                            date_obj = datetime.strptime(processing_date_str, date_format).date()
                            break
                        except ValueError:
                            continue

                    if not date_obj:
                        continue

                    # Check if date is within selected range
                    if self.date_from_selected <= date_obj <= self.date_to_selected:
                        # Get duration for special Ad Ras handling
                        duration = row.get(duration_column, 0)
                        try:
                            duration_val = float(duration) if duration is not None else 0
                        except (ValueError, TypeError):
                            duration_val = 0

                        # Normalize motif for categorization
                        motif_normalized = motif.upper().strip()

                        # Special handling for Ad Ras motifs
                        if 'AD' in motif_normalized and 'RAS' in motif_normalized:
                            if duration_val > 0:
                                category = 'AD RAS avec temps'
                            else:
                                category = 'AD RAS sans temps'
                        else:
                            # Map other motifs to standard categories
                            category = self._normalize_acts_motif(motif_normalized)

                        # Count this motif category
                        if category in motif_counts:
                            motif_counts[category] += 1
                        else:
                            motif_counts[category] = 1

                        total_processed += 1

                except Exception as e:
                    self.logger.debug(f"Error processing Acts row {index}: {e}")
                    continue

            self.logger.info(f"Acts extraction completed: {total_processed} records processed")
            self.logger.info(f"  Unique motif categories found: {len(motif_counts)}")

            # Log all found categories with counts
            for category, count in sorted(motif_counts.items(), key=lambda x: x[1], reverse=True):
                self.logger.info(f"    {category}: {count}")

            # Special logging for NOK/UPR categories to debug the issue
            nok_count = motif_counts.get('NOK', 0)
            upr_nok_count = motif_counts.get('UPR NOK', 0)
            upr_ok_count = motif_counts.get('UPR OK', 0)

            self.logger.info(f"  Debug - Specific categories:")
            self.logger.info(f"    NOK: {nok_count}")
            self.logger.info(f"    UPR NOK: {upr_nok_count}")
            self.logger.info(f"    UPR OK: {upr_ok_count}")

            if nok_count == 0 and upr_nok_count == 0 and upr_ok_count == 0:
                self.logger.warning(f"  ⚠️ No NOK/UPR categories found - this may explain 0 values in dashboard")
                self.logger.warning(f"  Check if these motifs exist in the data or need different normalization")

            # Prepare data for dashboard - MUST match the exact order of labels in script.js AND HTML
            # Order matches HTML structure: AD RAS sans temps first (most frequent), then avec temps, etc.
            chart_labels_order = [
                'AD RAS sans temps',
                'AD RAS avec temps',
                'OK',
                'NOK',
                'AD Non jointe',
                'UPR RAS',
                'AD Non trouvée',
                'Hors commune',
                'UPR NOK',
                'UPR OK'
            ]

            # Map extracted data to the fixed chart order
            acts_chart_data = []
            acts_labels = []

            for label in chart_labels_order:
                count = motif_counts.get(label, 0)  # Get count or 0 if category not found
                acts_chart_data.append(count)
                acts_labels.append(label)

            self.logger.info(f"Acts data mapped to chart order:")
            for i, (label, count) in enumerate(zip(acts_labels, acts_chart_data)):
                self.logger.info(f"  {i+1}. {label}: {count}")

            return {
                'data': acts_chart_data,
                'labels': acts_labels,
                'total_records': total_processed,
                'motif_breakdown': motif_counts
            }

        except Exception as e:
            self.logger.error(f"Error extracting Acts data for dashboard: {e}")
            return None

    def _normalize_acts_motif(self, motif_upper):
        """Normalize Acts motifs to standard categories that match Chart.js labels."""
        # Define motif mappings based on the exact Chart.js labels order
        # Chart labels: ['AD RAS\navec temps', 'AD RAS\nsans temps', 'AD Non jointe', 'AD Non trouvée', 'Hors commune', 'NOK', 'OK', 'UPR RAS', 'UPR NOK', 'UPR OK']

        # Log the original motif for debugging
        self.logger.debug(f"Normalizing motif: '{motif_upper}'")

        # More comprehensive motif mappings with extensive variations
        # IMPORTANT: Order matters! More specific patterns first to avoid false matches
        motif_mappings = [
            # UPR patterns first (most specific)
            ('UPR NOK', ['UPR NOK', 'UPR-NOK', 'UPR_NOK', 'UPRNOK', 'UPR KO', 'UPR-KO', 'UPR_KO', 'UPRKO', 'UPR NOT OK', 'UPR NOTOK']),
            ('UPR OK', ['UPR OK', 'UPR-OK', 'UPR_OK', 'UPROK', 'UPR VALIDE', 'UPR CORRECT', 'UPR GOOD']),
            ('UPR RAS', ['UPR RAS', 'UPR-RAS', 'UPR_RAS', 'UPRRAS', 'UPR RIEN', 'UPR NOTHING']),

            # AD patterns
            ('AD Non jointe', ['AD NON JOINTE', 'ADRESSE NON JOINTE', 'AD_NON_JOINTE', 'ADRESSE_NON_JOINTE', 'AD NON JOINT', 'ADRESSE NON JOINT']),
            ('AD Non trouvée', ['AD NON TROUVEE', 'AD NON TROUVÉE', 'ADRESSE NON TROUVEE', 'ADRESSE NON TROUVÉE', 'AD_NON_TROUVEE', 'ADRESSE_NON_TROUVEE', 'AD NON TROUVE', 'ADRESSE NON TROUVE']),

            # Other patterns
            ('Hors commune', ['HORS COMMUNE', 'HORS_COMMUNE', 'HORSCOMMUNE', 'HORS COM', 'HORS-COMMUNE']),

            # NOK patterns (before OK to avoid false matches)
            ('NOK', ['NOK', 'KO', 'NOT OK', 'NOTOK', 'INVALIDE', 'INCORRECT', 'ERREUR', 'ERROR', 'MAUVAIS', 'BAD']),

            # OK patterns (last to avoid matching NOK)
            ('OK', ['OK', 'VALIDE', 'CORRECT', 'GOOD'])
        ]

        # Check each mapping with exact match first, then contains
        for category, patterns in motif_mappings:
            # First try exact matches
            for pattern in patterns:
                if motif_upper == pattern:
                    self.logger.debug(f"Exact match: '{motif_upper}' → '{category}'")
                    return category

            # Then try contains matches (but be careful with overlaps)
            for pattern in patterns:
                if pattern in motif_upper:
                    # Special check to avoid false matches
                    if category == 'OK' and ('NOK' in motif_upper or 'KO' in motif_upper.replace('OK', '')):
                        continue  # Skip OK match if NOK is also present
                    self.logger.debug(f"Contains match: '{motif_upper}' contains '{pattern}' → '{category}'")
                    return category

        # Special handling for UPR without suffix
        if motif_upper == 'UPR':
            self.logger.debug(f"UPR fallback: '{motif_upper}' → 'UPR RAS'")
            return 'UPR RAS'

        # Special handling for standalone patterns that might be missed
        if 'UPR' in motif_upper:
            if any(nok_word in motif_upper for nok_word in ['NOK', 'KO', 'NOT', 'ERREUR', 'ERROR']):
                self.logger.debug(f"UPR+NOK pattern: '{motif_upper}' → 'UPR NOK'")
                return 'UPR NOK'
            elif any(ok_word in motif_upper for ok_word in ['OK', 'VALIDE', 'CORRECT', 'GOOD']):
                self.logger.debug(f"UPR+OK pattern: '{motif_upper}' → 'UPR OK'")
                return 'UPR OK'
            else:
                self.logger.debug(f"UPR generic: '{motif_upper}' → 'UPR RAS'")
                return 'UPR RAS'

        # If no match found, log and return the original motif (cleaned)
        self.logger.debug(f"No match found: '{motif_upper}' → '{motif_upper}' (unchanged)")
        return motif_upper

    def _extract_upr_data_for_dashboard(self):
        """Extract UPR tickets data from Sheet 1 (Suivi Tickets) for dashboard population."""
        try:
            from datetime import datetime

            # Check if we have the required data
            if not hasattr(self, 'global_suivi_data') or not self.global_suivi_data:
                self.logger.warning("No global_suivi_data available for UPR extraction")
                return None

            if 'Suivi Tickets' not in self.global_suivi_data:
                self.logger.warning("Sheet 'Suivi Tickets' not found in global_suivi_data")
                return None

            df_tickets = self.global_suivi_data['Suivi Tickets']
            if df_tickets.empty:
                self.logger.warning("Sheet 'Suivi Tickets' is empty")
                return None

            # Check if we have the required date range
            if not hasattr(self, 'date_from_selected') or not hasattr(self, 'date_to_selected'):
                self.logger.warning("Date range not set for UPR data extraction")
                return None

            if not self.date_from_selected or not self.date_to_selected:
                self.logger.warning("Date range values are None")
                return None

            self.logger.info(f"Extracting UPR data for period: {self.date_from_selected} to {self.date_to_selected}")

            # Check column structure - need at least 19 columns (A-S)
            if len(df_tickets.columns) < 19:
                self.logger.warning(f"UPR sheet: Not enough columns (found {len(df_tickets.columns)}, need at least 19)")
                return None

            # Column mapping based on analysis
            upr_motif_column = df_tickets.columns[18]  # Column S (Dépose Ticket UPR)
            delivery_date_column = df_tickets.columns[14]  # Column O (Date Livraison)

            self.logger.info(f"UPR extraction: Using motif column '{upr_motif_column}', date column '{delivery_date_column}'")

            # Extract and count UPR tickets within the date range
            upr_counts = {'Créé': 0, 'Non': 0}
            total_processed = 0

            for index, row in df_tickets.iterrows():
                try:
                    # Get UPR motif value
                    upr_motif = row.get(upr_motif_column, '')
                    if not upr_motif or str(upr_motif).strip() == '' or str(upr_motif).lower() == 'nan':
                        continue
                    upr_motif = str(upr_motif).strip()

                    # Get delivery date value
                    delivery_date = row.get(delivery_date_column, '')
                    if not delivery_date or str(delivery_date).strip() == '' or str(delivery_date).lower() == 'nan':
                        continue

                    # Parse delivery date (handle format: "2025-05-22 00:00:00")
                    date_obj = None
                    delivery_date_str = str(delivery_date).strip()

                    # Try multiple date formats
                    date_formats = [
                        '%Y-%m-%d %H:%M:%S',  # "2025-05-22 00:00:00"
                        '%Y-%m-%d',           # "2025-05-22"
                        '%d/%m/%Y %H:%M:%S',  # "22/05/2025 00:00:00"
                        '%d/%m/%Y',           # "22/05/2025"
                        '%d-%m-%Y %H:%M:%S',  # "22-05-2025 00:00:00"
                        '%d-%m-%Y'            # "22-05-2025"
                    ]

                    for date_format in date_formats:
                        try:
                            date_obj = datetime.strptime(delivery_date_str, date_format).date()
                            break
                        except ValueError:
                            continue

                    if not date_obj:
                        continue

                    # Check if date is within selected range
                    if self.date_from_selected <= date_obj <= self.date_to_selected:
                        # Normalize UPR motif
                        upr_motif_normalized = upr_motif.strip()

                        # Count UPR tickets by status
                        if upr_motif_normalized in upr_counts:
                            upr_counts[upr_motif_normalized] += 1
                        else:
                            # Handle unexpected values
                            upr_counts[upr_motif_normalized] = 1

                        total_processed += 1

                except Exception as e:
                    self.logger.debug(f"Error processing UPR row {index}: {e}")
                    continue

            self.logger.info(f"UPR extraction completed: {total_processed} records processed")
            self.logger.info(f"  UPR ticket counts: {upr_counts}")

            # Prepare data for dashboard - fixed order for chart consistency
            # Order: ['Créé', 'Non'] to match expected chart structure
            chart_labels_order = ['Créé', 'Non']

            upr_chart_data = []
            upr_labels = []

            for label in chart_labels_order:
                count = upr_counts.get(label, 0)
                upr_chart_data.append(count)
                upr_labels.append(label)

            self.logger.info(f"UPR data mapped to chart order:")
            for i, (label, count) in enumerate(zip(upr_labels, upr_chart_data)):
                self.logger.info(f"  {i+1}. {label}: {count}")

            return {
                'data': upr_chart_data,
                'labels': upr_labels,
                'total_records': total_processed,
                'upr_breakdown': upr_counts
            }

        except Exception as e:
            self.logger.error(f"Error extracting UPR data for dashboard: {e}")
            return None

    def _extract_501511_data_for_dashboard(self):
        """Extract 501/511 tickets data from Sheet 1 (Suivi Tickets) for dashboard population."""
        try:
            from datetime import datetime

            # Check if we have the required data
            if not hasattr(self, 'global_suivi_data') or not self.global_suivi_data:
                self.logger.warning("No global_suivi_data available for 501/511 extraction")
                return None

            if 'Suivi Tickets' not in self.global_suivi_data:
                self.logger.warning("Sheet 'Suivi Tickets' not found in global_suivi_data")
                return None

            df_tickets = self.global_suivi_data['Suivi Tickets']
            if df_tickets.empty:
                self.logger.warning("Sheet 'Suivi Tickets' is empty")
                return None

            # Check if we have the required date range
            if not hasattr(self, 'date_from_selected') or not hasattr(self, 'date_to_selected'):
                self.logger.warning("Date range not set for 501/511 data extraction")
                return None

            if not self.date_from_selected or not self.date_to_selected:
                self.logger.warning("Date range values are None")
                return None

            self.logger.info(f"Extracting 501/511 data for period: {self.date_from_selected} to {self.date_to_selected}")

            # Check column structure - need at least 18 columns (A-R)
            if len(df_tickets.columns) < 18:
                self.logger.warning(f"501/511 sheet: Not enough columns (found {len(df_tickets.columns)}, need at least 18)")
                return None

            # Column mapping based on analysis
            deposit_date_column = df_tickets.columns[17]  # Column R (Date Dépose Ticket 501/511)

            self.logger.info(f"501/511 extraction: Using deposit date column '{deposit_date_column}'")

            # Extract and count 501/511 tickets within the date range
            tickets_501511_count = 0
            total_processed = 0

            for index, row in df_tickets.iterrows():
                try:
                    # Get deposit date value
                    deposit_date = row.get(deposit_date_column, '')
                    if not deposit_date or str(deposit_date).strip() == '' or str(deposit_date).lower() == 'nan':
                        continue

                    # Parse deposit date (handle format: "2025-07-02")
                    date_obj = None
                    deposit_date_str = str(deposit_date).strip()

                    # Try multiple date formats
                    date_formats = [
                        '%Y-%m-%d',           # "2025-07-02"
                        '%Y-%m-%d %H:%M:%S',  # "2025-07-02 00:00:00"
                        '%d/%m/%Y',           # "02/07/2025"
                        '%d/%m/%Y %H:%M:%S',  # "02/07/2025 00:00:00"
                        '%d-%m-%Y',           # "02-07-2025"
                        '%d-%m-%Y %H:%M:%S'   # "02-07-2025 00:00:00"
                    ]

                    for date_format in date_formats:
                        try:
                            date_obj = datetime.strptime(deposit_date_str, date_format).date()
                            break
                        except ValueError:
                            continue

                    if not date_obj:
                        continue

                    # Check if date is within selected range
                    if self.date_from_selected <= date_obj <= self.date_to_selected:
                        tickets_501511_count += 1
                        total_processed += 1

                except Exception as e:
                    self.logger.debug(f"Error processing 501/511 row {index}: {e}")
                    continue

            self.logger.info(f"501/511 extraction completed: {total_processed} records processed")
            self.logger.info(f"  501/511 tickets count: {tickets_501511_count}")

            # Prepare data for dashboard - simple count for now
            # Can be extended later if subcategories are needed
            tickets_501511_data = [tickets_501511_count]
            tickets_501511_labels = ['501/511 Tickets']

            return {
                'data': tickets_501511_data,
                'labels': tickets_501511_labels,
                'total_records': total_processed,
                'tickets_breakdown': {'501/511 Tickets': tickets_501511_count}
            }

        except Exception as e:
            self.logger.error(f"Error extracting 501/511 data for dashboard: {e}")
            return None

    def _extract_rip_data_for_dashboard(self):
        """Extract RIP (P0 P1) data from Sheet 4 (Traitement RIP) for dashboard population."""
        try:
            from datetime import datetime

            # Check if we have the required data
            if not hasattr(self, 'global_suivi_data') or not self.global_suivi_data:
                self.logger.warning("No global_suivi_data available for RIP extraction")
                return None

            if 'Traitement RIP' not in self.global_suivi_data:
                self.logger.warning("Sheet 'Traitement RIP' not found in global_suivi_data")
                return None

            df_rip = self.global_suivi_data['Traitement RIP']
            if df_rip.empty:
                self.logger.warning("Sheet 'Traitement RIP' is empty")
                return None

            # Check if we have the required date range
            if not hasattr(self, 'date_from_selected') or not hasattr(self, 'date_to_selected'):
                self.logger.warning("Date range not set for RIP data extraction")
                return None

            if not self.date_from_selected or not self.date_to_selected:
                self.logger.warning("Date range values are None")
                return None

            self.logger.info(f"Extracting RIP data for period: {self.date_from_selected} to {self.date_to_selected}")

            # Check column structure - need at least 9 columns (A-I)
            if len(df_rip.columns) < 9:
                self.logger.warning(f"RIP sheet: Not enough columns (found {len(df_rip.columns)}, need at least 9)")
                return None

            # Column mapping based on analysis
            type_column = df_rip.columns[3]  # Column D (Type)
            motif_column = df_rip.columns[4]  # Column E (Acte de traitement)
            delivery_date_column = df_rip.columns[8]  # Column I (Date de livraison)

            self.logger.info(f"RIP extraction: Using type column '{type_column}', motif column '{motif_column}', date column '{delivery_date_column}'")

            # Extract and count RIP motifs within the date range for P0/P1 types
            rip_motif_counts = {}
            total_processed = 0

            for index, row in df_rip.iterrows():
                try:
                    # Get type value and filter for P0/P1
                    type_value = row.get(type_column, '')
                    if not type_value or str(type_value).strip() == '' or str(type_value).lower() == 'nan':
                        continue

                    type_str = str(type_value).upper().strip()
                    if 'P0' not in type_str and 'P1' not in type_str:
                        continue  # Skip non-P0/P1 records

                    # Get motif value
                    motif_value = row.get(motif_column, '')
                    if not motif_value or str(motif_value).strip() == '' or str(motif_value).lower() == 'nan':
                        continue
                    motif_value = str(motif_value).strip()

                    # Get delivery date value
                    delivery_date = row.get(delivery_date_column, '')
                    if not delivery_date or str(delivery_date).strip() == '' or str(delivery_date).lower() == 'nan':
                        continue

                    # Parse delivery date
                    date_obj = None
                    delivery_date_str = str(delivery_date).strip()

                    # Try multiple date formats
                    date_formats = [
                        '%Y-%m-%d %H:%M:%S',  # "2025-05-22 00:00:00"
                        '%Y-%m-%d',           # "2025-05-22"
                        '%d/%m/%Y %H:%M:%S',  # "22/05/2025 00:00:00"
                        '%d/%m/%Y',           # "22/05/2025"
                        '%d-%m-%Y %H:%M:%S',  # "22-05-2025 00:00:00"
                        '%d-%m-%Y'            # "22-05-2025"
                    ]

                    for date_format in date_formats:
                        try:
                            date_obj = datetime.strptime(delivery_date_str, date_format).date()
                            break
                        except ValueError:
                            continue

                    if not date_obj:
                        continue

                    # Check if date is within selected range
                    if self.date_from_selected <= date_obj <= self.date_to_selected:
                        # Normalize RIP motif
                        motif_normalized = self._normalize_rip_motif(motif_value.upper().strip())

                        # Count RIP motifs
                        if motif_normalized in rip_motif_counts:
                            rip_motif_counts[motif_normalized] += 1
                        else:
                            rip_motif_counts[motif_normalized] = 1

                        total_processed += 1

                except Exception as e:
                    self.logger.debug(f"Error processing RIP row {index}: {e}")
                    continue

            self.logger.info(f"RIP extraction completed: {total_processed} records processed")
            self.logger.info(f"  RIP motif counts: {rip_motif_counts}")

            # Prepare data for dashboard - fixed order for chart consistency
            # Order: ['Rien à faire', 'Modification', 'Création'] to match expected structure
            chart_labels_order = ['Rien à faire', 'Modification', 'Création']

            rip_chart_data = []
            rip_labels = []

            for label in chart_labels_order:
                count = rip_motif_counts.get(label, 0)
                rip_chart_data.append(count)
                rip_labels.append(label)

            self.logger.info(f"RIP data mapped to chart order:")
            for i, (label, count) in enumerate(zip(rip_labels, rip_chart_data)):
                self.logger.info(f"  {i+1}. {label}: {count}")

            return {
                'data': rip_chart_data,
                'labels': rip_labels,
                'total_records': total_processed,
                'rip_breakdown': rip_motif_counts
            }

        except Exception as e:
            self.logger.error(f"Error extracting RIP data for dashboard: {e}")
            return None

    def _normalize_rip_motif(self, motif_upper):
        """Normalize RIP motifs to standard categories."""
        # Log the original motif for debugging
        self.logger.debug(f"Normalizing RIP motif: '{motif_upper}'")

        # RIP motif mappings with extensive variations
        # Order matters! More specific patterns first to avoid false matches
        motif_mappings = [
            ('Rien à faire', ['RIEN A FAIRE', 'RIEN À FAIRE', 'RIEN_A_FAIRE', 'RIENAfaire', 'NOTHING TO DO', 'NO ACTION']),
            ('Modification', ['MODIFICATION', 'MODIF', 'MODIFY', 'UPDATE', 'CHANGE', 'EDIT']),
            ('Création', ['CREATION', 'CRÉATION', 'CREATE', 'NEW', 'ADD', 'CREER', 'CRÉER'])
        ]

        # Check each mapping with exact match first, then contains
        for category, patterns in motif_mappings:
            # First try exact matches
            for pattern in patterns:
                if motif_upper == pattern:
                    self.logger.debug(f"Exact match: '{motif_upper}' → '{category}'")
                    return category

            # Then try contains matches
            for pattern in patterns:
                if pattern in motif_upper:
                    self.logger.debug(f"Contains match: '{motif_upper}' contains '{pattern}' → '{category}'")
                    return category

        # If no match found, log and return the original motif (cleaned)
        self.logger.debug(f"No match found: '{motif_upper}' → '{motif_upper}' (unchanged)")
        return motif_upper

    def _map_motifs_to_cm_categories(self, motif_counts):
        """Map actual motifs found to CM chart categories."""
        try:
            # Default CM categories from the original dashboard
            default_categories = ['RAF', 'MODIF', 'CREA']

            # Create mapping from real motif names to dashboard categories
            motif_mapping = {
                # RAF category - "Rien à faire"
                'RAF': ['RAF', 'RAS', 'RIEN À FAIRE', 'RIEN A FAIRE', 'RIEN_A_FAIRE'],
                # MODIF category - "Modification"
                'MODIF': ['MODIF', 'MODIFICATION', 'MODIFICATION VOIE', 'CORRIGER', 'MODIFIER'],
                # CREA category - "Création"
                'CREA': ['CREA', 'CREATION', 'CRÉATION', 'CRÉATION VOIE', 'CREATION VOIE', 'CREER', 'CRÉER', 'NOUVEAU']
            }

            cm_data = []

            # Strategy 1: Map using the motif mapping
            self.logger.info("Mapping motifs using predefined categories...")
            for category in default_categories:
                count = 0
                patterns = motif_mapping.get(category, [])

                # Check each motif in the data
                for motif, motif_count in motif_counts.items():
                    motif_upper = motif.upper().strip()

                    # Check if this motif matches any pattern for this category
                    for pattern in patterns:
                        if pattern.upper() in motif_upper or motif_upper in pattern.upper():
                            count += motif_count
                            self.logger.info(f"Matched '{motif}' to category '{category}' (pattern: '{pattern}')")
                            break

                cm_data.append(count)
                if count > 0:
                    self.logger.info(f"CM category '{category}': {count} records")

            # Strategy 2: If no matches found, try exact category names
            if sum(cm_data) == 0:
                self.logger.info("No pattern matches found, trying exact category names...")
                for category in default_categories:
                    count = motif_counts.get(category, 0)
                    cm_data.append(count)
                    if count > 0:
                        self.logger.info(f"CM exact category '{category}': {count} records")

            # Strategy 3: If still no matches, use top 3 motifs
            if sum(cm_data) == 0 and motif_counts:
                self.logger.info("No category matches found, using top 3 motifs")
                sorted_motifs = sorted(motif_counts.items(), key=lambda x: x[1], reverse=True)

                cm_data = []
                for i in range(3):
                    if i < len(sorted_motifs):
                        motif, count = sorted_motifs[i]
                        cm_data.append(count)
                        self.logger.info(f"CM top motif {i+1} '{motif}': {count} records")
                    else:
                        cm_data.append(0)

            # Ensure we always have exactly 3 values
            while len(cm_data) < 3:
                cm_data.append(0)
            cm_data = cm_data[:3]

            self.logger.info(f"Final CM chart data: {cm_data}")

            # Log the mapping for debugging
            if sum(cm_data) > 0:
                self.logger.info("CM mapping successful:")
                for i, (category, count) in enumerate(zip(default_categories, cm_data)):
                    self.logger.info(f"  {category}: {count}")
            else:
                self.logger.warning("CM mapping resulted in all zeros - check motif names and date range")

            return cm_data

        except Exception as e:
            self.logger.error(f"Error mapping motifs to CM categories: {e}")
            # Return default values if mapping fails
            return [0, 0, 0]

    def _update_chart_data(self, script_content, chart_context, new_data):
        """Update Chart.js data array in the script content."""
        try:
            import re

            # Find the chart configuration for the specific context
            # Look for pattern like: const communesCtx = ... data: [56, 4]
            pattern = rf'(const {chart_context}.*?data:\s*\[)[^\]]*(\])'

            # Convert new_data to string format
            data_str = ', '.join(str(x) for x in new_data)

            # Replace the data array
            replacement = rf'\g<1>{data_str}\g<2>'
            updated_script = re.sub(pattern, replacement, script_content, flags=re.DOTALL)

            if updated_script != script_content:
                self.logger.info(f"Updated {chart_context} data: {new_data}")
            else:
                self.logger.warning(f"Could not find data array for {chart_context}")

            return updated_script

        except Exception as e:
            self.logger.error(f"Error updating chart data for {chart_context}: {e}")
            return script_content

    def _format_stats_as_cards(self, stats_html):
        """Format statistics to match the existing card-based layout of pres stats dashboard."""
        try:
            # Extract key statistics from the stats_data
            if not hasattr(self, 'filtered_statistics') or not self.filtered_statistics:
                return stats_html

            stats = self.filtered_statistics

            # Create cards that match the existing dashboard style
            cards_html = ""

            # Period Summary Card
            cards_html += f'''
            <!-- Pladria Statistics Card - Period Summary -->
            <div class="card">
                <h2>📊 Statistiques Pladria - Période Filtrée</h2>
                <div class="stats-summary">
                    <div class="stat-item">
                        <span class="stat-value" style="color: #3498db;">{stats['data_summary']['total_records']}</span>
                        <span class="stat-label">Total enregistrements</span>
                    </div>
                    <div class="stat-item">
                        <span class="stat-value" style="color: #27ae60;">{stats['period']['total_days']}</span>
                        <span class="stat-label">Jours analysés</span>
                    </div>
                    <div class="stat-item">
                        <span class="stat-value" style="color: #f39c12;">{stats['motifs'].get('total_unique', 0)}</span>
                        <span class="stat-label">Motifs uniques</span>
                    </div>
                </div>
                <div style="margin-top: 15px; padding: 10px; background: #f8f9fa; border-radius: 5px; font-size: 0.9em;">
                    <strong>Période:</strong> {stats['period']['start_date']} à {stats['period']['end_date']}
                </div>
            </div>'''

            # Top Motifs Card (if available)
            if stats['motifs'].get('sorted'):
                top_motifs = stats['motifs']['sorted'][:5]  # Top 5 motifs
                motifs_html = ""
                for i, (motif, count) in enumerate(top_motifs):
                    percentage = round((count / stats['data_summary']['total_records']) * 100, 1)
                    color = ['#e74c3c', '#f39c12', '#3498db', '#27ae60', '#9b59b6'][i % 5]
                    motifs_html += f'''
                    <div class="stat-item">
                        <span class="stat-value" style="color: {color};">{count}</span>
                        <span class="stat-label">{motif[:20]}... ({percentage}%)</span>
                    </div>'''

                cards_html += f'''
            <!-- Pladria Statistics Card - Top Motifs -->
            <div class="card">
                <h2>🎯 Top Motifs Pladria</h2>
                <div class="stats-summary">
                    {motifs_html}
                </div>
            </div>'''

            # Collaborateurs Card (if available)
            if stats['collaborateurs']:
                sorted_collabs = sorted(stats['collaborateurs'].items(), key=lambda x: x[1]['count'], reverse=True)[:4]
                collabs_html = ""
                for i, (collab, data) in enumerate(sorted_collabs):
                    percentage = round((data['count'] / stats['data_summary']['total_records']) * 100, 1)
                    color = ['#27ae60', '#3498db', '#f39c12', '#e74c3c'][i % 4]
                    collabs_html += f'''
                    <div class="stat-item">
                        <span class="stat-value" style="color: {color};">{data['count']}</span>
                        <span class="stat-label">{collab[:15]}... ({percentage}%)</span>
                    </div>'''

                cards_html += f'''
            <!-- Pladria Statistics Card - Collaborateurs -->
            <div class="card">
                <h2>👥 Top Collaborateurs Pladria</h2>
                <div class="stats-summary">
                    {collabs_html}
                </div>
            </div>'''

            # Processing Times Card (if available)
            if stats['processing_times'].get('count', 0) > 0:
                pt = stats['processing_times']
                cards_html += f'''
            <!-- Pladria Statistics Card - Processing Times -->
            <div class="card">
                <h2>⏱️ Temps de Traitement Pladria</h2>
                <div class="stats-summary">
                    <div class="stat-item">
                        <span class="stat-value" style="color: #3498db;">{pt['average']:.1f}</span>
                        <span class="stat-label">Temps moyen</span>
                    </div>
                    <div class="stat-item">
                        <span class="stat-value" style="color: #27ae60;">{pt['median']:.1f}</span>
                        <span class="stat-label">Temps médian</span>
                    </div>
                    <div class="stat-item">
                        <span class="stat-value" style="color: #f39c12;">{pt['count']}</span>
                        <span class="stat-label">Total traité</span>
                    </div>
                </div>
            </div>'''

            return cards_html

        except Exception as e:
            self.logger.error(f"Error formatting stats as cards: {e}")
            # Fallback to original stats_html
            return stats_html

    def _generate_html_statistics(self, stats_data):
        """Generate HTML content for statistics injection."""
        try:
            html = f"""
<div id="pladria-statistics" style="margin: 20px 0; padding: 20px; border: 2px solid #007acc; border-radius: 8px; background-color: #f8f9fa;">
    <h2 style="color: #007acc; margin-bottom: 15px;">📊 Statistiques Pladria - Période Filtrée</h2>

    <div style="margin-bottom: 20px;">
        <h3 style="color: #333; font-size: 16px;">📋 Résumé</h3>
        <table style="width: 100%; border-collapse: collapse; margin-bottom: 15px;">
            <tr style="background-color: #e9ecef;">
                <td style="padding: 8px; border: 1px solid #ddd; font-weight: bold;">Période</td>
                <td style="padding: 8px; border: 1px solid #ddd;">{stats_data['metadata']['period_start']} à {stats_data['metadata']['period_end']}</td>
            </tr>
            <tr>
                <td style="padding: 8px; border: 1px solid #ddd; font-weight: bold;">Durée</td>
                <td style="padding: 8px; border: 1px solid #ddd;">{stats_data['metadata']['total_days']} jour(s)</td>
            </tr>
            <tr style="background-color: #e9ecef;">
                <td style="padding: 8px; border: 1px solid #ddd; font-weight: bold;">Total enregistrements</td>
                <td style="padding: 8px; border: 1px solid #ddd;">{stats_data['summary']['total_records']}</td>
            </tr>
            <tr>
                <td style="padding: 8px; border: 1px solid #ddd; font-weight: bold;">Motifs uniques</td>
                <td style="padding: 8px; border: 1px solid #ddd;">{stats_data['summary']['unique_motifs']}</td>
            </tr>
            <tr style="background-color: #e9ecef;">
                <td style="padding: 8px; border: 1px solid #ddd; font-weight: bold;">Généré le</td>
                <td style="padding: 8px; border: 1px solid #ddd;">{stats_data['metadata']['generated_at']}</td>
            </tr>
        </table>
    </div>
"""

            # Top motifs section
            if stats_data['top_motifs']:
                html += """
    <div style="margin-bottom: 20px;">
        <h3 style="color: #333; font-size: 16px;">🎯 Top Motifs</h3>
        <table style="width: 100%; border-collapse: collapse;">
            <thead>
                <tr style="background-color: #007acc; color: white;">
                    <th style="padding: 8px; border: 1px solid #ddd; text-align: left;">Motif</th>
                    <th style="padding: 8px; border: 1px solid #ddd; text-align: center;">Nombre</th>
                    <th style="padding: 8px; border: 1px solid #ddd; text-align: center;">%</th>
                </tr>
            </thead>
            <tbody>
"""
                for i, motif_data in enumerate(stats_data['top_motifs'][:5]):
                    bg_color = "#f8f9fa" if i % 2 == 0 else "#ffffff"
                    html += f"""
                <tr style="background-color: {bg_color};">
                    <td style="padding: 8px; border: 1px solid #ddd;">{motif_data['motif']}</td>
                    <td style="padding: 8px; border: 1px solid #ddd; text-align: center;">{motif_data['count']}</td>
                    <td style="padding: 8px; border: 1px solid #ddd; text-align: center;">{motif_data['percentage']}%</td>
                </tr>
"""
                html += """
            </tbody>
        </table>
    </div>
"""

            # Collaborateurs section
            if stats_data['collaborateurs']:
                html += """
    <div style="margin-bottom: 20px;">
        <h3 style="color: #333; font-size: 16px;">👥 Top Collaborateurs</h3>
        <table style="width: 100%; border-collapse: collapse;">
            <thead>
                <tr style="background-color: #28a745; color: white;">
                    <th style="padding: 8px; border: 1px solid #ddd; text-align: left;">Collaborateur</th>
                    <th style="padding: 8px; border: 1px solid #ddd; text-align: center;">Nombre</th>
                    <th style="padding: 8px; border: 1px solid #ddd; text-align: center;">%</th>
                </tr>
            </thead>
            <tbody>
"""
                for i, collab_data in enumerate(stats_data['collaborateurs'][:5]):
                    bg_color = "#f8f9fa" if i % 2 == 0 else "#ffffff"
                    html += f"""
                <tr style="background-color: {bg_color};">
                    <td style="padding: 8px; border: 1px solid #ddd;">{collab_data['name']}</td>
                    <td style="padding: 8px; border: 1px solid #ddd; text-align: center;">{collab_data['count']}</td>
                    <td style="padding: 8px; border: 1px solid #ddd; text-align: center;">{collab_data['percentage']}%</td>
                </tr>
"""
                html += """
            </tbody>
        </table>
    </div>
"""

            # Processing times section
            if stats_data['processing_times']:
                pt = stats_data['processing_times']
                html += f"""
    <div style="margin-bottom: 20px;">
        <h3 style="color: #333; font-size: 16px;">⏱️ Temps de Traitement</h3>
        <div style="display: flex; gap: 15px; flex-wrap: wrap;">
            <div style="background-color: #e3f2fd; padding: 10px; border-radius: 5px; min-width: 120px;">
                <strong>Moyenne:</strong> {pt['average']}
            </div>
            <div style="background-color: #f3e5f5; padding: 10px; border-radius: 5px; min-width: 120px;">
                <strong>Médiane:</strong> {pt['median']}
            </div>
            <div style="background-color: #e8f5e8; padding: 10px; border-radius: 5px; min-width: 120px;">
                <strong>Min/Max:</strong> {pt['min']} / {pt['max']}
            </div>
        </div>
    </div>
"""

            # Daily stats section
            if stats_data['daily_stats']:
                html += """
    <div style="margin-bottom: 20px;">
        <h3 style="color: #333; font-size: 16px;">📈 Évolution Quotidienne (7 derniers jours)</h3>
        <table style="width: 100%; border-collapse: collapse;">
            <thead>
                <tr style="background-color: #ffc107; color: black;">
                    <th style="padding: 8px; border: 1px solid #ddd; text-align: left;">Date</th>
                    <th style="padding: 8px; border: 1px solid #ddd; text-align: center;">Nombre</th>
                    <th style="padding: 8px; border: 1px solid #ddd; text-align: left;">Graphique</th>
                </tr>
            </thead>
            <tbody>
"""
                max_count = max([day['count'] for day in stats_data['daily_stats']]) if stats_data['daily_stats'] else 1
                for i, day_data in enumerate(stats_data['daily_stats']):
                    bg_color = "#f8f9fa" if i % 2 == 0 else "#ffffff"
                    bar_width = int((day_data['count'] / max_count) * 100) if max_count > 0 else 0
                    html += f"""
                <tr style="background-color: {bg_color};">
                    <td style="padding: 8px; border: 1px solid #ddd;">{day_data['date']}</td>
                    <td style="padding: 8px; border: 1px solid #ddd; text-align: center;">{day_data['count']}</td>
                    <td style="padding: 8px; border: 1px solid #ddd;">
                        <div style="background-color: #007acc; height: 20px; width: {bar_width}%; border-radius: 3px;"></div>
                    </td>
                </tr>
"""
                html += """
            </tbody>
        </table>
    </div>
"""

            html += """
    <div style="margin-top: 20px; padding: 10px; background-color: #d4edda; border-radius: 5px; font-size: 12px; color: #155724;">
        <strong>Note:</strong> Ces statistiques ont été générées automatiquement par Pladria v3.0 - Module Team Statistics
    </div>
</div>
"""

            return html

        except Exception as e:
            self.logger.error(f"Error generating HTML statistics: {e}")
            return "<div>Erreur lors de la génération des statistiques HTML</div>"

    def _prepare_dashboard_data(self, statistics):
        """Prepare data for the dashboard view."""
        try:
            dashboard_data = {
                'summary': {
                    'period': f"{statistics['period']['start_date']} à {statistics['period']['end_date']}",
                    'total_days': statistics['period']['total_days'],
                    'total_records': statistics['data_summary']['total_records'],
                    'stats_files': statistics['data_summary']['stats_files']
                },
                'charts': {},
                'tables': {},
                'key_metrics': {}
            }

            # Prepare motifs chart data
            if statistics['motifs'].get('sorted'):
                dashboard_data['charts']['motifs'] = {
                    'type': 'bar',
                    'title': 'Répartition des motifs',
                    'data': statistics['motifs']['sorted'][:10]  # Top 10
                }

            # Prepare processing times metrics
            if statistics['processing_times'].get('count', 0) > 0:
                dashboard_data['key_metrics']['processing_times'] = {
                    'average': round(statistics['processing_times']['average'], 2),
                    'median': round(statistics['processing_times']['median'], 2),
                    'total_processed': statistics['processing_times']['count']
                }

            # Prepare collaborateurs table
            if statistics['collaborateurs']:
                collab_data = []
                for collab, data in statistics['collaborateurs'].items():
                    collab_data.append({
                        'collaborateur': collab,
                        'count': data['count'],
                        'percentage': round((data['count'] / statistics['data_summary']['total_records']) * 100, 1)
                    })
                dashboard_data['tables']['collaborateurs'] = sorted(collab_data, key=lambda x: x['count'], reverse=True)

            # Prepare communes table
            if statistics['communes']:
                commune_data = []
                for commune, data in statistics['communes'].items():
                    commune_data.append({
                        'commune': commune,
                        'count': data['count'],
                        'percentage': round((data['count'] / statistics['data_summary']['total_records']) * 100, 1)
                    })
                dashboard_data['tables']['communes'] = sorted(commune_data, key=lambda x: x['count'], reverse=True)

            # Prepare daily chart data
            if statistics['daily_stats']:
                daily_data = []
                for date_str, data in sorted(statistics['daily_stats'].items()):
                    daily_data.append((date_str, data['count']))
                dashboard_data['charts']['daily'] = {
                    'type': 'line',
                    'title': 'Évolution quotidienne',
                    'data': daily_data
                }

            return dashboard_data

        except Exception as e:
            self.logger.error(f"Error preparing dashboard data: {e}")
            return {}

    def _update_statistics_display_with_filtered_data(self):
        """Update the statistics display with filtered data."""
        try:
            if not hasattr(self, 'filtered_statistics') or not self.filtered_statistics:
                return

            # Update the statistics text area if it exists
            if hasattr(self, 'stats_text') and self.stats_text:
                # Clear existing content
                self.stats_text.delete(1.0, tk.END)

                # Add filtered statistics summary
                stats = self.filtered_statistics
                summary_text = f"""📊 STATISTIQUES FILTRÉES
{'='*50}

📅 Période: {stats['period']['start_date']} à {stats['period']['end_date']}
📈 Durée: {stats['period']['total_days']} jour(s)
📋 Total enregistrements: {stats['data_summary']['total_records']}
📁 Fichiers stats: {stats['data_summary']['stats_files']}

🎯 MOTIFS LES PLUS FRÉQUENTS:
"""

                # Add top motifs
                if stats['motifs'].get('sorted'):
                    for i, (motif, count) in enumerate(stats['motifs']['sorted'][:5], 1):
                        percentage = round((count / stats['data_summary']['total_records']) * 100, 1)
                        summary_text += f"{i}. {motif}: {count} ({percentage}%)\n"

                # Add processing times if available
                if stats['processing_times'].get('count', 0) > 0:
                    pt = stats['processing_times']
                    summary_text += f"""
⏱️ TEMPS DE TRAITEMENT:
   • Moyenne: {pt['average']:.2f}
   • Médiane: {pt['median']:.2f}
   • Min/Max: {pt['min']:.2f} / {pt['max']:.2f}
"""

                # Add collaborateurs summary
                if stats['collaborateurs']:
                    summary_text += f"\n👥 COLLABORATEURS ({len(stats['collaborateurs'])}):\n"
                    sorted_collabs = sorted(stats['collaborateurs'].items(), key=lambda x: x[1]['count'], reverse=True)
                    for collab, data in sorted_collabs[:5]:
                        percentage = round((data['count'] / stats['data_summary']['total_records']) * 100, 1)
                        summary_text += f"   • {collab}: {data['count']} ({percentage}%)\n"

                # Add communes summary
                if stats['communes']:
                    summary_text += f"\n🏘️ COMMUNES ({len(stats['communes'])}):\n"
                    sorted_communes = sorted(stats['communes'].items(), key=lambda x: x[1]['count'], reverse=True)
                    for commune, data in sorted_communes[:5]:
                        percentage = round((data['count'] / stats['data_summary']['total_records']) * 100, 1)
                        summary_text += f"   • {commune}: {data['count']} ({percentage}%)\n"

                # Insert the text
                self.stats_text.insert(1.0, summary_text)

            self.logger.info("Statistics display updated with filtered data")

        except Exception as e:
            self.logger.error(f"Error updating statistics display: {e}")









    def _create_archive_zip(self, treated_communes: list) -> str:
        """Create ZIP archive with complete folders of treated communes."""
        try:
            import zipfile
            from datetime import datetime

            # Create backup directory on D:\ drive
            backup_dir = "D:\\BackUP Plan Adressage"
            os.makedirs(backup_dir, exist_ok=True)

            # Generate archive filename using selected month
            selected_month = self.archive_month_var.get() if hasattr(self, 'archive_month_var') else None
            if not selected_month:
                current_date = datetime.now()
                month_name = current_date.strftime("%B")  # Full month name
            else:
                month_name = selected_month
            commune_count = len(treated_communes)

            archive_filename = f"{commune_count}.CommunesTraités_{month_name}.zip"
            archive_path = os.path.join(backup_dir, archive_filename)

            self.logger.info(f"Creating archive: {archive_filename}")
            self.logger.info(f"Archive will contain {commune_count} commune folders")

            # Debug: Log commune information
            self.logger.info("📋 COMMUNE DETAILS:")
            for i, commune in enumerate(treated_communes):
                commune_name = commune['nom']
                id_tache = commune.get('id_tache', 'N/A')
                folder_path = commune.get('folder_path', 'NOT_SET')
                folder_exists = os.path.exists(folder_path) if folder_path else False

                self.logger.info(f"  {i+1}. {commune_name} (ID: {id_tache})")
                self.logger.info(f"     Folder: {folder_path}")
                self.logger.info(f"     Exists: {folder_exists}")

            # Calculate total size for progress tracking
            total_size = sum(commune.get('folder_size', 0) for commune in treated_communes)
            processed_size = 0

            # Create ZIP file with standard compression
            with zipfile.ZipFile(archive_path, 'w', zipfile.ZIP_DEFLATED, compresslevel=6) as zipf:

                # Add each commune's complete folder without any filtering
                folders_added = 0
                for i, commune in enumerate(treated_communes):
                    commune_name = commune['nom']
                    id_tache = commune.get('id_tache', 'N/A')
                    folder_path = commune.get('folder_path')

                    self.logger.info(f"📁 Processing commune {i+1}/{commune_count}: {commune_name} (ID: {id_tache})")

                    if not folder_path:
                        self.logger.warning(f"❌ No folder path set for commune {commune_name}")
                        continue

                    if not os.path.exists(folder_path):
                        self.logger.warning(f"❌ Folder does not exist for commune {commune_name}: {folder_path}")
                        continue

                    if not os.path.isdir(folder_path):
                        self.logger.warning(f"❌ Path is not a directory for commune {commune_name}: {folder_path}")
                        continue

                    self.logger.info(f"📂 Source folder: {folder_path}")
                    self.logger.info(f"📂 Folder is accessible: {os.access(folder_path, os.R_OK)}")

                    try:
                        # Add complete folder recursively - copy EVERYTHING
                        added_files = self._add_folder_to_zip(zipf, folder_path, commune_name)

                        # Add separate commune files if any
                        separate_files = commune.get('separate_files', [])
                        if separate_files:
                            self.logger.info(f"📄 Adding {len(separate_files)} separate files for {commune_name}")
                            for separate_file in separate_files:
                                try:
                                    # Add separate file to commune folder in archive
                                    file_name = os.path.basename(separate_file)
                                    archive_path = f"{commune_name}/Fichiers_Separes/{file_name}"
                                    zipf.write(separate_file, archive_path)
                                    added_files += 1
                                    self.logger.debug(f"Added separate file: {archive_path}")
                                except Exception as e:
                                    self.logger.warning(f"Could not add separate file {separate_file}: {e}")

                        # Update progress
                        processed_size += commune.get('folder_size', 0)
                        progress = (processed_size / total_size * 100) if total_size > 0 else 0

                        total_files_msg = f"{added_files} fichiers"
                        if separate_files:
                            total_files_msg += f" (+ {len(separate_files)} séparés)"

                        self.logger.info(f"✅ Successfully added {total_files_msg} for {commune_name} "
                                       f"(Progress: {progress:.1f}%)")

                        folders_added += 1

                        # Update status label if it exists
                        if hasattr(self, 'archive_status_label') and self.archive_status_label:
                            try:
                                self.archive_status_label.config(
                                    text=f"Archivage: {commune_name} ({i+1}/{commune_count}) - {total_files_msg}"
                                )
                                self.archive_status_label.update()
                            except:
                                pass  # Ignore UI update errors

                    except Exception as e:
                        self.logger.error(f"❌ Error adding folder for {commune_name}: {e}")
                        # Continue with next commune even if this one fails
                        continue

                # Log summary of folders added
                self.logger.info(f"📊 ARCHIVE SUMMARY: {folders_added}/{commune_count} commune folders added successfully")

                # Add global tracking file for treated communes
                self.logger.info("📊 Adding global tracking file...")
                self._add_treated_communes_tracking_to_zip(zipf, treated_communes)

            # Get final archive size
            archive_size = os.path.getsize(archive_path)
            compression_ratio = (1 - archive_size / total_size) * 100 if total_size > 0 else 0

            self.logger.info(f"✅ Archive created successfully: {archive_path}")
            self.logger.info(f"📊 Archive size: {archive_size/1024/1024:.1f} MB "
                           f"(Compression: {compression_ratio:.1f}%)")

            return archive_path

        except Exception as e:
            self.logger.error(f"❌ Error creating archive ZIP: {e}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            return None

    def _add_folder_to_zip(self, zipf, folder_path: str, commune_name: str) -> int:
        """Add a complete folder and ALL its contents to the ZIP archive without filtering."""
        try:
            added_files = 0
            error_files = 0

            self.logger.info(f"📁 Adding complete folder for {commune_name}: {folder_path}")

            # Walk through ALL files and subdirectories without exception
            for root, dirs, files in os.walk(folder_path):
                # Calculate relative path from the commune folder
                rel_path = os.path.relpath(root, folder_path)

                # Create the archive path structure
                if rel_path == '.':
                    # Root of commune folder
                    archive_dir = commune_name
                else:
                    # Subdirectory
                    archive_dir = f"{commune_name}/{rel_path.replace(os.sep, '/')}"

                # Add ALL files in current directory without any filtering
                for file in files:
                    file_path = os.path.join(root, file)

                    try:
                        # Create archive path
                        archive_file_path = f"{archive_dir}/{file}"

                        # Add file to ZIP - copy everything without analysis
                        zipf.write(file_path, archive_file_path)
                        added_files += 1

                        # Log only every 50 files to avoid spam
                        if added_files % 50 == 0:
                            self.logger.debug(f"Added {added_files} files so far...")

                    except (OSError, IOError, PermissionError) as e:
                        error_files += 1
                        self.logger.warning(f"Could not access file {file_path}: {e}")
                        continue
                    except Exception as e:
                        error_files += 1
                        self.logger.warning(f"Error adding file {file_path}: {e}")
                        continue

                # Add empty directories to preserve complete structure
                if not files and not dirs:
                    try:
                        zipf.writestr(f"{archive_dir}/", "")
                        self.logger.debug(f"Added empty directory: {archive_dir}/")
                    except Exception as e:
                        self.logger.debug(f"Could not add empty directory {archive_dir}: {e}")

            # Log final summary
            self.logger.info(f"✅ Folder {commune_name}: {added_files} files added, {error_files} errors")

            return added_files

        except Exception as e:
            self.logger.error(f"Error adding folder to ZIP: {e}")
            return 0

    def _is_file_accessible(self, file_path: str) -> bool:
        """Check if a file is accessible for copying (minimal check)."""
        try:
            # Simple existence and readability check
            return os.path.isfile(file_path) and os.access(file_path, os.R_OK)
        except Exception:
            return False

    def _add_treated_communes_tracking_to_zip(self, zipf, treated_communes: list):
        """Add a copy of the global tracking file to the ZIP archive."""
        try:
            from datetime import datetime
            from config.constants import TeamsConfig

            # CORRECTION: Simplement copier le fichier de suivi global existant
            # Chemin vers le fichier de suivi global
            global_teams_path = TeamsConfig.get_global_teams_path()
            global_file_path = os.path.join(global_teams_path, self.global_excel_filename)

            self.logger.info(f"Looking for global tracking file: {global_file_path}")

            # Vérifier que le fichier existe
            if not os.path.exists(global_file_path):
                self.logger.warning(f"Global tracking file not found: {global_file_path}")
                return

            # Vérifier que le fichier est accessible
            if not os.access(global_file_path, os.R_OK):
                self.logger.warning(f"Global tracking file not readable: {global_file_path}")
                return

            # Générer le nom du fichier dans l'archive
            current_date = datetime.now()
            month_name = current_date.strftime("%B")
            tracking_filename = f"Suivi_Global_CommunesTraités_{month_name}.xlsx"

            # Copier directement le fichier dans l'archive
            zipf.write(global_file_path, tracking_filename)

            # Obtenir la taille du fichier pour les logs
            file_size = os.path.getsize(global_file_path)

            self.logger.info(f"✅ Added complete global tracking file to archive:")
            self.logger.info(f"   Source: {global_file_path}")
            self.logger.info(f"   Archive name: {tracking_filename}")
            self.logger.info(f"   File size: {file_size/1024/1024:.1f} MB")

        except Exception as e:
            self.logger.error(f"Error adding global tracking file to ZIP: {e}")
            # Don't raise - this is optional, archive can continue without it
