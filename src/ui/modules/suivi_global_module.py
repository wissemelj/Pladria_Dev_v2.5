"""
Suivi Global Tickets module - Aggregate commune suivi files into global Excel.
"""

import tkinter as tk
from tkinter import messagebox, filedialog, ttk
import logging
import os
from typing import Optional, List, Dict, Any
from pathlib import Path
import sys
from datetime import datetime

# Ensure src directory is in path
src_path = Path(__file__).parent.parent.parent
if str(src_path) not in sys.path:
    sys.path.insert(0, str(src_path))

from config.constants import COLORS, UIConfig
from core import FileProcessor, DataValidator, ExcelGenerator
from utils.file_utils import get_icon_path, check_file_access, is_excel_file_open
from utils.lazy_imports import get_pandas
from utils.performance import run_async_task

from ui.styles import StyleManager, create_card_frame, create_section_header
from ui.responsive_utils import get_responsive_manager
from ui.components.scrollable_frame import create_scrollable_container
from ui.keyboard_shortcuts import KeyboardShortcutManager

logger = logging.getLogger(__name__)


class SuiviGlobalModule:
    """Suivi Global Tickets module for aggregating commune suivi files."""
    
    def __init__(self, parent: tk.Widget, navigation_manager=None):
        """
        Initialize the Suivi Global module.

        Args:
            parent: Parent widget
            navigation_manager: Navigation manager instance
        """
        self.parent = parent
        self.navigation_manager = navigation_manager
        self.logger = logging.getLogger(__name__)

        # Initialize responsive manager
        self.responsive_manager = get_responsive_manager()

        # Initialize core components
        try:
            self.file_processor = FileProcessor()
            self.data_validator = DataValidator()
            self.excel_generator = ExcelGenerator()
            self.logger.info("Core components initialized")
        except Exception as e:
            self.logger.error(f"Error initializing core components: {e}")
            self.file_processor = None
            self.data_validator = None
            self.excel_generator = None

        # Module data
        self.commune_folders = []
        self.processed_data = []
        self.existing_communes = {}  # Store existing commune data for comparison
        self.new_communes = []  # Track new communes found during scan
        self.updated_communes = []  # Track communes that will be updated
        # Get dynamic Teams path for current user
        from config.constants import TeamsConfig
        self.teams_folder_path = TeamsConfig.get_global_teams_path()
        self.global_excel_filename = "Suivis Global Tickets CMS Adr_PA.xlsx"

        # UI components
        self.progress_var = None
        self.progress_bar = None
        self.status_label = None
        self.generate_button = None
        self.summary_text = None
        self.scan_button = None

        # Keyboard shortcuts (optional)
        self.keyboard_manager = None

        # Create UI first
        self._create_module_ui()

        # Initialize optional features after UI is created
        self.parent.after(100, self._initialize_optional_features)

        # Start automatic scanning after UI is fully loaded
        self.parent.after(500, self._start_automatic_scan)

    def _initialize_optional_features(self):
        """Initialize optional features after UI is ready."""
        try:
            # Setup keyboard shortcuts
            if self.navigation_manager and hasattr(self.navigation_manager, 'root'):
                self.keyboard_manager = KeyboardShortcutManager(self.navigation_manager.root)
                self._setup_module_shortcuts()

            self.logger.info("Optional features initialized successfully")

        except Exception as e:
            self.logger.error(f"Error initializing optional features: {e}")

    def _start_automatic_scan(self):
        """Start automatic scanning when the module opens."""
        try:
            # Only start automatic scan if no scan is already in progress
            if hasattr(self, 'scan_button') and self.scan_button and self.scan_button.cget('state') == 'normal':
                self.logger.info("Starting automatic scan on module load")
                self._scan_and_process_folders()
        except Exception as e:
            self.logger.error(f"Error starting automatic scan: {e}")

    def _create_module_ui(self):
        """Create the module user interface."""
        try:
            # Skip title bar to eliminate blank space - go directly to content
            # self._create_title_bar()  # Commented out to remove blank space

            # Main content area
            self._create_content_area()

            self.logger.info("Module UI created successfully")

        except Exception as e:
            self.logger.error(f"Error creating module UI: {e}")
            self._create_error_ui(str(e))

    def _create_error_ui(self, error_message: str):
        """Create a simple error display when module fails to load."""
        error_frame = tk.Frame(self.parent, bg=COLORS['BG'])
        error_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        # Error icon and message
        error_label = tk.Label(
            error_frame,
            text=f"❌ Erreur lors du chargement du module:\n\n{error_message}",
            font=UIConfig.FONT_SUBTITLE,
            fg=COLORS['DANGER'],
            bg=COLORS['BG'],
            justify=tk.CENTER
        )
        error_label.pack(expand=True)

        # Retry button
        retry_btn = tk.Button(
            error_frame,
            text="🔄 Réessayer",
            command=self._retry_initialization,
            bg=COLORS['PRIMARY'],
            fg='white',
            font=UIConfig.FONT_BUTTON,
            relief='flat',
            padx=20,
            pady=10
        )
        retry_btn.pack(pady=10)

    def _retry_initialization(self):
        """Retry module initialization."""
        try:
            # Clear the parent
            for widget in self.parent.winfo_children():
                widget.destroy()

            # Reinitialize
            self.__init__(self.parent, self.navigation_manager)

        except Exception as e:
            self.logger.error(f"Retry failed: {e}")
            messagebox.showerror("Erreur", f"Impossible de réinitialiser le module:\n{e}")

    def _create_title_bar(self):
        """Create the module title bar - Ultra compact."""
        title_frame = tk.Frame(self.parent, bg=COLORS['CARD'], height=35)  # Reduced from 60
        title_frame.pack(fill=tk.X, padx=3, pady=(3, 0))  # Ultra minimal padding
        title_frame.pack_propagate(False)

        # Title content - Ultra compact
        title_content = tk.Frame(title_frame, bg=COLORS['CARD'])
        title_content.pack(fill=tk.BOTH, expand=True, padx=6, pady=3)  # Ultra minimal padding
        
        # Module icon and title
        title_left = tk.Frame(title_content, bg=COLORS['CARD'])
        title_left.pack(side=tk.LEFT, fill=tk.Y)
        
        # Icon - Ultra compact
        icon_label = tk.Label(
            title_left,
            text="🌐",
            font=("Segoe UI", 12),  # Reduced from 20
            fg=COLORS['PRIMARY'],
            bg=COLORS['CARD']
        )
        icon_label.pack(side=tk.LEFT, padx=(0, 4))  # Ultra minimal spacing
        
        # Title and description
        title_text_frame = tk.Frame(title_left, bg=COLORS['CARD'])
        title_text_frame.pack(side=tk.LEFT, fill=tk.Y)
        
        title_label = tk.Label(
            title_text_frame,
            text="Suivi Global Tickets",
            font=UIConfig.FONT_HEADER,
            fg=COLORS['PRIMARY'],
            bg=COLORS['CARD']
        )
        title_label.pack(anchor=tk.W)
        
        subtitle_label = tk.Label(
            title_text_frame,
            text="Agrégation automatique des suivis de communes",
            font=UIConfig.FONT_SMALL,
            fg=COLORS['INFO'],
            bg=COLORS['CARD']
        )
        subtitle_label.pack(anchor=tk.W)
        
        # Action buttons on the right
        title_right = tk.Frame(title_content, bg=COLORS['CARD'])
        title_right.pack(side=tk.RIGHT, fill=tk.Y)

        # Reset button
        reset_btn = ttk.Button(
            title_right,
            text="🔄 Réinitialiser",
            command=self._reset_module,
            style='CompactWarning.TButton'
        )
        reset_btn.pack(side=tk.RIGHT, padx=(2, 0))  # Ultra minimal spacing

    def _create_content_area(self):
        """Create the main content area with responsive layout and scrolling."""
        # Get responsive configuration with corporate-optimized spacing
        column_config = self.responsive_manager.get_responsive_column_config()
        corporate_spacing = self.responsive_manager.get_corporate_spacing()
        padding = corporate_spacing['section_margin']

        # Create scrollable container for the entire content
        scrollable_container = create_scrollable_container(
            self.parent,
            self.responsive_manager,
            bg=COLORS['BG']
        )
        scrollable_container.pack(fill=tk.BOTH, expand=True, padx=1, pady=1)  # Ultra minimal padding

        # Get the scrollable frame to add content to
        content_container = scrollable_container.get_scrollable_frame()
        content_container.configure(bg=COLORS['BG'])

        # Configure grid layout for responsive design - Ultra compact
        if column_config['columns'] <= 2:
            # Single or two column layout for smaller screens - Ultra compact
            content_container.grid_columnconfigure(0, weight=1, minsize=250)  # Reduced minsize
            if column_config['columns'] == 2:
                content_container.grid_columnconfigure(1, weight=1, minsize=250)  # Reduced minsize
        else:
            # Two column layout for larger screens - Ultra compact
            content_container.grid_columnconfigure(0, weight=1, minsize=300)  # Reduced minsize
            content_container.grid_columnconfigure(1, weight=1, minsize=300)  # Reduced minsize
        content_container.grid_rowconfigure(0, weight=0)  # Don't expand row, use natural height

        # Create columns based on responsive configuration
        if column_config['columns'] == 1:
            # Single column layout - stack everything vertically with no padding
            main_column = tk.Frame(content_container, bg=COLORS['BG'])
            main_column.grid(row=0, column=0, sticky="new", padx=0, pady=0)  # Stick to top, no padding

            # Create sections in single column
            self._create_enhanced_scanning_section(main_column)
            self._create_enhanced_generation_section(main_column)

        else:
            # Two column layout - Ultra compact, stick to top
            left_column = tk.Frame(content_container, bg=COLORS['BG'])
            left_column.grid(row=0, column=0, sticky="new", padx=(0, 1), pady=0)  # Stick to top

            right_column = tk.Frame(content_container, bg=COLORS['BG'])
            right_column.grid(row=0, column=1, sticky="new", padx=(1, 0), pady=0)  # Stick to top

            # Create enhanced sections in two columns
            self._create_enhanced_scanning_section(left_column)
            self._create_enhanced_generation_section(right_column)

    def _setup_module_shortcuts(self):
        """Set up keyboard shortcuts specific to this module."""
        self.keyboard_manager.set_callback("Control-s", self._scan_shortcut)
        self.keyboard_manager.set_callback("Control-g", self._generate_shortcut)
        self.keyboard_manager.set_callback("F5", self._refresh_shortcut)



    def _reset_module(self):
        """Reset the module to initial state."""
        try:
            self.commune_folders.clear()
            self.processed_data.clear()
            self.existing_communes.clear()
            self.new_communes.clear()
            self.updated_communes.clear()

            if self.summary_text:
                self.summary_text.delete(1.0, tk.END)

            if self.status_label:
                self.status_label.config(text="Prêt à scanner les dossiers")

            if self.progress_var:
                self.progress_var.set(0)

            if hasattr(self, 'progress_percentage_label'):
                self.progress_percentage_label.config(text="0%")

            if self.generate_button:
                self.generate_button.config(state=tk.DISABLED)

            if self.scan_button:
                self.scan_button.config(
                    state=tk.NORMAL,
                    text="🚀 Lancer le Scan Complet",
                    bg=COLORS['SUCCESS']
                )

            if hasattr(self, 'update_button'):
                self.update_button.config(state=tk.NORMAL)

            # Reset status indicators
            if hasattr(self, 'scan_status_indicator'):
                self.scan_status_indicator.config(text="⚪", fg=COLORS['INFO'])

            if hasattr(self, 'status_icon_label'):
                self.status_icon_label.config(text="⏳", fg=COLORS['INFO'])

            # Clear statistics display
            if hasattr(self, 'stats_label'):
                self.stats_label.config(text="")

            self.logger.info("Module reset successfully")

        except Exception as e:
            self.logger.error(f"Error resetting module: {e}")
            messagebox.showerror("Erreur", f"Erreur lors de la réinitialisation:\n{e}")

    def _create_enhanced_scanning_section(self, parent: tk.Widget):
        """Create the enhanced folder scanning section with integrated progress tracking."""
        # Main scanning card with gradient-like effect - Compact, don't expand
        scan_card = create_card_frame(parent)
        scan_card.pack(fill=tk.X, expand=False, pady=(0, 5))  # Don't expand vertically

        # Enhanced section header with status indicator - Ultra compact
        header_frame = tk.Frame(scan_card, bg=COLORS['CARD'])
        header_frame.pack(fill=tk.X, padx=6, pady=(6, 3))  # Ultra minimal padding

        # Header left side - icon and title
        header_left = tk.Frame(header_frame, bg=COLORS['CARD'])
        header_left.pack(side=tk.LEFT, fill=tk.Y)

        # Animated scan icon - Ultra compact
        self.scan_icon_label = tk.Label(
            header_left,
            text="🔍",
            font=("Segoe UI", 10),  # Reduced from 16
            fg=COLORS['PRIMARY'],
            bg=COLORS['CARD']
        )
        self.scan_icon_label.pack(side=tk.LEFT, padx=(0, 4))  # Ultra minimal spacing

        # Title and subtitle
        title_frame = tk.Frame(header_left, bg=COLORS['CARD'])
        title_frame.pack(side=tk.LEFT, fill=tk.Y)

        title_label = tk.Label(
            title_frame,
            text="Scan Automatique",
            font=UIConfig.FONT_HEADER,
            fg=COLORS['PRIMARY'],
            bg=COLORS['CARD']
        )
        title_label.pack(anchor=tk.W)

        subtitle_label = tk.Label(
            title_frame,
            text="Détection et traitement des communes",
            font=UIConfig.FONT_SMALL,
            fg=COLORS['INFO'],
            bg=COLORS['CARD']
        )
        subtitle_label.pack(anchor=tk.W)

        # Status indicator on the right
        self.scan_status_indicator = tk.Label(
            header_frame,
            text="⚪",
            font=("Segoe UI", 12),
            fg=COLORS['INFO'],
            bg=COLORS['CARD']
        )
        self.scan_status_indicator.pack(side=tk.RIGHT)

        # Content frame with better spacing - Ultra compact, don't expand
        content_frame = tk.Frame(scan_card, bg=COLORS['CARD'])
        content_frame.pack(fill=tk.X, expand=False, padx=6, pady=(0, 6))  # Don't expand vertically

        # Enhanced description with features list - Ultra compact
        desc_frame = tk.Frame(content_frame, bg=COLORS['CARD'])
        desc_frame.pack(fill=tk.X, pady=(0, 5))  # Ultra minimal spacing

        desc_label = tk.Label(
            desc_frame,
            text="Analyse automatique du dossier Teams pour identifier et traiter les fichiers de suivi:",
            font=UIConfig.FONT_SMALL,
            fg=COLORS['TEXT_PRIMARY'],
            bg=COLORS['CARD'],
            wraplength=350,
            justify=tk.LEFT
        )
        desc_label.pack(anchor=tk.W, pady=(0, 3))  # Ultra minimal spacing

        # Features list
        features = [
            "📁 Détection automatique des dossiers communes",
            "📊 Extraction des données de suivi",
            "🔄 Comparaison avec les données existantes",
            "✅ Identification des nouvelles communes"
        ]

        for feature in features:
            feature_label = tk.Label(
                desc_frame,
                text=feature,
                font=("Segoe UI", 9),
                fg=COLORS['INFO'],
                bg=COLORS['CARD']
            )
            feature_label.pack(anchor=tk.W, pady=1)

        # Enhanced scan button with better styling - Ultra compact
        button_frame = tk.Frame(content_frame, bg=COLORS['CARD'])
        button_frame.pack(fill=tk.X, pady=(5, 0))  # Ultra minimal spacing

        self.scan_button = tk.Button(
            button_frame,
            text="🚀 Lancer le Scan Complet",
            command=self._scan_and_process_folders,
            bg=COLORS['SUCCESS'],
            fg='white',
            font=("Segoe UI", 9, "bold"),
            relief='flat',
            padx=15,
            pady=8,
            cursor='hand2'
        )
        self.scan_button.pack(side=tk.LEFT)

        # Add update button next to scan button
        self.update_button = tk.Button(
            button_frame,
            text="🔄 Actualiser",
            command=self._refresh_shortcut,
            bg=COLORS['BORDER'],
            fg=COLORS['TEXT_PRIMARY'],
            font=("Segoe UI", 9),
            relief='flat',
            padx=12,
            pady=8,
            cursor='hand2'
        )
        self.update_button.pack(side=tk.LEFT, padx=(8, 0))

        # INTEGRATED PROGRESS SECTION
        # Progress section header
        progress_header_frame = tk.Frame(content_frame, bg=COLORS['CARD'])
        progress_header_frame.pack(fill=tk.X, pady=(20, 10))

        # Progress icon and title
        progress_icon = tk.Label(
            progress_header_frame,
            text="📊",
            font=("Segoe UI", 12),
            fg=COLORS['PRIMARY'],
            bg=COLORS['CARD']
        )
        progress_icon.pack(side=tk.LEFT, padx=(0, 8))

        progress_title = tk.Label(
            progress_header_frame,
            text="Progression",
            font=("Segoe UI", 10, "bold"),
            fg=COLORS['PRIMARY'],
            bg=COLORS['CARD']
        )
        progress_title.pack(side=tk.LEFT)

        # Progress percentage on the right
        self.progress_percentage_label = tk.Label(
            progress_header_frame,
            text="0%",
            font=("Segoe UI", 10, "bold"),
            fg=COLORS['PRIMARY'],
            bg=COLORS['CARD']
        )
        self.progress_percentage_label.pack(side=tk.RIGHT)

        # Progress bar container
        progress_container = tk.Frame(content_frame, bg=COLORS['CARD'])
        progress_container.pack(fill=tk.X, pady=(0, 10))

        # Initialize progress variable
        self.progress_var = tk.DoubleVar()

        # Create styled progress bar
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("Custom.Horizontal.TProgressbar",
                       background=COLORS['SUCCESS'],
                       troughcolor=COLORS['BORDER'],
                       borderwidth=0,
                       lightcolor=COLORS['SUCCESS'],
                       darkcolor=COLORS['SUCCESS'])

        self.progress_bar = ttk.Progressbar(
            progress_container,
            variable=self.progress_var,
            maximum=100,
            length=350,
            mode='determinate',
            style="Custom.Horizontal.TProgressbar"
        )
        self.progress_bar.pack(fill=tk.X)

        # Status display
        status_container = tk.Frame(content_frame, bg=COLORS['CARD'])
        status_container.pack(fill=tk.X, pady=(0, 15))

        self.status_icon_label = tk.Label(
            status_container,
            text="⏳",
            font=("Segoe UI", 9),
            fg=COLORS['INFO'],
            bg=COLORS['CARD']
        )
        self.status_icon_label.pack(side=tk.LEFT, padx=(0, 8))

        self.status_label = tk.Label(
            status_container,
            text="Prêt à scanner les dossiers",
            font=UIConfig.FONT_SMALL,
            fg=COLORS['TEXT_PRIMARY'],
            bg=COLORS['CARD']
        )
        self.status_label.pack(side=tk.LEFT)

    def _create_enhanced_progress_section(self, parent: tk.Widget):
        """Create the enhanced progress monitoring section - DEPRECATED: Now integrated into scanning section."""
        # This method is no longer used as progress components are now integrated
        # into the _create_enhanced_scanning_section method for better UX
        pass

    def _create_enhanced_generation_section(self, parent: tk.Widget):
        """Create the enhanced Excel generation section."""
        # Generation card with modern styling - Expand to take remaining space
        generation_card = create_card_frame(parent)
        generation_card.pack(fill=tk.BOTH, expand=True)  # This should take all remaining space

        # Enhanced section header with dynamic statistics - Ultra compact
        self.generation_header_frame = tk.Frame(generation_card, bg=COLORS['CARD'])
        self.generation_header_frame.pack(fill=tk.X, padx=6, pady=(6, 3))  # Ultra minimal padding

        # Header left side
        header_left = tk.Frame(self.generation_header_frame, bg=COLORS['CARD'])
        header_left.pack(side=tk.LEFT, fill=tk.Y)

        # Animated generation icon - Ultra compact
        self.generation_icon_label = tk.Label(
            header_left,
            text="📈",
            font=("Segoe UI", 10),  # Reduced from 16
            fg=COLORS['PRIMARY'],
            bg=COLORS['CARD']
        )
        self.generation_icon_label.pack(side=tk.LEFT, padx=(0, 12))

        title_frame = tk.Frame(header_left, bg=COLORS['CARD'])
        title_frame.pack(side=tk.LEFT, fill=tk.Y)

        self.generation_header_label = tk.Label(
            title_frame,
            text="Génération Excel",
            font=UIConfig.FONT_HEADER,
            fg=COLORS['PRIMARY'],
            bg=COLORS['CARD']
        )
        self.generation_header_label.pack(anchor=tk.W)

        subtitle_label = tk.Label(
            title_frame,
            text="Fichier global consolidé",
            font=UIConfig.FONT_SMALL,
            fg=COLORS['INFO'],
            bg=COLORS['CARD']
        )
        subtitle_label.pack(anchor=tk.W)

        # Enhanced statistics display on the right
        stats_container = tk.Frame(self.generation_header_frame, bg=COLORS['CARD'])
        stats_container.pack(side=tk.RIGHT)

        self.stats_label = tk.Label(
            stats_container,
            text="",
            font=("Segoe UI", 9, "bold"),
            fg=COLORS['SUCCESS'],
            bg=COLORS['CARD'],
            justify=tk.RIGHT
        )
        self.stats_label.pack()

        # Content frame with better spacing - Expand to fill available space
        content_frame = tk.Frame(generation_card, bg=COLORS['CARD'])
        content_frame.pack(fill=tk.BOTH, expand=True, padx=6, pady=(0, 6))  # Ultra minimal padding

        # Enhanced summary section with tabs-like appearance
        summary_header_frame = tk.Frame(content_frame, bg=COLORS['CARD'])
        summary_header_frame.pack(fill=tk.X, pady=(0, 10))

        summary_icon = tk.Label(
            summary_header_frame,
            text="📋",
            font=("Segoe UI", 10),
            fg=COLORS['PRIMARY'],
            bg=COLORS['CARD']
        )
        summary_icon.pack(side=tk.LEFT, padx=(0, 8))

        summary_label = tk.Label(
            summary_header_frame,
            text="Résumé Détaillé des Données",
            font=("Segoe UI", 10, "bold"),
            fg=COLORS['TEXT_PRIMARY'],
            bg=COLORS['CARD']
        )
        summary_label.pack(side=tk.LEFT)

        # Toggle button for detailed view
        self.toggle_details_btn = tk.Button(
            summary_header_frame,
            text="📊 Vue Détaillée",
            command=self._toggle_detailed_view,
            bg=COLORS['BORDER'],
            fg=COLORS['TEXT_SECONDARY'],
            font=("Segoe UI", 8),
            relief='flat',
            padx=8,
            pady=2
        )
        self.toggle_details_btn.pack(side=tk.RIGHT)

        # Enhanced text widget with better styling - Expand to fill all available space
        text_container = tk.Frame(content_frame, bg=COLORS['CARD'], relief=tk.SOLID, bd=1)
        text_container.pack(fill=tk.BOTH, expand=True, pady=(0, 3))  # Take all remaining space

        # Text frame with scrollbar
        text_frame = tk.Frame(text_container, bg=COLORS['WHITE'])
        text_frame.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)

        text_scrollbar = tk.Scrollbar(text_frame, bg=COLORS['BORDER'])
        text_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.summary_text = tk.Text(
            text_frame,
            yscrollcommand=text_scrollbar.set,
            font=("Consolas", 8),  # Slightly smaller font for more content
            bg=COLORS['WHITE'],
            fg=COLORS['TEXT_PRIMARY'],
            # Remove fixed height to allow expansion
            wrap=tk.WORD,
            state=tk.DISABLED,
            relief='flat',
            padx=6,  # Reduced padding
            pady=4,  # Reduced padding
            selectbackground=COLORS['PRIMARY'],
            selectforeground='white'
        )
        self.summary_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        text_scrollbar.config(command=self.summary_text.yview)

        # Configure text tags for better formatting - Smaller fonts for more content
        self.summary_text.tag_configure("header", font=("Consolas", 8, "bold"), foreground=COLORS['PRIMARY'])
        self.summary_text.tag_configure("new_commune", font=("Consolas", 8), foreground=COLORS['SUCCESS'])
        self.summary_text.tag_configure("update_commune", font=("Consolas", 8), foreground=COLORS['WARNING'])
        self.summary_text.tag_configure("info", font=("Consolas", 7), foreground=COLORS['INFO'])

        # Enhanced file status indicator with better visualization - Minimal space
        status_container = tk.Frame(content_frame, bg=COLORS['LIGHT'], relief=tk.SOLID, bd=1)
        status_container.pack(fill=tk.X, pady=(0, 3))  # Minimal spacing to give more room to text

        # Status header
        status_header = tk.Frame(status_container, bg=COLORS['LIGHT'])
        status_header.pack(fill=tk.X, padx=10, pady=(8, 5))

        status_title = tk.Label(
            status_header,
            text="📁 Statut du Fichier Global",
            font=("Segoe UI", 9, "bold"),
            fg=COLORS['PRIMARY'],
            bg=COLORS['LIGHT']
        )
        status_title.pack(side=tk.LEFT)

        # Refresh status button with better styling
        refresh_status_btn = tk.Button(
            status_header,
            text="🔄 Actualiser",
            command=self._update_file_status_indicator,
            bg=COLORS['BORDER'],
            fg=COLORS['TEXT_SECONDARY'],
            font=("Segoe UI", 8),
            relief='flat',
            padx=8,
            pady=2,
            cursor='hand2'
        )
        refresh_status_btn.pack(side=tk.RIGHT)

        # Status content
        status_content = tk.Frame(status_container, bg=COLORS['LIGHT'])
        status_content.pack(fill=tk.X, padx=10, pady=(0, 8))

        self.file_status_icon = tk.Label(
            status_content,
            text="⏳",
            font=("Segoe UI", 10),
            bg=COLORS['LIGHT'],
            fg=COLORS['INFO']
        )
        self.file_status_icon.pack(side=tk.LEFT, padx=(0, 8))

        self.file_status_label = tk.Label(
            status_content,
            text="Vérification du statut du fichier...",
            font=UIConfig.FONT_SMALL,
            bg=COLORS['LIGHT'],
            fg=COLORS['TEXT_PRIMARY']
        )
        self.file_status_label.pack(side=tk.LEFT)

        # Enhanced generation button with action indicators
        button_container = tk.Frame(content_frame, bg=COLORS['CARD'])
        button_container.pack(fill=tk.X)

        self.generate_button = tk.Button(
            button_container,
            text="📊 Générer Excel Global",
            command=self._update_global_excel,
            bg=COLORS['PRIMARY'],
            fg='white',
            font=("Segoe UI", 9, "bold"),
            relief='flat',
            padx=15,
            pady=8,
            state=tk.DISABLED,
            cursor='hand2'
        )
        self.generate_button.pack(side=tk.LEFT)

        # Initialize file status
        self._update_file_status_indicator()

        # Initialize detailed view state
        self.detailed_view_enabled = False

    def _toggle_detailed_view(self):
        """Toggle between simple and detailed view in the summary."""
        self.detailed_view_enabled = not self.detailed_view_enabled

        if self.detailed_view_enabled:
            self.toggle_details_btn.config(text="📋 Vue Simple")
            self.progress_details_frame.pack(fill=tk.X, pady=(10, 0))
        else:
            self.toggle_details_btn.config(text="📊 Vue Détaillée")
            self.progress_details_frame.pack_forget()

        # Refresh the summary display if data exists
        if hasattr(self, 'processed_data') and self.processed_data:
            self._update_summary_display()

    def _update_progress_indicators(self, progress: float, status: str, details: str = ""):
        """Update all progress indicators with enhanced visualization."""
        # Update progress bar and percentage
        self.progress_var.set(progress)
        self.progress_percentage_label.config(text=f"{int(progress)}%")

        # Update status with appropriate icon
        if progress == 0:
            icon = "⏳"
            self.scan_status_indicator.config(text="⚪", fg=COLORS['INFO'])
        elif progress < 100:
            icon = "⚡"
            self.scan_status_indicator.config(text="🟡", fg=COLORS['WARNING'])
        else:
            icon = "✅"
            self.scan_status_indicator.config(text="🟢", fg=COLORS['SUCCESS'])

        self.status_icon_label.config(text=icon)
        self.status_label.config(text=status)

        # Update progress details if provided
        if details and self.detailed_view_enabled:
            self.progress_details_label.config(text=details)

    def _animate_icons(self):
        """Add subtle animation to icons during processing."""
        if hasattr(self, '_animation_active') and self._animation_active:
            # Rotate through different icons for animation effect
            icons = ["🔍", "🔎", "🔍", "🔎"]
            current_icon = self.scan_icon_label.cget('text')
            try:
                current_index = icons.index(current_icon)
                next_index = (current_index + 1) % len(icons)
                self.scan_icon_label.config(text=icons[next_index])
            except ValueError:
                self.scan_icon_label.config(text=icons[0])

            # Schedule next animation frame
            self.parent.after(500, self._animate_icons)

    def _start_animation(self):
        """Start icon animation during processing."""
        self._animation_active = True
        self._animate_icons()

    def _stop_animation(self):
        """Stop icon animation."""
        self._animation_active = False
        if hasattr(self, 'scan_icon_label'):
            self.scan_icon_label.config(text="🔍")
        if hasattr(self, 'generation_icon_label'):
            self.generation_icon_label.config(text="📈")







    def _load_existing_communes(self):
        """Load existing commune data from the global Excel file for comparison."""
        try:
            global_file_path = os.path.join(self.teams_folder_path, self.global_excel_filename)

            if not os.path.exists(global_file_path):
                self.logger.info("No existing global file found - all communes will be new")
                return {}

            pd = get_pandas()
            existing_df = pd.read_excel(
                global_file_path,
                sheet_name='Suivi Tickets',
                dtype={'Code INSEE': str, 'Insee': str},
                date_format=None  # CRITICAL: Prevent automatic date parsing
            )

            existing_communes = {}
            if not existing_df.empty and 'Nom Commune' in existing_df.columns:
                for _, row in existing_df.iterrows():
                    commune_name = str(row.get('Nom Commune', '')).strip()
                    insee_code = str(row.get('Code INSEE', '')).strip()
                    if commune_name:
                        # Use commune name as key, store additional info
                        existing_communes[commune_name] = {
                            'insee': insee_code,
                            'data': row.to_dict()
                        }

            self.logger.info(f"Loaded {len(existing_communes)} existing communes for comparison")
            return existing_communes

        except Exception as e:
            self.logger.error(f"Error loading existing communes: {e}")
            return {}

    def _analyze_commune_changes(self, processed_data):
        """Analyze which communes are new vs existing."""
        self.new_communes.clear()
        self.updated_communes.clear()

        for commune_data in processed_data:
            commune_name = commune_data['nom_commune']

            if commune_name in self.existing_communes:
                # Existing commune - will be updated
                self.updated_communes.append(commune_data)
            else:
                # New commune
                self.new_communes.append(commune_data)

        self.logger.info(f"Analysis: {len(self.new_communes)} new communes, {len(self.updated_communes)} communes to update")

    def _scan_and_process_folders(self):
        """Automatically scan Teams folder and process commune folders in one step."""
        try:
            # Start enhanced progress tracking
            self._start_animation()
            self._update_progress_indicators(5, "Initialisation du scan...", "Préparation de l'analyse des dossiers")
            self.scan_button.config(state=tk.DISABLED, text="🔄 Scan en cours...")
            self.generate_button.config(state=tk.DISABLED)

            def scan_and_process():
                # Step 1: Load existing commune data for comparison
                self.existing_communes = self._load_existing_communes()

                # Step 2: Auto-scan Teams folder
                from config.constants import TeamsConfig
                base_path = TeamsConfig.get_teams_base_path()

                if not os.path.exists(base_path):
                    raise Exception(f"Le dossier Teams n'a pas été trouvé: {base_path}")

                # Find commune folders
                found_folders = []
                for item in os.listdir(base_path):
                    item_path = os.path.join(base_path, item)
                    if os.path.isdir(item_path) and '_' in item:
                        # Check if folder contains Excel files
                        excel_files = [f for f in os.listdir(item_path) if f.endswith(('.xlsx', '.xls'))]
                        if excel_files:
                            found_folders.append(item_path)

                if not found_folders:
                    raise Exception("Aucun dossier de commune trouvé dans le répertoire Teams.")

                self.commune_folders = found_folders

                # Step 3: Process the folders
                processed_data = self._process_commune_folders()

                # Step 4: Analyze changes (new vs existing)
                self._analyze_commune_changes(processed_data)

                return processed_data

            def on_success(result):
                self._stop_animation()
                self.processed_data = result
                self._update_summary_display()

                # Enhanced success indicators
                new_count = len(self.new_communes)
                updated_count = len(self.updated_communes)
                total_count = len(result)

                # Update all UI elements
                self.generate_button.config(
                    state=tk.NORMAL,
                    text="📊 Générer Excel Global",
                    bg=COLORS['SUCCESS']
                )
                self.scan_button.config(
                    state=tk.NORMAL,
                    text="🚀 Lancer le Scan Complet",
                    bg=COLORS['SUCCESS']
                )

                # Enhanced status message
                status_msg = f"✅ Scan terminé - {total_count} communes analysées"
                details_msg = f"Nouvelles communes: {new_count}\nMises à jour: {updated_count}\nPrêt pour génération Excel"
                self._update_progress_indicators(100, status_msg, details_msg)



                # Update file status indicator
                self._update_file_status_indicator()

            def on_error(error):
                self._stop_animation()
                self.logger.error(f"Error in scan and process: {error}")

                # Enhanced error indicators
                self._update_progress_indicators(0, "❌ Erreur lors du scan", f"Détails: {str(error)[:100]}...")
                self.scan_button.config(
                    state=tk.NORMAL,
                    text="🚀 Lancer le Scan Complet",
                    bg=COLORS['SUCCESS']
                )
                self.scan_status_indicator.config(text="🔴", fg=COLORS['DANGER'])
                messagebox.showerror("Erreur", f"Erreur lors du scan:\n{error}")

            # Process asynchronously
            run_async_task(scan_and_process, on_success, on_error, "Scan and process")

        except Exception as e:
            self.logger.error(f"Error initiating scan and process: {e}")
            self.scan_button.config(state=tk.NORMAL)
            messagebox.showerror("Erreur", f"Erreur lors du lancement du scan:\n{e}")

    def _process_commune_folders(self) -> List[Dict[str, Any]]:
        """Process commune folders and extract data from page 3 of suivi files."""
        processed_data = []
        total_folders = len(self.commune_folders)

        for i, folder_path in enumerate(self.commune_folders):
            try:
                # Enhanced progress tracking
                progress = 20 + (i / total_folders) * 60  # Reserve 20% for initial scan, 60% for processing
                folder_name = os.path.basename(folder_path)

                # Update enhanced progress indicators
                status_msg = f"📁 Traitement: {folder_name} ({i+1}/{total_folders})"
                details_msg = f"Analyse du dossier: {folder_name}\nRecherche des fichiers Excel de suivi\nExtraction des données..."
                self._update_progress_indicators(progress, status_msg, details_msg)

                self.parent.update()

                # Find Excel files in the folder
                excel_files = []
                for file in os.listdir(folder_path):
                    if file.endswith(('.xlsx', '.xls')) and 'suivi' in file.lower():
                        excel_files.append(os.path.join(folder_path, file))

                if not excel_files:
                    self.logger.warning(f"No suivi Excel files found in {folder_path}")
                    continue

                # Process the most recent suivi file
                latest_file = max(excel_files, key=os.path.getmtime)
                commune_data = self._extract_commune_data(latest_file, folder_path)

                if commune_data:
                    processed_data.append(commune_data)
                    self.logger.info(f"Processed commune: {commune_data.get('nom_commune', 'Unknown')}")

            except Exception as e:
                self.logger.error(f"Error processing folder {folder_path}: {e}")
                continue

        # Enhanced analysis phase
        analysis_msg = "🔍 Analyse des nouvelles communes..."
        analysis_details = f"Comparaison avec les données existantes\nIdentification des nouvelles communes\nPréparation du résumé"
        self._update_progress_indicators(85, analysis_msg, analysis_details)
        self.parent.update()

        return processed_data

    def _extract_commune_data(self, excel_file_path: str, folder_path: str) -> Optional[Dict[str, Any]]:
        """Extract data from page 3 of a commune suivi file and detect RIP sheet if present."""
        try:
            pd = get_pandas()

            # Read page 3 of the Excel file with INSEE as string to preserve leading zeros
            # Also prevent automatic date parsing to avoid date format issues
            df = pd.read_excel(
                excel_file_path,
                sheet_name=2,
                dtype={'Insee': str, 'insee': str, 'Code INSEE': str},
                date_format=None  # CRITICAL: Prevent automatic date parsing
            )  # Page 3 (0-indexed)

            if df.empty:
                self.logger.warning(f"Page 3 is empty in {excel_file_path}")
                return None

            # Check if this is a RIP commune by looking at the Domaine field
            is_rip_commune = self._check_if_rip_commune(df)

            # Extract commune information from the first row or filename
            folder_name = os.path.basename(folder_path)
            parts = folder_name.split('_')

            commune_data = {
                'nom_commune': parts[0] if len(parts) > 0 else 'Unknown',
                'id_tache': parts[1] if len(parts) > 1 else 'Unknown',
                'insee_code': self._extract_insee_from_data(df) or 'Unknown',
                'file_path': excel_file_path,
                'folder_path': folder_path,
                'last_modified': os.path.getmtime(excel_file_path),
                'data_summary': self._summarize_page3_data(df),
                'is_rip_commune': is_rip_commune,
                'has_rip_sheet': False,
                'rip_data_summary': None
            }

            # If this is a RIP commune, try to read the RIP sheet (4th page)
            if is_rip_commune:
                rip_data = self._extract_rip_sheet_data(excel_file_path)
                if rip_data:
                    commune_data['has_rip_sheet'] = True
                    commune_data['rip_data_summary'] = rip_data
                    self.logger.info(f"RIP sheet detected and processed for {commune_data['nom_commune']}")

            return commune_data

        except Exception as e:
            self.logger.error(f"Error extracting data from {excel_file_path}: {e}")
            return None

    def _check_if_rip_commune(self, df) -> bool:
        """Check if this is a RIP commune by looking at the Domaine field."""
        try:
            # Look for Domaine column in page 3 data
            domaine_columns = ['Domaine', 'domaine', 'DOMAINE']
            for col in domaine_columns:
                if col in df.columns and not df[col].empty:
                    domaine_value = df[col].iloc[0]
                    if domaine_value is not None:
                        domaine_str = str(domaine_value).strip().upper()
                        return domaine_str == 'RIP'
            return False
        except:
            return False

    def _extract_rip_sheet_data(self, excel_file_path: str) -> Optional[Dict[str, Any]]:
        """Extract data from the RIP sheet (4th page) if it exists."""
        try:
            pd = get_pandas()

            # Try to read the 4th sheet (index 3)
            rip_df = pd.read_excel(
                excel_file_path,
                sheet_name=3,
                dtype={'Code INSEE': str, 'Insee': str},
                date_format=None  # Prevent automatic date parsing
            )

            if rip_df.empty:
                self.logger.warning(f"RIP sheet (page 4) is empty in {excel_file_path}")
                return None

            # Summarize RIP sheet data
            rip_summary = {
                'total_rows': len(rip_df),
                'columns': list(rip_df.columns),
                'has_duration_data': 'Durée' in rip_df.columns,
                'total_duration': 0
            }

            # Calculate total duration if duration column exists
            if 'Durée' in rip_df.columns:
                try:
                    # Convert duration values to numeric, handling empty/invalid values
                    duration_values = pd.to_numeric(rip_df['Durée'], errors='coerce').fillna(0)
                    rip_summary['total_duration'] = float(duration_values.sum())
                except:
                    rip_summary['total_duration'] = 0

            self.logger.info(f"RIP sheet processed: {rip_summary['total_rows']} rows, total duration: {rip_summary['total_duration']}")
            return rip_summary

        except Exception as e:
            # This is expected for non-RIP files or files without 4th sheet
            self.logger.debug(f"No RIP sheet found in {excel_file_path}: {e}")
            return None

    def _extract_insee_from_data(self, df) -> Optional[str]:
        """Extract INSEE code from the dataframe."""
        try:
            # Look for INSEE code in common column names
            insee_columns = ['insee', 'code_insee', 'INSEE', 'Code INSEE', 'Insee']
            for col in insee_columns:
                if col in df.columns and not df[col].empty:
                    insee_value = df[col].iloc[0]
                    if insee_value is not None:
                        # Format INSEE code to preserve leading zeros (5 digits)
                        insee_str = str(insee_value).strip()
                        if insee_str.isdigit() and len(insee_str) <= 5:
                            return insee_str.zfill(5)  # Pad with leading zeros to 5 digits
                        return insee_str
            return None
        except:
            return None

    def _summarize_page3_data(self, df) -> Dict[str, Any]:
        """Summarize the data from page 3."""
        try:
            return {
                'total_rows': len(df),
                'columns': list(df.columns),
                'non_empty_rows': len(df.dropna(how='all'))
            }
        except:
            return {'total_rows': 0, 'columns': [], 'non_empty_rows': 0}

    def _update_summary_display(self):
        """Update the summary text display with enhanced formatting and detailed analysis."""
        try:
            self.summary_text.config(state=tk.NORMAL)
            self.summary_text.delete(1.0, tk.END)

            if not self.processed_data:
                self.summary_text.insert(tk.END, "Aucune donnée trouvée.", "info")
                self.summary_text.config(state=tk.DISABLED)
                return

            new_count = len(self.new_communes)
            updated_count = len(self.updated_communes)
            total_count = len(self.processed_data)

            # Enhanced header with better formatting - Wider for more space
            self.summary_text.insert(tk.END, "╔════════════════════════════════════════════════════════════════╗\n", "header")
            self.summary_text.insert(tk.END, "║                        ANALYSE DÉTAILLÉE DU SCAN                        ║\n", "header")
            self.summary_text.insert(tk.END, "╚════════════════════════════════════════════════════════════════╝\n\n", "header")

            # Summary statistics with icons - More detailed
            self.summary_text.insert(tk.END, f"📊 Total communes trouvées: {total_count}\n", "info")
            self.summary_text.insert(tk.END, f"🆕 Nouvelles communes: {new_count}\n", "new_commune")
            self.summary_text.insert(tk.END, f"🔄 Communes à mettre à jour: {updated_count}\n", "update_commune")

            # Add more detailed statistics
            total_rows = sum(data['data_summary']['total_rows'] for data in self.processed_data)
            self.summary_text.insert(tk.END, f"📋 Total lignes de données: {total_rows}\n", "info")

            # Show scan time
            from datetime import datetime
            scan_time = datetime.now().strftime("%d/%m/%Y à %H:%M:%S")
            self.summary_text.insert(tk.END, f"⏰ Scan effectué le: {scan_time}\n\n", "info")

            # Enhanced new communes section - Wider format
            if self.new_communes:
                self.summary_text.insert(tk.END, f"┌─ 🆕 NOUVELLES COMMUNES ({new_count}) ─────────────────────────────────────────────┐\n", "header")

                for i, data in enumerate(self.new_communes, 1):
                    commune_info = (
                        f"│ {i:2d}. {data['nom_commune']:<25} │ ID: {data['id_tache']:<12} │ INSEE: {data['insee_code']:<8} │\n"
                        f"│     Lignes: {data['data_summary']['total_rows']:>4} │ Fichier: {os.path.basename(data['file_path'])[:35]:<35} │\n"
                    )
                    self.summary_text.insert(tk.END, commune_info, "new_commune")

                    if i < len(self.new_communes):
                        self.summary_text.insert(tk.END, "├─────────────────────────────────────────────────────────────────────────────┤\n", "header")

                self.summary_text.insert(tk.END, "└─────────────────────────────────────────────────────────────────────────────┘\n\n", "header")

            # Enhanced updated communes section - Wider format
            if self.updated_communes:
                self.summary_text.insert(tk.END, f"┌─ 🔄 COMMUNES À METTRE À JOUR ({updated_count}) ─────────────────────────────────────┐\n", "header")

                for i, data in enumerate(self.updated_communes, 1):
                    commune_info = (
                        f"│ {i:2d}. {data['nom_commune']:<25} │ ID: {data['id_tache']:<12} │ INSEE: {data['insee_code']:<8} │\n"
                        f"│     Lignes: {data['data_summary']['total_rows']:>4} │ Fichier: {os.path.basename(data['file_path'])[:35]:<35} │\n"
                    )
                    self.summary_text.insert(tk.END, commune_info, "update_commune")

                    if i < len(self.updated_communes):
                        self.summary_text.insert(tk.END, "├─────────────────────────────────────────────────────────────────────────────┤\n", "header")

                self.summary_text.insert(tk.END, "└─────────────────────────────────────────────────────────────────────────────┘\n\n", "header")

            # Count RIP communes for summary
            rip_count = len([data for data in self.processed_data if data.get('is_rip_commune', False)])
            rip_with_sheets = len([data for data in self.processed_data if data.get('has_rip_sheet', False)])

            # Enhanced Excel generation summary
            global_file_path = os.path.join(self.teams_folder_path, self.global_excel_filename)
            file_exists = os.path.exists(global_file_path)
            action = "mis à jour" if file_exists else "créé"

            self.summary_text.insert(tk.END, "┌─ 📊 GÉNÉRATION EXCEL GLOBAL ─────────────────────────────────────────────────┐\n", "header")
            self.summary_text.insert(tk.END, f"│ Fichier sera {action:<60} │\n", "info")
            self.summary_text.insert(tk.END, "├─────────────────────────────────────────────────────────────────────────────┤\n", "header")
            self.summary_text.insert(tk.END, f"│ 📄 Nom: {self.global_excel_filename:<65} │\n", "info")
            self.summary_text.insert(tk.END, f"│ 📊 Page 1: Suivi Tickets (source: page 3 des fichiers communes)            │\n", "info")
            self.summary_text.insert(tk.END, f"│ 🏢 Page 2: Traitement CMS Adr (source: page 1 des fichiers communes)      │\n", "info")
            self.summary_text.insert(tk.END, f"│ 📍 Page 3: Traitement PA (source: page 2 des fichiers communes)           │\n", "info")
            if rip_with_sheets > 0:
                self.summary_text.insert(tk.END, f"│ 🔧 Page 4: Traitement RIP (source: page 4 des fichiers communes)          │\n", "info")
            self.summary_text.insert(tk.END, "├─────────────────────────────────────────────────────────────────────────────┤\n", "header")
            self.summary_text.insert(tk.END, f"│ 🆕 Nouvelles communes: {new_count:>3} ajoutées                                      │\n", "new_commune")
            self.summary_text.insert(tk.END, f"│ 🔄 Communes existantes: {updated_count:>3} mises à jour                              │\n", "update_commune")
            if rip_count > 0:
                self.summary_text.insert(tk.END, f"│ 🔧 Communes RIP: {rip_count:>3} détectées ({rip_with_sheets} avec données RIP complètes)        │\n", "info")
            self.summary_text.insert(tk.END, "├─────────────────────────────────────────────────────────────────────────────┤\n", "header")
            self.summary_text.insert(tk.END, f"│ 📁 Emplacement: Teams/Global/                                               │\n", "info")

            status_text = "Fichier existant - données préservées" if file_exists else "Nouveau fichier - toutes communes ajoutées"
            self.summary_text.insert(tk.END, f"│ 📋 Statut: {status_text[:30]:<30} │\n", "info")
            self.summary_text.insert(tk.END, "└─────────────────────────────────────────────────┘\n", "header")

            self.summary_text.config(state=tk.DISABLED)

            # Update statistics in header
            if hasattr(self, 'stats_label'):
                stats_text = f"🆕 {new_count} nouvelles | 🔄 {updated_count} à mettre à jour"
                self.stats_label.config(text=stats_text)

        except Exception as e:
            self.logger.error(f"Error updating summary display: {e}")

    def _update_global_excel(self):
        """Update or create the global Excel file."""
        if not self.processed_data:
            messagebox.showwarning("Aucune donnée", "Aucune donnée à traiter. Veuillez d'abord scanner des dossiers.")
            return

        try:
            self.status_label.config(text="Mise à jour du fichier Excel...")
            self.progress_var.set(0)

            def update_process():
                return self._create_or_update_global_excel_file()

            def on_success(result):
                file_path, is_new = result
                action = "créé" if is_new else "mis à jour"
                self.status_label.config(text=f"Fichier Excel {action} avec succès")
                self.progress_var.set(100)

                # Prepare detailed success message
                new_count = len(self.new_communes)
                updated_count = len(self.updated_communes)
                total_count = len(self.processed_data)

                if is_new:
                    details = f"Nouveau fichier créé avec {total_count} communes."
                else:
                    details = f"Fichier mis à jour:\n• {new_count} nouvelles communes ajoutées\n• {updated_count} communes existantes mises à jour\n• Total: {total_count} communes traitées"

                # Show enhanced success message with statistics
                response = messagebox.askyesno(
                    "Mise à jour terminée",
                    f"Le fichier Excel global a été {action} avec succès!\n\n"
                    f"📊 Statistiques:\n{details}\n\n"
                    f"📁 Fichier: {os.path.basename(file_path)}\n"
                    f"📍 Emplacement: {os.path.dirname(file_path)}\n\n"
                    f"Voulez-vous ouvrir le fichier?"
                )

                if response:
                    self._open_generated_file(file_path)

                # Update file status indicator after successful update
                self._update_file_status_indicator()

            def on_error(error):
                self.logger.error(f"Error updating Excel: {error}")
                self.status_label.config(text="Erreur lors de la mise à jour")
                self.progress_var.set(0)

                # Provide user-friendly error messages
                error_str = str(error)
                if "Permission denied" in error_str or "WinError 33" in error_str or "annulée par l'utilisateur" in error_str:
                    # File access issues - user already saw the dialog
                    if "annulée par l'utilisateur" not in error_str:
                        messagebox.showwarning(
                            "Fichier en cours d'utilisation",
                            "🔒 Le fichier Excel global est actuellement ouvert.\n\n"
                            "Veuillez fermer Excel et cliquer sur 'Mettre à jour Excel Global' pour réessayer."
                        )
                else:
                    # Other errors
                    messagebox.showerror("Erreur", f"Erreur lors de la mise à jour:\n{error}")

            # Update asynchronously
            run_async_task(update_process, on_success, on_error, "Excel update")

        except Exception as e:
            self.logger.error(f"Error initiating Excel update: {e}")
            messagebox.showerror("Erreur", f"Erreur lors du lancement de la mise à jour:\n{e}")

    def _create_or_update_global_excel_file(self) -> tuple:
        """Create or update the global Excel file with aggregated data."""
        try:
            pd = get_pandas()

            # Ensure output directory exists
            os.makedirs(self.teams_folder_path, exist_ok=True)

            # Define the fixed filename
            file_path = os.path.join(self.teams_folder_path, self.global_excel_filename)

            # Check if file exists
            is_new_file = not os.path.exists(file_path)

            # Check file access before proceeding
            if not is_new_file:
                access_result = check_file_access(file_path, 'rw')
                if not access_result['accessible']:
                    self.logger.warning(f"File access issue: {access_result['error_message']}")

                    # Show user-friendly dialog and get retry decision
                    if access_result['error_type'] in ['file_locked', 'permission_denied']:
                        # This is a file-in-use situation, show the custom dialog
                        retry = self._show_file_access_dialog(access_result, file_path)
                        if not retry:
                            raise Exception("Opération annulée par l'utilisateur")

                        # User wants to retry, check again
                        retry_access = check_file_access(file_path, 'rw')
                        if not retry_access['accessible']:
                            # Still not accessible, raise with user-friendly message
                            raise Exception(f"{retry_access['user_message']}\n\nVeuillez fermer Excel et réessayer.")
                    else:
                        # Other types of errors, raise with user message
                        raise Exception(access_result['user_message'])

            if is_new_file:
                # Create new file
                self.logger.info("Creating new global Excel file")
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    # Page 1: Aggregated data from page 3 of all communes (Suivi Tickets)
                    self._create_page1_aggregated_data(writer, None)

                    # Page 2: Aggregated data from page 1 of all communes (CM Adresse)
                    self._create_page2_cm_adresse_data(writer, None)

                    # Page 3: Aggregated data from page 2 of all communes (Plan Adressage)
                    self._create_page3_plan_adressage_data(writer, None)

                    # Page 4: Aggregated data from RIP sheets (if any RIP communes exist)
                    self._create_page4_rip_data(writer, None)
            else:
                # Update existing file
                self.logger.info("Updating existing global Excel file")
                # Read existing data from all sheets
                existing_page1_df = None
                existing_page2_df = None
                existing_page3_df = None
                existing_page4_df = None

                try:
                    existing_page1_df = pd.read_excel(
                        file_path,
                        sheet_name='Suivi Tickets',
                        dtype={'Code INSEE': str, 'Insee': str},
                        date_format=None  # CRITICAL: Prevent automatic date parsing
                    )
                    # Immediately apply date formatting to existing data
                    if existing_page1_df is not None and not existing_page1_df.empty:
                        existing_page1_df = self._format_date_columns(existing_page1_df)
                        self.logger.info("Applied date formatting to existing Suivi Tickets data")
                except:
                    pass
                try:
                    existing_page2_df = pd.read_excel(
                        file_path,
                        sheet_name='Traitement CMS Adr',
                        dtype={'Code INSEE': str, 'Insee': str},
                        date_format=None  # CRITICAL: Prevent automatic date parsing
                    )
                    # Immediately apply date formatting to existing data
                    if existing_page2_df is not None and not existing_page2_df.empty:
                        existing_page2_df = self._format_date_columns(existing_page2_df)
                        self.logger.info("Applied date formatting to existing Traitement CMS Adr data")
                except:
                    pass
                try:
                    existing_page3_df = pd.read_excel(
                        file_path,
                        sheet_name='Traitement PA',
                        dtype={'Code INSEE': str, 'Insee': str},
                        date_format=None  # CRITICAL: Prevent automatic date parsing
                    )
                    # Immediately apply date formatting to existing data
                    if existing_page3_df is not None and not existing_page3_df.empty:
                        existing_page3_df = self._format_date_columns(existing_page3_df)
                        self.logger.info("Applied date formatting to existing Traitement PA data")
                except:
                    pass
                try:
                    existing_page4_df = pd.read_excel(
                        file_path,
                        sheet_name='Traitement RIP',
                        dtype={'Code INSEE': str, 'Insee': str},
                        date_format=None  # CRITICAL: Prevent automatic date parsing
                    )
                    # Immediately apply date formatting to existing data
                    if existing_page4_df is not None and not existing_page4_df.empty:
                        existing_page4_df = self._format_date_columns(existing_page4_df)
                        self.logger.info("Applied date formatting to existing Traitement RIP data")
                except:
                    pass

                # Update with new data
                with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    self._create_page1_aggregated_data(writer, existing_page1_df)
                    self._create_page2_cm_adresse_data(writer, existing_page2_df)
                    self._create_page3_plan_adressage_data(writer, existing_page3_df)
                    self._create_page4_rip_data(writer, existing_page4_df)

            # Apply formatting similar to individual suivi files
            self._format_global_excel(file_path)

            self.logger.info(f"Global Excel file {'created' if is_new_file else 'updated'}: {file_path}")
            return file_path, is_new_file

        except Exception as e:
            self.logger.error(f"Error creating/updating global Excel file: {e}")
            raise

    def _format_date_columns(self, df):
        """Keep date columns exactly as they are in the source files."""
        try:
            pd = get_pandas()

            # List of column names that should be treated as dates (expanded list)
            date_column_keywords = ['date', 'livraison', 'affectation', 'dépose', 'traitement']

            # List of column names that should NOT be treated as dates (duration/time columns)
            duration_column_keywords = ['durée', 'duration', 'temps', 'time', 'traitement optimum', 'finale', 'motif']

            for column in df.columns:
                column_lower = str(column).lower()

                # First check if it's a duration column - if so, skip it
                is_duration = any(keyword in column_lower for keyword in duration_column_keywords)
                if is_duration:
                    continue

                # Then check if it's a date column
                is_date = any(keyword in column_lower for keyword in date_column_keywords)
                if is_date:
                    # Only convert datetime objects to strings, keep string dates as-is
                    for idx in df.index:
                        value = df.at[idx, column]
                        if pd.notna(value) and value != '':
                            try:
                                # Convert datetime objects to ISO format
                                if hasattr(value, 'strftime'):
                                    df.at[idx, column] = value.strftime('%Y-%m-%d')
                                # Convert string values to ISO format
                                else:
                                    converted_date = self._normalize_date_value(value, column)
                                    df.at[idx, column] = converted_date
                            except Exception as e:
                                self.logger.warning(f"Could not process date in column {column}: {e}")
                                continue

            return df

        except Exception as e:
            self.logger.error(f"Error formatting date columns: {e}")
            return df

    def _create_page1_aggregated_data(self, writer, existing_df=None):
        """Create page 1 with aggregated data from all communes, maintaining exact structure from source files."""
        try:
            pd = get_pandas()

            # Get the structure from the first commune's page 3 to maintain consistency
            template_structure = None
            if self.processed_data:
                try:
                    first_commune = self.processed_data[0]
                    template_df = pd.read_excel(first_commune['file_path'], sheet_name=2)
                    if not template_df.empty:
                        template_structure = list(template_df.columns)
                except Exception as e:
                    self.logger.warning(f"Could not get template structure: {e}")

            # Prepare aggregated data maintaining original structure
            aggregated_rows = []

            for commune_data in self.processed_data:
                try:
                    # Read the original page 3 data with INSEE as string to preserve leading zeros
                    # Also ensure date columns are read as strings to prevent automatic date conversion
                    original_df = pd.read_excel(
                        commune_data['file_path'],
                        sheet_name=2,
                        dtype={'Insee': str, 'insee': str, 'Code INSEE': str},
                        date_format=None  # Prevent automatic date parsing
                    )

                    if not original_df.empty:
                        # Format date columns to remove time component
                        original_df = self._format_date_columns(original_df)

                        # Take all rows from page 3 and add commune identification
                        for _, row in original_df.iterrows():
                            row_data = row.to_dict()

                            # Ensure we have commune identification columns at the beginning
                            # Handle different column name variations
                            if 'Nom Commune' not in row_data and 'Nom de commune' not in row_data:
                                row_data['Nom Commune'] = commune_data['nom_commune']
                            if 'Code INSEE' not in row_data:
                                row_data['Code INSEE'] = commune_data['insee_code']

                            # Normalize column names for consistency
                            if 'Nom de commune' in row_data and 'Nom Commune' not in row_data:
                                row_data['Nom Commune'] = row_data.pop('Nom de commune')

                            aggregated_rows.append(row_data)

                except Exception as e:
                    self.logger.error(f"Error processing commune {commune_data['nom_commune']}: {e}")
                    continue

            # Create DataFrame
            if aggregated_rows:
                new_df = pd.DataFrame(aggregated_rows)

                # If we have existing data, merge it
                if existing_df is not None and not existing_df.empty:
                    # Format date columns in existing data
                    existing_df = self._format_date_columns(existing_df)

                    # Remove existing entries for communes that are being updated
                    commune_names = [data['nom_commune'] for data in self.processed_data]
                    if 'Nom Commune' in existing_df.columns:
                        existing_df = existing_df[~existing_df['Nom Commune'].isin(commune_names)]

                    # Combine existing and new data
                    combined_df = pd.concat([existing_df, new_df], ignore_index=True, sort=False)
                else:
                    combined_df = new_df

                # Ensure Nom Commune and Code INSEE are first columns
                columns = list(combined_df.columns)
                # Define the expected column order including the new columns
                column_order = ['Nom Commune', 'ID tâche Plan Adressage', 'Code INSEE', 'Domaine', 'Type de Commune', 'Type de base',
                               'Nbr des voies CM', 'Nbr des IMB PA', 'Date d\'affectation', 'Temps préparation QGis',
                               'Durée Totale CM', 'Duréé Totale PA', 'Traitement Optimum', 'Durée Finale',
                               'Date Livraison', 'Etat Ticket PA ', 'ID Tache 501/511', 'Date Dépose Ticket 501/511',
                               'Dépose Ticket UPR', 'ID tâche UPR', 'Collaborateur']

                # Reorder columns, keeping any additional columns at the end
                existing_columns = list(combined_df.columns)
                final_columns = []
                for col in column_order:
                    if col in existing_columns:
                        final_columns.append(col)
                        existing_columns.remove(col)
                final_columns.extend(existing_columns)  # Add any remaining columns

                combined_df = combined_df.reindex(columns=final_columns)

                # Final validation and formatting of date columns before writing
                combined_df = self._validate_and_format_dates_before_writing(combined_df, 'Suivi Tickets')

                # Write to Excel with the correct sheet name
                combined_df.to_excel(writer, sheet_name='Suivi Tickets', index=False)

                # Format INSEE columns as text to preserve leading zeros
                self._format_insee_columns_as_text(writer, 'Suivi Tickets', combined_df)

                # Format date columns as text to prevent Excel auto-formatting
                self._format_date_columns_as_text(writer, 'Suivi Tickets', combined_df)
            else:
                # Create empty DataFrame with basic headers
                empty_df = pd.DataFrame(columns=['Nom Commune', 'Code INSEE'])
                empty_df.to_excel(writer, sheet_name='Suivi Tickets', index=False)

        except Exception as e:
            self.logger.error(f"Error creating page 1 data: {e}")
            raise

    def _create_page2_cm_adresse_data(self, writer, existing_df=None):
        """Create page 2 with aggregated CM Adresse data from page 1 of all commune files."""
        try:
            pd = get_pandas()

            # Prepare aggregated data from CM Adresse sheets (page 1)
            aggregated_rows = []

            for commune_data in self.processed_data:
                try:
                    # Read the CM Adresse sheet (page 1) data with INSEE as string to preserve leading zeros
                    # Also ensure date columns are read as strings to prevent automatic date conversion
                    cm_df = pd.read_excel(
                        commune_data['file_path'],
                        sheet_name=0,
                        dtype={'Insee': str, 'insee': str, 'Code INSEE': str},
                        date_format=None  # Prevent automatic date parsing
                    )  # First sheet

                    # Debug: Log the actual columns found
                    self.logger.info(f"CM Adresse columns in {commune_data['nom_commune']}: {list(cm_df.columns)}")

                    if not cm_df.empty:
                        # Format date columns to remove time component
                        cm_df = self._format_date_columns(cm_df)

                        # Extract specific columns for Page 2 based on actual CM Adresse structure
                        for _, row in cm_df.iterrows():
                            row_data = {}

                            # Map columns based on actual CM Adresse structure from excel_generator.py
                            columns = cm_df.columns.tolist()
                            self.logger.debug(f"CM Adresse columns found: {columns}")

                            # Extract required columns based on actual structure
                            # From excel_generator.py: CM Adresse columns are:
                            # A: Nom commune, B: Insee, C: ID Tache, D: Voie demandé, E: Motif Voie,
                            # F: CODE RIVOLI, G: GPS (X,Y), H: Centre/Zone, I: Status PC, J: Descriptif Commentaire,
                            # K: Collaborateur, L: Date affectation, M: Date traitement, N: Date livraison, O: Durée, P: STATUT Ticket

                            row_data['Nom commune'] = self._get_column_value(row, columns, ['Nom commune'], commune_data['nom_commune'])
                            row_data['Insee'] = self._get_column_value(row, columns, ['Insee'], commune_data['insee_code'])
                            row_data['ID Tache'] = self._get_column_value(row, columns, ['ID Tache'], '')
                            row_data['Motif Voie'] = self._get_column_value(row, columns, ['Motif Voie'], '')  # NEW: Add Motif Voie column
                            row_data['Collaborateur'] = self._get_column_value(row, columns, ['Collaborateur'], '')
                            row_data['Date affectation'] = self._get_column_value(row, columns, ['Date affectation'], '')
                            row_data['Date traitement'] = self._get_column_value(row, columns, ['Date traitement'], '')
                            row_data['Date livraison'] = self._get_column_value(row, columns, ['Date livraison'], '')
                            row_data['STATUT Ticket'] = self._get_column_value(row, columns, ['STATUT Ticket'], '')
                            row_data['Durée'] = self._get_column_value(row, columns, ['Durée'], '')

                            aggregated_rows.append(row_data)

                except Exception as e:
                    self.logger.error(f"Error processing CM Adresse data for commune {commune_data['nom_commune']}: {e}")
                    continue

            # Create DataFrame
            if aggregated_rows:
                new_df = pd.DataFrame(aggregated_rows)

                # Apply comprehensive date formatting to the new DataFrame
                new_df = self._format_date_columns(new_df)
                self.logger.info(f"Applied date formatting to new CM Adresse data: {len(new_df)} rows")

                # If we have existing data, merge it
                if existing_df is not None and not existing_df.empty:
                    # Format date columns in existing data
                    existing_df = self._format_date_columns(existing_df)

                    # Remove existing entries for communes that are being updated
                    commune_names = [data['nom_commune'] for data in self.processed_data]
                    if 'Nom commune' in existing_df.columns:
                        existing_df = existing_df[~existing_df['Nom commune'].isin(commune_names)]

                    # Combine existing and new data
                    combined_df = pd.concat([existing_df, new_df], ignore_index=True, sort=False)
                else:
                    combined_df = new_df

                # Ensure proper column order with new Motif Voie column
                column_order = ['Nom commune', 'Insee', 'ID Tache', 'Motif Voie', 'Collaborateur',
                               'Date affectation', 'Date traitement', 'Date livraison', 'STATUT Ticket']

                # Reorder columns, keeping any additional columns at the end
                existing_columns = list(combined_df.columns)
                final_columns = []
                for col in column_order:
                    if col in existing_columns:
                        final_columns.append(col)
                        existing_columns.remove(col)
                final_columns.extend(existing_columns)  # Add any remaining columns

                combined_df = combined_df.reindex(columns=final_columns)

                # Apply aggressive date formatting specifically for problematic columns
                combined_df = self._force_date_formatting_for_pages_2_3(combined_df, 'Traitement CMS Adr')

                # Final validation and formatting of date columns before writing
                combined_df = self._validate_and_format_dates_before_writing(combined_df, 'Traitement CMS Adr')

                # Write to Excel with the correct sheet name
                combined_df.to_excel(writer, sheet_name='Traitement CMS Adr', index=False)

                # Format INSEE columns as text to preserve leading zeros
                self._format_insee_columns_as_text(writer, 'Traitement CMS Adr', combined_df)

                # Format date columns as text to prevent Excel auto-formatting
                self._format_date_columns_as_text(writer, 'Traitement CMS Adr', combined_df)
            else:
                # Create empty DataFrame with headers including new Motif Voie and Durée columns
                empty_df = pd.DataFrame(columns=['Nom commune', 'Insee', 'ID Tache', 'Motif Voie', 'Collaborateur',
                                               'Date affectation', 'Date traitement', 'Date livraison', 'STATUT Ticket', 'Durée'])
                empty_df.to_excel(writer, sheet_name='Traitement CMS Adr', index=False)

        except Exception as e:
            self.logger.error(f"Error creating page 2 CM Adresse data: {e}")
            raise

    def _create_page3_plan_adressage_data(self, writer, existing_df=None):
        """Create page 3 with aggregated Plan Adressage data from page 2 of all commune files."""
        try:
            pd = get_pandas()

            # Prepare aggregated data from Plan Adressage sheets (page 2)
            aggregated_rows = []

            for commune_data in self.processed_data:
                try:
                    # Read the Plan Adressage sheet (page 2) data with INSEE as string to preserve leading zeros
                    # Also ensure date columns are read as strings to prevent automatic date conversion
                    plan_df = pd.read_excel(
                        commune_data['file_path'],
                        sheet_name=1,
                        dtype={'Insee': str, 'insee': str, 'Code INSEE': str},
                        date_format=None  # Prevent automatic date parsing
                    )  # Second sheet

                    # Debug: Log the actual columns found
                    self.logger.info(f"Plan Adressage columns in {commune_data['nom_commune']}: {list(plan_df.columns)}")

                    if not plan_df.empty:
                        # Format date columns to remove time component (but preserve duration columns)
                        plan_df = self._format_date_columns(plan_df)

                        # Extract specific columns for Page 3
                        for _, row in plan_df.iterrows():
                            row_data = {}

                            # Map columns based on expected structure
                            columns = plan_df.columns.tolist()

                            # Extract required columns based on actual Plan Adressage structure
                            # From excel_generator.py: Plan Adressage has commune info added at beginning:
                            # A: Nom commune, B: Insee, then original columns including:
                            # C: Num Dossier Site, D: Num Voie Site, E: Comp Voie Site, F: Libelle Voie Site,
                            # G: Batiment IMB, H: Motif, I: Même Adresse, J: Numero Voie BAN, K: Repondant Voie BAN,
                            # L: Libelle Voie BAN, M: Collaborateur, N: Date traitement, O: Durée, P: Traitement Optimum

                            row_data['Nom commune'] = self._get_column_value(row, columns, ['Nom commune'], commune_data['nom_commune'])
                            row_data['Insee'] = self._get_column_value(row, columns, ['Insee'], commune_data['insee_code'])
                            row_data['Num Dossier Site'] = self._get_column_value(row, columns, ['Num Dossier Site'], '')
                            row_data['Motif'] = self._get_column_value(row, columns, ['Motif'], '')
                            row_data['Adresse BAN'] = self._get_column_value(row, columns, ['Adresse BAN'], '')
                            row_data['Collaborateur'] = self._get_column_value(row, columns, ['Collaborateur'], '')
                            row_data['Date traitement'] = self._get_column_value(row, columns, ['Date traitement'], '')
                            row_data['Durée'] = self._get_column_value(row, columns, ['Durée'], '')

                            aggregated_rows.append(row_data)

                except Exception as e:
                    self.logger.error(f"Error processing Plan Adressage data for commune {commune_data['nom_commune']}: {e}")
                    continue

            # Create DataFrame
            if aggregated_rows:
                new_df = pd.DataFrame(aggregated_rows)

                # Apply comprehensive date formatting to the new DataFrame
                new_df = self._format_date_columns(new_df)
                self.logger.info(f"Applied date formatting to new Plan Adressage data: {len(new_df)} rows")

                # If we have existing data, merge it
                if existing_df is not None and not existing_df.empty:
                    # Format date columns in existing data
                    existing_df = self._format_date_columns(existing_df)

                    # Remove existing entries for communes that are being updated
                    commune_names = [data['nom_commune'] for data in self.processed_data]
                    if 'Nom commune' in existing_df.columns:
                        existing_df = existing_df[~existing_df['Nom commune'].isin(commune_names)]

                    # Combine existing and new data
                    combined_df = pd.concat([existing_df, new_df], ignore_index=True, sort=False)
                else:
                    combined_df = new_df

                # Remove obsolete columns if they exist
                obsolete_columns = ['Batiment IMB', 'Traitement Optimum']
                for col in obsolete_columns:
                    if col in combined_df.columns:
                        combined_df = combined_df.drop(columns=[col])
                        self.logger.info(f"Removed obsolete column: {col}")

                # Ensure proper column order for Plan Adressage page
                column_order = ['Nom commune', 'Insee', 'Num Dossier Site', 'Motif', 'Adresse BAN',
                               'Collaborateur', 'Date traitement', 'Durée']

                # Reorder columns, keeping any additional columns at the end
                existing_columns = list(combined_df.columns)
                final_columns = []
                for col in column_order:
                    if col in existing_columns:
                        final_columns.append(col)
                        existing_columns.remove(col)
                final_columns.extend(existing_columns)  # Add any remaining columns

                combined_df = combined_df.reindex(columns=final_columns)

                # Apply aggressive date formatting specifically for problematic columns
                combined_df = self._force_date_formatting_for_pages_2_3(combined_df, 'Traitement PA')

                # Final validation and formatting of date columns before writing
                combined_df = self._validate_and_format_dates_before_writing(combined_df, 'Traitement PA')

                # Write to Excel with the correct sheet name
                combined_df.to_excel(writer, sheet_name='Traitement PA', index=False)

                # Format INSEE columns as text to preserve leading zeros
                self._format_insee_columns_as_text(writer, 'Traitement PA', combined_df)

                # Format date columns as text to prevent Excel auto-formatting
                self._format_date_columns_as_text(writer, 'Traitement PA', combined_df)
            else:
                # Create empty DataFrame with headers for Plan Adressage
                empty_df = pd.DataFrame(columns=['Nom commune', 'Insee', 'Num Dossier Site', 'Motif', 'Adresse BAN',
                                               'Collaborateur', 'Date traitement', 'Durée'])
                empty_df.to_excel(writer, sheet_name='Traitement PA', index=False)

        except Exception as e:
            self.logger.error(f"Error creating page 3 Plan Adressage data: {e}")
            raise

    def _create_page4_rip_data(self, writer, existing_df=None):
        """Create page 4 with aggregated RIP data from RIP sheets of all RIP commune files."""
        try:
            pd = get_pandas()

            # Check if we have any RIP communes
            rip_communes = [data for data in self.processed_data if data.get('is_rip_commune', False) and data.get('has_rip_sheet', False)]

            if not rip_communes:
                # No RIP communes, create empty sheet
                empty_df = pd.DataFrame(columns=['Nom commune', 'Code INSEE', 'ID tâche', 'Type', 'Acte de traitement',
                                               'Commentaire', 'Date d\'affectation', 'Date de traitement', 'Date de livraison',
                                               'Collaborateur', 'Durée'])
                empty_df.to_excel(writer, sheet_name='Traitement RIP', index=False)
                self.logger.info("Created empty RIP sheet - no RIP communes found")
                return

            # Prepare aggregated data from RIP sheets (page 4)
            aggregated_rows = []

            for commune_data in rip_communes:
                try:
                    # Read the RIP sheet (page 4) data with INSEE as string to preserve leading zeros
                    rip_df = pd.read_excel(
                        commune_data['file_path'],
                        sheet_name=3,
                        dtype={'Code INSEE': str, 'Insee': str},
                        date_format=None  # Prevent automatic date parsing
                    )  # Fourth sheet (0-indexed)

                    self.logger.info(f"RIP sheet columns in {commune_data['nom_commune']}: {list(rip_df.columns)}")

                    if not rip_df.empty:
                        # Format date columns to remove time component
                        rip_df = self._format_date_columns(rip_df)

                        # Extract specific columns for Page 4 based on RIP sheet structure
                        for _, row in rip_df.iterrows():
                            row_data = {}

                            # Map columns based on RIP sheet structure from excel_generator.py
                            columns = rip_df.columns.tolist()
                            self.logger.debug(f"RIP sheet columns found: {columns}")

                            # Extract required columns based on RIP sheet structure:
                            # A: Nom commune, B: Code INSEE, C: ID tâche, D: Type, E: Acte de traitement,
                            # F: Commentaire, G: Date d'affectation, H: Date de traitement, I: Date de livraison,
                            # J: Collaborateur, K: Durée

                            row_data['Nom commune'] = self._get_column_value(row, columns, ['Nom commune', 'nom_commune'], commune_data['nom_commune'])
                            row_data['Code INSEE'] = self._get_column_value(row, columns, ['Code INSEE', 'insee', 'Insee'], commune_data['insee_code'])
                            row_data['ID tâche'] = self._get_column_value(row, columns, ['ID tâche', 'id_tache'])
                            row_data['Type'] = self._get_column_value(row, columns, ['Type', 'type'])
                            row_data['Acte de traitement'] = self._get_column_value(row, columns, ['Acte de traitement', 'acte_traitement'])
                            row_data['Commentaire'] = self._get_column_value(row, columns, ['Commentaire', 'commentaire'])
                            row_data['Date d\'affectation'] = self._get_column_value(row, columns, ['Date d\'affectation', 'date_affectation'])
                            row_data['Date de traitement'] = self._get_column_value(row, columns, ['Date de traitement', 'date_traitement'])
                            row_data['Date de livraison'] = self._get_column_value(row, columns, ['Date de livraison', 'date_livraison'])
                            row_data['Collaborateur'] = self._get_column_value(row, columns, ['Collaborateur', 'collaborateur'])
                            row_data['Durée'] = self._get_column_value(row, columns, ['Durée', 'duree', 'duration'])

                            aggregated_rows.append(row_data)

                except Exception as e:
                    self.logger.error(f"Error processing RIP sheet for commune {commune_data['nom_commune']}: {e}")
                    continue

            # Create DataFrame
            if aggregated_rows:
                new_df = pd.DataFrame(aggregated_rows)

                # If we have existing data, merge it
                if existing_df is not None and not existing_df.empty:
                    # Format date columns in existing data
                    existing_df = self._format_date_columns(existing_df)

                    # Remove existing entries for communes that are being updated
                    commune_names = [data['nom_commune'] for data in rip_communes]
                    if 'Nom commune' in existing_df.columns:
                        existing_df = existing_df[~existing_df['Nom commune'].isin(commune_names)]

                    # Combine existing and new data
                    combined_df = pd.concat([existing_df, new_df], ignore_index=True, sort=False)
                else:
                    combined_df = new_df

                # Ensure proper column order for RIP data
                column_order = ['Nom commune', 'Code INSEE', 'ID tâche', 'Type', 'Acte de traitement',
                               'Commentaire', 'Date d\'affectation', 'Date de traitement', 'Date de livraison',
                               'Collaborateur', 'Durée']

                # Reorder columns
                existing_columns = [col for col in combined_df.columns if col not in column_order]
                final_columns = [col for col in column_order if col in combined_df.columns]
                final_columns.extend(existing_columns)  # Add any remaining columns

                combined_df = combined_df.reindex(columns=final_columns)

                # Apply date formatting
                combined_df = self._format_date_columns(combined_df)

                # Write to Excel with the correct sheet name
                combined_df.to_excel(writer, sheet_name='Traitement RIP', index=False)

                # Format INSEE columns as text to preserve leading zeros
                self._format_insee_columns_as_text(writer, 'Traitement RIP', combined_df)

                # Format date columns as text to prevent Excel auto-formatting
                self._format_date_columns_as_text(writer, 'Traitement RIP', combined_df)

                self.logger.info(f"Created RIP sheet with {len(aggregated_rows)} rows from {len(rip_communes)} RIP communes")
            else:
                # Create empty DataFrame with headers for RIP data
                empty_df = pd.DataFrame(columns=['Nom commune', 'Code INSEE', 'ID tâche', 'Type', 'Acte de traitement',
                                               'Commentaire', 'Date d\'affectation', 'Date de traitement', 'Date de livraison',
                                               'Collaborateur', 'Durée'])
                empty_df.to_excel(writer, sheet_name='Traitement RIP', index=False)

        except Exception as e:
            self.logger.error(f"Error creating page 4 RIP data: {e}")
            raise

    def _get_column_value(self, row, columns, possible_names, default_value=''):
        """Get value from row using possible column names, with fallback to default."""
        try:
            pd = get_pandas()
            for name in possible_names:
                if name in columns:
                    value = row[name]
                    # Return the value if it's not null/empty, otherwise continue to next possible name
                    if pd.notna(value) and str(value).strip() != '':
                        # Special handling for INSEE codes to preserve leading zeros
                        if any(insee_keyword in name.lower() for insee_keyword in ['insee', 'code insee']):
                            # Format INSEE code to preserve leading zeros (5 digits)
                            insee_str = str(value).strip()
                            if insee_str.isdigit() and len(insee_str) <= 5:
                                return insee_str.zfill(5)  # Pad with leading zeros to 5 digits
                            return insee_str

                        # Special handling for date columns to ensure consistent dd/mm/yyyy format
                        elif any(date_keyword in name.lower() for date_keyword in ['date', 'livraison', 'affectation', 'dépose', 'traitement']):
                            # Skip duration columns
                            if not any(duration_keyword in name.lower() for duration_keyword in ['durée', 'duration', 'temps', 'time', 'optimum']):
                                normalized_value = self._normalize_date_value(value, name)
                                # Debug logging for date conversion
                                if str(value) != normalized_value:
                                    self.logger.debug(f"Date conversion in column '{name}': '{value}' -> '{normalized_value}'")
                                return normalized_value

                        return value

            # If no valid value found in any of the possible columns, return default
            return default_value

        except Exception as e:
            self.logger.warning(f"Error getting column value for {possible_names}: {e}")
            return default_value

    def _format_insee_columns_as_text(self, writer, sheet_name: str, df):
        """Format INSEE columns as text to preserve leading zeros."""
        try:
            # Get the worksheet
            worksheet = writer.sheets[sheet_name]

            # Find INSEE columns
            insee_columns = []
            for i, col_name in enumerate(df.columns):
                if any(insee_keyword in col_name.lower() for insee_keyword in ['insee', 'code insee']):
                    insee_columns.append(i + 1)  # Excel columns are 1-indexed

            # Format INSEE columns as text
            for col_idx in insee_columns:
                col_letter = chr(64 + col_idx)  # Convert to Excel column letter (A, B, C, etc.)

                # Set the entire column format to text
                for row_idx in range(2, len(df) + 2):  # Start from row 2 (skip header)
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        # Ensure the value is formatted as text with leading zeros
                        insee_str = str(cell.value).strip()
                        if insee_str.isdigit() and len(insee_str) <= 5:
                            cell.value = insee_str.zfill(5)  # Pad with leading zeros
                        cell.number_format = '@'  # Text format

            self.logger.info(f"Formatted INSEE columns as text in sheet: {sheet_name}")

        except Exception as e:
            self.logger.warning(f"Error formatting INSEE columns in {sheet_name}: {e}")

    def _format_date_columns_as_text(self, writer, sheet_name: str, df):
        """Format date columns as text to prevent Excel auto-formatting."""
        try:
            # Define date columns for each sheet
            date_columns_map = {
                'Suivi Tickets': {
                    'columns': ['Date d\'affectation', 'Date Livraison', 'Date Dépose Ticket', 'Date traitement', 'Date livraison'],
                    'excel_columns': ['I', 'O', 'R']  # Colonnes I, O, R
                },
                'Traitement CMS Adr': {
                    'columns': ['Date affectation', 'Date traitement', 'Date livraison'],
                    'excel_columns': ['F', 'G', 'H']  # Colonnes F, G, H (décalées à cause de la nouvelle colonne Motif Voie en D)
                },
                'Traitement PA': {
                    'columns': ['Date traitement'],
                    'excel_columns': ['G']  # Colonne G
                }
            }

            # Get the worksheet
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            # Get date column configuration for this sheet
            sheet_config = None
            for sheet_type, config in date_columns_map.items():
                if sheet_type.lower() in sheet_name.lower():
                    sheet_config = config
                    break

            if not sheet_config:
                self.logger.warning(f"No date column configuration found for sheet: {sheet_name}")
                return

            # Find actual date columns in the dataframe
            date_columns_to_format = []

            # Method 1: Use exact column names
            for col_name in sheet_config['columns']:
                if col_name in df.columns:
                    col_idx = df.columns.get_loc(col_name) + 1  # +1 for Excel 1-based indexing
                    date_columns_to_format.append(col_idx)
                    self.logger.info(f"Found date column '{col_name}' at index {col_idx} in {sheet_name}")

            # Method 2: Use Excel column letters if available
            if not date_columns_to_format:
                for excel_col in sheet_config['excel_columns']:
                    col_idx = ord(excel_col.upper()) - ord('A') + 1  # Convert A=1, B=2, etc.
                    if col_idx <= len(df.columns):
                        date_columns_to_format.append(col_idx)
                        self.logger.info(f"Using Excel column {excel_col} (index {col_idx}) as date column in {sheet_name}")

            # Method 3: Search for columns containing date keywords
            if not date_columns_to_format:
                date_keywords = ['date', 'livraison', 'affectation', 'dépose', 'traitement']
                duration_keywords = ['durée', 'duration', 'temps', 'time', 'optimum', 'motif']

                for i, column in enumerate(df.columns):
                    column_lower = str(column).lower()

                    # Skip duration columns
                    is_duration = any(keyword in column_lower for keyword in duration_keywords)
                    if is_duration:
                        continue

                    # Check if it's a date column
                    is_date_column = any(keyword in column_lower for keyword in date_keywords)
                    if is_date_column:
                        col_idx = i + 1  # +1 for Excel 1-based indexing
                        date_columns_to_format.append(col_idx)
                        self.logger.info(f"Found date column by keyword '{column}' at index {col_idx} in {sheet_name}")

            # Apply text formatting to date columns
            if date_columns_to_format:
                self.logger.info(f"Formatting {len(date_columns_to_format)} date columns as text in {sheet_name}: {date_columns_to_format}")

                for col_idx in date_columns_to_format:
                    col_letter = chr(64 + col_idx)  # Convert to Excel column letter (A, B, C, etc.)

                    # Set the entire column format to text
                    for row_idx in range(1, len(df) + 2):  # Include header row
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.number_format = '@'  # Text format

                        # Ensure the value is stored as text
                        if cell.value is not None:
                            cell.value = str(cell.value)

                    self.logger.debug(f"Applied text format to column {col_letter} (index {col_idx}) in {sheet_name}")
            else:
                self.logger.warning(f"No date columns found to format in {sheet_name}")

        except Exception as e:
            self.logger.error(f"Error formatting date columns as text in {sheet_name}: {e}")

    def _format_global_excel(self, file_path: str):
        """Apply formatting to the global Excel file similar to individual suivi files."""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            # Load workbook
            wb = load_workbook(file_path)

            # Format all sheets with consistent styling
            sheets_to_format = ['Suivi Tickets', 'Traitement CMS Adr', 'Traitement PA', 'Traitement RIP']

            for sheet_name in sheets_to_format:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    self._format_sheet(ws, sheet_name)
                    # Apply additional date formatting to ensure consistency
                    self._apply_comprehensive_date_formatting(ws, sheet_name)

            # Save changes
            wb.save(file_path)
            self.logger.info(f"Global Excel file formatted successfully: {file_path}")

        except Exception as e:
            self.logger.error(f"Error formatting Excel file: {e}")
            # Don't raise - formatting is optional

    def _format_sheet(self, ws, sheet_name):
        """Apply consistent formatting to a worksheet."""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment

            # Header formatting (blue background, centered, frozen)
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            center_alignment = Alignment(horizontal='center', vertical='center')

            # Apply header formatting
            if ws.max_row > 0:
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_alignment

                # Freeze header row
                ws.freeze_panes = 'A2'

            # Center all content and apply date formatting
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        cell.alignment = center_alignment

                        # Apply date formatting to date columns
                        if cell.row > 1:  # Skip header row
                            column_letter = cell.column_letter
                            header_cell = ws[f'{column_letter}1']
                            header_value = str(header_cell.value).lower() if header_cell.value else ""

                            # First check if it's a duration column - if so, skip date formatting
                            duration_keywords = ['durée', 'duration', 'temps', 'time', 'traitement optimum', 'finale', 'motif']
                            is_duration = any(keyword in header_value for keyword in duration_keywords)

                            # Only apply date formatting if it's not a duration column
                            if not is_duration:
                                date_keywords = ['date', 'livraison', 'affectation', 'dépose', 'traitement']
                                is_date = any(keyword in header_value for keyword in date_keywords)
                                if is_date and cell.value is not None and str(cell.value).strip() != '':
                                    # Apply YYYY-MM-DD format
                                    cell.number_format = 'YYYY-MM-DD'

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            self.logger.info(f"Formatting applied to sheet: {sheet_name}")

        except Exception as e:
            self.logger.error(f"Error formatting sheet {sheet_name}: {e}")

    def _apply_comprehensive_date_formatting(self, ws, sheet_name):
        """Apply comprehensive date formatting to ensure all date columns use YYYY-MM-DD format."""
        try:
            # Define date column patterns for each sheet type
            date_patterns = {
                'Suivi Tickets': ['date', 'livraison', 'affectation', 'dépose'],
                'Traitement CMS Adr': ['date', 'livraison', 'affectation', 'traitement'],
                'Traitement PA': ['date', 'traitement']
            }

            # Get patterns for this sheet
            patterns = []
            for sheet_type, sheet_patterns in date_patterns.items():
                if sheet_type.lower() in sheet_name.lower():
                    patterns = sheet_patterns
                    break

            # If no specific patterns, use general date patterns
            if not patterns:
                patterns = ['date', 'livraison', 'affectation', 'dépose', 'traitement']

            # Duration columns to exclude
            duration_keywords = ['durée', 'duration', 'temps', 'time', 'traitement optimum', 'finale', 'motif']

            # Check each column
            for col in range(1, ws.max_column + 1):
                header_cell = ws.cell(row=1, column=col)
                if header_cell.value:
                    header_text = str(header_cell.value).lower()

                    # Skip duration columns
                    is_duration = any(keyword in header_text for keyword in duration_keywords)
                    if is_duration:
                        continue

                    # Check if it's a date column
                    is_date = any(pattern in header_text for pattern in patterns)
                    if is_date:
                        # Apply date formatting to all cells in this column
                        for row in range(2, ws.max_row + 1):
                            cell = ws.cell(row=row, column=col)
                            if cell.value is not None and str(cell.value).strip() != '':
                                cell.number_format = 'YYYY-MM-DD'

                        self.logger.debug(f"Applied date formatting to column {header_text} in {sheet_name}")

        except Exception as e:
            self.logger.warning(f"Error applying comprehensive date formatting to {sheet_name}: {e}")

    def _normalize_date_value(self, value, column_name=''):
        """Convert all dates to ISO format YYYY-MM-DD."""
        try:
            pd = get_pandas()

            if pd.isna(value) or value == '' or value is None:
                return ''

            # If it's a datetime object, convert to YYYY-MM-DD string
            if hasattr(value, 'strftime'):
                return value.strftime('%Y-%m-%d')

            # For string values, try to parse and convert to ISO format
            value_str = str(value).strip()

            # Remove any time components if present
            if ' ' in value_str:
                value_str = value_str.split(' ')[0]

            if not value_str:
                return ''

            # Try to parse various date formats and convert to ISO
            try:
                # Try dd/mm/yyyy format first
                if '/' in value_str and len(value_str.split('/')) == 3:
                    parts = value_str.split('/')
                    if len(parts[0]) <= 2 and len(parts[1]) <= 2 and len(parts[2]) == 4:
                        # dd/mm/yyyy format
                        day, month, year = parts
                        iso_date = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                        self.logger.debug(f"Converted date: '{value_str}' -> '{iso_date}' in column '{column_name}'")
                        return iso_date

                # Try dd-mm-yyyy format
                if '-' in value_str and len(value_str.split('-')) == 3:
                    parts = value_str.split('-')
                    if len(parts[0]) <= 2 and len(parts[1]) <= 2 and len(parts[2]) == 4:
                        # dd-mm-yyyy format
                        day, month, year = parts
                        iso_date = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                        self.logger.debug(f"Converted date: '{value_str}' -> '{iso_date}' in column '{column_name}'")
                        return iso_date
                    elif len(parts[0]) == 4 and len(parts[1]) <= 2 and len(parts[2]) <= 2:
                        # Already in YYYY-MM-DD format
                        year, month, day = parts
                        iso_date = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                        self.logger.debug(f"Date already in ISO format: '{value_str}' -> '{iso_date}' in column '{column_name}'")
                        return iso_date

                # Try using pandas to parse the date with dayfirst=True
                parsed_date = pd.to_datetime(value_str, dayfirst=True, errors='coerce')
                if pd.notna(parsed_date):
                    iso_date = parsed_date.strftime('%Y-%m-%d')
                    self.logger.debug(f"Parsed and converted date: '{value_str}' -> '{iso_date}' in column '{column_name}'")
                    return iso_date

            except Exception as e:
                self.logger.warning(f"Could not parse date '{value_str}' in column '{column_name}': {e}")

            # If all parsing fails, return original value
            self.logger.warning(f"Could not convert date '{value_str}' to ISO format in column '{column_name}'")
            return value_str

        except Exception as e:
            self.logger.error(f"Error processing date value '{value}' in column '{column_name}': {e}")
            return str(value) if value is not None else ''

    def _validate_and_format_dates_before_writing(self, df, sheet_name):
        """Final validation and formatting of date columns before writing to Excel."""
        try:
            pd = get_pandas()

            # Define date columns for each sheet type
            date_columns_map = {
                'Suivi Tickets': ['Date d\'affectation', 'Date Livraison', 'Date Dépose Ticket', 'Date traitement', 'Date livraison'],
                'Traitement CMS Adr': ['Date affectation', 'Date traitement', 'Date livraison'],
                'Traitement PA': ['Date traitement']
            }

            # Get relevant date columns for this sheet
            relevant_date_columns = []
            for sheet_type, columns in date_columns_map.items():
                if sheet_type.lower() in sheet_name.lower():
                    relevant_date_columns = columns
                    break

            # Also check for any column containing date keywords
            date_keywords = ['date', 'livraison', 'affectation', 'dépose', 'traitement']
            duration_keywords = ['durée', 'duration', 'temps', 'time', 'optimum', 'motif']

            for column in df.columns:
                column_lower = str(column).lower()

                # Skip duration columns
                is_duration = any(keyword in column_lower for keyword in duration_keywords)
                if is_duration:
                    continue

                # Check if it's a date column (either in the specific list or contains date keywords)
                is_date_column = (column in relevant_date_columns or
                                any(keyword in column_lower for keyword in date_keywords))

                if is_date_column:
                    self.logger.info(f"Validating and formatting date column '{column}' in {sheet_name}")

                    # Process each value in the date column
                    for idx in df.index:
                        value = df.at[idx, column]
                        if pd.notna(value) and str(value).strip() != '':
                            original_value = str(value)
                            normalized_value = self._normalize_date_value(value, column)

                            if normalized_value != original_value:
                                df.at[idx, column] = normalized_value
                                self.logger.debug(f"Final date formatting in {sheet_name}, column '{column}': '{original_value}' -> '{normalized_value}'")

            return df

        except Exception as e:
            self.logger.error(f"Error validating dates before writing {sheet_name}: {e}")
            return df

    def _force_date_formatting_for_pages_2_3(self, df, sheet_name):
        """Aggressively format date columns specifically for pages 2 and 3 to fix inversion issues."""
        try:
            pd = get_pandas()

            # Define the exact problematic columns for each sheet
            problematic_date_columns = {
                'Traitement CMS Adr': ['Date affectation', 'Date traitement', 'Date livraison'],
                'Traitement PA': ['Date traitement']
            }

            # Get the columns to fix for this sheet
            columns_to_fix = []
            for sheet_type, columns in problematic_date_columns.items():
                if sheet_type.lower() in sheet_name.lower():
                    columns_to_fix = columns
                    break

            if not columns_to_fix:
                return df

            self.logger.info(f"Applying aggressive date formatting to {sheet_name} for columns: {columns_to_fix}")

            # Process each problematic column
            for column in columns_to_fix:
                if column in df.columns:
                    self.logger.info(f"Processing column '{column}' in {sheet_name}")

                    for idx in df.index:
                        value = df.at[idx, column]
                        if pd.notna(value) and str(value).strip() != '':
                            original_value = str(value)

                            # Apply multiple strategies to ensure European format
                            normalized_value = self._aggressive_european_date_normalization(value, column)

                            if normalized_value != original_value:
                                df.at[idx, column] = normalized_value
                                self.logger.debug(f"Date formatting applied - Column '{column}': '{original_value}' -> '{normalized_value}'")
                            else:
                                self.logger.debug(f"No change needed for '{original_value}' in column '{column}'")

            return df

        except Exception as e:
            self.logger.error(f"Error in aggressive date formatting for {sheet_name}: {e}")
            return df

    def _aggressive_european_date_normalization(self, value, column_name):
        """Convert dates to ISO format YYYY-MM-DD."""
        try:
            pd = get_pandas()

            if pd.isna(value) or value == '' or value is None:
                return ''

            value_str = str(value).strip()
            if not value_str:
                return ''

            # Handle datetime objects - convert to YYYY-MM-DD string
            if hasattr(value, 'strftime'):
                return value.strftime('%Y-%m-%d')

            # For string values, try to parse and convert to ISO format
            # Remove any time components if present
            if ' ' in value_str:
                value_str = value_str.split(' ')[0]

            # Try to parse various date formats and convert to ISO
            try:
                # Try dd/mm/yyyy format first
                if '/' in value_str and len(value_str.split('/')) == 3:
                    parts = value_str.split('/')
                    if len(parts[0]) <= 2 and len(parts[1]) <= 2 and len(parts[2]) == 4:
                        # dd/mm/yyyy format
                        day, month, year = parts
                        iso_date = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                        self.logger.debug(f"Converted date: '{value_str}' -> '{iso_date}' in column '{column_name}'")
                        return iso_date

                # Try dd-mm-yyyy format
                if '-' in value_str and len(value_str.split('-')) == 3:
                    parts = value_str.split('-')
                    if len(parts[0]) <= 2 and len(parts[1]) <= 2 and len(parts[2]) == 4:
                        # dd-mm-yyyy format
                        day, month, year = parts
                        iso_date = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                        self.logger.debug(f"Converted date: '{value_str}' -> '{iso_date}' in column '{column_name}'")
                        return iso_date
                    elif len(parts[0]) == 4 and len(parts[1]) <= 2 and len(parts[2]) <= 2:
                        # Already in YYYY-MM-DD format
                        year, month, day = parts
                        iso_date = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                        self.logger.debug(f"Date already in ISO format: '{value_str}' -> '{iso_date}' in column '{column_name}'")
                        return iso_date

                # Try using pandas to parse the date with dayfirst=True
                parsed_date = pd.to_datetime(value_str, dayfirst=True, errors='coerce')
                if pd.notna(parsed_date):
                    iso_date = parsed_date.strftime('%Y-%m-%d')
                    self.logger.debug(f"Parsed and converted date: '{value_str}' -> '{iso_date}' in column '{column_name}'")
                    return iso_date

            except Exception as e:
                self.logger.warning(f"Could not parse date '{value_str}' in column '{column_name}': {e}")

            # If all parsing fails, return original value
            self.logger.warning(f"Could not convert date '{value_str}' to ISO format in column '{column_name}'")
            return value_str

        except Exception as e:
            self.logger.error(f"Error processing date '{value}' in column '{column_name}': {e}")
            return str(value) if value is not None else ''



    def _show_file_access_dialog(self, access_result: dict, file_path: str) -> bool:
        """
        Show a user-friendly dialog for file access issues.

        Args:
            access_result: Result from check_file_access()
            file_path: Path to the file

        Returns:
            True if user wants to retry, False otherwise
        """
        filename = os.path.basename(file_path)

        # Create custom dialog
        dialog = tk.Toplevel(self.parent)
        dialog.title("Fichier en cours d'utilisation")
        dialog.geometry("500x400")
        dialog.configure(bg=COLORS['BG'])
        dialog.resizable(False, False)

        # Center the dialog
        dialog.transient(self.parent.winfo_toplevel())
        dialog.grab_set()

        # Main frame
        main_frame = tk.Frame(dialog, bg=COLORS['BG'], padx=20, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Icon and title
        title_frame = tk.Frame(main_frame, bg=COLORS['BG'])
        title_frame.pack(fill=tk.X, pady=(0, 15))

        icon_label = tk.Label(
            title_frame,
            text="🔒",
            font=("Segoe UI", 24),
            bg=COLORS['BG'],
            fg=COLORS['WARNING']
        )
        icon_label.pack(side=tk.LEFT, padx=(0, 10))

        title_label = tk.Label(
            title_frame,
            text="Fichier Excel en cours d'utilisation",
            font=UIConfig.FONT_HEADER,
            bg=COLORS['BG'],
            fg=COLORS['PRIMARY']
        )
        title_label.pack(side=tk.LEFT)

        # Message
        message_text = access_result.get('user_message', f"Le fichier '{filename}' est actuellement utilisé.")
        message_label = tk.Label(
            main_frame,
            text=message_text,
            font=UIConfig.FONT_SUBTITLE,
            bg=COLORS['BG'],
            fg=COLORS['TEXT_SECONDARY'],
            wraplength=450,
            justify=tk.LEFT
        )
        message_label.pack(fill=tk.X, pady=(0, 15))

        # Suggestions
        suggestions = access_result.get('suggestions', [])
        if suggestions:
            suggestions_label = tk.Label(
                main_frame,
                text="💡 Solutions recommandées:",
                font=UIConfig.FONT_SUBTITLE,
                bg=COLORS['BG'],
                fg=COLORS['PRIMARY']
            )
            suggestions_label.pack(anchor=tk.W, pady=(0, 5))

            for i, suggestion in enumerate(suggestions, 1):
                suggestion_label = tk.Label(
                    main_frame,
                    text=f"  {i}. {suggestion}",
                    font=UIConfig.FONT_SMALL,
                    bg=COLORS['BG'],
                    fg=COLORS['TEXT_SECONDARY'],
                    wraplength=450,
                    justify=tk.LEFT
                )
                suggestion_label.pack(anchor=tk.W, pady=1)

        # Additional info
        info_frame = tk.Frame(main_frame, bg=COLORS['LIGHT'], relief=tk.RAISED, bd=1)
        info_frame.pack(fill=tk.X, pady=(15, 0))

        info_label = tk.Label(
            info_frame,
            text="ℹ️ Astuce: Vous pouvez consulter le fichier Excel pendant que l'application fonctionne,\nmais fermez-le avant de faire des mises à jour.",
            font=UIConfig.FONT_SMALL,
            bg=COLORS['LIGHT'],
            fg=COLORS['INFO'],
            justify=tk.CENTER,
            padx=10,
            pady=8
        )
        info_label.pack()

        # Buttons
        button_frame = tk.Frame(main_frame, bg=COLORS['BG'])
        button_frame.pack(fill=tk.X, pady=(20, 0))

        result = {'retry': False}

        def on_retry():
            result['retry'] = True
            dialog.destroy()

        def on_cancel():
            result['retry'] = False
            dialog.destroy()

        # Retry button
        retry_btn = tk.Button(
            button_frame,
            text="🔄 Réessayer",
            command=on_retry,
            bg=COLORS['PRIMARY'],
            fg='white',
            font=UIConfig.FONT_BUTTON,
            relief='flat',
            padx=20,
            pady=8
        )
        retry_btn.pack(side=tk.RIGHT, padx=(10, 0))

        # Cancel button
        cancel_btn = tk.Button(
            button_frame,
            text="❌ Annuler",
            command=on_cancel,
            bg=COLORS['BORDER'],
            fg=COLORS['TEXT_SECONDARY'],
            font=UIConfig.FONT_BUTTON,
            relief='flat',
            padx=20,
            pady=8
        )
        cancel_btn.pack(side=tk.RIGHT)

        # Wait for dialog to close
        dialog.wait_window()

        return result['retry']

    def _check_global_file_status(self):
        """Check if the global Excel file is accessible and update UI accordingly."""
        try:
            file_path = os.path.join(self.teams_folder_path, self.global_excel_filename)

            if not os.path.exists(file_path):
                return "not_exists"

            if is_excel_file_open(file_path):
                return "in_use"

            access_result = check_file_access(file_path, 'rw')
            if access_result['accessible']:
                return "available"
            else:
                return "locked"

        except Exception as e:
            self.logger.error(f"Error checking global file status: {e}")
            return "unknown"

    def _update_file_status_indicator(self):
        """Update the file status indicator in the UI."""
        try:
            status = self._check_global_file_status()

            if hasattr(self, 'file_status_label'):
                status_config = {
                    'not_exists': {
                        'text': 'Nouveau fichier Excel global - sera créé',
                        'color': COLORS['INFO'],
                        'icon': '📄'
                    },
                    'available': {
                        'text': 'Fichier Excel global existant - sera mis à jour',
                        'color': COLORS['SUCCESS'],
                        'icon': '✅'
                    },
                    'in_use': {
                        'text': '⚠️ Fichier Excel ouvert - fermez-le avant de continuer',
                        'color': COLORS['WARNING'],
                        'icon': '🔒'
                    },
                    'locked': {
                        'text': '⚠️ Fichier verrouillé - impossible d\'accéder',
                        'color': COLORS['DANGER'],
                        'icon': '🔒'
                    },
                    'unknown': {
                        'text': 'Statut du fichier inconnu',
                        'color': COLORS['INFO'],
                        'icon': '❓'
                    }
                }

                config = status_config.get(status, status_config['unknown'])
                self.file_status_label.config(text=config['text'], fg=config['color'])

                # Update icon if available
                if hasattr(self, 'file_status_icon'):
                    self.file_status_icon.config(text=config['icon'], fg=config['color'])

                # Update generate button state based on file status
                if hasattr(self, 'generate_button'):
                    if status == 'in_use' or status == 'locked':
                        self.generate_button.config(
                            text="⚠️ Fermer Excel d'abord",
                            bg=COLORS['WARNING']
                        )
                    elif hasattr(self, 'processed_data') and self.processed_data:
                        self.generate_button.config(
                            text="📊 Générer Excel Global",
                            bg=COLORS['PRIMARY']
                        )

        except Exception as e:
            self.logger.error(f"Error updating file status indicator: {e}")

    def _open_generated_file(self, file_path: str):
        """Open the generated Excel file."""
        try:
            import subprocess
            import platform

            if platform.system() == 'Windows':
                os.startfile(file_path)
            elif platform.system() == 'Darwin':  # macOS
                subprocess.call(['open', file_path])
            else:  # Linux
                subprocess.call(['xdg-open', file_path])

        except Exception as e:
            self.logger.error(f"Error opening file: {e}")
            messagebox.showerror("Erreur", f"Impossible d'ouvrir le fichier:\n{e}")



    # Keyboard shortcut methods
    def _scan_shortcut(self):
        """Keyboard shortcut for scanning folders."""
        if self.scan_button['state'] == tk.NORMAL:
            self._scan_and_process_folders()

    def _generate_shortcut(self):
        """Keyboard shortcut for generating Excel."""
        if self.generate_button['state'] == tk.NORMAL:
            self._update_global_excel()

    def _refresh_shortcut(self):
        """Keyboard shortcut for refreshing."""
        self._reset_module()
