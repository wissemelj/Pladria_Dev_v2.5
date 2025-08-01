"""
Excel file generation module.
Handles creation and formatting of Excel output files.
"""

import logging
from datetime import datetime
from typing import Dict, Any, Optional
import sys
from pathlib import Path

# Ensure src directory is in path
src_path = Path(__file__).parent.parent
if str(src_path) not in sys.path:
    sys.path.insert(0, str(src_path))

from utils.lazy_imports import get_pandas
from config.constants import VALIDATION_LISTS


class ExcelGenerator:
    """Handles Excel file generation and formatting."""
    
    def __init__(self):
        """Initialize the Excel generator."""
        self.logger = logging.getLogger(__name__)
    
    def generate_excel_file(self,
                          moai_data: Dict[str, Any],
                          plan_df: 'pd.DataFrame',
                          project_info: Dict[str, str],
                          output_path: str) -> bool:
        """
        Generate the complete Excel file with all sheets.

        Args:
            moai_data: Processed MOAI data
            plan_df: Processed Plan Adressage DataFrame
            project_info: Project information (domaine, commune, insee, id_tache)
            output_path: Path to save the Excel file

        Returns:
            True if successful, False otherwise
        """
        try:
            pd = get_pandas()

            # Ensure output directory exists
            from utils.file_utils import ensure_directory_exists
            import os
            output_dir = os.path.dirname(output_path)
            if output_dir and not ensure_directory_exists(output_dir):
                raise Exception(f"Cannot create output directory: {output_dir}")

            # Create the three main DataFrames
            df_cm = self._create_cm_adresse_sheet(moai_data, project_info)
            df_plan = self._create_plan_adressage_sheet(plan_df, project_info)
            df_commune = self._create_commune_info_sheet(project_info, df_cm)

            # Check if we need to create RIP sheet (4th page) for RIP commune type
            is_rip_commune = project_info.get('domaine', '').upper() == 'RIP'
            df_rip = None
            if is_rip_commune:
                df_rip = self._create_rip_sheet(project_info)

            # Generate sheet names
            sheet_names = self._generate_sheet_names(project_info['id_tache'], is_rip_commune)

            # Write to Excel with formatting
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Write the main sheets
                df_cm.to_excel(writer, sheet_name=sheet_names['cm'], index=False)
                df_plan.to_excel(writer, sheet_name=sheet_names['plan'], index=False)
                df_commune.to_excel(writer, sheet_name=sheet_names['commune'], index=False)

                # Write RIP sheet if applicable
                if is_rip_commune and df_rip is not None:
                    df_rip.to_excel(writer, sheet_name=sheet_names['rip'], index=False)

                # Apply styling to all sheets
                self._apply_sheet_styling(writer, sheet_names['cm'])
                self._apply_sheet_styling(writer, sheet_names['plan'])
                self._apply_sheet_styling(writer, sheet_names['commune'])

                # Apply styling to RIP sheet if applicable
                if is_rip_commune and 'rip' in sheet_names:
                    self._apply_sheet_styling(writer, sheet_names['rip'])

                # Apply special styling to Plan Adressage sheet (page 2)
                self._apply_plan_adressage_special_styling(writer, sheet_names['plan'], df_plan)

                # Add data validations
                self._add_data_validations(writer, df_cm, sheet_names['cm'])
                self._create_validation_sheet(writer)

                # Pass RIP sheet name to duration formula if RIP commune
                rip_sheet_name = sheet_names.get('rip') if is_rip_commune else None
                self._add_duration_formula(writer, len(df_cm), df_plan,
                                         sheet_names['cm'], sheet_names['plan'], sheet_names['commune'], rip_sheet_name)
                self._add_commune_validations(writer, sheet_names['commune'])
                self._add_plan_adressage_validations(writer, df_plan, sheet_names['plan'])

                # Add RIP sheet validations if applicable
                if is_rip_commune and 'rip' in sheet_names:
                    self._add_rip_validations(writer, sheet_names['rip'])

            self.logger.info(f"Excel file generated successfully: {output_path}")
            return True

        except Exception as e:
            self.logger.error(f"Error generating Excel file: {e}")
            return False
    
    def _create_cm_adresse_sheet(self, moai_data: Dict[str, Any], project_info: Dict[str, str]) -> 'pd.DataFrame':
        """
        Create the CM Adresse sheet DataFrame.

        Args:
            moai_data: Processed MOAI data
            project_info: Project information

        Returns:
            CM Adresse DataFrame
        """
        pd = get_pandas()

        date_affectation = datetime.now().strftime("%Y-%m-%d")
        num_rows = moai_data['num_rows']

        # Use communes from MOAI if available, otherwise use project commune
        communes = (moai_data['communes_moai'] if len(moai_data['communes_moai']) > 0
                   else [project_info['nom_commune']] * num_rows)

        df_cm = pd.DataFrame({
            'Nom commune': [project_info['nom_commune']] * num_rows,  # Colonne A
            'Insee': [project_info['insee']] * num_rows,              # Colonne B
            'ID Tache': moai_data['id_taches'],                       # Colonne C
            'Voie demandé': moai_data['voies_demandees'],             # Colonne D
            'Motif Voie': [''] * num_rows,                            # Colonne E
            'CODE RIVOLI': [''] * num_rows,                           # Colonne F
            'GPS (X,Y)': [''] * num_rows,                             # Colonne G
            'Centre/Zone': [''] * num_rows,                           # Colonne H
            'Status PC': [''] * num_rows,                             # Colonne I
            'Descriptif Commentaire': [''] * num_rows,                # Colonne J
            'Collaborateur': [''] * num_rows,                         # Colonne K
            'Date affectation': [date_affectation] * num_rows,        # Colonne L
            'Date traitement': [''] * num_rows,                       # Colonne M
            'Date livraison': [''] * num_rows,                        # Colonne N
            'Durée': [''] * num_rows,                                 # Colonne O
            'STATUT Ticket': [''] * num_rows                          # Colonne P
        })

        return df_cm

    def _create_plan_adressage_sheet(self, plan_df: 'pd.DataFrame', project_info: Dict[str, str]) -> 'pd.DataFrame':
        """
        Create the Plan Adressage sheet DataFrame with additional commune info columns.

        Args:
            plan_df: Original Plan Adressage DataFrame
            project_info: Project information

        Returns:
            Modified Plan Adressage DataFrame
        """
        pd = get_pandas()
        from datetime import datetime

        # Create a copy of the original DataFrame
        df_plan = plan_df.copy()

        # Add commune info columns at the beginning
        num_rows = len(df_plan)
        df_plan.insert(0, 'Insee', [project_info['insee']] * num_rows)
        df_plan.insert(0, 'Nom commune', [project_info['nom_commune']] * num_rows)

        # Auto-populate Date traitement and Durée for specific motifs
        # Motifs that should get automatic values (case-insensitive matching)
        auto_motifs_patterns = [
            "ad non trouvee", "ad non trouvée",  # Variations de "Ad Non Trouvée"
            "sans geometrie", "sans géométrie",   # Variations de "Sans Géométrie"
            "ad non joint", "ad non jointe"       # Variations de "Ad Non Jointe"
        ]
        current_date = datetime.now().strftime("%d/%m/%Y")

        # Check if we have the required columns
        if 'Date traitement' in df_plan.columns and 'Durée' in df_plan.columns and 'Motif' in df_plan.columns:
            for index, row in df_plan.iterrows():
                motif_original = str(row['Motif']).strip() if pd.notna(row['Motif']) else ''
                motif_normalized = motif_original.lower().replace('é', 'e').replace('è', 'e')

                # Check if motif matches any of the auto-motifs patterns (case-insensitive)
                should_auto_populate = False
                for pattern in auto_motifs_patterns:
                    pattern_normalized = pattern.lower().replace('é', 'e').replace('è', 'e')
                    if pattern_normalized in motif_normalized or motif_normalized in pattern_normalized:
                        should_auto_populate = True
                        break

                if should_auto_populate:
                    df_plan.at[index, 'Date traitement'] = current_date
                    df_plan.at[index, 'Durée'] = 0
                    self.logger.info(f"Auto-populated row {index + 1}: Motif '{motif_original}' -> Date: {current_date}, Durée: 0")

        return df_plan

    def _create_commune_info_sheet(self, project_info: Dict[str, str], df_cm: 'pd.DataFrame') -> 'pd.DataFrame':
        """
        Create the commune information sheet DataFrame.

        Args:
            project_info: Project information
            df_cm: CM Adresse DataFrame to extract data from

        Returns:
            Commune information DataFrame
        """
        pd = get_pandas()

        date_affectation = datetime.now().strftime("%Y-%m-%d")

        df_commune = pd.DataFrame({
            'Nom de commune': [project_info['nom_commune']],           # Colonne A
            'ID tâche Plan Adressage': [project_info['id_tache']],     # Colonne B
            'Code INSEE': [project_info['insee']],                     # Colonne C
            'Domaine': [project_info.get('domaine', '')],             # Colonne D - Rempli avec la valeur sélectionnée
            'Type de Commune': [''],                                  # Colonne E - Nouvelle colonne
            'Type de base': [''],                                     # Colonne F - Nouvelle colonne
            'Nbr des voies CM': [''],                                 # Colonne G - Will be calculated automatically
            'Nbr des IMB PA': [''],                                   # Colonne H - Will be calculated automatically
            'Date d\'affectation': [date_affectation],                # Colonne I
            'Temps préparation QGis': [0],                            # Colonne J - Nouvelle colonne
            'Durée Totale CM': [5],                                   # Colonne K - Initialized to 5 min
            'Duréé Totale PA': [0],                                   # Colonne L - Initialized to 0
            'Traitement Optimum': [0],                                # Colonne M - Nouvelle colonne
            'Durée Finale': [''],                                     # Colonne N - Will be calculated with Excel formula
            'Date Livraison': [''],                                   # Colonne O
            'Etat Ticket PA ': [''],                                  # Colonne P
            'ID Tache 501/511': [''],                                 # Colonne Q
            'Date Dépose Ticket 501/511': [''],                       # Colonne R
            'Dépose Ticket UPR': ['Non Créé'],                       # Colonne S
            'ID tâche UPR': [''],                                     # Colonne T
            'Collaborateur': ['']                                     # Colonne U
        })

        return df_commune

    def _create_rip_sheet(self, project_info: Dict[str, str]) -> 'pd.DataFrame':
        """
        Create the RIP sheet DataFrame for RIP commune type.
        This 4th page contains specific fields for RIP communes.

        Args:
            project_info: Project information

        Returns:
            RIP DataFrame
        """
        pd = get_pandas()

        df_rip = pd.DataFrame({
            'Nom commune': [project_info['nom_commune']],              # Column A: Commune name
            'Code INSEE': [project_info['insee']],                     # Column B: INSEE code
            'ID tâche': [''],                                          # Column C: Task ID (text input)
            'Type': [''],                                              # Column D: Type (P0/P1 dropdown)
            'Acte de traitement': [''],                                # Column E: Processing action (Motif Voie validation)
            'Commentaire': [''],                                       # Column F: Comment (free text input)
            'Date d\'affectation': [''],                               # Column G: Assignment date (no auto-fill)
            'Date de traitement': [''],                                # Column H: Processing date
            'Date de livraison': [''],                                 # Column I: Delivery date
            'Collaborateur': [''],                                     # Column J: Collaborator/Employee
            'Durée': ['']                                              # Column K: Duration
        })

        return df_rip

    def _apply_sheet_styling(self, writer, sheet_name: str):
        """
        Apply default styling to a sheet: center alignment, freeze first row, blue header.

        Args:
            writer: Excel writer object
            sheet_name: Name of the sheet to style
        """
        try:
            from openpyxl.styles import Alignment, PatternFill, Font

            worksheet = writer.sheets[sheet_name]

            # Freeze the first row
            worksheet.freeze_panes = 'A2'

            # Define styles
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')  # Blue froid
            header_font = Font(color='FFFFFF', bold=True)  # White text, bold

            # Apply center alignment to all cells
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = center_alignment

            # Apply header styling to first row
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment

            # Apply date formatting to date columns
            self._apply_date_formatting(worksheet, sheet_name)

            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)  # Max width of 30
                worksheet.column_dimensions[column_letter].width = adjusted_width

            self.logger.info(f"Styling applied to sheet: {sheet_name}")

        except Exception as e:
            self.logger.error(f"Error applying styling to sheet {sheet_name}: {e}")

    def _apply_plan_adressage_special_styling(self, writer, sheet_name: str, plan_df: 'pd.DataFrame'):
        """
        Apply special styling to Plan Adressage sheet (page 2):
        - Num Dossier Site (colonne C): doublons en remplissage bleu ciel froid
        - Même Adresse (colonne H): 'oui' en vert froid, 'non' en rose
        - Adresse BAN (colonne M): doublons en remplissage rose
        """
        try:
            from openpyxl.styles import PatternFill

            worksheet = writer.sheets[sheet_name]
            columns = plan_df.columns.tolist()

            # Define colors
            light_blue_fill = PatternFill(start_color='B3E5FC', end_color='B3E5FC', fill_type='solid')  # Bleu ciel froid
            light_green_fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')  # Vert froid
            light_pink_fill = PatternFill(start_color='F8BBD9', end_color='F8BBD9', fill_type='solid')  # Rose

            # Find column indices (accounting for added commune columns at beginning)
            num_dossier_col = None
            meme_adresse_col = None
            adresse_ban_col = None

            # Look for columns by name
            for i, col_name in enumerate(columns):
                col_name_lower = str(col_name).lower()
                if 'num' in col_name_lower and 'dossier' in col_name_lower and 'site' in col_name_lower:
                    num_dossier_col = i + 1  # Excel is 1-indexed
                elif 'même' in col_name_lower and 'adresse' in col_name_lower:
                    meme_adresse_col = i + 1
                elif 'adresse' in col_name_lower and 'ban' in col_name_lower:
                    adresse_ban_col = i + 1

            # If not found by name, use position-based approach (C, H, M after adding commune columns)
            if num_dossier_col is None:
                num_dossier_col = 3  # Column C
            if meme_adresse_col is None:
                meme_adresse_col = 8  # Column H
            if adresse_ban_col is None:
                adresse_ban_col = 13  # Column M

            # Apply styling for Num Dossier Site (Column C) - highlight duplicates in light blue
            if num_dossier_col and num_dossier_col <= len(columns):
                self._highlight_duplicates(worksheet, num_dossier_col, light_blue_fill)

            # Apply styling for Même Adresse (Column H) - 'oui' in green, 'non' in pink
            if meme_adresse_col and meme_adresse_col <= len(columns):
                self._highlight_oui_non_values(worksheet, meme_adresse_col, light_green_fill, light_pink_fill)

            # Apply styling for Adresse BAN (Column M) - highlight duplicates in pink
            if adresse_ban_col and adresse_ban_col <= len(columns):
                self._highlight_duplicates(worksheet, adresse_ban_col, light_pink_fill)

            self.logger.info(f"Special Plan Adressage styling applied to sheet: {sheet_name}")

        except Exception as e:
            self.logger.error(f"Error applying special Plan Adressage styling to sheet {sheet_name}: {e}")

    def _highlight_duplicates(self, worksheet, col_num: int, fill_color):
        """Highlight duplicate values in a column with the specified fill color."""
        try:
            # Get all values in the column (skip header)
            values = []
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=col_num)
                if cell.value is not None and str(cell.value).strip() != '':
                    values.append((row, str(cell.value).strip()))

            # Find duplicates
            value_counts = {}
            for row, value in values:
                if value in value_counts:
                    value_counts[value].append(row)
                else:
                    value_counts[value] = [row]

            # Highlight cells with duplicate values
            for value, rows in value_counts.items():
                if len(rows) > 1:  # More than one occurrence = duplicate
                    for row in rows:
                        cell = worksheet.cell(row=row, column=col_num)
                        cell.fill = fill_color

        except Exception as e:
            self.logger.error(f"Error highlighting duplicates in column {col_num}: {e}")

    def _highlight_oui_non_values(self, worksheet, col_num: int, oui_fill, non_fill):
        """Highlight 'oui' values with one color and 'non' values with another."""
        try:
            # Check all values in the column (skip header)
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=col_num)
                if cell.value is not None:
                    value = str(cell.value).strip().lower()
                    if value == 'oui':
                        cell.fill = oui_fill
                    elif value == 'non':
                        cell.fill = non_fill

        except Exception as e:
            self.logger.error(f"Error highlighting oui/non values in column {col_num}: {e}")

    def _apply_date_formatting(self, worksheet, sheet_name: str):
        """Apply date formatting to date columns (format as date only, no time)."""
        try:
            # Define specific date columns for each sheet type (by position)
            date_columns = {
                'CM Adresse': ['L', 'M', 'N'],  # Date affectation, Date traitement, Date livraison
                'Plan Adressage': [],  # No date columns in Plan Adressage sheet
                'Suivi Tickets': ['I', 'O', 'R'],  # Date d'affectation, Date Livraison, Date Dépose Ticket 501/511
                'RIP': ['G', 'H', 'I']  # Date d'affectation, Date de traitement, Date de livraison
            }

            # Get the appropriate date columns for this sheet
            sheet_date_columns = []
            for sheet_type, columns in date_columns.items():
                if sheet_type in sheet_name:
                    sheet_date_columns = columns
                    break

            # If no predefined columns, use intelligent detection based on header names
            if not sheet_date_columns:
                sheet_date_columns = self._detect_date_columns_by_header(worksheet)

            # Apply date formatting to identified columns
            if sheet_date_columns:
                for col_letter in sheet_date_columns:
                    # Apply to all rows in the column (starting from row 2 to skip header)
                    for row in range(2, worksheet.max_row + 1):
                        cell = worksheet[f'{col_letter}{row}']
                        if cell.value is not None and str(cell.value).strip() != '':
                            # Set date format (YYYY-MM-DD)
                            cell.number_format = 'YYYY-MM-DD'

                self.logger.info(f"Date formatting applied to columns {sheet_date_columns} in sheet: {sheet_name}")

        except Exception as e:
            self.logger.error(f"Error applying date formatting to sheet {sheet_name}: {e}")

    def _detect_date_columns_by_header(self, worksheet):
        """Detect date columns by analyzing header names, excluding duration columns."""
        date_columns = []

        try:
            # Check each column header
            for col in range(1, worksheet.max_column + 1):
                header_cell = worksheet.cell(row=1, column=col)
                if header_cell.value:
                    header_text = str(header_cell.value).lower()

                    # Exclude duration/time columns first
                    duration_keywords = ['durée', 'duration', 'temps', 'time', 'traitement optimum', 'finale', 'motif']
                    is_duration = any(keyword in header_text for keyword in duration_keywords)

                    # Check for date keywords only if it's not a duration column
                    if not is_duration:
                        date_keywords = ['date', 'livraison', 'affectation', 'dépose', 'traitement']
                        is_date = any(keyword in header_text for keyword in date_keywords)

                        if is_date:
                            col_letter = self._get_column_letter(col)
                            date_columns.append(col_letter)

            return date_columns

        except Exception as e:
            self.logger.error(f"Error detecting date columns: {e}")
            return []

    def _generate_sheet_names(self, id_tache: str, include_rip: bool = False) -> Dict[str, str]:
        """
        Generate sheet names based on task ID.

        Args:
            id_tache: Task ID
            include_rip: Whether to include RIP sheet name

        Returns:
            Dictionary with sheet names
        """
        id_tache_clean = id_tache.replace(" ", "").replace("/", "-").replace("\\", "-")

        sheet_names = {
            'cm': f"{id_tache_clean}-CM Adresse",
            'plan': f"{id_tache_clean}-Plan Adressage",
            'commune': f"{id_tache_clean}-Informations Commune"
        }

        # Add RIP sheet if needed
        if include_rip:
            sheet_names['rip'] = f"{id_tache_clean}-RIP"

        return sheet_names
    
    def _add_data_validations(self, writer, df_cm: 'pd.DataFrame', sheet_name: str):
        """Add data validation lists to CM Adresse sheet."""
        try:
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            from openpyxl.worksheet.datavalidation import DataValidation

            # Get column indices for validation
            columns = df_cm.columns.tolist()

            # Add validations for specific columns (same mappings, positions will be automatically adjusted)
            validation_mappings = {
                'Domaine': 'Domaine',
                'Type de Commune': 'Type de Commune',
                'Type de base': 'Type de base',
                'Motif Voie': 'Motif Voie',
                'Status PC': 'PC Status',
                'Descriptif Commentaire': 'XY Status',
                'Collaborateur': 'Collaborateur',
                'STATUT Ticket': 'STATUT Ticket'
            }

            for col_name, validation_key in validation_mappings.items():
                if col_name in columns:
                    col_idx = columns.index(col_name) + 1  # Excel is 1-indexed
                    col_letter = self._get_column_letter(col_idx)

                    # Create validation list using reference to validation sheet
                    validation_list = VALIDATION_LISTS.get(validation_key, [])
                    if validation_list:
                        # Use reference to validation sheet instead of direct formula
                        # This prevents Excel formula length issues
                        list_length = len(validation_list)
                        validation_range = f"ValidationLists!${self._get_validation_column_letter(validation_key)}$2:${self._get_validation_column_letter(validation_key)}${list_length + 1}"

                        dv = DataValidation(
                            type="list",
                            formula1=validation_range,
                            allow_blank=True
                        )
                        dv.error = f'Valeur non valide pour {col_name}'
                        dv.errorTitle = 'Erreur de validation'

                        # Apply to all data rows (skip header)
                        range_str = f"{col_letter}2:{col_letter}{len(df_cm) + 1}"
                        dv.add(range_str)
                        worksheet.add_data_validation(dv)

            self.logger.info("Data validations added to CM Adresse sheet")

        except Exception as e:
            self.logger.error(f"Error adding data validations: {e}")
    
    def _create_validation_sheet(self, writer):
        """Create a hidden sheet with validation lists."""
        try:
            pd = get_pandas()
            
            # Create DataFrame with all validation lists
            max_length = max(len(lst) for lst in VALIDATION_LISTS.values())
            validation_data = {}
            
            for key, values in VALIDATION_LISTS.items():
                # Pad shorter lists with empty strings
                padded_values = values + [''] * (max_length - len(values))
                validation_data[key] = padded_values
            
            df_validation = pd.DataFrame(validation_data)
            df_validation.to_excel(writer, sheet_name='ValidationLists', index=False)
            
            # Hide the validation sheet
            if 'ValidationLists' in writer.sheets:
                writer.sheets['ValidationLists'].sheet_state = 'hidden'
            
            self.logger.info("Validation sheet created")
            
        except Exception as e:
            self.logger.error(f"Error creating validation sheet: {e}")
    
    def _get_column_letter(self, col_num: int) -> str:
        """Convert column number to Excel column letter."""
        result = ""
        while col_num > 0:
            col_num -= 1
            result = chr(col_num % 26 + ord('A')) + result
            col_num //= 26
        return result

    def _get_validation_column_letter(self, validation_key: str) -> str:
        """Get the column letter for a validation key in the ValidationLists sheet."""
        # Map validation keys to their column positions in the validation sheet
        validation_columns = list(VALIDATION_LISTS.keys())
        if validation_key in validation_columns:
            col_index = validation_columns.index(validation_key) + 1
            return self._get_column_letter(col_index)
        return 'A'  # Default to column A if not found

    def _add_duration_formula(self, writer, cm_rows: int, plan_df: 'pd.DataFrame',
                            cm_sheet: str, plan_sheet: str, commune_sheet: str, rip_sheet: str = None):
        """Add duration calculation formulas."""
        try:
            workbook = writer.book
            commune_ws = writer.sheets[commune_sheet]

            # Add formula to calculate total CM duration (sum of CM sheet duration column)
            # 'Durée' is now column O (15th column) in CM sheet after removing columns
            cm_duration_formula = f"=SUM('{cm_sheet}'!O2:O{cm_rows + 1})+5"
            commune_ws['K2'] = cm_duration_formula  # Durée Totale CM (column K)

            # Add formula to count unique voies in CM sheet (count non-empty rows)
            voies_count_formula = f"=COUNTA('{cm_sheet}'!D2:D{cm_rows + 1})"  # Voie demandé column (now column D)
            commune_ws['G2'] = voies_count_formula  # Nbr des voies CM (column G)

            # Add formula to count unique IMB in Plan Adressage (excluding empty and duplicates)
            plan_rows = len(plan_df)
            if plan_rows > 0:
                # Count unique non-empty values in column C of Plan Adressage sheet (IMB codes like IMB/87193/X/0015)
                # This formula counts each unique IMB only once, excluding duplicates and empty cells
                # Example: 52 IMB with 2 duplicates = 50 unique IMB
                imb_count_formula = f"=SUMPRODUCT(('{plan_sheet}'!C2:C{plan_rows + 1}<>\"\")*1/COUNTIF('{plan_sheet}'!C2:C{plan_rows + 1},'{plan_sheet}'!C2:C{plan_rows + 1}&\"\"))"
                commune_ws['H2'] = imb_count_formula  # Nbr des IMB PA (column H)

            # Add formula to calculate total PA duration (sum of PA sheet duration column)
            if plan_rows > 0:
                # Find the 'Durée' column in plan sheet (should be the last column after adding commune/insee)
                plan_columns = plan_df.columns.tolist()
                if 'Durée' in plan_columns:
                    duree_col_idx = plan_columns.index('Durée') +1
                    duree_col_letter = self._get_column_letter(duree_col_idx)
                    pa_duration_formula = f"=SUM('{plan_sheet}'!{duree_col_letter}2:{duree_col_letter}{plan_rows + 1})"
                    commune_ws['L2'] = pa_duration_formula  # Durée Totale PA (column L)

            # Add formula for final duration including RIP sheet duration if applicable
            if rip_sheet:
                # Include RIP sheet duration sum in final calculation
                # RIP duration is in column K, sum all rows from K2 to K1000 (dynamic range for user additions)
                commune_ws['N2'] = f"=J2+K2+L2+M2+SUM('{rip_sheet}'!K:K)"  # Durée Finale includes RIP duration sum
            else:
                # Original formula without RIP sheet
                commune_ws['N2'] = "=J2+K2+L2+M2"  # Durée Finale = Temps préparation QGis + Durée Totale CM + Durée Totale PA + Traitement Optimum

            self.logger.info("Duration formulas added")

        except Exception as e:
            self.logger.error(f"Error adding duration formulas: {e}")

    def _add_commune_validations(self, writer, sheet_name: str):
        """Add data validations to commune information sheet."""
        try:
            from openpyxl.worksheet.datavalidation import DataValidation

            worksheet = writer.sheets[sheet_name]

            # Add validation for 'Domaine' (column D)
            domaine_list_length = len(VALIDATION_LISTS["Domaine"])
            domaine_range = f"ValidationLists!${self._get_validation_column_letter('Domaine')}$2:${self._get_validation_column_letter('Domaine')}${domaine_list_length + 1}"
            domaine_validation = DataValidation(
                type="list",
                formula1=domaine_range,
                allow_blank=True
            )
            domaine_validation.add('D2')
            worksheet.add_data_validation(domaine_validation)

            # Add validation for 'Type de Commune' (column E)
            type_commune_list_length = len(VALIDATION_LISTS["Type de Commune"])
            type_commune_range = f"ValidationLists!${self._get_validation_column_letter('Type de Commune')}$2:${self._get_validation_column_letter('Type de Commune')}${type_commune_list_length + 1}"
            type_commune_validation = DataValidation(
                type="list",
                formula1=type_commune_range,
                allow_blank=True
            )
            type_commune_validation.add('E2')
            worksheet.add_data_validation(type_commune_validation)

            # Add validation for 'Type de base' (column F)
            type_base_list_length = len(VALIDATION_LISTS["Type de base"])
            type_base_range = f"ValidationLists!${self._get_validation_column_letter('Type de base')}$2:${self._get_validation_column_letter('Type de base')}${type_base_list_length + 1}"
            type_base_validation = DataValidation(
                type="list",
                formula1=type_base_range,
                allow_blank=True
            )
            type_base_validation.add('F2')
            worksheet.add_data_validation(type_base_validation)

            # Add validation for 'Etat Ticket PA' (now column P)
            etat_list_length = len(VALIDATION_LISTS["Etat"])
            etat_range = f"ValidationLists!${self._get_validation_column_letter('Etat')}$2:${self._get_validation_column_letter('Etat')}${etat_list_length + 1}"
            etat_validation = DataValidation(
                type="list",
                formula1=etat_range,
                allow_blank=True
            )
            etat_validation.add('P2')
            worksheet.add_data_validation(etat_validation)

            # Add validation for 'Dépose Ticket UPR' (now column S)
            upr_list_length = len(VALIDATION_LISTS["Depose Ticket UPR"])
            upr_range = f"ValidationLists!${self._get_validation_column_letter('Depose Ticket UPR')}$2:${self._get_validation_column_letter('Depose Ticket UPR')}${upr_list_length + 1}"
            upr_validation = DataValidation(
                type="list",
                formula1=upr_range,
                allow_blank=True
            )
            upr_validation.add('S2')
            worksheet.add_data_validation(upr_validation)

            # Add validation for 'Collaborateur' (now column U)
            collab_list_length = len(VALIDATION_LISTS["Collaborateur"])
            collab_range = f"ValidationLists!${self._get_validation_column_letter('Collaborateur')}$2:${self._get_validation_column_letter('Collaborateur')}${collab_list_length + 1}"
            collab_validation = DataValidation(
                type="list",
                formula1=collab_range,
                allow_blank=True
            )
            collab_validation.add('U2')
            worksheet.add_data_validation(collab_validation)

            self.logger.info("Commune validations added")

        except Exception as e:
            self.logger.error(f"Error adding commune validations: {e}")

    def _add_plan_adressage_validations(self, writer, plan_df: 'pd.DataFrame', sheet_name: str):
        """Add data validations to Plan Adressage sheet."""
        try:
            from openpyxl.worksheet.datavalidation import DataValidation

            worksheet = writer.sheets[sheet_name]
            columns = plan_df.columns.tolist()

            # Add validation for 'Motif' column (column I) if it exists
            if len(columns) >= 9:  # Ensure we have at least 9 columns (A-I)
                # Column I is index 8 (0-based), so column letter is I
                motif_col_letter = 'I'

                # Use reference to validation sheet for Motif
                motif_list_length = len(VALIDATION_LISTS["Motif"])
                motif_range = f"ValidationLists!${self._get_validation_column_letter('Motif')}$2:${self._get_validation_column_letter('Motif')}${motif_list_length + 1}"

                motif_validation = DataValidation(
                    type="list",
                    formula1=motif_range,
                    allow_blank=True
                )
                motif_validation.error = 'Valeur non valide pour Motif'
                motif_validation.errorTitle = 'Erreur de validation'

                # Apply to all data rows in column I
                range_str = f"{motif_col_letter}2:{motif_col_letter}{len(plan_df) + 1}"
                motif_validation.add(range_str)
                worksheet.add_data_validation(motif_validation)

            # Add validation for 'Collaborateur' column if it exists
            if 'Collaborateur' in columns:
                col_idx = columns.index('Collaborateur') + 1
                col_letter = self._get_column_letter(col_idx)

                # Use reference to validation sheet
                collab_list_length = len(VALIDATION_LISTS["Collaborateur"])
                collab_range = f"ValidationLists!${self._get_validation_column_letter('Collaborateur')}$2:${self._get_validation_column_letter('Collaborateur')}${collab_list_length + 1}"

                collab_validation = DataValidation(
                    type="list",
                    formula1=collab_range,
                    allow_blank=True
                )

                # Apply to all data rows
                range_str = f"{col_letter}2:{col_letter}{len(plan_df) + 1}"
                collab_validation.add(range_str)
                worksheet.add_data_validation(collab_validation)

            self.logger.info("Plan Adressage validations added")

        except Exception as e:
            self.logger.error(f"Error adding Plan Adressage validations: {e}")

    def _add_rip_validations(self, writer, sheet_name: str):
        """Add data validations to RIP sheet."""
        try:
            from openpyxl.worksheet.datavalidation import DataValidation

            worksheet = writer.sheets[sheet_name]

            # Add validation for 'Type' (column D) - P0 or P1
            type_validation = DataValidation(
                type="list",
                formula1='"P0,P1"',
                allow_blank=True
            )
            type_validation.add('D2')
            worksheet.add_data_validation(type_validation)

            # Add validation for 'Acte de traitement' (column E) using Motif Voie validation list
            motif_voie_list_length = len(VALIDATION_LISTS["Motif Voie"])
            motif_voie_range = f"ValidationLists!${self._get_validation_column_letter('Motif Voie')}$2:${self._get_validation_column_letter('Motif Voie')}${motif_voie_list_length + 1}"
            acte_traitement_validation = DataValidation(
                type="list",
                formula1=motif_voie_range,
                allow_blank=True
            )
            acte_traitement_validation.add('E2')
            worksheet.add_data_validation(acte_traitement_validation)

            # Column F (Commentaire) has no validation - free text input

            # Add validation for 'Collaborateur' (column J) using existing validation list
            collaborateur_list_length = len(VALIDATION_LISTS["Collaborateur"])
            collaborateur_range = f"ValidationLists!${self._get_validation_column_letter('Collaborateur')}$2:${self._get_validation_column_letter('Collaborateur')}${collaborateur_list_length + 1}"
            collaborateur_validation = DataValidation(
                type="list",
                formula1=collaborateur_range,
                allow_blank=True
            )
            collaborateur_validation.add('J2')
            worksheet.add_data_validation(collaborateur_validation)

            self.logger.info("RIP validations added")

        except Exception as e:
            self.logger.error(f"Error adding RIP validations: {e}")

    def generate_filename(self, nom_commune: str, id_tache: str, insee: str) -> str:
        """
        Generate standardized filename for the Excel output.

        Args:
            nom_commune: Commune name
            id_tache: Task ID
            insee: INSEE code

        Returns:
            Generated filename
        """
        return f"Suivi_{nom_commune}_{id_tache}_{insee}.xlsx"
